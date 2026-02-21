#!/usr/bin/env python
"""
Update Initial Conditions liquid holdups so Francis hydraulics match the input
steady liquid-flow profile on internal stages (2..N-1).
"""

from __future__ import annotations

import argparse
import datetime as dt
from pathlib import Path
import shutil
import sys
from typing import Any, Dict, List, Optional

import numpy as np
import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_PATH = PROJECT_ROOT / "src"
if str(SRC_PATH) not in sys.path:
    sys.path.insert(0, str(SRC_PATH))

from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case  # noqa: E402
from dynamic_distillation.dynamic_run_scaffold_v1 import RunnerConfig, build_inputs_for_runner  # noqa: E402
from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel  # noqa: E402


FRANCIS_C = 3.33
SEC_PER_HOUR = 3600.0
INCHES_PER_FOOT = 12.0


def _resolve_path(raw: str) -> Path:
    p = Path(str(raw))
    if p.is_absolute():
        return p.resolve()
    return (PROJECT_ROOT / p).resolve()


def _compute_rho_profile_lbmol_ft3(*, col: Any, provider: Any, rho_default: float) -> np.ndarray:
    n = int(col.n_stages)
    rho = np.full(n, np.nan, dtype=float)
    has_density = bool(provider is not None and hasattr(provider, "liquid_density_lbmol_ft3"))

    if has_density:
        for i in range(n):
            try:
                rho_i = float(
                    provider.liquid_density_lbmol_ft3(
                        float(col.T_f[i]),
                        float(col.P_psia[i]),
                        np.asarray(col.x0[i, :], dtype=float).reshape((col.n_components,)),
                    )
                )
                if np.isfinite(rho_i) and rho_i > 0.0:
                    rho[i] = rho_i
            except Exception:
                pass

    finite = np.isfinite(rho) & (rho > 0.0)
    if not np.any(finite):
        rho[:] = float(rho_default)
        return rho

    med = float(np.median(rho[finite]))
    fill = med if np.isfinite(med) and med > 0.0 else float(rho_default)
    rho[~finite] = fill
    return rho


def _find_column_index(ws: openpyxl.worksheet.worksheet.Worksheet, name: str) -> int:
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        if str(v).strip().lower() == name.strip().lower():
            return c
    raise RuntimeError(f"Column '{name}' not found in Initial Conditions.")


def _backup_path(p: Path) -> Path:
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    return p.with_name(f"{p.stem}.holdup_update_{ts}.bak{p.suffix}")


def _fmt(x: float) -> str:
    if not np.isfinite(x):
        return "nan"
    return f"{x:.6g}"


def main() -> int:
    ap = argparse.ArgumentParser(description="Update Initial Conditions holdups from Francis inversion.")
    ap.add_argument("--excel", dest="excel_path", default="distillation_column_template.xlsx")
    ap.add_argument("--thermo", dest="thermo_mode", choices=["stub", "table", "table-pool", "dwsim"], default="table-pool")
    ap.add_argument("--thermo-table", dest="thermo_table_path", default="cache/thermo_table.json")
    ap.add_argument("--thermo-pool-workers", dest="thermo_pool_workers", type=int, default=6)
    ap.add_argument("--thermo-pool-chunk-size", dest="thermo_pool_chunk_size", type=int, default=4)
    ap.add_argument("--rho-default", dest="rho_default", type=float, default=1.0)
    ap.add_argument("--no-backup", dest="no_backup", action="store_true")
    ap.add_argument("--dry-run", dest="dry_run", action="store_true")
    args = ap.parse_args()

    excel_path = _resolve_path(str(args.excel_path))
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    thermo_table_path: Optional[Path] = None
    if str(args.thermo_mode).strip().lower() in ("table", "table-pool"):
        thermo_table_path = _resolve_path(str(args.thermo_table_path))
        if not thermo_table_path.exists():
            raise FileNotFoundError(f"Thermo table not found: {thermo_table_path}")

    case = load_case_from_excel(str(excel_path))
    col = build_column_spec_from_case(case)
    geom = getattr(col, "geometry", None)
    if geom is None:
        raise RuntimeError("No geometry found in case. Add Geometry Sections before holdup update.")

    n = int(col.n_stages)
    ml_old = np.asarray(col.M_L_lbmol, dtype=float).reshape((n,))
    l_target = np.asarray(col.L_lbmolph, dtype=float).reshape((n,))
    weir_h_in = np.asarray(geom.weir_height_in_per_stage, dtype=float).reshape((n,))
    weir_l_ft = np.asarray(geom.weir_length_ft_per_stage, dtype=float).reshape((n,))
    active_area_ft2 = np.asarray(geom.active_area_ft2_per_stage, dtype=float).reshape((n,))
    c_fac = np.asarray(
        getattr(geom, "hydraulic_c_factor_per_stage", np.ones(n, dtype=float)),
        dtype=float,
    ).reshape((n,))

    cfg = RunnerConfig(
        excel_path=str(excel_path),
        thermo_mode=str(args.thermo_mode),
        thermo_table_path=(None if thermo_table_path is None else str(thermo_table_path)),
        thermo_pool_workers=int(args.thermo_pool_workers),
        thermo_pool_chunk_size=max(int(args.thermo_pool_chunk_size), 1),
        include_temperature=True,
        include_energy=False,
        write_logs=False,
    )
    _inputs, provider = build_inputs_for_runner(case, col, cfg)
    rho = _compute_rho_profile_lbmol_ft3(col=col, provider=provider, rho_default=float(args.rho_default))

    ml_new = ml_old.copy()
    l_reconstructed = np.full(n, np.nan, dtype=float)
    h_ow = np.full(n, np.nan, dtype=float)

    internal = np.zeros(n, dtype=bool)
    if n >= 3:
        internal[1 : n - 1] = True

    for i in range(n):
        if not bool(internal[i]):
            continue

        r = float(rho[i])
        a = float(active_area_ft2[i])
        wh = float(weir_h_in[i]) / INCHES_PER_FOOT
        wl = float(weir_l_ft[i])
        c = float(c_fac[i])
        lt = float(l_target[i])

        if (not np.isfinite(r)) or r <= 0.0:
            raise RuntimeError(f"Invalid rho at stage {i + 1}: {r}")
        if (not np.isfinite(a)) or a <= 0.0:
            raise RuntimeError(f"Invalid active area at stage {i + 1}: {a}")
        if (not np.isfinite(wl)) or wl <= 0.0:
            raise RuntimeError(f"Invalid weir length at stage {i + 1}: {wl}")
        if (not np.isfinite(c)) or c <= 0.0:
            raise RuntimeError(f"Invalid hydraulic/system factor at stage {i + 1}: {c}")
        if (not np.isfinite(lt)) or lt < 0.0:
            raise RuntimeError(f"Invalid target liquid flow at stage {i + 1}: {lt}")

        den = FRANCIS_C * c * wl * r * SEC_PER_HOUR
        hov = 0.0 if lt <= 0.0 else float((lt / den) ** (2.0 / 3.0))
        htot = wh + hov
        ml_new[i] = r * a * htot
        h_ow[i] = hov
        l_reconstructed[i] = FRANCIS_C * c * wl * (max(hov, 0.0) ** 1.5) * r * SEC_PER_HOUR

    # Persist only Stage 2..N-1 holdup values in Initial Conditions.
    if not bool(args.dry_run):
        if not bool(args.no_backup):
            bkp = _backup_path(excel_path)
            shutil.copy2(excel_path, bkp)
            print(f"backup: {bkp}")

        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Initial Conditions"]
        c_stage = _find_column_index(ws, "Stage")
        c_holdup = _find_column_index(ws, "Liquid Holdup (lbmol)")

        updated_rows = 0
        for r in range(2, ws.max_row + 1):
            stage_val = ws.cell(row=r, column=c_stage).value
            if stage_val is None:
                continue
            try:
                st = int(float(stage_val))
            except Exception:
                continue
            i = st - 1
            if i < 0 or i >= n:
                continue
            if not bool(internal[i]):
                continue
            ws.cell(row=r, column=c_holdup, value=float(ml_new[i]))
            updated_rows += 1

        wb.save(excel_path)
        print(f"updated workbook: {excel_path}")
        print(f"updated rows: {updated_rows}")

    # Summary
    delta = ml_new - ml_old
    idx = np.where(internal)[0]
    max_abs_delta = float(np.nanmax(np.abs(delta[idx]))) if idx.size else 0.0
    mae_recon = float(np.nanmean(np.abs(l_reconstructed[idx] - l_target[idx]))) if idx.size else 0.0
    max_abs_recon = float(np.nanmax(np.abs(l_reconstructed[idx] - l_target[idx]))) if idx.size else 0.0

    print("")
    print("summary:")
    print(f"stages updated: 2..{n - 1}")
    print(f"max |holdup delta| lbmol: {_fmt(max_abs_delta)}")
    print(f"mean |L_recon - L_target| lbmol/h: {_fmt(mae_recon)}")
    print(f"max |L_recon - L_target| lbmol/h: {_fmt(max_abs_recon)}")
    print("")
    print("stage,ML_old,ML_new,delta,L_target,L_recon,rho,h_ow_ft")
    for i in idx:
        print(
            f"{i + 1},{_fmt(float(ml_old[i]))},{_fmt(float(ml_new[i]))},{_fmt(float(delta[i]))},"
            f"{_fmt(float(l_target[i]))},{_fmt(float(l_reconstructed[i]))},{_fmt(float(rho[i]))},{_fmt(float(h_ow[i]))}"
        )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

