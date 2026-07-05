#!/usr/bin/env python
"""
Create a workbook with tray phase compositions nudged toward fixed-flow closure.

This is a narrow initializer aid for profile-flow parity cases. It evaluates the
dynamic RHS at t=0, then updates each tray vapor composition using:

    y_required ~= y_current + dM_V/dt / V_out

for trays with a finite positive vapor outflow. With
--update-liquid-compositions, the analogous liquid update is also applied. The
absolute tray holdups are preserved; only normalized composition profiles are
changed.
"""

from __future__ import annotations

import argparse
import datetime as _dt
import json
from pathlib import Path
import shutil
import sys
from typing import Any, Dict, List, Optional

import numpy as np
from openpyxl import load_workbook

_PROJECT_ROOT = Path(__file__).resolve().parents[1]
_SRC_PATH = _PROJECT_ROOT / "src"
if str(_SRC_PATH) not in sys.path:
    sys.path.insert(0, str(_SRC_PATH))

from dynamic_distillation.column_rhs_v1 import column_rhs  # noqa: E402
from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case  # noqa: E402
from dynamic_distillation.dynamic_run_scaffold_v1 import (  # noqa: E402
    RunnerConfig,
    _clear_initial_tray_vapor_holdup,
    _initialize_vapor_holdup_from_spec_pressure,
    build_inputs_for_runner,
)
from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel  # noqa: E402
from dynamic_distillation.state_vector_layout_v1 import StateVectorLayout  # noqa: E402


def _tag() -> str:
    return _dt.datetime.now().strftime("%Y%m%d_%H%M%S")


def _resolve(raw: str | Path) -> Path:
    p = Path(str(raw))
    if p.is_absolute():
        return p.resolve()
    return (_PROJECT_ROOT / p).resolve()


def _normalize(values: np.ndarray, fallback: Optional[np.ndarray] = None) -> np.ndarray:
    arr = np.asarray(values, dtype=float).reshape((-1,))
    arr = np.where(np.isfinite(arr), arr, 0.0)
    arr = np.clip(arr, 0.0, None)
    s = float(np.sum(arr))
    if s > 1.0e-300:
        return arr / s
    if fallback is not None:
        return _normalize(fallback)
    return np.full(arr.size, 1.0 / float(arr.size), dtype=float)


def _find_header(ws: Any, header: str) -> tuple[int, int]:
    target = str(header).strip().lower()
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value).strip().lower() == target:
                return int(cell.row), int(cell.column)
    raise ValueError(f"Could not find header {header!r} in sheet {ws.title!r}")


def _component_columns(ws: Any, prefix: str, n_components: int) -> List[int]:
    cols: List[int] = []
    for k in range(1, int(n_components) + 1):
        _row, col = _find_header(ws, f"{prefix} Component {k}")
        cols.append(col)
    return cols


def _append_note(wb: Any, field: str, value: str) -> None:
    ws = wb["Notes"] if "Notes" in wb.sheetnames else wb.create_sheet("Notes")
    if ws.max_row < 1:
        ws.cell(1, 1).value = "Field"
        ws.cell(1, 2).value = "Value"
    row = int(ws.max_row) + 1
    ws.cell(row, 1).value = field
    ws.cell(row, 2).value = value


def _norm_label(value: Any) -> str:
    return "".join(ch for ch in str(value).strip().lower() if ch.isalnum())


def _stream_component_array(stream: Any, component_names: List[str]) -> np.ndarray:
    flows = getattr(stream, "component_molar_flows_lbmolph", None)
    if not isinstance(flows, dict):
        return np.zeros(len(component_names), dtype=float)
    return np.array([float(flows.get(name, 0.0) or 0.0) for name in component_names], dtype=float)


def _feed_split_vectors(col: Any) -> tuple[int, np.ndarray, np.ndarray]:
    n_comp = int(col.n_components)
    names = [str(c) for c in getattr(col, "components_excel", [])]
    if not names:
        names = [f"component_{i + 1}" for i in range(n_comp)]
    streams = getattr(col, "streams", {})
    feed = streams.get("Feed") if isinstance(streams, dict) else None
    if feed is None:
        return -1, np.zeros(n_comp, dtype=float), np.zeros(n_comp, dtype=float)
    stage0 = int(getattr(feed, "stage_1based", 1) or 1) - 1
    total = float(getattr(feed, "total_molar_flow_lbmolph", 0.0) or 0.0)
    vf = min(max(float(getattr(feed, "vapor_fraction", 0.0) or 0.0), 0.0), 1.0)
    comp = _stream_component_array(feed, names)
    if float(np.sum(comp)) <= 0.0:
        comp = np.full(n_comp, total / float(n_comp), dtype=float)
    # Convert to lbmol/s phase component flows.
    z = comp / max(float(np.sum(comp)), 1.0e-300)
    feed_l = total * (1.0 - vf) * z / 3600.0
    feed_v = total * vf * z / 3600.0
    return stage0, feed_l, feed_v


def _update_product_stream_compositions(wb: Any, col: Any, x_top: np.ndarray, x_bottom: np.ndarray) -> bool:
    if "Streams" not in wb.sheetnames:
        return False
    ws = wb["Streams"]
    n_comp = int(col.n_components)
    comp_names = [str(c) for c in getattr(col, "components_excel", [])]
    if not comp_names:
        return False

    stream_cols: Dict[str, int] = {}
    for c in range(2, ws.max_column + 1):
        key = _norm_label(ws.cell(1, c).value)
        if "distillate" in key or key.startswith("dist") or key == "top":
            stream_cols["distillate"] = int(c)
        elif "bottom" in key:
            stream_cols["bottom"] = int(c)
    if not stream_cols:
        return False

    comp_rows: Dict[str, int] = {}
    for r in range(1, ws.max_row + 1):
        label = _norm_label(ws.cell(r, 1).value)
        for name in comp_names:
            if label == _norm_label(name):
                comp_rows[name] = int(r)
    if len(comp_rows) < n_comp:
        return False

    def _total_for_col(col_idx: int) -> float:
        for r in range(1, ws.max_row + 1):
            label = _norm_label(ws.cell(r, 1).value)
            if "totalmolarflow" in label:
                return float(ws.cell(r, col_idx).value or 0.0)
        vals = [float(ws.cell(comp_rows[name], col_idx).value or 0.0) for name in comp_names]
        return float(np.sum(vals))

    if "distillate" in stream_cols:
        c = stream_cols["distillate"]
        total = _total_for_col(c)
        if total > 0.0:
            z = _normalize(x_top)
            for k, name in enumerate(comp_names):
                ws.cell(comp_rows[name], c).value = float(total * z[k])
    if "bottom" in stream_cols:
        c = stream_cols["bottom"]
        total = _total_for_col(c)
        if total > 0.0:
            z = _normalize(x_bottom)
            for k, name in enumerate(comp_names):
                ws.cell(comp_rows[name], c).value = float(total * z[k])
    return True


def _build_initial_state(
    *,
    col: Any,
    layout: StateVectorLayout,
    inputs: Any,
    include_temperature: bool,
    use_excel_vapor_holdup: bool,
) -> np.ndarray:
    y = layout.pack_y0(col)
    if not bool(use_excel_vapor_holdup):
        y = _clear_initial_tray_vapor_holdup(y, layout)
    y = _initialize_vapor_holdup_from_spec_pressure(
        col=col,
        layout=layout,
        y=y,
        inputs=inputs,
        include_temperature=bool(include_temperature),
        preserve_tray_vapor_holdup=bool(use_excel_vapor_holdup),
    )
    return np.asarray(y, dtype=float)


def _run_rhs(
    *,
    excel_path: Path,
    thermo_mode: str,
    clapeyron_model: str,
    dwsim_property_package: str,
    runtime_mode: str,
    include_temperature: bool,
    include_energy: bool,
    include_boundary_states: bool,
    use_excel_vapor_holdup: bool,
    enable_equilibrium_relaxation: bool,
    flash_feed_at_stage_conditions: Optional[bool],
    vapor_holdup_relaxation_sec: Optional[float],
) -> tuple[Any, StateVectorLayout, np.ndarray, np.ndarray, Dict[str, Any]]:
    case = load_case_from_excel(str(excel_path))
    col = build_column_spec_from_case(case)
    layout = StateVectorLayout(
        n_stages=int(col.n_stages),
        n_components=int(col.n_components),
        include_top=bool(include_boundary_states),
        include_bottom=bool(include_boundary_states),
        include_vapor=True,
        include_temperature=bool(include_temperature),
        include_energy=bool(include_energy),
    )
    cfg = RunnerConfig(
        excel_path=str(excel_path),
        runtime_mode=str(runtime_mode),
        include_temperature=bool(include_temperature),
        include_energy=bool(include_energy),
        include_boundary_states=bool(include_boundary_states),
        include_vapor_states=True,
        thermo_mode=str(thermo_mode),
        clapeyron_model=str(clapeyron_model),
        dwsim_property_package=str(dwsim_property_package),
        enable_equilibrium_relaxation=bool(enable_equilibrium_relaxation),
        flash_feed_at_stage_conditions=flash_feed_at_stage_conditions,
        vapor_holdup_relaxation_sec=vapor_holdup_relaxation_sec,
        write_logs=False,
    )
    inputs, provider = build_inputs_for_runner(case, col, cfg)
    try:
        y0 = _build_initial_state(
            col=col,
            layout=layout,
            inputs=inputs,
            include_temperature=bool(include_temperature),
            use_excel_vapor_holdup=bool(use_excel_vapor_holdup),
        )
        dydt, diag = column_rhs(0.0, y0, col, layout, inputs)
        return col, layout, np.asarray(y0, dtype=float), np.asarray(dydt, dtype=float), diag
    finally:
        if provider is not None and hasattr(provider, "close"):
            try:
                provider.close()
            except Exception:
                pass


def _max_tray_v_rate(layout: StateVectorLayout, dydt: np.ndarray, n: int, nc: int) -> float:
    du = layout.unpack(np.asarray(dydt, dtype=float))
    if "tray_V" not in du:
        return 0.0
    d_v = np.asarray(du["tray_V"], dtype=float).reshape((n, nc))
    return float(np.max(np.abs(d_v)))


def _coupled_profile_closure(
    *,
    col: Any,
    u: Dict[str, Any],
    du: Dict[str, Any],
    diag: Dict[str, Any],
    x_workbook: np.ndarray,
    y_workbook: np.ndarray,
    skip_condenser_stage: bool,
) -> tuple[np.ndarray, np.ndarray, Dict[str, Any]]:
    n = int(col.n_stages)
    nc = int(col.n_components)
    try:
        v_out = np.asarray(diag["V_out_lbmolph"], dtype=float).reshape((n,)) / 3600.0
        l_out = np.asarray(diag["L_out_lbmolph"], dtype=float).reshape((n,)) / 3600.0
    except Exception as exc:
        raise RuntimeError("RHS diagnostics must include L_out_lbmolph and V_out_lbmolph.") from exc

    y_state = np.asarray(u["y_tray"], dtype=float).reshape((n, nc))
    x_state = np.asarray(u["x_tray"], dtype=float).reshape((n, nc))
    feed_stage0, feed_l, feed_v = _feed_split_vectors(col)

    y_new = np.asarray(y_workbook, dtype=float).reshape((n, nc)).copy()
    x_new = np.asarray(x_workbook, dtype=float).reshape((n, nc)).copy()
    d_v = np.asarray(du["tray_V"], dtype=float).reshape((n, nc))

    # Bottom tray uses the current reboiler vapor source implicitly via its
    # existing residual; trays above use the newly reconciled lower-stage vapor.
    for i in range(n - 1, -1, -1):
        if bool(skip_condenser_stage) and i == 0:
            continue
        vo = float(v_out[i])
        if not np.isfinite(vo) or vo <= 1.0e-12:
            continue
        if i == n - 1:
            rhs = vo * _normalize(y_state[i, :], fallback=y_workbook[i, :]) + d_v[i, :]
        else:
            rhs = float(v_out[i + 1]) * y_new[i + 1, :]
            if feed_stage0 == i:
                rhs = rhs + feed_v
        y_new[i, :] = _normalize(rhs / vo, fallback=y_workbook[i, :])

    if not bool(skip_condenser_stage) and n > 1:
        x_new[0, :] = _normalize(y_new[1, :], fallback=x_workbook[0, :])
    elif n > 1:
        # Even when stage 1 vapor is skipped, its liquid acts as condenser
        # transfer inventory. Match it to incoming condensed vapor so the liquid
        # march starts from a physically compatible terminal composition.
        x_new[0, :] = _normalize(y_new[1, :], fallback=x_workbook[0, :])

    if "top_L" in u:
        x_top = _normalize(np.asarray(u["top_L"], dtype=float).reshape((nc,)), fallback=x_new[0, :])
    else:
        x_top = x_new[0, :].copy()

    for i in range(1, n):
        lo = float(l_out[i])
        if not np.isfinite(lo) or lo <= 1.0e-12:
            continue
        if i == 1:
            lin = float(l_out[0])
            xin = x_top
        else:
            lin = float(l_out[i - 1])
            xin = x_new[i - 1, :]
        rhs = lin * xin
        if feed_stage0 == i:
            rhs = rhs + feed_l
        x_new[i, :] = _normalize(rhs / lo, fallback=x_workbook[i, :])

    meta = {
        "feed_stage_1based": int(feed_stage0 + 1) if feed_stage0 >= 0 else None,
        "max_abs_delta_x": float(np.max(np.abs(x_new - x_workbook))),
        "max_abs_delta_y": float(np.max(np.abs(y_new - y_workbook))),
        "method": "coupled",
    }
    return x_new, y_new, meta


def main() -> int:
    ap = argparse.ArgumentParser(description="Reconcile tray vapor composition seed toward fixed-flow vapor closure.")
    ap.add_argument("--input", required=True, help="Input workbook.")
    ap.add_argument("--output", required=True, help="Output workbook.")
    ap.add_argument("--thermo", choices=["stub", "relative-volatility", "simple-rv", "constant-alpha", "clapeyron", "table", "table-pool", "dwsim"], default="clapeyron")
    ap.add_argument("--clapeyron-model", default="PR")
    ap.add_argument("--dwsim-property-package", default="pr")
    ap.add_argument("--runtime-mode", choices=["legacy", "parity", "calibration", "hydraulic"], default="parity")
    ap.add_argument("--no-temperature", dest="include_temperature", action="store_false")
    ap.add_argument("--include-energy", dest="include_energy", action="store_true")
    ap.add_argument("--no-energy", dest="include_energy", action="store_false")
    ap.set_defaults(include_temperature=True, include_energy=False)
    ap.add_argument("--disable-boundary-states", dest="include_boundary_states", action="store_false")
    ap.set_defaults(include_boundary_states=True)
    ap.add_argument("--use-excel-vapor-holdup", action="store_true")
    ap.add_argument("--no-equilibrium", dest="enable_equilibrium_relaxation", action="store_false")
    ap.set_defaults(enable_equilibrium_relaxation=True)
    ap.add_argument("--no-flash-feed-at-stage-conditions", dest="flash_feed_at_stage_conditions", action="store_false")
    ap.add_argument("--flash-feed-at-stage-conditions", dest="flash_feed_at_stage_conditions", action="store_true")
    ap.set_defaults(flash_feed_at_stage_conditions=None)
    ap.add_argument("--vapor-holdup-relaxation-sec", type=float, default=None)
    ap.add_argument(
        "--method",
        choices=["local", "coupled"],
        default="local",
        help="local=residual nudge; coupled=bottom-up vapor/top-down liquid fixed-flow closure.",
    )
    ap.add_argument("--blend", type=float, default=1.0, help="0..1 blend between old and inferred vapor composition.")
    ap.add_argument(
        "--update-liquid-compositions",
        action="store_true",
        help="Also update tray liquid compositions from fixed-flow liquid residuals.",
    )
    ap.add_argument("--skip-condenser-stage", action="store_true", default=True)
    ap.add_argument("--include-condenser-stage", dest="skip_condenser_stage", action="store_false")
    ap.add_argument("--backup", action="store_true")
    ap.add_argument("--summary-json", default=None)
    args = ap.parse_args()

    in_path = _resolve(args.input)
    out_path = _resolve(args.output)
    if not in_path.exists():
        raise FileNotFoundError(in_path)

    if bool(args.backup) and out_path.exists():
        backup_path = out_path.with_name(f"{out_path.stem}__backup_{_tag()}{out_path.suffix}")
        shutil.copy2(out_path, backup_path)

    col, layout, y0, dydt, diag = _run_rhs(
        excel_path=in_path,
        thermo_mode=str(args.thermo),
        clapeyron_model=str(args.clapeyron_model),
        dwsim_property_package=str(args.dwsim_property_package),
        runtime_mode=str(args.runtime_mode),
        include_temperature=bool(args.include_temperature),
        include_energy=bool(args.include_energy),
        include_boundary_states=bool(args.include_boundary_states),
        use_excel_vapor_holdup=bool(args.use_excel_vapor_holdup),
        enable_equilibrium_relaxation=bool(args.enable_equilibrium_relaxation),
        flash_feed_at_stage_conditions=args.flash_feed_at_stage_conditions,
        vapor_holdup_relaxation_sec=args.vapor_holdup_relaxation_sec,
    )
    n = int(col.n_stages)
    nc = int(col.n_components)
    u = layout.unpack(y0)
    du = layout.unpack(dydt)
    tray_v = np.asarray(u["tray_V"], dtype=float).reshape((n, nc))
    d_v = np.asarray(du["tray_V"], dtype=float).reshape((n, nc))
    mv = np.sum(tray_v, axis=1)
    x_workbook = np.asarray(col.x0, dtype=float).reshape((n, nc))
    y_workbook = np.asarray(col.y0, dtype=float).reshape((n, nc))
    x_state = np.asarray(u["x_tray"], dtype=float).reshape((n, nc))
    y_state = np.asarray(u["y_tray"], dtype=float).reshape((n, nc))

    try:
        v_out = np.asarray(diag["V_out_lbmolph"], dtype=float).reshape((n,)) / 3600.0
    except Exception as exc:
        raise RuntimeError("RHS diagnostics did not include V_out_lbmolph; cannot infer vapor closure.") from exc
    try:
        l_out = np.asarray(diag["L_out_lbmolph"], dtype=float).reshape((n,)) / 3600.0
    except Exception:
        l_out = np.zeros(n, dtype=float)

    blend = min(max(float(args.blend), 0.0), 1.0)
    y_new = y_workbook.copy()
    x_new = x_workbook.copy()
    updated_stages: List[int] = []
    updated_liquid_stages: List[int] = []
    max_abs_delta_y = 0.0
    max_abs_delta_x = 0.0
    max_raw_negative = 0.0
    max_inferred_sum_error = 0.0
    method_meta: Dict[str, Any] = {"method": str(args.method)}
    if str(args.method) == "coupled":
        x_target, y_target, method_meta = _coupled_profile_closure(
            col=col,
            u=u,
            du=du,
            diag=diag,
            x_workbook=x_workbook,
            y_workbook=y_workbook,
            skip_condenser_stage=bool(args.skip_condenser_stage),
        )
        y_new = np.vstack(
            [
                _normalize((1.0 - blend) * y_workbook[i, :] + blend * y_target[i, :], fallback=y_workbook[i, :])
                for i in range(n)
            ]
        )
        x_new = np.vstack(
            [
                _normalize((1.0 - blend) * x_workbook[i, :] + blend * x_target[i, :], fallback=x_workbook[i, :])
                for i in range(n)
            ]
        )
        max_abs_delta_y = float(np.max(np.abs(y_new - y_workbook)))
        max_abs_delta_x = float(np.max(np.abs(x_new - x_workbook)))
        updated_stages = [int(i + 1) for i in range(n) if float(np.max(np.abs(y_new[i, :] - y_workbook[i, :]))) > 0.0]
        updated_liquid_stages = [
            int(i + 1) for i in range(n) if float(np.max(np.abs(x_new[i, :] - x_workbook[i, :]))) > 0.0
        ]
    else:
        for i in range(n):
            if bool(args.skip_condenser_stage) and i == 0:
                continue
            vo = float(v_out[i])
            if not np.isfinite(vo) or vo <= 1.0e-12:
                continue
            y_basis = _normalize(y_state[i, :], fallback=y_workbook[i, :])
            inferred = y_basis + d_v[i, :] / vo
            max_raw_negative = min(float(max_raw_negative), float(np.min(inferred)))
            max_inferred_sum_error = max(float(max_inferred_sum_error), abs(float(np.sum(inferred)) - 1.0))
            inferred = _normalize(inferred, fallback=y_workbook[i, :])
            candidate = _normalize((1.0 - blend) * y_workbook[i, :] + blend * inferred, fallback=y_workbook[i, :])
            delta = float(np.max(np.abs(candidate - y_workbook[i, :])))
            if delta > 0.0:
                updated_stages.append(int(i + 1))
                max_abs_delta_y = max(max_abs_delta_y, delta)
                y_new[i, :] = candidate

        if bool(args.update_liquid_compositions):
            d_l = np.asarray(du["tray_L"], dtype=float).reshape((n, nc))
            for i in range(n):
                lo = float(l_out[i])
                if not np.isfinite(lo) or lo <= 1.0e-12:
                    continue
                x_basis = _normalize(x_state[i, :], fallback=x_workbook[i, :])
                inferred = _normalize(x_basis + d_l[i, :] / lo, fallback=x_workbook[i, :])
                candidate = _normalize((1.0 - blend) * x_workbook[i, :] + blend * inferred, fallback=x_workbook[i, :])
                delta = float(np.max(np.abs(candidate - x_workbook[i, :])))
                if delta > 0.0:
                    updated_liquid_stages.append(int(i + 1))
                    max_abs_delta_x = max(max_abs_delta_x, delta)
                    x_new[i, :] = candidate

    wb = load_workbook(in_path)
    if "Initial Conditions" not in wb.sheetnames:
        raise KeyError("Initial Conditions sheet not found")
    ws = wb["Initial Conditions"]
    stage_hdr_row, stage_col = _find_header(ws, "Stage")
    x_cols = _component_columns(ws, "Liquid Composition", nc)
    y_cols = _component_columns(ws, "Vapor Composition", nc)
    mv_col = None
    try:
        _r, mv_col = _find_header(ws, "Vapor Holdup (lbmol)")
    except Exception:
        mv_col = None

    stage_to_row: Dict[int, int] = {}
    for row in range(stage_hdr_row + 1, ws.max_row + 1):
        val = ws.cell(row, stage_col).value
        if val is None:
            continue
        try:
            stage_to_row[int(float(val))] = int(row)
        except Exception:
            continue

    for i in range(n):
        row = stage_to_row.get(i + 1)
        if row is None:
            raise ValueError(f"Missing Initial Conditions row for stage {i + 1}.")
        for k, col_idx in enumerate(x_cols):
            ws.cell(row, col_idx).value = float(x_new[i, k])
        for k, col_idx in enumerate(y_cols):
            ws.cell(row, col_idx).value = float(y_new[i, k])
        if mv_col is not None:
            ws.cell(row, mv_col).value = float(mv[i])

    if "Boundary State" in wb.sheetnames:
        ws_b = wb["Boundary State"]
        top_row = None
        bottom_row = None
        for row in range(1, ws_b.max_row + 1):
            label = str(ws_b.cell(row, 1).value).strip().lower()
            if label == "top_l":
                top_row = int(row)
            elif label == "bottom_l":
                bottom_row = int(row)
        if top_row is not None:
            old = np.array([float(ws_b.cell(top_row, k + 2).value or 0.0) for k in range(nc)], dtype=float)
            total = float(np.sum(old))
            if total > 0.0:
                for k in range(nc):
                    ws_b.cell(top_row, k + 2).value = float(total * x_new[0, k])
        if bottom_row is not None:
            old = np.array([float(ws_b.cell(bottom_row, k + 2).value or 0.0) for k in range(nc)], dtype=float)
            total = float(np.sum(old))
            if total > 0.0:
                for k in range(nc):
                    ws_b.cell(bottom_row, k + 2).value = float(total * x_new[-1, k])
    else:
        _update_product_stream_compositions(wb, col, x_new[0, :], x_new[-1, :])

    _append_note(
        wb,
        "Vapor closure seed reconciliation",
        (
            f"{_tag()}: Updated tray vapor compositions from RHS fixed-flow vapor residuals. "
            f"runtime_mode={args.runtime_mode}; thermo={args.thermo}; "
            f"method={args.method}; "
            f"equilibrium_relaxation={bool(args.enable_equilibrium_relaxation)}; "
            f"flash_feed_at_stage_conditions={args.flash_feed_at_stage_conditions}; "
            f"blend={blend:.6g}; updated_stages={updated_stages}; "
            f"max_abs_delta_y={max_abs_delta_y:.9g}; "
            f"update_liquid_compositions={bool(args.update_liquid_compositions)}; "
            f"updated_liquid_stages={updated_liquid_stages}; "
            f"max_abs_delta_x={max_abs_delta_x:.9g}; "
            f"max_raw_negative={max_raw_negative:.9g}; "
            f"max_inferred_sum_error={max_inferred_sum_error:.9g}."
        ),
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    # Evaluate the output workbook once so the tool reports whether the narrow
    # closure pass actually helped under the same runtime options.
    _col2, layout2, _y02, dydt2, _diag2 = _run_rhs(
        excel_path=out_path,
        thermo_mode=str(args.thermo),
        clapeyron_model=str(args.clapeyron_model),
        dwsim_property_package=str(args.dwsim_property_package),
        runtime_mode=str(args.runtime_mode),
        include_temperature=bool(args.include_temperature),
        include_energy=bool(args.include_energy),
        include_boundary_states=bool(args.include_boundary_states),
        use_excel_vapor_holdup=bool(args.use_excel_vapor_holdup),
        enable_equilibrium_relaxation=bool(args.enable_equilibrium_relaxation),
        flash_feed_at_stage_conditions=args.flash_feed_at_stage_conditions,
        vapor_holdup_relaxation_sec=args.vapor_holdup_relaxation_sec,
    )
    before_max = _max_tray_v_rate(layout, dydt, n, nc)
    after_max = _max_tray_v_rate(layout2, dydt2, n, nc)

    summary = {
        "input": str(in_path),
        "output": str(out_path),
        "updated_stages": updated_stages,
        "updated_liquid_stages": updated_liquid_stages,
        "method": str(args.method),
        "method_meta": method_meta,
        "blend": blend,
        "max_abs_delta_y": max_abs_delta_y,
        "max_abs_delta_x": max_abs_delta_x,
        "max_raw_negative_before_clip": max_raw_negative,
        "max_inferred_sum_error": max_inferred_sum_error,
        "max_abs_tray_V_rate_before_lbmolps": before_max,
        "max_abs_tray_V_rate_after_lbmolps": after_max,
    }
    summary_path = _resolve(args.summary_json) if args.summary_json else out_path.with_suffix(".vapor_closure_summary.json")
    summary_path.parent.mkdir(parents=True, exist_ok=True)
    summary_path.write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")

    print(f"Wrote: {out_path}")
    print(f"Summary: {summary_path}")
    print(f"method: {args.method}")
    print(f"updated_stages: {updated_stages}")
    print(f"updated_liquid_stages: {updated_liquid_stages}")
    print(f"max_abs_delta_y: {max_abs_delta_y:.8g}")
    print(f"max_abs_delta_x: {max_abs_delta_x:.8g}")
    print(f"max_abs_tray_V_rate_before_lbmolps: {before_max:.8g}")
    print(f"max_abs_tray_V_rate_after_lbmolps: {after_max:.8g}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
