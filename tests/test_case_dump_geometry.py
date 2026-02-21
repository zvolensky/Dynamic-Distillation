"""
test_case_dump_geometry.py

Dynamic Distillation - Geometry Case-Dump Tests

PURPOSE
-------
Verify geometry parsing and stage-range expansion behavior used by
`case_dump.py`, including compatibility with runtime module loading in tests.

SCOPE
-----
- Geometry-table extraction from workbook-like inputs
- Expanded per-stage geometry vectors and expected normalization behavior

KEY DEPENDENCIES
----------------
- openpyxl workbook fixtures
- dynamic import path for `case_dump.py`
"""


import sys
import tempfile
import importlib.util
from pathlib import Path

from openpyxl import Workbook


def _load_case_dump_module():
    """
    Load case_dump.py from repository root (one level above /tests).
    Returns the loaded module object.
    """
    repo_root = Path(__file__).resolve().parents[1]
    case_dump_py = repo_root / "case_dump.py"
    if not case_dump_py.exists():
        raise FileNotFoundError(f"case_dump.py not found at expected location: {case_dump_py}")

    spec = importlib.util.spec_from_file_location("case_dump", case_dump_py)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Unable to create module spec for: {case_dump_py}")

    mod = importlib.util.module_from_spec(spec)

    # CRITICAL: register before executing (dataclasses needs sys.modules entry)
    sys.modules[spec.name] = mod

    spec.loader.exec_module(mod)  # type: ignore[attr-defined]
    return mod


def _write_min_template(xlsx_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Specs"

    # Parameter/Value table
    ws["A1"] = "Parameter"
    ws["B1"] = "Value"
    ws["A2"] = "Number of Stages"
    ws["B2"] = 20
    ws["A3"] = "Number of Components"
    ws["B3"] = 3

    # Component names row
    ws["A5"] = "Component Name"
    ws["B5"] = "n-Propane"
    ws["C5"] = "n-Butane"
    ws["D5"] = "n-Pentane"

    # Stage Geometry header row
    ws["A10"] = "Stage Geometry"
    ws["A11"] = "Start Stage"
    ws["B11"] = "End Stage"
    ws["C11"] = "Diameter (ft)"
    ws["D11"] = "Tray Spacing (ft)"
    ws["E11"] = "Gas Void Fraction"
    ws["F11"] = "Weir Height (in)"
    ws["G11"] = "Weir Length (ft)"
    ws["H11"] = "Active Area Fraction"

    # Two range rows like your template screenshot
    ws["A12"] = 2
    ws["B12"] = 15
    ws["C12"] = 16.9619
    ws["D12"] = 1.5
    ws["E12"] = 0.75
    ws["F12"] = 2
    ws["G12"] = 10
    ws["H12"] = 0.75

    ws["A13"] = 16
    ws["B13"] = 20
    ws["C13"] = 18.1759
    ws["D13"] = 1.5
    ws["E13"] = 0.75
    ws["F13"] = 2
    ws["G13"] = 9
    ws["H13"] = 0.75

    # Stage Profile sheet (minimal but complete)
    ws_prof = wb.create_sheet(title="Stage Profile")
    ws_prof["A1"] = "Stage"
    ws_prof["B1"] = "Pressure (psia)"
    ws_prof["C1"] = "Temperature (F)"
    ws_prof["D1"] = "Holdup (lbmol)"
    ws_prof["E1"] = "n-Propane"
    ws_prof["F1"] = "n-Butane"
    ws_prof["G1"] = "n-Pentane"

    for i in range(1, 21):
        r = i + 1
        ws_prof[f"A{r}"] = i
        ws_prof[f"B{r}"] = 200.0
        ws_prof[f"C{r}"] = 100.0
        ws_prof[f"D{r}"] = 5.0
        ws_prof[f"E{r}"] = 0.5
        ws_prof[f"F{r}"] = 0.3
        ws_prof[f"G{r}"] = 0.2

    # Streams sheet (minimal feed/distillate/bottoms)
    ws_str = wb.create_sheet(title="Streams")
    ws_str["A1"] = "Stream"
    ws_str["B1"] = "Stage"
    ws_str["C1"] = "Flow (lbmol/hr)"
    ws_str["D1"] = "n-Propane"
    ws_str["E1"] = "n-Butane"
    ws_str["F1"] = "n-Pentane"

    ws_str["A2"] = "Feed"
    ws_str["B2"] = 10
    ws_str["C2"] = 100.0
    ws_str["D2"] = 0.5
    ws_str["E2"] = 0.3
    ws_str["F2"] = 0.2

    ws_str["A3"] = "Distillate"
    ws_str["B3"] = 1
    ws_str["C3"] = 40.0
    ws_str["D3"] = 0.5
    ws_str["E3"] = 0.3
    ws_str["F3"] = 0.2

    ws_str["A4"] = "Bottoms"
    ws_str["B4"] = 20
    ws_str["C4"] = 60.0
    ws_str["D4"] = 0.5
    ws_str["E4"] = 0.3
    ws_str["F4"] = 0.2

    wb.save(xlsx_path)


def test_stage_geometry_expansion_writes_vectors():
    case_dump = _load_case_dump_module()
    build_case_dump = case_dump.build_case_dump

    with tempfile.TemporaryDirectory() as td:
        xlsx = Path(td) / "template.xlsx"
        _write_min_template(xlsx)

        lines = build_case_dump(xlsx)

        # Parse key=value lines
        kv = {}
        for ln in lines:
            if "=" in ln:
                k, v = ln.split("=", 1)
                kv[k.strip()] = v.strip()

        assert kv["N"] == "20"
        assert kv["NC"] == "3"

        # Diameter vector should be length 20 with zeros for uncovered stage 1
        diam = [float(x) for x in kv["TRAY_DIAM_FT"].split(",")]
        assert len(diam) == 20
        assert diam[0] == 0.0  # stage 1 not covered by geometry table
        assert abs(diam[1] - 16.9619) < 1e-6   # stage 2
        assert abs(diam[14] - 16.9619) < 1e-6  # stage 15
        assert abs(diam[15] - 18.1759) < 1e-6  # stage 16
        assert abs(diam[19] - 18.1759) < 1e-6  # stage 20

        # Weir length should switch from 10 to 9 at stage 16
        weirL = [float(x) for x in kv["WEIR_LENGTH_FT"].split(",")]
        assert len(weirL) == 20
        assert weirL[1] == 10.0   # stage 2
        assert weirL[14] == 10.0  # stage 15
        assert weirL[15] == 9.0   # stage 16
        assert weirL[19] == 9.0   # stage 20

        # Active area fraction should be either 0.0 (uncovered stage 1) or 0.75 (covered stages)
        aaf = [float(x) for x in kv["ACTIVE_AREA_FRACTION"].split(",")]
        assert len(aaf) == 20
        assert aaf[0] == 0.0
        assert all(v in (0.0, 0.75) for v in aaf)
