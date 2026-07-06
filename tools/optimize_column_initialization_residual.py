#!/usr/bin/env python
"""
Bounded least-squares initializer for column composition profiles.

This experimental tool minimizes the same t=0 residuals reported by
column_initialization_residual_audit.py, while preserving phase holdup totals
and writing a new Excel seed workbook.
"""

from __future__ import annotations

import argparse
from dataclasses import replace
import json
import math
from pathlib import Path
import shutil
import subprocess
import sys
from typing import Any, Dict, List, Optional

import numpy as np
from openpyxl import load_workbook
from scipy.optimize import least_squares

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_PATH = PROJECT_ROOT / "src"
if str(SRC_PATH) not in sys.path:
    sys.path.insert(0, str(SRC_PATH))

from dynamic_distillation.column_rhs_v1 import BoundaryFlows, column_rhs  # noqa: E402
from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case  # noqa: E402
from dynamic_distillation.dynamic_run_scaffold_v1 import (  # noqa: E402
    RunnerConfig,
    _clear_initial_tray_vapor_holdup,
    _initialize_vapor_holdup_from_spec_pressure,
    build_inputs_for_runner,
)
from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel  # noqa: E402
from dynamic_distillation.state_vector_layout_v1 import StateVectorLayout  # noqa: E402


def _resolve(raw: str | Path) -> Path:
    p = Path(str(raw))
    if p.is_absolute():
        return p.resolve()
    return (PROJECT_ROOT / p).resolve()


def _normalize(v: np.ndarray, fallback: Optional[np.ndarray] = None) -> np.ndarray:
    arr = np.asarray(v, dtype=float).reshape((-1,))
    arr = np.where(np.isfinite(arr), arr, 0.0)
    arr = np.clip(arr, 0.0, None)
    s = float(np.sum(arr))
    if s > 1.0e-300:
        return arr / s
    if fallback is not None:
        return _normalize(fallback)
    return np.full(arr.size, 1.0 / float(arr.size), dtype=float)


def _softmax(logits: np.ndarray) -> np.ndarray:
    z = np.asarray(logits, dtype=float).reshape((-1,))
    z = z - float(np.max(z))
    e = np.exp(np.clip(z, -60.0, 60.0))
    return e / max(float(np.sum(e)), 1.0e-300)


def _logit(comp: np.ndarray) -> np.ndarray:
    c = _normalize(comp)
    return np.log(np.clip(c, 1.0e-12, 1.0))


def _find_header(ws: Any, header: str) -> tuple[int, int]:
    target = str(header).strip().lower()
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value).strip().lower() == target:
                return int(cell.row), int(cell.column)
    raise ValueError(f"Could not find header {header!r} in sheet {ws.title!r}")


def _component_columns(ws: Any, prefix: str, n_components: int) -> List[int]:
    cols: List[int] = []
    for k in range(1, int(n_components) + 1):
        _row, col = _find_header(ws, f"{prefix} Component {k}")
        cols.append(col)
    return cols


def _norm_label(value: Any) -> str:
    return "".join(ch for ch in str(value).strip().lower() if ch.isalnum())


def _update_product_stream_compositions(wb: Any, col: Any, x_top: np.ndarray, x_bottom: np.ndarray) -> None:
    if "Streams" not in wb.sheetnames:
        return
    ws = wb["Streams"]
    names = [str(c) for c in getattr(col, "components_excel", [])]
    if not names:
        return
    stream_cols: Dict[str, int] = {}
    for c in range(2, ws.max_column + 1):
        key = _norm_label(ws.cell(1, c).value)
        if "distillate" in key or key.startswith("dist") or key == "top":
            stream_cols["distillate"] = int(c)
        elif "bottom" in key:
            stream_cols["bottom"] = int(c)
    comp_rows: Dict[str, int] = {}
    for r in range(1, ws.max_row + 1):
        label = _norm_label(ws.cell(r, 1).value)
        for name in names:
            if label == _norm_label(name):
                comp_rows[name] = int(r)
    if len(comp_rows) != len(names):
        return

    def total_for_col(col_idx: int) -> float:
        for r in range(1, ws.max_row + 1):
            if "totalmolarflow" in _norm_label(ws.cell(r, 1).value):
                return float(ws.cell(r, col_idx).value or 0.0)
        return float(sum(float(ws.cell(comp_rows[name], col_idx).value or 0.0) for name in names))

    for key, comp in (("distillate", x_top), ("bottom", x_bottom)):
        col_idx = stream_cols.get(key)
        if col_idx is None:
            continue
        total = total_for_col(col_idx)
        if total <= 0.0:
            continue
        z = _normalize(comp)
        for k, name in enumerate(names):
            ws.cell(comp_rows[name], col_idx).value = float(total * z[k])


def _stream_total(col: Any, key: str) -> float:
    streams = getattr(col, "streams", {}) or {}
    stream = None
    for name, candidate in streams.items():
        norm = _norm_label(name)
        if key == "distillate" and ("distillate" in norm or norm.startswith("top")):
            stream = candidate
            break
        if key == "bottom" and "bottom" in norm:
            stream = candidate
            break
    if stream is None:
        return 0.0
    total = getattr(stream, "total_molar_flow_lbmolph", None)
    if total is not None:
        try:
            val = float(total)
            if np.isfinite(val):
                return val
        except Exception:
            pass
    flows = getattr(stream, "component_molar_flows_lbmolph", None)
    if isinstance(flows, dict):
        return float(sum(float(v or 0.0) for v in flows.values()))
    return 0.0


def _update_stream_total(wb: Any, key: str, total: float) -> None:
    if "Streams" not in wb.sheetnames:
        return
    ws = wb["Streams"]
    stream_col = None
    for c in range(2, ws.max_column + 1):
        label = _norm_label(ws.cell(1, c).value)
        if key == "distillate" and ("distillate" in label or label.startswith("top")):
            stream_col = int(c)
            break
        if key == "bottom" and "bottom" in label:
            stream_col = int(c)
            break
    if stream_col is None:
        return
    total_row = None
    comp_rows: List[int] = []
    for r in range(1, ws.max_row + 1):
        label = _norm_label(ws.cell(r, 1).value)
        if "totalmolarflow" in label:
            total_row = int(r)
        elif label and label not in {
            "stage",
            "pressurepsia",
            "vapourfraction",
            "vaporfraction",
            "temperaturef",
            "moleflowslbmolh",
        }:
            try:
                float(ws.cell(r, stream_col).value or 0.0)
                comp_rows.append(int(r))
            except Exception:
                pass
    old_vals = np.array([float(ws.cell(r, stream_col).value or 0.0) for r in comp_rows], dtype=float)
    old_total = float(np.sum(old_vals))
    if total_row is not None:
        ws.cell(total_row, stream_col).value = float(total)
    if comp_rows and old_total > 0.0:
        z = old_vals / old_total
        for r, frac in zip(comp_rows, z):
            ws.cell(r, stream_col).value = float(total) * float(frac)


def _write_boundary_state(
    wb: Any,
    col: Any,
    top_l: Optional[np.ndarray],
    top_v: Optional[np.ndarray],
    bottom_l: Optional[np.ndarray],
    bottom_v: Optional[np.ndarray],
) -> None:
    names = [str(c) for c in getattr(col, "components_excel", [])]
    nc = int(getattr(col, "n_components", len(names)))
    if not names or len(names) != nc:
        names = [f"Component {i}" for i in range(1, nc + 1)]
    existing: Dict[str, np.ndarray] = {}
    if "Boundary State" in wb.sheetnames:
        ws_old = wb["Boundary State"]
        for row in range(2, ws_old.max_row + 1):
            key = str(ws_old.cell(row, 1).value or "").strip()
            if key:
                vals = [float(ws_old.cell(row, k + 2).value or 0.0) for k in range(nc)]
                existing[key] = np.asarray(vals, dtype=float)
        del wb["Boundary State"]
    rows = {
        "top_L": top_l if top_l is not None else existing.get("top_L"),
        "top_V": top_v if top_v is not None else existing.get("top_V"),
        "bottom_L": bottom_l if bottom_l is not None else existing.get("bottom_L"),
        "bottom_V": bottom_v if bottom_v is not None else existing.get("bottom_V"),
    }
    ws = wb.create_sheet("Boundary State")
    for c, value in enumerate(["State", *names], start=1):
        ws.cell(1, c).value = value
    r = 2
    for key in ("top_L", "top_V", "bottom_L", "bottom_V"):
        arr = rows.get(key)
        if arr is None:
            arr = np.zeros(nc, dtype=float)
        arr = np.asarray(arr, dtype=float).reshape((nc,))
        ws.cell(r, 1).value = key
        for k in range(nc):
            ws.cell(r, k + 2).value = float(arr[k])
        r += 1


def _update_spec_value(wb: Any, label: str, value: Any) -> None:
    if "Specifications" not in wb.sheetnames:
        return
    ws = wb["Specifications"]
    target = _norm_label(label)
    for r in range(1, ws.max_row + 1):
        if _norm_label(ws.cell(r, 1).value) == target:
            ws.cell(r, 2).value = value
            return
    row = ws.max_row + 1
    ws.cell(row, 1).value = label
    ws.cell(row, 2).value = value


def _build_case_runtime(args: argparse.Namespace) -> tuple[Any, StateVectorLayout, Any, np.ndarray, Dict[str, Any]]:
    excel_path = _resolve(args.input)
    case = load_case_from_excel(str(excel_path))
    col = build_column_spec_from_case(case)
    layout = StateVectorLayout(
        n_stages=int(col.n_stages),
        n_components=int(col.n_components),
        include_top=bool(args.include_boundary_states),
        include_bottom=bool(args.include_boundary_states),
        include_vapor=True,
        include_temperature=bool(args.include_temperature),
        include_energy=bool(args.include_energy),
    )
    cfg = RunnerConfig(
        excel_path=str(excel_path),
        runtime_mode=str(args.runtime_mode),
        include_temperature=bool(args.include_temperature),
        include_energy=bool(args.include_energy),
        include_boundary_states=bool(args.include_boundary_states),
        include_vapor_states=True,
        thermo_mode=str(args.thermo),
        clapeyron_model=str(args.clapeyron_model),
        condenser_duty_mode=str(args.condenser_duty_mode),
        condenser_duty_btu_per_h=args.condenser_duty_btuph,
        enable_equilibrium_relaxation=not bool(args.no_equilibrium),
        flash_feed_at_stage_conditions=(False if bool(args.no_flash_feed_at_stage_conditions) else None),
        vapor_holdup_relaxation_sec=float(args.vapor_holdup_relaxation_sec),
        write_logs=False,
    )
    inputs, provider = build_inputs_for_runner(case, col, cfg)
    inputs = replace(
        inputs,
        feed_stage_flash_prev=None,
        feed_stage_flash_reuse_dT_F=0.0,
        feed_stage_flash_reuse_dP_psia=0.0,
        feed_stage_flash_reuse_dx=0.0,
    )
    y0 = layout.pack_y0(col)
    if not bool(args.use_excel_vapor_holdup):
        y0 = _clear_initial_tray_vapor_holdup(y0, layout)
    y0 = _initialize_vapor_holdup_from_spec_pressure(
        col=col,
        layout=layout,
        y=y0,
        inputs=inputs,
        include_temperature=bool(args.include_temperature),
        preserve_tray_vapor_holdup=bool(args.use_excel_vapor_holdup),
    )
    return col, layout, inputs, np.asarray(y0, dtype=float), {"provider": provider}


def _selected_stages(raw: str, n: int) -> List[int]:
    key = str(raw).strip().lower()
    if key in ("all", "*"):
        return list(range(n))
    out: List[int] = []
    for part in key.split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            a, b = [int(x) for x in part.split("-", 1)]
            out.extend(range(a - 1, b))
        else:
            out.append(int(part) - 1)
    return sorted({i for i in out if 0 <= i < n})


def _selected_names(raw: str, allowed: List[str], *, default: List[str]) -> List[str]:
    key = str(raw or "").strip()
    if not key or key.lower() == "default":
        return list(default)
    if key.lower() in ("all", "*"):
        return list(allowed)
    allowed_set = set(allowed)
    out: List[str] = []
    for part in key.split(","):
        name = part.strip()
        if not name:
            continue
        if name not in allowed_set:
            raise ValueError(f"Unsupported residual block {name!r}; allowed: {', '.join(allowed)}")
        if name not in out:
            out.append(name)
    if not out:
        raise ValueError(f"No valid residual blocks selected from {raw!r}")
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description="Bounded least-squares t=0 initializer.")
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--stages", default="all", help="1-based stages to vary, e.g. all or 2-5,18-20.")
    ap.add_argument("--vary-liquid", action="store_true")
    ap.add_argument("--vary-vapor", action="store_true")
    ap.add_argument("--vary-liquid-flow", action="store_true")
    ap.add_argument("--vary-vapor-flow", action="store_true")
    ap.add_argument("--vary-boundary-flows", action="store_true")
    ap.add_argument(
        "--chemsep-product-specs",
        action="store_true",
        help=(
            "Use ChemSep-like product DOFs: vary distillate and enforce reflux ratio; "
            "bottom-side DOFs remain fixed unless explicitly varied."
        ),
    )
    ap.add_argument("--reflux-ratio", type=float, default=None)
    ap.add_argument("--vary-reflux", action="store_true")
    ap.add_argument("--vary-boilup", action="store_true")
    ap.add_argument("--vary-distillate", action="store_true")
    ap.add_argument("--vary-bottoms", action="store_true")
    ap.add_argument("--vary-feed-temperature", action="store_true")
    ap.add_argument("--vary-feed-vapor-fraction", action="store_true")
    ap.add_argument("--vary-tray-energy", action="store_true")
    ap.add_argument("--vary-condenser-duty", action="store_true")
    ap.add_argument("--vary-top-liquid", action="store_true")
    ap.add_argument("--vary-top-vapor", action="store_true")
    ap.add_argument("--vary-bottom-liquid", action="store_true")
    ap.add_argument("--vary-bottom-vapor", action="store_true")
    ap.add_argument("--vary-top-vapor-total", action="store_true")
    ap.add_argument(
        "--residual-stages",
        default="all",
        help="1-based tray stages to include in tray_L/tray_V residuals; default all.",
    )
    ap.add_argument(
        "--residual-top-boundary-only",
        action="store_true",
        help="Include top_L/top_V boundary residuals but omit bottom_L/bottom_V from the objective.",
    )
    ap.add_argument(
        "--residual-state-blocks",
        default="default",
        help=(
            "Comma-separated material state blocks for the objective. "
            "Default is tray_L,tray_V,top_L,top_V,bottom_L,bottom_V; "
            "use tray_V for global vapor-only closure."
        ),
    )
    ap.add_argument(
        "--residual-energy-blocks",
        default="default",
        help=(
            "Comma-separated energy state blocks for the objective when --include-energy. "
            "Default is tray_EL_BTU,tray_EV_BTU; use tray_EL_BTU for liquid-enthalpy-only closure."
        ),
    )
    ap.add_argument("--max-logit-delta", type=float, default=1.0)
    ap.add_argument("--max-top-vapor-total-log-delta", type=float, default=0.5)
    ap.add_argument("--max-flow-log-delta", type=float, default=0.15)
    ap.add_argument("--max-boundary-log-delta", type=float, default=0.15)
    ap.add_argument("--max-feed-temperature-delta-f", type=float, default=25.0)
    ap.add_argument("--max-feed-vapor-fraction-delta", type=float, default=0.5)
    ap.add_argument("--max-energy-rel-delta", type=float, default=0.25)
    ap.add_argument("--max-condenser-duty-rel-delta", type=float, default=0.25)
    ap.add_argument("--max-nfev", type=int, default=40)
    ap.add_argument("--denom-floor-lbmol", type=float, default=1.0)
    ap.add_argument("--energy-denom-floor-btu", type=float, default=1.0)
    ap.add_argument("--residual-scale", type=float, default=1.0)
    ap.add_argument("--energy-residual-scale", type=float, default=1.0)
    ap.add_argument("--tray-v-residual-weight", type=float, default=1.0)
    ap.add_argument("--tray-l-residual-weight", type=float, default=1.0)
    ap.add_argument("--top-l-residual-weight", type=float, default=1.0)
    ap.add_argument("--top-v-residual-weight", type=float, default=1.0)
    ap.add_argument("--bottom-l-residual-weight", type=float, default=1.0)
    ap.add_argument("--bottom-v-residual-weight", type=float, default=1.0)
    ap.add_argument("--bottom-boundary-balance-weight", type=float, default=0.0)
    ap.add_argument("--bottom-boundary-total-weight", type=float, default=0.0)
    ap.add_argument("--profile-penalty", type=float, default=0.02)
    ap.add_argument("--flow-penalty", type=float, default=0.02)
    ap.add_argument("--boundary-penalty", type=float, default=0.02)
    ap.add_argument("--feed-temperature-penalty", type=float, default=0.02)
    ap.add_argument("--feed-vapor-fraction-penalty", type=float, default=0.02)
    ap.add_argument("--energy-penalty", type=float, default=0.02)
    ap.add_argument("--tray-total-penalty", type=float, default=0.0)
    ap.add_argument("--thermo", default="clapeyron")
    ap.add_argument("--clapeyron-model", default="PR")
    ap.add_argument("--runtime-mode", default="parity")
    ap.add_argument("--condenser-duty-mode", default="total-condense")
    ap.add_argument("--condenser-duty-btuph", type=float, default=None)
    ap.add_argument("--no-equilibrium", action="store_true")
    ap.add_argument("--no-flash-feed-at-stage-conditions", action="store_true")
    ap.add_argument("--vapor-holdup-relaxation-sec", type=float, default=0.0)
    ap.add_argument("--use-excel-vapor-holdup", action="store_true")
    ap.add_argument("--no-temperature", dest="include_temperature", action="store_false")
    ap.add_argument("--include-energy", dest="include_energy", action="store_true")
    ap.set_defaults(include_temperature=True, include_energy=False)
    ap.add_argument("--disable-boundary-states", dest="include_boundary_states", action="store_false")
    ap.set_defaults(include_boundary_states=True)
    ap.add_argument("--audit-output-dir", default=None)
    args = ap.parse_args()
    if not args.vary_liquid and not args.vary_vapor:
        if not args.vary_liquid_flow and not args.vary_vapor_flow:
            if not args.vary_boundary_flows and not any(
                bool(v)
                for v in (
                    args.vary_reflux,
                    args.vary_boilup,
                    args.vary_distillate,
                    args.vary_bottoms,
                    args.vary_feed_temperature,
                    args.vary_feed_vapor_fraction,
                    args.vary_tray_energy,
                    args.vary_condenser_duty,
                    args.vary_top_liquid,
                    args.vary_top_vapor,
                    args.vary_bottom_liquid,
                    args.vary_bottom_vapor,
                    args.vary_top_vapor_total,
                )
            ):
                args.vary_liquid = True
                args.vary_vapor = True
    if bool(args.vary_boundary_flows):
        args.vary_reflux = True
        args.vary_boilup = True
        args.vary_distillate = True
        args.vary_bottoms = True
    explicit_boundary_flags = {
        "reflux": "--vary-reflux" in sys.argv[1:],
        "boilup": "--vary-boilup" in sys.argv[1:],
        "distillate": "--vary-distillate" in sys.argv[1:],
        "bottoms": "--vary-bottoms" in sys.argv[1:],
    }
    if bool(args.chemsep_product_specs):
        if not explicit_boundary_flags["reflux"]:
            args.vary_reflux = False
        if not explicit_boundary_flags["boilup"]:
            args.vary_boilup = False
        if not explicit_boundary_flags["distillate"]:
            args.vary_distillate = True
        if not explicit_boundary_flags["bottoms"]:
            args.vary_bottoms = False

    col, layout, inputs, y_base, extra = _build_case_runtime(args)
    n = int(col.n_stages)
    nc = int(col.n_components)
    stages = _selected_stages(args.stages, n)
    residual_stages = _selected_stages(args.residual_stages, n)
    state_residual_blocks = _selected_names(
        args.residual_state_blocks,
        ["tray_L", "tray_V", "top_L", "top_V", "bottom_L", "bottom_V"],
        default=["tray_L", "tray_V", "top_L", "top_V", "bottom_L", "bottom_V"],
    )
    energy_residual_blocks = _selected_names(
        args.residual_energy_blocks,
        ["tray_EL_BTU", "tray_EV_BTU"],
        default=["tray_EL_BTU", "tray_EV_BTU"],
    )
    u_base = layout.unpack(y_base)
    tray_l_base = np.asarray(u_base["tray_L"], dtype=float).reshape((n, nc))
    tray_v_base = np.asarray(u_base["tray_V"], dtype=float).reshape((n, nc))
    top_l_base = np.asarray(u_base.get("top_L", np.zeros(nc)), dtype=float).reshape((nc,))
    top_v_base = np.asarray(u_base.get("top_V", np.zeros(nc)), dtype=float).reshape((nc,))
    bottom_l_base = np.asarray(u_base.get("bottom_L", np.zeros(nc)), dtype=float).reshape((nc,))
    bottom_v_base = np.asarray(u_base.get("bottom_V", np.zeros(nc)), dtype=float).reshape((nc,))
    top_l_total_base = float(np.sum(top_l_base))
    top_v_total_base = float(np.sum(top_v_base))
    bottom_l_total_base = float(np.sum(bottom_l_base))
    bottom_v_total_base = float(np.sum(bottom_v_base))
    x0 = np.asarray(col.x0, dtype=float).reshape((n, nc))
    yv0 = np.asarray(col.y0, dtype=float).reshape((n, nc))
    x_top_base = _normalize(top_l_base, fallback=x0[0, :])
    y_top_base = _normalize(top_v_base, fallback=yv0[0, :])
    x_bottom_base = _normalize(bottom_l_base, fallback=x0[-1, :])
    y_bottom_base = _normalize(bottom_v_base, fallback=yv0[-1, :])
    tray_el_base = np.asarray(u_base.get("tray_EL_BTU", np.zeros(n)), dtype=float).reshape((n,))
    tray_ev_base = np.asarray(u_base.get("tray_EV_BTU", np.zeros(n)), dtype=float).reshape((n,))
    ml = np.sum(tray_l_base, axis=1)
    mv = np.sum(tray_v_base, axis=1)
    L0_lbmolph = np.asarray(col.L_lbmolph, dtype=float).reshape((n,))
    V0_lbmolph = np.asarray(col.V_lbmolph, dtype=float).reshape((n,))
    boundary_base = {
        "reflux": float(L0_lbmolph[0]) if L0_lbmolph.size else 0.0,
        "boilup": float(V0_lbmolph[-1]) if V0_lbmolph.size else 0.0,
        "distillate": _stream_total(col, "distillate"),
        "bottoms": _stream_total(col, "bottom"),
    }
    condenser_duty_base = args.condenser_duty_btuph
    if condenser_duty_base is None:
        condenser_duty_base = getattr(inputs, "condenser_duty_btu_per_h", None)
    if condenser_duty_base is None:
        specs_raw = getattr(col, "specs_raw", {}) or {}
        for key in ("Condenser Duty (Btu/h)", "Condenser Duty", "Q Condenser (Btu/h)"):
            try:
                val = specs_raw.get(key)
                if val is not None and str(val).strip() != "":
                    condenser_duty_base = float(val)
                    break
            except Exception:
                pass
    if condenser_duty_base is None:
        condenser_duty_base = 0.0
    condenser_duty_base = float(condenser_duty_base)
    reflux_ratio = None
    if bool(args.chemsep_product_specs):
        if args.reflux_ratio is not None:
            reflux_ratio = float(args.reflux_ratio)
        elif float(boundary_base["distillate"]) > 1.0e-12:
            reflux_ratio = float(boundary_base["reflux"]) / float(boundary_base["distillate"])
        else:
            raise ValueError("--chemsep-product-specs requires nonzero distillate or explicit --reflux-ratio")
    feed_stream_base = (getattr(col, "streams", {}) or {}).get("Feed")
    feed_temperature_base = None
    feed_vapor_fraction_base = None
    if feed_stream_base is not None and getattr(feed_stream_base, "temperature_f", None) is not None:
        try:
            feed_temperature_base = float(getattr(feed_stream_base, "temperature_f"))
        except Exception:
            feed_temperature_base = None
    if feed_stream_base is not None and getattr(feed_stream_base, "vapor_fraction", None) is not None:
        try:
            feed_vapor_fraction_base = float(getattr(feed_stream_base, "vapor_fraction"))
        except Exception:
            feed_vapor_fraction_base = None
    state_residual_weights = {
        "tray_V": float(args.tray_v_residual_weight),
        "tray_L": float(args.tray_l_residual_weight),
        "top_L": float(args.top_l_residual_weight),
        "top_V": float(args.top_v_residual_weight),
        "bottom_L": float(args.bottom_l_residual_weight),
        "bottom_V": float(args.bottom_v_residual_weight),
    }

    blocks: List[tuple[str, int]] = []
    for i in stages:
        if bool(args.vary_liquid):
            blocks.append(("L", i))
        if bool(args.vary_vapor) and mv[i] > 1.0e-12:
            blocks.append(("V", i))
    top_comp_blocks: List[str] = []
    if bool(args.vary_top_liquid) and top_l_total_base > 1.0e-12:
        top_comp_blocks.append("top_L")
    if bool(args.vary_top_vapor) and top_v_total_base > 1.0e-12:
        top_comp_blocks.append("top_V")
    bottom_comp_blocks: List[str] = []
    if bool(args.vary_bottom_liquid) and bottom_l_total_base > 1.0e-12:
        bottom_comp_blocks.append("bottom_L")
    if bool(args.vary_bottom_vapor) and bottom_v_total_base > 1.0e-12:
        bottom_comp_blocks.append("bottom_V")
    flow_blocks: List[tuple[str, int]] = []
    for i in stages:
        if bool(args.vary_liquid_flow) and L0_lbmolph[i] > 1.0e-12:
            flow_blocks.append(("LF", i))
        if bool(args.vary_vapor_flow) and V0_lbmolph[i] > 1.0e-12:
            flow_blocks.append(("VF", i))
    boundary_blocks: List[str] = []
    for key, enabled in (
        ("reflux", args.vary_reflux),
        ("boilup", args.vary_boilup),
        ("distillate", args.vary_distillate),
        ("bottoms", args.vary_bottoms),
    ):
        if bool(enabled) and float(boundary_base.get(key, 0.0)) > 1.0e-12:
            boundary_blocks.append(key)
    energy_blocks: List[tuple[str, int]] = []
    if bool(args.vary_tray_energy) and bool(args.include_energy):
        for i in stages:
            energy_blocks.append(("EL", i))
            if i > 0 or abs(float(tray_ev_base[i])) > 1.0e-12:
                energy_blocks.append(("EV", i))
    n_comp_var = len(blocks) * nc
    n_top_comp_var = len(top_comp_blocks) * nc
    n_bottom_comp_var = len(bottom_comp_blocks) * nc
    n_top_vapor_total_var = 1 if bool(args.vary_top_vapor_total) and top_v_total_base > 1.0e-12 else 0
    n_profile_var = n_comp_var + n_top_comp_var + n_bottom_comp_var
    n_state_var = n_profile_var + n_top_vapor_total_var
    n_flow_var = len(flow_blocks)
    n_boundary_var = len(boundary_blocks)
    n_feed_temp_var = 1 if bool(args.vary_feed_temperature) and feed_temperature_base is not None else 0
    n_feed_vf_var = 1 if bool(args.vary_feed_vapor_fraction) and feed_vapor_fraction_base is not None else 0
    n_condenser_duty_var = 1 if bool(args.vary_condenser_duty) and abs(float(condenser_duty_base)) > 1.0e-12 else 0
    n_energy_var = len(energy_blocks)
    nvar = (
        n_state_var
        + n_flow_var
        + n_boundary_var
        + n_feed_temp_var
        + n_feed_vf_var
        + n_condenser_duty_var
        + n_energy_var
    )
    z0 = np.zeros(nvar, dtype=float)
    lb = np.concatenate(
        [
            np.full(n_comp_var, -abs(float(args.max_logit_delta)), dtype=float),
            np.full(n_top_comp_var, -abs(float(args.max_logit_delta)), dtype=float),
            np.full(n_bottom_comp_var, -abs(float(args.max_logit_delta)), dtype=float),
            np.full(n_top_vapor_total_var, -abs(float(args.max_top_vapor_total_log_delta)), dtype=float),
            np.full(n_flow_var, -abs(float(args.max_flow_log_delta)), dtype=float),
            np.full(n_boundary_var, -abs(float(args.max_boundary_log_delta)), dtype=float),
            np.full(n_feed_temp_var, -abs(float(args.max_feed_temperature_delta_f)), dtype=float),
            np.full(n_feed_vf_var, -abs(float(args.max_feed_vapor_fraction_delta)), dtype=float),
            np.full(n_condenser_duty_var, -abs(float(args.max_condenser_duty_rel_delta)), dtype=float),
            np.full(n_energy_var, -abs(float(args.max_energy_rel_delta)), dtype=float),
        ]
    )
    ub = np.concatenate(
        [
            np.full(n_comp_var, abs(float(args.max_logit_delta)), dtype=float),
            np.full(n_top_comp_var, abs(float(args.max_logit_delta)), dtype=float),
            np.full(n_bottom_comp_var, abs(float(args.max_logit_delta)), dtype=float),
            np.full(n_top_vapor_total_var, abs(float(args.max_top_vapor_total_log_delta)), dtype=float),
            np.full(n_flow_var, abs(float(args.max_flow_log_delta)), dtype=float),
            np.full(n_boundary_var, abs(float(args.max_boundary_log_delta)), dtype=float),
            np.full(n_feed_temp_var, abs(float(args.max_feed_temperature_delta_f)), dtype=float),
            np.full(n_feed_vf_var, abs(float(args.max_feed_vapor_fraction_delta)), dtype=float),
            np.full(n_condenser_duty_var, abs(float(args.max_condenser_duty_rel_delta)), dtype=float),
            np.full(n_energy_var, abs(float(args.max_energy_rel_delta)), dtype=float),
        ]
    )

    eval_count = 0
    best = {"norm": float("inf"), "z": z0.copy(), "max_rel": float("inf")}

    def make_state(z: np.ndarray) -> np.ndarray:
        y_state = y_base.copy()
        unpacked = layout.unpack(y_state)
        tray_l = np.asarray(unpacked["tray_L"], dtype=float).reshape((n, nc)).copy()
        tray_v = np.asarray(unpacked["tray_V"], dtype=float).reshape((n, nc)).copy()
        # Ensure untouched zero-holdup vapor rows still unpack to the workbook
        # composition if they later become nonzero through initialization.
        for ii in range(n):
            if mv[ii] <= 1.0e-12:
                tray_v[ii, :] = 0.0
        idx = 0
        for phase, stage in blocks:
            delta = np.asarray(z[idx: idx + nc], dtype=float)
            idx += nc
            if phase == "L":
                comp = _softmax(_logit(x0[stage, :]) + delta)
                tray_l[stage, :] = float(ml[stage]) * comp
            else:
                comp = _softmax(_logit(yv0[stage, :]) + delta)
                tray_v[stage, :] = float(mv[stage]) * comp
        sl = layout.slices()
        y_state[sl["tray_L"]] = tray_l.reshape(-1)
        y_state[sl["tray_V"]] = tray_v.reshape(-1)
        top_l = np.asarray(unpacked.get("top_L", top_l_base), dtype=float).reshape((nc,)).copy()
        top_v = np.asarray(unpacked.get("top_V", top_v_base), dtype=float).reshape((nc,)).copy()
        bottom_l = np.asarray(unpacked.get("bottom_L", bottom_l_base), dtype=float).reshape((nc,)).copy()
        bottom_v = np.asarray(unpacked.get("bottom_V", bottom_v_base), dtype=float).reshape((nc,)).copy()
        for key in top_comp_blocks:
            delta = np.asarray(z[idx: idx + nc], dtype=float)
            idx += nc
            if key == "top_L":
                comp = _softmax(_logit(x_top_base) + delta)
                top_l = float(top_l_total_base) * comp
            elif key == "top_V":
                comp = _softmax(_logit(y_top_base) + delta)
                top_v = float(top_v_total_base) * comp
        for key in bottom_comp_blocks:
            delta = np.asarray(z[idx: idx + nc], dtype=float)
            idx += nc
            if key == "bottom_L":
                comp = _softmax(_logit(x_bottom_base) + delta)
                bottom_l = float(bottom_l_total_base) * comp
            elif key == "bottom_V":
                comp = _softmax(_logit(y_bottom_base) + delta)
                bottom_v = float(bottom_v_total_base) * comp
        if n_top_vapor_total_var > 0:
            scale = math.exp(float(z[idx]))
            idx += 1
            top_v = float(top_v_total_base) * float(scale) * _normalize(top_v, fallback=y_top_base)
        if "top_L" in sl:
            y_state[sl["top_L"]] = top_l.reshape(-1)
        if "top_V" in sl:
            y_state[sl["top_V"]] = top_v.reshape(-1)
        if "bottom_L" in sl:
            y_state[sl["bottom_L"]] = bottom_l.reshape(-1)
        if "bottom_V" in sl:
            y_state[sl["bottom_V"]] = bottom_v.reshape(-1)
        if n_energy_var > 0:
            tray_el = tray_el_base.copy()
            tray_ev = tray_ev_base.copy()
            idx_e = (
                n_state_var
                + n_flow_var
                + n_boundary_var
                + n_feed_temp_var
                + n_feed_vf_var
                + n_condenser_duty_var
            )
            for phase, stage in energy_blocks:
                delta = float(z[idx_e])
                idx_e += 1
                if phase == "EL":
                    scale = max(abs(float(tray_el_base[stage])), float(args.energy_denom_floor_btu))
                    tray_el[stage] = float(tray_el_base[stage]) + delta * scale
                else:
                    scale = max(abs(float(tray_ev_base[stage])), float(args.energy_denom_floor_btu))
                    tray_ev[stage] = float(tray_ev_base[stage]) + delta * scale
            if "tray_EL_BTU" in sl:
                y_state[sl["tray_EL_BTU"]] = tray_el.reshape(-1)
            if "tray_EV_BTU" in sl:
                y_state[sl["tray_EV_BTU"]] = tray_ev.reshape(-1)
        return y_state

    def apply_flows(z: np.ndarray) -> tuple[np.ndarray, np.ndarray]:
        L_new = L0_lbmolph.copy()
        V_new = V0_lbmolph.copy()
        idx = n_state_var
        for phase, stage in flow_blocks:
            scale = math.exp(float(z[idx]))
            idx += 1
            if phase == "LF":
                L_new[stage] = float(L0_lbmolph[stage]) * scale
            else:
                V_new[stage] = float(V0_lbmolph[stage]) * scale
        object.__setattr__(col, "L_lbmolph", L_new)
        object.__setattr__(col, "V_lbmolph", V_new)
        return L_new, V_new

    def apply_boundary(z: np.ndarray) -> tuple[Any, Dict[str, float]]:
        idx = n_state_var + n_flow_var
        vals = dict(boundary_base)
        for key in boundary_blocks:
            vals[key] = float(boundary_base[key]) * math.exp(float(z[idx]))
            idx += 1
        if reflux_ratio is not None:
            vals["reflux"] = float(reflux_ratio) * float(vals["distillate"])
        b = BoundaryFlows(
            reflux_lbmolph=float(vals["reflux"]),
            boilup_lbmolph=float(vals["boilup"]),
            distillate_lbmolph=float(vals["distillate"]),
            bottoms_lbmolph=float(vals["bottoms"]),
        )
        condenser_duty = condenser_duty_base
        idx_q = n_state_var + n_flow_var + n_boundary_var + n_feed_temp_var + n_feed_vf_var
        if n_condenser_duty_var > 0:
            condenser_duty = float(condenser_duty_base) * (1.0 + float(z[idx_q]))
        inputs_out = replace(
            inputs,
            boundary=b,
            condenser_duty_mode=str(args.condenser_duty_mode),
            condenser_duty_btu_per_h=(
                float(condenser_duty)
                if (n_condenser_duty_var > 0 or args.condenser_duty_btuph is not None)
                else getattr(inputs, "condenser_duty_btu_per_h", None)
            ),
        )
        vals["condenser_duty_btu_per_h"] = float(condenser_duty)
        return inputs_out, vals

    def apply_feed_temperature(z: np.ndarray) -> Optional[float]:
        if feed_stream_base is None:
            return None
        streams = dict(getattr(col, "streams", {}) or {})
        current = streams.get("Feed", feed_stream_base)
        feed_temperature = None
        if n_feed_temp_var > 0 and feed_temperature_base is not None:
            idx = n_state_var + n_flow_var + n_boundary_var
            feed_temperature = float(feed_temperature_base) + float(z[idx])
            current = replace(current, temperature_f=float(feed_temperature))
        feed_vf = None
        if n_feed_vf_var > 0 and feed_vapor_fraction_base is not None:
            idx = n_state_var + n_flow_var + n_boundary_var + n_feed_temp_var
            feed_vf = float(np.clip(float(feed_vapor_fraction_base) + float(z[idx]), 0.0, 1.0))
            current = replace(current, vapor_fraction=float(feed_vf))
        streams["Feed"] = current
        object.__setattr__(col, "streams", streams)
        return float(feed_temperature) if feed_temperature is not None else None

    def current_feed_vf() -> Optional[float]:
        try:
            feed = (getattr(col, "streams", {}) or {}).get("Feed")
            if feed is not None and getattr(feed, "vapor_fraction", None) is not None:
                return float(getattr(feed, "vapor_fraction"))
        except Exception:
            pass
        return None

    def residual(z: np.ndarray) -> np.ndarray:
        nonlocal eval_count, best
        eval_count += 1
        y_state = make_state(z)
        L_eval, V_eval = apply_flows(z)
        feed_temperature_eval = apply_feed_temperature(z)
        inputs_eval, boundary_eval = apply_boundary(z)
        dydt, _diag = column_rhs(0.0, y_state, col, layout, inputs_eval)
        u = layout.unpack(y_state)
        du = layout.unpack(np.asarray(dydt, dtype=float))
        parts: List[np.ndarray] = []
        max_rel = 0.0
        for key in state_residual_blocks:
            if key not in u or key not in du:
                continue
            val = np.asarray(u[key], dtype=float)
            rate = np.asarray(du[key], dtype=float)
            if key in ("tray_L", "tray_V"):
                val = val.reshape((n, nc))[residual_stages, :]
                rate = rate.reshape((n, nc))[residual_stages, :]
            if bool(args.residual_top_boundary_only) and key in ("bottom_L", "bottom_V"):
                continue
            denom = np.abs(val) + float(args.denom_floor_lbmol)
            rel = rate / np.maximum(denom, 1.0e-300)
            max_rel = max(max_rel, float(np.max(np.abs(rel))))
            block_weight = max(float(state_residual_weights.get(key, 1.0)), 0.0)
            parts.append(block_weight * rel.reshape(-1) / max(float(args.residual_scale), 1.0e-300))
        if float(args.tray_total_penalty) > 0.0 and "tray_L" in u and "tray_V" in u and "tray_L" in du and "tray_V" in du:
            tray_l_val = np.asarray(u["tray_L"], dtype=float).reshape((n, nc))
            tray_v_val = np.asarray(u["tray_V"], dtype=float).reshape((n, nc))
            tray_l_rate = np.asarray(du["tray_L"], dtype=float).reshape((n, nc))
            tray_v_rate = np.asarray(du["tray_V"], dtype=float).reshape((n, nc))
            total_inv = np.sum(np.abs(tray_l_val), axis=1) + np.sum(np.abs(tray_v_val), axis=1)
            total_rate = np.sum(tray_l_rate, axis=1) + np.sum(tray_v_rate, axis=1)
            total_rel = total_rate / np.maximum(total_inv + float(args.denom_floor_lbmol), 1.0e-300)
            parts.append(float(args.tray_total_penalty) * total_rel.reshape(-1))
        if float(args.bottom_boundary_balance_weight) > 0.0 and "bottom_L" in u and "bottom_L" in du and "tray_L" in u:
            bottom_rate = np.asarray(du["bottom_L"], dtype=float).reshape((nc,))
            bottom_l_val = np.asarray(u["bottom_L"], dtype=float).reshape((nc,))
            tray_l_val = np.asarray(u["tray_L"], dtype=float).reshape((n, nc))
            x_bottom = _normalize(bottom_l_val, fallback=x_bottom_base)
            x_bottom_tray = _normalize(tray_l_val[-1, :], fallback=x0[-1, :])
            liquid_in = max(float(L_eval[-1]) / 3600.0, 0.0) * x_bottom_tray
            bottoms_out = max(float(boundary_eval.get("bottoms", boundary_base["bottoms"])) / 3600.0, 0.0) * x_bottom
            boilup_out = max(float(boundary_eval.get("boilup", boundary_base["boilup"])) / 3600.0, 0.0) * x_bottom
            flow_denom = np.abs(liquid_in) + np.abs(bottoms_out) + np.abs(boilup_out) + float(args.denom_floor_lbmol) / 3600.0
            parts.append(float(args.bottom_boundary_balance_weight) * bottom_rate / np.maximum(flow_denom, 1.0e-300))
        if float(args.bottom_boundary_total_weight) > 0.0 and "bottom_L" in du:
            bottom_total_rate = float(np.sum(np.asarray(du["bottom_L"], dtype=float).reshape((nc,))))
            liquid_in_total = max(float(L_eval[-1]) / 3600.0, 0.0)
            bottoms_total = max(float(boundary_eval.get("bottoms", boundary_base["bottoms"])) / 3600.0, 0.0)
            boilup_total = max(float(boundary_eval.get("boilup", boundary_base["boilup"])) / 3600.0, 0.0)
            total_denom = liquid_in_total + bottoms_total + boilup_total + float(args.denom_floor_lbmol) / 3600.0
            parts.append(
                np.array(
                    [float(args.bottom_boundary_total_weight) * bottom_total_rate / max(total_denom, 1.0e-300)],
                    dtype=float,
                )
            )
        if float(args.profile_penalty) > 0.0:
            parts.append(float(args.profile_penalty) * np.asarray(z[:n_profile_var], dtype=float).reshape(-1))
        if float(args.boundary_penalty) > 0.0 and n_top_vapor_total_var > 0:
            parts.append(
                float(args.boundary_penalty)
                * np.asarray(z[n_profile_var:n_state_var], dtype=float).reshape(-1)
            )
        if float(args.flow_penalty) > 0.0 and n_flow_var > 0:
            parts.append(float(args.flow_penalty) * np.asarray(z[n_state_var: n_state_var + n_flow_var], dtype=float).reshape(-1))
        if float(args.boundary_penalty) > 0.0 and n_boundary_var > 0:
            parts.append(
                float(args.boundary_penalty)
                * np.asarray(z[n_state_var + n_flow_var: n_state_var + n_flow_var + n_boundary_var], dtype=float).reshape(-1)
            )
        if float(args.feed_temperature_penalty) > 0.0 and n_feed_temp_var > 0:
            idx = n_state_var + n_flow_var + n_boundary_var
            scale = max(abs(float(args.max_feed_temperature_delta_f)), 1.0)
            parts.append(np.array([float(args.feed_temperature_penalty) * float(z[idx]) / scale], dtype=float))
        if float(args.feed_vapor_fraction_penalty) > 0.0 and n_feed_vf_var > 0:
            idx = n_state_var + n_flow_var + n_boundary_var + n_feed_temp_var
            scale = max(abs(float(args.max_feed_vapor_fraction_delta)), 1.0e-6)
            parts.append(np.array([float(args.feed_vapor_fraction_penalty) * float(z[idx]) / scale], dtype=float))
        if float(args.boundary_penalty) > 0.0 and n_condenser_duty_var > 0:
            idx = n_state_var + n_flow_var + n_boundary_var + n_feed_temp_var + n_feed_vf_var
            scale = max(abs(float(args.max_condenser_duty_rel_delta)), 1.0e-6)
            parts.append(np.array([float(args.boundary_penalty) * float(z[idx]) / scale], dtype=float))
        if bool(args.include_energy):
            for key in energy_residual_blocks:
                if key not in u or key not in du:
                    continue
                val = np.asarray(u[key], dtype=float)
                rate = np.asarray(du[key], dtype=float)
                if key in ("tray_EL_BTU", "tray_EV_BTU"):
                    val = val.reshape((n,))[residual_stages]
                    rate = rate.reshape((n,))[residual_stages]
                denom = np.abs(val) + float(args.energy_denom_floor_btu)
                rel = rate / np.maximum(denom, 1.0e-300)
                max_rel = max(max_rel, float(np.max(np.abs(rel))))
                parts.append(rel.reshape(-1) / max(float(args.energy_residual_scale), 1.0e-300))
        if float(args.energy_penalty) > 0.0 and n_energy_var > 0:
            start = (
                n_state_var
                + n_flow_var
                + n_boundary_var
                + n_feed_temp_var
                + n_feed_vf_var
                + n_condenser_duty_var
            )
            parts.append(float(args.energy_penalty) * np.asarray(z[start: start + n_energy_var], dtype=float).reshape(-1))
        r = np.concatenate(parts) if parts else np.zeros(0, dtype=float)
        norm = float(np.linalg.norm(r))
        if norm < float(best["norm"]):
            best = {"norm": norm, "z": np.asarray(z, dtype=float).copy(), "max_rel": max_rel}
        if eval_count == 1 or eval_count % 5 == 0:
            print(f"eval={eval_count} norm={norm:.8g} max_rel={max_rel:.8g} best_max_rel={best['max_rel']:.8g}")
        return r

    result = least_squares(
        residual,
        z0,
        bounds=(lb, ub),
        max_nfev=max(int(args.max_nfev), 1),
        verbose=1,
        x_scale="jac",
    )
    z_best = np.asarray(best["z"], dtype=float)
    y_opt = make_state(z_best)
    L_opt_lbmolph, V_opt_lbmolph = apply_flows(z_best)
    feed_temperature_opt = apply_feed_temperature(z_best)
    feed_vapor_fraction_opt = current_feed_vf()
    _inputs_opt, boundary_opt = apply_boundary(z_best)
    u_opt = layout.unpack(y_opt)
    x_opt = np.asarray(u_opt["x_tray"], dtype=float).reshape((n, nc))
    yv_opt = np.asarray(u_opt["y_tray"], dtype=float).reshape((n, nc))
    top_l_opt = np.asarray(u_opt.get("top_L", top_l_base), dtype=float).reshape((nc,))
    top_v_opt = np.asarray(u_opt.get("top_V", top_v_base), dtype=float).reshape((nc,))
    bottom_l_opt = np.asarray(u_opt.get("bottom_L", np.zeros(nc)), dtype=float).reshape((nc,))
    bottom_v_opt = np.asarray(u_opt.get("bottom_V", np.zeros(nc)), dtype=float).reshape((nc,))
    tray_el_opt = np.asarray(u_opt.get("tray_EL_BTU", tray_el_base), dtype=float).reshape((n,))
    tray_ev_opt = np.asarray(u_opt.get("tray_EV_BTU", tray_ev_base), dtype=float).reshape((n,))
    for i in range(n):
        x_opt[i, :] = _normalize(x_opt[i, :], fallback=x0[i, :])
        if mv[i] <= 1.0e-12:
            yv_opt[i, :] = yv0[i, :]
        yv_opt[i, :] = _normalize(yv_opt[i, :], fallback=yv0[i, :])

    in_path = _resolve(args.input)
    out_path = _resolve(args.output)
    if in_path.resolve() != out_path.resolve():
        out_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(in_path, out_path)
    wb = load_workbook(out_path)
    ws = wb["Initial Conditions"]
    stage_hdr_row, stage_col = _find_header(ws, "Stage")
    x_cols = _component_columns(ws, "Liquid Composition", nc)
    y_cols = _component_columns(ws, "Vapor Composition", nc)
    _row, vapor_flow_col = _find_header(ws, "Vapor Flow (lbmol/h)")
    _row, liquid_flow_col = _find_header(ws, "Liquid Flow (lbmol/h)")
    stage_to_row: Dict[int, int] = {}
    for row in range(stage_hdr_row + 1, ws.max_row + 1):
        val = ws.cell(row, stage_col).value
        if val is None:
            continue
        try:
            stage_to_row[int(float(val))] = int(row)
        except Exception:
            pass
    for i in range(n):
        row = stage_to_row[i + 1]
        ws.cell(row, liquid_flow_col).value = float(L_opt_lbmolph[i])
        ws.cell(row, vapor_flow_col).value = float(V_opt_lbmolph[i])
        for k, col_idx in enumerate(x_cols):
            ws.cell(row, col_idx).value = float(x_opt[i, k])
        for k, col_idx in enumerate(y_cols):
            ws.cell(row, col_idx).value = float(yv_opt[i, k])
    _update_product_stream_compositions(wb, col, x_opt[0, :], x_opt[-1, :])
    if "reflux" in boundary_opt and n > 0:
        first_row = stage_to_row.get(1)
        if first_row is not None:
            ws.cell(first_row, liquid_flow_col).value = float(boundary_opt["reflux"])
    if "boilup" in boundary_opt and n > 0:
        last_row = stage_to_row.get(n)
        if last_row is not None:
            ws.cell(last_row, vapor_flow_col).value = float(boundary_opt["boilup"])
    if "distillate" in boundary_opt:
        _update_stream_total(wb, "distillate", float(boundary_opt["distillate"]))
    if "bottoms" in boundary_opt:
        _update_stream_total(wb, "bottom", float(boundary_opt["bottoms"]))
    if feed_temperature_opt is not None and "Streams" in wb.sheetnames:
        ws_streams = wb["Streams"]
        feed_col = None
        temp_row = None
        for c in range(2, ws_streams.max_column + 1):
            if "feed" in _norm_label(ws_streams.cell(1, c).value):
                feed_col = int(c)
                break
        for r in range(1, ws_streams.max_row + 1):
            if "temperaturef" in _norm_label(ws_streams.cell(r, 1).value):
                temp_row = int(r)
                break
        if feed_col is not None and temp_row is not None:
            ws_streams.cell(temp_row, feed_col).value = float(feed_temperature_opt)
    if feed_vapor_fraction_opt is not None and "Streams" in wb.sheetnames:
        ws_streams = wb["Streams"]
        feed_col = None
        vf_row = None
        for c in range(2, ws_streams.max_column + 1):
            if "feed" in _norm_label(ws_streams.cell(1, c).value):
                feed_col = int(c)
                break
        for r in range(1, ws_streams.max_row + 1):
            label = _norm_label(ws_streams.cell(r, 1).value)
            if "vapourfraction" in label or "vaporfraction" in label:
                vf_row = int(r)
                break
        if feed_col is not None and vf_row is not None:
            ws_streams.cell(vf_row, feed_col).value = float(feed_vapor_fraction_opt)
    if n_energy_var > 0 and "Energy State" in wb.sheetnames:
        ws_energy = wb["Energy State"]
        energy_stage_hdr_row, energy_stage_col = _find_header(ws_energy, "Stage")
        _row, tray_el_col = _find_header(ws_energy, "Tray EL (BTU)")
        _row, tray_ev_col = _find_header(ws_energy, "Tray EV (BTU)")
        energy_stage_to_row: Dict[int, int] = {}
        for row in range(energy_stage_hdr_row + 1, ws_energy.max_row + 1):
            val = ws_energy.cell(row, energy_stage_col).value
            if val is None:
                continue
            try:
                energy_stage_to_row[int(float(val))] = int(row)
            except Exception:
                pass
        for i in range(n):
            row = energy_stage_to_row.get(i + 1)
            if row is None:
                continue
            ws_energy.cell(row, tray_el_col).value = float(tray_el_opt[i])
            ws_energy.cell(row, tray_ev_col).value = float(tray_ev_opt[i])
    if (
        bool(args.vary_top_liquid)
        or bool(args.vary_top_vapor)
        or bool(args.vary_bottom_liquid)
        or bool(args.vary_bottom_vapor)
        or bool(args.vary_top_vapor_total)
    ):
        _write_boundary_state(wb, col, top_l_opt, top_v_opt, bottom_l_opt, bottom_v_opt)
    if n_condenser_duty_var > 0 or args.condenser_duty_btuph is not None:
        _update_spec_value(wb, "Condenser Duty Mode", str(args.condenser_duty_mode))
        _update_spec_value(wb, "Condenser Duty (Btu/h)", float(boundary_opt["condenser_duty_btu_per_h"]))
    wb.save(out_path)

    audit_dir = _resolve(args.audit_output_dir) if args.audit_output_dir else out_path.with_suffix("")
    audit_cmd = [
        sys.executable,
        str(PROJECT_ROOT / "tools" / "column_initialization_residual_audit.py"),
        "--excel",
        str(out_path),
        "--thermo",
        str(args.thermo),
        "--clapeyron-model",
        str(args.clapeyron_model),
        "--runtime-mode",
        str(args.runtime_mode),
        "--condenser-duty-mode",
        str(args.condenser_duty_mode),
        "--vapor-holdup-relaxation-sec",
        str(float(args.vapor_holdup_relaxation_sec)),
        "--output-dir",
        str(audit_dir),
    ]
    if bool(args.no_equilibrium):
        audit_cmd.append("--no-equilibrium")
    if bool(args.no_flash_feed_at_stage_conditions):
        audit_cmd.append("--no-flash-feed-at-stage-conditions")
    if bool(args.include_energy):
        audit_cmd.append("--include-energy")
    if not bool(args.include_temperature):
        audit_cmd.append("--no-temperature")
    if bool(args.use_excel_vapor_holdup):
        audit_cmd.append("--use-excel-vapor-holdup")
    if n_condenser_duty_var > 0 or args.condenser_duty_btuph is not None:
        audit_cmd.extend(["--condenser-duty-btuph", str(float(boundary_opt["condenser_duty_btu_per_h"]))])
    subprocess.run(audit_cmd, cwd=str(PROJECT_ROOT), check=True)

    summary = {
        "input": str(in_path),
        "output": str(out_path),
        "stages": [i + 1 for i in stages],
        "residual_stages": [i + 1 for i in residual_stages],
        "state_residual_blocks": state_residual_blocks,
        "state_residual_weights": state_residual_weights,
        "bottom_boundary_balance_weight": float(args.bottom_boundary_balance_weight),
        "bottom_boundary_total_weight": float(args.bottom_boundary_total_weight),
        "energy_residual_blocks": energy_residual_blocks,
        "blocks": blocks,
        "top_comp_blocks": top_comp_blocks,
        "bottom_comp_blocks": bottom_comp_blocks,
        "flow_blocks": flow_blocks,
        "boundary_blocks": boundary_blocks,
        "energy_blocks": energy_blocks,
        "boundary_base_lbmolph": boundary_base,
        "boundary_opt_lbmolph": boundary_opt,
        "chemsep_product_specs": bool(args.chemsep_product_specs),
        "reflux_ratio": reflux_ratio,
        "n_variables": nvar,
        "n_composition_variables": n_comp_var,
        "n_top_composition_variables": n_top_comp_var,
        "n_bottom_composition_variables": n_bottom_comp_var,
        "n_top_vapor_total_variables": n_top_vapor_total_var,
        "n_flow_variables": n_flow_var,
        "n_boundary_variables": n_boundary_var,
        "n_feed_temperature_variables": n_feed_temp_var,
        "n_feed_vapor_fraction_variables": n_feed_vf_var,
        "n_condenser_duty_variables": n_condenser_duty_var,
        "n_energy_variables": n_energy_var,
        "condenser_duty_mode": str(args.condenser_duty_mode),
        "condenser_duty_base_BTUph": float(condenser_duty_base),
        "condenser_duty_opt_BTUph": float(boundary_opt["condenser_duty_btu_per_h"]),
        "rel_delta_condenser_duty": float(
            abs(float(boundary_opt["condenser_duty_btu_per_h"]) - float(condenser_duty_base))
            / max(abs(float(condenser_duty_base)), 1.0)
        ),
        "eval_count": eval_count,
        "least_squares_success": bool(result.success),
        "least_squares_message": str(result.message),
        "best_objective_norm": float(best["norm"]),
        "best_max_relative_rate_per_s_inprocess": float(best["max_rel"]),
        "max_abs_delta_x": float(np.max(np.abs(x_opt - x0))),
        "max_abs_delta_y": float(np.max(np.abs(yv_opt - yv0))),
        "max_abs_delta_top_L_x": float(
            np.max(np.abs(_normalize(top_l_opt, fallback=x_top_base) - x_top_base))
        ),
        "max_abs_delta_top_V_y": float(
            np.max(np.abs(_normalize(top_v_opt, fallback=y_top_base) - y_top_base))
        ),
        "max_abs_delta_bottom_L_x": float(
            np.max(np.abs(_normalize(bottom_l_opt, fallback=x_bottom_base) - x_bottom_base))
        ),
        "max_abs_delta_bottom_V_y": float(
            np.max(np.abs(_normalize(bottom_v_opt, fallback=y_bottom_base) - y_bottom_base))
        ),
        "rel_delta_top_V_total": float(
            abs(float(np.sum(top_v_opt)) - float(top_v_total_base)) / max(abs(float(top_v_total_base)), 1.0)
        ),
        "max_rel_delta_L_flow": float(np.max(np.abs(L_opt_lbmolph - L0_lbmolph) / np.maximum(np.abs(L0_lbmolph), 1.0))),
        "max_rel_delta_V_flow": float(np.max(np.abs(V_opt_lbmolph - V0_lbmolph) / np.maximum(np.abs(V0_lbmolph), 1.0))),
        "max_rel_delta_boundary_flow": float(
            max(
                [
                    abs(float(boundary_opt[k]) - float(boundary_base[k])) / max(abs(float(boundary_base[k])), 1.0)
                    for k in boundary_blocks
                ]
                or [0.0]
            )
        ),
        "max_rel_delta_tray_EL": float(
            np.max(np.abs(tray_el_opt - tray_el_base) / np.maximum(np.abs(tray_el_base), float(args.energy_denom_floor_btu)))
        ) if n_energy_var > 0 else 0.0,
        "max_rel_delta_tray_EV": float(
            np.max(np.abs(tray_ev_opt - tray_ev_base) / np.maximum(np.abs(tray_ev_base), float(args.energy_denom_floor_btu)))
        ) if n_energy_var > 0 else 0.0,
        "feed_temperature_base_F": feed_temperature_base,
        "feed_temperature_opt_F": feed_temperature_opt,
        "feed_vapor_fraction_base": feed_vapor_fraction_base,
        "feed_vapor_fraction_opt": feed_vapor_fraction_opt,
        "audit_dir": str(audit_dir),
    }
    summary_path = out_path.with_suffix(".optimizer_summary.json")
    summary_path.write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(f"Wrote: {out_path}")
    print(f"Summary: {summary_path}")
    print(f"Audit: {audit_dir}")
    try:
        provider = extra.get("provider")
        if provider is not None and hasattr(provider, "close"):
            provider.close()
    except Exception:
        pass
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
