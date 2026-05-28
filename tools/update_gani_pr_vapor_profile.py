#!/usr/bin/env python
"""
Update the Gani debutanizer workbook's seeded vapor profile from Clapeyron PR.

The Gani source does not publish a tray profile.  The workbook therefore uses a
constructed liquid profile between the published product compositions.  This
tool keeps that liquid profile, T/P profile, flows, and holdups intact, but
replaces each tray vapor composition with the vapor composition returned by a
live PR flash at the tray liquid composition, temperature, and pressure.
"""

from __future__ import annotations

import argparse
import datetime as dt
from pathlib import Path
import shutil
import sys
from typing import Sequence

import numpy as np
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_PATH = PROJECT_ROOT / "src"
if str(SRC_PATH) not in sys.path:
    sys.path.insert(0, str(SRC_PATH))

from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case  # noqa: E402
from dynamic_distillation.dynamic_run_scaffold_v1 import RunnerConfig, build_inputs_for_runner  # noqa: E402
from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel  # noqa: E402


def _tag() -> str:
    return dt.datetime.now().strftime("%Y%m%d_%H%M%S")


def _resolve(raw: str) -> Path:
    p = Path(str(raw))
    if p.is_absolute():
        return p.resolve()
    return (PROJECT_ROOT / p).resolve()


def _normalize(values: Sequence[float]) -> np.ndarray:
    arr = np.asarray(values, dtype=float).reshape((-1,))
    arr = np.where(np.isfinite(arr) & (arr > 0.0), arr, 0.0)
    total = float(np.sum(arr))
    if total <= 0.0:
        raise ValueError("composition total must be positive")
    return arr / total


def _append_note(wb, field: str, value: str) -> None:
    if "Notes" not in wb.sheetnames:
        ws = wb.create_sheet("Notes")
        ws.cell(1, 1).value = "Field"
        ws.cell(1, 2).value = "Value"
    ws = wb["Notes"]
    row = int(ws.max_row) + 1
    ws.cell(row, 1).value = field
    ws.cell(row, 2).value = value


def main() -> int:
    ap = argparse.ArgumentParser(description="Replace Gani workbook seeded vapor y with PR flash vapor y.")
    ap.add_argument("--excel", default="validation_gani_1986_debutanizer.xlsx")
    ap.add_argument("--clapeyron-model", default="PR")
    ap.add_argument("--backup-dir", default="logs")
    ap.add_argument("--no-backup", action="store_true")
    args = ap.parse_args()

    excel_path = _resolve(args.excel)
    if not excel_path.exists():
        raise FileNotFoundError(excel_path)

    backup_path = None
    if not bool(args.no_backup):
        backup_dir = _resolve(args.backup_dir)
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup_path = backup_dir / f"{excel_path.stem}__before_pr_vapor_profile_{_tag()}{excel_path.suffix}"
        shutil.copy2(excel_path, backup_path)

    case = load_case_from_excel(str(excel_path))
    col = build_column_spec_from_case(case)
    cfg = RunnerConfig(
        excel_path=str(excel_path),
        thermo_mode="clapeyron",
        clapeyron_model=str(args.clapeyron_model),
        runtime_mode="parity",
        include_temperature=True,
        include_energy=True,
        write_logs=False,
    )
    _inputs, provider = build_inputs_for_runner(case, col, cfg)

    wb = load_workbook(excel_path)
    if "Initial Conditions" not in wb.sheetnames:
        raise KeyError("Initial Conditions sheet not found")
    ws = wb["Initial Conditions"]

    n_components = int(col.n_components)
    vapor_start_col = 8
    liquid_start_col = vapor_start_col + n_components
    max_abs_delta = 0.0
    max_delta_stage = 0
    max_delta_component = 0

    for stage in range(1, int(col.n_stages) + 1):
        row = stage + 1
        t_f = float(ws.cell(row, 2).value)
        p_psia = float(ws.cell(row, 3).value)
        x = _normalize([ws.cell(row, liquid_start_col + k).value for k in range(n_components)])
        old_y = _normalize([ws.cell(row, vapor_start_col + k).value for k in range(n_components)])
        flash = provider.flash_TP_full(float(t_f), float(p_psia), x.tolist())
        new_y = _normalize(flash.y)
        delta = np.abs(new_y - old_y)
        kmax = int(np.argmax(delta))
        if float(delta[kmax]) > max_abs_delta:
            max_abs_delta = float(delta[kmax])
            max_delta_stage = int(stage)
            max_delta_component = int(kmax + 1)
        for k in range(n_components):
            ws.cell(row, vapor_start_col + k).value = float(new_y[k])

    _append_note(
        wb,
        "Seeded vapor profile update",
        (
            f"{_tag()}: Vapor composition profile replaced with Clapeyron "
            f"{args.clapeyron_model} flash vapor y at each seeded liquid x/T/P. "
            "Liquid composition, T/P, flows, holdups, and source endpoints were preserved."
        ),
    )
    _append_note(
        wb,
        "Seeded vapor profile max change",
        f"max |delta y|={max_abs_delta:.6g} at stage {max_delta_stage}, component {max_delta_component}.",
    )

    wb.save(excel_path)
    print(f"updated: {excel_path}")
    if backup_path is not None:
        print(f"backup:  {backup_path}")
    print(
        "max_abs_delta_y="
        f"{max_abs_delta:.6g} stage={max_delta_stage} component={max_delta_component}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
