"""
dynamic_run_scaffold_v1.py

Dynamic Distillation - Simulation Runner and CLI

PURPOSE
-------
Execute end-to-end dynamic column runs: case load/build/validation,
state initialization/startup conditioning, controller updates, time
integration (explicit/stiff-implicit), diagnostics/logging, and
experiment-ledger registration.

INPUTS
------
RunnerConfig (main runtime configuration), including:
- case/log settings: excel_path, n_steps, dt_sec, log_every_n_steps, logs_dir
- model toggles: include_temperature, include_energy, equilibrium relaxation
- thermo settings: thermo_mode (stub/dwsim/table/table-pool), refresh cadence,
  optional refresh thresholds, thermo table/cache/pool settings
- boundary/control overrides: reflux/boilup, duty modes, level/pressure/
  composition controller options, PSV options, top-drum volume settings

OUTPUTS
-------
run_smoke_simulation(cfg) returns a dict with run artifacts and runtime state,
including summary/profile CSV paths, validation status, final state/time,
last diagnostics, and startup initialization diagnostics.

CLI (`python -m dynamic_distillation.dynamic_run_scaffold_v1`) parses
arguments into RunnerConfig and writes CSV outputs when logging is enabled.

KEY DEPENDENCIES
----------------
- excel_case_loader_v1 / column_spec_builder_v1 / excel_case_validator_v1
- state_vector_layout_v1 / column_rhs_v1
- thermo_provider_v1 / thermo_surrogate_v1 / thermo_table_pool_v1
- experiment_ledger_v1 (duplicate-command guard + ledger refresh)

ASSUMPTIONS & CONSTRAINTS
-------------------------
- Default integrator is explicit Euler; optional stiff solvers are available
  when SciPy is installed.
- Excel case must be loadable and pass blocking validation checks.
- Thermo mode and required assets must be consistent (e.g., table path).
- This runner prioritizes diagnostic transparency and workflow reproducibility.

NOTES / CURRENT BEHAVIOR
------------------------
- Supports startup thermo-consistent conditioning and top-drum startup steadying.
- Supports pressure, level, distillate-composition, and bottoms-composition PI loops.
- Supports top-drum PSV vent diagnostics and mass-closure diagnostics.
- Auto-registers logged runs and rebuilds experiment ledger artifacts.
"""

from __future__ import annotations

from collections import deque
from contextlib import nullcontext
from dataclasses import dataclass, replace
from pathlib import Path
from types import SimpleNamespace
from typing import Any, Dict, List, Optional, Sequence, Tuple

import csv
import datetime as _dt
import json
import os
import sys
import time
import traceback

import numpy as np
try:
    from scipy.integrate import solve_ivp as _solve_ivp
except Exception:  # pragma: no cover - optional dependency
    _solve_ivp = None

from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel, CaseData
from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case, ColumnSpec
from dynamic_distillation.excel_case_validator_v1 import validate_loaded_case, print_validation_report
from dynamic_distillation.state_vector_layout_v1 import StateVectorLayout
from dynamic_distillation.column_rhs_v1 import (
    BottomSumpCpPacket,
    BoundaryFlows,
    CondenserDutyPacket,
    ColumnInputs,
    FeedStageFlashPacket,
    TrayThermoPacket,
    VolumeModel,
    _compatible_feed_stage_flash_packet,
    _component_molar_flows_vector_lbmolps,
    _compute_top_drum_pressure_psia,
    _bubble_point_T_F,
    column_rhs,
)
from dynamic_distillation.stage_hydraulics_francis_v1 import compute_francis_weir_liquid_outflow
from dynamic_distillation.dae_pilot_v1 import (
    default_algebraic_seed,
    evaluate_pilot_residual,
    finite_difference_jacobian,
    inf_norm,
)
from dynamic_distillation.experiment_ledger_v1 import (
    append_run_registry_entry,
    compose_cli_command,
    compose_cli_command_identity,
    find_exact_command_matches,
    rebuild_experiment_ledger,
)
from dynamic_distillation.thermo_backend_factory_v1 import (
    build_equilibrium_relaxation_pr_provider,
    build_primary_thermo_backend,
)
from dynamic_distillation.thermo_step_coordinator_v1 import refresh_tray_tp_packet


# -------------------------
# Small helpers
# -------------------------


def _timestamp_tag() -> str:
    return _dt.datetime.now().strftime("%Y%m%d_%H%M%S")


def _ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


_PROGRESS_TRACE_PATH: Optional[Path] = None
_PROGRESS_TRACE_T0: Optional[float] = None


def _configure_progress_streams() -> None:
    """Best-effort line buffering for detached/redirected CLI runs."""
    for stream in (getattr(sys, "stdout", None), getattr(sys, "stderr", None)):
        if stream is None:
            continue
        reconfigure = getattr(stream, "reconfigure", None)
        if not callable(reconfigure):
            continue
        try:
            reconfigure(line_buffering=True, write_through=True)
        except TypeError:
            try:
                reconfigure(line_buffering=True)
            except Exception:
                pass
        except Exception:
            pass


def _set_progress_trace_path(path: Optional[Path]) -> None:
    global _PROGRESS_TRACE_PATH, _PROGRESS_TRACE_T0
    _PROGRESS_TRACE_PATH = path
    _PROGRESS_TRACE_T0 = time.perf_counter()
    if path is None:
        return
    try:
        _ensure_dir(path.parent)
        path.write_text("", encoding="utf-8")
    except Exception:
        pass


def _emit_progress(message: str, *, echo: bool = True) -> None:
    text = str(message)
    if echo:
        try:
            print(text, flush=True)
        except Exception:
            pass
    path = _PROGRESS_TRACE_PATH
    if path is None:
        return
    wall = np.nan
    try:
        if _PROGRESS_TRACE_T0 is not None:
            wall = float(time.perf_counter() - float(_PROGRESS_TRACE_T0))
    except Exception:
        wall = np.nan
    stamp = _dt.datetime.now().isoformat(timespec="seconds")
    prefix = f"[{stamp}]"
    if np.isfinite(wall):
        prefix += f"[wall={wall:8.2f}s]"
    try:
        with path.open("a", encoding="utf-8") as fh:
            fh.write(f"{prefix} {text}\n")
    except Exception:
        pass


def _clean_optional_text(value: Any) -> Optional[str]:
    txt = str(value or "").strip()
    return txt if txt else None


def _write_json_atomic(path: Path, payload: Dict[str, Any]) -> None:
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    tmp_path.write_text(json.dumps(payload, indent=2, ensure_ascii=True), encoding="utf-8")
    tmp_path.replace(path)


def _read_json_if_exists(path: Path) -> Dict[str, Any]:
    try:
        if not path.exists():
            return {}
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _normalize_comp(z: np.ndarray) -> np.ndarray:
    z = np.asarray(z, dtype=float).reshape((-1,))
    s = float(np.sum(z))
    if not np.isfinite(s) or s <= 0.0:
        n = z.size
        return np.full(n, 1.0 / max(n, 1), dtype=float)
    return z / s


def _get_first_mapping_value(d: Dict[str, Any], keys: Sequence[str]) -> Optional[Any]:
    for k in keys:
        if k in d:
            return d[k]
    return None


def _as_float(x: Any) -> Optional[float]:
    try:
        v = float(x)
        return v
    except Exception:
        return None


def _as_int(x: Any) -> Optional[int]:
    try:
        return int(float(x))
    except Exception:
        return None


def _json_optional_array(value: Any, *, shape: Optional[Tuple[int, ...]] = None) -> Optional[Any]:
    if value is None:
        return None
    arr = np.asarray(value, dtype=float)
    if shape is not None:
        arr = arr.reshape(shape)
    return arr.tolist()


def _array_from_json(value: Any, *, shape: Optional[Tuple[int, ...]] = None) -> Optional[np.ndarray]:
    if value is None:
        return None
    arr = np.asarray(value, dtype=float)
    if shape is not None:
        arr = arr.reshape(shape)
    return arr.copy()


def _snapshot_thermo_call_counters(*providers: Any) -> Dict[str, Dict[str, Any]]:
    merged: Dict[str, Dict[str, Any]] = {}
    seen_provider_ids: set[int] = set()
    for provider in providers:
        if provider is None:
            continue
        try:
            provider_id = id(provider)
        except Exception:
            provider_id = -1
        if provider_id in seen_provider_ids:
            continue
        seen_provider_ids.add(provider_id)
        getter = getattr(provider, "get_call_counters", None)
        if not callable(getter):
            continue
        try:
            counters = getter()
        except Exception:
            continue
        if not isinstance(counters, dict):
            continue
        for category, metrics in counters.items():
            cat = str(category).strip() or "uncategorized"
            if cat not in merged:
                merged[cat] = {}
            if not isinstance(metrics, dict):
                continue
            for metric, value in metrics.items():
                name = str(metric).strip()
                if not name:
                    continue
                if name.endswith("_sec"):
                    try:
                        amt = float(value)
                    except Exception:
                        continue
                    merged[cat][name] = float(merged[cat].get(name, 0.0)) + float(amt)
                else:
                    try:
                        amt = int(value)
                    except Exception:
                        continue
                    merged[cat][name] = int(merged[cat].get(name, 0)) + int(amt)
    return merged


def _tray_thermo_packet_from_diag(
    diag: Dict[str, Any],
    *,
    n_stages: int,
    n_components: int,
    T_tray_F: Optional[np.ndarray] = None,
    P_tray_psia: Optional[np.ndarray] = None,
) -> Optional[TrayThermoPacket]:
    required = ("z_overall_tray", "K_tray", "HL_BTU_lbmol_tray", "HV_BTU_lbmol_tray", "Z_tray")
    if not all(key in diag for key in required):
        return None
    try:
        return TrayThermoPacket(
            z_overall_tray=np.asarray(diag["z_overall_tray"], dtype=float).reshape((n_stages, n_components)).copy(),
            K_tray=np.asarray(diag["K_tray"], dtype=float).reshape((n_stages, n_components)).copy(),
            HL_BTU_lbmol_tray=np.asarray(diag["HL_BTU_lbmol_tray"], dtype=float).reshape((n_stages,)).copy(),
            HV_BTU_lbmol_tray=np.asarray(diag["HV_BTU_lbmol_tray"], dtype=float).reshape((n_stages,)).copy(),
            Z_tray=np.asarray(diag["Z_tray"], dtype=float).reshape((n_stages,)).copy(),
            cpL_BTU_lbmolF_tray=(
                np.asarray(diag["cpL_BTU_lbmolF_tray"], dtype=float).reshape((n_stages,)).copy()
                if "cpL_BTU_lbmolF_tray" in diag
                else None
            ),
            cpV_BTU_lbmolF_tray=(
                np.asarray(diag["cpV_BTU_lbmolF_tray"], dtype=float).reshape((n_stages,)).copy()
                if "cpV_BTU_lbmolF_tray" in diag
                else None
            ),
            T_tray_F=(
                None
                if T_tray_F is None
                else np.asarray(T_tray_F, dtype=float).reshape((n_stages,)).copy()
            ),
            P_tray_psia=(
                None
                if P_tray_psia is None
                else np.asarray(P_tray_psia, dtype=float).reshape((n_stages,)).copy()
            ),
            x_equilibrium_tray=(
                np.asarray(diag["x_eq_thermo_tray"], dtype=float).reshape((n_stages, n_components)).copy()
                if "x_eq_thermo_tray" in diag
                else None
            ),
            y_equilibrium_tray=(
                np.asarray(diag["y_eq_thermo_tray"], dtype=float).reshape((n_stages, n_components)).copy()
                if "y_eq_thermo_tray" in diag
                else None
            ),
        )
    except Exception:
        return None


def _resolve_step0_startup_packet_reuse_thresholds(
    *,
    startup_seed_loaded: bool,
    runtime_mode: str,
    step: int,
    last_tray_thermo_packet: Optional[TrayThermoPacket],
    last_T_tray: Optional[np.ndarray],
    last_P_hyd: Optional[np.ndarray],
    last_P_diag: Optional[np.ndarray],
    last_z_overall: Optional[np.ndarray],
    base_inputs: ColumnInputs,
) -> tuple[bool, Optional[float], Optional[float], Optional[float]]:
    startup_packet_mainflash_reuse = bool(
        str(runtime_mode).strip().lower() == "hydraulic"
        and int(step) == 0
        and last_tray_thermo_packet is not None
        and last_T_tray is not None
        and (last_P_hyd is not None or last_P_diag is not None)
        and last_z_overall is not None
    )

    thermo_refresh_dT_step = base_inputs.thermo_refresh_dT_F
    thermo_refresh_dP_step = base_inputs.thermo_refresh_dP_psia
    thermo_refresh_dx_step = base_inputs.thermo_refresh_dx
    if startup_packet_mainflash_reuse:
        default_dT = 0.5 if startup_seed_loaded else 1.0e-3
        default_dP = 5.0 if startup_seed_loaded else 1.0e-3
        default_dx = 1.0e-5 if startup_seed_loaded else 1.0e-6
        if thermo_refresh_dT_step is None:
            thermo_refresh_dT_step = float(default_dT)
        if thermo_refresh_dP_step is None:
            thermo_refresh_dP_step = float(default_dP)
        if thermo_refresh_dx_step is None:
            thermo_refresh_dx_step = float(default_dx)

    return (
        bool(startup_packet_mainflash_reuse),
        (None if thermo_refresh_dT_step is None else float(thermo_refresh_dT_step)),
        (None if thermo_refresh_dP_step is None else float(thermo_refresh_dP_step)),
        (None if thermo_refresh_dx_step is None else float(thermo_refresh_dx_step)),
    )


def _resolve_step0_startup_packet_phase_reuse_settings(
    *,
    startup_seed_loaded: bool,
    runtime_mode: str,
    step: int,
    last_tray_thermo_packet: Optional[TrayThermoPacket],
    last_T_tray: Optional[np.ndarray],
    last_P_hyd: Optional[np.ndarray],
    last_P_diag: Optional[np.ndarray],
    last_z_overall: Optional[np.ndarray],
    base_inputs: ColumnInputs,
) -> tuple[Optional[float], Optional[float], Optional[float], Optional[float]]:
    phase_dx = float(getattr(base_inputs, "thermo_packet_phase_reuse_dx", 0.0) or 0.0)
    vapor_dx_raw = getattr(base_inputs, "thermo_packet_vapor_reuse_dx", None)
    vapor_dx = None if vapor_dx_raw is None else float(vapor_dx_raw)
    phase_dT = float(getattr(base_inputs, "thermo_packet_phase_reuse_dT_F", 0.0) or 0.0)
    phase_dP = float(getattr(base_inputs, "thermo_packet_phase_reuse_dP_psia", 0.0) or 0.0)
    startup_packet_mainflash_reuse = bool(
        startup_seed_loaded
        and str(runtime_mode).strip().lower() == "hydraulic"
        and int(step) == 0
        and last_tray_thermo_packet is not None
        and last_T_tray is not None
        and (last_P_hyd is not None or last_P_diag is not None)
        and last_z_overall is not None
    )
    if startup_packet_mainflash_reuse:
        vapor_floor = 0.25
        if vapor_dx is None or (np.isfinite(vapor_dx) and vapor_dx < vapor_floor):
            vapor_dx = float(vapor_floor)
    return float(phase_dx), (None if vapor_dx is None else float(vapor_dx)), float(phase_dT), float(phase_dP)


def _resolve_runtime_thermo_execution_plan(
    *,
    step: int,
    thermo_every: int,
    col: ColumnSpec,
    layout: StateVectorLayout,
    y: np.ndarray,
    include_temperature: bool,
    pressure_model: str,
    base_inputs: ColumnInputs,
    last_T_tray: Optional[np.ndarray],
    last_P_hyd: Optional[np.ndarray],
    last_P_diag: Optional[np.ndarray],
    last_Zfac: Optional[np.ndarray],
    last_z_overall: Optional[np.ndarray],
    last_K_tray: Optional[np.ndarray],
) -> tuple[bool, str]:
    cadence_n = max(int(thermo_every), 1)
    do_thermo = (int(step) % cadence_n) == 0
    if do_thermo:
        return True, "cadence"

    refresh_by_state = (
        (base_inputs.thermo_refresh_dT_F is not None)
        or (base_inputs.thermo_refresh_dP_psia is not None)
        or (base_inputs.thermo_refresh_dx is not None)
    )
    if refresh_by_state:
        thermo_event = False
        event_reason = "hold"
        missing_history = False
        try:
            N = int(col.n_stages)
            Nc = int(col.n_components)
            u_probe = layout.unpack(y)
            T_now = _tray_temperature_F(
                col,
                layout,
                y,
                include_temperature=bool(include_temperature),
            )

            if base_inputs.thermo_refresh_dT_F is not None:
                if last_T_tray is None:
                    missing_history = True
                else:
                    dT_max = float(
                        np.nanmax(
                            np.abs(np.asarray(T_now, dtype=float) - np.asarray(last_T_tray, dtype=float))
                        )
                    )
                    if np.isfinite(dT_max) and dT_max >= float(base_inputs.thermo_refresh_dT_F):
                        thermo_event = True
                        event_reason = "state_dT"

            if (not thermo_event) and base_inputs.thermo_refresh_dP_psia is not None:
                P_prev = last_P_hyd if last_P_hyd is not None else last_P_diag
                if P_prev is None:
                    missing_history = True
                else:
                    MV_now = np.asarray(u_probe["MV_tot_tray"], dtype=float).reshape((N,))
                    Z_for_p = (
                        np.asarray(last_Zfac, dtype=float).reshape((N,))
                        if last_Zfac is not None
                        else np.ones(N, dtype=float)
                    )
                    P_now = _pressure_diag_psia(
                        col,
                        base_inputs.volume_model,
                        np.asarray(T_now, dtype=float).reshape((N,)),
                        MV_now,
                        Z_for_p,
                    )
                    dP_max = float(
                        np.nanmax(
                            np.abs(
                                np.asarray(P_now, dtype=float).reshape((N,))
                                - np.asarray(P_prev, dtype=float).reshape((N,))
                            )
                        )
                    )
                    if np.isfinite(dP_max) and dP_max >= float(base_inputs.thermo_refresh_dP_psia):
                        thermo_event = True
                        event_reason = "state_dP"

            if (not thermo_event) and base_inputs.thermo_refresh_dx is not None:
                if last_z_overall is None:
                    missing_history = True
                else:
                    tray_L_now = np.asarray(u_probe["tray_L"], dtype=float).reshape((N, Nc))
                    tray_V_now = np.asarray(u_probe["tray_V"], dtype=float).reshape((N, Nc))
                    x_now = np.asarray(u_probe["x_tray"], dtype=float).reshape((N, Nc))
                    z_now = np.zeros((N, Nc), dtype=float)
                    for i in range(N):
                        z_i = tray_L_now[i, :] + tray_V_now[i, :]
                        s_i = float(np.sum(z_i))
                        if (not np.isfinite(s_i)) or s_i <= 1e-300:
                            z_i = x_now[i, :]
                            s_i = float(np.sum(z_i))
                        z_now[i, :] = z_i / max(s_i, 1e-300)
                    dx_max = float(
                        np.nanmax(
                            np.abs(
                                np.asarray(z_now, dtype=float).reshape((N, Nc))
                                - np.asarray(last_z_overall, dtype=float).reshape((N, Nc))
                            )
                        )
                    )
                    if np.isfinite(dx_max) and dx_max >= float(base_inputs.thermo_refresh_dx):
                        thermo_event = True
                        event_reason = "state_dx"
        except Exception:
            missing_history = True

        if missing_history:
            return True, "guard_missing_history"
        if thermo_event:
            return True, str(event_reason)

    if bool(base_inputs.equilibrium_relaxation) and (last_K_tray is None):
        return True, "missing_K"
    return False, "hold"


def _ensure_startup_packet_equilibrium_arrays(
    packet: TrayThermoPacket,
    *,
    n_stages: int,
    n_components: int,
) -> None:
    if packet.x_eq is None:
        packet.x_equilibrium_tray = np.full((n_stages, n_components), np.nan, dtype=float)
    if packet.y_eq is None:
        packet.y_equilibrium_tray = np.full((n_stages, n_components), np.nan, dtype=float)


def _startup_flash_TP_full_stage_F_psia(
    provider: Any,
    stage_index0: int,
    T_F: float,
    P_psia: float,
    z: Sequence[float],
    *,
    n_components: int,
    thermo_call_category: Optional[str] = None,
):
    _ = int(stage_index0)
    category_fn = getattr(provider, "thermo_call_category", None)
    context = nullcontext()
    if thermo_call_category and callable(category_fn):
        try:
            context = category_fn(thermo_call_category)
        except Exception:
            context = nullcontext()
    with context:
        flash_fn = getattr(provider, "flash_TP_full", None)
        if callable(flash_fn):
            return flash_fn(float(T_F), float(P_psia), z)
        flash_tuple_fn = getattr(provider, "flash_TP_full_F_psia", None)
        if not callable(flash_tuple_fn):
            raise RuntimeError("Thermo provider does not expose flash_TP_full or flash_TP_full_F_psia")
        out = flash_tuple_fn(float(T_F), float(P_psia), z)
    if isinstance(out, (tuple, list)):
        if len(out) == 5:
            x_i, y_i, K_i, HL_i, HV_i = out
            Z_i = None
        elif len(out) == 6:
            x_i, y_i, K_i, HL_i, HV_i, Z_i = out
        else:
            raise RuntimeError("flash_TP_full_F_psia tuple rows must be length 5 or 6")
        return SimpleNamespace(
            x=np.asarray(x_i, dtype=float).reshape((n_components,)),
            y=np.asarray(y_i, dtype=float).reshape((n_components,)),
            K=np.asarray(K_i, dtype=float).reshape((n_components,)),
            HL_BTU_lbmol=float(HL_i),
            HV_BTU_lbmol=float(HV_i),
            Z=(None if Z_i is None else float(Z_i)),
        )
    return out


def _direct_startup_tray_thermo_diag(
    *,
    col: ColumnSpec,
    layout: StateVectorLayout,
    y: np.ndarray,
    inputs: ColumnInputs,
    include_temperature: bool,
) -> Dict[str, Any]:
    if inputs.thermo_provider is None:
        raise RuntimeError("Direct startup tray thermo refresh requires a thermo provider.")

    N = int(col.n_stages)
    Nc = int(col.n_components)
    u = layout.unpack(np.asarray(y, dtype=float))
    T_tray = _tray_temperature_F(col, layout, y, include_temperature)
    P_tray = np.asarray(getattr(col, "P_psia"), dtype=float).reshape((N,))
    tray_L = np.asarray(u["tray_L"], dtype=float).reshape((N, Nc))
    tray_V = np.asarray(u["tray_V"], dtype=float).reshape((N, Nc))
    x_tray = np.asarray(u["x_tray"], dtype=float).reshape((N, Nc))
    x_seed = None
    try:
        x_seed = np.asarray(getattr(col, "x0"), dtype=float).reshape((N, Nc))
    except Exception:
        x_seed = None
    z_overall = np.zeros((N, Nc), dtype=float)
    for i in range(N):
        z_i = np.asarray(tray_L[i, :] + tray_V[i, :], dtype=float)
        s_i = float(np.sum(z_i))
        if (not np.isfinite(s_i)) or s_i <= 1.0e-300:
            z_i = np.asarray(x_tray[i, :], dtype=float)
            s_i = float(np.sum(z_i))
        if ((not np.isfinite(s_i)) or s_i <= 1.0e-300) and x_seed is not None:
            z_i = np.asarray(x_seed[i, :], dtype=float)
            s_i = float(np.sum(z_i))
        if (not np.isfinite(s_i)) or s_i <= 1.0e-300:
            z_i = np.ones((Nc,), dtype=float) / float(max(Nc, 1))
            s_i = float(np.sum(z_i))
        z_overall[i, :] = z_i / max(s_i, 1.0e-300)

    packet = TrayThermoPacket(
        z_overall_tray=z_overall.copy(),
        K_tray=np.ones((N, Nc), dtype=float),
        HL_BTU_lbmol_tray=np.zeros(N, dtype=float),
        HV_BTU_lbmol_tray=np.zeros(N, dtype=float),
        Z_tray=np.ones(N, dtype=float),
        T_tray_F=np.asarray(T_tray, dtype=float).reshape((N,)).copy(),
        P_tray_psia=np.asarray(P_tray, dtype=float).reshape((N,)).copy(),
    )

    refresh_t0 = time.perf_counter()
    category_fn = getattr(inputs.thermo_provider, "thermo_call_category", None)
    context = nullcontext()
    if callable(category_fn):
        try:
            context = category_fn("startup_vapor_holdup_tray_refresh")
        except Exception:
            context = nullcontext()
    with context:
        refresh = refresh_tray_tp_packet(
            packet=packet,
            provider=inputs.thermo_provider,
            T_tray_F=T_tray,
            P_tray_psia=P_tray,
            z_overall_tray=z_overall,
            n_stages=N,
            n_components=Nc,
            dT_thresh_F=None,
            dP_thresh_psia=None,
            dX_thresh=None,
            T_prev_F=None,
            P_prev_psia=None,
            z_prev=None,
            ensure_packet_equilibrium_arrays=_ensure_startup_packet_equilibrium_arrays,
            flash_stage_fn=_startup_flash_TP_full_stage_F_psia,
            thermo_call_category="startup_vapor_holdup_tray_refresh",
        )
    wall_sec = float(time.perf_counter() - refresh_t0)
    return {
        "startup_vapor_holdup_refresh_source": "direct-tray-refresh",
        "startup_vapor_holdup_refresh_wall_sec": wall_sec,
        "startup_vapor_holdup_refresh_batch_used": bool(refresh.batch_used),
        "startup_vapor_holdup_refresh_rows": int(len(refresh.refresh_indices)),
        "z_overall_tray": np.asarray(packet.z_overall_tray, dtype=float).copy(),
        "K_tray": np.asarray(packet.K_tray, dtype=float).copy(),
        "HL_BTU_lbmol_tray": np.asarray(packet.HL, dtype=float).copy(),
        "HV_BTU_lbmol_tray": np.asarray(packet.HV, dtype=float).copy(),
        "Z_tray": np.asarray(packet.Zfac_tray, dtype=float).copy(),
        "cpL_BTU_lbmolF_tray": (
            None
            if packet.cpL_tray is None
            else np.asarray(packet.cpL_tray, dtype=float).copy()
        ),
        "cpV_BTU_lbmolF_tray": (
            None
            if packet.cpV_tray is None
            else np.asarray(packet.cpV_tray, dtype=float).copy()
        ),
        "x_eq_thermo_tray": np.asarray(packet.x_eq, dtype=float).copy(),
        "y_eq_thermo_tray": np.asarray(packet.y_eq, dtype=float).copy(),
        "P_psia_diag": np.asarray(P_tray, dtype=float).copy(),
        "P_psia_hyd": np.asarray(P_tray, dtype=float).copy(),
        "thermo_flash_refreshed": np.asarray(refresh.flash_refreshed, dtype=float).copy(),
        "thermo_flash_skipped": np.asarray(refresh.flash_skipped, dtype=float).copy(),
    }


def _condenser_duty_packet_from_diag(
    diag: Dict[str, Any],
    *,
    n_components: int,
) -> Optional[CondenserDutyPacket]:
    required = (
        "condenser_duty_cache_V_vapor_in_lbmolps",
        "condenser_duty_cache_T_vapor_in_F",
        "condenser_duty_cache_P_vapor_in_psia",
        "condenser_duty_cache_P_condenser_psia",
        "condenser_duty_cache_y_vapor_in",
    )
    if not all(key in diag for key in required):
        return None
    try:
        q_calc = None
        if "condenser_duty_cache_q_calc_BTUph" in diag:
            q_try = float(np.asarray(diag["condenser_duty_cache_q_calc_BTUph"], dtype=float).reshape((-1,))[0])
            if np.isfinite(q_try):
                q_calc = float(q_try)
        t_bub = None
        if "condenser_duty_cache_T_bubble_F" in diag:
            t_try = float(np.asarray(diag["condenser_duty_cache_T_bubble_F"], dtype=float).reshape((-1,))[0])
            if np.isfinite(t_try):
                t_bub = float(t_try)
        if q_calc is None and t_bub is None:
            return None
        mode = "specified"
        if "condenser_duty_cache_mode_total_condense" in diag:
            mode_flag = float(
                np.asarray(diag["condenser_duty_cache_mode_total_condense"], dtype=float).reshape((-1,))[0]
            )
            mode = "total-condense" if np.isfinite(mode_flag) and mode_flag >= 0.5 else "specified"
        return CondenserDutyPacket(
            q_calc_BTUph=q_calc,
            T_bubble_F=t_bub,
            mode=mode,
            V_vapor_in_lbmolps=float(
                np.asarray(diag["condenser_duty_cache_V_vapor_in_lbmolps"], dtype=float).reshape((-1,))[0]
            ),
            T_vapor_in_F=float(
                np.asarray(diag["condenser_duty_cache_T_vapor_in_F"], dtype=float).reshape((-1,))[0]
            ),
            P_vapor_in_psia=float(
                np.asarray(diag["condenser_duty_cache_P_vapor_in_psia"], dtype=float).reshape((-1,))[0]
            ),
            P_condenser_psia=float(
                np.asarray(diag["condenser_duty_cache_P_condenser_psia"], dtype=float).reshape((-1,))[0]
            ),
            y_vapor_in=np.asarray(diag["condenser_duty_cache_y_vapor_in"], dtype=float).reshape((n_components,)).copy(),
            hL_cond_BTU_lbmol=_as_float(_get_first_mapping_value(diag, ["condenser_duty_cache_hL_cond_BTU_lbmol"])),
        )
    except Exception:
        return None


def _normalize_startup_condenser_duty_mode(mode: Optional[str]) -> str:
    s = str(mode or "").strip().lower().replace("_", "-")
    if s in ("", "auto", "total", "total-condense", "total-condensing", "total-condenser"):
        return "total-condense"
    if s in ("specified", "spec", "fixed", "manual"):
        return "specified"
    return "total-condense"


def _seed_startup_condenser_duty_packet(
    *,
    col: ColumnSpec,
    layout: StateVectorLayout,
    y: np.ndarray,
    startup_packet: Optional[TrayThermoPacket],
    condenser_duty_mode: Optional[str],
) -> Optional[CondenserDutyPacket]:
    if startup_packet is None:
        return None
    try:
        N = int(col.n_stages)
        Nc = int(col.n_components)
    except Exception:
        return None
    if N <= 0 or Nc <= 0:
        return None

    try:
        u = layout.unpack(np.asarray(y, dtype=float))
        y_tray = np.asarray(u["y_tray"], dtype=float).reshape((N, Nc))
    except Exception:
        return None

    T_state = getattr(startup_packet, "T_state", None)
    P_state = getattr(startup_packet, "P_state", None)
    HL = getattr(startup_packet, "HL", None)
    HV = getattr(startup_packet, "HV", None)
    if T_state is None or P_state is None or HL is None or HV is None:
        return None

    try:
        T_arr = np.asarray(T_state, dtype=float).reshape((N,))
        P_arr = np.asarray(P_state, dtype=float).reshape((N,))
        HL_arr = np.asarray(HL, dtype=float).reshape((N,))
        HV_arr = np.asarray(HV, dtype=float).reshape((N,))
    except Exception:
        return None

    src_i = 1 if N > 1 else 0
    try:
        V_profile = np.asarray(getattr(col, "V_lbmolph"), dtype=float).reshape((N,))
        V_vapor_in_lbmolps = float(V_profile[src_i]) / 3600.0
    except Exception:
        return None
    if not np.isfinite(V_vapor_in_lbmolps) or V_vapor_in_lbmolps < 0.0:
        return None

    try:
        y_vapor_in = np.asarray(y_tray[src_i, :], dtype=float).reshape((Nc,))
    except Exception:
        return None
    if not np.all(np.isfinite(y_vapor_in)):
        return None
    sy = float(np.sum(y_vapor_in))
    if sy <= 1.0e-12:
        return None
    y_vapor_in = y_vapor_in / sy

    if not (
        np.isfinite(float(T_arr[src_i]))
        and np.isfinite(float(P_arr[src_i]))
        and np.isfinite(float(P_arr[0]))
    ):
        return None

    mode_norm = _normalize_startup_condenser_duty_mode(condenser_duty_mode)
    T_bubble_F = None
    if np.isfinite(float(T_arr[0])):
        T_bubble_F = float(T_arr[0])

    q_calc_BTUph = None
    if np.isfinite(float(HL_arr[0])) and np.isfinite(float(HV_arr[src_i])):
        q_try = float(V_vapor_in_lbmolps) * (float(HL_arr[0]) - float(HV_arr[src_i])) * 3600.0
        if np.isfinite(q_try):
            q_calc_BTUph = float(q_try)

    if q_calc_BTUph is None and T_bubble_F is None:
        return None

    return CondenserDutyPacket(
        q_calc_BTUph=q_calc_BTUph,
        T_bubble_F=T_bubble_F,
        mode=str(mode_norm),
        V_vapor_in_lbmolps=float(V_vapor_in_lbmolps),
        T_vapor_in_F=float(T_arr[src_i]),
        P_vapor_in_psia=float(P_arr[src_i]),
        P_condenser_psia=float(P_arr[0]),
        y_vapor_in=y_vapor_in.copy(),
        hL_cond_BTU_lbmol=float(HL_arr[0]) if np.isfinite(float(HL_arr[0])) else None,
    )


def _tray_thermo_packet_to_json_doc(packet: Optional[TrayThermoPacket]) -> Optional[Dict[str, Any]]:
    if packet is None:
        return None
    try:
        z_overall = np.asarray(packet.z_overall, dtype=float)
        n_stages, n_components = z_overall.shape
        return {
            "z_overall_tray": _json_optional_array(z_overall, shape=(n_stages, n_components)),
            "K_tray": _json_optional_array(packet.K_tray, shape=(n_stages, n_components)),
            "HL_BTU_lbmol_tray": _json_optional_array(packet.HL, shape=(n_stages,)),
            "HV_BTU_lbmol_tray": _json_optional_array(packet.HV, shape=(n_stages,)),
            "Z_tray": _json_optional_array(packet.Zfac_tray, shape=(n_stages,)),
            "cpL_BTU_lbmolF_tray": _json_optional_array(packet.cpL_tray, shape=(n_stages,)),
            "cpV_BTU_lbmolF_tray": _json_optional_array(packet.cpV_tray, shape=(n_stages,)),
            "T_tray_F": _json_optional_array(packet.T_state, shape=(n_stages,)),
            "P_tray_psia": _json_optional_array(packet.P_state, shape=(n_stages,)),
            "x_equilibrium_tray": _json_optional_array(packet.x_eq, shape=(n_stages, n_components)),
            "y_equilibrium_tray": _json_optional_array(packet.y_eq, shape=(n_stages, n_components)),
        }
    except Exception:
        return None


def _tray_thermo_packet_from_json_doc(
    doc: Optional[Dict[str, Any]],
    *,
    n_stages: int,
    n_components: int,
) -> Optional[TrayThermoPacket]:
    if not isinstance(doc, dict):
        return None
    try:
        z_overall = _array_from_json(doc.get("z_overall_tray"), shape=(n_stages, n_components))
        K_tray = _array_from_json(doc.get("K_tray"), shape=(n_stages, n_components))
        HL = _array_from_json(doc.get("HL_BTU_lbmol_tray"), shape=(n_stages,))
        HV = _array_from_json(doc.get("HV_BTU_lbmol_tray"), shape=(n_stages,))
        Z = _array_from_json(doc.get("Z_tray"), shape=(n_stages,))
        if z_overall is None or K_tray is None or HL is None or HV is None or Z is None:
            return None
        return TrayThermoPacket(
            z_overall_tray=z_overall,
            K_tray=K_tray,
            HL_BTU_lbmol_tray=HL,
            HV_BTU_lbmol_tray=HV,
            Z_tray=Z,
            cpL_BTU_lbmolF_tray=_array_from_json(doc.get("cpL_BTU_lbmolF_tray"), shape=(n_stages,)),
            cpV_BTU_lbmolF_tray=_array_from_json(doc.get("cpV_BTU_lbmolF_tray"), shape=(n_stages,)),
            T_tray_F=_array_from_json(doc.get("T_tray_F"), shape=(n_stages,)),
            P_tray_psia=_array_from_json(doc.get("P_tray_psia"), shape=(n_stages,)),
            x_equilibrium_tray=_array_from_json(doc.get("x_equilibrium_tray"), shape=(n_stages, n_components)),
            y_equilibrium_tray=_array_from_json(doc.get("y_equilibrium_tray"), shape=(n_stages, n_components)),
        )
    except Exception:
        return None


def _condenser_duty_packet_to_json_doc(packet: Optional[CondenserDutyPacket]) -> Optional[Dict[str, Any]]:
    if packet is None:
        return None
    try:
        return {
            "q_calc_BTUph": None if packet.q_calc_BTUph is None else float(packet.q_calc_BTUph),
            "T_bubble_F": None if packet.T_bubble_F is None else float(packet.T_bubble_F),
            "mode": str(packet.mode),
            "V_vapor_in_lbmolps": float(packet.V_vapor_in_lbmolps),
            "T_vapor_in_F": float(packet.T_vapor_in_F),
            "P_vapor_in_psia": float(packet.P_vapor_in_psia),
            "P_condenser_psia": float(packet.P_condenser_psia),
            "y_vapor_in": _json_optional_array(packet.y_vapor_in),
            "hL_cond_BTU_lbmol": None if packet.hL_cond_BTU_lbmol is None else float(packet.hL_cond_BTU_lbmol),
        }
    except Exception:
        return None


def _condenser_duty_packet_from_json_doc(
    doc: Optional[Dict[str, Any]],
    *,
    n_components: int,
) -> Optional[CondenserDutyPacket]:
    if not isinstance(doc, dict):
        return None
    try:
        y_vapor_in = _array_from_json(doc.get("y_vapor_in"), shape=(n_components,))
        if y_vapor_in is None:
            return None
        return CondenserDutyPacket(
            q_calc_BTUph=_as_float(doc.get("q_calc_BTUph")),
            T_bubble_F=_as_float(doc.get("T_bubble_F")),
            mode=str(doc.get("mode") or "total-condense"),
            V_vapor_in_lbmolps=float(doc.get("V_vapor_in_lbmolps")),
            T_vapor_in_F=float(doc.get("T_vapor_in_F")),
            P_vapor_in_psia=float(doc.get("P_vapor_in_psia")),
            P_condenser_psia=float(doc.get("P_condenser_psia")),
            y_vapor_in=y_vapor_in,
            hL_cond_BTU_lbmol=_as_float(doc.get("hL_cond_BTU_lbmol")),
        )
    except Exception:
        return None


def _startup_feed_rachford_rice_beta(K: np.ndarray, z: np.ndarray) -> float:
    K = np.asarray(K, dtype=float).reshape((-1,))
    z = np.asarray(z, dtype=float).reshape((-1,))
    if K.size != z.size or K.size <= 0:
        return 0.0
    if np.all(K <= 1.0):
        return 0.0
    if np.all(K >= 1.0):
        return 1.0
    lo = 0.0
    hi = 1.0
    for _ in range(60):
        beta = 0.5 * (lo + hi)
        denom = 1.0 + beta * (K - 1.0)
        denom = np.where(np.abs(denom) < 1.0e-12, 1.0e-12, denom)
        rr = float(np.sum(z * (K - 1.0) / denom))
        if abs(rr) <= 1.0e-12:
            return float(np.clip(beta, 0.0, 1.0))
        if rr > 0.0:
            lo = beta
        else:
            hi = beta
    return float(np.clip(0.5 * (lo + hi), 0.0, 1.0))


def _feed_stage_flash_packet_to_json_doc(packet: Optional[FeedStageFlashPacket]) -> Optional[Dict[str, Any]]:
    if packet is None:
        return None
    try:
        return {
            "stage0": int(packet.stage0),
            "T_feed_F": float(packet.T_feed_F),
            "P_feed_psia": float(packet.P_feed_psia),
            "z_feed": _json_optional_array(packet.z_feed),
            "Fk_L_lbmolps": _json_optional_array(packet.Fk_L_lbmolps),
            "Fk_V_lbmolps": _json_optional_array(packet.Fk_V_lbmolps),
            "hL_BTU_lbmol": (None if packet.hL_BTU_lbmol is None else float(packet.hL_BTU_lbmol)),
            "hV_BTU_lbmol": (None if packet.hV_BTU_lbmol is None else float(packet.hV_BTU_lbmol)),
        }
    except Exception:
        return None


def _feed_stage_flash_packet_from_json_doc(
    doc: Optional[Dict[str, Any]],
    *,
    n_components: int,
) -> Optional[FeedStageFlashPacket]:
    if not isinstance(doc, dict):
        return None
    try:
        z_feed = _array_from_json(doc.get("z_feed"), shape=(n_components,))
        fk_l = _array_from_json(doc.get("Fk_L_lbmolps"), shape=(n_components,))
        fk_v = _array_from_json(doc.get("Fk_V_lbmolps"), shape=(n_components,))
        if z_feed is None or fk_l is None or fk_v is None:
            return None
        return FeedStageFlashPacket(
            stage0=int(doc.get("stage0")),
            T_feed_F=float(doc.get("T_feed_F")),
            P_feed_psia=float(doc.get("P_feed_psia")),
            z_feed=z_feed,
            Fk_L_lbmolps=fk_l,
            Fk_V_lbmolps=fk_v,
            hL_BTU_lbmol=_as_float(doc.get("hL_BTU_lbmol")),
            hV_BTU_lbmol=_as_float(doc.get("hV_BTU_lbmol")),
        )
    except Exception:
        return None


def _bottom_sump_cp_packet_to_json_doc(packet: Optional[BottomSumpCpPacket]) -> Optional[Dict[str, Any]]:
    if packet is None:
        return None
    try:
        return {
            "T_sump_F": float(packet.T_sump_F),
            "P_sump_psia": float(packet.P_sump_psia),
            "x_sump": _json_optional_array(packet.x_sump),
            "cpL_BTU_lbmolF": float(packet.cpL_BTU_lbmolF),
        }
    except Exception:
        return None


def _bottom_sump_cp_packet_from_json_doc(
    doc: Optional[Dict[str, Any]],
    *,
    n_components: int,
) -> Optional[BottomSumpCpPacket]:
    if not isinstance(doc, dict):
        return None
    try:
        x_sump = _array_from_json(doc.get("x_sump"), shape=(n_components,))
        if x_sump is None:
            return None
        return BottomSumpCpPacket(
            T_sump_F=float(doc.get("T_sump_F")),
            P_sump_psia=float(doc.get("P_sump_psia")),
            x_sump=x_sump,
            cpL_BTU_lbmolF=float(doc.get("cpL_BTU_lbmolF")),
        )
    except Exception:
        return None


def _bottom_sump_cp_packet_from_diag(
    diag: Dict[str, Any],
    *,
    n_components: int,
) -> Optional[BottomSumpCpPacket]:
    required = (
        "bottom_sump_cp_cache_T_F",
        "bottom_sump_cp_cache_P_psia",
        "bottom_sump_cp_cache_x",
        "bottom_sump_cp_cache_cpL_BTU_lbmolF",
    )
    if not all(key in diag for key in required):
        return None
    try:
        return BottomSumpCpPacket(
            T_sump_F=float(np.asarray(diag["bottom_sump_cp_cache_T_F"], dtype=float).reshape((-1,))[0]),
            P_sump_psia=float(np.asarray(diag["bottom_sump_cp_cache_P_psia"], dtype=float).reshape((-1,))[0]),
            x_sump=np.asarray(diag["bottom_sump_cp_cache_x"], dtype=float).reshape((n_components,)).copy(),
            cpL_BTU_lbmolF=float(
                np.asarray(diag["bottom_sump_cp_cache_cpL_BTU_lbmolF"], dtype=float).reshape((-1,))[0]
            ),
        )
    except Exception:
        return None


def _startup_feed_stage_flash_state(
    *,
    col: ColumnSpec,
    P_tray_psia: Optional[np.ndarray],
) -> Optional[Tuple[int, np.ndarray, float, float, float]]:
    try:
        s = (getattr(col, "streams", {}) or {}).get("Feed")
    except Exception:
        s = None
    if s is None or getattr(s, "stage_1based", None) is None or getattr(s, "total_molar_flow_lbmolph", None) is None:
        return None
    if getattr(s, "temperature_f", None) is None:
        return None
    try:
        stage0 = int(s.stage_1based) - 1
        Nc = int(col.n_components)
        Ft = float(s.total_molar_flow_lbmolph) / 3600.0
        if (not np.isfinite(Ft)) or Ft <= 1.0e-300:
            return None
        if getattr(s, "component_molar_flows_lbmolph", None):
            Fk = _component_molar_flows_vector_lbmolps(
                dict(getattr(s, "component_molar_flows_lbmolph", {}) or {}),
                np.asarray(col.components_excel, dtype=object),
            )
        else:
            z_fallback = np.asarray(col.x0[stage0, :], dtype=float).reshape((Nc,))
            z_fallback = z_fallback / max(float(np.sum(z_fallback)), 1.0e-300)
            Fk = Ft * z_fallback
        Ft_comp = float(np.sum(Fk))
        if (not np.isfinite(Ft_comp)) or Ft_comp <= 1.0e-300:
            return None
        z_feed = Fk / Ft_comp
        T_feed = float(s.temperature_f)
        P_feed = None
        if P_tray_psia is not None:
            p_arr = np.asarray(P_tray_psia, dtype=float).reshape((-1,))
            if 0 <= stage0 < p_arr.size and np.isfinite(p_arr[stage0]) and p_arr[stage0] > 0.0:
                P_feed = float(p_arr[stage0])
        if P_feed is None:
            p_arr = np.asarray(
                getattr(col, "P_psia", np.full(col.n_stages, 200.0, dtype=float)),
                dtype=float,
            ).reshape((-1,))
            if 0 <= stage0 < p_arr.size and np.isfinite(p_arr[stage0]) and p_arr[stage0] > 0.0:
                P_feed = float(p_arr[stage0])
        if P_feed is None or (not np.isfinite(P_feed)) or P_feed <= 0.0:
            return None
    except Exception:
        return None
    return int(stage0), np.asarray(z_feed, dtype=float).reshape((Nc,)).copy(), float(T_feed), float(P_feed), float(Ft_comp)


def _seed_startup_feed_stage_flash_packet(
    *,
    col: ColumnSpec,
    thermo_provider: Optional[Any],
    P_tray_psia: Optional[np.ndarray],
) -> Optional[FeedStageFlashPacket]:
    if thermo_provider is None:
        return None
    feed_state = _startup_feed_stage_flash_state(
        col=col,
        P_tray_psia=P_tray_psia,
    )
    if feed_state is None:
        return None
    try:
        stage0, z_feed, T_feed, P_feed, Ft_comp = feed_state
        Nc = int(col.n_components)
        with getattr(thermo_provider, "thermo_call_category", lambda *_args, **_kwargs: nullcontext())("feed_stage_flash"):
            fres = getattr(thermo_provider, "flash_TP_full", None)
            if callable(fres):
                fres = fres(float(T_feed), float(P_feed), z_feed)
            else:
                fres = _startup_flash_TP_full_stage_F_psia(
                    thermo_provider,
                    stage0,
                    T_feed,
                    P_feed,
                    z_feed,
                    n_components=Nc,
                    thermo_call_category="feed_stage_flash",
                )
        K = np.asarray(fres.K, dtype=float).reshape((Nc,))
        hL_feed = None
        hV_feed = None
        try:
            h_try_L = getattr(fres, "HL_BTU_lbmol", None)
            if h_try_L is None:
                h_try_L = getattr(fres, "HL", None)
            if h_try_L is not None and np.isfinite(float(h_try_L)):
                hL_feed = float(h_try_L)
        except Exception:
            hL_feed = None
        try:
            h_try_V = getattr(fres, "HV_BTU_lbmol", None)
            if h_try_V is None:
                h_try_V = getattr(fres, "HV", None)
            if h_try_V is not None and np.isfinite(float(h_try_V)):
                hV_feed = float(h_try_V)
        except Exception:
            hV_feed = None
        beta = float(np.clip(_startup_feed_rachford_rice_beta(K, z_feed), 0.0, 1.0))
        denom = 1.0 + beta * (K - 1.0)
        denom = np.where(np.abs(denom) < 1e-12, np.sign(denom) * 1e-12 + (denom == 0) * 1e-12, denom)
        x = np.clip(z_feed / denom, 0.0, None)
        sx = float(np.sum(x))
        x = z_feed.copy() if ((not np.isfinite(sx)) or sx <= 1.0e-300) else (x / sx)
        y = np.clip(K * x, 0.0, None)
        sy = float(np.sum(y))
        y = z_feed.copy() if ((not np.isfinite(sy)) or sy <= 1.0e-300) else (y / sy)
        return FeedStageFlashPacket(
            stage0=int(stage0),
            T_feed_F=float(T_feed),
            P_feed_psia=float(P_feed),
            z_feed=z_feed.copy(),
            Fk_L_lbmolps=((1.0 - beta) * Ft_comp * x).copy(),
            Fk_V_lbmolps=(beta * Ft_comp * y).copy(),
            hL_BTU_lbmol=hL_feed,
            hV_BTU_lbmol=hV_feed,
        )
    except Exception:
        return None


def _resolve_startup_seed_cache_path(cfg: "RunnerConfig") -> Optional[Path]:
    raw = _clean_optional_text(getattr(cfg, "startup_seed_cache_path", None))
    if raw:
        path = Path(raw)
    elif bool(getattr(cfg, "enable_startup_seed_cache", False)):
        excel_stem = Path(str(getattr(cfg, "excel_path", "case.xlsx"))).stem or "case"
        runtime_mode = _normalize_runtime_mode(getattr(cfg, "runtime_mode", None), default="legacy")
        thermo_mode = str(getattr(cfg, "thermo_mode", "thermo") or "thermo").strip().lower()
        model_suffix = ""
        if thermo_mode == "clapeyron":
            model_txt = str(getattr(cfg, "clapeyron_model", "") or "").strip().lower()
            if model_txt:
                model_suffix = f"_{model_txt}"
            source_txt = str(getattr(cfg, "clapeyron_pr_parameter_source", "") or "").strip().lower()
            if source_txt and source_txt not in {"default", "clapeyron"}:
                model_suffix += f"_{source_txt}"
        path = Path(str(getattr(cfg, "logs_dir", "logs"))) / (
            f"{excel_stem}__startup_seed_{runtime_mode}_{thermo_mode}{model_suffix}.json"
        )
    else:
        return None
    if not path.is_absolute():
        path = Path.cwd() / path
    return path


def _startup_seed_cache_signature(
    *,
    cfg: "RunnerConfig",
    col: ColumnSpec,
    layout: StateVectorLayout,
    base_inputs: ColumnInputs,
) -> Dict[str, Any]:
    return {
        "excel_path": str(Path(cfg.excel_path).resolve()),
        "runtime_mode": _normalize_runtime_mode(getattr(cfg, "runtime_mode", None), default="legacy"),
        "thermo_mode": str(getattr(cfg, "thermo_mode", "") or "").strip().lower(),
        "clapeyron_model": str(getattr(cfg, "clapeyron_model", "") or "").strip(),
        "clapeyron_pr_parameter_source": str(
            getattr(cfg, "clapeyron_pr_parameter_source", "default") or "default"
        ).strip().lower(),
        "include_temperature": bool(getattr(cfg, "include_temperature", True)),
        "include_energy": bool(getattr(cfg, "include_energy", False)),
        "n_stages": int(col.n_stages),
        "n_components": int(col.n_components),
        "components_excel": [str(x) for x in list(getattr(col, "components_excel", []) or [])],
        "pressure_model": str(getattr(base_inputs, "pressure_model", "") or "").strip().lower(),
        "vapor_flow_model": str(getattr(base_inputs, "vapor_flow_model", "") or "").strip().lower(),
        "condenser_duty_mode": _normalize_startup_condenser_duty_mode(
            getattr(base_inputs, "condenser_duty_mode", None)
        ),
        "layout_include_top": bool(getattr(layout, "include_top", False)),
        "layout_include_bottom": bool(getattr(layout, "include_bottom", False)),
        "layout_include_vapor": bool(getattr(layout, "include_vapor", False)),
    }


def _startup_seed_cache_payload(
    *,
    cfg: "RunnerConfig",
    col: ColumnSpec,
    layout: StateVectorLayout,
    base_inputs: ColumnInputs,
    y: np.ndarray,
    last_T_tray: Optional[np.ndarray],
    last_P_diag: Optional[np.ndarray],
    last_P_hyd: Optional[np.ndarray],
    last_K_tray: Optional[np.ndarray],
    last_HL: Optional[np.ndarray],
    last_HV: Optional[np.ndarray],
    last_Zfac: Optional[np.ndarray],
    last_z_overall: Optional[np.ndarray],
    last_tray_bubble_target_F: Optional[np.ndarray],
    last_tray_thermo_packet: Optional[TrayThermoPacket],
    last_condenser_duty_packet: Optional[CondenserDutyPacket],
    last_feed_stage_flash_packet: Optional[FeedStageFlashPacket],
    last_bottom_sump_cp_packet: Optional[BottomSumpCpPacket],
    last_reb_T: Optional[float],
    last_reb_x: Optional[np.ndarray],
    last_reb_y: Optional[np.ndarray],
    last_reb_beta: Optional[float],
    startup_seeded_condenser_duty_packet: bool,
) -> Dict[str, Any]:
    n_stages = int(col.n_stages)
    n_components = int(col.n_components)
    return {
        "format_version": 1,
        "saved_at_local": _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "signature": _startup_seed_cache_signature(
            cfg=cfg,
            col=col,
            layout=layout,
            base_inputs=base_inputs,
        ),
        "state": {
            "y": _json_optional_array(y),
            "last_T_tray": _json_optional_array(last_T_tray, shape=(n_stages,)),
            "last_P_diag": _json_optional_array(last_P_diag, shape=(n_stages,)),
            "last_P_hyd": _json_optional_array(last_P_hyd, shape=(n_stages,)),
            "last_K_tray": _json_optional_array(last_K_tray, shape=(n_stages, n_components)),
            "last_HL": _json_optional_array(last_HL, shape=(n_stages,)),
            "last_HV": _json_optional_array(last_HV, shape=(n_stages,)),
            "last_Zfac": _json_optional_array(last_Zfac, shape=(n_stages,)),
            "last_z_overall": _json_optional_array(last_z_overall, shape=(n_stages, n_components)),
            "last_tray_bubble_target_F": _json_optional_array(last_tray_bubble_target_F, shape=(n_stages,)),
            "last_reb_T": (None if last_reb_T is None else float(last_reb_T)),
            "last_reb_x": _json_optional_array(last_reb_x, shape=(n_components,)),
            "last_reb_y": _json_optional_array(last_reb_y, shape=(n_components,)),
            "last_reb_beta": (None if last_reb_beta is None else float(last_reb_beta)),
            "startup_seeded_condenser_duty_packet": bool(startup_seeded_condenser_duty_packet),
        },
        "tray_thermo_packet": _tray_thermo_packet_to_json_doc(last_tray_thermo_packet),
        "condenser_duty_packet": _condenser_duty_packet_to_json_doc(last_condenser_duty_packet),
        "feed_stage_flash_packet": _feed_stage_flash_packet_to_json_doc(last_feed_stage_flash_packet),
        "bottom_sump_cp_packet": _bottom_sump_cp_packet_to_json_doc(last_bottom_sump_cp_packet),
    }


def _load_startup_seed_cache(
    *,
    path: Path,
    cfg: "RunnerConfig",
    col: ColumnSpec,
    layout: StateVectorLayout,
    base_inputs: ColumnInputs,
) -> Tuple[Optional[Dict[str, Any]], Dict[str, Any]]:
    info: Dict[str, Any] = {
        "enabled": True,
        "path": str(path),
        "loaded": False,
        "saved": False,
        "reason": "",
    }
    if not path.exists():
        info["reason"] = "missing"
        return None, info
    try:
        doc = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        info["reason"] = f"read_error:{exc}"
        return None, info
    if int(_as_int(doc.get("format_version")) or 0) != 1:
        info["reason"] = "unsupported_format"
        return None, info
    expected_signature = _startup_seed_cache_signature(
        cfg=cfg,
        col=col,
        layout=layout,
        base_inputs=base_inputs,
    )
    actual_signature = doc.get("signature")
    if not isinstance(actual_signature, dict):
        info["reason"] = "missing_signature"
        return None, info
    mismatches = sorted(
        key for key, value in expected_signature.items()
        if actual_signature.get(key) != value
    )
    if mismatches:
        info["reason"] = "signature_mismatch:" + ",".join(mismatches)
        info["signature_mismatch_keys"] = list(mismatches)
        return None, info
    state = doc.get("state")
    if not isinstance(state, dict):
        info["reason"] = "missing_state"
        return None, info
    try:
        y = _array_from_json(state.get("y"))
        if y is None:
            info["reason"] = "missing_y"
            return None, info
        loaded = {
            "y": y.reshape((layout.n_states(),)).copy(),
            "last_T_tray": _array_from_json(state.get("last_T_tray"), shape=(col.n_stages,)),
            "last_P_diag": _array_from_json(state.get("last_P_diag"), shape=(col.n_stages,)),
            "last_P_hyd": _array_from_json(state.get("last_P_hyd"), shape=(col.n_stages,)),
            "last_K_tray": _array_from_json(state.get("last_K_tray"), shape=(col.n_stages, col.n_components)),
            "last_HL": _array_from_json(state.get("last_HL"), shape=(col.n_stages,)),
            "last_HV": _array_from_json(state.get("last_HV"), shape=(col.n_stages,)),
            "last_Zfac": _array_from_json(state.get("last_Zfac"), shape=(col.n_stages,)),
            "last_z_overall": _array_from_json(
                state.get("last_z_overall"),
                shape=(col.n_stages, col.n_components),
            ),
            "last_tray_bubble_target_F": _array_from_json(
                state.get("last_tray_bubble_target_F"),
                shape=(col.n_stages,),
            ),
            "last_tray_thermo_packet": _tray_thermo_packet_from_json_doc(
                doc.get("tray_thermo_packet"),
                n_stages=col.n_stages,
                n_components=col.n_components,
            ),
            "last_condenser_duty_packet": _condenser_duty_packet_from_json_doc(
                doc.get("condenser_duty_packet"),
                n_components=col.n_components,
            ),
            "last_feed_stage_flash_packet": _feed_stage_flash_packet_from_json_doc(
                doc.get("feed_stage_flash_packet"),
                n_components=col.n_components,
            ),
            "last_bottom_sump_cp_packet": _bottom_sump_cp_packet_from_json_doc(
                doc.get("bottom_sump_cp_packet"),
                n_components=col.n_components,
            ),
            "last_reb_T": _as_float(state.get("last_reb_T")),
            "last_reb_x": _array_from_json(state.get("last_reb_x"), shape=(col.n_components,)),
            "last_reb_y": _array_from_json(state.get("last_reb_y"), shape=(col.n_components,)),
            "last_reb_beta": _as_float(state.get("last_reb_beta")),
            "startup_seeded_condenser_duty_packet": bool(
                state.get("startup_seeded_condenser_duty_packet", False)
            ),
        }
    except Exception as exc:
        info["reason"] = f"decode_error:{exc}"
        return None, info
    info["loaded"] = True
    info["reason"] = "loaded"
    return loaded, info


def _write_startup_seed_cache(
    *,
    path: Path,
    cfg: "RunnerConfig",
    col: ColumnSpec,
    layout: StateVectorLayout,
    base_inputs: ColumnInputs,
    y: np.ndarray,
    last_T_tray: Optional[np.ndarray],
    last_P_diag: Optional[np.ndarray],
    last_P_hyd: Optional[np.ndarray],
    last_K_tray: Optional[np.ndarray],
    last_HL: Optional[np.ndarray],
    last_HV: Optional[np.ndarray],
    last_Zfac: Optional[np.ndarray],
    last_z_overall: Optional[np.ndarray],
    last_tray_bubble_target_F: Optional[np.ndarray],
    last_tray_thermo_packet: Optional[TrayThermoPacket],
    last_condenser_duty_packet: Optional[CondenserDutyPacket],
    last_feed_stage_flash_packet: Optional[FeedStageFlashPacket],
    last_bottom_sump_cp_packet: Optional[BottomSumpCpPacket],
    last_reb_T: Optional[float],
    last_reb_x: Optional[np.ndarray],
    last_reb_y: Optional[np.ndarray],
    last_reb_beta: Optional[float],
    startup_seeded_condenser_duty_packet: bool,
) -> Dict[str, Any]:
    info: Dict[str, Any] = {
        "enabled": True,
        "path": str(path),
        "loaded": False,
        "saved": False,
        "reason": "",
    }
    try:
        _ensure_dir(path.parent)
        payload = _startup_seed_cache_payload(
            cfg=cfg,
            col=col,
            layout=layout,
            base_inputs=base_inputs,
            y=y,
            last_T_tray=last_T_tray,
            last_P_diag=last_P_diag,
            last_P_hyd=last_P_hyd,
            last_K_tray=last_K_tray,
            last_HL=last_HL,
            last_HV=last_HV,
            last_Zfac=last_Zfac,
            last_z_overall=last_z_overall,
            last_tray_bubble_target_F=last_tray_bubble_target_F,
            last_tray_thermo_packet=last_tray_thermo_packet,
            last_condenser_duty_packet=last_condenser_duty_packet,
            last_feed_stage_flash_packet=last_feed_stage_flash_packet,
            last_bottom_sump_cp_packet=last_bottom_sump_cp_packet,
            last_reb_T=last_reb_T,
            last_reb_x=last_reb_x,
            last_reb_y=last_reb_y,
            last_reb_beta=last_reb_beta,
            startup_seeded_condenser_duty_packet=startup_seeded_condenser_duty_packet,
        )
        _write_json_atomic(path, payload)
        info["saved"] = True
        info["reason"] = "saved"
    except Exception as exc:
        info["reason"] = f"write_error:{exc}"
    return info


def _mapping_scalar(d: Dict[str, Any], key: str, default: float = np.nan) -> float:
    """Best-effort extraction of a scalar float from a mapping entry."""
    if key not in d:
        return float(default)
    try:
        val = float(np.asarray(d[key], dtype=float).reshape((-1,))[0])
        return float(val)
    except Exception:
        return float(default)


def _pack_top_drum_vapor_to_pressure(
    *,
    y: np.ndarray,
    col: ColumnSpec,
    layout: StateVectorLayout,
    inputs: ColumnInputs,
    target_pressure_psia: Optional[float],
) -> Tuple[np.ndarray, Dict[str, Any]]:
    """Scale explicit top-drum vapor inventory so raw drum pressure starts at target."""
    info: Dict[str, Any] = {
        "enabled": True,
        "applied": False,
        "target_pressure_psia": np.nan,
        "raw_pressure_initial_psia": np.nan,
        "raw_pressure_final_psia": np.nan,
        "top_vapor_initial_lbmol": np.nan,
        "top_vapor_final_lbmol": np.nan,
        "scale_factor": np.nan,
        "reason": "",
    }
    if (not getattr(layout, "include_top", False)) or (not getattr(layout, "include_vapor", False)):
        info["reason"] = "missing_top_or_vapor_state"
        return np.asarray(y, dtype=float), info
    sl = layout.slices()
    if "top_V" not in sl:
        info["reason"] = "missing_top_V_slice"
        return np.asarray(y, dtype=float), info
    try:
        p_target = float(target_pressure_psia) if target_pressure_psia is not None else np.nan
    except Exception:
        p_target = np.nan
    if (not np.isfinite(p_target)) or p_target <= 0.0:
        info["reason"] = "invalid_target_pressure"
        return np.asarray(y, dtype=float), info
    info["target_pressure_psia"] = float(p_target)

    y_arr = np.asarray(y, dtype=float).reshape((-1,)).copy()
    u = layout.unpack(y_arr)
    top_v = np.asarray(y_arr[sl["top_V"]], dtype=float).reshape((-1,))
    m0 = float(np.sum(np.where(np.isfinite(top_v), top_v, 0.0)))
    if (not np.isfinite(m0)) or m0 <= 0.0:
        info["reason"] = "nonpositive_top_vapor_inventory"
        return y_arr, info
    top_extra = 0.0
    try:
        vextra = getattr(inputs, "top_drum_extra_vapor_volume_ft3", None)
        if vextra is not None and np.isfinite(float(vextra)) and float(vextra) > 0.0:
            top_extra = float(vextra)
    except Exception:
        top_extra = 0.0
    top_vol = None
    top_total_vol = getattr(inputs, "top_drum_total_volume_ft3", None)
    try:
        top_total_vol_f = float(top_total_vol) if top_total_vol is not None else np.nan
    except Exception:
        top_total_vol_f = np.nan
    if np.isfinite(top_total_vol_f) and top_total_vol_f > 0.0:
        rho_top = None
        if getattr(inputs, "thermo_provider", None) is not None and hasattr(inputs.thermo_provider, "liquid_density_lbmol_ft3"):
            try:
                top_L = np.asarray(u.get("top_L", np.zeros((col.n_components,), dtype=float)), dtype=float).reshape((col.n_components,))
                x_top = _normalize_comp(np.where(np.isfinite(top_L), top_L, 0.0))
                if "tray_T_f" in u:
                    T_rho = float(np.asarray(u["tray_T_f"], dtype=float).reshape((-1,))[0])
                else:
                    T_rho = float(np.asarray(getattr(col, "T_f"), dtype=float).reshape((-1,))[0])
                P_rho = float(p_target)
                rho_try = float(inputs.thermo_provider.liquid_density_lbmol_ft3(T_rho, P_rho, x_top))
                if np.isfinite(rho_try) and rho_try > 1.0e-12:
                    rho_top = float(rho_try)
            except Exception:
                rho_top = None
        if rho_top is not None:
            try:
                m_top_L = float(np.sum(np.asarray(u.get("top_L", []), dtype=float).reshape((-1,))))
                liq_vol = float(np.clip(max(m_top_L / float(rho_top), 0.0), 0.0, top_total_vol_f))
                top_vol = float(top_total_vol_f) - float(liq_vol) + float(top_extra)
            except Exception:
                top_vol = None
    if top_vol is None:
        top_vol = getattr(inputs, "top_drum_vapor_volume_ft3", None)
        if top_vol is not None:
            try:
                top_vol = float(top_vol) + float(top_extra)
            except Exception:
                pass
    if top_vol is None:
        try:
            vv = _vapor_volume_ft3_per_stage(inputs.volume_model, int(col.n_stages))
            top_vol = float(np.asarray(vv, dtype=float).reshape((-1,))[0])
        except Exception:
            top_vol = None
    try:
        top_vol_f = float(top_vol) if top_vol is not None else np.nan
    except Exception:
        top_vol_f = np.nan
    if (not np.isfinite(top_vol_f)) or top_vol_f <= 0.0:
        info["reason"] = "invalid_top_vapor_volume"
        return y_arr, info
    try:
        if "tray_T_f" in u:
            top_T_F = float(np.asarray(u["tray_T_f"], dtype=float).reshape((-1,))[0])
        else:
            top_T_F = float(np.asarray(getattr(col, "T_f"), dtype=float).reshape((-1,))[0])
    except Exception:
        top_T_F = np.nan
    if (not np.isfinite(top_T_F)) or (top_T_F + 459.67) <= 1.0e-6:
        info["reason"] = "invalid_top_temperature"
        return y_arr, info
    y_top = _normalize_comp(np.where(np.isfinite(top_v), top_v, 0.0))
    p_top_res = _compute_top_drum_pressure_psia(
        top_V=np.asarray(top_v, dtype=float).reshape((-1,)),
        top_T_F=float(top_T_F),
        Z_top=1.0,
        top_vapor_volume_ft3=float(top_vol_f),
        thermo_provider=getattr(inputs, "thermo_provider", None),
        y_top=np.asarray(y_top, dtype=float).reshape((-1,)),
        P_seed_psia=float(p_target),
        return_details=True,
        allow_flash_fallback_on_refine_failure=False,
    )
    if isinstance(p_top_res, tuple):
        p_raw = float(p_top_res[0])
    else:
        p_raw = float(p_top_res)
    if (not np.isfinite(float(p_raw))) or float(p_raw) <= 0.0:
        info["reason"] = "invalid_raw_pressure"
        return y_arr, info

    scale = float(p_target) / float(p_raw)
    if (not np.isfinite(scale)) or scale <= 0.0:
        info["reason"] = "invalid_scale"
        return y_arr, info
    top_v_new = np.clip(top_v * float(scale), 0.0, None)
    # One correction pass accounts for mild Z changes in the pressure helper.
    try:
        p_mid_res = _compute_top_drum_pressure_psia(
            top_V=np.asarray(top_v_new, dtype=float).reshape((-1,)),
            top_T_F=float(top_T_F),
            Z_top=1.0,
            top_vapor_volume_ft3=float(top_vol_f),
            thermo_provider=getattr(inputs, "thermo_provider", None),
            y_top=np.asarray(y_top, dtype=float).reshape((-1,)),
            P_seed_psia=float(p_target),
            return_details=True,
            allow_flash_fallback_on_refine_failure=False,
        )
        p_mid = float(p_mid_res[0] if isinstance(p_mid_res, tuple) else p_mid_res)
        if np.isfinite(p_mid) and p_mid > 0.0:
            scale2 = float(p_target) / float(p_mid)
            if np.isfinite(scale2) and scale2 > 0.0:
                top_v_new = np.clip(top_v_new * float(scale2), 0.0, None)
                scale *= float(scale2)
    except Exception:
        pass
    y_arr[sl["top_V"]] = top_v_new

    info["raw_pressure_initial_psia"] = float(p_raw)
    info["top_vapor_initial_lbmol"] = float(m0)
    info["top_vapor_final_lbmol"] = float(np.sum(np.asarray(y_arr[sl["top_V"]], dtype=float)))
    info["scale_factor"] = float(scale)
    try:
        p_final_res = _compute_top_drum_pressure_psia(
            top_V=np.asarray(top_v_new, dtype=float).reshape((-1,)),
            top_T_F=float(top_T_F),
            Z_top=1.0,
            top_vapor_volume_ft3=float(top_vol_f),
            thermo_provider=getattr(inputs, "thermo_provider", None),
            y_top=np.asarray(y_top, dtype=float).reshape((-1,)),
            P_seed_psia=float(p_target),
            return_details=True,
            allow_flash_fallback_on_refine_failure=False,
        )
        info["raw_pressure_final_psia"] = float(p_final_res[0] if isinstance(p_final_res, tuple) else p_final_res)
    except Exception:
        info["raw_pressure_final_psia"] = np.nan
    info["applied"] = True
    info["reason"] = "applied"
    return y_arr, info


def _linear_trend_slope_per_s(times_s: Sequence[float], values: Sequence[float]) -> float:
    """Least-squares slope estimate d(value)/dt over a time window."""
    if len(times_s) != len(values):
        return float("nan")
    if len(times_s) < 2:
        return float("nan")
    t = np.asarray(times_s, dtype=float).reshape((-1,))
    v = np.asarray(values, dtype=float).reshape((-1,))
    mask = np.isfinite(t) & np.isfinite(v)
    if np.sum(mask) < 2:
        return float("nan")
    t = t[mask]
    v = v[mask]
    if t.size < 2:
        return float("nan")
    t0 = t - float(np.mean(t))
    v0 = v - float(np.mean(v))
    denom = float(np.dot(t0, t0))
    if denom <= 1e-18:
        dt = float(t[-1] - t[0])
        if abs(dt) <= 1e-18:
            return float("nan")
        return float((v[-1] - v[0]) / dt)
    return float(np.dot(t0, v0) / denom)


def _max_rel_inventory_rate_per_s(
    layout: StateVectorLayout,
    y: np.ndarray,
    dydt: np.ndarray,
    *,
    denom_floor_lbmol: float = 1.0,
) -> float:
    """
    Maximum relative rate |dM/dt|/(|M|+denom_floor) over inventory states.
    """
    detail = _max_rel_inventory_rate_detail_per_s(
        layout,
        y,
        dydt,
        denom_floor_lbmol=denom_floor_lbmol,
    )
    return float(detail.get("max_rel_rate_per_s", np.nan))


def _max_rel_inventory_rate_detail_per_s(
    layout: StateVectorLayout,
    y: np.ndarray,
    dydt: np.ndarray,
    *,
    denom_floor_lbmol: float = 1.0,
) -> Dict[str, Any]:
    """
    Detail for maximum relative rate |dM/dt|/(|M|+denom_floor) over inventory states.
    Returns state family and indices so transient bursts can be traced.
    """
    floor = float(denom_floor_lbmol)
    if (not np.isfinite(floor)) or floor < 0.0:
        floor = 1.0
    u = layout.unpack(np.asarray(y, dtype=float).reshape((-1,)))
    ud = layout.unpack(np.asarray(dydt, dtype=float).reshape((-1,)))
    keys = ("tray_L", "tray_V", "top_L", "top_V", "bottom_L", "bottom_V")
    max_rel = np.nan
    best: Dict[str, Any] = {
        "max_rel_rate_per_s": np.nan,
        "state_key": "",
        "stage_1based": np.nan,
        "component_1based": np.nan,
    }
    for key in keys:
        if key not in u or key not in ud:
            continue
        try:
            x_raw = np.asarray(u[key], dtype=float)
            dx_raw = np.asarray(ud[key], dtype=float)
        except Exception:
            continue
        x = np.asarray(x_raw, dtype=float).reshape((-1,))
        dx = np.asarray(dx_raw, dtype=float).reshape((-1,))
        n = min(x.size, dx.size)
        if n <= 0:
            continue
        x = x[:n]
        dx = dx[:n]
        denom = np.abs(x) + float(floor)
        rel = np.abs(dx) / np.maximum(denom, 1e-300)
        finite_mask = np.isfinite(rel)
        if np.sum(finite_mask) <= 0:
            continue
        rel_f = np.where(finite_mask, rel, -np.inf)
        idx_flat = int(np.argmax(rel_f))
        cand = float(rel_f[idx_flat])
        if (not np.isfinite(max_rel)) or cand > max_rel:
            max_rel = cand
            stage_1based = np.nan
            component_1based = np.nan
            if x_raw.ndim == 2:
                try:
                    i_stage, i_comp = np.unravel_index(idx_flat, x_raw.shape)
                    stage_1based = float(i_stage + 1)
                    component_1based = float(i_comp + 1)
                except Exception:
                    pass
            elif key in ("top_L", "top_V"):
                component_1based = float(idx_flat + 1)
                stage_1based = 0.0
            elif key in ("bottom_L", "bottom_V"):
                component_1based = float(idx_flat + 1)
                stage_1based = float(layout.n_stages + 1)
            best = {
                "max_rel_rate_per_s": float(cand),
                "state_key": str(key),
                "stage_1based": stage_1based,
                "component_1based": component_1based,
            }
    return best


def _max_rel_inventory_fd_rate_per_s(
    layout: StateVectorLayout,
    y_prev: np.ndarray,
    y_now: np.ndarray,
    *,
    dt_sec: float,
    denom_floor_lbmol: float = 1.0,
) -> float:
    """
    Maximum relative inventory rate estimated by finite differences:
    |(x_now-x_prev)/dt| / (|x_now| + denom_floor).
    """
    detail = _max_rel_inventory_fd_rate_detail_per_s(
        layout,
        y_prev,
        y_now,
        dt_sec=dt_sec,
        denom_floor_lbmol=denom_floor_lbmol,
    )
    return float(detail.get("max_rel_rate_per_s", np.nan))


def _max_rel_inventory_fd_rate_detail_per_s(
    layout: StateVectorLayout,
    y_prev: np.ndarray,
    y_now: np.ndarray,
    *,
    dt_sec: float,
    denom_floor_lbmol: float = 1.0,
) -> Dict[str, Any]:
    """
    Detail for maximum finite-difference relative inventory rate:
    |(x_now-x_prev)/dt| / (|x_now| + denom_floor).
    """
    try:
        dt = float(dt_sec)
    except Exception:
        return {
            "max_rel_rate_per_s": np.nan,
            "state_key": "",
            "stage_1based": np.nan,
            "component_1based": np.nan,
        }
    if (not np.isfinite(dt)) or dt <= 0.0:
        return {
            "max_rel_rate_per_s": np.nan,
            "state_key": "",
            "stage_1based": np.nan,
            "component_1based": np.nan,
        }

    floor = float(denom_floor_lbmol)
    if (not np.isfinite(floor)) or floor < 0.0:
        floor = 1.0

    u_prev = layout.unpack(np.asarray(y_prev, dtype=float).reshape((-1,)))
    u_now = layout.unpack(np.asarray(y_now, dtype=float).reshape((-1,)))
    keys = ("tray_L", "tray_V", "top_L", "top_V", "bottom_L", "bottom_V")
    max_rel = np.nan
    best: Dict[str, Any] = {
        "max_rel_rate_per_s": np.nan,
        "state_key": "",
        "stage_1based": np.nan,
        "component_1based": np.nan,
    }
    for key in keys:
        if key not in u_prev or key not in u_now:
            continue
        try:
            x0_raw = np.asarray(u_prev[key], dtype=float)
            x1_raw = np.asarray(u_now[key], dtype=float)
        except Exception:
            continue
        x0 = np.asarray(x0_raw, dtype=float).reshape((-1,))
        x1 = np.asarray(x1_raw, dtype=float).reshape((-1,))
        n = min(x0.size, x1.size)
        if n <= 0:
            continue
        x0 = x0[:n]
        x1 = x1[:n]
        rate = (x1 - x0) / dt
        denom = np.abs(x1) + float(floor)
        rel = np.abs(rate) / np.maximum(denom, 1e-300)
        finite_mask = np.isfinite(rel)
        if np.sum(finite_mask) <= 0:
            continue
        rel_f = np.where(finite_mask, rel, -np.inf)
        idx_flat = int(np.argmax(rel_f))
        cand = float(rel_f[idx_flat])
        if (not np.isfinite(max_rel)) or cand > max_rel:
            max_rel = cand
            stage_1based = np.nan
            component_1based = np.nan
            if x1_raw.ndim == 2:
                try:
                    i_stage, i_comp = np.unravel_index(idx_flat, x1_raw.shape)
                    stage_1based = float(i_stage + 1)
                    component_1based = float(i_comp + 1)
                except Exception:
                    pass
            elif key in ("top_L", "top_V"):
                component_1based = float(idx_flat + 1)
                stage_1based = 0.0
            elif key in ("bottom_L", "bottom_V"):
                component_1based = float(idx_flat + 1)
                stage_1based = float(layout.n_stages + 1)
            best = {
                "max_rel_rate_per_s": float(cand),
                "state_key": str(key),
                "stage_1based": stage_1based,
                "component_1based": component_1based,
            }
    return best


def _max_abs_temperature_fd_rate_per_s(
    layout: StateVectorLayout,
    y_prev: np.ndarray,
    y_now: np.ndarray,
    *,
    dt_sec: float,
) -> float:
    """Maximum absolute finite-difference tray temperature rate (F/s)."""
    try:
        dt = float(dt_sec)
    except Exception:
        return float("nan")
    if (not np.isfinite(dt)) or dt <= 0.0:
        return float("nan")

    u_prev = layout.unpack(np.asarray(y_prev, dtype=float).reshape((-1,)))
    u_now = layout.unpack(np.asarray(y_now, dtype=float).reshape((-1,)))
    # Steady-state criterion is defined on tray temperatures only.
    keys = ("tray_T_f",)
    max_rate = np.nan
    for key in keys:
        if key not in u_prev or key not in u_now:
            continue
        try:
            t0 = np.asarray(u_prev[key], dtype=float).reshape((-1,))
            t1 = np.asarray(u_now[key], dtype=float).reshape((-1,))
        except Exception:
            continue
        n = min(t0.size, t1.size)
        if n <= 0:
            continue
        rate = np.abs((t1[:n] - t0[:n]) / dt)
        rate_f = rate[np.isfinite(rate)]
        if rate_f.size <= 0:
            continue
        cand = float(np.max(rate_f))
        if (not np.isfinite(max_rate)) or cand > max_rate:
            max_rate = cand
    return float(max_rate)


# -------------------------
# Stream extraction for logs
# -------------------------


@dataclass(frozen=True)
class StreamTag:
    name: str
    flow_lbmolph: Optional[float]
    stage_1based: Optional[int]


@dataclass
class PIController:
    kc: float
    ti_sec: float
    bias: float
    out_min: float
    out_max: float
    integ: float = 0.0


def _as_bool(x: Any) -> Optional[bool]:
    if isinstance(x, bool):
        return x
    if x is None:
        return None
    s = str(x).strip().lower()
    if s in ("1", "true", "yes", "y", "on"):
        return True
    if s in ("0", "false", "no", "n", "off"):
        return False
    return None


def _pi_update(
    controller: PIController,
    *,
    pv: float,
    sp: float,
    dt_sec: float,
    out_min: Optional[float] = None,
    out_max: Optional[float] = None,
    debug: Optional[Dict[str, float]] = None,
) -> float:
    """
    One-step PI with anti-windup (conditional integration).
    """
    e = float(pv) - float(sp)
    if not np.isfinite(e):
        e = 0.0
    if not np.isfinite(controller.kc):
        controller.kc = 0.0
    if not np.isfinite(controller.ti_sec) or controller.ti_sec <= 0.0:
        controller.ti_sec = 1e12
    if not np.isfinite(controller.integ):
        controller.integ = 0.0

    umin = float(controller.out_min if out_min is None else out_min)
    umax = float(controller.out_max if out_max is None else out_max)
    if not np.isfinite(umin):
        umin = float(controller.out_min)
    if not np.isfinite(umax):
        umax = float(controller.out_max)
    if umax < umin:
        umax = umin

    # Tentative unclamped output
    i_next = controller.integ + e * float(dt_sec) / max(float(controller.ti_sec), 1e-9)
    u_unclamped = float(controller.bias) + float(controller.kc) * (e + i_next)
    u = float(np.clip(u_unclamped, umin, umax))

    # Anti-windup: accept integrator update only when output is unsaturated,
    # or integration would move the output back toward the unsaturated region.
    # This must account for gain sign (reverse-acting loops use negative Kc).
    sat_hi = u_unclamped > umax + 1e-12
    sat_lo = u_unclamped < umin - 1e-12
    du_from_int = float(controller.kc) * float(e)
    allow_int = (not sat_hi and not sat_lo) or (sat_hi and du_from_int < 0.0) or (sat_lo and du_from_int > 0.0)
    if allow_int:
        controller.integ = float(i_next)
        u = float(np.clip(float(controller.bias) + float(controller.kc) * (e + controller.integ), umin, umax))
    if debug is not None:
        debug.clear()
        debug.update(
            {
                "error": float(e),
                "p_term": float(controller.kc) * float(e),
                "i_term": float(controller.kc) * float(controller.integ),
                "u_unclamped": float(u_unclamped),
                "u_clamped": float(u),
                "out_min": float(umin),
                "out_max": float(umax),
                "bias": float(controller.bias),
                "kc": float(controller.kc),
                "ti_sec": float(controller.ti_sec),
                "integ": float(controller.integ),
                "sat_hi": 1.0 if sat_hi else 0.0,
                "sat_lo": 1.0 if sat_lo else 0.0,
                "allow_int": 1.0 if allow_int else 0.0,
            }
        )
    return u


def _pressure_resid_gain_scale(
    *,
    resid_abs_btups: Optional[float],
    resid_ref_btups: Optional[float],
    min_gain: float = 0.25,
) -> float:
    """
    Map a pressure-loop energy residual magnitude to a PI gain scale in [min_gain, 1].
    """
    if resid_ref_btups is None:
        return 1.0
    try:
        ref = float(resid_ref_btups)
    except Exception:
        return 1.0
    if (not np.isfinite(ref)) or ref <= 0.0:
        return 1.0

    try:
        resid = float(resid_abs_btups) if resid_abs_btups is not None else np.nan
    except Exception:
        resid = np.nan
    if (not np.isfinite(resid)) or resid <= 0.0:
        return 1.0

    g_min = float(np.clip(float(min_gain), 0.0, 1.0))
    g = 1.0 / (1.0 + (resid / ref))
    if not np.isfinite(g):
        return 1.0
    return float(np.clip(g, g_min, 1.0))


def _apply_slew_limit(
    *,
    cmd: float,
    prev_cmd: Optional[float],
    rate_limit_per_s: Optional[float],
    dt_sec: float,
) -> float:
    """
    Symmetric rate limiter: |cmd - prev_cmd| <= rate_limit_per_s * dt_sec.
    """
    try:
        u = float(cmd)
    except Exception:
        return cmd
    if not np.isfinite(u):
        return u
    if prev_cmd is None or rate_limit_per_s is None:
        return u
    try:
        u_prev = float(prev_cmd)
        r = float(rate_limit_per_s)
        dt = float(dt_sec)
    except Exception:
        return u
    if (not np.isfinite(u_prev)) or (not np.isfinite(r)) or (not np.isfinite(dt)):
        return u
    if r <= 0.0 or dt <= 0.0:
        return u
    du_max = abs(r) * dt
    return float(np.clip(u, u_prev - du_max, u_prev + du_max))


def _extract_stream_from_obj_or_dict(obj: Any) -> Tuple[Optional[float], Optional[int]]:
    """
    Extract (total_molar_flow_lbmolph, stage_1based) from either:
      - an object with attributes
      - a dict with known keys
    """
    if obj is None:
        return (None, None)

    # object-like
    flow = None
    stage = None
    for attr in ("total_molar_flow_lbmolph", "flow_lbmolph", "flow", "TotalMolarFlow"):
        if hasattr(obj, attr):
            flow = _as_float(getattr(obj, attr))
            if flow is not None:
                break
    for attr in ("stage_1based", "stage", "Stage", "StageNumber"):
        if hasattr(obj, attr):
            stage = _as_int(getattr(obj, attr))
            if stage is not None:
                break

    # dict-like
    if isinstance(obj, dict):
        flow = flow if flow is not None else _as_float(
            _get_first_mapping_value(
                obj,
                (
                    "total_molar_flow_lbmolph",
                    "Total Molar Flow (lbmol/h)",
                    "Total Flow (lbmol/h)",
                    "flow_lbmolph",
                    "flow",
                ),
            )
        )
        stage = stage if stage is not None else _as_int(
            _get_first_mapping_value(
                obj,
                (
                    "stage_1based",
                    "Stage (1-based)",
                    "Stage",
                    "stage",
                ),
            )
        )

    return (flow, stage)


def _lookup_named_stream(
    *,
    col: ColumnSpec,
    case: CaseData,
    aliases: Sequence[str],
) -> StreamTag:
    """
    Find a stream by name (case-insensitive match against aliases).
    Searches ColumnSpec.streams first, then CaseData.streams.
    Returns StreamTag(name, flow_lbmolph, stage_1based).
    """
    alias_lc = [str(a).strip().lower() for a in aliases if str(a).strip()]
    alias_clean = [''.join(ch for ch in a if ch.isalnum()) for a in alias_lc]
    if not alias_lc:
        return StreamTag(name=str(aliases[0]) if aliases else "stream", flow_lbmolph=None, stage_1based=None)

    def match(nm: str) -> bool:

        nmlc = str(nm).strip().lower()

        def _clean(s: str) -> str:
            # Keep only [a-z0-9] to make comparisons robust to spaces, dashes, etc.
            return ''.join(ch for ch in s.lower() if ch.isalnum())

        nmc = _clean(nmlc)

        # 1) Exact match (strict) on both raw and cleaned forms
        if nmlc in alias_lc or nmc in alias_clean:
            return True

        # 2) Relaxed matching ONLY for non-trivial aliases.
        # Short aliases (e.g. "D", "B") must be exact; otherwise they can match "FeeD", "Bottoms", etc.
        for a_raw, a_clean in zip(alias_lc, alias_clean):
            if len(a_clean) < 3:
                continue
            if nmc.startswith(a_clean):
                return True
        for a_raw, a_clean in zip(alias_lc, alias_clean):
            if len(a_clean) < 3:
                continue
            if a_clean in nmc:
                return True
        return False

    # 1) ColumnSpec.streams
    streams = getattr(col, "streams", None)
    if isinstance(streams, dict):
        for nm, obj in streams.items():
            if match(str(nm)):
                f, st = _extract_stream_from_obj_or_dict(obj)
                return StreamTag(name=str(nm), flow_lbmolph=f, stage_1based=st)

    # 2) CaseData.streams
    streams2 = getattr(case, "streams", None)
    if isinstance(streams2, dict):
        for nm, obj in streams2.items():
            if match(str(nm)):
                f, st = _extract_stream_from_obj_or_dict(obj)
                return StreamTag(name=str(nm), flow_lbmolph=f, stage_1based=st)

    return StreamTag(name=str(aliases[0]), flow_lbmolph=None, stage_1based=None)


def _lookup_spec_int(col: ColumnSpec, keys: Sequence[str]) -> Optional[int]:
    d = getattr(col, "specs_raw", None)
    if not isinstance(d, dict):
        return None
    for k in keys:
        if k in d:
            v = _as_int(d[k])
            if v is not None:
                return v
    return None


def _geometry_has_explicit_hydraulic_c_factor(col: ColumnSpec) -> bool:
    specs = getattr(col, "specs_raw", None) or {}
    sections = specs.get("Geometry Sections", None)
    if not isinstance(sections, list):
        return False
    for row in sections:
        if not isinstance(row, dict):
            continue
        for key in ("hydraulic_c_factor", "system_factor"):
            val = row.get(key, None)
            try:
                f = float(val)
            except Exception:
                continue
            if np.isfinite(f) and f > 0.0:
                return True
    return False


def _autocalibrate_francis_hydraulic_c_factors_from_seed(
    *,
    col: ColumnSpec,
    thermo_provider: Any,
) -> bool:
    """
    Fit per-stage Francis hydraulic C multipliers so the hydraulic liquid-flow
    diagnostic matches the seeded tray liquid profile at t=0.

    This is only intended when the workbook did not already provide explicit
    hydraulic/system factors. It updates the in-memory frozen ColumnSpec via
    object.__setattr__ so the rest of the runner sees the calibrated geometry.
    """
    geom = getattr(col, "geometry", None)
    if geom is None:
        return False
    if _geometry_has_explicit_hydraulic_c_factor(col):
        return False

    try:
        N = int(col.n_stages)
        ML = np.asarray(col.M_L_lbmol, dtype=float).reshape((N,))
        T = np.asarray(col.T_f, dtype=float).reshape((N,))
        P = np.asarray(col.P_psia, dtype=float).reshape((N,))
        x = np.asarray(col.x0, dtype=float).reshape((N, int(col.n_components)))
        L_target = np.asarray(col.L_lbmolph, dtype=float).reshape((N,))
        active_area = np.asarray(geom.active_area_ft2_per_stage, dtype=float).reshape((N,))
        holdup_area = np.asarray(geom.area_ft2_per_stage, dtype=float).reshape((N,))
        weir_h = np.asarray(geom.weir_height_in_per_stage, dtype=float).reshape((N,))
        weir_L = np.asarray(geom.weir_length_ft_per_stage, dtype=float).reshape((N,))
    except Exception:
        return False

    if (
        ML.size != N
        or T.size != N
        or P.size != N
        or L_target.size != N
        or active_area.size != N
        or holdup_area.size != N
        or weir_h.size != N
        or weir_L.size != N
    ):
        return False

    rhoL = np.full(N, np.nan, dtype=float)
    for i in range(N):
        try:
            rho_try = thermo_provider.liquid_density_lbmol_ft3(float(T[i]), float(P[i]), x[i, :].tolist())
            rho_try = float(rho_try)
        except Exception:
            rho_try = np.nan
        if np.isfinite(rho_try) and rho_try > 0.0:
            rhoL[i] = rho_try

    if not np.any(np.isfinite(rhoL[1:max(N - 1, 1)]) & (rhoL[1:max(N - 1, 1)] > 0.0)):
        return False

    try:
        hyd_base = compute_francis_weir_liquid_outflow(
            ML_lbmol=ML,
            rhoL_lbmol_ft3=rhoL,
            active_area_ft2=active_area,
            holdup_area_ft2=holdup_area,
            weir_height_in=weir_h,
            weir_length_ft=weir_L,
            c_multiplier=None,
        )
    except Exception:
        return False

    L_base = np.asarray(hyd_base.ML_lbmolph, dtype=float).reshape((N,))
    cfac_old = np.asarray(getattr(geom, "hydraulic_c_factor_per_stage", np.ones(N, dtype=float)), dtype=float).reshape((N,))
    cfac_new = np.asarray(cfac_old, dtype=float).copy()

    fitted = False
    for i in range(1, N - 1):
        target = float(L_target[i])
        base = float(L_base[i])
        if (
            np.isfinite(target)
            and target > 0.0
            and np.isfinite(base)
            and base > 1.0e-12
        ):
            cfit = target / base
            if np.isfinite(cfit) and cfit > 0.0:
                cfac_new[i] = float(np.clip(cfit, 1.0e-6, 1.0e6))
                fitted = True

    if not fitted:
        return False

    new_geom = replace(geom, hydraulic_c_factor_per_stage=cfac_new)
    object.__setattr__(col, "geometry", new_geom)
    return True


def _stage_value(i_1based: int, target_stage_1based: Optional[int], value: Optional[float]) -> float:
    """
    Returns value on the target stage, else 0.0.
    If target stage exists but value missing -> NaN (signals incomplete stream spec).
    """
    if target_stage_1based is None:
        return 0.0
    if int(i_1based) != int(target_stage_1based):
        return 0.0
    return float(value) if value is not None else float("nan")


def _normalized_component_holdup_comp(
    comp_holdup: Optional[np.ndarray],
    n_components: int,
) -> Optional[np.ndarray]:
    if comp_holdup is None:
        return None
    try:
        arr = np.asarray(comp_holdup, dtype=float).reshape((n_components,))
    except Exception:
        return None
    if not np.all(np.isfinite(arr)):
        return None
    total = float(np.sum(arr))
    if total <= 1e-12:
        return None
    return arr / total


def _composition_tv_distance(a: Optional[np.ndarray], b: Optional[np.ndarray]) -> float:
    if a is None or b is None:
        return float("nan")
    try:
        aa = np.asarray(a, dtype=float).reshape((-1,))
        bb = np.asarray(b, dtype=float).reshape((-1,))
    except Exception:
        return float("nan")
    if aa.shape != bb.shape or not np.all(np.isfinite(aa)) or not np.all(np.isfinite(bb)):
        return float("nan")
    return 0.5 * float(np.sum(np.abs(aa - bb)))


# -------------------------
# Thermo provider stub
# -------------------------


class StubThermoProvider:
    """Deterministic stub provider.

    stage_thermo_v1.flash_TP_full_F_psia() accepts tuple/list:
      (x, y, K, HL, HV) or (x, y, K, HL, HV, Z)
    """

    def __init__(self, K: Sequence[float], Z: float = 1.0):
        self._K = np.asarray(K, dtype=float).ravel()
        self._Z = float(Z)

    def flash_TP_full_F_psia(self, T_F: float, P_psia: float, z: List[float]):
        z_arr = _normalize_comp(np.asarray(z, dtype=float))
        K = self._K.copy()
        x = z_arr
        y = _normalize_comp(K * x)
        HL = 0.0
        HV = 0.0
        return (x.tolist(), y.tolist(), K.tolist(), HL, HV, float(self._Z))

    def flash_TP_full(self, T_F: float, P_psia: float, z: List[float]):
        return self.flash_TP_full_F_psia(T_F, P_psia, z)

    def liquid_density_lbmol_ft3(self, T_F: float, P_psia: float, x: List[float]) -> float:
        # Simple stub: constant density to keep hydraulics running in stub mode.
        return 1.0

    def cp_liq_vap_btu_per_lbmolF(self, T_F: float, P_psia: float, z: List[float]):
        # Simple stub: constant Cp values to keep energy balance running in stub mode.
        return (30.0, 20.0)


# -------------------------
# Runner config
# -------------------------


@dataclass(frozen=True)
class RunnerConfig:
    excel_path: str
    n_steps: int = 600
    dt_sec: Optional[float] = None
    log_every_n_steps: Optional[int] = None
    runtime_mode: str = "legacy"  # legacy | parity | calibration | hydraulic | total-reflux
    integrator: str = "explicit-euler"  # explicit-euler | bdf | radau | ida
    integrator_rtol: float = 1.0e-3
    integrator_atol: float = 1.0e-6
    integrator_max_step_sec: Optional[float] = None
    integrator_substep_sec: Optional[float] = None
    integrator_max_rhs_evals_per_step: Optional[int] = 24
    integrator_step_wall_limit_sec: Optional[float] = 15.0
    ida_max_iter: int = 8
    ida_relax: float = 1.0

    include_temperature: bool = True
    include_energy: bool = False
    include_boundary_states: bool = True
    include_vapor_states: bool = True
    enable_equilibrium_relaxation: bool = True
    equilibrium_relaxation_mode: str = "auto"  # auto | phase-holdup | composition-only
    equilibrium_tau_sec: Optional[float] = None
    equilibrium_tau_ramp_initial_sec: Optional[float] = None
    equilibrium_tau_ramp_final_sec: Optional[float] = None
    equilibrium_tau_ramp_decay_sec: Optional[float] = None
    equilibrium_phase_holdup_guard_lbmol: Optional[float] = None
    equilibrium_energy_damping_gain: Optional[float] = None
    hydraulic_energy_temperature_damping: Optional[float] = None
    hydraulic_energy_temperature_mode: Optional[str] = None
    hydraulic_energy_temperature_follow_tau_sec: Optional[float] = None
    hydraulic_energy_temperature_resid_frac: Optional[float] = None
    hydraulic_energy_temperature_pressure_slope_F_per_psi: Optional[float] = None
    hydraulic_energy_temperature_target_refresh_steps: Optional[int] = None

    thermo_mode: str = "table-pool"  # 'stub', 'relative-volatility', 'clapeyron', 'dwsim', 'table', or 'table-pool'
    clapeyron_model: str = "PR"
    clapeyron_ideal_model: Optional[str] = None
    clapeyron_pr_parameter_source: str = "default"  # default | dwsim
    dwsim_property_package: str = "pr"
    thermo_every_n_steps: int = 1  # 1=every step
    enable_thermo_cadence_guardrails: bool = True
    thermo_refresh_dT_F: Optional[float] = None
    thermo_refresh_dP_psia: Optional[float] = None
    thermo_refresh_dx: Optional[float] = None
    equilibrium_relaxation_live_pr: bool = False
    flash_feed_at_stage_conditions: Optional[bool] = None
    thermo_table_path: Optional[str] = r"cache/thermo_table.json"
    thermo_top_saturation_table_path: Optional[str] = None
    thermo_upper_section_table_path: Optional[str] = None
    thermo_upper_section_stage_count: int = 5
    thermo_table_n_anchor_blend: int = 3
    thermo_table_anchor_blend_power: float = 2.0
    thermo_pool_workers: Optional[int] = 2
    thermo_pool_chunk_size: int = 4
    thermo_pool_task_timeout_sec: Optional[float] = None
    reboiler_neighbor_vflow_hi_ratio: Optional[float] = None
    reboiler_neighbor_vflow_lo_ratio: Optional[float] = None
    vapor_holdup_relaxation_sec: Optional[float] = None
    debug_freeze_tray_vapor_derivatives: bool = False
    debug_override_reflux_composition: bool = False
    debug_clamp_top_drum_pressure_psia: Optional[float] = None
    debug_clamp_top_drum_pressure_duration_sec: Optional[float] = None
    startup_total_reflux_washout_sec: Optional[float] = None
    total_reflux_startup_ramp_tau_sec: Optional[float] = None
    total_reflux_startup_min_ramp_fraction: float = 0.0
    total_reflux_scale_reflux_with_startup_factor: bool = False
    total_reflux_boundary_ramp_duration_sec: Optional[float] = None
    hydraulic_pressure_relaxation_sec: Optional[float] = None
    top_drum_pressure_temperature_relaxation_sec: Optional[float] = None
    vapor_flow_relaxation_sec: Optional[float] = None
    conductance_vflow_nominal_hi_ratio: Optional[float] = None
    # Optional smooth-clamp width (lbmol/h) used only for stiff integrator RHS
    # regularization in hydraulic vapor-flow closures.
    # None -> auto (small default in stiff hydraulic runs), <=0 disables.
    stiff_vflow_smooth_clamp_lbmolph: Optional[float] = None
    pv_inner_max_iter: int = 1
    pv_inner_p_tol_psia: Optional[float] = 0.05
    pv_inner_v_tol_lbmolph: Optional[float] = 25.0
    enable_dae_pilot_algebraic_solve: bool = False
    dae_pilot_max_iter: int = 3
    dae_pilot_p_tol_psia: Optional[float] = 0.05
    dae_pilot_v_tol_lbmolph: Optional[float] = 25.0
    dae_pilot_jac_rel_step: float = 1.0e-6
    dae_pilot_line_search_max: int = 4
    enable_liquid_hydraulic_override: Optional[bool] = None
    liquid_hydraulic_override_alpha: Optional[float] = None
    liquid_hydraulic_model: Optional[str] = None
    liquid_hydraulic_htc_sec: Optional[float] = None

    reflux_lbmolph: Optional[float] = None
    boilup_lbmolph: Optional[float] = None
    condenser_duty_mode: str = "total-condense"
    condenser_duty_btu_per_h: Optional[float] = None
    condenser_duty_trim_btu_per_h: Optional[float] = None
    init_pack_top_drum_vapor_to_pressure: bool = False
    init_top_drum_vapor_pressure_psia: Optional[float] = None
    init_match_condenser_duty: bool = False
    init_align_top_liquid_to_condensate: bool = False
    condenser_pressure_drop_psi: Optional[float] = None
    top_drum_vapor_volume_ft3: Optional[float] = None
    top_drum_total_volume_ft3: Optional[float] = None
    enforce_top_drum_pressure_gate: bool = True
    top_drum_pressure_gate_soft_psi: Optional[float] = None
    enforce_top_pressure_ordering: bool = True
    top_pressure_ordering_margin_psi: float = 0.0
    enable_top_psv: bool = False
    top_psv_setpoint_psia: Optional[float] = None
    top_psv_gain_lbmolps_per_psi: Optional[float] = None
    top_psv_max_vent_lbmolps: Optional[float] = None
    enable_level_control: bool = False
    top_level_pv_mode: str = "molar-holdup"  # molar-holdup|true-level
    ignore_workbook_level_pv_mode: bool = False
    top_level_sp_lbmol: Optional[float] = None
    top_level_sp_frac: Optional[float] = None
    bottom_level_pv_mode: str = "molar-holdup"  # molar-holdup|true-level
    bottom_level_sp_lbmol: Optional[float] = None
    bottom_level_sp_frac: Optional[float] = None
    top_level_kc: Optional[float] = None
    top_level_ti_sec: Optional[float] = None
    enable_top_level_feedforward: bool = False
    top_level_feedforward_gain: Optional[float] = None
    bottom_level_kc: Optional[float] = None
    bottom_level_ti_sec: Optional[float] = None
    bottom_sump_total_volume_ft3: Optional[float] = None
    enable_pressure_control: bool = False
    pressure_control_mv: str = "auto"  # auto|condenser-duty|top-anchor
    allow_coupled_pressure_duty: bool = False
    enable_coupled_total_condenser_partial_condense: bool = True
    top_pressure_sp_psia: Optional[float] = None
    top_pressure_kc: Optional[float] = None
    top_pressure_ti_sec: Optional[float] = None
    top_pressure_pv_filter_tau_sec: Optional[float] = None
    top_pressure_mv_slew_limit_per_s: Optional[float] = None
    top_pressure_resid_ref_btups: Optional[float] = None
    top_pressure_resid_min_gain: float = 0.25
    condenser_duty_min_btu_per_h: Optional[float] = None
    condenser_duty_max_btu_per_h: Optional[float] = None
    top_pressure_anchor_min_psia: Optional[float] = None
    top_pressure_anchor_max_psia: Optional[float] = None
    enable_distillate_composition_control: bool = False
    distillate_composition_component: str = "C4"
    distillate_composition_sp_molfrac: Optional[float] = None
    distillate_composition_kc: Optional[float] = None
    distillate_composition_ti_sec: Optional[float] = None
    reflux_cmd_min_lbmolph: Optional[float] = None
    reflux_cmd_max_lbmolph: Optional[float] = None
    enable_reflux_feasibility_cap: bool = True
    reflux_ratio_min: Optional[float] = None
    reflux_ratio_max: Optional[float] = None
    enable_bottoms_composition_control: bool = False
    bottoms_composition_component: str = "C5"
    bottoms_composition_sp_molfrac: Optional[float] = None
    bottoms_composition_kc: Optional[float] = None
    bottoms_composition_ti_sec: Optional[float] = None
    bottoms_composition_mv: str = "boilup"  # boilup|reboiler-duty
    boilup_cmd_min_lbmolph: Optional[float] = None
    boilup_cmd_max_lbmolph: Optional[float] = None
    reboiler_duty_cmd_min_btu_per_h: Optional[float] = None
    reboiler_duty_cmd_max_btu_per_h: Optional[float] = None
    reboiler_duty_btu_per_h: Optional[float] = None
    feed_step_time_sec: Optional[float] = None
    feed_step_scale: Optional[float] = None

    logs_dir: str = "logs"
    run_name: Optional[str] = None
    run_description: Optional[str] = None
    write_logs: bool = True
    use_excel_vapor_holdup: bool = False
    fast_startup: bool = False
    enable_startup_seed_cache: bool = False
    refresh_startup_seed_cache: bool = False
    startup_seed_cache_path: Optional[str] = None
    enable_primary_thermo_startup_prewarm: bool = False
    enable_startup_thermo_conditioning: bool = True
    startup_thermo_conditioning_iters: int = 2
    startup_thermo_conditioning_relaxation: float = 1.0
    enable_startup_hydraulic_energy_consistency: bool = False
    startup_hydraulic_energy_consistency_iters: int = 6
    startup_hydraulic_energy_consistency_dt_sec: float = 0.5
    startup_hydraulic_energy_consistency_mass_tol_lbmolph: Optional[float] = 5.0
    startup_hydraulic_energy_consistency_energy_tol_btups: Optional[float] = 1000.0
    enable_startup_hydraulic_sequence: bool = False
    startup_sequence_energy_on_sec: float = 30.0
    startup_sequence_liquid_on_sec: float = 120.0
    startup_sequence_liquid_ramp_sec: float = 180.0
    startup_sequence_mass_resid_gate_lbmolph: Optional[float] = 250.0
    startup_sequence_liquid_backoff_sec: Optional[float] = None
    enable_startup_vapor_homotopy: bool = False
    startup_sequence_profile_hold_sec: float = 0.0
    startup_sequence_vapor_on_sec: Optional[float] = None
    startup_sequence_vapor_ramp_sec: float = 60.0
    startup_sequence_vapor_rel_rate_gate_per_s: Optional[float] = 1.0e-2
    startup_sequence_vapor_backoff_sec: Optional[float] = None
    fast_startup_thermo_conditioning_iters: int = 1
    fast_startup_thermo_conditioning_relaxation: float = 1.0
    fast_startup_top_drum_max_iter: int = 2
    fast_startup_top_drum_tol_lbmolps: float = 1.0e-4
    fast_startup_top_drum_wall_limit_sec: float = 30.0
    enable_restart_reentry_settling: bool = True
    restart_reentry_thermo_conditioning_iters: int = 1
    restart_reentry_thermo_conditioning_relaxation: float = 1.0
    restart_reentry_top_drum_max_iter: int = 2
    restart_reentry_top_drum_tol_lbmolps: float = 1.0e-4
    restart_reentry_top_drum_wall_limit_sec: float = 10.0
    enable_restart_reentry_dynamic_warmup: bool = False
    restart_reentry_dynamic_warmup_sec: float = 5.0

    # Runtime steady-state detector (diagnostic-only)
    enable_steady_state_detection: bool = True
    steady_state_window_sec: float = 30.0
    steady_state_min_time_sec: float = 60.0
    steady_state_rel_state_rate_tol_per_s: Optional[float] = 3.0e-3
    steady_state_kpi_slope_tol_per_s: Optional[float] = 1.0e-4
    steady_state_mv_rate_tol_per_s: Optional[float] = 20.0
    steady_state_temp_rate_tol_F_per_s: Optional[float] = 0.15
    steady_state_sp_error_tol: Optional[float] = 0.02
    steady_state_require_sp: bool = False
    steady_state_rate_denom_floor_lbmol: float = 1.0


def _resolve_startup_execution_flags(cfg: RunnerConfig) -> Dict[str, bool]:
    """
    Resolve which expensive startup conditioning passes should run.

    `fast_startup` is a conservative shortcut mode intended to reduce wall-clock
    spent before integration begins without changing the main runtime path.
    In fast mode it:
    - skips startup thermo-consistent conditioning
    - skips bounded hydraulic-energy startup consistency relaxation
    - skips top-drum startup steadying
    to minimize pre-integration overhead.
    """
    fast = bool(getattr(cfg, "fast_startup", False))
    runtime_mode = _normalize_runtime_mode(getattr(cfg, "runtime_mode", None), default="legacy")
    parity_mode = runtime_mode == "parity"
    parity_like = runtime_mode in ("parity", "calibration")
    startup_thermo_enabled = (
        bool(getattr(cfg, "enable_startup_thermo_conditioning", False))
        and (not parity_mode)
        and (not fast)
    )
    return {
        "fast_startup": fast,
        "enable_startup_thermo_conditioning": startup_thermo_enabled,
        "startup_thermo_conditioning_iters": (
            int(getattr(cfg, "fast_startup_thermo_conditioning_iters", 1))
            if fast
            else int(getattr(cfg, "startup_thermo_conditioning_iters", 2))
        ),
        "startup_thermo_conditioning_relaxation": (
            float(getattr(cfg, "fast_startup_thermo_conditioning_relaxation", 1.0))
            if fast
            else float(getattr(cfg, "startup_thermo_conditioning_relaxation", 1.0))
        ),
        "enable_startup_hydraulic_energy_consistency": bool(
            getattr(cfg, "enable_startup_hydraulic_energy_consistency", False)
        ) and (not fast),
        "enable_top_drum_startup_steadying": ((not parity_like) and startup_thermo_enabled),
        "top_drum_steady_max_iter": (
            int(getattr(cfg, "fast_startup_top_drum_max_iter", 2)) if fast else 6
        ),
        "top_drum_steady_tol_lbmolps": (
            float(getattr(cfg, "fast_startup_top_drum_tol_lbmolps", 1.0e-4)) if fast else 1.0e-6
        ),
        "top_drum_steady_wall_limit_sec": (
            float(getattr(cfg, "fast_startup_top_drum_wall_limit_sec", 30.0)) if fast else None
        ),
        "enable_restart_reentry_settling": bool(getattr(cfg, "enable_restart_reentry_settling", True)),
        "restart_reentry_thermo_conditioning_iters": int(
            getattr(cfg, "restart_reentry_thermo_conditioning_iters", 1)
        ),
        "restart_reentry_thermo_conditioning_relaxation": float(
            getattr(cfg, "restart_reentry_thermo_conditioning_relaxation", 1.0)
        ),
        "restart_reentry_top_drum_max_iter": int(getattr(cfg, "restart_reentry_top_drum_max_iter", 2)),
        "restart_reentry_top_drum_tol_lbmolps": float(
            getattr(cfg, "restart_reentry_top_drum_tol_lbmolps", 1.0e-4)
        ),
        "restart_reentry_top_drum_wall_limit_sec": float(
            getattr(cfg, "restart_reentry_top_drum_wall_limit_sec", 10.0)
        ),
    }
def _resolve_parity_runtime_thermo_defer_visible_steps(cfg: RunnerConfig, *, log_every_n_steps: int) -> int:
    runtime_mode = _normalize_runtime_mode(getattr(cfg, "runtime_mode", None), default="legacy")
    if runtime_mode != "parity":
        return 0
    try:
        log_every = int(log_every_n_steps)
    except Exception:
        log_every = 1
    return max(1, log_every)


# -------------------------
# Boundary + volume model
# -------------------------


def _case_stream_lookup(case: CaseData, name: str) -> Optional[Dict[str, Any]]:
    if not getattr(case, "streams", None) or not isinstance(case.streams, dict):
        return None
    target = name.strip().lower()
    for k, v in case.streams.items():
        if str(k).strip().lower() == target:
            return v if isinstance(v, dict) else None
    return None


def _stream_total_flow_lbmolph(stream: Dict[str, Any]) -> Optional[float]:
    for key in (
        "Total Molar Flow (lbmol/h)",
        "total molar flow (lbmol/h)",
        "total_molar_flow_lbmolph",
        "Total Flow (lbmol/h)",
    ):
        if key in stream:
            try:
                return float(stream[key])
            except Exception:
                return None
    return None


def _infer_boundary_flows(case: CaseData, col: ColumnSpec, cfg: RunnerConfig) -> BoundaryFlows:
    reflux = cfg.reflux_lbmolph
    boilup = cfg.boilup_lbmolph

    if reflux is None:
        s = _case_stream_lookup(case, "Reflux")
        if s is not None:
            reflux = _stream_total_flow_lbmolph(s)

    if boilup is None:
        s = _case_stream_lookup(case, "Boilup")
        if s is not None:
            boilup = _stream_total_flow_lbmolph(s)

    if reflux is None:
        reflux = float(col.L_lbmolph[0])
    if boilup is None:
        boilup = float(col.V_lbmolph[-1])

    return BoundaryFlows(reflux_lbmolph=float(reflux), boilup_lbmolph=float(boilup))


def _build_volume_model(col: ColumnSpec, default_vapor_volume_ft3: float = 1.0) -> VolumeModel:
    vv = None

    # Prefer geometry container
    geom = getattr(col, "geometry", None)
    if geom is not None:
        vv = getattr(geom, "vapor_volume_ft3_per_stage", None)

    # Fallback
    if vv is None:
        vv = getattr(col, "vapor_volume_ft3_per_stage", None)

    if vv is not None:
        vv = np.asarray(vv, dtype=float).reshape((col.n_stages,)).copy()

    return VolumeModel(vapor_volume_ft3_per_stage=vv, default_vapor_volume_ft3=float(default_vapor_volume_ft3))


def _normalize_spec_key(key: str) -> str:
    return "".join(ch for ch in str(key).lower() if ch.isalnum())


def _spec_get(specs: Dict[str, Any], key: str, *aliases: str) -> Any:
    if key in specs:
        return specs[key]

    norm_map = {_normalize_spec_key(k): v for k, v in specs.items()}

    norm = _normalize_spec_key(key)
    if norm:
        if norm in norm_map:
            return norm_map[norm]

    for alias in aliases:
        if alias in specs:
            return specs[alias]
        alias_norm = _normalize_spec_key(alias)
        if alias_norm and alias_norm in norm_map:
            return norm_map[alias_norm]
    return None


def _spec_float(specs: Dict[str, Any], key: str, *aliases: str) -> Optional[float]:
    v_raw = _spec_get(specs, key, *aliases)
    if v_raw is None:
        return None
    try:
        v = float(v_raw)
        if not np.isfinite(v):
            return None
        return v
    except Exception:
        return None


def _spec_bool(specs: Dict[str, Any], key: str, *aliases: str) -> Optional[bool]:
    v_raw = _spec_get(specs, key, *aliases)
    if v_raw is None:
        return None
    if isinstance(v_raw, bool):
        return bool(v_raw)
    txt = str(v_raw).strip().lower()
    if txt in {"1", "true", "yes", "y", "on", "enabled"}:
        return True
    if txt in {"0", "false", "no", "n", "off", "disabled"}:
        return False
    return None


def _normalize_startup_comp_row(comp: Any, *, n_components: int) -> np.ndarray:
    arr = np.asarray(comp, dtype=float).reshape((int(n_components),))
    s = float(np.sum(arr))
    if np.isfinite(s) and s > 1.0e-300:
        return arr / s
    return np.ones((int(n_components),), dtype=float) / float(max(int(n_components), 1))


def _collect_startup_prewarm_rows(col: ColumnSpec, *, max_rows: int = 3) -> list[tuple[float, float, list[float]]]:
    n_stages = int(getattr(col, "n_stages", 0) or 0)
    n_components = int(getattr(col, "n_components", 0) or 0)
    if n_stages <= 0 or n_components <= 0:
        return []
    try:
        T = np.asarray(getattr(col, "T_f"), dtype=float).reshape((n_stages,))
        P = np.asarray(getattr(col, "P_psia"), dtype=float).reshape((n_stages,))
        x = np.asarray(getattr(col, "x0"), dtype=float).reshape((n_stages, n_components))
    except Exception:
        return []

    row_indices: list[int] = []
    for idx in (0, max(0, n_stages // 2), max(0, n_stages - 1)):
        idx_i = int(np.clip(int(idx), 0, max(n_stages - 1, 0)))
        if idx_i not in row_indices:
            row_indices.append(idx_i)
    if max_rows > 0:
        row_indices = row_indices[: int(max_rows)]

    out: list[tuple[float, float, list[float]]] = []
    for idx in row_indices:
        T_i = float(T[idx])
        P_i = float(P[idx])
        if (not np.isfinite(T_i)) or (not np.isfinite(P_i)):
            continue
        z_i = _normalize_startup_comp_row(x[idx, :], n_components=n_components)
        out.append((T_i, P_i, z_i.tolist()))
    return out


def _prewarm_primary_thermo_backend_for_startup(
    *,
    col: ColumnSpec,
    provider: Any,
    emit_progress,
) -> Dict[str, Any]:
    warm_fn = getattr(provider, "warm_startup_kernels", None)
    if not callable(warm_fn):
        return {"executed": False, "reason": "provider-no-prewarm-hook"}

    flash_rows = _collect_startup_prewarm_rows(col, max_rows=3)
    density_state = flash_rows[0] if flash_rows else None
    if density_state is None and not flash_rows:
        return {"executed": False, "reason": "no-valid-startup-states"}

    category_fn = getattr(provider, "thermo_call_category", None)
    context = nullcontext()
    if callable(category_fn):
        try:
            context = category_fn("startup_prewarm")
        except Exception:
            context = nullcontext()
    t0 = time.perf_counter()
    with context:
        result = warm_fn(
            density_state=density_state,
            flash_rows=flash_rows,
        )
    wall = float(time.perf_counter() - t0)
    payload = dict(result or {})
    payload["executed"] = True
    payload["wall_sec"] = wall
    payload["flash_rows_requested"] = int(len(flash_rows))
    emit_progress(
        "[Init] Startup thermo prewarm complete  "
        f"wall={wall:.2f}s  density={'yes' if density_state is not None else 'no'}  "
        f"flash_rows={len(flash_rows)}"
    )
    return payload


def _clip_unit(x: Any, default: float = 0.0) -> float:
    try:
        v = float(x)
    except Exception:
        v = float(default)
    if not np.isfinite(v):
        v = float(default)
    return float(np.clip(v, 0.0, 1.0))


def _normalize_runtime_mode(mode: Any, default: str = "legacy") -> str:
    m = str(mode).strip().lower() if mode is not None else str(default).strip().lower()
    m = m.replace("_", "-")
    if m in ("legacy", "parity", "calibration", "hydraulic", "total-reflux"):
        return m
    return str(default).strip().lower()


def _normalize_integrator_mode(mode: Any, default: str = "explicit-euler") -> str:
    m = str(mode).strip().lower().replace("_", "-") if mode is not None else str(default).strip().lower()
    if m in ("explicit", "euler", "explicit-euler"):
        return "explicit-euler"
    if m in ("bdf",):
        return "bdf"
    if m in ("radau",):
        return "radau"
    if m in ("ida", "dae", "simultaneous-dae"):
        return "ida"
    return str(default).strip().lower()


def _normalize_equilibrium_relaxation_mode(mode: Any, default: str = "phase-holdup") -> str:
    m = str(mode).strip().lower().replace("_", "-") if mode is not None else str(default).strip().lower()
    if m in ("", "auto"):
        return str(default).strip().lower()
    if m in ("phase-holdup", "phase", "legacy"):
        return "phase-holdup"
    if m in ("composition-only", "composition", "comp-only", "y-only"):
        return "composition-only"
    return str(default).strip().lower()


def _effective_hydraulic_ida_profile(
    cfg: RunnerConfig,
    *,
    runtime_mode: str,
    integrator_mode: str,
) -> Dict[str, Any]:
    """
    Resolve effective IDA/DAE pilot settings.

    For hydraulic+IDA, apply tuned defaults when legacy defaults are still in
    place, while preserving explicit user overrides.
    """
    try:
        ida_max_iter = int(getattr(cfg, "ida_max_iter", 8))
    except Exception:
        ida_max_iter = 8
    ida_max_iter = max(1, ida_max_iter)

    try:
        dae_pilot_max_iter = int(getattr(cfg, "dae_pilot_max_iter", 3))
    except Exception:
        dae_pilot_max_iter = 3
    dae_pilot_max_iter = max(1, dae_pilot_max_iter)

    dae_pilot_enabled = bool(getattr(cfg, "enable_dae_pilot_algebraic_solve", False))
    dae_pilot_p_tol = getattr(cfg, "dae_pilot_p_tol_psia", None)
    dae_pilot_v_tol = getattr(cfg, "dae_pilot_v_tol_lbmolph", None)
    defaults_applied: List[str] = []

    is_hyd_ida = (
        str(runtime_mode).strip().lower() == "hydraulic"
        and str(integrator_mode).strip().lower() == "ida"
    )
    if is_hyd_ida:
        if not bool(dae_pilot_enabled):
            dae_pilot_enabled = True
            defaults_applied.append("enable_dae_pilot_algebraic_solve=True")
        if int(ida_max_iter) == 8:
            ida_max_iter = 12
            defaults_applied.append("ida_max_iter=12")
        v_try = np.nan
        try:
            if dae_pilot_v_tol is not None:
                v_try = float(dae_pilot_v_tol)
        except Exception:
            v_try = np.nan
        if (dae_pilot_v_tol is None) or ((not np.isfinite(v_try)) or abs(float(v_try) - 25.0) <= 1.0e-12):
            dae_pilot_v_tol = 100.0
            defaults_applied.append("dae_pilot_v_tol_lbmolph=100")

    return {
        "ida_max_iter": int(ida_max_iter),
        "dae_pilot_enabled": bool(dae_pilot_enabled),
        "dae_pilot_max_iter": int(dae_pilot_max_iter),
        "dae_pilot_p_tol_psia": dae_pilot_p_tol,
        "dae_pilot_v_tol_lbmolph": dae_pilot_v_tol,
        "defaults_applied": list(defaults_applied),
    }


def _resolve_startup_hydraulic_sequence_step(
    *,
    t_s: float,
    dt_sec: float,
    base_inputs: ColumnInputs,
    enable_sequence: bool,
    energy_on_sec: float,
    liquid_on_sec: float,
    liquid_ramp_sec: float,
    liquid_resid_gate_lbmolph: Optional[float],
    liquid_backoff_sec: Optional[float],
    liquid_alpha_state: float,
    last_mass_resid_max_lbmolph: Optional[float],
) -> Tuple[str, str, float, str]:
    """
    Runtime startup sequencing:
      1) pressure-only (hydraulic pressure + profile vapor/liquid traffic)
      2) pressure+energy vapor
      3) residual-gated liquid hydraulics ramp
    """
    p_base = str(base_inputs.pressure_model or "spec").strip().lower()
    if p_base not in ("spec", "hydraulic"):
        p_base = "spec"
    v_base = str(base_inputs.vapor_flow_model or "profile").strip().lower()
    if v_base not in ("profile", "energy", "conductance"):
        v_base = "profile"

    liq_enabled = bool(base_inputs.enable_liquid_hydraulic_override)
    liq_alpha_max = _clip_unit(getattr(base_inputs, "liquid_hydraulic_override_alpha", 1.0), default=1.0)
    if (not liq_enabled) or liq_alpha_max <= 0.0:
        liq_enabled = False
        liq_alpha_max = 0.0

    alpha = _clip_unit(liquid_alpha_state, default=liq_alpha_max)
    if (not enable_sequence) or p_base != "hydraulic":
        return p_base, v_base, liq_alpha_max, "base"

    t_now = max(float(t_s), 0.0)
    dt = max(float(dt_sec), 0.0)
    t_energy = max(float(energy_on_sec), 0.0)
    t_liq = max(float(liquid_on_sec), float(t_energy))
    t_ramp = max(float(liquid_ramp_sec), 1e-9)

    p_eff = "hydraulic"
    v_eff = v_base
    if v_base in ("energy", "conductance") and t_now < t_energy:
        v_eff = "profile"

    if (not liq_enabled) or liq_alpha_max <= 0.0:
        alpha = 0.0
    elif t_now < t_liq:
        alpha = 0.0
    else:
        a_time = float(np.clip((t_now - t_liq) / t_ramp, 0.0, 1.0)) * liq_alpha_max
        gate = None
        if liquid_resid_gate_lbmolph is not None:
            try:
                gate_try = float(liquid_resid_gate_lbmolph)
            except Exception:
                gate_try = np.nan
            if np.isfinite(gate_try) and gate_try > 0.0:
                gate = gate_try
        resid = np.nan
        if last_mass_resid_max_lbmolph is not None:
            try:
                resid = float(last_mass_resid_max_lbmolph)
            except Exception:
                resid = np.nan
        above_gate = bool(gate is not None and np.isfinite(resid) and resid > float(gate))
        if above_gate:
            tau_back = liquid_backoff_sec
            if tau_back is None:
                tau_back = 0.5 * t_ramp
            try:
                tau_back = float(tau_back)
            except Exception:
                tau_back = 0.5 * t_ramp
            tau_back = max(tau_back, 1e-9)
            alpha = max(0.0, alpha - (dt / tau_back) * liq_alpha_max)
        else:
            if alpha < a_time:
                alpha = min(a_time, alpha + (dt / t_ramp) * liq_alpha_max)
            else:
                alpha = min(alpha, liq_alpha_max)

    if t_now < t_energy:
        phase = "pressure_only"
    elif t_now < t_liq:
        phase = "pressure_energy"
    else:
        phase = "pressure_energy_liquid_ramp"
    return p_eff, v_eff, _clip_unit(alpha, default=0.0), phase


def _cosine_startup_fraction(t_since_start_s: float, ramp_sec: float) -> float:
    ramp = max(float(ramp_sec), 1.0e-9)
    frac = float(np.clip(float(t_since_start_s) / ramp, 0.0, 1.0))
    return float(0.5 * (1.0 - np.cos(np.pi * frac)))


def _startup_total_reflux_washout_active(t_s: float, cfg: "RunnerConfig") -> bool:
    raw = getattr(cfg, "startup_total_reflux_washout_sec", None)
    if raw is None:
        return False
    try:
        duration = float(raw)
    except Exception:
        return False
    if (not np.isfinite(duration)) or duration <= 0.0:
        return False
    return float(t_s) < float(duration)


def _resolve_startup_vapor_homotopy_beta(
    *,
    t_s: float,
    dt_sec: float,
    base_inputs: ColumnInputs,
    enable_homotopy: bool,
    liquid_on_sec: float,
    liquid_ramp_sec: float,
    vapor_on_sec: Optional[float],
    vapor_ramp_sec: float,
    vapor_rel_rate_gate_per_s: Optional[float],
    vapor_backoff_sec: Optional[float],
    beta_state: float,
    last_rel_state_rate_per_s: Optional[float],
) -> Tuple[float, str]:
    v_base = str(base_inputs.vapor_flow_model or "profile").strip().lower()
    if v_base not in ("energy", "conductance"):
        return 1.0, "disabled"
    if not bool(enable_homotopy):
        return 1.0, "disabled"

    t_now = max(float(t_s), 0.0)
    dt = max(float(dt_sec), 0.0)
    beta = _clip_unit(beta_state, default=0.0)
    if vapor_on_sec is None:
        t_vapor = max(float(liquid_on_sec), 0.0) + max(float(liquid_ramp_sec), 0.0)
    else:
        try:
            t_vapor = float(vapor_on_sec)
        except Exception:
            t_vapor = max(float(liquid_on_sec), 0.0) + max(float(liquid_ramp_sec), 0.0)
    t_vapor = max(t_vapor, 0.0)
    ramp = max(float(vapor_ramp_sec), 1.0e-9)

    if t_now < t_vapor:
        return 0.0, "profile_hold"

    gate = None
    if vapor_rel_rate_gate_per_s is not None:
        try:
            gate_try = float(vapor_rel_rate_gate_per_s)
        except Exception:
            gate_try = np.nan
        if np.isfinite(gate_try) and gate_try > 0.0:
            gate = gate_try
    rel_rate = np.nan
    if last_rel_state_rate_per_s is not None:
        try:
            rel_rate = float(last_rel_state_rate_per_s)
        except Exception:
            rel_rate = np.nan
    if gate is not None and np.isfinite(rel_rate) and rel_rate > float(gate):
        if vapor_backoff_sec is not None:
            try:
                backoff = max(float(vapor_backoff_sec), 1.0e-9)
            except Exception:
                backoff = ramp
            beta = max(0.0, beta - dt / backoff)
            return _clip_unit(beta, default=0.0), "backoff"
        return _clip_unit(beta, default=0.0), "guard_hold"

    target = _cosine_startup_fraction(t_now - t_vapor, ramp)
    if beta < target:
        beta = min(target, beta + dt / ramp)
        return _clip_unit(beta, default=0.0), "ramp"
    beta = min(beta, target)
    if beta >= 1.0 - 1.0e-12:
        return 1.0, "complete"
    return _clip_unit(beta, default=0.0), "hold"


def _resolve_residual_guarded_liquid_hydraulic_alpha(
    *,
    dt_sec: float,
    base_inputs: ColumnInputs,
    liquid_resid_gate_lbmolph: Optional[float],
    liquid_backoff_sec: Optional[float],
    liquid_recover_sec: float,
    liquid_alpha_state: float,
    last_mass_resid_max_lbmolph: Optional[float],
) -> Tuple[float, str]:
    """
    Continuously back off explicit liquid hydraulics when tray mass residuals
    become large, then recover toward the configured alpha once residuals calm.
    """
    liq_enabled = bool(base_inputs.enable_liquid_hydraulic_override)
    liq_alpha_max = _clip_unit(getattr(base_inputs, "liquid_hydraulic_override_alpha", 1.0), default=1.0)
    if (not liq_enabled) or liq_alpha_max <= 0.0:
        return 0.0, "disabled"

    alpha = _clip_unit(liquid_alpha_state, default=liq_alpha_max)
    dt = max(float(dt_sec), 0.0)
    if dt <= 0.0:
        return alpha, "hold"

    gate = None
    if liquid_resid_gate_lbmolph is not None:
        try:
            gate_try = float(liquid_resid_gate_lbmolph)
        except Exception:
            gate_try = np.nan
        if np.isfinite(gate_try) and gate_try > 0.0:
            gate = gate_try
    if gate is None:
        return liq_alpha_max, "no-gate"

    resid = np.nan
    if last_mass_resid_max_lbmolph is not None:
        try:
            resid = float(last_mass_resid_max_lbmolph)
        except Exception:
            resid = np.nan
    above_gate = bool(np.isfinite(resid) and resid > float(gate))
    if above_gate:
        # When tray mass closure deteriorates badly, cap the hydraulic blend
        # directly from the current residual ratio instead of relying on a slow
        # linear backoff that may be too weak to matter at small dt.
        alpha = liq_alpha_max * float(
            np.clip(float(gate) / max(float(resid), float(gate)), 0.0, 1.0)
        )
        return _clip_unit(alpha, default=0.0), "backoff"

    tau_recover = max(float(liquid_recover_sec), 1e-9)
    if alpha < liq_alpha_max:
        alpha = min(liq_alpha_max, alpha + (dt / tau_recover) * liq_alpha_max)
        return _clip_unit(alpha, default=0.0), "recover"
    return _clip_unit(alpha, default=0.0), "hold"


def _resolve_residual_guarded_liquid_hydraulic_alpha_per_stage(
    *,
    dt_sec: float,
    base_inputs: ColumnInputs,
    liquid_resid_gate_lbmolph: Optional[float],
    liquid_recover_sec: float,
    liquid_alpha_state: Optional[np.ndarray],
    last_mass_resid_lbmolph_per_stage: Optional[np.ndarray],
) -> Tuple[np.ndarray, str]:
    """
    Apply residual-based liquid-hydraulic backoff per tray instead of using a
    single whole-column alpha. This keeps ChemSep startup profiles as an
    initializer without letting one bad tray drag the full column back toward
    the seeded profile during runtime.
    """
    liq_enabled = bool(base_inputs.enable_liquid_hydraulic_override)
    liq_alpha_max = _clip_unit(getattr(base_inputs, "liquid_hydraulic_override_alpha", 1.0), default=1.0)
    try:
        n_stages = int(getattr(base_inputs.boundary, "n_stages", 0))
    except Exception:
        n_stages = 0
    if liquid_alpha_state is not None:
        try:
            state = np.asarray(liquid_alpha_state, dtype=float).reshape((-1,)).copy()
        except Exception:
            state = np.full((max(n_stages, 1),), float(liq_alpha_max), dtype=float)
    else:
        state = np.full((max(n_stages, 1),), float(liq_alpha_max), dtype=float)
    if n_stages <= 0:
        n_stages = int(state.size)
    if state.size != n_stages:
        state = np.full((n_stages,), float(liq_alpha_max), dtype=float)
    if (not liq_enabled) or liq_alpha_max <= 0.0:
        return np.zeros((n_stages,), dtype=float), "disabled"

    state = np.clip(np.where(np.isfinite(state), state, liq_alpha_max), 0.0, 1.0)
    dt = max(float(dt_sec), 0.0)
    if dt <= 0.0:
        return state, "hold"

    gate = None
    if liquid_resid_gate_lbmolph is not None:
        try:
            gate_try = float(liquid_resid_gate_lbmolph)
        except Exception:
            gate_try = np.nan
        if np.isfinite(gate_try) and gate_try > 0.0:
            gate = gate_try
    if gate is None:
        return np.full((n_stages,), float(liq_alpha_max), dtype=float), "no-gate"

    resid = np.full((n_stages,), np.nan, dtype=float)
    if last_mass_resid_lbmolph_per_stage is not None:
        try:
            resid = np.asarray(last_mass_resid_lbmolph_per_stage, dtype=float).reshape((n_stages,))
        except Exception:
            resid = np.full((n_stages,), np.nan, dtype=float)

    alpha = state.copy()
    phase = "hold"
    tau_recover = max(float(liquid_recover_sec), 1e-9)
    recover_step = (dt / tau_recover) * liq_alpha_max
    for i in range(n_stages):
        r_i = resid[i]
        if np.isfinite(r_i) and r_i > float(gate):
            alpha[i] = liq_alpha_max * float(np.clip(float(gate) / max(float(r_i), float(gate)), 0.0, 1.0))
            phase = "backoff"
        elif alpha[i] < liq_alpha_max:
            alpha[i] = min(liq_alpha_max, alpha[i] + recover_step)
            if phase != "backoff":
                phase = "recover"
    return np.clip(alpha, 0.0, 1.0), phase


def build_inputs_for_runner(case: CaseData, col: ColumnSpec, cfg: RunnerConfig) -> Tuple[ColumnInputs, Any]:
    specs = getattr(col, "specs_raw", None) or {}
    startup_build_timing_sec: Dict[str, float] = {}
    startup_build_info: Dict[str, Any] = {}
    backend_build_t0 = time.perf_counter()
    backend_build = build_primary_thermo_backend(
        cfg=cfg,
        col=col,
        emit_progress=_emit_progress,
    )
    startup_build_timing_sec["primary_backend_build"] = float(time.perf_counter() - backend_build_t0)
    _emit_progress(
        "[Init] Primary thermo backend build complete  "
        f"wall={float(startup_build_timing_sec['primary_backend_build']):.2f}s"
    )
    thermo_mode = str(backend_build.thermo_mode)
    dwsim_pkg = backend_build.dwsim_property_package
    prov = backend_build.provider
    if bool(getattr(cfg, "enable_primary_thermo_startup_prewarm", False)):
        prewarm_t0 = time.perf_counter()
        prewarm_info = _prewarm_primary_thermo_backend_for_startup(
            col=col,
            provider=prov,
            emit_progress=_emit_progress,
        )
        startup_build_timing_sec["primary_backend_prewarm"] = float(time.perf_counter() - prewarm_t0)
        if bool(prewarm_info.get("executed", False)):
            startup_build_info["primary_backend_prewarm"] = dict(prewarm_info)
            get_counters_fn = getattr(prov, "get_call_counters", None)
            if callable(get_counters_fn):
                try:
                    prewarm_counters = dict(get_counters_fn().get("startup_prewarm", {}) or {})
                except Exception:
                    prewarm_counters = {}
                if prewarm_counters:
                    startup_build_info["primary_backend_prewarm_call_counters"] = prewarm_counters
            _emit_progress(
                "[Init] Primary thermo backend prewarm recorded  "
                f"wall={float(startup_build_timing_sec['primary_backend_prewarm']):.2f}s"
            )

    eq_relax_live_pr = bool(getattr(cfg, "equilibrium_relaxation_live_pr", False))
    if not eq_relax_live_pr:
        eq_relax_live_pr = bool(
            _spec_bool(
                specs,
                "Equilibrium Relaxation Live PR",
                "Equilibrium Live PR",
                "Selective PR for Equilibrium Relaxation",
            )
        )

    eq_relax_build_t0 = time.perf_counter()
    eq_relax_prov = build_equilibrium_relaxation_pr_provider(
        enabled=eq_relax_live_pr,
        col=col,
        primary_thermo_mode=thermo_mode,
        primary_dwsim_property_package=dwsim_pkg,
        emit_progress=_emit_progress,
    )
    startup_build_timing_sec["equilibrium_relaxation_backend_build"] = float(
        time.perf_counter() - eq_relax_build_t0
    )

    _emit_progress("[Init] Starting Francis hydraulic autocalibration from seeded liquid profile")
    autocal_t0 = time.perf_counter()
    calibrated_hyd_c = _autocalibrate_francis_hydraulic_c_factors_from_seed(
        col=col,
        thermo_provider=prov,
    )
    startup_build_timing_sec["francis_autocalibration"] = float(time.perf_counter() - autocal_t0)
    if calibrated_hyd_c:
        _emit_progress(
            "[Init] Calibrated Francis hydraulic C factors from seeded ChemSep liquid profile  "
            f"wall={float(startup_build_timing_sec['francis_autocalibration']):.2f}s"
        )
    else:
        _emit_progress(
            "[Init] Francis hydraulic autocalibration complete  no seeded adjustments applied  "
            f"wall={float(startup_build_timing_sec['francis_autocalibration']):.2f}s"
        )

    boundary = _infer_boundary_flows(case, col, cfg)
    vol = _build_volume_model(col, default_vapor_volume_ft3=1.0)

    pressure_model = str(_spec_get(specs, "Pressure Model") or "").strip().lower()
    if not pressure_model:
        pressure_model = "hydraulic" if col.geometry is not None else "spec"
    if pressure_model not in ("spec", "hydraulic"):
        pressure_model = "spec"

    vapor_flow_model = str(_spec_get(specs, "Vapor Flow Model") or "").strip().lower()
    if not vapor_flow_model:
        vapor_flow_model = "energy" if pressure_model == "hydraulic" else "profile"
    if vapor_flow_model not in ("profile", "energy", "conductance"):
        vapor_flow_model = "profile"

    runtime_mode = _normalize_runtime_mode(getattr(cfg, "runtime_mode", None), default="legacy")
    if runtime_mode in ("parity", "calibration"):
        pressure_model = "spec"
        vapor_flow_model = "profile"
    elif runtime_mode in ("hydraulic", "total-reflux"):
        pressure_model = "hydraulic"
        vapor_flow_model = "energy"

    eq_mode_default = "composition-only" if runtime_mode in ("hydraulic", "total-reflux") else "phase-holdup"
    eq_mode_spec = _spec_get(
        specs,
        "Equilibrium Relaxation Mode",
        "Equilibrium Mode",
        "Equilibrium Transfer Mode",
    )
    eq_mode = _normalize_equilibrium_relaxation_mode(
        cfg.equilibrium_relaxation_mode if cfg.equilibrium_relaxation_mode is not None else eq_mode_spec,
        default=eq_mode_default,
    )
    # If CLI left mode at default "auto", fall back to spec/default behavior.
    if str(getattr(cfg, "equilibrium_relaxation_mode", "auto")).strip().lower() in ("", "auto"):
        eq_mode = _normalize_equilibrium_relaxation_mode(eq_mode_spec, default=eq_mode_default)
    eq_phase_guard_lbmol = cfg.equilibrium_phase_holdup_guard_lbmol
    if eq_phase_guard_lbmol is None:
        eq_phase_guard_lbmol = 0.0
        if eq_mode == "phase-holdup" and runtime_mode in ("hydraulic", "total-reflux"):
            eq_phase_guard_lbmol = 1.0
    try:
        eq_phase_guard_lbmol = float(eq_phase_guard_lbmol)
    except Exception:
        eq_phase_guard_lbmol = 0.0
    if (not np.isfinite(eq_phase_guard_lbmol)) or eq_phase_guard_lbmol < 0.0:
        eq_phase_guard_lbmol = 0.0
    eq_tau_spec = _spec_float(
        specs,
        "Equilibrium Tau (sec)",
        "Equilibrium Relaxation Tau (sec)",
        "Equilibrium Relaxation Time Constant (sec)",
    )
    eq_energy_damping_gain = cfg.equilibrium_energy_damping_gain
    if eq_energy_damping_gain is None:
        eq_energy_damping_gain = _spec_float(
            specs,
            "Equilibrium Energy Damping Gain",
            "Eq Energy Damping Gain",
        )
    if eq_energy_damping_gain is None:
        eq_energy_damping_gain = 0.0
    try:
        eq_energy_damping_gain = float(eq_energy_damping_gain)
    except Exception:
        eq_energy_damping_gain = 0.0
    if (not np.isfinite(eq_energy_damping_gain)) or eq_energy_damping_gain < 0.0:
        eq_energy_damping_gain = 0.0
    hydraulic_energy_temp_damping = cfg.hydraulic_energy_temperature_damping
    if hydraulic_energy_temp_damping is None:
        hydraulic_energy_temp_damping = 1.0
    try:
        hydraulic_energy_temp_damping = float(hydraulic_energy_temp_damping)
    except Exception:
        hydraulic_energy_temp_damping = 1.0
    if (not np.isfinite(hydraulic_energy_temp_damping)) or hydraulic_energy_temp_damping < 0.0:
        hydraulic_energy_temp_damping = 1.0
    hydraulic_energy_temp_mode = str(
        cfg.hydraulic_energy_temperature_mode
        if cfg.hydraulic_energy_temperature_mode is not None
        else "legacy"
    ).strip().lower()
    if hydraulic_energy_temp_mode not in (
        "legacy",
        "bubble-point-follower",
        "pressure-correction-follower",
        "enthalpy-state-follower",
    ):
        hydraulic_energy_temp_mode = "legacy"
    hydraulic_energy_temp_follow_tau_sec = cfg.hydraulic_energy_temperature_follow_tau_sec
    if hydraulic_energy_temp_follow_tau_sec is None:
        hydraulic_energy_temp_follow_tau_sec = _spec_float(
            specs,
            "Hydraulic Energy Temperature Follow Tau (sec)",
            "Hydraulic Energy Temperature Follow Time Constant (sec)",
            "Condenser Transfer Temperature Tau (sec)",
        )
    if hydraulic_energy_temp_follow_tau_sec is None:
        hydraulic_energy_temp_follow_tau_sec = 0.5
    try:
        hydraulic_energy_temp_follow_tau_sec = float(hydraulic_energy_temp_follow_tau_sec)
    except Exception:
        hydraulic_energy_temp_follow_tau_sec = 0.5
    if (not np.isfinite(hydraulic_energy_temp_follow_tau_sec)) or hydraulic_energy_temp_follow_tau_sec <= 0.0:
        hydraulic_energy_temp_follow_tau_sec = 0.5
    hydraulic_energy_temp_resid_frac = cfg.hydraulic_energy_temperature_resid_frac
    if hydraulic_energy_temp_resid_frac is None:
        hydraulic_energy_temp_resid_frac = 0.01
    try:
        hydraulic_energy_temp_resid_frac = float(hydraulic_energy_temp_resid_frac)
    except Exception:
        hydraulic_energy_temp_resid_frac = 0.01
    if (not np.isfinite(hydraulic_energy_temp_resid_frac)) or hydraulic_energy_temp_resid_frac < 0.0:
        hydraulic_energy_temp_resid_frac = 0.01
    hydraulic_energy_temp_pressure_slope = cfg.hydraulic_energy_temperature_pressure_slope_F_per_psi
    if hydraulic_energy_temp_pressure_slope is None:
        hydraulic_energy_temp_pressure_slope = 2.0
    try:
        hydraulic_energy_temp_pressure_slope = float(hydraulic_energy_temp_pressure_slope)
    except Exception:
        hydraulic_energy_temp_pressure_slope = 2.0
    if not np.isfinite(hydraulic_energy_temp_pressure_slope):
        hydraulic_energy_temp_pressure_slope = 2.0
    hydraulic_energy_temp_target_refresh_steps = cfg.hydraulic_energy_temperature_target_refresh_steps
    if hydraulic_energy_temp_target_refresh_steps is None:
        hydraulic_energy_temp_target_refresh_steps = 20
    try:
        hydraulic_energy_temp_target_refresh_steps = int(hydraulic_energy_temp_target_refresh_steps)
    except Exception:
        hydraulic_energy_temp_target_refresh_steps = 20
    if hydraulic_energy_temp_target_refresh_steps <= 0:
        hydraulic_energy_temp_target_refresh_steps = 20

    coupled_hydraulic_energy = (pressure_model == "hydraulic") and (vapor_flow_model == "energy")

    dry_tray_k = _spec_float(specs, "Dry Tray K")
    if dry_tray_k is None or not np.isfinite(dry_tray_k):
        dry_tray_k = 1.0

    tau_v = _spec_float(specs, "Vapor Holdup Relaxation (sec)")
    if tau_v is None:
        tau_v = _spec_float(specs, "Stage time constant [tau] (sec)")
    tau_v_cli_explicit = cfg.vapor_holdup_relaxation_sec is not None
    if cfg.vapor_holdup_relaxation_sec is not None:
        tau_v = float(cfg.vapor_holdup_relaxation_sec)
    if tau_v is not None and (not np.isfinite(tau_v) or tau_v <= 0.0):
        tau_v = None
    if tau_v is None and coupled_hydraulic_energy:
        # Stabilizing default for explicit integration under hydraulic/energy coupling.
        tau_v = 10.0

    tau_p_hyd = _spec_float(
        specs,
        "Hydraulic Pressure Relaxation (sec)",
        "Pressure Relaxation (sec)",
        "Hydraulic Pressure Tau (sec)",
    )
    if cfg.hydraulic_pressure_relaxation_sec is not None:
        tau_p_hyd = float(cfg.hydraulic_pressure_relaxation_sec)
    if tau_p_hyd is not None and (not np.isfinite(tau_p_hyd) or tau_p_hyd <= 0.0):
        tau_p_hyd = None
    if tau_p_hyd is None:
        # Backward-compatible default behavior: share vapor-holdup timescale.
        tau_p_hyd = tau_v

    tau_top_pT = _spec_float(
        specs,
        "Top Drum Pressure Temperature Relaxation (sec)",
        "Top Drum Pressure Temperature Tau (sec)",
    )
    if cfg.top_drum_pressure_temperature_relaxation_sec is not None:
        tau_top_pT = float(cfg.top_drum_pressure_temperature_relaxation_sec)
    if tau_top_pT is not None and (not np.isfinite(tau_top_pT)):
        tau_top_pT = None

    tau_vflow = _spec_float(specs, "Vapor Flow Relaxation (sec)")
    if cfg.vapor_flow_relaxation_sec is not None:
        tau_vflow = float(cfg.vapor_flow_relaxation_sec)
    if tau_vflow is not None and (not np.isfinite(tau_vflow) or tau_vflow <= 0.0):
        tau_vflow = None

    conductance_vflow_nominal_hi_ratio = _spec_float(
        specs,
        "Conductance Vapor Nominal Hi Ratio",
        "Conductance Vflow Nominal Hi Ratio",
        "Conductance Vapor Profile Hi Ratio",
    )
    if cfg.conductance_vflow_nominal_hi_ratio is not None:
        conductance_vflow_nominal_hi_ratio = float(cfg.conductance_vflow_nominal_hi_ratio)
    if (
        conductance_vflow_nominal_hi_ratio is not None
        and (not np.isfinite(conductance_vflow_nominal_hi_ratio) or conductance_vflow_nominal_hi_ratio <= 0.0)
    ):
        conductance_vflow_nominal_hi_ratio = None

    reb_nbr_hi = _spec_float(
        specs,
        "Reboiler Neighbor Vapor Hi Ratio",
        "Reboiler Neighbor Vflow Hi Ratio",
    )
    reb_nbr_lo = _spec_float(
        specs,
        "Reboiler Neighbor Vapor Lo Ratio",
        "Reboiler Neighbor Vflow Lo Ratio",
    )
    if cfg.reboiler_neighbor_vflow_hi_ratio is not None:
        reb_nbr_hi = float(cfg.reboiler_neighbor_vflow_hi_ratio)
    if cfg.reboiler_neighbor_vflow_lo_ratio is not None:
        reb_nbr_lo = float(cfg.reboiler_neighbor_vflow_lo_ratio)
    if reb_nbr_hi is not None and (not np.isfinite(reb_nbr_hi) or reb_nbr_hi <= 0.0):
        reb_nbr_hi = None
    if reb_nbr_lo is not None and (not np.isfinite(reb_nbr_lo) or reb_nbr_lo <= 0.0):
        reb_nbr_lo = None

    thermo_refresh_dT = _spec_float(
        specs,
        "Thermo Refresh dT (F)",
        "Thermo Refresh Delta T (F)",
        "Thermo Refresh Delta (F)",
        "Thermo Refresh \u0394T (F)",
    )
    thermo_refresh_dP = _spec_float(
        specs,
        "Thermo Refresh dP (psia)",
        "Thermo Refresh Delta P (psia)",
    )
    thermo_refresh_dX = _spec_float(
        specs,
        "Thermo Refresh dX",
        "Thermo Refresh Delta X",
    )
    feed_flash_at_stage = _as_bool(
        _spec_get(
            specs,
            "Flash Feed At Stage Conditions",
            "Feed Flash At Stage Conditions",
            "Flash Feed At Tray Conditions",
            "Feed Flash At Tray Conditions",
        )
    )
    feed_flash_explicit = feed_flash_at_stage is not None
    if cfg.flash_feed_at_stage_conditions is not None:
        feed_flash_at_stage = bool(cfg.flash_feed_at_stage_conditions)
        feed_flash_explicit = True
    if feed_flash_at_stage is None:
        feed_flash_at_stage = True

    if cfg.thermo_refresh_dT_F is not None:
        thermo_refresh_dT = float(cfg.thermo_refresh_dT_F)
    if cfg.thermo_refresh_dP_psia is not None:
        thermo_refresh_dP = float(cfg.thermo_refresh_dP_psia)
    if cfg.thermo_refresh_dx is not None:
        thermo_refresh_dX = float(cfg.thermo_refresh_dx)

    if thermo_refresh_dT is not None and (not np.isfinite(thermo_refresh_dT) or thermo_refresh_dT <= 0.0):
        thermo_refresh_dT = None
    if thermo_refresh_dP is not None and (not np.isfinite(thermo_refresh_dP) or thermo_refresh_dP <= 0.0):
        thermo_refresh_dP = None
    if thermo_refresh_dX is not None and (not np.isfinite(thermo_refresh_dX) or thermo_refresh_dX <= 0.0):
        thermo_refresh_dX = None
    if (
        int(getattr(cfg, "thermo_every_n_steps", 1) or 1) > 1
        and bool(getattr(cfg, "enable_thermo_cadence_guardrails", True))
        and str(runtime_mode).strip().lower() == "hydraulic"
        and cfg.thermo_refresh_dT_F is None
        and cfg.thermo_refresh_dP_psia is None
        and cfg.thermo_refresh_dx is None
    ):
        # Give frozen-thermo runs a conservative safety-refresh floor without
        # auto-enabling hydraulic pressure drift checks. Pressure profiles can
        # legitimately move every step in dynamic runs, which defeats cadence
        # if we make dP a default guardrail instead of an explicit opt-in.
        thermo_refresh_dT = 1.0 if bool(cfg.include_temperature) else None
        thermo_refresh_dP = None
        thermo_refresh_dX = 5.0e-3
    if (
        coupled_hydraulic_energy
        and cfg.thermo_refresh_dT_F is None
        and cfg.thermo_refresh_dP_psia is None
        and cfg.thermo_refresh_dx is None
        and (not (int(getattr(cfg, "thermo_every_n_steps", 1) or 1) > 1 and bool(getattr(cfg, "enable_thermo_cadence_guardrails", True))))
    ):
        # Avoid thermo/controller dead-band aliasing during steady-state approach.
        thermo_refresh_dT = None
        thermo_refresh_dP = None
        thermo_refresh_dX = None

    condenser_dp_psi = cfg.condenser_pressure_drop_psi
    if condenser_dp_psi is None:
        condenser_dp_psi = _spec_float(
            specs,
            "Condenser Pressure Drop (psi)",
            "Condenser Pressure Drop (psia)",
            "Condenser dP (psi)",
            "Condenser dP (psia)",
            "Condenser Delta P (psi)",
            "Condenser Delta P (psia)",
        )
    if condenser_dp_psi is not None and (not np.isfinite(condenser_dp_psi) or condenser_dp_psi < 0.0):
        condenser_dp_psi = None

    enable_top_psv = bool(cfg.enable_top_psv)
    if not enable_top_psv:
        b_psv = _as_bool(
            _spec_get(
                specs,
                "Enable Top PSV",
                "Top PSV Enabled",
                "Top Drum PSV Enabled",
                "Distillate Drum PSV Enabled",
                "Enable PSV",
            )
        )
        if b_psv is not None:
            enable_top_psv = bool(b_psv)
    top_psv_setpoint_psia = cfg.top_psv_setpoint_psia
    if top_psv_setpoint_psia is None:
        top_psv_setpoint_psia = _spec_float(
            specs,
            "Top PSV SP (psia)",
            "Top PSV Setpoint (psia)",
            "Top Drum PSV SP (psia)",
            "Top Drum PSV Setpoint (psia)",
            "Distillate Drum PSV SP (psia)",
            "Distillate Drum PSV Setpoint (psia)",
        )
    if top_psv_setpoint_psia is not None and (
        (not np.isfinite(top_psv_setpoint_psia)) or top_psv_setpoint_psia <= 0.0
    ):
        top_psv_setpoint_psia = None

    top_psv_gain_lbmolps_per_psi = cfg.top_psv_gain_lbmolps_per_psi
    if top_psv_gain_lbmolps_per_psi is None:
        top_psv_gain_lbmolps_per_psi = _spec_float(
            specs,
            "Top PSV Gain (lbmol/s/psi)",
            "Top PSV Gain (lbmolps/psi)",
            "Top Drum PSV Gain (lbmol/s/psi)",
            "Distillate Drum PSV Gain (lbmol/s/psi)",
        )
    if top_psv_gain_lbmolps_per_psi is not None and (
        (not np.isfinite(top_psv_gain_lbmolps_per_psi)) or top_psv_gain_lbmolps_per_psi <= 0.0
    ):
        top_psv_gain_lbmolps_per_psi = None

    top_psv_max_vent_lbmolps = cfg.top_psv_max_vent_lbmolps
    if top_psv_max_vent_lbmolps is None:
        top_psv_max_vent_lbmolps = _spec_float(
            specs,
            "Top PSV Max Vent (lbmol/s)",
            "Top PSV Max (lbmol/s)",
            "Top Drum PSV Max Vent (lbmol/s)",
            "Distillate Drum PSV Max Vent (lbmol/s)",
        )
    if top_psv_max_vent_lbmolps is not None and (
        (not np.isfinite(top_psv_max_vent_lbmolps)) or top_psv_max_vent_lbmolps < 0.0
    ):
        top_psv_max_vent_lbmolps = None

    overhead_line_vapor_volume_ft3 = _spec_float(
        specs,
        "Overhead Vapor Line Volume (ft3)",
        "Overhead Vapour Line Volume (ft3)",
        "Overhead Line Vapor Volume (ft3)",
        "Overhead Line Volume (ft3)",
    )
    if (
        overhead_line_vapor_volume_ft3 is not None
        and ((not np.isfinite(overhead_line_vapor_volume_ft3)) or overhead_line_vapor_volume_ft3 < 0.0)
    ):
        overhead_line_vapor_volume_ft3 = None

    condenser_vapor_volume_ft3 = _spec_float(
        specs,
        "Condenser Vapor Volume (ft3)",
        "Condenser Vapour Volume (ft3)",
        "Condenser Vapor Space (ft3)",
        "Condenser Vapour Space (ft3)",
    )
    if (
        condenser_vapor_volume_ft3 is not None
        and ((not np.isfinite(condenser_vapor_volume_ft3)) or condenser_vapor_volume_ft3 < 0.0)
    ):
        condenser_vapor_volume_ft3 = None

    overhead_vapor_adders_ft3 = 0.0
    if overhead_line_vapor_volume_ft3 is not None:
        overhead_vapor_adders_ft3 += float(overhead_line_vapor_volume_ft3)
    if condenser_vapor_volume_ft3 is not None:
        overhead_vapor_adders_ft3 += float(condenser_vapor_volume_ft3)
    condenser_type_txt = str(
        _spec_get(specs, "Condenser Type")
        or getattr(getattr(col, "duties", None), "condenser_type", "")
        or ""
    ).strip().lower()
    is_total_condenser = condenser_type_txt.startswith("total")

    top_drum_total_volume_ft3 = cfg.top_drum_total_volume_ft3
    if top_drum_total_volume_ft3 is None:
        top_drum_total_volume_ft3 = _spec_float(
            specs,
            "Top Drum Total Volume (ft3)",
            "Top Accumulator Total Volume (ft3)",
            "Reflux Drum Total Volume (ft3)",
            "Distillate Drum Total Volume (ft3)",
            "Top Drum Volume (ft3)",
            "Reflux Drum Volume (ft3)",
            "Distillate Drum Volume (ft3)",
        )
    if top_drum_total_volume_ft3 is None:
        d_ft = _spec_float(
            specs,
            "Top Drum Diameter (ft)",
            "Top Accumulator Diameter (ft)",
            "Reflux Drum Diameter (ft)",
            "Distillate Drum Diameter (ft)",
            "Top Drum ID (ft)",
            "Reflux Drum ID (ft)",
            "Distillate Drum ID (ft)",
        )
        l_ft = _spec_float(
            specs,
            "Top Drum Length (ft)",
            "Top Accumulator Length (ft)",
            "Reflux Drum Length (ft)",
            "Distillate Drum Length (ft)",
        )
        if d_ft is not None and l_ft is not None and d_ft > 0.0 and l_ft > 0.0:
            top_drum_total_volume_ft3 = float(np.pi * 0.25 * float(d_ft) * float(d_ft) * float(l_ft))

    top_drum_vapor_volume_ft3 = cfg.top_drum_vapor_volume_ft3
    if top_drum_vapor_volume_ft3 is None:
        top_drum_vapor_volume_ft3 = _spec_float(
            specs,
            "Top Drum Vapor Volume (ft3)",
            "Top Accumulator Vapor Volume (ft3)",
            "Reflux Drum Vapor Volume (ft3)",
            "Distillate Drum Vapor Volume (ft3)",
            "Top Vapor Volume (ft3)",
        )
    if top_drum_vapor_volume_ft3 is None:
        if top_drum_total_volume_ft3 is not None and np.isfinite(top_drum_total_volume_ft3) and top_drum_total_volume_ft3 > 0.0:
            top_holdup_lbmol = _spec_float(
                specs,
                "Top Accumulator Holdup (lbmol)",
                "Top Drum Holdup (lbmol)",
                "Reflux Drum Holdup (lbmol)",
            )
            top_liq_frac = _spec_float(
                specs,
                "Top Drum Liquid Fraction (-)",
                "Top Drum Liquid Volume Fraction",
                "Top Drum Liquid Fraction",
                "Top Accumulator Liquid Volume Fraction",
                "Top Accumulator Liquid Fraction",
                "Reflux Drum Liquid Volume Fraction",
                "Reflux Drum Liquid Fraction",
                "Distillate Drum Liquid Volume Fraction",
                "Distillate Drum Liquid Fraction",
                "Top Drum Fill Fraction",
                "Reflux Drum Fill Fraction",
                "Distillate Drum Fill Fraction",
            )
            if top_liq_frac is not None and top_liq_frac > 1.0 and top_liq_frac <= 100.0:
                top_liq_frac = float(top_liq_frac) / 100.0
            if top_liq_frac is not None and (top_liq_frac < 0.0 or top_liq_frac > 1.0):
                top_liq_frac = None

            top_liq_vol_ft3 = None
            if (
                top_holdup_lbmol is not None
                and top_holdup_lbmol > 0.0
                and hasattr(prov, "liquid_density_lbmol_ft3")
                and hasattr(col, "x0")
                and hasattr(col, "T_f")
                and hasattr(col, "P_psia")
            ):
                try:
                    x_top = np.asarray(getattr(col, "x0"), dtype=float).reshape((col.n_stages, col.n_components))[0, :]
                    T_top = float(np.asarray(getattr(col, "T_f"), dtype=float).reshape((col.n_stages,))[0])
                    P_top = float(np.asarray(getattr(col, "P_psia"), dtype=float).reshape((col.n_stages,))[0])
                    rho_top = float(prov.liquid_density_lbmol_ft3(T_top, P_top, x_top))
                    if np.isfinite(rho_top) and rho_top > 1e-12:
                        top_liq_vol_ft3 = float(top_holdup_lbmol) / float(rho_top)
                except Exception:
                    top_liq_vol_ft3 = None
            elif top_liq_frac is not None:
                top_liq_vol_ft3 = float(top_liq_frac) * float(top_drum_total_volume_ft3)

            if top_liq_vol_ft3 is None or (not np.isfinite(top_liq_vol_ft3)):
                # If only geometry is provided, default to half-full.
                top_liq_vol_ft3 = 0.5 * float(top_drum_total_volume_ft3)
            top_liq_vol_ft3 = float(np.clip(top_liq_vol_ft3, 0.0, float(top_drum_total_volume_ft3)))
            top_drum_vapor_volume_ft3 = float(top_drum_total_volume_ft3) - float(top_liq_vol_ft3)
    if (
        top_drum_total_volume_ft3 is None
        and top_drum_vapor_volume_ft3 is not None
        and np.isfinite(top_drum_vapor_volume_ft3)
        and top_drum_vapor_volume_ft3 > 0.0
    ):
        top_liq_frac = _spec_float(
            specs,
            "Top Drum Liquid Fraction (-)",
            "Top Drum Liquid Volume Fraction",
            "Top Drum Liquid Fraction",
            "Top Accumulator Liquid Volume Fraction",
            "Top Accumulator Liquid Fraction",
            "Reflux Drum Liquid Volume Fraction",
            "Reflux Drum Liquid Fraction",
            "Distillate Drum Liquid Volume Fraction",
            "Distillate Drum Liquid Fraction",
            "Top Drum Fill Fraction",
            "Reflux Drum Fill Fraction",
            "Distillate Drum Fill Fraction",
        )
        if top_liq_frac is not None and top_liq_frac > 1.0 and top_liq_frac <= 100.0:
            top_liq_frac = float(top_liq_frac) / 100.0
        if top_liq_frac is not None and 0.0 <= float(top_liq_frac) < 1.0:
            try:
                top_drum_total_volume_ft3 = float(top_drum_vapor_volume_ft3) / max(1.0 - float(top_liq_frac), 1e-12)
            except Exception:
                top_drum_total_volume_ft3 = None
        elif (
            hasattr(prov, "liquid_density_lbmol_ft3")
            and hasattr(col, "x0")
            and hasattr(col, "T_f")
            and hasattr(col, "P_psia")
        ):
            # Infer total drum volume from initial liquid holdup + explicit initial vapor volume.
            top_holdup_lbmol = _spec_float(
                specs,
                "Top Accumulator Holdup (lbmol)",
                "Top Drum Holdup (lbmol)",
                "Reflux Drum Holdup (lbmol)",
            )
            if top_holdup_lbmol is not None and top_holdup_lbmol >= 0.0:
                try:
                    x_top = np.asarray(getattr(col, "x0"), dtype=float).reshape((col.n_stages, col.n_components))[0, :]
                    T_top = float(np.asarray(getattr(col, "T_f"), dtype=float).reshape((col.n_stages,))[0])
                    P_top = float(np.asarray(getattr(col, "P_psia"), dtype=float).reshape((col.n_stages,))[0])
                    rho_top = float(prov.liquid_density_lbmol_ft3(T_top, P_top, x_top))
                    if np.isfinite(rho_top) and rho_top > 1e-12:
                        top_liq_vol_ft3 = float(top_holdup_lbmol) / float(rho_top)
                        total_try = float(top_drum_vapor_volume_ft3) + max(float(top_liq_vol_ft3), 0.0)
                        if np.isfinite(total_try) and total_try > float(top_drum_vapor_volume_ft3):
                            top_drum_total_volume_ft3 = total_try
                except Exception:
                    top_drum_total_volume_ft3 = None

    if top_drum_vapor_volume_ft3 is None:
        try:
            vv = _vapor_volume_ft3_per_stage(vol, int(col.n_stages))
            top_drum_vapor_volume_ft3 = float(vv[0])
        except Exception:
            top_drum_vapor_volume_ft3 = None
    if (
        top_drum_total_volume_ft3 is None
        and top_drum_vapor_volume_ft3 is not None
        and np.isfinite(top_drum_vapor_volume_ft3)
        and top_drum_vapor_volume_ft3 > 0.0
        and hasattr(prov, "liquid_density_lbmol_ft3")
        and hasattr(col, "x0")
        and hasattr(col, "T_f")
        and hasattr(col, "P_psia")
    ):
        top_holdup_lbmol = _spec_float(
            specs,
            "Top Accumulator Holdup (lbmol)",
            "Top Drum Holdup (lbmol)",
            "Reflux Drum Holdup (lbmol)",
        )
        top_liq_frac = _spec_float(
            specs,
            "Top Drum Liquid Fraction (-)",
            "Top Drum Liquid Volume Fraction",
            "Top Drum Liquid Fraction",
            "Top Accumulator Liquid Volume Fraction",
            "Top Accumulator Liquid Fraction",
            "Reflux Drum Liquid Volume Fraction",
            "Reflux Drum Liquid Fraction",
            "Distillate Drum Liquid Volume Fraction",
            "Distillate Drum Liquid Fraction",
            "Top Drum Fill Fraction",
            "Reflux Drum Fill Fraction",
            "Distillate Drum Fill Fraction",
        )
        if top_liq_frac is not None and top_liq_frac > 1.0 and top_liq_frac <= 100.0:
            top_liq_frac = float(top_liq_frac) / 100.0
        if top_holdup_lbmol is not None and top_holdup_lbmol >= 0.0:
            try:
                x_top = np.asarray(getattr(col, "x0"), dtype=float).reshape((col.n_stages, col.n_components))[0, :]
                T_top = float(np.asarray(getattr(col, "T_f"), dtype=float).reshape((col.n_stages,))[0])
                P_top = float(np.asarray(getattr(col, "P_psia"), dtype=float).reshape((col.n_stages,))[0])
                rho_top = float(prov.liquid_density_lbmol_ft3(T_top, P_top, x_top))
                if np.isfinite(rho_top) and rho_top > 1e-12:
                    top_liq_vol_ft3 = float(top_holdup_lbmol) / float(rho_top)
                    total_try = float(top_drum_vapor_volume_ft3) + max(float(top_liq_vol_ft3), 0.0)
                    if np.isfinite(total_try) and total_try > float(top_drum_vapor_volume_ft3):
                        top_drum_total_volume_ft3 = total_try
            except Exception:
                top_drum_total_volume_ft3 = None
        elif top_liq_frac is not None and 0.0 <= float(top_liq_frac) < 1.0:
            try:
                top_drum_total_volume_ft3 = float(top_drum_vapor_volume_ft3) / max(1.0 - float(top_liq_frac), 1e-12)
            except Exception:
                top_drum_total_volume_ft3 = None
    if top_drum_vapor_volume_ft3 is not None and (
        (not np.isfinite(top_drum_vapor_volume_ft3)) or top_drum_vapor_volume_ft3 <= 0.0
    ):
        top_drum_vapor_volume_ft3 = None
    if top_drum_total_volume_ft3 is not None and (
        (not np.isfinite(top_drum_total_volume_ft3)) or top_drum_total_volume_ft3 <= 0.0
    ):
        top_drum_total_volume_ft3 = None
    top_drum_extra_vapor_volume_ft3 = None
    if overhead_vapor_adders_ft3 > 0.0:
        if top_drum_total_volume_ft3 is not None:
            # Keep the physical drum shell volume active and add overhead-line /
            # condenser vapor capacitance separately on top of the live drum
            # headspace.
            top_drum_extra_vapor_volume_ft3 = float(overhead_vapor_adders_ft3)
        else:
            top_drum_vapor_volume_ft3 = float(overhead_vapor_adders_ft3)
    if (
        top_drum_total_volume_ft3 is not None
        and top_drum_vapor_volume_ft3 is not None
        and top_drum_extra_vapor_volume_ft3 is None
        and top_drum_vapor_volume_ft3 > top_drum_total_volume_ft3
    ):
        top_drum_vapor_volume_ft3 = float(top_drum_total_volume_ft3)

    bottom_sump_total_volume_ft3 = cfg.bottom_sump_total_volume_ft3
    if bottom_sump_total_volume_ft3 is None:
        bottom_sump_total_volume_ft3 = _spec_float(
            specs,
            "Bottom Sump Total Volume (ft3)",
            "Bottom Sump Volume (ft3)",
            "Bottom Total Volume (ft3)",
            "Bottom Vessel Total Volume (ft3)",
            "Bottom Vessel Volume (ft3)",
            "Bottom Drum Total Volume (ft3)",
            "Bottom Drum Volume (ft3)",
        )
    if bottom_sump_total_volume_ft3 is None:
        d_ft = _spec_float(
            specs,
            "Bottom Sump Diameter (ft)",
            "Bottom Sump ID (ft)",
            "Bottom Vessel Diameter (ft)",
            "Bottom Vessel ID (ft)",
            "Bottom Drum Diameter (ft)",
            "Bottom Drum ID (ft)",
        )
        l_ft = _spec_float(
            specs,
            "Bottom Sump Height (ft)",
            "Bottom Sump height (ft)",
            "Bottom Sump Length (ft)",
            "Bottom Vessel Height (ft)",
            "Bottom Vessel Length (ft)",
            "Bottom Drum Height (ft)",
            "Bottom Drum Length (ft)",
        )
        if d_ft is not None and l_ft is not None and d_ft > 0.0 and l_ft > 0.0:
            bottom_sump_total_volume_ft3 = float(np.pi * 0.25 * float(d_ft) * float(d_ft) * float(l_ft))
    if bottom_sump_total_volume_ft3 is None and hasattr(prov, "liquid_density_lbmol_ft3"):
        bottom_liq_frac = _spec_float(
            specs,
            "Bottom Sump Liquid Fraction (-)",
            "Bottom Sump Liquid Volume Fraction",
            "Bottom Sump Liquid Fraction",
            "Bottom Sump Fill Fraction",
            "Bottom Liquid Fraction (-)",
            "Bottom Liquid Volume Fraction",
            "Bottom Liquid Fraction",
            "Bottom Fill Fraction",
        )
        if bottom_liq_frac is not None and bottom_liq_frac > 1.0 and bottom_liq_frac <= 100.0:
            bottom_liq_frac = float(bottom_liq_frac) / 100.0
        if bottom_liq_frac is not None and 0.0 < float(bottom_liq_frac) <= 1.0:
            bottom_holdup_lbmol = _spec_float(
                specs,
                "Bottom Holdup (lbmol)",
                "Bottom Sump Holdup (lbmol)",
                "Bottom Level Holdup (lbmol)",
            )
            if bottom_holdup_lbmol is not None and bottom_holdup_lbmol >= 0.0 and hasattr(col, "x0") and hasattr(col, "T_f") and hasattr(col, "P_psia"):
                try:
                    x_bot = np.asarray(getattr(col, "x0"), dtype=float).reshape((col.n_stages, col.n_components))[-1, :]
                    T_bot = float(np.asarray(getattr(col, "T_f"), dtype=float).reshape((col.n_stages,))[-1])
                    P_bot = float(np.asarray(getattr(col, "P_psia"), dtype=float).reshape((col.n_stages,))[-1])
                    rho_bot = float(prov.liquid_density_lbmol_ft3(T_bot, P_bot, x_bot))
                    if np.isfinite(rho_bot) and rho_bot > 1e-12:
                        bottom_liq_vol_ft3 = float(bottom_holdup_lbmol) / float(rho_bot)
                        total_try = float(bottom_liq_vol_ft3) / max(float(bottom_liq_frac), 1e-12)
                        if np.isfinite(total_try) and total_try > 0.0:
                            bottom_sump_total_volume_ft3 = float(total_try)
                except Exception:
                    bottom_sump_total_volume_ft3 = None
    if bottom_sump_total_volume_ft3 is not None and (
        (not np.isfinite(bottom_sump_total_volume_ft3)) or bottom_sump_total_volume_ft3 <= 0.0
    ):
        bottom_sump_total_volume_ft3 = None

    mw_components = None
    if hasattr(prov, "component_mw_lbm_per_lbmol"):
        try:
            mw_raw = prov.component_mw_lbm_per_lbmol()
            if mw_raw is not None:
                mw_try = np.asarray(mw_raw, dtype=float).reshape((col.n_components,))
                if np.all(np.isfinite(mw_try)) and np.all(mw_try > 0.0):
                    mw_components = mw_try
        except Exception:
            mw_components = None

    top_pressure_ordering_margin_psi = 0.0
    try:
        top_pressure_ordering_margin_psi = float(cfg.top_pressure_ordering_margin_psi)
    except Exception:
        top_pressure_ordering_margin_psi = 0.0
    if (not np.isfinite(top_pressure_ordering_margin_psi)) or top_pressure_ordering_margin_psi < 0.0:
        top_pressure_ordering_margin_psi = 0.0

    liq_hyd_override_enable = _as_bool(
        _spec_get(
            specs,
            "Enable Liquid Hydraulic Override",
            "Enable Liquid Hydraulics Override",
            "Liquid Hydraulic Override Enabled",
            "Liquid Hydraulics Enabled",
        )
    )
    liq_hyd_override_enable_explicit = liq_hyd_override_enable is not None
    if cfg.enable_liquid_hydraulic_override is not None:
        liq_hyd_override_enable = bool(cfg.enable_liquid_hydraulic_override)
        liq_hyd_override_enable_explicit = True
    if liq_hyd_override_enable is None:
        liq_hyd_override_enable = True

    liq_hyd_override_alpha = cfg.liquid_hydraulic_override_alpha
    liq_hyd_override_alpha_explicit = liq_hyd_override_alpha is not None
    if liq_hyd_override_alpha is None:
        liq_hyd_override_alpha = _spec_float(
            specs,
            "Liquid Hydraulic Override Alpha",
            "Liquid Hydraulics Override Alpha",
            "Liquid Hydraulic Blend",
            "Liquid Hydraulics Blend",
            "Liquid Hydraulics Alpha",
        )
        liq_hyd_override_alpha_explicit = liq_hyd_override_alpha is not None
    liq_hyd_override_alpha = _clip_unit(liq_hyd_override_alpha, default=1.0)

    liq_hyd_model = str(
        cfg.liquid_hydraulic_model
        if cfg.liquid_hydraulic_model is not None
        else (_spec_get(
            specs,
            "Liquid Hydraulic Model",
            "Liquid Hydraulics Model",
            "Liquid Downflow Model",
        ) or "francis")
    ).strip().lower()
    if liq_hyd_model in {"linear", "linear_holdup", "linear-holdup", "skogestad", "skogestad-linear"}:
        liq_hyd_model = "linear-holdup"
    elif liq_hyd_model != "francis":
        liq_hyd_model = "francis"

    liq_hyd_htc_sec = cfg.liquid_hydraulic_htc_sec
    if liq_hyd_htc_sec is None:
        liq_hyd_htc_sec = _spec_float(
            specs,
            "Liquid Hydraulic HTC (sec)",
            "Hydraulic Time Constant (sec)",
        )
        if liq_hyd_htc_sec is None:
            liq_hyd_htc_sec = _spec_float(specs, "Stage time constant [tau] (sec)")
    if liq_hyd_htc_sec is not None and (not np.isfinite(liq_hyd_htc_sec) or liq_hyd_htc_sec <= 0.0):
        liq_hyd_htc_sec = None

    if runtime_mode in ("parity", "calibration"):
        if not (
            liq_hyd_override_enable_explicit
            and bool(liq_hyd_override_enable)
            and str(liq_hyd_model).strip().lower() == "linear-holdup"
        ):
            liq_hyd_override_enable = False
            liq_hyd_override_alpha = 0.0
            liq_hyd_model = "francis"
    elif runtime_mode in ("hydraulic", "total-reflux"):
        # Preserve the documented hydraulic-mode behavior: pressure/hydraulic
        # plus vapor/energy are enabled, but internal liquid hydraulics remain
        # opt-in unless the workbook or CLI explicitly requested them.
        if not liq_hyd_override_enable_explicit:
            liq_hyd_override_enable = False
            liq_hyd_override_alpha = 0.0
        if not liq_hyd_override_enable:
            liq_hyd_override_alpha = 0.0
        # Do not inherit a generic stage time constant as a startup
        # vapor-holdup forcing term in hydraulic mode unless the case or CLI
        # explicitly requested vapor-holdup relaxation.
        if not tau_v_cli_explicit:
            tau_v = None
        # ChemSep/Excel steady profiles already reflect the intended feed-stage
        # split. Re-flashing the feed during hydraulic startup can double-count
        # that effect and pull the initialized state away from the seed.
        if not feed_flash_explicit:
            feed_flash_at_stage = False
    top_drum_gate_soft_psi = cfg.top_drum_pressure_gate_soft_psi
    if top_drum_gate_soft_psi is None:
        top_drum_gate_soft_psi = 0.25

    inputs = ColumnInputs(
        boundary=boundary,
        volume_model=vol,
        runtime_mode=str(runtime_mode),
        condenser_duty_mode=str(cfg.condenser_duty_mode),
        condenser_duty_btu_per_h=(float(cfg.condenser_duty_btu_per_h) if cfg.condenser_duty_btu_per_h is not None else None),
        condenser_duty_trim_btu_per_h=(
            float(cfg.condenser_duty_trim_btu_per_h)
            if cfg.condenser_duty_trim_btu_per_h is not None
            else None
        ),
        enable_live_total_condenser_duty=(runtime_mode != "parity"),
        thermo_provider=prov,
        enable_legacy_temperature_state=(runtime_mode != "parity"),
        equilibrium_relaxation_thermo_provider=eq_relax_prov,
        compute_thermo_diag=True,
        equilibrium_relaxation=bool(cfg.enable_equilibrium_relaxation),
        equilibrium_relaxation_mode=str(eq_mode),
        equilibrium_tau_ramp_initial_sec=(
            float(cfg.equilibrium_tau_ramp_initial_sec)
            if cfg.equilibrium_tau_ramp_initial_sec is not None
            else None
        ),
        equilibrium_tau_ramp_final_sec=(
            float(cfg.equilibrium_tau_ramp_final_sec)
            if cfg.equilibrium_tau_ramp_final_sec is not None
            else None
        ),
        equilibrium_tau_ramp_decay_sec=(
            float(cfg.equilibrium_tau_ramp_decay_sec)
            if cfg.equilibrium_tau_ramp_decay_sec is not None
            else None
        ),
        equilibrium_phase_holdup_guard_lbmol=float(eq_phase_guard_lbmol),
        equilibrium_energy_damping_gain=float(eq_energy_damping_gain),
        hydraulic_energy_temperature_damping=float(hydraulic_energy_temp_damping),
        hydraulic_energy_temperature_mode=str(hydraulic_energy_temp_mode),
        hydraulic_energy_temperature_follow_tau_sec=float(hydraulic_energy_temp_follow_tau_sec),
        hydraulic_energy_temperature_resid_frac=float(hydraulic_energy_temp_resid_frac),
        hydraulic_energy_temperature_pressure_slope_F_per_psi=float(hydraulic_energy_temp_pressure_slope),
        tau_eq_sec=(
            float(cfg.equilibrium_tau_sec)
            if cfg.equilibrium_tau_sec is not None and np.isfinite(float(cfg.equilibrium_tau_sec))
            else (
                float(eq_tau_spec)
                if eq_tau_spec is not None and np.isfinite(float(eq_tau_spec))
                else getattr(col, "tau_eq_sec", None)
            )
        ),
        reboiler_duty_btu_per_h=(float(cfg.reboiler_duty_btu_per_h) if cfg.reboiler_duty_btu_per_h is not None else None),
        reboiler_equilibrium=(runtime_mode != "parity"),
        reboiler_mode=("specified" if runtime_mode == "parity" else "auto"),
        pressure_model=str(pressure_model),
        pressure_top_anchor_psia=None,
        hydraulic_use_top_drum_pressure_as_anchor=False,
        condenser_pressure_drop_psi=(float(condenser_dp_psi) if condenser_dp_psi is not None else None),
        top_drum_vapor_volume_ft3=(
            float(top_drum_vapor_volume_ft3)
            if top_drum_vapor_volume_ft3 is not None
            else None
        ),
        top_drum_extra_vapor_volume_ft3=(
            float(top_drum_extra_vapor_volume_ft3)
            if top_drum_extra_vapor_volume_ft3 is not None
            else None
        ),
        top_drum_total_volume_ft3=(
            float(top_drum_total_volume_ft3)
            if top_drum_total_volume_ft3 is not None
            else None
        ),
        bottom_sump_total_volume_ft3=(
            float(bottom_sump_total_volume_ft3)
            if bottom_sump_total_volume_ft3 is not None
            else None
        ),
        enforce_top_drum_pressure_gate=bool(cfg.enforce_top_drum_pressure_gate),
        top_drum_pressure_gate_soft_psi=(
            float(top_drum_gate_soft_psi)
            if top_drum_gate_soft_psi is not None
            else None
        ),
        enable_top_drum_psv=bool(enable_top_psv),
        top_drum_psv_setpoint_psia=(
            float(top_psv_setpoint_psia)
            if top_psv_setpoint_psia is not None
            else None
        ),
        top_drum_psv_gain_lbmolps_per_psi=(
            float(top_psv_gain_lbmolps_per_psi)
            if top_psv_gain_lbmolps_per_psi is not None
            else None
        ),
        top_drum_psv_max_vent_lbmolps=(
            float(top_psv_max_vent_lbmolps)
            if top_psv_max_vent_lbmolps is not None
            else None
        ),
        vapor_flow_model=str(vapor_flow_model),
        debug_freeze_tray_vapor_derivatives=bool(cfg.debug_freeze_tray_vapor_derivatives),
        debug_override_reflux_composition=bool(cfg.debug_override_reflux_composition),
        debug_clamp_top_drum_pressure_psia=(
            float(cfg.debug_clamp_top_drum_pressure_psia)
            if cfg.debug_clamp_top_drum_pressure_psia is not None
            else None
        ),
        debug_clamp_top_drum_pressure_duration_sec=(
            float(cfg.debug_clamp_top_drum_pressure_duration_sec)
            if cfg.debug_clamp_top_drum_pressure_duration_sec is not None
            else None
        ),
        total_reflux_startup_ramp_tau_sec=(
            float(cfg.total_reflux_startup_ramp_tau_sec)
            if cfg.total_reflux_startup_ramp_tau_sec is not None
            else None
        ),
        total_reflux_startup_min_ramp_fraction=float(cfg.total_reflux_startup_min_ramp_fraction),
        total_reflux_scale_reflux_with_startup_factor=bool(
            cfg.total_reflux_scale_reflux_with_startup_factor
        ),
        total_reflux_boundary_ramp_duration_sec=(
            float(cfg.total_reflux_boundary_ramp_duration_sec)
            if cfg.total_reflux_boundary_ramp_duration_sec is not None
            else None
        ),
        dry_tray_K=float(dry_tray_k),
        vapor_holdup_relaxation_sec=(float(tau_v) if tau_v is not None else None),
        hydraulic_pressure_relaxation_sec=(float(tau_p_hyd) if tau_p_hyd is not None else None),
        top_drum_pressure_temperature_relaxation_sec=(
            float(tau_top_pT) if tau_top_pT is not None else None
        ),
        vapor_flow_relaxation_sec=(float(tau_vflow) if tau_vflow is not None else None),
        conductance_vflow_nominal_hi_ratio=(
            float(conductance_vflow_nominal_hi_ratio)
            if conductance_vflow_nominal_hi_ratio is not None
            else None
        ),
        enforce_top_pressure_ordering=bool(cfg.enforce_top_pressure_ordering),
        top_pressure_ordering_margin_psi=float(top_pressure_ordering_margin_psi),
        reboiler_neighbor_vflow_hi_ratio=(float(reb_nbr_hi) if reb_nbr_hi is not None else 1.20),
        reboiler_neighbor_vflow_lo_ratio=(float(reb_nbr_lo) if reb_nbr_lo is not None else 0.80),
        thermo_refresh_dT_F=(float(thermo_refresh_dT) if thermo_refresh_dT is not None else None),
        thermo_refresh_dP_psia=(float(thermo_refresh_dP) if thermo_refresh_dP is not None else None),
        thermo_refresh_dx=(float(thermo_refresh_dX) if thermo_refresh_dX is not None else None),
        flash_feed_at_stage_conditions=bool(feed_flash_at_stage),
        enable_liquid_hydraulic_override=bool(liq_hyd_override_enable),
        liquid_hydraulic_override_alpha=float(liq_hyd_override_alpha),
        liquid_hydraulic_model=str(liq_hyd_model),
        liquid_hydraulic_htc_sec=(float(liq_hyd_htc_sec) if liq_hyd_htc_sec is not None else None),
        component_mw_lbm_per_lbmol=mw_components,
        progress_hook=_emit_progress,
    )
    try:
        object.__setattr__(inputs, "startup_build_timing_sec", dict(startup_build_timing_sec))
    except Exception:
        pass
    try:
        object.__setattr__(inputs, "startup_build_info", dict(startup_build_info))
    except Exception:
        pass
    return inputs, prov


# -------------------------
# State utilities
# -------------------------


def _clamp_nonnegative_holdups(y: np.ndarray, layout: StateVectorLayout) -> np.ndarray:
    y = np.asarray(y, dtype=float).copy()
    sl = layout.slices()

    y[sl["tray_L"]] = np.clip(y[sl["tray_L"]], 0.0, None)
    if layout.include_vapor:
        y[sl["tray_V"]] = np.clip(y[sl["tray_V"]], 0.0, None)

    if layout.include_top:
        y[sl["top_L"]] = np.clip(y[sl["top_L"]], 0.0, None)
        if layout.include_vapor:
            y[sl["top_V"]] = np.clip(y[sl["top_V"]], 0.0, None)

    if layout.include_bottom:
        y[sl["bottom_L"]] = np.clip(y[sl["bottom_L"]], 0.0, None)
        if layout.include_vapor:
            y[sl["bottom_V"]] = np.clip(y[sl["bottom_V"]], 0.0, None)

    return y


def _clip_temperature_states_to_provider_bounds(
    y: np.ndarray,
    layout: StateVectorLayout,
    thermo_provider: Optional[Any],
) -> np.ndarray:
    """
    Clip temperature states to provider table bounds when available.
    """
    if thermo_provider is None:
        return y
    t_grid = getattr(thermo_provider, "T_grid_F", None)
    if t_grid is None:
        return y
    try:
        g = np.asarray(t_grid, dtype=float).reshape((-1,))
    except Exception:
        return y
    if g.size < 2 or (not np.all(np.isfinite(g))):
        return y
    t_min = float(np.min(g))
    t_max = float(np.max(g))
    if (not np.isfinite(t_min)) or (not np.isfinite(t_max)) or (t_max <= t_min):
        return y

    y_new = np.asarray(y, dtype=float).copy()
    sl = layout.slices()
    if "tray_T_f" in sl:
        y_new[sl["tray_T_f"]] = np.clip(y_new[sl["tray_T_f"]], t_min, t_max)
    if "bottom_T_f" in sl:
        y_new[sl["bottom_T_f"]] = np.clip(y_new[sl["bottom_T_f"]], t_min, t_max)
    return y_new


def _sync_algebraic_tray_temperature_state(
    y: np.ndarray,
    layout: StateVectorLayout,
    diag: Mapping[str, Any],
    thermo_provider: Optional[Any],
) -> np.ndarray:
    """
    Project tray temperature state onto an algebraic tray-temperature diagnostic.

    This is a diagnostic bridge for experimental hydraulic-energy modes where
    RHS thermo is evaluated on an algebraic temperature manifold. Keeping the
    stored tray_T_f state synchronized avoids a ghost-state split between the
    integrator state and the property evaluation temperature.
    """
    sl = layout.slices()
    if "tray_T_f" not in sl:
        return np.asarray(y, dtype=float)
    T_alg = None
    for key in ("T_enthalpy_algebraic_F_tray", "T_bubble_target_F_tray"):
        try:
            arr = np.asarray(diag.get(key), dtype=float).reshape((layout.n_stages,))
        except Exception:
            continue
        if arr.size == int(layout.n_stages) and np.any(np.isfinite(arr)):
            T_alg = arr
            break
    if T_alg is None:
        return np.asarray(y, dtype=float)

    y_new = np.asarray(y, dtype=float).copy()
    T_state = np.asarray(y_new[sl["tray_T_f"]], dtype=float).reshape((layout.n_stages,)).copy()
    mask = np.isfinite(T_alg)
    T_state[mask] = T_alg[mask]
    y_new[sl["tray_T_f"]] = T_state
    return _clip_temperature_states_to_provider_bounds(y_new, layout, thermo_provider)


def _advance_explicit_euler_step(
    *,
    y: np.ndarray,
    dydt: np.ndarray,
    dt_sec: float,
    layout: StateVectorLayout,
    thermo_provider: Optional[Any],
    requested_mode: str = "explicit-euler",
) -> Tuple[np.ndarray, Dict[str, Any]]:
    """
    Advance one explicit-Euler step from a precomputed outer-step derivative.

    The runtime loop already evaluates the outer RHS once per visible step for
    diagnostics and logging. Reusing that derivative avoids a second identical
    RHS evaluation when explicit Euler is selected.
    """

    y0 = np.asarray(y, dtype=float).reshape((-1,))
    dydt0 = np.asarray(dydt, dtype=float).reshape((-1,))
    dt = float(dt_sec)
    if dt <= 0.0:
        raise ValueError("dt_sec must be > 0")
    if y0.shape != dydt0.shape:
        raise ValueError("dydt shape must match y shape")

    y1 = y0 + dt * dydt0
    y1 = _clamp_nonnegative_holdups(y1, layout)
    y1 = _clip_temperature_states_to_provider_bounds(y1, layout, thermo_provider)
    return y1, {
        "requested_mode": str(requested_mode),
        "used_mode": "explicit-euler",
        "fallback_used": False,
        "fallback_reason": "",
        "nfev": 0.0,
        "njev": np.nan,
        "nlu": np.nan,
        "status": np.nan,
        "message": "advanced with precomputed outer-step RHS",
        "n_substeps": 1.0,
        "used_precomputed_rhs": True,
    }


def _integrate_one_step(
    *,
    t_s: float,
    y: np.ndarray,
    dt_sec: float,
    rhs_eval: Any,
    rhs_eval_fallback: Optional[Any],
    layout: StateVectorLayout,
    thermo_provider: Optional[Any],
    integrator_mode: str,
    rtol: Optional[float],
    atol: Optional[float],
    max_step_sec: Optional[float],
    substep_sec: Optional[float],
    max_rhs_evals_per_step: Optional[int],
    step_wall_limit_sec: Optional[float],
) -> Tuple[np.ndarray, Dict[str, Any]]:
    y0 = np.asarray(y, dtype=float).reshape((-1,))
    dt = float(dt_sec)
    if dt <= 0.0:
        raise ValueError("dt_sec must be > 0")

    mode = _normalize_integrator_mode(integrator_mode, default="explicit-euler")
    info: Dict[str, Any] = {
        "requested_mode": str(mode),
        "used_mode": str(mode),
        "fallback_used": False,
        "fallback_reason": "",
        "nfev": np.nan,
        "njev": np.nan,
        "nlu": np.nan,
        "status": np.nan,
        "message": "",
        "n_substeps": np.nan,
    }
    step_t0 = time.perf_counter()
    rhs_eval_count = 0

    def _project_state(y_in: np.ndarray) -> np.ndarray:
        y_out = _clamp_nonnegative_holdups(np.asarray(y_in, dtype=float), layout)
        y_out = _clip_temperature_states_to_provider_bounds(y_out, layout, thermo_provider)
        return y_out

    y0 = _project_state(y0)

    def _explicit_step(reason: str = "") -> Tuple[np.ndarray, Dict[str, Any]]:
        nonlocal rhs_eval_count
        rhs_eval_count += 1
        rhs_cb = rhs_eval_fallback if rhs_eval_fallback is not None else rhs_eval
        dydt, _diag = rhs_cb(float(t_s), y0)
        y1 = y0 + dt * np.asarray(dydt, dtype=float).reshape((-1,))
        y1 = _project_state(y1)
        if reason:
            info["used_mode"] = "explicit-euler"
            info["fallback_used"] = True
            info["fallback_reason"] = str(reason)
        info["nfev"] = float(rhs_eval_count)
        return y1, info

    if mode == "explicit-euler":
        return _explicit_step()

    if _solve_ivp is None:
        return _explicit_step("SciPy solve_ivp unavailable")

    method = "BDF" if mode == "bdf" else "Radau"

    try:
        rtol_use = float(rtol) if rtol is not None else 1.0e-3
    except Exception:
        rtol_use = 1.0e-3
    if (not np.isfinite(rtol_use)) or rtol_use <= 0.0:
        rtol_use = 1.0e-3

    try:
        atol_use = float(atol) if atol is not None else 1.0e-6
    except Exception:
        atol_use = 1.0e-6
    if (not np.isfinite(atol_use)) or atol_use <= 0.0:
        atol_use = 1.0e-6

    max_step_use = float("inf")
    if max_step_sec is not None:
        try:
            ms = float(max_step_sec)
        except Exception:
            ms = np.nan
        if np.isfinite(ms) and ms > 0.0:
            max_step_use = min(float(ms), float(dt))

    sub_dt = float(dt)
    if substep_sec is not None:
        try:
            sub_try = float(substep_sec)
        except Exception:
            sub_try = np.nan
        if np.isfinite(sub_try) and sub_try > 0.0:
            sub_dt = min(float(dt), float(sub_try))
    n_substeps = max(1, int(np.ceil(float(dt) / max(float(sub_dt), 1.0e-12))))
    info["n_substeps"] = float(n_substeps)

    max_rhs = None
    if max_rhs_evals_per_step is not None:
        try:
            max_rhs_try = int(max_rhs_evals_per_step)
        except Exception:
            max_rhs_try = 0
        if max_rhs_try > 0:
            max_rhs = int(max_rhs_try)

    wall_limit = None
    if step_wall_limit_sec is not None:
        try:
            wall_try = float(step_wall_limit_sec)
        except Exception:
            wall_try = np.nan
        if np.isfinite(wall_try) and wall_try > 0.0:
            wall_limit = float(wall_try)

    def _rhs_ivp(t_eval: float, y_eval: np.ndarray) -> np.ndarray:
        nonlocal rhs_eval_count
        rhs_eval_count += 1
        if max_rhs is not None and rhs_eval_count > int(max_rhs):
            raise RuntimeError(f"max RHS evals exceeded ({int(max_rhs)})")
        if wall_limit is not None and (time.perf_counter() - step_t0) > float(wall_limit):
            raise RuntimeError(f"step wall-time limit exceeded ({float(wall_limit):.3g}s)")
        y_rhs = np.asarray(y_eval, dtype=float).reshape((-1,))
        if not np.all(np.isfinite(y_rhs)):
            raise RuntimeError("non-finite state encountered in stiff substep")
        dydt, _diag = rhs_eval(float(t_eval), y_rhs)
        dydt_out = np.asarray(dydt, dtype=float).reshape((-1,))
        if not np.all(np.isfinite(dydt_out)):
            raise RuntimeError("non-finite derivative encountered in stiff substep")
        return dydt_out

    y_curr = np.asarray(y0, dtype=float).reshape((-1,))
    t_curr = float(t_s)
    nfev_total = 0.0
    njev_total = 0.0
    nlu_total = 0.0
    status_last = 0.0
    msg_last = ""

    for k in range(n_substeps):
        t_next = float(min(float(t_s) + float(dt), t_curr + float(sub_dt)))
        dt_k = float(t_next - t_curr)
        if dt_k <= 0.0:
            continue
        max_step_k = min(float(max_step_use), float(dt_k))
        try:
            sol = _solve_ivp(
                fun=_rhs_ivp,
                t_span=(float(t_curr), float(t_next)),
                y0=np.asarray(y_curr, dtype=float).reshape((-1,)),
                method=method,
                t_eval=[float(t_next)],
                rtol=float(rtol_use),
                atol=float(atol_use),
                max_step=float(max_step_k),
                vectorized=False,
            )
        except Exception as exc:
            return _explicit_step(f"{method} integrator exception (substep {k+1}/{n_substeps}): {exc}")

        nfev_total += float(getattr(sol, "nfev", 0.0))
        njev_total += float(getattr(sol, "njev", 0.0))
        nlu_total += float(getattr(sol, "nlu", 0.0))
        status_last = float(getattr(sol, "status", np.nan))
        msg_last = str(getattr(sol, "message", ""))

        y_sol = getattr(sol, "y", None)
        if (
            (not bool(getattr(sol, "success", False)))
            or y_sol is None
            or np.asarray(y_sol).ndim != 2
            or np.asarray(y_sol).shape[1] < 1
        ):
            msg = str(getattr(sol, "message", "")).strip()
            if not msg:
                msg = f"{method} integrator failed at substep {k+1}/{n_substeps}"
            return _explicit_step(msg)

        y_next = np.asarray(y_sol, dtype=float)[:, -1].reshape((-1,))
        if not np.all(np.isfinite(y_next)):
            return _explicit_step(f"{method} produced non-finite state at substep {k+1}/{n_substeps}")
        y_curr = _project_state(y_next)
        t_curr = float(t_next)

    info["nfev"] = float(nfev_total)
    info["njev"] = float(njev_total)
    info["nlu"] = float(nlu_total)
    info["status"] = float(status_last)
    info["message"] = str(msg_last)
    return y_curr, info


def _integrate_one_step_ida(
    *,
    t_s: float,
    y: np.ndarray,
    dt_sec: float,
    rhs_eval: Any,
    layout: StateVectorLayout,
    thermo_provider: Optional[Any],
    substep_sec: Optional[float],
    max_iter: Optional[int],
    relax: Optional[float],
    rtol: Optional[float],
    atol: Optional[float],
    max_rhs_evals_per_step: Optional[int],
    step_wall_limit_sec: Optional[float],
    alg_p_tol_psia: Optional[float] = None,
    alg_v_tol_lbmolph: Optional[float] = None,
) -> Tuple[np.ndarray, Dict[str, Any]]:
    """
    Pilot simultaneous-DAE stepper (IDA-style fixed-point implicit Euler).

    For each substep it solves:
      y_{k+1} = y_k + dt * f(t_{k+1}, y_{k+1})
    by fixed-point iteration. When rhs_eval includes DAE algebraic closure
    (e.g., pilot Newton solve for P/V), this yields a lightweight
    simultaneous differential+algebraic step without SciPy solve_ivp.
    """
    y0 = np.asarray(y, dtype=float).reshape((-1,))
    dt = float(dt_sec)
    if dt <= 0.0:
        raise ValueError("dt_sec must be > 0")

    info: Dict[str, Any] = {
        "requested_mode": "ida",
        "used_mode": "ida",
        "fallback_used": False,
        "fallback_reason": "",
        "nfev": np.nan,
        "njev": np.nan,
        "nlu": np.nan,
        "status": np.nan,
        "message": "",
        "n_substeps": np.nan,
        "ida_iter_max": np.nan,
        "ida_iter_mean": np.nan,
        "ida_converged": 0.0,
        "ida_last_err": np.nan,
        "ida_alg_p_inf_psia": np.nan,
        "ida_alg_v_inf_lbmolph": np.nan,
        "ida_alg_weighted": np.nan,
        "ida_alg_converged": np.nan,
        "ida_resid_energy_btups": np.nan,
    }
    step_t0 = time.perf_counter()

    def _project_state(y_in: np.ndarray) -> np.ndarray:
        y_out = _clamp_nonnegative_holdups(np.asarray(y_in, dtype=float), layout)
        y_out = _clip_temperature_states_to_provider_bounds(y_out, layout, thermo_provider)
        return y_out

    y0 = _project_state(y0)

    try:
        n_iter = int(max_iter) if max_iter is not None else 8
    except Exception:
        n_iter = 8
    n_iter = max(1, n_iter)

    try:
        relax_use = float(relax) if relax is not None else 1.0
    except Exception:
        relax_use = 1.0
    if (not np.isfinite(relax_use)) or relax_use <= 0.0:
        relax_use = 1.0
    if relax_use > 1.0:
        relax_use = 1.0

    try:
        rtol_use = float(rtol) if rtol is not None else 1.0e-3
    except Exception:
        rtol_use = 1.0e-3
    if (not np.isfinite(rtol_use)) or rtol_use <= 0.0:
        rtol_use = 1.0e-3

    try:
        atol_use = float(atol) if atol is not None else 1.0e-6
    except Exception:
        atol_use = 1.0e-6
    if (not np.isfinite(atol_use)) or atol_use <= 0.0:
        atol_use = 1.0e-6

    try:
        alg_p_tol_use = float(alg_p_tol_psia) if alg_p_tol_psia is not None else 0.05
    except Exception:
        alg_p_tol_use = 0.05
    if (not np.isfinite(alg_p_tol_use)) or alg_p_tol_use <= 0.0:
        alg_p_tol_use = 0.05

    try:
        alg_v_tol_use = float(alg_v_tol_lbmolph) if alg_v_tol_lbmolph is not None else 25.0
    except Exception:
        alg_v_tol_use = 25.0
    if (not np.isfinite(alg_v_tol_use)) or alg_v_tol_use <= 0.0:
        alg_v_tol_use = 25.0

    sub_dt = float(dt)
    if substep_sec is not None:
        try:
            sub_try = float(substep_sec)
        except Exception:
            sub_try = np.nan
        if np.isfinite(sub_try) and sub_try > 0.0:
            sub_dt = min(float(dt), float(sub_try))
    n_substeps = max(1, int(np.ceil(float(dt) / max(float(sub_dt), 1.0e-12))))
    info["n_substeps"] = float(n_substeps)

    max_rhs = None
    if max_rhs_evals_per_step is not None:
        try:
            max_rhs_try = int(max_rhs_evals_per_step)
        except Exception:
            max_rhs_try = 0
        if max_rhs_try > 0:
            max_rhs = int(max_rhs_try)

    wall_limit = None
    if step_wall_limit_sec is not None:
        try:
            wall_try = float(step_wall_limit_sec)
        except Exception:
            wall_try = np.nan
        if np.isfinite(wall_try) and wall_try > 0.0:
            wall_limit = float(wall_try)

    rhs_eval_count = 0
    iter_counts: List[int] = []
    last_err = float("nan")
    last_alg_p = float("nan")
    last_alg_v = float("nan")
    last_alg_weighted = float("nan")
    last_energy_resid = float("nan")
    converged_all = True

    def _rhs(t_eval: float, y_eval: np.ndarray) -> Tuple[np.ndarray, Dict[str, Any]]:
        nonlocal rhs_eval_count
        rhs_eval_count += 1
        if max_rhs is not None and rhs_eval_count > int(max_rhs):
            raise RuntimeError(f"max RHS evals exceeded ({int(max_rhs)})")
        if wall_limit is not None and (time.perf_counter() - step_t0) > float(wall_limit):
            raise RuntimeError(f"step wall-time limit exceeded ({float(wall_limit):.3g}s)")
        dydt, diag = rhs_eval(float(t_eval), np.asarray(y_eval, dtype=float).reshape((-1,)))
        dydt_out = np.asarray(dydt, dtype=float).reshape((-1,))
        if not np.all(np.isfinite(dydt_out)):
            raise RuntimeError("non-finite derivative in ida fixed-point step")
        diag_out = dict(diag) if isinstance(diag, dict) else {}
        return dydt_out, diag_out

    y_curr = np.asarray(y0, dtype=float).reshape((-1,))
    t_curr = float(t_s)
    fallback_reason = ""

    try:
        for k in range(n_substeps):
            t_next = float(min(float(t_s) + float(dt), t_curr + float(sub_dt)))
            dt_k = float(t_next - t_curr)
            if dt_k <= 0.0:
                continue

            # Explicit predictor from current state.
            dydt_pred, _diag_pred = _rhs(float(t_curr), y_curr)
            y_guess = _project_state(y_curr + float(dt_k) * dydt_pred)

            converged = False
            it_used = 0
            for it in range(n_iter):
                it_used = int(it + 1)
                dydt_it, diag_it = _rhs(float(t_next), y_guess)
                y_impl = _project_state(y_curr + float(dt_k) * dydt_it)
                y_new = _project_state(y_guess + float(relax_use) * (y_impl - y_guess))

                dy = np.asarray(y_new, dtype=float) - np.asarray(y_guess, dtype=float)
                scale = float(atol_use) + float(rtol_use) * np.maximum(
                    np.abs(np.asarray(y_new, dtype=float)),
                    np.abs(np.asarray(y_guess, dtype=float)),
                )
                scale = np.where(scale > 1.0e-14, scale, 1.0e-14)
                err = float(np.max(np.abs(dy) / scale)) if dy.size > 0 else 0.0
                last_err = err

                alg_p_inf = abs(_mapping_scalar(diag_it, "dae_pilot_alg_p_inf_psia"))
                alg_v_inf = abs(_mapping_scalar(diag_it, "dae_pilot_alg_v_inf_lbmolph"))
                resid_energy = abs(_mapping_scalar(diag_it, "resid_energy_btups"))

                terms: List[float] = []
                if np.isfinite(alg_p_inf):
                    last_alg_p = float(alg_p_inf)
                    terms.append(float(alg_p_inf) / float(alg_p_tol_use))
                if np.isfinite(alg_v_inf):
                    last_alg_v = float(alg_v_inf)
                    terms.append(float(alg_v_inf) / float(alg_v_tol_use))
                if np.isfinite(resid_energy):
                    last_energy_resid = float(resid_energy)
                alg_weighted = float(np.max(np.asarray(terms, dtype=float))) if terms else float("nan")
                if np.isfinite(alg_weighted):
                    last_alg_weighted = float(alg_weighted)
                alg_ok = (not np.isfinite(alg_weighted)) or (alg_weighted <= 1.0)

                y_guess = y_new
                if np.isfinite(err) and err <= 1.0 and bool(alg_ok):
                    converged = True
                    break

            iter_counts.append(int(it_used))
            if not converged:
                converged_all = False
                fallback_reason = (
                    f"IDA fixed-point did not converge at substep {k+1}/{n_substeps} "
                    f"(iters={it_used}, err={last_err:.3g}, alg={last_alg_weighted:.3g})"
                )
                break

            y_curr = np.asarray(y_guess, dtype=float).reshape((-1,))
            t_curr = float(t_next)
    except Exception as exc:
        converged_all = False
        fallback_reason = str(exc)

    if not converged_all:
        # Step-level explicit fallback from original y0.
        try:
            dydt_fb, _diag_fb = _rhs(float(t_s), y0)
            y_fb = _project_state(y0 + float(dt) * np.asarray(dydt_fb, dtype=float).reshape((-1,)))
        except Exception:
            y_fb = _project_state(np.asarray(y0, dtype=float))
        info["used_mode"] = "explicit-euler"
        info["fallback_used"] = True
        info["fallback_reason"] = str(fallback_reason or "ida step failed")
        info["nfev"] = float(rhs_eval_count)
        if iter_counts:
            info["ida_iter_max"] = float(np.max(np.asarray(iter_counts, dtype=float)))
            info["ida_iter_mean"] = float(np.mean(np.asarray(iter_counts, dtype=float)))
        info["ida_converged"] = 0.0
        info["ida_last_err"] = float(last_err)
        info["ida_alg_p_inf_psia"] = float(last_alg_p)
        info["ida_alg_v_inf_lbmolph"] = float(last_alg_v)
        info["ida_alg_weighted"] = float(last_alg_weighted)
        info["ida_alg_converged"] = (
            1.0 if ((not np.isfinite(last_alg_weighted)) or (last_alg_weighted <= 1.0)) else 0.0
        )
        info["ida_resid_energy_btups"] = float(last_energy_resid)
        return y_fb, info

    info["nfev"] = float(rhs_eval_count)
    if iter_counts:
        info["ida_iter_max"] = float(np.max(np.asarray(iter_counts, dtype=float)))
        info["ida_iter_mean"] = float(np.mean(np.asarray(iter_counts, dtype=float)))
    info["ida_converged"] = 1.0
    info["ida_last_err"] = float(last_err)
    info["ida_alg_p_inf_psia"] = float(last_alg_p)
    info["ida_alg_v_inf_lbmolph"] = float(last_alg_v)
    info["ida_alg_weighted"] = float(last_alg_weighted)
    info["ida_alg_converged"] = (
        1.0 if ((not np.isfinite(last_alg_weighted)) or (last_alg_weighted <= 1.0)) else 0.0
    )
    info["ida_resid_energy_btups"] = float(last_energy_resid)
    return y_curr, info


def _clear_initial_tray_vapor_holdup(y: np.ndarray, layout: StateVectorLayout) -> np.ndarray:
    """Clear initial tray vapor holdup states before pressure-based MV initialization."""
    if not layout.include_vapor:
        return y
    y_new = np.asarray(y, dtype=float).copy()
    sl = layout.slices()
    y_new[sl["tray_V"]] = 0.0
    return y_new


def _tray_temperature_F(col: ColumnSpec, layout: StateVectorLayout, y: np.ndarray, include_temperature: bool) -> np.ndarray:
    N = col.n_stages
    if include_temperature:
        u = layout.unpack(y)
        if "tray_T_f" in u:
            return np.asarray(u["tray_T_f"], dtype=float).reshape((N,))
    return np.asarray(getattr(col, "T_f", np.full(N, np.nan, dtype=float)), dtype=float).reshape((N,))


def _total_inventory_lbmol(layout: StateVectorLayout, y: np.ndarray) -> float:
    """Total molar inventory over all holdup states (lbmol)."""
    u = layout.unpack(np.asarray(y, dtype=float))
    total = 0.0
    # When a separate top drum state is present, tray-1 liquid/vapor are tied to
    # that same inventory in the RHS. Exclude tray-1 from global totals to avoid
    # double counting in mass-closure diagnostics.
    has_top_drum = bool(getattr(layout, "include_top", False)) and ("top_L" in u)
    for key in ("tray_L", "tray_V", "top_L", "top_V", "bottom_L", "bottom_V"):
        if key not in u:
            continue
        try:
            arr = np.asarray(u[key], dtype=float)
            if has_top_drum and key in ("tray_L", "tray_V") and arr.ndim == 2 and arr.shape[0] > 0:
                arr = arr[1:, :]
            total += float(np.nansum(arr))
        except Exception:
            pass
    return float(total)


def _total_inventory_rate_lbmolps(layout: StateVectorLayout, dydt: np.ndarray) -> float:
    """Total inventory time derivative over molar holdup states (lbmol/s)."""
    ud = layout.unpack(np.asarray(dydt, dtype=float))
    total = 0.0
    has_top_drum = bool(getattr(layout, "include_top", False)) and ("top_L" in ud)
    for key in ("tray_L", "tray_V", "top_L", "top_V", "bottom_L", "bottom_V"):
        if key not in ud:
            continue
        try:
            arr = np.asarray(ud[key], dtype=float)
            if has_top_drum and key in ("tray_L", "tray_V") and arr.ndim == 2 and arr.shape[0] > 0:
                arr = arr[1:, :]
            total += float(np.nansum(arr))
        except Exception:
            pass
    return float(total)


def _vapor_volume_ft3_per_stage(vol: VolumeModel, N: int) -> np.ndarray:
    if vol.vapor_volume_ft3_per_stage is not None:
        V = np.asarray(vol.vapor_volume_ft3_per_stage, dtype=float).reshape((N,))
    else:
        V = np.full(N, float(vol.default_vapor_volume_ft3), dtype=float)
    V = np.where(~np.isfinite(V) | (V <= 0.0), 1.0, V)
    return V


def _pressure_diag_psia(
    col: ColumnSpec,
    vol: VolumeModel,
    T_F: np.ndarray,
    MV_tot_tray: np.ndarray,
    Z_tray: np.ndarray,
) -> np.ndarray:
    N = col.n_stages
    MV = np.asarray(MV_tot_tray, dtype=float).reshape((N,))
    Z = np.asarray(Z_tray, dtype=float).reshape((N,))
    Z = np.where(~np.isfinite(Z) | (Z <= 0.0), 1.0, Z)

    V = _vapor_volume_ft3_per_stage(vol, N)

    R = 10.7316  # (psia*ft3)/(lbmol*R)
    T_R = np.asarray(T_F, dtype=float).reshape((N,)) + 459.67
    T_R = np.where(~np.isfinite(T_R) | (T_R <= 1e-6), 559.67, T_R)

    P = MV * Z * R * T_R / V
    P = np.where(np.isfinite(P), P, np.nan)

    # Condenser stage pin to spec, if present
    try:
        P_spec0 = float(np.asarray(getattr(col, "P_psia", [np.nan]), dtype=float).reshape((-1,))[0])
        if np.isfinite(P_spec0):
            P[0] = P_spec0
    except Exception:
        pass

    return P


def _as_stage_vector(
    arr: Optional[np.ndarray],
    n_stages: int,
    *,
    positive_only: bool = False,
) -> Optional[np.ndarray]:
    if arr is None:
        return None
    try:
        out = np.asarray(arr, dtype=float).reshape((int(n_stages),)).copy()
    except Exception:
        return None
    if out.size != int(n_stages):
        return None
    valid = np.isfinite(out)
    if positive_only:
        valid = valid & (out > 0.0)
    if not np.any(valid):
        return None
    return out


def _diag_stage_vector(
    diag: Dict[str, np.ndarray],
    key: str,
    n_stages: int,
    *,
    positive_only: bool = False,
) -> Optional[np.ndarray]:
    if key not in diag:
        return None
    return _as_stage_vector(diag.get(key), n_stages, positive_only=positive_only)


def _max_abs_delta(a: Optional[np.ndarray], b: Optional[np.ndarray], *, positive_only: bool = False) -> float:
    if a is None or b is None:
        return float("nan")
    try:
        aa = np.asarray(a, dtype=float).reshape((-1,))
        bb = np.asarray(b, dtype=float).reshape((-1,))
    except Exception:
        return float("nan")
    if aa.size != bb.size or aa.size == 0:
        return float("nan")
    mask = np.isfinite(aa) & np.isfinite(bb)
    if positive_only:
        mask = mask & (aa > 0.0) & (bb > 0.0)
    if not np.any(mask):
        return float("nan")
    return float(np.max(np.abs(aa[mask] - bb[mask])))


def _column_rhs_with_inner_pv_coupling(
    *,
    t_s: float,
    y: np.ndarray,
    col: ColumnSpec,
    layout: StateVectorLayout,
    inputs: ColumnInputs,
    max_iter: int,
    p_tol_psia: Optional[float],
    v_tol_lbmolph: Optional[float],
) -> Tuple[np.ndarray, Dict[str, np.ndarray]]:
    """
    Fixed-point inner coupling for pressure and vapor flow within one timestep.

    This iterates RHS evaluation by feeding the newly computed tray pressure and
    vapor outflow back as `P_tray_prev` and `V_out_prev_lbmolph` before the next
    inner pass. It is intentionally lightweight and keeps the outer integrator
    explicit.
    """
    try:
        n_iter_req = int(max_iter)
    except Exception:
        n_iter_req = 1
    n_iter_req = max(n_iter_req, 1)

    if p_tol_psia is not None:
        try:
            p_tol_psia = float(p_tol_psia)
        except Exception:
            p_tol_psia = None
        if p_tol_psia is not None and ((not np.isfinite(p_tol_psia)) or p_tol_psia <= 0.0):
            p_tol_psia = None
    if v_tol_lbmolph is not None:
        try:
            v_tol_lbmolph = float(v_tol_lbmolph)
        except Exception:
            v_tol_lbmolph = None
        if v_tol_lbmolph is not None and ((not np.isfinite(v_tol_lbmolph)) or v_tol_lbmolph <= 0.0):
            v_tol_lbmolph = None

    n_stages = int(getattr(col, "n_stages", 0))
    inputs_iter = inputs
    p_prev = _as_stage_vector(inputs.P_tray_prev, n_stages, positive_only=True)
    v_prev = _as_stage_vector(inputs.V_out_prev_lbmolph, n_stages, positive_only=False)

    converged = False
    dp_last = float("nan")
    dv_last = float("nan")
    iter_count = 0
    dydt = np.zeros_like(np.asarray(y, dtype=float))
    diag: Dict[str, np.ndarray] = {}

    for it in range(n_iter_req):
        dydt, diag = column_rhs(t_s, y, col, layout, inputs=inputs_iter)
        iter_count = int(it + 1)

        p_next = _diag_stage_vector(diag, "P_psia_hyd", n_stages, positive_only=True)
        if p_next is None:
            p_next = _diag_stage_vector(diag, "P_psia_diag", n_stages, positive_only=True)
        v_next = _diag_stage_vector(diag, "V_out_lbmolph", n_stages, positive_only=False)

        dp_last = _max_abs_delta(p_next, p_prev, positive_only=True)
        dv_last = _max_abs_delta(v_next, v_prev, positive_only=False)

        if it >= 1:
            p_ok = True if p_tol_psia is None else (np.isfinite(dp_last) and dp_last <= float(p_tol_psia))
            v_ok = True if v_tol_lbmolph is None else (np.isfinite(dv_last) and dv_last <= float(v_tol_lbmolph))
            if p_ok and v_ok:
                converged = True
                break

        if it >= (n_iter_req - 1):
            break

        updates: Dict[str, Any] = {}
        if p_next is not None:
            updates["P_tray_prev"] = np.asarray(p_next, dtype=float).copy()
            p_prev = np.asarray(p_next, dtype=float).copy()
        if v_next is not None:
            updates["V_out_prev_lbmolph"] = np.asarray(v_next, dtype=float).copy()
            v_prev = np.asarray(v_next, dtype=float).copy()
        z_next = _diag_stage_vector(diag, "Z_tray", n_stages, positive_only=True)
        if z_next is not None:
            updates["Zfac_prev"] = np.asarray(z_next, dtype=float).copy()
        if "T_top_drum_pressure_used_F" in diag:
            try:
                top_t = float(np.asarray(diag["T_top_drum_pressure_used_F"], dtype=float).reshape((-1,))[0])
                if np.isfinite(top_t):
                    updates["top_drum_pressure_T_prev_F"] = float(top_t)
            except Exception:
                pass
        if not updates:
            break
        inputs_iter = replace(inputs_iter, **updates)

    diag["pv_inner_iter_count"] = np.array([float(iter_count)], dtype=float)
    diag["pv_inner_converged"] = np.array([1.0 if converged else 0.0], dtype=float)
    diag["pv_inner_dp_max_psia"] = np.array([float(dp_last)], dtype=float)
    diag["pv_inner_dv_max_lbmolph"] = np.array([float(dv_last)], dtype=float)
    return dydt, diag


def _solve_dae_pilot_algebraic(
    *,
    t_s: float,
    y: np.ndarray,
    col: ColumnSpec,
    layout: StateVectorLayout,
    inputs: ColumnInputs,
    max_iter: int,
    p_tol_psia: Optional[float],
    v_tol_lbmolph: Optional[float],
    jac_rel_step: float,
    line_search_max: int,
    rhs_func: Optional[Any] = None,
) -> Tuple[np.ndarray, Dict[str, np.ndarray]]:
    """
    Pilot simultaneous algebraic solve for z=[P_tray, V_out].

    Differential update remains explicit in time. Within a timestep, this
    routine iterates algebraic consistency using finite-difference Newton
    updates on:
      r_p = z_p - P_rhs(t, y, z)
      r_v = z_v - V_rhs(t, y, z)
    """
    rhs_callable = rhs_func if rhs_func is not None else column_rhs
    y_arr = np.asarray(y, dtype=float).reshape((-1,))
    n_stages = int(getattr(col, "n_stages", 0))
    if n_stages <= 0:
        dydt, diag = rhs_callable(float(t_s), y_arr, col, layout, inputs=inputs)
        return np.asarray(dydt, dtype=float).reshape((-1,)), dict(diag)

    try:
        n_iter = max(1, int(max_iter))
    except Exception:
        n_iter = 1
    try:
        ls_max = max(1, int(line_search_max))
    except Exception:
        ls_max = 1

    def _norm_tol(val: Optional[float]) -> Optional[float]:
        if val is None:
            return None
        try:
            v = float(val)
        except Exception:
            return None
        if not np.isfinite(v) or v <= 0.0:
            return None
        return v

    p_tol = _norm_tol(p_tol_psia)
    v_tol = _norm_tol(v_tol_lbmolph)
    try:
        jac_rel = float(jac_rel_step)
    except Exception:
        jac_rel = 1.0e-6
    if (not np.isfinite(jac_rel)) or jac_rel <= 0.0:
        jac_rel = 1.0e-6

    p_fallback = None
    if hasattr(col, "P_psia"):
        try:
            p_fallback = np.asarray(getattr(col, "P_psia"), dtype=float).reshape((n_stages,))
        except Exception:
            p_fallback = None
    if p_fallback is None:
        p_fallback = np.full(n_stages, 200.0, dtype=float)

    v_fallback = None
    if hasattr(col, "V_lbmolph"):
        try:
            v_fallback = np.asarray(getattr(col, "V_lbmolph"), dtype=float).reshape((n_stages,))
        except Exception:
            v_fallback = None
    if v_fallback is None:
        v_fallback = np.zeros(n_stages, dtype=float)

    seed_diag: Dict[str, np.ndarray] = {}
    if inputs.P_tray_prev is not None:
        try:
            seed_diag["P_psia_hyd"] = np.asarray(inputs.P_tray_prev, dtype=float).reshape((n_stages,))
        except Exception:
            pass
    if inputs.V_out_prev_lbmolph is not None:
        try:
            seed_diag["V_out_lbmolph"] = np.asarray(inputs.V_out_prev_lbmolph, dtype=float).reshape((n_stages,))
        except Exception:
            pass
    z_seed = default_algebraic_seed(
        n_stages=n_stages,
        diag=(seed_diag if seed_diag else None),
        p_fallback_psia=p_fallback,
        v_fallback_lbmolph=v_fallback,
    )
    z_work = np.asarray(z_seed, dtype=float).copy()

    def _alg_residual(z_trial: np.ndarray) -> np.ndarray:
        rr = evaluate_pilot_residual(
            t_s=float(t_s),
            y=y_arr,
            ydot=np.zeros_like(y_arr),
            z=np.asarray(z_trial, dtype=float),
            col=col,
            layout=layout,
            inputs=inputs,
            rhs_func=rhs_callable,
        )
        return np.concatenate([rr.alg_pressure, rr.alg_vapor], axis=0)

    converged = False
    iter_count = 0
    final_rr = None
    alg_p_inf = float("nan")
    alg_v_inf = float("nan")
    alg_full_inf = float("nan")
    failed = False

    try:
        for it in range(n_iter):
            rr = evaluate_pilot_residual(
                t_s=float(t_s),
                y=y_arr,
                ydot=np.zeros_like(y_arr),
                z=z_work,
                col=col,
                layout=layout,
                inputs=inputs,
                rhs_func=rhs_callable,
            )
            iter_count = int(it + 1)
            alg_vec = np.concatenate([rr.alg_pressure, rr.alg_vapor], axis=0)
            alg_p_inf = inf_norm(rr.alg_pressure)
            alg_v_inf = inf_norm(rr.alg_vapor)
            alg_full_inf = inf_norm(alg_vec)
            final_rr = rr

            p_ok = True if p_tol is None else (np.isfinite(alg_p_inf) and alg_p_inf <= float(p_tol))
            v_ok = True if v_tol is None else (np.isfinite(alg_v_inf) and alg_v_inf <= float(v_tol))
            if p_ok and v_ok:
                converged = True
                break
            if it >= (n_iter - 1):
                break

            J = finite_difference_jacobian(_alg_residual, z_work, rel_step=jac_rel)
            if J.shape[0] != alg_vec.size or J.shape[1] != z_work.size:
                failed = True
                break
            if (not np.all(np.isfinite(J))) or (not np.all(np.isfinite(alg_vec))):
                failed = True
                break
            delta, *_ = np.linalg.lstsq(J, -alg_vec, rcond=None)
            delta = np.asarray(delta, dtype=float).reshape((z_work.size,))
            if not np.all(np.isfinite(delta)):
                failed = True
                break

            base_norm = inf_norm(alg_vec)
            accepted = False
            alpha = 1.0
            for _ls in range(ls_max):
                z_try = z_work + float(alpha) * delta
                alg_try = _alg_residual(z_try)
                n_try = inf_norm(alg_try)
                if np.isfinite(n_try) and ((not np.isfinite(base_norm)) or n_try <= base_norm):
                    z_work = np.asarray(z_try, dtype=float)
                    accepted = True
                    break
                alpha *= 0.5
            if not accepted:
                failed = True
                break

        if final_rr is None:
            final_rr = evaluate_pilot_residual(
                t_s=float(t_s),
                y=y_arr,
                ydot=np.zeros_like(y_arr),
                z=z_work,
                col=col,
                layout=layout,
                inputs=inputs,
                rhs_func=rhs_callable,
            )

        rr_consistent = evaluate_pilot_residual(
            t_s=float(t_s),
            y=y_arr,
            ydot=np.asarray(final_rr.dydt_rhs, dtype=float),
            z=z_work,
            col=col,
            layout=layout,
            inputs=inputs,
            rhs_func=rhs_callable,
        )
        dydt_out = np.asarray(rr_consistent.dydt_rhs, dtype=float).reshape((-1,))
        diag_out = dict(rr_consistent.diag)
    except Exception:
        failed = True
        dydt_out, diag_fallback = rhs_callable(float(t_s), y_arr, col, layout, inputs=inputs)
        dydt_out = np.asarray(dydt_out, dtype=float).reshape((-1,))
        diag_out = dict(diag_fallback)

    z_update_inf = inf_norm(np.asarray(z_work, dtype=float) - np.asarray(z_seed, dtype=float))
    diag_out["dae_pilot_enabled"] = np.array([1.0], dtype=float)
    diag_out["dae_pilot_iter_count"] = np.array([float(iter_count)], dtype=float)
    diag_out["dae_pilot_converged"] = np.array([1.0 if converged else 0.0], dtype=float)
    diag_out["dae_pilot_failed"] = np.array([1.0 if failed else 0.0], dtype=float)
    diag_out["dae_pilot_alg_p_inf_psia"] = np.array([float(alg_p_inf)], dtype=float)
    diag_out["dae_pilot_alg_v_inf_lbmolph"] = np.array([float(alg_v_inf)], dtype=float)
    diag_out["dae_pilot_alg_full_inf"] = np.array([float(alg_full_inf)], dtype=float)
    diag_out["dae_pilot_z_update_inf"] = np.array([float(z_update_inf)], dtype=float)
    return dydt_out, diag_out


def _initialize_vapor_holdup_from_spec_pressure(
    *,
    col: ColumnSpec,
    layout: StateVectorLayout,
    y: np.ndarray,
    inputs: ColumnInputs,
    include_temperature: bool,
    return_diag: bool = False,
    preserve_tray_vapor_holdup: bool = False,
) -> Any:
    """Initialize tray vapor holdup MV from P_spec using PV=nZRT/V.

    Done once at t=0 so PV diagnostic pressure starts near specified pressure profile.
    Stage index 0 (condenser) left as MV=0 unless tray vapor holdup is preserved.
    """
    if not layout.include_vapor:
        return (y, {}) if return_diag else y

    P_spec = getattr(col, "P_psia", None)
    if P_spec is None:
        return (y, {}) if return_diag else y
    P_spec = np.asarray(P_spec, dtype=float).reshape((col.n_stages,))

    N = col.n_stages
    Z0 = None
    startup_diag: Dict[str, Any] = {}
    if inputs.Zfac_prev is not None:
        try:
            Z0 = np.asarray(inputs.Zfac_prev, dtype=float).reshape((N,))
        except Exception:
            Z0 = None
    if Z0 is None:
        direct_startup_ok = False
        if inputs.thermo_provider is not None:
            _emit_progress("[Init] Vapor holdup initialization  solving startup tray Z factors from direct tray thermo refresh")
            try:
                startup_diag = _direct_startup_tray_thermo_diag(
                    col=col,
                    layout=layout,
                    y=y,
                    inputs=inputs,
                    include_temperature=include_temperature,
                )
                Z0 = np.asarray(startup_diag.get("Z_tray", np.ones(N, dtype=float)), dtype=float).reshape((N,))
                direct_startup_ok = True
                _emit_progress(
                    "[Init] Vapor holdup initialization  direct tray thermo refresh complete  "
                    f"wall={float(startup_diag.get('startup_vapor_holdup_refresh_wall_sec', float('nan'))):.2f}s  "
                    f"batch={'yes' if bool(startup_diag.get('startup_vapor_holdup_refresh_batch_used', False)) else 'no'}"
                )
            except Exception as exc:
                _emit_progress(
                    "[Warn] Vapor holdup initialization  direct tray thermo refresh failed; "
                    f"falling back to one thermo/RHS pass ({exc})"
                )
                startup_diag = {}
                Z0 = None
        if not direct_startup_ok:
            _emit_progress("[Init] Vapor holdup initialization  solving startup tray Z factors from one thermo/RHS pass")
            # One thermo pass at t=0 to get Z_tray (if available)
            init_inputs = replace(inputs, equilibrium_relaxation=False, compute_thermo_diag=True)
            _dydt0, diag0 = column_rhs(0.0, y, col, layout, inputs=init_inputs)
            if isinstance(diag0, dict):
                startup_diag = dict(diag0)
            Z0 = np.asarray(diag0.get("Z_tray", np.ones(N, dtype=float)), dtype=float).reshape((N,))
        try:
            z_min = float(np.nanmin(Z0))
            z_max = float(np.nanmax(Z0))
            _emit_progress(f"[Init] Vapor holdup initialization  tray Z factors ready  range={z_min:.4g}..{z_max:.4g}")
        except Exception:
            _emit_progress("[Init] Vapor holdup initialization  tray Z factors ready")
    Z0 = np.where(~np.isfinite(Z0) | (Z0 <= 0.0), 1.0, Z0)

    T_F = _tray_temperature_F(col, layout, y, include_temperature)
    T_R = T_F + 459.67
    T_R = np.where(~np.isfinite(T_R) | (T_R <= 1e-6), 559.67, T_R)

    V = _vapor_volume_ft3_per_stage(inputs.volume_model, N)

    R = 10.7316  # (psia*ft3)/(lbmol*R)

    MV_target = np.zeros(N, dtype=float)
    for i in range(N):
        if i == 0:
            MV_target[i] = 0.0
            continue
        Pi = float(P_spec[i])
        if not np.isfinite(Pi) or Pi <= 0.0:
            MV_target[i] = np.nan
            continue
        MV_target[i] = Pi * V[i] / (Z0[i] * R * T_R[i])

    u = layout.unpack(y)

    # Apply to tray vapor component holdups using vapor composition from col.y0
    # (fallback to current y state if col.y0 is unavailable).
    try:
        yfrac = np.asarray(getattr(col, "y0"), dtype=float).reshape((N, col.n_components))
        for i in range(N):
            yfrac[i, :] = _normalize_comp(np.where(np.isfinite(yfrac[i, :]), yfrac[i, :], 0.0))
    except Exception:
        yfrac = np.asarray(u["y_tray"], dtype=float).reshape((N, col.n_components))
        yfrac = np.where(np.isfinite(yfrac), yfrac, 0.0)

    tray_V = np.asarray(u["tray_V"], dtype=float).reshape((N, col.n_components)).copy()
    if not bool(preserve_tray_vapor_holdup):
        for i in range(N):
            if i == 0:
                tray_V[i, :] = 0.0
                continue
            if not np.isfinite(MV_target[i]) or MV_target[i] < 0.0:
                continue
            tray_V[i, :] = float(MV_target[i]) * _normalize_comp(yfrac[i, :])

    sl = layout.slices()
    y_new = np.asarray(y, dtype=float).copy()
    y_new[sl["tray_V"]] = tray_V.ravel(order="C")
    if layout.include_energy and ("tray_EV_BTU" in sl):
        try:
            tray_EV_prev = np.asarray(u["tray_EV_BTU"], dtype=float).reshape((N,))
            MV_prev = np.sum(np.asarray(u["tray_V"], dtype=float).reshape((N, col.n_components)), axis=1)
            hV_prev = tray_EV_prev / np.maximum(MV_prev, float(layout.epsilon_lbmol))
            hV_prev = np.where(np.isfinite(hV_prev), hV_prev, T_F)

            MV_new = np.sum(np.asarray(tray_V, dtype=float), axis=1)
            tray_EV_new = hV_prev * MV_new
            tray_EV_new = np.where(np.isfinite(tray_EV_new), tray_EV_new, 0.0)
            tray_EV_new[MV_new <= float(layout.epsilon_lbmol)] = 0.0
            y_new[sl["tray_EV_BTU"]] = tray_EV_new
        except Exception:
            pass

    # Optional top-drum vapor initialization from top pressure specification.
    if layout.include_top and layout.include_vapor and ("top_V" in sl):
        _emit_progress("[Init] Vapor holdup initialization  seeding top-drum vapor holdup from startup pressure")
        top_vol = inputs.top_drum_vapor_volume_ft3
        top_total_vol = inputs.top_drum_total_volume_ft3
        top_extra_vap_vol = 0.0
        if inputs.top_drum_extra_vapor_volume_ft3 is not None:
            try:
                vextra = float(inputs.top_drum_extra_vapor_volume_ft3)
                if np.isfinite(vextra) and vextra > 0.0:
                    top_extra_vap_vol = float(vextra)
            except Exception:
                top_extra_vap_vol = 0.0
        if top_total_vol is not None and np.isfinite(float(top_total_vol)) and float(top_total_vol) > 0.0:
            rho_top = None
            if inputs.thermo_provider is not None and hasattr(inputs.thermo_provider, "liquid_density_lbmol_ft3"):
                try:
                    x_top = np.asarray(u["top_L"], dtype=float).reshape((col.n_components,))
                    x_top = _normalize_comp(np.where(np.isfinite(x_top), x_top, 0.0))
                    rho_try = float(
                        inputs.thermo_provider.liquid_density_lbmol_ft3(
                            float(T_F[0]),
                            float(P_spec[0]) if np.isfinite(float(P_spec[0])) else 200.0,
                            x_top,
                        )
                    )
                    if np.isfinite(rho_try) and rho_try > 1e-12:
                        rho_top = rho_try
                except Exception:
                    rho_top = None
            if rho_top is not None:
                try:
                    m_top = float(np.sum(np.asarray(u["top_L"], dtype=float).reshape((col.n_components,))))
                    liq_vol = max(m_top / float(rho_top), 0.0)
                    liq_vol = float(np.clip(liq_vol, 0.0, float(top_total_vol)))
                    top_vol = float(top_total_vol) - liq_vol + float(top_extra_vap_vol)
                    if top_vol < 1e-3:
                        top_vol = 1e-3
                except Exception:
                    top_vol = inputs.top_drum_vapor_volume_ft3
        if top_vol is None:
            try:
                top_vol = float(V[0])
            except Exception:
                top_vol = None
        if top_vol is not None and np.isfinite(float(top_vol)) and float(top_vol) > 0.0:
            p_top = float(P_spec[0]) if np.isfinite(float(P_spec[0])) and float(P_spec[0]) > 0.0 else np.nan
            if np.isfinite(p_top):
                # Keep top-drum initialization on the same ideal-gas Z basis
                # used by dynamic top-drum pressure diagnostics.
                z_top = 1.0
                mv_top = p_top * float(top_vol) / max(z_top * R * T_R[0], 1e-12)
                mv_top = max(float(mv_top), 0.0)
                top_y_src_idx = 1 if N > 1 else 0
                top_y = _normalize_comp(yfrac[top_y_src_idx, :])
                y_new[sl["top_V"]] = (mv_top * top_y).reshape((-1,))
    _emit_progress("[Init] Vapor holdup initialization complete")
    if return_diag:
        return y_new, startup_diag
    return y_new


def _initialize_thermo_consistent_state(
    *,
    col: ColumnSpec,
    layout: StateVectorLayout,
    y: np.ndarray,
    inputs: ColumnInputs,
    include_temperature: bool,
    max_iter: int = 2,
    relaxation: float = 1.0,
    preserve_tray_vapor_holdup: bool = False,
) -> Tuple[np.ndarray, Dict[str, Any]]:
    """
    Startup conditioning of tray phase compositions (and optional energy states)
    using thermo-equilibrium targets at tray T/P.

    This keeps stage holdup totals fixed while projecting x/y toward x_eq/y_eq,
    then reapplies pressure-consistent vapor-holdup initialization.
    """
    info: Dict[str, Any] = {
        "attempted": False,
        "success": False,
        "n_iter": 0,
        "max_dx": np.nan,
        "max_dy": np.nan,
        "eq_phase_change_init_lbmolps": np.nan,
        "eq_phase_change_final_lbmolps": np.nan,
    }
    if (not layout.include_vapor) or (inputs.thermo_provider is None):
        return np.asarray(y, dtype=float), info

    try:
        n_iter = max(1, int(max_iter))
    except Exception:
        n_iter = 1
    try:
        lam = float(relaxation)
    except Exception:
        lam = 1.0
    lam = float(np.clip(lam, 0.0, 1.0))
    if lam <= 0.0:
        return np.asarray(y, dtype=float), info

    N = int(col.n_stages)
    Nc = int(col.n_components)
    sl = layout.slices()
    y_work = np.asarray(y, dtype=float).copy()
    info["attempted"] = True

    dx_max = 0.0
    dy_max = 0.0
    eq_init = np.nan

    for it in range(n_iter):
        _emit_progress(
            f"[Init] Thermo startup conditioning  iteration {int(it + 1)}/{int(n_iter)}  solving tray equilibrium targets"
        )
        eval_inputs = replace(
            inputs,
            compute_thermo_diag=True,
            equilibrium_relaxation=True,
            trace_stage_thermo=True,
            thermo_stage_trace_label=f"startup_eq_iter_{int(it + 1)}",
        )
        try:
            _dydt, diag = column_rhs(0.0, y_work, col, layout, inputs=eval_inputs)
        except Exception:
            _emit_progress(
                f"[Init] Thermo startup conditioning  iteration {int(it + 1)}/{int(n_iter)} failed during thermo/RHS evaluation"
            )
            break
        _emit_progress(
            f"[Init] Thermo startup conditioning  iteration {int(it + 1)}/{int(n_iter)}  thermo/RHS evaluation complete"
        )

        x_eq_raw = diag.get("x_eq_tray", None)
        y_eq_raw = diag.get("y_eq_tray", None)
        if x_eq_raw is None or y_eq_raw is None:
            break
        try:
            x_eq = np.asarray(x_eq_raw, dtype=float).reshape((N, Nc))
            y_eq = np.asarray(y_eq_raw, dtype=float).reshape((N, Nc))
        except Exception:
            break

        if np.isnan(eq_init):
            try:
                eq_rate = np.asarray(diag.get("eq_phase_change_lbmolps_tray"), dtype=float).reshape((N,))
                if np.any(np.isfinite(eq_rate)):
                    eq_init = float(np.nanmax(np.abs(eq_rate)))
            except Exception:
                pass

        u = layout.unpack(y_work)
        tray_L = np.asarray(u["tray_L"], dtype=float).reshape((N, Nc))
        tray_V = np.asarray(u["tray_V"], dtype=float).reshape((N, Nc))
        ML_tot = np.sum(tray_L, axis=1).reshape((N,))
        MV_tot = np.sum(tray_V, axis=1).reshape((N,))
        x_old = np.asarray(u["x_tray"], dtype=float).reshape((N, Nc))
        y_old = np.asarray(u["y_tray"], dtype=float).reshape((N, Nc))

        x_new_frac = np.zeros((N, Nc), dtype=float)
        y_new_frac = np.zeros((N, Nc), dtype=float)
        for i in range(N):
            x_tgt = np.where(np.isfinite(x_eq[i, :]), x_eq[i, :], x_old[i, :])
            y_tgt = np.where(np.isfinite(y_eq[i, :]), y_eq[i, :], y_old[i, :])
            x_new_frac[i, :] = _normalize_comp((1.0 - lam) * x_old[i, :] + lam * x_tgt)
            y_new_frac[i, :] = _normalize_comp((1.0 - lam) * y_old[i, :] + lam * y_tgt)

        tray_L_new = ML_tot[:, None] * x_new_frac
        tray_V_new = MV_tot[:, None] * y_new_frac
        if N > 0:
            tray_V_new[0, :] = 0.0

        y_new = y_work.copy()
        y_new[sl["tray_L"]] = tray_L_new.ravel(order="C")
        y_new[sl["tray_V"]] = tray_V_new.ravel(order="C")

        # Keep boundary liquid holdups on their seeded product compositions.
        # The reflux drum and bottoms sump are separate inventories; forcing
        # them to equal neighboring tray compositions during startup destroys
        # workbook/ChemSep product seeds before the first dynamic step.
        # Vapor holdups remain aligned with neighboring tray targets.
        if layout.include_top and ("top_L" in sl):
            top_L_vec = np.asarray(u.get("top_L", np.zeros(Nc, dtype=float)), dtype=float).reshape((Nc,))
            y_new[sl["top_L"]] = np.where(np.isfinite(top_L_vec), top_L_vec, 0.0)
            if layout.include_vapor and ("top_V" in sl):
                top_V_vec = np.asarray(u.get("top_V", np.zeros(Nc, dtype=float)), dtype=float).reshape((Nc,))
                m_top_V = max(float(np.sum(top_V_vec)), 0.0)
                src = 1 if N > 1 else 0
                y_new[sl["top_V"]] = m_top_V * y_new_frac[src, :]
        if layout.include_bottom and ("bottom_L" in sl):
            bot_L_vec = np.asarray(u.get("bottom_L", np.zeros(Nc, dtype=float)), dtype=float).reshape((Nc,))
            y_new[sl["bottom_L"]] = np.where(np.isfinite(bot_L_vec), bot_L_vec, 0.0)
            if layout.include_vapor and ("bottom_V" in sl):
                bot_V_vec = np.asarray(u.get("bottom_V", np.zeros(Nc, dtype=float)), dtype=float).reshape((Nc,))
                m_bot_V = max(float(np.sum(bot_V_vec)), 0.0)
                y_new[sl["bottom_V"]] = m_bot_V * y_new_frac[-1, :]

        y_new = _clamp_nonnegative_holdups(y_new, layout)
        y_new = _clip_temperature_states_to_provider_bounds(y_new, layout, inputs.thermo_provider)

        z_seed = None
        if "Z_tray" in diag:
            try:
                z_seed = np.asarray(diag["Z_tray"], dtype=float).reshape((N,))
            except Exception:
                z_seed = None
        init_inputs = replace(inputs, Zfac_prev=z_seed) if z_seed is not None else inputs
        y_new = _initialize_vapor_holdup_from_spec_pressure(
            col=col,
            layout=layout,
            y=y_new,
            inputs=init_inputs,
            include_temperature=include_temperature,
            preserve_tray_vapor_holdup=bool(preserve_tray_vapor_holdup),
        )

        if bool(getattr(layout, "include_energy", False)):
            try:
                diag_energy = diag
                try:
                    _dydt_energy, diag_energy_try = column_rhs(0.0, y_new, col, layout, inputs=eval_inputs)
                    if isinstance(diag_energy_try, dict):
                        diag_energy = diag_energy_try
                except Exception:
                    diag_energy = diag

                u_energy = layout.unpack(y_new)
                if ("tray_EL_BTU" in sl) and ("HL_BTU_lbmol_tray" in diag_energy):
                    HL = np.asarray(diag_energy["HL_BTU_lbmol_tray"], dtype=float).reshape((N,))
                    HL = np.where(np.isfinite(HL), HL, 0.0)
                    ML_now = np.asarray(u_energy["ML_tot_tray"], dtype=float).reshape((N,))
                    y_new[sl["tray_EL_BTU"]] = ML_now * HL
                if ("tray_EV_BTU" in sl) and ("HV_BTU_lbmol_tray" in diag_energy):
                    HV = np.asarray(diag_energy["HV_BTU_lbmol_tray"], dtype=float).reshape((N,))
                    HV = np.where(np.isfinite(HV), HV, 0.0)
                    MV_now = np.asarray(u_energy["MV_tot_tray"], dtype=float).reshape((N,))
                    y_new[sl["tray_EV_BTU"]] = MV_now * HV
            except Exception:
                pass

        try:
            u_after = layout.unpack(y_new)
            x_after = np.asarray(u_after["x_tray"], dtype=float).reshape((N, Nc))
            y_after = np.asarray(u_after["y_tray"], dtype=float).reshape((N, Nc))
            dx = float(np.nanmax(np.abs(x_after - x_old)))
            dy = float(np.nanmax(np.abs(y_after - y_old)))
            if np.isfinite(dx):
                dx_max = max(dx_max, dx)
            if np.isfinite(dy):
                dy_max = max(dy_max, dy)
        except Exception:
            pass

        y_work = y_new
        info["n_iter"] = int(it + 1)
        try:
            _emit_progress(
                "[Init] Thermo startup conditioning  "
                f"iteration {int(it + 1)}/{int(n_iter)} complete  "
                f"max_dx={float(dx_max):.3g}  max_dy={float(dy_max):.3g}"
            )
        except Exception:
            pass

    eq_final = np.nan
    try:
        eval_inputs = replace(inputs, compute_thermo_diag=True, equilibrium_relaxation=True)
        _dydt_f, diag_f = column_rhs(0.0, y_work, col, layout, inputs=eval_inputs)
        if "eq_phase_change_lbmolps_tray" in diag_f:
            eq_rate_f = np.asarray(diag_f["eq_phase_change_lbmolps_tray"], dtype=float).reshape((N,))
            if np.any(np.isfinite(eq_rate_f)):
                eq_final = float(np.nanmax(np.abs(eq_rate_f)))
    except Exception:
        pass

    info["max_dx"] = float(dx_max) if np.isfinite(dx_max) else np.nan
    info["max_dy"] = float(dy_max) if np.isfinite(dy_max) else np.nan
    info["eq_phase_change_init_lbmolps"] = float(eq_init) if np.isfinite(eq_init) else np.nan
    info["eq_phase_change_final_lbmolps"] = float(eq_final) if np.isfinite(eq_final) else np.nan

    success = bool(info["n_iter"] > 0)
    if np.isfinite(eq_init) and np.isfinite(eq_final):
        success = success and (float(eq_final) <= float(eq_init) + 1e-12)
    info["success"] = bool(success)
    return y_work, info


def _initialize_top_drum_dynamic_steady(
    *,
    col: ColumnSpec,
    layout: StateVectorLayout,
    y: np.ndarray,
    inputs: ColumnInputs,
    max_iter: int = 6,
    tol_lbmolps: float = 1e-6,
    wall_limit_sec: Optional[float] = None,
) -> Tuple[np.ndarray, Dict[str, Any]]:
    """
    Best-effort startup trim for top-drum holdups.

    Adjust top liquid/vapor total holdups at t=0 (holding compositions fixed)
    to reduce top-drum accumulation residuals computed by RHS:
      d(top_L_total)/dt and d(top_V_total)/dt.
    """
    info: Dict[str, Any] = {
        "attempted": False,
        "success": False,
        "n_iter": 0,
        "pressure_coupled": False,
        "hit_wall_limit": False,
        "d_top_L_init_lbmolps": np.nan,
        "d_top_V_init_lbmolps": np.nan,
        "d_top_L_final_lbmolps": np.nan,
        "d_top_V_final_lbmolps": np.nan,
    }

    sl = layout.slices()
    if (not layout.include_top) or (not layout.include_vapor):
        return np.asarray(y, dtype=float), info
    if ("top_L" not in sl) or ("top_V" not in sl):
        return np.asarray(y, dtype=float), info

    y_work = np.asarray(y, dtype=float).copy()
    try:
        u0 = layout.unpack(y_work)
    except Exception:
        return y_work, info

    top_L0 = np.asarray(u0.get("top_L", []), dtype=float).reshape((-1,))
    top_V0 = np.asarray(u0.get("top_V", []), dtype=float).reshape((-1,))
    Nc = int(getattr(col, "n_components", 0))
    Ns = int(getattr(col, "n_stages", 0))
    if top_L0.size != Nc or top_V0.size != Nc or Nc <= 0:
        return y_work, info

    x_top = _normalize_comp(np.where(np.isfinite(top_L0), top_L0, 0.0))
    if np.sum(np.where(np.isfinite(top_V0), top_V0, 0.0)) > float(layout.epsilon_lbmol):
        y_top = _normalize_comp(np.where(np.isfinite(top_V0), top_V0, 0.0))
    else:
        y_tray = np.asarray(u0.get("y_tray", []), dtype=float)
        if y_tray.size == (Ns * Nc):
            y_tray = y_tray.reshape((Ns, Nc))
            src = 1 if Ns > 1 else 0
            y_top = _normalize_comp(np.where(np.isfinite(y_tray[src, :]), y_tray[src, :], 0.0))
        else:
            y_top = x_top.copy()

    p_target = None
    try:
        p0 = float(np.asarray(getattr(col, "P_psia"), dtype=float).reshape((-1,))[0])
        if np.isfinite(p0) and p0 > 0.0:
            p_target = float(p0)
    except Exception:
        p_target = None

    try:
        t_top_f = float(np.asarray(u0.get("tray_T_f"), dtype=float).reshape((Ns,))[0])
    except Exception:
        try:
            t_top_f = float(np.asarray(getattr(col, "T_f"), dtype=float).reshape((Ns,))[0])
        except Exception:
            t_top_f = 100.0
    if (not np.isfinite(t_top_f)):
        t_top_f = 100.0

    rho_top = None
    if inputs.rhoL_tray_lbmol_ft3 is not None:
        try:
            rho_try = float(np.asarray(inputs.rhoL_tray_lbmol_ft3, dtype=float).reshape((Ns,))[0])
            if np.isfinite(rho_try) and rho_try > 1e-12:
                rho_top = float(rho_try)
        except Exception:
            rho_top = None
    if rho_top is None and inputs.thermo_provider is not None and hasattr(inputs.thermo_provider, "liquid_density_lbmol_ft3"):
        try:
            p_for_rho = float(p_target) if p_target is not None else 200.0
            rho_try = float(inputs.thermo_provider.liquid_density_lbmol_ft3(float(t_top_f), p_for_rho, x_top))
            if np.isfinite(rho_try) and rho_try > 1e-12:
                rho_top = float(rho_try)
        except Exception:
            rho_top = None

    top_total_vol = None
    if inputs.top_drum_total_volume_ft3 is not None:
        try:
            vtot = float(inputs.top_drum_total_volume_ft3)
            if np.isfinite(vtot) and vtot > 0.0:
                top_total_vol = float(vtot)
        except Exception:
            top_total_vol = None

    top_vap_fixed = None
    if inputs.top_drum_vapor_volume_ft3 is not None:
        try:
            vv = float(inputs.top_drum_vapor_volume_ft3)
            if np.isfinite(vv) and vv > 0.0:
                top_vap_fixed = float(vv)
        except Exception:
            top_vap_fixed = None
    top_extra_vap_vol = 0.0
    if inputs.top_drum_extra_vapor_volume_ft3 is not None:
        try:
            vv = float(inputs.top_drum_extra_vapor_volume_ft3)
            if np.isfinite(vv) and vv > 0.0:
                top_extra_vap_vol = float(vv)
        except Exception:
            top_extra_vap_vol = 0.0
    if top_vap_fixed is None:
        try:
            vv = _vapor_volume_ft3_per_stage(inputs.volume_model, Ns)
            top_vap_fixed = float(vv[0])
        except Exception:
            top_vap_fixed = None

    def _top_vapor_volume_for_mL(mL: float) -> Optional[float]:
        if top_total_vol is not None and rho_top is not None:
            liq_vol = float(np.clip(float(mL) / max(float(rho_top), 1e-12), 0.0, float(top_total_vol)))
            return max(float(top_total_vol) - liq_vol + float(top_extra_vap_vol), 1e-3)
        if top_vap_fixed is not None and np.isfinite(top_vap_fixed) and top_vap_fixed > 0.0:
            return max(float(top_vap_fixed) + float(top_extra_vap_vol), 1e-3)
        return None

    def _mV_anchor_from_mL(mL: float) -> Optional[float]:
        if p_target is None:
            return None
        vap_vol = _top_vapor_volume_for_mL(mL)
        if vap_vol is None:
            return None
        T_R = float(t_top_f) + 459.67
        if (not np.isfinite(T_R)) or T_R <= 0.0:
            return None
        R = 10.7316
        mv = float(p_target) * float(vap_vol) / max(float(R) * float(T_R), 1e-12)
        if (not np.isfinite(mv)) or mv < 0.0:
            return None
        if inputs.thermo_provider is None:
            return max(float(mv), 0.0)

        y_anchor = np.asarray(y_top, dtype=float).reshape((Nc,))
        y_anchor = _normalize_comp(np.where(np.isfinite(y_anchor), y_anchor, 0.0))
        for _ in range(6):
            top_v = max(float(mv), 0.0) * y_anchor
            try:
                p_res = _compute_top_drum_pressure_psia(
                    top_V=top_v,
                    top_T_F=float(t_top_f),
                    Z_top=1.0,
                    top_vapor_volume_ft3=float(vap_vol),
                    thermo_provider=inputs.thermo_provider,
                    y_top=y_anchor,
                    P_seed_psia=float(p_target),
                    return_details=True,
                )
            except Exception:
                p_res = None
            if isinstance(p_res, tuple):
                p_eval = p_res[0]
            else:
                p_eval = p_res
            if (p_eval is None) or (not np.isfinite(float(p_eval))) or float(p_eval) <= 0.0:
                break
            ratio = float(p_target) / max(float(p_eval), 1.0e-12)
            if abs(ratio - 1.0) <= 1.0e-6:
                return max(float(mv), 0.0)
            mv_new = float(mv) * float(np.clip(ratio, 0.25, 4.0))
            if (not np.isfinite(mv_new)) or mv_new < 0.0:
                break
            mv = float(mv_new)
        return max(float(mv), 0.0)

    pressure_coupled = _mV_anchor_from_mL(float(np.sum(top_L0))) is not None
    info["pressure_coupled"] = bool(pressure_coupled)
    print(
        "[Init] Top-drum startup steadying  "
        f"mode={'pressure-coupled' if pressure_coupled else 'two-holdup'}  "
        f"max_iter={int(max_iter)}"
    )

    def _pack_top_totals(base_y: np.ndarray, mL: float, mV: float) -> np.ndarray:
        y_new = np.asarray(base_y, dtype=float).copy()
        mL_eff = max(float(mL), 0.0)
        mV_eff = max(float(mV), 0.0)
        if pressure_coupled:
            mv_anchor = _mV_anchor_from_mL(mL_eff)
            if mv_anchor is not None:
                mV_eff = float(mv_anchor)
        y_new[sl["top_L"]] = mL_eff * x_top
        y_new[sl["top_V"]] = mV_eff * y_top
        return y_new

    def _residual(mL: float, mV: float) -> Tuple[Optional[np.ndarray], np.ndarray]:
        y_try = _pack_top_totals(y_work, mL, mV)
        try:
            dydt, _diag = column_rhs(0.0, y_try, col, layout, inputs=inputs)
            dL = float(np.sum(np.asarray(dydt[sl["top_L"]], dtype=float).reshape((-1,))))
            dV = float(np.sum(np.asarray(dydt[sl["top_V"]], dtype=float).reshape((-1,))))
            if np.isfinite(dL) and np.isfinite(dV):
                return np.array([dL, dV], dtype=float), y_try
        except Exception:
            pass
        return None, y_try

    mL = max(float(np.sum(top_L0)), 0.0)
    mV = max(float(np.sum(top_V0)), 0.0)
    if pressure_coupled:
        mv_anchor0 = _mV_anchor_from_mL(mL)
        if mv_anchor0 is not None:
            mV = float(mv_anchor0)

    f, y_best = _residual(mL, mV)
    if f is None:
        return y_work, info

    info["attempted"] = True
    info["d_top_L_init_lbmolps"] = float(f[0])
    info["d_top_V_init_lbmolps"] = float(f[1])
    best_norm = float(np.linalg.norm(f))
    f_best = f.copy()
    t_start_wall = time.perf_counter()

    var_idx = [0] if pressure_coupled else [0, 1]
    for it in range(int(max_iter)):
        info["n_iter"] = int(it + 1)
        print(f"[Init] Top-drum startup steadying  iteration {int(it + 1)}/{int(max_iter)}")
        if wall_limit_sec is not None:
            try:
                if float(time.perf_counter() - t_start_wall) >= float(wall_limit_sec):
                    info["hit_wall_limit"] = True
                    print("[Init] Top-drum startup steadying  wall-time cap reached")
                    break
            except Exception:
                pass
        if np.max(np.abs(f_best)) <= float(tol_lbmolps):
            info["success"] = True
            print("[Init] Top-drum startup steadying  converged within tolerance")
            break

        J = np.zeros((2, len(var_idx)), dtype=float)
        valid = np.zeros(len(var_idx), dtype=bool)
        for jj, var in enumerate(var_idx):
            base = mL if var == 0 else mV
            dvar = max(0.05 * max(abs(base), 1.0), 1.0e-3)
            mL_p = mL + (dvar if var == 0 else 0.0)
            mV_p = mV + (dvar if var == 1 else 0.0)
            f_p, _ = _residual(mL_p, mV_p)
            if f_p is None:
                continue
            J[:, jj] = (f_p - f_best) / dvar
            valid[jj] = True

        if not np.any(valid):
            break
        J_use = J[:, valid]
        try:
            delta_use, *_ = np.linalg.lstsq(J_use, -f_best, rcond=None)
        except Exception:
            break

        delta = np.zeros(2, dtype=float)
        for jj, var in enumerate(var_idx):
            if valid[jj]:
                delta[var] = float(np.asarray(delta_use, dtype=float).reshape((-1,))[int(np.sum(valid[:jj]))])

        max_dL = max(0.5 * max(mL, 1.0), 1.0)
        max_dV = max(0.5 * max(mV, 1.0), 1.0)
        delta[0] = float(np.clip(delta[0], -max_dL, max_dL))
        delta[1] = float(np.clip(delta[1], -max_dV, max_dV))

        improved = False
        for fac in (1.0, 0.5, 0.25, 0.1):
            mL_try = max(mL + fac * float(delta[0]), 0.0)
            mV_try = max(mV + fac * float(delta[1]), 0.0)
            f_try, y_try = _residual(mL_try, mV_try)
            if f_try is None:
                continue
            n_try = float(np.linalg.norm(f_try))
            if n_try + 1e-12 < best_norm:
                mL = float(mL_try)
                mV = float(np.sum(np.asarray(y_try[sl["top_V"]], dtype=float).reshape((-1,))))
                f_best = f_try.copy()
                y_best = y_try.copy()
                best_norm = n_try
                improved = True
                break
        if not improved:
            print("[Init] Top-drum startup steadying  no improving Newton step found")
            break
        try:
            print(
                "[Init] Top-drum startup steadying  "
                f"iteration {int(it + 1)}/{int(max_iter)} residuals  "
                f"dL={float(f_best[0]):+.4g} lbmol/s  dV={float(f_best[1]):+.4g} lbmol/s"
            )
        except Exception:
            pass

    info["d_top_L_final_lbmolps"] = float(f_best[0])
    info["d_top_V_final_lbmolps"] = float(f_best[1])
    if np.max(np.abs(f_best)) <= float(tol_lbmolps):
        info["success"] = True
    return y_best, info


def _initialize_restart_reentry_settling(
    *,
    col: ColumnSpec,
    layout: StateVectorLayout,
    y: np.ndarray,
    inputs: ColumnInputs,
    include_temperature: bool,
    thermo_max_iter: int = 1,
    thermo_relaxation: float = 1.0,
    top_drum_max_iter: int = 2,
    top_drum_tol_lbmolps: float = 1.0e-4,
    top_drum_wall_limit_sec: Optional[float] = 10.0,
    preserve_tray_vapor_holdup: bool = True,
) -> Tuple[np.ndarray, Dict[str, Any]]:
    """
    Apply a short hidden restart re-entry conditioning pass so an explicit
    runtime restart resumes closer to the pre-stop trajectory before normal
    logging begins.
    """
    info: Dict[str, Any] = {
        "attempted": False,
        "success": False,
        "thermo": {
            "attempted": False,
            "success": False,
            "n_iter": 0,
            "max_dx": np.nan,
            "max_dy": np.nan,
            "eq_phase_change_init_lbmolps": np.nan,
            "eq_phase_change_final_lbmolps": np.nan,
        },
        "top_drum": {
            "attempted": False,
            "success": False,
            "n_iter": 0,
            "pressure_coupled": False,
            "hit_wall_limit": False,
            "d_top_L_init_lbmolps": np.nan,
            "d_top_V_init_lbmolps": np.nan,
            "d_top_L_final_lbmolps": np.nan,
            "d_top_V_final_lbmolps": np.nan,
        },
    }
    y_work = np.asarray(y, dtype=float).copy()

    thermo_iters = max(0, int(thermo_max_iter))
    if thermo_iters > 0:
        y_work, thermo_info = _initialize_thermo_consistent_state(
            col=col,
            layout=layout,
            y=y_work,
            inputs=inputs,
            include_temperature=include_temperature,
            max_iter=thermo_iters,
            relaxation=float(thermo_relaxation),
            preserve_tray_vapor_holdup=bool(preserve_tray_vapor_holdup),
        )
        info["thermo"] = thermo_info
        info["attempted"] = bool(info["attempted"] or thermo_info.get("attempted", False))

    y_work, top_drum_info = _initialize_top_drum_dynamic_steady(
        col=col,
        layout=layout,
        y=y_work,
        inputs=inputs,
        max_iter=max(1, int(top_drum_max_iter)),
        tol_lbmolps=float(top_drum_tol_lbmolps),
        wall_limit_sec=top_drum_wall_limit_sec,
    )
    info["top_drum"] = top_drum_info
    info["attempted"] = bool(info["attempted"] or top_drum_info.get("attempted", False))
    info["success"] = bool(
        (not info["attempted"])
        or (
            bool(info["thermo"].get("success", False) or (not info["thermo"].get("attempted", False)))
            and bool(info["top_drum"].get("success", False) or (not info["top_drum"].get("attempted", False)))
        )
    )
    return y_work, info


def _has_explicit_top_boundary_restart_state(col: ColumnSpec) -> bool:
    for key in ("top_L0_lbmol", "top_V0_lbmol"):
        if not hasattr(col, key):
            continue
        try:
            arr = np.asarray(getattr(col, key), dtype=float).reshape((-1,))
        except Exception:
            continue
        if arr.size > 0 and np.any(np.isfinite(arr) & (arr > 0.0)):
            return True
    return False


def _has_explicit_energy_restart_state(col: ColumnSpec) -> bool:
    for key in ("tray_EL0_BTU", "tray_EV0_BTU"):
        if not hasattr(col, key):
            continue
        try:
            arr = np.asarray(getattr(col, key), dtype=float).reshape((-1,))
        except Exception:
            continue
        if arr.size > 0 and np.any(np.isfinite(arr)):
            return True
    return False


def _has_explicit_runtime_restart_state(col: ColumnSpec) -> bool:
    return bool(_has_explicit_top_boundary_restart_state(col) or _has_explicit_energy_restart_state(col))


def _get_controller_restart_value(col: ColumnSpec, key: str) -> Optional[float]:
    try:
        raw = getattr(col, "controller_state", None)
    except Exception:
        raw = None
    if not isinstance(raw, dict):
        return None
    try:
        v = float(raw.get(key))
    except Exception:
        return None
    return float(v) if np.isfinite(v) else None


def _refresh_tray_bubble_targets_F(
    *,
    col: ColumnSpec,
    layout: StateVectorLayout,
    y: np.ndarray,
    thermo_provider: Any,
    P_tray_psia: np.ndarray,
) -> Optional[np.ndarray]:
    if thermo_provider is None or (not bool(getattr(layout, "include_temperature", False))):
        return None
    try:
        u = layout.unpack(y)
        tray_T = np.asarray(u["tray_T_f"], dtype=float).reshape((col.n_stages,))
        x_tray = np.asarray(u["x_tray"], dtype=float).reshape((col.n_stages, col.n_components))
        P_arr = np.asarray(P_tray_psia, dtype=float).reshape((col.n_stages,))
    except Exception:
        return None
    targets = np.full((col.n_stages,), np.nan, dtype=float)
    for i in range(col.n_stages):
        try:
            if not np.isfinite(float(P_arr[i])):
                continue
            T_eq, _ = _bubble_point_T_F(
                thermo_provider=thermo_provider,
                P_psia=float(P_arr[i]),
                x=x_tray[i, :],
                T_guess_F=float(tray_T[i]),
            )
            if np.isfinite(float(T_eq)):
                targets[i] = float(T_eq)
        except Exception:
            continue
    return targets


def _update_tray_temp_pressure_slope_F_per_psi(
    *,
    prev_slope_F_per_psi: Optional[np.ndarray],
    prev_T_F: Optional[np.ndarray],
    curr_T_F: Optional[np.ndarray],
    prev_P_psia: Optional[np.ndarray],
    curr_P_psia: Optional[np.ndarray],
    default_slope_F_per_psi: float = 1.5,
    dp_min_psia: float = 0.05,
    slope_clip_min_F_per_psi: float = -2.0,
    slope_clip_max_F_per_psi: float = 2.0,
    blend_new: float = 0.5,
) -> Optional[np.ndarray]:
    try:
        curr_T = np.asarray(curr_T_F, dtype=float).reshape((-1,))
        curr_P = np.asarray(curr_P_psia, dtype=float).reshape((-1,))
    except Exception:
        return None
    n = int(curr_T.size)
    if int(curr_P.size) != n or n <= 0:
        return None
    slope = np.full(n, float(default_slope_F_per_psi), dtype=float)
    if prev_slope_F_per_psi is not None:
        try:
            prev_slope = np.asarray(prev_slope_F_per_psi, dtype=float).reshape((n,))
            valid_prev = np.isfinite(prev_slope)
            slope[valid_prev] = prev_slope[valid_prev]
        except Exception:
            pass
    if prev_T_F is None or prev_P_psia is None:
        return slope
    try:
        prev_T = np.asarray(prev_T_F, dtype=float).reshape((n,))
        prev_P = np.asarray(prev_P_psia, dtype=float).reshape((n,))
    except Exception:
        return slope
    try:
        blend = float(blend_new)
    except Exception:
        blend = 0.5
    if (not np.isfinite(blend)) or blend < 0.0:
        blend = 0.5
    if blend > 1.0:
        blend = 1.0
    valid = (
        np.isfinite(prev_T)
        & np.isfinite(curr_T)
        & np.isfinite(prev_P)
        & np.isfinite(curr_P)
        & (np.abs(curr_P - prev_P) >= float(dp_min_psia))
    )
    if np.any(valid):
        secant = (curr_T[valid] - prev_T[valid]) / (curr_P[valid] - prev_P[valid])
        secant = np.clip(
            secant,
            float(slope_clip_min_F_per_psi),
            float(slope_clip_max_F_per_psi),
        )
        slope[valid] = (1.0 - float(blend)) * slope[valid] + float(blend) * secant
    return slope


def _initialize_hydraulic_energy_consistent_state(
    *,
    col: ColumnSpec,
    layout: StateVectorLayout,
    y: np.ndarray,
    inputs: ColumnInputs,
    include_temperature: bool,
    max_iter: int = 6,
    pseudo_dt_sec: float = 0.5,
    mass_tol_lbmolph: Optional[float] = 5.0,
    energy_tol_btups: Optional[float] = 1000.0,
    dae_max_iter: int = 5,
    dae_p_tol_psia: Optional[float] = 0.05,
    dae_v_tol_lbmolph: Optional[float] = 25.0,
    dae_jac_rel_step: float = 1.0e-6,
    dae_line_search_max: int = 4,
) -> Tuple[np.ndarray, Dict[str, Any]]:
    """
    Best-effort startup consistency relaxation for hydraulic+energy runs.

    The ChemSep seed can match product specs while still being slightly off the
    hydraulic-energy differential/algebraic manifold. This initializer uses the
    existing pilot algebraic solve at t=0, then takes a few bounded pseudo-time
    relaxation steps only when they improve a normalized startup residual score.
    """
    info: Dict[str, Any] = {
        "attempted": False,
        "success": False,
        "n_iter": 0,
        "objective_init": np.nan,
        "objective_final": np.nan,
        "mass_resid_init_lbmolph": np.nan,
        "mass_resid_final_lbmolph": np.nan,
        "energy_resid_init_btups": np.nan,
        "energy_resid_final_btups": np.nan,
        "alg_p_init_psia": np.nan,
        "alg_p_final_psia": np.nan,
        "alg_v_init_lbmolph": np.nan,
        "alg_v_final_lbmolph": np.nan,
    }

    if str(getattr(inputs, "pressure_model", "")).strip().lower() != "hydraulic":
        return np.asarray(y, dtype=float), info
    if str(getattr(inputs, "vapor_flow_model", "")).strip().lower() != "energy":
        return np.asarray(y, dtype=float), info
    if not bool(getattr(layout, "include_vapor", False)):
        return np.asarray(y, dtype=float), info

    try:
        n_iter = max(1, int(max_iter))
    except Exception:
        n_iter = 1
    try:
        dt_relax = float(pseudo_dt_sec)
    except Exception:
        dt_relax = 0.5
    if (not np.isfinite(dt_relax)) or dt_relax <= 0.0:
        dt_relax = 0.5

    def _project_state(y_in: np.ndarray) -> np.ndarray:
        y_proj = _clamp_nonnegative_holdups(np.asarray(y_in, dtype=float), layout)
        return _clip_temperature_states_to_provider_bounds(y_proj, layout, inputs.thermo_provider)

    def _metric(diag: Dict[str, Any], dydt: np.ndarray) -> Tuple[float, float, float, float, float]:
        mass_resid_lbmolph = np.nan
        if "mass_balance_resid_lbmolps_tray" in diag:
            try:
                mr = np.asarray(diag["mass_balance_resid_lbmolps_tray"], dtype=float).reshape((-1,))
                mr = mr[np.isfinite(mr)]
                if mr.size > 0:
                    mass_resid_lbmolph = float(np.max(np.abs(mr))) * 3600.0
            except Exception:
                mass_resid_lbmolph = np.nan
        if not np.isfinite(mass_resid_lbmolph):
            try:
                mass_resid_lbmolph = abs(float(_total_inventory_rate_lbmolps(layout, dydt))) * 3600.0
            except Exception:
                mass_resid_lbmolph = np.nan

        energy_resid_btups = abs(_mapping_scalar(diag, "resid_energy_btups"))
        alg_p_psia = abs(_mapping_scalar(diag, "dae_pilot_alg_p_inf_psia"))
        alg_v_lbmolph = abs(_mapping_scalar(diag, "dae_pilot_alg_v_inf_lbmolph"))

        terms: List[float] = []
        if np.isfinite(mass_resid_lbmolph):
            m_tol = float(mass_tol_lbmolph) if mass_tol_lbmolph is not None else 5.0
            m_tol = m_tol if np.isfinite(m_tol) and m_tol > 0.0 else 5.0
            terms.append(float(mass_resid_lbmolph) / float(m_tol))
        if np.isfinite(energy_resid_btups):
            e_tol = float(energy_tol_btups) if energy_tol_btups is not None else 1000.0
            e_tol = e_tol if np.isfinite(e_tol) and e_tol > 0.0 else 1000.0
            terms.append(float(energy_resid_btups) / float(e_tol))
        if np.isfinite(alg_p_psia):
            p_tol = float(dae_p_tol_psia) if dae_p_tol_psia is not None else 0.05
            p_tol = p_tol if np.isfinite(p_tol) and p_tol > 0.0 else 0.05
            terms.append(float(alg_p_psia) / float(p_tol))
        if np.isfinite(alg_v_lbmolph):
            v_tol = float(dae_v_tol_lbmolph) if dae_v_tol_lbmolph is not None else 25.0
            v_tol = v_tol if np.isfinite(v_tol) and v_tol > 0.0 else 25.0
            terms.append(float(alg_v_lbmolph) / float(v_tol))
        objective = float(np.max(np.asarray(terms, dtype=float))) if terms else float("inf")
        return objective, mass_resid_lbmolph, energy_resid_btups, alg_p_psia, alg_v_lbmolph

    y_work = _project_state(y)
    inputs_work = inputs
    info["attempted"] = True
    print("[Init] Hydraulic-energy startup consistency  initial pilot algebraic solve")

    try:
        dydt_best, diag_best = _solve_dae_pilot_algebraic(
            t_s=0.0,
            y=y_work,
            col=col,
            layout=layout,
            inputs=inputs_work,
            max_iter=int(dae_max_iter),
            p_tol_psia=dae_p_tol_psia,
            v_tol_lbmolph=dae_v_tol_lbmolph,
            jac_rel_step=float(dae_jac_rel_step),
            line_search_max=int(dae_line_search_max),
        )
    except Exception:
        return y_work, info

    obj_best, mass_best, energy_best, alg_p_best, alg_v_best = _metric(diag_best, dydt_best)
    info["objective_init"] = float(obj_best)
    info["mass_resid_init_lbmolph"] = float(mass_best) if np.isfinite(mass_best) else np.nan
    info["energy_resid_init_btups"] = float(energy_best) if np.isfinite(energy_best) else np.nan
    info["alg_p_init_psia"] = float(alg_p_best) if np.isfinite(alg_p_best) else np.nan
    info["alg_v_init_lbmolph"] = float(alg_v_best) if np.isfinite(alg_v_best) else np.nan
    try:
        print(
            "[Init] Hydraulic-energy startup consistency  "
            f"initial objective={float(obj_best):.3g}  "
            f"mass={float(mass_best):.3g} lbmol/h  energy={float(energy_best):.3g} BTU/s"
        )
    except Exception:
        pass

    for it in range(n_iter):
        info["n_iter"] = int(it + 1)
        print(f"[Init] Hydraulic-energy startup consistency  relaxation iteration {int(it + 1)}/{int(n_iter)}")
        dydt_curr = np.asarray(dydt_best, dtype=float).reshape((-1,))
        improved = False
        for fac in (1.0, 0.5, 0.25, 0.1):
            y_try = _project_state(y_work + float(fac) * float(dt_relax) * dydt_curr)
            try:
                dydt_try, diag_try = _solve_dae_pilot_algebraic(
                    t_s=0.0,
                    y=y_try,
                    col=col,
                    layout=layout,
                    inputs=inputs_work,
                    max_iter=int(dae_max_iter),
                    p_tol_psia=dae_p_tol_psia,
                    v_tol_lbmolph=dae_v_tol_lbmolph,
                    jac_rel_step=float(dae_jac_rel_step),
                    line_search_max=int(dae_line_search_max),
                )
            except Exception:
                continue

            obj_try, mass_try, energy_try, alg_p_try, alg_v_try = _metric(diag_try, dydt_try)
            if np.isfinite(obj_try) and ((not np.isfinite(obj_best)) or obj_try + 1.0e-12 < obj_best):
                y_work = y_try
                dydt_best = np.asarray(dydt_try, dtype=float).reshape((-1,))
                diag_best = dict(diag_try)
                obj_best = float(obj_try)
                mass_best = float(mass_try) if np.isfinite(mass_try) else np.nan
                energy_best = float(energy_try) if np.isfinite(energy_try) else np.nan
                alg_p_best = float(alg_p_try) if np.isfinite(alg_p_try) else np.nan
                alg_v_best = float(alg_v_try) if np.isfinite(alg_v_try) else np.nan
                try:
                    p_prev = np.asarray(diag_best.get("P_psia_hyd"), dtype=float).reshape((col.n_stages,))
                    v_prev = np.asarray(diag_best.get("V_out_lbmolph"), dtype=float).reshape((col.n_stages,))
                    inputs_work = replace(inputs_work, P_tray_prev=p_prev, V_out_prev_lbmolph=v_prev)
                except Exception:
                    pass
                improved = True
                try:
                    print(
                        "[Init] Hydraulic-energy startup consistency  "
                        f"iteration {int(it + 1)}/{int(n_iter)} improved  "
                        f"objective={float(obj_best):.3g}  "
                        f"mass={float(mass_best):.3g} lbmol/h  energy={float(energy_best):.3g} BTU/s"
                    )
                except Exception:
                    pass
                break
        if (not improved) or (np.isfinite(obj_best) and obj_best <= 1.0):
            if not improved:
                print("[Init] Hydraulic-energy startup consistency  no improving relaxation step found")
            elif np.isfinite(obj_best) and obj_best <= 1.0:
                print("[Init] Hydraulic-energy startup consistency  objective is within startup target")
            break

    info["objective_final"] = float(obj_best) if np.isfinite(obj_best) else np.nan
    info["mass_resid_final_lbmolph"] = float(mass_best) if np.isfinite(mass_best) else np.nan
    info["energy_resid_final_btups"] = float(energy_best) if np.isfinite(energy_best) else np.nan
    info["alg_p_final_psia"] = float(alg_p_best) if np.isfinite(alg_p_best) else np.nan
    info["alg_v_final_lbmolph"] = float(alg_v_best) if np.isfinite(alg_v_best) else np.nan
    info["success"] = bool(
        np.isfinite(info["objective_final"])
        and np.isfinite(info["objective_init"])
        and float(info["objective_final"]) + 1.0e-12 < float(info["objective_init"])
    )
    return y_work, info


# -------------------------
# Logging row writers
# -------------------------


def _resolve_logging_streams(case: CaseData, col: ColumnSpec) -> Tuple[StreamTag, StreamTag, StreamTag]:
    """
    Resolve feed, distillate, bottoms for logging purposes.

    - Feed: prefers named stream "Feed" (or alias); if no stage in stream, tries specs_raw.
    - Distillate: prefers named stream "Distillate"; default stage=1 if missing.
    - Bottoms: prefers named stream "Bottoms"; default stage=N if missing.
    """
    N = col.n_stages

    feed = _lookup_named_stream(col=col, case=case, aliases=("Feed",))
    dist = _lookup_named_stream(col=col, case=case, aliases=("Distillate", "Dist", "Overhead", "D"))
    bots = _lookup_named_stream(col=col, case=case, aliases=("Bottoms", "Bottom", "B"))

    # Feed stage fallback from specs_raw
    if feed.stage_1based is None:
        feed_stage = _lookup_spec_int(col, ("Feed Stage", "Feed stage", "Feed Stage (1-based)"))
        feed = StreamTag(name=feed.name, flow_lbmolph=feed.flow_lbmolph, stage_1based=feed_stage)

    # Distillate/bottoms defaults
    if dist.stage_1based is None:
        dist = StreamTag(name=dist.name, flow_lbmolph=dist.flow_lbmolph, stage_1based=1)
    if bots.stage_1based is None:
        bots = StreamTag(name=bots.name, flow_lbmolph=bots.flow_lbmolph, stage_1based=N)

    # Clamp stages into [1, N]
    def clamp(tag: StreamTag) -> StreamTag:
        st = tag.stage_1based
        if st is None:
            return tag
        if 1 <= int(st) <= N:
            return StreamTag(name=tag.name, flow_lbmolph=tag.flow_lbmolph, stage_1based=int(st))
        return StreamTag(name=tag.name, flow_lbmolph=tag.flow_lbmolph, stage_1based=None)

    return clamp(feed), clamp(dist), clamp(bots)


def _infer_condenser_duty_bias_btu_per_h(col: ColumnSpec) -> float:
    specs = getattr(col, "specs_raw", None) or {}
    q = _spec_float(specs, "Condenser Duty (Btu/h)")
    if q is None:
        try:
            duties = getattr(col, "duties", None)
            if duties is not None and getattr(duties, "q_cond_btu_per_h", None) is not None:
                q = float(getattr(duties, "q_cond_btu_per_h"))
        except Exception:
            q = None
    if q is None or (not np.isfinite(float(q))):
        q = -5.0e7
    q = float(q)
    if q > 0.0:
        q = -abs(q)
    return q


def _build_pressure_controller(
    *,
    col: ColumnSpec,
    cfg: RunnerConfig,
) -> Tuple[bool, Optional[PIController], Optional[float], str, Optional[str]]:
    specs = getattr(col, "specs_raw", None) or {}
    enabled = bool(cfg.enable_pressure_control)
    if not enabled:
        b = _as_bool(_spec_get(specs, "Enable Pressure Control", "Pressure Control Enabled"))
        enabled = bool(b) if b is not None else False
    if not enabled:
        return False, None, None, "off", None

    sp = cfg.top_pressure_sp_psia
    if sp is None:
        sp = _spec_float(specs, "Top Pressure SP (psia)", "Condenser Pressure SP (psia)")
    if sp is None:
        try:
            p_ctrl_idx = 0
            sp = float(np.asarray(getattr(col, "P_psia"), dtype=float).reshape((-1,))[p_ctrl_idx])
        except Exception:
            sp = None
    if sp is None or (not np.isfinite(float(sp))) or float(sp) <= 0.0:
        return False, None, None, "off", None

    pressure_mode_note: Optional[str] = None
    mv_mode = str(cfg.pressure_control_mv or "auto").strip().lower().replace("_", "-")
    if mv_mode in ("", "auto"):
        mv_mode = str(_spec_get(specs, "Pressure Control MV", "Pressure Controller MV", "Pressure MV") or "auto").strip().lower().replace("_", "-")
    if mv_mode in ("", "auto"):
        if str(cfg.condenser_duty_mode or "").strip().lower() == "total-condense":
            mv_mode = "top-anchor"
        else:
            mv_mode = "condenser-duty"
    if mv_mode in ("duty", "cond-duty", "condenser-duty", "condenserduty"):
        mv_mode = "condenser-duty"
    elif mv_mode in ("top-anchor", "pressure-anchor", "p-anchor", "toppressure"):
        mv_mode = "top-anchor"
    else:
        mv_mode = "top-anchor"

    # Avoid hidden pressure/composition coupling by default:
    # condenser-duty MV with total-condense drives condenser mass-split authority.
    if (
        mv_mode == "condenser-duty"
        and str(cfg.condenser_duty_mode or "").strip().lower() == "total-condense"
        and (not bool(getattr(cfg, "allow_coupled_pressure_duty", False)))
    ):
        mv_mode = "top-anchor"
        pressure_mode_note = (
            "pressure-control-mv=condenser-duty with condenser-duty-mode=total-condense "
            "was auto-switched to top-anchor; use --allow-coupled-pressure-duty to keep coupled duty control."
        )

    kc = cfg.top_pressure_kc
    if kc is None:
        kc = _spec_float(specs, "Top Pressure Kc", "Pressure Controller Kc")
    if kc is None:
        kc = -0.5 if mv_mode == "top-anchor" else -5.0e5

    ti = cfg.top_pressure_ti_sec
    if ti is None:
        ti = _spec_float(specs, "Top Pressure Ti (sec)", "Pressure Controller Ti (sec)")
    if ti is None:
        ti = 120.0

    if mv_mode == "condenser-duty":
        q_bias = float(cfg.condenser_duty_btu_per_h) if cfg.condenser_duty_btu_per_h is not None else _infer_condenser_duty_bias_btu_per_h(col)
        q_min = cfg.condenser_duty_min_btu_per_h
        q_max = cfg.condenser_duty_max_btu_per_h

        if q_min is None:
            q_min = -3.0 * max(abs(q_bias), 1.0)
        if q_max is None:
            q_max = 0.0
        q_min = float(q_min)
        q_max = float(q_max)
        if q_min > q_max:
            q_min, q_max = q_max, q_min

        ctrl = PIController(
            kc=float(kc),
            ti_sec=float(ti),
            bias=float(q_bias),
            out_min=float(q_min),
            out_max=float(q_max),
            integ=float(_get_controller_restart_value(col, "top_pressure_integ") or 0.0),
        )
        return True, ctrl, float(sp), mv_mode, pressure_mode_note

    # top-anchor pressure control mode
    p_ctrl_idx = 0
    try:
        p_bias = float(np.asarray(getattr(col, "P_psia"), dtype=float).reshape((-1,))[p_ctrl_idx])
    except Exception:
        p_bias = float(sp)
    if not np.isfinite(p_bias) or p_bias <= 0.0:
        p_bias = float(sp)

    p_min = cfg.top_pressure_anchor_min_psia
    p_max = cfg.top_pressure_anchor_max_psia
    if p_min is None:
        p_min = max(1.0, 0.5 * float(sp))
    if p_max is None:
        p_max = max(1.5 * float(sp), float(sp) + 20.0)
    p_min = float(p_min)
    p_max = float(p_max)
    if p_min > p_max:
        p_min, p_max = p_max, p_min

    ctrl = PIController(
        kc=float(kc),
        ti_sec=float(ti),
        bias=float(p_bias),
        out_min=float(p_min),
        out_max=float(p_max),
        integ=float(_get_controller_restart_value(col, "top_pressure_integ") or 0.0),
    )
    return True, ctrl, float(sp), mv_mode, pressure_mode_note


def _allow_coupled_total_condenser_partial_condense(
    *,
    cfg: "RunnerConfig",
    pressure_control_mv: str,
    condenser_duty_mode: str,
) -> bool:
    return bool(
        bool(getattr(cfg, "enable_coupled_total_condenser_partial_condense", True))
        and str(pressure_control_mv).strip().lower() == "condenser-duty"
        and str(condenser_duty_mode).strip().lower() == "total-condense"
    )


def _build_level_controllers(
    *,
    col: ColumnSpec,
    cfg: RunnerConfig,
    layout: StateVectorLayout,
    y0: np.ndarray,
    dist_tag: StreamTag,
    bots_tag: StreamTag,
) -> Tuple[bool, Optional[PIController], Optional[PIController], Optional[float], Optional[float]]:
    specs = getattr(col, "specs_raw", None) or {}
    enabled = bool(cfg.enable_level_control)
    if not enabled:
        b = _as_bool(_spec_get(specs, "Enable Level Control", "Level Control Enabled"))
        enabled = bool(b) if b is not None else False
    if not enabled:
        return False, None, None, None, None

    u0 = layout.unpack(np.asarray(y0, dtype=float))
    top0 = float(np.sum(u0["top_L"])) if (layout.include_top and "top_L" in u0) else None
    bot0 = float(np.sum(u0["bottom_L"])) if (layout.include_bottom and "bottom_L" in u0) else None

    sp_top = cfg.top_level_sp_lbmol
    if sp_top is None:
        sp_top = _spec_float(specs, "Top Level SP (lbmol)", "Top Drum Level SP (lbmol)", "Reflux Drum Level SP (lbmol)")
    if sp_top is None and top0 is not None:
        sp_top = float(top0)

    sp_bot = cfg.bottom_level_sp_lbmol
    if sp_bot is None:
        sp_bot = _spec_float(specs, "Bottom Level SP (lbmol)", "Bottom Sump Level SP (lbmol)")
    if sp_bot is None and bot0 is not None:
        sp_bot = float(bot0)

    d_bias = float(dist_tag.flow_lbmolph) if dist_tag.flow_lbmolph is not None else 0.0
    b_bias = float(bots_tag.flow_lbmolph) if bots_tag.flow_lbmolph is not None else 0.0

    top_kc = cfg.top_level_kc
    if top_kc is None:
        top_kc = _spec_float(specs, "Top Level Kc", "Top Level Controller Kc")
    if top_kc is None:
        top_kc = 8.0

    top_ti = cfg.top_level_ti_sec
    if top_ti is None:
        top_ti = _spec_float(specs, "Top Level Ti (sec)", "Top Level Controller Ti (sec)")
    if top_ti is None:
        top_ti = 120.0

    bot_kc = cfg.bottom_level_kc
    if bot_kc is None:
        bot_kc = _spec_float(specs, "Bottom Level Kc", "Bottom Level Controller Kc")
    if bot_kc is None:
        bot_kc = 8.0

    bot_ti = cfg.bottom_level_ti_sec
    if bot_ti is None:
        bot_ti = _spec_float(specs, "Bottom Level Ti (sec)", "Bottom Level Controller Ti (sec)")
    if bot_ti is None:
        bot_ti = 120.0

    d_hi = max(2.5 * max(d_bias, 1.0), d_bias + 5000.0)
    b_hi = max(2.5 * max(b_bias, 1.0), b_bias + 5000.0)

    if sp_top is None or sp_bot is None:
        return False, None, None, None, None

    top_ctrl = PIController(
        kc=float(top_kc),
        ti_sec=float(top_ti),
        bias=float(d_bias),
        out_min=0.0,
        out_max=float(d_hi),
        integ=float(_get_controller_restart_value(col, "top_level_integ") or 0.0),
    )
    bot_ctrl = PIController(
        kc=float(bot_kc),
        ti_sec=float(bot_ti),
        bias=float(b_bias),
        out_min=0.0,
        out_max=float(b_hi),
        integ=float(_get_controller_restart_value(col, "bottom_level_integ") or 0.0),
    )
    return True, top_ctrl, bot_ctrl, (float(sp_top) if sp_top is not None else None), (float(sp_bot) if sp_bot is not None else None)


def _estimate_top_drum_liquid_volume_ft3(
    col: ColumnSpec,
    layout: StateVectorLayout,
    y_vec: np.ndarray,
    thermo_provider: Any,
    top_drum_total_volume_ft3: Optional[float],
    p_top_psia: Optional[float] = None,
) -> Tuple[Optional[float], Optional[float], Optional[float]]:
    if (
        top_drum_total_volume_ft3 is None
        or (not np.isfinite(float(top_drum_total_volume_ft3)))
        or float(top_drum_total_volume_ft3) <= 0.0
        or (not layout.include_top)
        or thermo_provider is None
        or not hasattr(thermo_provider, "liquid_density_lbmol_ft3")
    ):
        return None, None, None
    try:
        u = layout.unpack(np.asarray(y_vec, dtype=float))
    except Exception:
        return None, None, None
    if "top_L" not in u:
        return None, None, None
    try:
        top_L = np.asarray(u["top_L"], dtype=float).reshape((-1,))
    except Exception:
        return None, None, None
    m_top_liq = float(np.sum(top_L))
    if (not np.isfinite(m_top_liq)) or m_top_liq < 0.0:
        return None, None, None
    if m_top_liq <= 0.0:
        return 0.0, None, 0.0
    x_top = np.zeros_like(top_L, dtype=float)
    if m_top_liq > 1e-12:
        x_top = np.clip(top_L / m_top_liq, 0.0, None)
        s = float(np.sum(x_top))
        if s > 0.0:
            x_top = x_top / s
    try:
        if "tray_T_f" in u:
            T_top = float(np.asarray(u["tray_T_f"], dtype=float).reshape((col.n_stages,))[0])
        else:
            T_top = float(np.asarray(getattr(col, "T_f"), dtype=float).reshape((col.n_stages,))[0])
    except Exception:
        return None, None, None
    if p_top_psia is None:
        try:
            p_top_psia = float(np.asarray(getattr(col, "P_psia"), dtype=float).reshape((col.n_stages,))[0])
        except Exception:
            p_top_psia = None
    if p_top_psia is None or (not np.isfinite(float(p_top_psia))) or float(p_top_psia) <= 0.0:
        return None, None, None
    try:
        rho_top = float(
            thermo_provider.liquid_density_lbmol_ft3(
                float(T_top),
                float(p_top_psia),
                np.asarray(x_top, dtype=float),
            )
        )
    except Exception:
        return None, None, None
    if (not np.isfinite(rho_top)) or rho_top <= 1e-12:
        return None, None, None
    try:
        total_vol = float(top_drum_total_volume_ft3)
        top_liq_vol_ft3 = float(np.clip(m_top_liq / rho_top, 0.0, total_vol))
        top_liq_frac = float(np.clip(top_liq_vol_ft3 / total_vol, 0.0, 1.0))
    except Exception:
        return None, None, None
    return top_liq_vol_ft3, float(rho_top), top_liq_frac


def _estimate_bottom_sump_liquid_volume_ft3(
    col: ColumnSpec,
    layout: StateVectorLayout,
    y_vec: np.ndarray,
    thermo_provider: Any,
    bottom_sump_total_volume_ft3: Optional[float],
    p_bottom_psia: Optional[float] = None,
) -> Tuple[Optional[float], Optional[float], Optional[float]]:
    if (
        bottom_sump_total_volume_ft3 is None
        or (not np.isfinite(float(bottom_sump_total_volume_ft3)))
        or float(bottom_sump_total_volume_ft3) <= 0.0
        or (not layout.include_bottom)
        or thermo_provider is None
        or not hasattr(thermo_provider, "liquid_density_lbmol_ft3")
    ):
        return None, None, None
    try:
        u = layout.unpack(np.asarray(y_vec, dtype=float))
    except Exception:
        return None, None, None
    if "bottom_L" not in u:
        return None, None, None
    try:
        bottom_L = np.asarray(u["bottom_L"], dtype=float).reshape((-1,))
    except Exception:
        return None, None, None
    m_bottom_liq = float(np.sum(bottom_L))
    if (not np.isfinite(m_bottom_liq)) or m_bottom_liq < 0.0:
        return None, None, None
    if m_bottom_liq <= 0.0:
        return 0.0, None, 0.0
    x_bottom = np.zeros_like(bottom_L, dtype=float)
    if m_bottom_liq > 1e-12:
        x_bottom = np.clip(bottom_L / m_bottom_liq, 0.0, None)
        s = float(np.sum(x_bottom))
        if s > 0.0:
            x_bottom = x_bottom / s
    try:
        if "tray_T_f" in u:
            T_bottom = float(np.asarray(u["tray_T_f"], dtype=float).reshape((col.n_stages,))[-1])
        else:
            T_bottom = float(np.asarray(getattr(col, "T_f"), dtype=float).reshape((col.n_stages,))[-1])
    except Exception:
        return None, None, None
    if p_bottom_psia is None:
        try:
            p_bottom_psia = float(np.asarray(getattr(col, "P_psia"), dtype=float).reshape((col.n_stages,))[-1])
        except Exception:
            p_bottom_psia = None
    if p_bottom_psia is None or (not np.isfinite(float(p_bottom_psia))) or float(p_bottom_psia) <= 0.0:
        return None, None, None
    try:
        rho_bottom = float(
            thermo_provider.liquid_density_lbmol_ft3(
                float(T_bottom),
                float(p_bottom_psia),
                np.asarray(x_bottom, dtype=float),
            )
        )
    except Exception:
        return None, None, None
    if (not np.isfinite(rho_bottom)) or rho_bottom <= 1e-12:
        return None, None, None
    try:
        total_vol = float(bottom_sump_total_volume_ft3)
        bottom_liq_vol_ft3 = float(np.clip(m_bottom_liq / rho_bottom, 0.0, total_vol))
        bottom_liq_frac = float(np.clip(bottom_liq_vol_ft3 / total_vol, 0.0, 1.0))
    except Exception:
        return None, None, None
    return bottom_liq_vol_ft3, float(rho_bottom), bottom_liq_frac


def _horizontal_cylinder_volume_fraction_from_height_fraction(h_over_d: float) -> float:
    try:
        hf = float(h_over_d)
    except Exception:
        return np.nan
    if not np.isfinite(hf):
        return np.nan
    if hf <= 0.0:
        return 0.0
    if hf >= 1.0:
        return 1.0
    R = 0.5
    h = hf
    try:
        term = max(2.0 * R * h - h * h, 0.0)
        area = (R * R) * np.arccos((R - h) / R) - (R - h) * np.sqrt(term)
        frac = float(area / (np.pi * R * R))
    except Exception:
        return np.nan
    return float(np.clip(frac, 0.0, 1.0))


def _horizontal_cylinder_height_fraction_from_volume_fraction(v_over_v: float) -> float:
    try:
        vf = float(v_over_v)
    except Exception:
        return np.nan
    if not np.isfinite(vf):
        return np.nan
    if vf <= 0.0:
        return 0.0
    if vf >= 1.0:
        return 1.0
    if abs(vf - 0.5) <= 1e-14:
        return 0.5
    lo = 0.0
    hi = 1.0
    for _ in range(48):
        mid = 0.5 * (lo + hi)
        vmid = _horizontal_cylinder_volume_fraction_from_height_fraction(mid)
        if not np.isfinite(vmid):
            return np.nan
        if vmid < vf:
            lo = mid
        else:
            hi = mid
    return float(0.5 * (lo + hi))


def _desired_inventory_recovery_rate_lbmolph(
    *,
    total_lbmol: Optional[float],
    pv: Optional[float],
    sp: Optional[float],
    pv_mode: str,
    lbmol_per_volume_fraction_scale: Optional[float] = None,
    recover_tau_sec: float = 120.0,
) -> float:
    if recover_tau_sec is None or (not np.isfinite(float(recover_tau_sec))) or float(recover_tau_sec) <= 1e-9:
        recover_tau_sec = 120.0
    if str(pv_mode).strip().lower() == "true-level":
        if (
            lbmol_per_volume_fraction_scale is None
            or (not np.isfinite(float(lbmol_per_volume_fraction_scale)))
            or float(lbmol_per_volume_fraction_scale) <= 1e-12
            or pv is None
            or sp is None
            or (not np.isfinite(float(pv)))
            or (not np.isfinite(float(sp)))
        ):
            return 0.0
        vfrac_pv = _horizontal_cylinder_volume_fraction_from_height_fraction(float(pv))
        vfrac_sp = _horizontal_cylinder_volume_fraction_from_height_fraction(float(sp))
        if (not np.isfinite(float(vfrac_pv))) or (not np.isfinite(float(vfrac_sp))):
            return 0.0
        desired_delta_lbmol = float(lbmol_per_volume_fraction_scale) * (float(vfrac_sp) - float(vfrac_pv))
        return float(desired_delta_lbmol) * 3600.0 / float(recover_tau_sec)
    if (
        total_lbmol is None
        or sp is None
        or (not np.isfinite(float(total_lbmol)))
        or (not np.isfinite(float(sp)))
    ):
        return 0.0
    return (float(sp) - float(total_lbmol)) * 3600.0 / float(recover_tau_sec)


def _component_index_by_name(col: ColumnSpec, token: str) -> Optional[int]:
    tok = _normalize_spec_key(token)
    if not tok:
        return None

    excel_names = list(getattr(col, "components_excel", []) or [])
    dwsim_names = list(getattr(col, "components_dwsim", []) or [])
    n_comp = int(getattr(col, "n_components", len(excel_names)))

    def _aliases_for(i: int) -> List[str]:
        out: List[str] = []
        if i < len(excel_names):
            out.append(_normalize_spec_key(excel_names[i]))
        if i < len(dwsim_names):
            out.append(_normalize_spec_key(dwsim_names[i]))
        for v in list(out):
            if len(v) >= 3 and v.startswith("c") and "h" in v:
                h_pos = v.find("h")
                cnum = v[1:h_pos]
                if cnum.isdigit():
                    out.append(f"c{int(cnum)}")
        return [a for a in out if a]

    all_aliases = [_aliases_for(i) for i in range(n_comp)]
    for i, aliases in enumerate(all_aliases):
        if tok in aliases:
            return i

    carbon_alias = {
        "c1": ("methane",),
        "c2": ("ethane",),
        "c3": ("propane",),
        "c4": ("butane",),
        "c5": ("pentane",),
        "c6": ("hexane",),
        "c7": ("heptane",),
        "c8": ("octane",),
    }
    if tok in carbon_alias:
        hints = tuple(_normalize_spec_key(v) for v in carbon_alias[tok])
        for i, aliases in enumerate(all_aliases):
            for a in aliases:
                for h in hints:
                    if h and h in a:
                        return i

    for i, aliases in enumerate(all_aliases):
        for a in aliases:
            if len(tok) >= 2 and (a.startswith(tok) or tok in a):
                return i
    return None


def _build_distillate_composition_controller(
    *,
    col: ColumnSpec,
    cfg: RunnerConfig,
    boundary: BoundaryFlows,
    dist_tag: StreamTag,
) -> Tuple[bool, Optional[PIController], Optional[float], Optional[int], Optional[str]]:
    specs = getattr(col, "specs_raw", None) or {}
    enabled = bool(cfg.enable_distillate_composition_control)
    if (not enabled) and (cfg.distillate_composition_sp_molfrac is not None):
        enabled = True
    if not enabled:
        b = _as_bool(
            _spec_get(
                specs,
                "Enable Distillate Composition Control",
                "Distillate Composition Control Enabled",
            )
        )
        enabled = bool(b) if b is not None else False
    if not enabled:
        return False, None, None, None, None

    comp_name = str(cfg.distillate_composition_component or "").strip()
    if not comp_name:
        comp_name = str(
            _spec_get(
                specs,
                "Distillate Composition Component",
                "Distillate Controller Component",
            )
            or "C4"
        ).strip()

    comp_idx = _component_index_by_name(col, comp_name)
    if comp_idx is None:
        return False, None, None, None, None

    sp = cfg.distillate_composition_sp_molfrac
    if sp is None:
        sp = _spec_float(
            specs,
            "Distillate Composition SP",
            "Distillate C4 SP",
            "Distillate x SP",
        )
    if sp is None:
        return False, None, None, None, None
    sp = float(sp)
    if (not np.isfinite(sp)) or sp < 0.0 or sp > 1.0:
        return False, None, None, None, None

    kc = cfg.distillate_composition_kc
    if kc is None:
        kc = _spec_float(specs, "Distillate Composition Kc", "Distillate Controller Kc")
    if kc is None:
        kc = 1.0e4

    ti = cfg.distillate_composition_ti_sec
    if ti is None:
        ti = _spec_float(specs, "Distillate Composition Ti (sec)", "Distillate Controller Ti (sec)")
    if ti is None:
        ti = 240.0

    l_bias = float(boundary.reflux_lbmolph) if boundary.reflux_lbmolph is not None else np.nan
    d_bias = float(dist_tag.flow_lbmolph) if dist_tag.flow_lbmolph is not None else np.nan
    if (not np.isfinite(l_bias)) or l_bias < 0.0:
        l_bias = float(np.asarray(getattr(col, "L_lbmolph"), dtype=float).reshape((-1,))[0])

    l_min = cfg.reflux_cmd_min_lbmolph
    l_max = cfg.reflux_cmd_max_lbmolph

    # Backward-compatible fallback: if only ratio clamps are provided, map them
    # to reflux-flow clamps around current distillate flow.
    if l_min is None and cfg.reflux_ratio_min is not None and np.isfinite(d_bias) and d_bias > 0.0:
        l_min = float(cfg.reflux_ratio_min) * float(d_bias)
    if l_max is None and cfg.reflux_ratio_max is not None and np.isfinite(d_bias) and d_bias > 0.0:
        l_max = float(cfg.reflux_ratio_max) * float(d_bias)

    if l_min is None:
        l_min = _spec_float(
            specs,
            "Reflux Flow Min (lbmol/h)",
            "Distillate Composition Reflux Min (lbmol/h)",
        )
    if l_max is None:
        l_max = _spec_float(
            specs,
            "Reflux Flow Max (lbmol/h)",
            "Distillate Composition Reflux Max (lbmol/h)",
        )

    if l_min is None:
        l_min = 0.0
    if l_max is None:
        l_max = max(2.5 * max(float(l_bias), 1.0), float(l_bias) + 5000.0)
    l_min = float(l_min)
    l_max = float(l_max)
    if l_min > l_max:
        l_min, l_max = l_max, l_min

    ctrl = PIController(
        kc=float(kc),
        ti_sec=float(ti),
        bias=float(l_bias),
        out_min=float(l_min),
        out_max=float(l_max),
        integ=float(_get_controller_restart_value(col, "distillate_comp_integ") or 0.0),
    )

    comp_label = None
    try:
        comp_label = str(list(getattr(col, "components_excel", []))[int(comp_idx)])
    except Exception:
        comp_label = str(comp_name)
    return True, ctrl, float(sp), int(comp_idx), comp_label


def _build_bottoms_composition_controller(
    *,
    col: ColumnSpec,
    cfg: RunnerConfig,
    boundary: BoundaryFlows,
) -> Tuple[bool, Optional[str], Optional[PIController], Optional[float], Optional[int], Optional[str]]:
    specs = getattr(col, "specs_raw", None) or {}
    enabled = bool(cfg.enable_bottoms_composition_control)
    if (not enabled) and (cfg.bottoms_composition_sp_molfrac is not None):
        enabled = True
    if not enabled:
        b = _as_bool(
            _spec_get(
                specs,
                "Enable Bottoms Composition Control",
                "Bottoms Composition Control Enabled",
            )
        )
        enabled = bool(b) if b is not None else False
    if not enabled:
        return False, None, None, None, None, None

    comp_name = str(cfg.bottoms_composition_component or "").strip()
    if not comp_name:
        comp_name = str(
            _spec_get(
                specs,
                "Bottoms Composition Component",
                "Bottoms Controller Component",
            )
            or "C5"
        ).strip()

    comp_idx = _component_index_by_name(col, comp_name)
    if comp_idx is None:
        return False, None, None, None, None, None

    sp = cfg.bottoms_composition_sp_molfrac
    if sp is None:
        sp = _spec_float(
            specs,
            "Bottoms Composition SP",
            "Bottoms C5 SP",
            "Bottoms x SP",
        )
    if sp is None:
        return False, None, None, None, None, None
    sp = float(sp)
    if (not np.isfinite(sp)) or sp < 0.0 or sp > 1.0:
        return False, None, None, None, None, None

    mv_raw = str(cfg.bottoms_composition_mv or "").strip()
    if not mv_raw:
        mv_raw = str(
            _spec_get(
                specs,
                "Bottoms Composition MV",
                "Bottoms Controller MV",
            )
            or "boilup"
        ).strip()
    mv_norm = mv_raw.lower().replace("_", "-").replace(" ", "")
    if mv_norm in ("duty", "reboilerduty", "reboiler-duty", "qreb", "reboilerq"):
        mv_mode = "reboiler-duty"
    else:
        mv_mode = "boilup"

    kc = cfg.bottoms_composition_kc
    if kc is None:
        kc = _spec_float(specs, "Bottoms Composition Kc", "Bottoms Controller Kc")

    ti = cfg.bottoms_composition_ti_sec
    if ti is None:
        ti = _spec_float(specs, "Bottoms Composition Ti (sec)", "Bottoms Controller Ti (sec)")
    if ti is None:
        ti = 240.0

    if mv_mode == "reboiler-duty":
        q_bias = cfg.reboiler_duty_btu_per_h
        if q_bias is None:
            q_bias = _spec_float(specs, "Reboiler Duty (Btu/h)")
        if q_bias is None:
            try:
                q_bias = float(getattr(getattr(col, "duties", None), "q_reb_btu_per_h"))
            except Exception:
                q_bias = None
        if q_bias is None or (not np.isfinite(float(q_bias))):
            q_bias = 0.0
        q_bias = float(q_bias)

        if kc is None:
            # Preserve rough authority of legacy boilup-MV default by converting
            # with an estimated latent heat near the design point.
            v_bias = float(boundary.boilup_lbmolph) if boundary.boilup_lbmolph is not None else np.nan
            if (not np.isfinite(v_bias)) or v_bias <= 0.0:
                try:
                    v_bias = float(np.asarray(getattr(col, "V_lbmolph"), dtype=float).reshape((-1,))[-1])
                except Exception:
                    v_bias = np.nan
            latent_est = q_bias / max(v_bias, 1.0) if np.isfinite(v_bias) and v_bias > 0.0 else 15000.0
            kc = -1.0e4 * float(latent_est)

        q_min = cfg.reboiler_duty_cmd_min_btu_per_h
        q_max = cfg.reboiler_duty_cmd_max_btu_per_h
        if q_min is None:
            q_min = _spec_float(
                specs,
                "Reboiler Duty Min (Btu/h)",
                "Bottoms Composition Reboiler Duty Min (Btu/h)",
            )
        if q_max is None:
            q_max = _spec_float(
                specs,
                "Reboiler Duty Max (Btu/h)",
                "Bottoms Composition Reboiler Duty Max (Btu/h)",
            )
        if q_min is None:
            q_min = 0.0
        if q_max is None:
            q_max = max(2.5 * max(float(q_bias), 1.0), float(q_bias) + 5.0e7)
        q_min = float(q_min)
        q_max = float(q_max)
        if q_min > q_max:
            q_min, q_max = q_max, q_min

        ctrl = PIController(
            kc=float(kc),
            ti_sec=float(ti),
            bias=float(q_bias),
            out_min=float(q_min),
            out_max=float(q_max),
            integ=float(_get_controller_restart_value(col, "bottoms_comp_integ") or 0.0),
        )
    else:
        if kc is None:
            kc = -1.0e4
        v_bias = float(boundary.boilup_lbmolph) if boundary.boilup_lbmolph is not None else np.nan
        if (not np.isfinite(v_bias)) or v_bias < 0.0:
            v_bias = float(np.asarray(getattr(col, "V_lbmolph"), dtype=float).reshape((-1,))[-1])

        v_min = cfg.boilup_cmd_min_lbmolph
        v_max = cfg.boilup_cmd_max_lbmolph
        if v_min is None:
            v_min = _spec_float(
                specs,
                "Boilup Flow Min (lbmol/h)",
                "Bottoms Composition Boilup Min (lbmol/h)",
            )
        if v_max is None:
            v_max = _spec_float(
                specs,
                "Boilup Flow Max (lbmol/h)",
                "Bottoms Composition Boilup Max (lbmol/h)",
            )
        if v_min is None:
            v_min = 0.0
        if v_max is None:
            v_max = max(2.5 * max(float(v_bias), 1.0), float(v_bias) + 5000.0)
        v_min = float(v_min)
        v_max = float(v_max)
        if v_min > v_max:
            v_min, v_max = v_max, v_min

        ctrl = PIController(
            kc=float(kc),
            ti_sec=float(ti),
            bias=float(v_bias),
            out_min=float(v_min),
            out_max=float(v_max),
            integ=float(_get_controller_restart_value(col, "bottoms_comp_integ") or 0.0),
        )

    comp_label = None
    try:
        comp_label = str(list(getattr(col, "components_excel", []))[int(comp_idx)])
    except Exception:
        comp_label = str(comp_name)
    return True, mv_mode, ctrl, float(sp), int(comp_idx), comp_label


def _profile_rows(
    t_s: float,
    case: CaseData,
    col: ColumnSpec,
    layout: StateVectorLayout,
    y: np.ndarray,
    diag: Dict[str, np.ndarray],
    *,
    include_temperature: bool,
    volume_model: VolumeModel,
    wall_clock_iso: str,
    wall_elapsed_s: float,
    feed_tag: StreamTag,
    dist_tag: StreamTag,
    bots_tag: StreamTag,
) -> List[Dict[str, Any]]:
    u = layout.unpack(y)
    N = col.n_stages
    Nc = col.n_components

    def _comp_suffix(name: str) -> str:
        out = "".join(ch if ch.isalnum() else "_" for ch in str(name).strip())
        out = out.strip("_")
        return out or "comp"

    comp_labels = [_comp_suffix(nm) for nm in getattr(col, "components_excel", [f"c{i+1}" for i in range(Nc)])]

    x = diag["x_tray"] if "x_tray" in diag else u["x_tray"]
    if "y_tray" in diag:
        yv = diag["y_tray"]
    elif "y_eq_thermo_tray" in diag:
        yv = diag["y_eq_thermo_tray"]
    elif "y0" in dir(col):
        yv = np.asarray(getattr(col, "y0"), dtype=float).reshape((N, Nc))
    else:
        yv = np.zeros((N, Nc), dtype=float)
    ML = u["ML_tot_tray"]
    MV = u["MV_tot_tray"] if "MV_tot_tray" in u else np.zeros(N, dtype=float)
    top_L_total = float(np.sum(u["top_L"])) if (layout.include_top and "top_L" in u) else None
    top_x = None
    bottom_L_total = float(np.sum(u["bottom_L"])) if (layout.include_bottom and "bottom_L" in u) else None
    bottom_x = None
    if layout.include_top and "top_L" in u:
        denom = max(float(np.sum(u["top_L"])), 1e-300)
        top_x = np.asarray(u["top_L"], dtype=float).reshape((Nc,)) / denom
    if layout.include_bottom and "bottom_L" in u:
        denom = max(float(np.sum(u["bottom_L"])), 1e-300)
        bottom_x = np.asarray(u["bottom_L"], dtype=float).reshape((Nc,)) / denom

    P_spec = np.asarray(getattr(col, "P_psia", np.full(N, np.nan, dtype=float)), dtype=float).reshape((N,))

    Z_raw = diag.get("Z_tray", np.full(N, np.nan, dtype=float))
    Z = np.asarray(Z_raw, dtype=float).reshape((N,))
    Z = np.where(~np.isfinite(Z) | (Z <= 0.0), 1.0, Z)
    P_hyd = None
    if "P_psia_hyd" in diag:
        try:
            P_hyd = np.asarray(diag["P_psia_hyd"], dtype=float).reshape((N,))
        except Exception:
            P_hyd = None
    L_out_hyd = None
    if "L_out_hyd_lbmolph" in diag:
        try:
            L_out_hyd = np.asarray(diag["L_out_hyd_lbmolph"], dtype=float).reshape((N,))
        except Exception:
            L_out_hyd = None
    h_ow = None
    if "h_ow_ft" in diag:
        try:
            h_ow = np.asarray(diag["h_ow_ft"], dtype=float).reshape((N,))
        except Exception:
            h_ow = None
    V_out_calc = None
    if "V_out_lbmolph" in diag:
        try:
            V_out_calc = np.asarray(diag["V_out_lbmolph"], dtype=float).reshape((N,))
        except Exception:
            V_out_calc = None
    hyd_dp_raw = None
    hyd_dp_dry_raw = None
    hyd_dp_liq_raw = None
    hyd_dp_used = None
    if "hydraulic_dp_raw_psia" in diag:
        try:
            hyd_dp_raw = np.asarray(diag["hydraulic_dp_raw_psia"], dtype=float).reshape((N,))
        except Exception:
            hyd_dp_raw = None
    if "hydraulic_dp_dry_raw_psia" in diag:
        try:
            hyd_dp_dry_raw = np.asarray(diag["hydraulic_dp_dry_raw_psia"], dtype=float).reshape((N,))
        except Exception:
            hyd_dp_dry_raw = None
    if "hydraulic_dp_liq_raw_psia" in diag:
        try:
            hyd_dp_liq_raw = np.asarray(diag["hydraulic_dp_liq_raw_psia"], dtype=float).reshape((N,))
        except Exception:
            hyd_dp_liq_raw = None
    if "hydraulic_dp_used_psia" in diag:
        try:
            hyd_dp_used = np.asarray(diag["hydraulic_dp_used_psia"], dtype=float).reshape((N,))
        except Exception:
            hyd_dp_used = None
    vflow_ok = None
    vflow_denom = None
    vflow_calc = None
    vflow_used = None
    vflow_clamped = None
    vflow_limit_hi = None
    vflow_limit_lo = None
    vflow_alpha = None
    if "vflow_energy_ok" in diag:
        try:
            vflow_ok = np.asarray(diag["vflow_energy_ok"], dtype=float).reshape((N,))
        except Exception:
            vflow_ok = None
    if "vflow_energy_denom_BTU_per_lbmol" in diag:
        try:
            vflow_denom = np.asarray(diag["vflow_energy_denom_BTU_per_lbmol"], dtype=float).reshape((N,))
        except Exception:
            vflow_denom = None
    if "vflow_energy_calc_lbmolph" in diag:
        try:
            vflow_calc = np.asarray(diag["vflow_energy_calc_lbmolph"], dtype=float).reshape((N,))
        except Exception:
            vflow_calc = None
    if "vflow_energy_used_lbmolph" in diag:
        try:
            vflow_used = np.asarray(diag["vflow_energy_used_lbmolph"], dtype=float).reshape((N,))
        except Exception:
            vflow_used = None
    if "vflow_energy_clamped" in diag:
        try:
            vflow_clamped = np.asarray(diag["vflow_energy_clamped"], dtype=float).reshape((N,))
        except Exception:
            vflow_clamped = None
    if "vflow_energy_limit_hi_lbmolph" in diag:
        try:
            vflow_limit_hi = np.asarray(diag["vflow_energy_limit_hi_lbmolph"], dtype=float).reshape((N,))
        except Exception:
            vflow_limit_hi = None
    if "vflow_energy_limit_lo_lbmolph" in diag:
        try:
            vflow_limit_lo = np.asarray(diag["vflow_energy_limit_lo_lbmolph"], dtype=float).reshape((N,))
        except Exception:
            vflow_limit_lo = None
    if "vflow_relax_alpha" in diag:
        try:
            vflow_alpha = np.asarray(diag["vflow_relax_alpha"], dtype=float).reshape((N,))
        except Exception:
            vflow_alpha = None
    mass_resid = None
    if "mass_balance_resid_lbmolps_tray" in diag:
        try:
            mass_resid = np.asarray(diag["mass_balance_resid_lbmolps_tray"], dtype=float).reshape((N,))
        except Exception:
            mass_resid = None
    energy_resid = None
    if "energy_balance_resid_BTUps_tray" in diag:
        try:
            energy_resid = np.asarray(diag["energy_balance_resid_BTUps_tray"], dtype=float).reshape((N,))
        except Exception:
            energy_resid = None
    dMLdt_transport = None
    if "dMLdt_transport_lbmolps_tray" in diag:
        try:
            dMLdt_transport = np.asarray(diag["dMLdt_transport_lbmolps_tray"], dtype=float).reshape((N,))
        except Exception:
            dMLdt_transport = None
    dMLdt_phase_relax = None
    if "dMLdt_phase_relax_lbmolps_tray" in diag:
        try:
            dMLdt_phase_relax = np.asarray(diag["dMLdt_phase_relax_lbmolps_tray"], dtype=float).reshape((N,))
        except Exception:
            dMLdt_phase_relax = None
    dMLdt_total = None
    if "dMLdt_total_lbmolps_tray" in diag:
        try:
            dMLdt_total = np.asarray(diag["dMLdt_total_lbmolps_tray"], dtype=float).reshape((N,))
        except Exception:
            dMLdt_total = None
    dMLdt_feed = None
    if "dMLdt_feed_lbmolps_tray" in diag:
        try:
            dMLdt_feed = np.asarray(diag["dMLdt_feed_lbmolps_tray"], dtype=float).reshape((N,))
        except Exception:
            dMLdt_feed = None
    beta_eq_tray = None
    if "beta_eq_tray" in diag:
        try:
            beta_eq_tray = np.asarray(diag["beta_eq_tray"], dtype=float).reshape((N,))
        except Exception:
            beta_eq_tray = None
    x_eq_tray = None
    if "x_eq_tray" in diag:
        try:
            x_eq_tray = np.asarray(diag["x_eq_tray"], dtype=float).reshape((N, Nc))
        except Exception:
            x_eq_tray = None
    y_eq_tray = None
    if "y_eq_tray" in diag:
        try:
            y_eq_tray = np.asarray(diag["y_eq_tray"], dtype=float).reshape((N, Nc))
        except Exception:
            y_eq_tray = None
    y_target_tray = None
    if "y_target_tray" in diag:
        try:
            y_target_tray = np.asarray(diag["y_target_tray"], dtype=float).reshape((N, Nc))
        except Exception:
            y_target_tray = None
    eq_target_mv_total = None
    if "eq_target_mv_total_lbmol_tray" in diag:
        try:
            eq_target_mv_total = np.asarray(diag["eq_target_mv_total_lbmol_tray"], dtype=float).reshape((N,))
        except Exception:
            eq_target_mv_total = None
    eq_flash_mv_total = None
    if "eq_flash_mv_total_lbmol_tray" in diag:
        try:
            eq_flash_mv_total = np.asarray(diag["eq_flash_mv_total_lbmol_tray"], dtype=float).reshape((N,))
        except Exception:
            eq_flash_mv_total = None
    eq_target_vapor_total = None
    if "eq_target_vapor_total_lbmol_tray" in diag:
        try:
            eq_target_vapor_total = np.asarray(diag["eq_target_vapor_total_lbmol_tray"], dtype=float).reshape((N,))
        except Exception:
            eq_target_vapor_total = None
    eq_target_vapor_delta = None
    if "eq_target_vapor_delta_lbmol_tray" in diag:
        try:
            eq_target_vapor_delta = np.asarray(diag["eq_target_vapor_delta_lbmol_tray"], dtype=float).reshape((N,))
        except Exception:
            eq_target_vapor_delta = None
    eq_target_vapor_fraction = None
    if "eq_target_vapor_fraction_tray" in diag:
        try:
            eq_target_vapor_fraction = np.asarray(diag["eq_target_vapor_fraction_tray"], dtype=float).reshape((N,))
        except Exception:
            eq_target_vapor_fraction = None
    eq_current_vapor_fraction = None
    if "eq_current_vapor_fraction_tray" in diag:
        try:
            eq_current_vapor_fraction = np.asarray(diag["eq_current_vapor_fraction_tray"], dtype=float).reshape((N,))
        except Exception:
            eq_current_vapor_fraction = None
    eq_phase_change = None
    if "eq_phase_change_lbmolps_tray" in diag:
        try:
            eq_phase_change = np.asarray(diag["eq_phase_change_lbmolps_tray"], dtype=float).reshape((N,))
        except Exception:
            eq_phase_change = None
    eq_phase_rate_guard_scale = None
    if "eq_phase_rate_guard_scale_tray" in diag:
        try:
            eq_phase_rate_guard_scale = np.asarray(diag["eq_phase_rate_guard_scale_tray"], dtype=float).reshape((N,))
        except Exception:
            eq_phase_rate_guard_scale = None
    eq_phase_rate_guard_limit = None
    if "eq_phase_rate_guard_limit_lbmolps_tray" in diag:
        try:
            eq_phase_rate_guard_limit = np.asarray(
                diag["eq_phase_rate_guard_limit_lbmolps_tray"], dtype=float
            ).reshape((N,))
        except Exception:
            eq_phase_rate_guard_limit = None
    eq_phase_weight = None
    if "eq_phase_holdup_guard_weight_tray" in diag:
        try:
            eq_phase_weight = np.asarray(diag["eq_phase_holdup_guard_weight_tray"], dtype=float).reshape((N,))
        except Exception:
            eq_phase_weight = None
    eq_phase_cap = None
    if "eq_phase_holdup_guard_cap_tray" in diag:
        try:
            eq_phase_cap = np.asarray(diag["eq_phase_holdup_guard_cap_tray"], dtype=float).reshape((N,))
        except Exception:
            eq_phase_cap = None
    eq_phase_energy_damping = None
    if "eq_phase_energy_damping_tray" in diag:
        try:
            eq_phase_energy_damping = np.asarray(diag["eq_phase_energy_damping_tray"], dtype=float).reshape((N,))
        except Exception:
            eq_phase_energy_damping = None
    T_bubble_target = None
    if "T_bubble_target_F_tray" in diag:
        try:
            T_bubble_target = np.asarray(diag["T_bubble_target_F_tray"], dtype=float).reshape((N,))
        except Exception:
            T_bubble_target = None
    HL_tray = None
    HV_tray = None
    q_phase_latent = None
    dT_energy_raw = None
    dT_mode_correction = None
    dT_phase_latent_equiv = None
    tray_heat_capacity = None
    tray_effective_heat_capacity = None
    tray_temperature_guard_active = None
    tray_temperature_rate_limit = None
    K_state_tray = None
    K_thermo_tray = None
    K_ratio_tray = None
    if "HL_BTU_lbmol_tray" in diag:
        try:
            HL_tray = np.asarray(diag["HL_BTU_lbmol_tray"], dtype=float).reshape((N,))
        except Exception:
            HL_tray = None
    if "HV_BTU_lbmol_tray" in diag:
        try:
            HV_tray = np.asarray(diag["HV_BTU_lbmol_tray"], dtype=float).reshape((N,))
        except Exception:
            HV_tray = None
    if "Q_phase_relax_latent_BTUps_tray" in diag:
        try:
            q_phase_latent = np.asarray(diag["Q_phase_relax_latent_BTUps_tray"], dtype=float).reshape((N,))
        except Exception:
            q_phase_latent = None
    if "dT_energy_raw_F_per_s_tray" in diag:
        try:
            dT_energy_raw = np.asarray(diag["dT_energy_raw_F_per_s_tray"], dtype=float).reshape((N,))
        except Exception:
            dT_energy_raw = None
    if "dT_mode_correction_F_per_s_tray" in diag:
        try:
            dT_mode_correction = np.asarray(diag["dT_mode_correction_F_per_s_tray"], dtype=float).reshape((N,))
        except Exception:
            dT_mode_correction = None
    if "dT_phase_latent_equiv_F_per_s_tray" in diag:
        try:
            dT_phase_latent_equiv = np.asarray(diag["dT_phase_latent_equiv_F_per_s_tray"], dtype=float).reshape((N,))
        except Exception:
            dT_phase_latent_equiv = None
    if "tray_heat_capacity_BTU_per_F_tray" in diag:
        try:
            tray_heat_capacity = np.asarray(diag["tray_heat_capacity_BTU_per_F_tray"], dtype=float).reshape((N,))
        except Exception:
            tray_heat_capacity = None
    if "tray_effective_heat_capacity_BTU_per_F_tray" in diag:
        try:
            tray_effective_heat_capacity = np.asarray(
                diag["tray_effective_heat_capacity_BTU_per_F_tray"], dtype=float
            ).reshape((N,))
        except Exception:
            tray_effective_heat_capacity = None
    if "tray_temperature_guard_active_tray" in diag:
        try:
            tray_temperature_guard_active = np.asarray(
                diag["tray_temperature_guard_active_tray"], dtype=float
            ).reshape((N,))
        except Exception:
            tray_temperature_guard_active = None
    if "tray_temperature_rate_limit_F_per_s_tray" in diag:
        try:
            tray_temperature_rate_limit = np.asarray(
                diag["tray_temperature_rate_limit_F_per_s_tray"], dtype=float
            ).reshape((N,))
        except Exception:
            tray_temperature_rate_limit = None
    if "K_state_y_over_x_tray" in diag:
        try:
            K_state_tray = np.asarray(diag["K_state_y_over_x_tray"], dtype=float).reshape((N, Nc))
        except Exception:
            K_state_tray = None
    if "K_tray" in diag:
        try:
            K_thermo_tray = np.asarray(diag["K_tray"], dtype=float).reshape((N, Nc))
        except Exception:
            K_thermo_tray = None
    if "K_state_over_K_thermo_tray" in diag:
        try:
            K_ratio_tray = np.asarray(diag["K_state_over_K_thermo_tray"], dtype=float).reshape((N, Nc))
        except Exception:
            K_ratio_tray = None
    L_out_used = None
    reflux_ratio = None
    if "L_out_lbmolph" in diag:
        try:
            L_out_used = np.asarray(diag["L_out_lbmolph"], dtype=float).reshape((N,))
        except Exception:
            L_out_used = None
    if L_out_used is not None and dist_tag.flow_lbmolph is not None:
        try:
            D_flow = float(dist_tag.flow_lbmolph)
            if np.isfinite(D_flow) and D_flow > 0.0 and np.isfinite(L_out_used[0]):
                reflux_ratio = float(L_out_used[0]) / D_flow
        except Exception:
            reflux_ratio = None

    T = _tray_temperature_F(col, layout, y, include_temperature)
    # Condenser/distillate temperature is stage-1 tray temperature.
    T_distillate = float(T[0])
    Q_cond_calc_BTUph = np.nan
    if "Q_cond_calc_BTUph" in diag:
        try:
            Q_cond_calc_BTUph = float(np.asarray(diag["Q_cond_calc_BTUph"], dtype=float).reshape((-1,))[0])
        except Exception:
            Q_cond_calc_BTUph = np.nan
    Q_cond_used_BTUph = np.nan
    if "Q_cond_used_BTUph" in diag:
        try:
            Q_cond_used_BTUph = float(np.asarray(diag["Q_cond_used_BTUph"], dtype=float).reshape((-1,))[0])
        except Exception:
            Q_cond_used_BTUph = np.nan
    Q_reb_used_BTUph = np.nan
    if "Q_reb_used_BTUph" in diag:
        try:
            Q_reb_used_BTUph = float(np.asarray(diag["Q_reb_used_BTUph"], dtype=float).reshape((-1,))[0])
        except Exception:
            Q_reb_used_BTUph = np.nan
    Q_cond_cmd_BTUph = np.nan
    if "Q_cond_cmd_BTUph" in diag:
        try:
            Q_cond_cmd_BTUph = float(np.asarray(diag["Q_cond_cmd_BTUph"], dtype=float).reshape((-1,))[0])
        except Exception:
            Q_cond_cmd_BTUph = np.nan
    P_top_drum_psia = np.nan
    if "P_top_drum_psia" in diag:
        try:
            P_top_drum_psia = float(np.asarray(diag["P_top_drum_psia"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_drum_psia = np.nan
    P_top_drum_psia_raw = np.nan
    if "P_top_drum_psia_raw" in diag:
        try:
            P_top_drum_psia_raw = float(np.asarray(diag["P_top_drum_psia_raw"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_drum_psia_raw = np.nan
    Z_top_drum_vapor = np.nan
    if "Z_top_drum_vapor" in diag:
        try:
            Z_top_drum_vapor = float(np.asarray(diag["Z_top_drum_vapor"], dtype=float).reshape((-1,))[0])
        except Exception:
            Z_top_drum_vapor = np.nan
    MV_top_drum_lbmol = np.nan
    if "MV_top_drum_lbmol" in diag:
        try:
            MV_top_drum_lbmol = float(np.asarray(diag["MV_top_drum_lbmol"], dtype=float).reshape((-1,))[0])
        except Exception:
            MV_top_drum_lbmol = np.nan
    V_condensed_in_lbmolph = np.nan
    if "V_condensed_in_lbmolph" in diag:
        try:
            V_condensed_in_lbmolph = float(np.asarray(diag["V_condensed_in_lbmolph"], dtype=float).reshape((-1,))[0])
        except Exception:
            V_condensed_in_lbmolph = np.nan
    V_to_top_drum_lbmolph = np.nan
    if "V_to_top_drum_lbmolph" in diag:
        try:
            V_to_top_drum_lbmolph = float(np.asarray(diag["V_to_top_drum_lbmolph"], dtype=float).reshape((-1,))[0])
        except Exception:
            V_to_top_drum_lbmolph = np.nan
    dP_stage2_to_top_drum_psia = np.nan
    if "dP_stage2_to_top_drum_psia" in diag:
        try:
            dP_stage2_to_top_drum_psia = float(
                np.asarray(diag["dP_stage2_to_top_drum_psia"], dtype=float).reshape((-1,))[0]
            )
        except Exception:
            dP_stage2_to_top_drum_psia = np.nan
    V_to_top_drum_pressure_gate_scale = np.nan
    if "V_to_top_drum_pressure_gate_scale" in diag:
        try:
            V_to_top_drum_pressure_gate_scale = float(
                np.asarray(diag["V_to_top_drum_pressure_gate_scale"], dtype=float).reshape((-1,))[0]
            )
        except Exception:
            V_to_top_drum_pressure_gate_scale = np.nan
    V_to_top_drum_blocked_lbmolph = np.nan
    if "V_to_top_drum_blocked_lbmolph" in diag:
        try:
            V_to_top_drum_blocked_lbmolph = float(
                np.asarray(diag["V_to_top_drum_blocked_lbmolph"], dtype=float).reshape((-1,))[0]
            )
        except Exception:
            V_to_top_drum_blocked_lbmolph = np.nan
    V_condensed_top_lbmolph = np.nan
    if "V_condensed_top_lbmolph" in diag:
        try:
            V_condensed_top_lbmolph = float(np.asarray(diag["V_condensed_top_lbmolph"], dtype=float).reshape((-1,))[0])
        except Exception:
            V_condensed_top_lbmolph = np.nan
    V_psv_top_lbmolph = np.nan
    if "V_psv_top_lbmolph" in diag:
        try:
            V_psv_top_lbmolph = float(np.asarray(diag["V_psv_top_lbmolph"], dtype=float).reshape((-1,))[0])
        except Exception:
            V_psv_top_lbmolph = np.nan
    PSV_open_flag = np.nan
    if "PSV_open_flag" in diag:
        try:
            PSV_open_flag = float(np.asarray(diag["PSV_open_flag"], dtype=float).reshape((-1,))[0])
        except Exception:
            PSV_open_flag = np.nan
    PSV_setpoint_psia = np.nan
    if "PSV_setpoint_psia" in diag:
        try:
            PSV_setpoint_psia = float(np.asarray(diag["PSV_setpoint_psia"], dtype=float).reshape((-1,))[0])
        except Exception:
            PSV_setpoint_psia = np.nan
    PSV_pv_psia = np.nan
    if "PSV_pv_psia" in diag:
        try:
            PSV_pv_psia = float(np.asarray(diag["PSV_pv_psia"], dtype=float).reshape((-1,))[0])
        except Exception:
            PSV_pv_psia = np.nan
    V_top_drum_vapor_ft3 = np.nan
    if "V_top_drum_vapor_ft3" in diag:
        try:
            V_top_drum_vapor_ft3 = float(np.asarray(diag["V_top_drum_vapor_ft3"], dtype=float).reshape((-1,))[0])
        except Exception:
            V_top_drum_vapor_ft3 = np.nan
    V_top_drum_liquid_ft3 = np.nan
    if "V_top_drum_liquid_ft3" in diag:
        try:
            V_top_drum_liquid_ft3 = float(np.asarray(diag["V_top_drum_liquid_ft3"], dtype=float).reshape((-1,))[0])
        except Exception:
            V_top_drum_liquid_ft3 = np.nan
    rho_top_drum_liq_lbmol_ft3 = np.nan
    if "rho_top_drum_liq_lbmol_ft3" in diag:
        try:
            rho_top_drum_liq_lbmol_ft3 = float(np.asarray(diag["rho_top_drum_liq_lbmol_ft3"], dtype=float).reshape((-1,))[0])
        except Exception:
            rho_top_drum_liq_lbmol_ft3 = np.nan
    Q_reb_cmd_BTUph = np.nan
    if "Q_reb_cmd_BTUph" in diag:
        try:
            Q_reb_cmd_BTUph = float(np.asarray(diag["Q_reb_cmd_BTUph"], dtype=float).reshape((-1,))[0])
        except Exception:
            Q_reb_cmd_BTUph = np.nan
    P_top_anchor_cmd_psia = np.nan
    if "P_top_anchor_cmd_psia" in diag:
        try:
            P_top_anchor_cmd_psia = float(np.asarray(diag["P_top_anchor_cmd_psia"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_anchor_cmd_psia = np.nan
    P_top_ctrl_pv_raw_psia = np.nan
    if "P_top_ctrl_pv_raw_psia" in diag:
        try:
            P_top_ctrl_pv_raw_psia = float(np.asarray(diag["P_top_ctrl_pv_raw_psia"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_pv_raw_psia = np.nan
    P_top_ctrl_pv_filt_psia = np.nan
    if "P_top_ctrl_pv_filt_psia" in diag:
        try:
            P_top_ctrl_pv_filt_psia = float(np.asarray(diag["P_top_ctrl_pv_filt_psia"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_pv_filt_psia = np.nan
    P_top_ctrl_gain_scale = np.nan
    if "P_top_ctrl_gain_scale" in diag:
        try:
            P_top_ctrl_gain_scale = float(np.asarray(diag["P_top_ctrl_gain_scale"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_gain_scale = np.nan
    P_top_ctrl_energy_resid_abs_BTUps = np.nan
    if "P_top_ctrl_energy_resid_abs_BTUps" in diag:
        try:
            P_top_ctrl_energy_resid_abs_BTUps = float(
                np.asarray(diag["P_top_ctrl_energy_resid_abs_BTUps"], dtype=float).reshape((-1,))[0]
            )
        except Exception:
            P_top_ctrl_energy_resid_abs_BTUps = np.nan
    P_top_ctrl_err_psia = np.nan
    if "P_top_ctrl_err_psia" in diag:
        try:
            P_top_ctrl_err_psia = float(np.asarray(diag["P_top_ctrl_err_psia"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_err_psia = np.nan
    P_top_ctrl_p_term = np.nan
    if "P_top_ctrl_p_term" in diag:
        try:
            P_top_ctrl_p_term = float(np.asarray(diag["P_top_ctrl_p_term"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_p_term = np.nan
    P_top_ctrl_i_term = np.nan
    if "P_top_ctrl_i_term" in diag:
        try:
            P_top_ctrl_i_term = float(np.asarray(diag["P_top_ctrl_i_term"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_i_term = np.nan
    P_top_ctrl_u_unclamped_BTUph = np.nan
    if "P_top_ctrl_u_unclamped_BTUph" in diag:
        try:
            P_top_ctrl_u_unclamped_BTUph = float(
                np.asarray(diag["P_top_ctrl_u_unclamped_BTUph"], dtype=float).reshape((-1,))[0]
            )
        except Exception:
            P_top_ctrl_u_unclamped_BTUph = np.nan
    P_top_ctrl_sat_hi = np.nan
    if "P_top_ctrl_sat_hi" in diag:
        try:
            P_top_ctrl_sat_hi = float(np.asarray(diag["P_top_ctrl_sat_hi"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_sat_hi = np.nan
    P_top_ctrl_sat_lo = np.nan
    if "P_top_ctrl_sat_lo" in diag:
        try:
            P_top_ctrl_sat_lo = float(np.asarray(diag["P_top_ctrl_sat_lo"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_sat_lo = np.nan
    P_top_ctrl_allow_int = np.nan
    if "P_top_ctrl_allow_int" in diag:
        try:
            P_top_ctrl_allow_int = float(np.asarray(diag["P_top_ctrl_allow_int"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_allow_int = np.nan
    P_top_ctrl_err_psia = np.nan
    if "P_top_ctrl_err_psia" in diag:
        try:
            P_top_ctrl_err_psia = float(np.asarray(diag["P_top_ctrl_err_psia"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_err_psia = np.nan
    P_top_ctrl_p_term = np.nan
    if "P_top_ctrl_p_term" in diag:
        try:
            P_top_ctrl_p_term = float(np.asarray(diag["P_top_ctrl_p_term"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_p_term = np.nan
    P_top_ctrl_i_term = np.nan
    if "P_top_ctrl_i_term" in diag:
        try:
            P_top_ctrl_i_term = float(np.asarray(diag["P_top_ctrl_i_term"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_i_term = np.nan
    P_top_ctrl_u_unclamped_BTUph = np.nan
    if "P_top_ctrl_u_unclamped_BTUph" in diag:
        try:
            P_top_ctrl_u_unclamped_BTUph = float(
                np.asarray(diag["P_top_ctrl_u_unclamped_BTUph"], dtype=float).reshape((-1,))[0]
            )
        except Exception:
            P_top_ctrl_u_unclamped_BTUph = np.nan
    P_top_ctrl_sat_hi = np.nan
    if "P_top_ctrl_sat_hi" in diag:
        try:
            P_top_ctrl_sat_hi = float(np.asarray(diag["P_top_ctrl_sat_hi"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_sat_hi = np.nan
    P_top_ctrl_sat_lo = np.nan
    if "P_top_ctrl_sat_lo" in diag:
        try:
            P_top_ctrl_sat_lo = float(np.asarray(diag["P_top_ctrl_sat_lo"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_sat_lo = np.nan
    P_top_ctrl_allow_int = np.nan
    if "P_top_ctrl_allow_int" in diag:
        try:
            P_top_ctrl_allow_int = float(np.asarray(diag["P_top_ctrl_allow_int"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_allow_int = np.nan
    P_top_ctrl_pv_raw_psia = np.nan
    if "P_top_ctrl_pv_raw_psia" in diag:
        try:
            P_top_ctrl_pv_raw_psia = float(np.asarray(diag["P_top_ctrl_pv_raw_psia"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_pv_raw_psia = np.nan
    P_top_ctrl_pv_filt_psia = np.nan
    if "P_top_ctrl_pv_filt_psia" in diag:
        try:
            P_top_ctrl_pv_filt_psia = float(np.asarray(diag["P_top_ctrl_pv_filt_psia"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_pv_filt_psia = np.nan
    P_top_ctrl_gain_scale = np.nan
    if "P_top_ctrl_gain_scale" in diag:
        try:
            P_top_ctrl_gain_scale = float(np.asarray(diag["P_top_ctrl_gain_scale"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_gain_scale = np.nan
    P_top_ctrl_energy_resid_abs_BTUps = np.nan
    if "P_top_ctrl_energy_resid_abs_BTUps" in diag:
        try:
            P_top_ctrl_energy_resid_abs_BTUps = float(
                np.asarray(diag["P_top_ctrl_energy_resid_abs_BTUps"], dtype=float).reshape((-1,))[0]
            )
        except Exception:
            P_top_ctrl_energy_resid_abs_BTUps = np.nan
    P_top_ctrl_err_psia = np.nan
    if "P_top_ctrl_err_psia" in diag:
        try:
            P_top_ctrl_err_psia = float(np.asarray(diag["P_top_ctrl_err_psia"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_err_psia = np.nan
    P_top_ctrl_p_term = np.nan
    if "P_top_ctrl_p_term" in diag:
        try:
            P_top_ctrl_p_term = float(np.asarray(diag["P_top_ctrl_p_term"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_p_term = np.nan
    P_top_ctrl_i_term = np.nan
    if "P_top_ctrl_i_term" in diag:
        try:
            P_top_ctrl_i_term = float(np.asarray(diag["P_top_ctrl_i_term"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_i_term = np.nan
    P_top_ctrl_u_unclamped_BTUph = np.nan
    if "P_top_ctrl_u_unclamped_BTUph" in diag:
        try:
            P_top_ctrl_u_unclamped_BTUph = float(
                np.asarray(diag["P_top_ctrl_u_unclamped_BTUph"], dtype=float).reshape((-1,))[0]
            )
        except Exception:
            P_top_ctrl_u_unclamped_BTUph = np.nan
    P_top_ctrl_sat_hi = np.nan
    if "P_top_ctrl_sat_hi" in diag:
        try:
            P_top_ctrl_sat_hi = float(np.asarray(diag["P_top_ctrl_sat_hi"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_sat_hi = np.nan
    P_top_ctrl_sat_lo = np.nan
    if "P_top_ctrl_sat_lo" in diag:
        try:
            P_top_ctrl_sat_lo = float(np.asarray(diag["P_top_ctrl_sat_lo"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_sat_lo = np.nan
    P_top_ctrl_allow_int = np.nan
    if "P_top_ctrl_allow_int" in diag:
        try:
            P_top_ctrl_allow_int = float(np.asarray(diag["P_top_ctrl_allow_int"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_allow_int = np.nan
    P_top_ctrl_err_psia = np.nan
    if "P_top_ctrl_err_psia" in diag:
        try:
            P_top_ctrl_err_psia = float(np.asarray(diag["P_top_ctrl_err_psia"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_err_psia = np.nan
    P_top_ctrl_p_term = np.nan
    if "P_top_ctrl_p_term" in diag:
        try:
            P_top_ctrl_p_term = float(np.asarray(diag["P_top_ctrl_p_term"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_p_term = np.nan
    P_top_ctrl_i_term = np.nan
    if "P_top_ctrl_i_term" in diag:
        try:
            P_top_ctrl_i_term = float(np.asarray(diag["P_top_ctrl_i_term"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_i_term = np.nan
    P_top_ctrl_u_unclamped_BTUph = np.nan
    if "P_top_ctrl_u_unclamped_BTUph" in diag:
        try:
            P_top_ctrl_u_unclamped_BTUph = float(
                np.asarray(diag["P_top_ctrl_u_unclamped_BTUph"], dtype=float).reshape((-1,))[0]
            )
        except Exception:
            P_top_ctrl_u_unclamped_BTUph = np.nan
    P_top_ctrl_sat_hi = np.nan
    if "P_top_ctrl_sat_hi" in diag:
        try:
            P_top_ctrl_sat_hi = float(np.asarray(diag["P_top_ctrl_sat_hi"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_sat_hi = np.nan
    P_top_ctrl_sat_lo = np.nan
    if "P_top_ctrl_sat_lo" in diag:
        try:
            P_top_ctrl_sat_lo = float(np.asarray(diag["P_top_ctrl_sat_lo"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_sat_lo = np.nan
    P_top_ctrl_allow_int = np.nan
    if "P_top_ctrl_allow_int" in diag:
        try:
            P_top_ctrl_allow_int = float(np.asarray(diag["P_top_ctrl_allow_int"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_allow_int = np.nan
    P_top_ctrl_sat_lo = np.nan
    if "P_top_ctrl_sat_lo" in diag:
        try:
            P_top_ctrl_sat_lo = float(np.asarray(diag["P_top_ctrl_sat_lo"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_sat_lo = np.nan
    P_top_ctrl_allow_int = np.nan
    if "P_top_ctrl_allow_int" in diag:
        try:
            P_top_ctrl_allow_int = float(np.asarray(diag["P_top_ctrl_allow_int"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_allow_int = np.nan
    xD_comp_sp = np.nan
    if "xD_comp_sp" in diag:
        try:
            xD_comp_sp = float(np.asarray(diag["xD_comp_sp"], dtype=float).reshape((-1,))[0])
        except Exception:
            xD_comp_sp = np.nan
    xD_comp_pv = np.nan
    if "xD_comp_pv" in diag:
        try:
            xD_comp_pv = float(np.asarray(diag["xD_comp_pv"], dtype=float).reshape((-1,))[0])
        except Exception:
            xD_comp_pv = np.nan
    xD_comp_err = np.nan
    if "xD_comp_err" in diag:
        try:
            xD_comp_err = float(np.asarray(diag["xD_comp_err"], dtype=float).reshape((-1,))[0])
        except Exception:
            xD_comp_err = np.nan
    RR_comp_cmd = np.nan
    if "RR_comp_cmd" in diag:
        try:
            RR_comp_cmd = float(np.asarray(diag["RR_comp_cmd"], dtype=float).reshape((-1,))[0])
        except Exception:
            RR_comp_cmd = np.nan
    Reflux_cmd_lbmolph = np.nan
    if "Reflux_cmd_lbmolph" in diag:
        try:
            Reflux_cmd_lbmolph = float(np.asarray(diag["Reflux_cmd_lbmolph"], dtype=float).reshape((-1,))[0])
        except Exception:
            Reflux_cmd_lbmolph = np.nan
    Reflux_cmd_unclamped_lbmolph = np.nan
    if "Reflux_cmd_unclamped_lbmolph" in diag:
        try:
            Reflux_cmd_unclamped_lbmolph = float(np.asarray(diag["Reflux_cmd_unclamped_lbmolph"], dtype=float).reshape((-1,))[0])
        except Exception:
            Reflux_cmd_unclamped_lbmolph = np.nan
    Reflux_cmd_active_max_lbmolph = np.nan
    if "Reflux_cmd_active_max_lbmolph" in diag:
        try:
            Reflux_cmd_active_max_lbmolph = float(np.asarray(diag["Reflux_cmd_active_max_lbmolph"], dtype=float).reshape((-1,))[0])
        except Exception:
            Reflux_cmd_active_max_lbmolph = np.nan
    Reflux_cap_active_flag = np.nan
    if "Reflux_cap_active_flag" in diag:
        try:
            Reflux_cap_active_flag = float(np.asarray(diag["Reflux_cap_active_flag"], dtype=float).reshape((-1,))[0])
        except Exception:
            Reflux_cap_active_flag = np.nan
    xB_comp_sp = np.nan
    if "xB_comp_sp" in diag:
        try:
            xB_comp_sp = float(np.asarray(diag["xB_comp_sp"], dtype=float).reshape((-1,))[0])
        except Exception:
            xB_comp_sp = np.nan
    xB_comp_pv = np.nan
    if "xB_comp_pv" in diag:
        try:
            xB_comp_pv = float(np.asarray(diag["xB_comp_pv"], dtype=float).reshape((-1,))[0])
        except Exception:
            xB_comp_pv = np.nan
    Boilup_cmd_lbmolph = np.nan
    if "Boilup_cmd_lbmolph" in diag:
        try:
            Boilup_cmd_lbmolph = float(np.asarray(diag["Boilup_cmd_lbmolph"], dtype=float).reshape((-1,))[0])
        except Exception:
            Boilup_cmd_lbmolph = np.nan
    M_total_lbmol = np.nan
    if "M_total_lbmol" in diag:
        try:
            M_total_lbmol = float(np.asarray(diag["M_total_lbmol"], dtype=float).reshape((-1,))[0])
        except Exception:
            M_total_lbmol = np.nan
    dM_total_dt_lbmolph = np.nan
    if "dM_total_dt_lbmolph" in diag:
        try:
            dM_total_dt_lbmolph = float(np.asarray(diag["dM_total_dt_lbmolph"], dtype=float).reshape((-1,))[0])
        except Exception:
            dM_total_dt_lbmolph = np.nan
    net_F_minus_D_minus_B_lbmolph = np.nan
    if "net_F_minus_D_minus_B_lbmolph" in diag:
        try:
            net_F_minus_D_minus_B_lbmolph = float(np.asarray(diag["net_F_minus_D_minus_B_lbmolph"], dtype=float).reshape((-1,))[0])
        except Exception:
            net_F_minus_D_minus_B_lbmolph = np.nan
    global_mass_closure_error_lbmolph = np.nan
    if "global_mass_closure_error_lbmolph" in diag:
        try:
            global_mass_closure_error_lbmolph = float(np.asarray(diag["global_mass_closure_error_lbmolph"], dtype=float).reshape((-1,))[0])
        except Exception:
            global_mass_closure_error_lbmolph = np.nan
    global_mass_closure_cum_lbmol = np.nan
    if "global_mass_closure_cum_lbmol" in diag:
        try:
            global_mass_closure_cum_lbmol = float(np.asarray(diag["global_mass_closure_cum_lbmol"], dtype=float).reshape((-1,))[0])
        except Exception:
            global_mass_closure_cum_lbmol = np.nan
    stage_mass_resid_sum_lbmolps = np.nan
    if "stage_mass_resid_sum_lbmolps" in diag:
        try:
            stage_mass_resid_sum_lbmolps = float(np.asarray(diag["stage_mass_resid_sum_lbmolps"], dtype=float).reshape((-1,))[0])
        except Exception:
            stage_mass_resid_sum_lbmolps = np.nan
    pv_inner_iter_count = np.nan
    if "pv_inner_iter_count" in diag:
        try:
            pv_inner_iter_count = float(np.asarray(diag["pv_inner_iter_count"], dtype=float).reshape((-1,))[0])
        except Exception:
            pv_inner_iter_count = np.nan
    pv_inner_converged = np.nan
    if "pv_inner_converged" in diag:
        try:
            pv_inner_converged = float(np.asarray(diag["pv_inner_converged"], dtype=float).reshape((-1,))[0])
        except Exception:
            pv_inner_converged = np.nan
    pv_inner_dp_max_psia = np.nan
    if "pv_inner_dp_max_psia" in diag:
        try:
            pv_inner_dp_max_psia = float(np.asarray(diag["pv_inner_dp_max_psia"], dtype=float).reshape((-1,))[0])
        except Exception:
            pv_inner_dp_max_psia = np.nan
    pv_inner_dv_max_lbmolph = np.nan
    if "pv_inner_dv_max_lbmolph" in diag:
        try:
            pv_inner_dv_max_lbmolph = float(np.asarray(diag["pv_inner_dv_max_lbmolph"], dtype=float).reshape((-1,))[0])
        except Exception:
            pv_inner_dv_max_lbmolph = np.nan
    T_sump = None
    if layout.include_bottom and "bottom_T_f" in u:
        try:
            T_sump = float(u["bottom_T_f"][0])
        except Exception:
            T_sump = None

    # Prefer diag-provided P_psia_diag if present; else compute from PV
    if "P_psia_diag" in diag:
        P_diag = np.asarray(diag["P_psia_diag"], dtype=float).reshape((N,))
    else:
        P_diag = _pressure_diag_psia(col, volume_model, T, MV, Z)

    # Condenser stage report pin
    if N >= 1:
        try:
            if np.isfinite(P_spec[0]):
                P_diag[0] = float(P_spec[0])
        except Exception:
            pass

    rows: List[Dict[str, Any]] = []
    for i in range(N):
        i1 = int(i + 1)

        r: Dict[str, Any] = {
            "wall_clock_iso": wall_clock_iso,
            "wall_elapsed_s": float(wall_elapsed_s),
            "time_s": float(t_s),
            "stage": i1,
            "node_type": "stage",
            "T_F": float(T[i]),
            "P_psia_hyd": float(P_hyd[i]) if P_hyd is not None and np.isfinite(P_hyd[i]) else np.nan,
            "L_out_used_lbmolph": (
                float(L_out_used[i]) if L_out_used is not None and np.isfinite(L_out_used[i]) else np.nan
            ),
            "L_out_hyd_lbmolph": float(L_out_hyd[i]) if L_out_hyd is not None and np.isfinite(L_out_hyd[i]) else np.nan,
            "V_out_lbmolph": float(V_out_calc[i]) if V_out_calc is not None and np.isfinite(V_out_calc[i]) else np.nan,
            "hydraulic_dp_raw_psia": float(hyd_dp_raw[i]) if hyd_dp_raw is not None and np.isfinite(hyd_dp_raw[i]) else np.nan,
            "hydraulic_dp_dry_raw_psia": float(hyd_dp_dry_raw[i]) if hyd_dp_dry_raw is not None and np.isfinite(hyd_dp_dry_raw[i]) else np.nan,
            "hydraulic_dp_liq_raw_psia": float(hyd_dp_liq_raw[i]) if hyd_dp_liq_raw is not None and np.isfinite(hyd_dp_liq_raw[i]) else np.nan,
            "hydraulic_dp_used_psia": float(hyd_dp_used[i]) if hyd_dp_used is not None and np.isfinite(hyd_dp_used[i]) else np.nan,
            "vflow_energy_ok": float(vflow_ok[i]) if vflow_ok is not None and np.isfinite(vflow_ok[i]) else np.nan,
            "vflow_energy_denom_BTU_per_lbmol": float(vflow_denom[i]) if vflow_denom is not None and np.isfinite(vflow_denom[i]) else np.nan,
            "vflow_energy_calc_lbmolph": float(vflow_calc[i]) if vflow_calc is not None and np.isfinite(vflow_calc[i]) else np.nan,
            "vflow_energy_used_lbmolph": float(vflow_used[i]) if vflow_used is not None and np.isfinite(vflow_used[i]) else np.nan,
            "vflow_energy_clamped": float(vflow_clamped[i]) if vflow_clamped is not None and np.isfinite(vflow_clamped[i]) else np.nan,
            "vflow_energy_limit_hi_lbmolph": float(vflow_limit_hi[i]) if vflow_limit_hi is not None and np.isfinite(vflow_limit_hi[i]) else np.nan,
            "vflow_energy_limit_lo_lbmolph": float(vflow_limit_lo[i]) if vflow_limit_lo is not None and np.isfinite(vflow_limit_lo[i]) else np.nan,
            "vflow_relax_alpha": float(vflow_alpha[i]) if vflow_alpha is not None and np.isfinite(vflow_alpha[i]) else np.nan,
            "h_ow_ft": float(h_ow[i]) if h_ow is not None and np.isfinite(h_ow[i]) else np.nan,
            "ML_lbmol": float(ML[i]),
            "MV_lbmol": float(MV[i]),
            "dMLdt_transport_lbmolps": (
                float(dMLdt_transport[i]) if dMLdt_transport is not None and np.isfinite(dMLdt_transport[i]) else np.nan
            ),
            "dMLdt_phase_relax_lbmolps": (
                float(dMLdt_phase_relax[i]) if dMLdt_phase_relax is not None and np.isfinite(dMLdt_phase_relax[i]) else np.nan
            ),
            "dMLdt_total_lbmolps": (
                float(dMLdt_total[i]) if dMLdt_total is not None and np.isfinite(dMLdt_total[i]) else np.nan
            ),
            "dMLdt_feed_lbmolps": (
                float(dMLdt_feed[i]) if dMLdt_feed is not None and np.isfinite(dMLdt_feed[i]) else np.nan
            ),
            "beta_eq_tray": float(beta_eq_tray[i]) if beta_eq_tray is not None and np.isfinite(beta_eq_tray[i]) else np.nan,
            "eq_flash_mv_total_lbmol_tray": (
                float(eq_flash_mv_total[i])
                if eq_flash_mv_total is not None and np.isfinite(eq_flash_mv_total[i])
                else np.nan
            ),
            "eq_target_mv_total_lbmol_tray": (
                float(eq_target_mv_total[i])
                if eq_target_mv_total is not None and np.isfinite(eq_target_mv_total[i])
                else np.nan
            ),
            "eq_target_vapor_total_lbmol_tray": (
                float(eq_target_vapor_total[i])
                if eq_target_vapor_total is not None and np.isfinite(eq_target_vapor_total[i])
                else np.nan
            ),
            "eq_target_vapor_delta_lbmol_tray": (
                float(eq_target_vapor_delta[i])
                if eq_target_vapor_delta is not None and np.isfinite(eq_target_vapor_delta[i])
                else np.nan
            ),
            "eq_target_vapor_fraction_tray": (
                float(eq_target_vapor_fraction[i])
                if eq_target_vapor_fraction is not None and np.isfinite(eq_target_vapor_fraction[i])
                else np.nan
            ),
            "eq_current_vapor_fraction_tray": (
                float(eq_current_vapor_fraction[i])
                if eq_current_vapor_fraction is not None and np.isfinite(eq_current_vapor_fraction[i])
                else np.nan
            ),
            "eq_phase_change_lbmolps_tray": (
                float(eq_phase_change[i]) if eq_phase_change is not None and np.isfinite(eq_phase_change[i]) else np.nan
            ),
            "eq_phase_rate_guard_scale_tray": (
                float(eq_phase_rate_guard_scale[i])
                if eq_phase_rate_guard_scale is not None and np.isfinite(eq_phase_rate_guard_scale[i])
                else np.nan
            ),
            "eq_phase_rate_guard_limit_lbmolps_tray": (
                float(eq_phase_rate_guard_limit[i])
                if eq_phase_rate_guard_limit is not None and np.isfinite(eq_phase_rate_guard_limit[i])
                else np.nan
            ),
            "eq_phase_holdup_guard_weight_tray": (
                float(eq_phase_weight[i]) if eq_phase_weight is not None and np.isfinite(eq_phase_weight[i]) else np.nan
            ),
            "eq_phase_holdup_guard_cap_tray": (
                float(eq_phase_cap[i]) if eq_phase_cap is not None and np.isfinite(eq_phase_cap[i]) else np.nan
            ),
            "eq_phase_energy_damping_tray": (
                float(eq_phase_energy_damping[i])
                if eq_phase_energy_damping is not None and np.isfinite(eq_phase_energy_damping[i])
                else np.nan
            ),
            "T_bubble_target_F_tray": (
                float(T_bubble_target[i])
                if T_bubble_target is not None and np.isfinite(T_bubble_target[i])
                else np.nan
            ),
            "stage_mass_balance_resid_lbmolps": float(mass_resid[i]) if mass_resid is not None and np.isfinite(mass_resid[i]) else np.nan,
            "stage_energy_balance_resid_BTUps": float(energy_resid[i]) if energy_resid is not None and np.isfinite(energy_resid[i]) else np.nan,
            "Q_phase_relax_latent_BTUps": (
                float(q_phase_latent[i]) if q_phase_latent is not None and np.isfinite(q_phase_latent[i]) else np.nan
            ),
            "dT_energy_raw_F_per_s": (
                float(dT_energy_raw[i]) if dT_energy_raw is not None and np.isfinite(dT_energy_raw[i]) else np.nan
            ),
            "dT_mode_correction_F_per_s": (
                float(dT_mode_correction[i])
                if dT_mode_correction is not None and np.isfinite(dT_mode_correction[i])
                else np.nan
            ),
            "dT_phase_latent_equiv_F_per_s": (
                float(dT_phase_latent_equiv[i])
                if dT_phase_latent_equiv is not None and np.isfinite(dT_phase_latent_equiv[i])
                else np.nan
            ),
            "tray_heat_capacity_BTU_per_F": (
                float(tray_heat_capacity[i]) if tray_heat_capacity is not None and np.isfinite(tray_heat_capacity[i]) else np.nan
            ),
            "tray_effective_heat_capacity_BTU_per_F": (
                float(tray_effective_heat_capacity[i])
                if tray_effective_heat_capacity is not None and np.isfinite(tray_effective_heat_capacity[i])
                else np.nan
            ),
            "tray_temperature_guard_active_tray": (
                float(tray_temperature_guard_active[i])
                if tray_temperature_guard_active is not None and np.isfinite(tray_temperature_guard_active[i])
                else np.nan
            ),
            "tray_temperature_rate_limit_F_per_s_tray": (
                float(tray_temperature_rate_limit[i])
                if tray_temperature_rate_limit is not None and np.isfinite(tray_temperature_rate_limit[i])
                else np.nan
            ),
            "HL_BTU_lbmol_tray": float(HL_tray[i]) if HL_tray is not None and np.isfinite(HL_tray[i]) else np.nan,
            "HV_BTU_lbmol_tray": float(HV_tray[i]) if HV_tray is not None and np.isfinite(HV_tray[i]) else np.nan,
            "reflux_ratio": _stage_value(i1, 1, reflux_ratio),
            # New: stream flow columns placed on their stages
            "F_lbmolph": _stage_value(i1, feed_tag.stage_1based, feed_tag.flow_lbmolph),
            "D_lbmolph": _stage_value(i1, dist_tag.stage_1based, dist_tag.flow_lbmolph),
            "B_lbmolph": _stage_value(i1, bots_tag.stage_1based, bots_tag.flow_lbmolph),
            "Distillate_L_lbmol": _stage_value(i1, 1 if layout.include_top else None, top_L_total),
            "Bottoms_L_lbmol": _stage_value(i1, N if layout.include_bottom else None, bottom_L_total),
            "T_Distillate_F": _stage_value(i1, 1, T_distillate),
            "Q_cond_calc_BTUph": _stage_value(i1, 1, Q_cond_calc_BTUph),
            "Q_cond_used_BTUph": _stage_value(i1, 1, Q_cond_used_BTUph),
            "Q_reb_used_BTUph": _stage_value(i1, N, Q_reb_used_BTUph),
            "Q_cond_cmd_BTUph": _stage_value(i1, 1, Q_cond_cmd_BTUph),
            "P_top_drum_psia": _stage_value(i1, 1, P_top_drum_psia),
            "P_top_drum_psia_raw": _stage_value(i1, 1, P_top_drum_psia_raw),
            "Z_top_drum_vapor": _stage_value(i1, 1, Z_top_drum_vapor),
            "MV_top_drum_lbmol": _stage_value(i1, 1, MV_top_drum_lbmol),
            "V_condensed_in_lbmolph": _stage_value(i1, 1, V_condensed_in_lbmolph),
            "V_to_top_drum_lbmolph": _stage_value(i1, 1, V_to_top_drum_lbmolph),
            "dP_stage2_to_top_drum_psia": _stage_value(i1, 1, dP_stage2_to_top_drum_psia),
            "V_to_top_drum_pressure_gate_scale": _stage_value(i1, 1, V_to_top_drum_pressure_gate_scale),
            "V_to_top_drum_blocked_lbmolph": _stage_value(i1, 1, V_to_top_drum_blocked_lbmolph),
            "V_condensed_top_lbmolph": _stage_value(i1, 1, V_condensed_top_lbmolph),
            "V_psv_top_lbmolph": _stage_value(i1, 1, V_psv_top_lbmolph),
            "PSV_open_flag": _stage_value(i1, 1, PSV_open_flag),
            "PSV_setpoint_psia": _stage_value(i1, 1, PSV_setpoint_psia),
            "PSV_pv_psia": _stage_value(i1, 1, PSV_pv_psia),
            "V_top_drum_vapor_ft3": _stage_value(i1, 1, V_top_drum_vapor_ft3),
            "V_top_drum_liquid_ft3": _stage_value(i1, 1, V_top_drum_liquid_ft3),
            "rho_top_drum_liq_lbmol_ft3": _stage_value(i1, 1, rho_top_drum_liq_lbmol_ft3),
            "Q_reb_cmd_BTUph": _stage_value(i1, N, Q_reb_cmd_BTUph),
            "P_top_anchor_cmd_psia": _stage_value(i1, 1, P_top_anchor_cmd_psia),
            "P_top_ctrl_err_psia": _stage_value(i1, 1, P_top_ctrl_err_psia),
            "P_top_ctrl_p_term": _stage_value(i1, 1, P_top_ctrl_p_term),
            "P_top_ctrl_i_term": _stage_value(i1, 1, P_top_ctrl_i_term),
            "P_top_ctrl_u_unclamped_BTUph": _stage_value(i1, 1, P_top_ctrl_u_unclamped_BTUph),
            "P_top_ctrl_sat_hi": _stage_value(i1, 1, P_top_ctrl_sat_hi),
            "P_top_ctrl_sat_lo": _stage_value(i1, 1, P_top_ctrl_sat_lo),
            "P_top_ctrl_allow_int": _stage_value(i1, 1, P_top_ctrl_allow_int),
            "Top_level_ctrl_pv": _stage_value(i1, 1, _mapping_scalar(diag, "Top_level_ctrl_pv")),
            "Top_level_ctrl_sp": _stage_value(i1, 1, _mapping_scalar(diag, "Top_level_ctrl_sp")),
            "Top_level_ctrl_err": _stage_value(i1, 1, _mapping_scalar(diag, "Top_level_ctrl_err")),
            "Top_level_ctrl_p_term": _stage_value(i1, 1, _mapping_scalar(diag, "Top_level_ctrl_p_term")),
            "Top_level_ctrl_i_term": _stage_value(i1, 1, _mapping_scalar(diag, "Top_level_ctrl_i_term")),
            "Top_level_ctrl_pi_term_lbmolph": _stage_value(i1, 1, _mapping_scalar(diag, "Top_level_ctrl_pi_term_lbmolph")),
            "Top_level_ctrl_ff_term_lbmolph": _stage_value(i1, 1, _mapping_scalar(diag, "Top_level_ctrl_ff_term_lbmolph")),
            "Top_level_ctrl_u_unclamped_lbmolph": _stage_value(i1, 1, _mapping_scalar(diag, "Top_level_ctrl_u_unclamped_lbmolph")),
            "Top_level_ctrl_sat_hi": _stage_value(i1, 1, _mapping_scalar(diag, "Top_level_ctrl_sat_hi")),
            "Top_level_ctrl_sat_lo": _stage_value(i1, 1, _mapping_scalar(diag, "Top_level_ctrl_sat_lo")),
            "Top_level_ctrl_allow_int": _stage_value(i1, 1, _mapping_scalar(diag, "Top_level_ctrl_allow_int")),
            "Bottom_level_ctrl_pv": _stage_value(i1, N, _mapping_scalar(diag, "Bottom_level_ctrl_pv")),
            "Bottom_level_ctrl_sp": _stage_value(i1, N, _mapping_scalar(diag, "Bottom_level_ctrl_sp")),
            "Bottom_level_ctrl_err": _stage_value(i1, N, _mapping_scalar(diag, "Bottom_level_ctrl_err")),
            "Bottom_level_ctrl_p_term": _stage_value(i1, N, _mapping_scalar(diag, "Bottom_level_ctrl_p_term")),
            "Bottom_level_ctrl_i_term": _stage_value(i1, N, _mapping_scalar(diag, "Bottom_level_ctrl_i_term")),
            "Bottom_level_ctrl_u_unclamped_lbmolph": _stage_value(i1, N, _mapping_scalar(diag, "Bottom_level_ctrl_u_unclamped_lbmolph")),
            "Bottom_level_ctrl_sat_hi": _stage_value(i1, N, _mapping_scalar(diag, "Bottom_level_ctrl_sat_hi")),
            "Bottom_level_ctrl_sat_lo": _stage_value(i1, N, _mapping_scalar(diag, "Bottom_level_ctrl_sat_lo")),
            "Bottom_level_ctrl_allow_int": _stage_value(i1, N, _mapping_scalar(diag, "Bottom_level_ctrl_allow_int")),
            "xD_comp_sp": _stage_value(i1, 1, xD_comp_sp),
            "xD_comp_pv": _stage_value(i1, 1, xD_comp_pv),
            "xD_comp_err": _stage_value(i1, 1, xD_comp_err),
            "RR_comp_cmd": _stage_value(i1, 1, RR_comp_cmd),
            "Reflux_cmd_lbmolph": _stage_value(i1, 1, Reflux_cmd_lbmolph),
            "Reflux_cmd_unclamped_lbmolph": _stage_value(i1, 1, Reflux_cmd_unclamped_lbmolph),
            "Reflux_cmd_active_max_lbmolph": _stage_value(i1, 1, Reflux_cmd_active_max_lbmolph),
            "Reflux_cap_active_flag": _stage_value(i1, 1, Reflux_cap_active_flag),
            "xB_comp_sp": _stage_value(i1, N, xB_comp_sp),
            "xB_comp_pv": _stage_value(i1, N, xB_comp_pv),
            "Boilup_cmd_lbmolph": _stage_value(i1, N, Boilup_cmd_lbmolph),
            "M_total_lbmol": _stage_value(i1, 1, M_total_lbmol),
            "dM_total_dt_lbmolph": _stage_value(i1, 1, dM_total_dt_lbmolph),
            "net_F_minus_D_minus_B_lbmolph": _stage_value(i1, 1, net_F_minus_D_minus_B_lbmolph),
            "global_mass_closure_error_lbmolph": _stage_value(i1, 1, global_mass_closure_error_lbmolph),
            "global_mass_closure_cum_lbmol": _stage_value(i1, 1, global_mass_closure_cum_lbmol),
            "stage_mass_resid_sum_lbmolps": _stage_value(i1, 1, stage_mass_resid_sum_lbmolps),
            "T_sump_F": _stage_value(i1, N if layout.include_bottom else None, T_sump),
        }

        for k in range(Nc):
            label = comp_labels[k]
            r[f"x_{label}"] = float(x[i, k])
            r[f"y_{label}"] = float(yv[i, k])
            r[f"x_eq_{label}"] = (
                float(x_eq_tray[i, k]) if x_eq_tray is not None and np.isfinite(x_eq_tray[i, k]) else np.nan
            )
            r[f"y_eq_{label}"] = (
                float(y_eq_tray[i, k]) if y_eq_tray is not None and np.isfinite(y_eq_tray[i, k]) else np.nan
            )
            r[f"y_target_{label}"] = (
                float(y_target_tray[i, k]) if y_target_tray is not None and np.isfinite(y_target_tray[i, k]) else np.nan
            )
            r[f"K_state_{label}"] = (
                float(K_state_tray[i, k])
                if K_state_tray is not None and np.isfinite(K_state_tray[i, k])
                else np.nan
            )
            r[f"K_thermo_{label}"] = (
                float(K_thermo_tray[i, k])
                if K_thermo_tray is not None and np.isfinite(K_thermo_tray[i, k])
                else np.nan
            )
            r[f"K_state_over_K_thermo_{label}"] = (
                float(K_ratio_tray[i, k])
                if K_ratio_tray is not None and np.isfinite(K_ratio_tray[i, k])
                else np.nan
            )
            if top_x is not None:
                r[f"Distillate_x_{label}"] = _stage_value(i1, 1, float(top_x[k]))
            if bottom_x is not None:
                r[f"Bottoms_x_{label}"] = _stage_value(i1, N, float(bottom_x[k]))

        rows.append(r)

    if rows:
        drum_fields = {
            "D_lbmolph",
            "Distillate_L_lbmol",
            "T_Distillate_F",
            "Q_cond_calc_BTUph",
            "Q_cond_used_BTUph",
            "Q_cond_cmd_BTUph",
            "P_top_drum_psia",
            "P_top_drum_psia_raw",
            "Z_top_drum_vapor",
            "MV_top_drum_lbmol",
            "V_condensed_in_lbmolph",
            "V_to_top_drum_lbmolph",
            "dP_stage2_to_top_drum_psia",
            "V_to_top_drum_pressure_gate_scale",
            "V_to_top_drum_blocked_lbmolph",
            "V_condensed_top_lbmolph",
            "V_psv_top_lbmolph",
            "PSV_open_flag",
            "PSV_setpoint_psia",
            "PSV_pv_psia",
            "V_top_drum_vapor_ft3",
            "V_top_drum_liquid_ft3",
            "rho_top_drum_liq_lbmol_ft3",
            "P_top_anchor_cmd_psia",
            "xD_comp_sp",
            "xD_comp_pv",
            "xD_comp_err",
            "RR_comp_cmd",
            "Reflux_cmd_lbmolph",
            "Reflux_cmd_unclamped_lbmolph",
            "Reflux_cmd_active_max_lbmolph",
            "Reflux_cap_active_flag",
        }
        sump_fields = {
            "B_lbmolph",
            "Bottoms_L_lbmol",
            "Q_reb_used_BTUph",
            "Q_reb_cmd_BTUph",
            "xB_comp_sp",
            "xB_comp_pv",
            "Boilup_cmd_lbmolph",
            "T_sump_F",
        }
        for label in comp_labels:
            drum_fields.add(f"Distillate_x_{label}")
            sump_fields.add(f"Bottoms_x_{label}")

        moved_fields = set(drum_fields) | set(sump_fields)
        top_stage_row = dict(rows[0])
        bottom_stage_row = dict(rows[-1])

        for r in rows:
            for key in moved_fields:
                if key in r:
                    r[key] = np.nan
        drum_row = {k: np.nan for k in top_stage_row.keys()}
        sump_row = {k: np.nan for k in top_stage_row.keys()}

        drum_row["wall_clock_iso"] = wall_clock_iso
        drum_row["wall_elapsed_s"] = float(wall_elapsed_s)
        drum_row["time_s"] = float(t_s)
        drum_row["stage"] = 0
        drum_row["node_type"] = "distillate_drum"
        for key in drum_fields:
            if key in top_stage_row:
                drum_row[key] = top_stage_row.get(key, np.nan)

        sump_row["wall_clock_iso"] = wall_clock_iso
        sump_row["wall_elapsed_s"] = float(wall_elapsed_s)
        sump_row["time_s"] = float(t_s)
        sump_row["stage"] = int(N + 1)
        sump_row["node_type"] = "bottoms_sump"
        for key in sump_fields:
            if key in bottom_stage_row:
                sump_row[key] = bottom_stage_row.get(key, np.nan)

        # Distillate and bottoms draws are inventories on drum/sump unit rows.
        drum_row["D_lbmolph"] = float(dist_tag.flow_lbmolph) if dist_tag.flow_lbmolph is not None else np.nan
        sump_row["B_lbmolph"] = float(bots_tag.flow_lbmolph) if bots_tag.flow_lbmolph is not None else np.nan

        rows = [drum_row, *rows, sump_row]

    return rows


def _summary_row(
    t_s: float,
    case: CaseData,
    col: ColumnSpec,
    layout: StateVectorLayout,
    y: np.ndarray,
    diag: Dict[str, np.ndarray],
    *,
    include_temperature: bool,
    volume_model: VolumeModel,
    wall_clock_iso: str,
    wall_elapsed_s: float,
    feed_tag: StreamTag,
    dist_tag: StreamTag,
    bots_tag: StreamTag,
    integrator_info: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    u = layout.unpack(y)
    N = col.n_stages
    Nc = col.n_components

    def _comp_suffix(name: str) -> str:
        out = "".join(ch if ch.isalnum() else "_" for ch in str(name).strip())
        out = out.strip("_")
        return out or "comp"

    comp_labels = [_comp_suffix(nm) for nm in getattr(col, "components_excel", [f"c{i+1}" for i in range(Nc)])]

    x = diag["x_tray"] if "x_tray" in diag else u["x_tray"]
    if "y_tray" in diag:
        yv = diag["y_tray"]
    elif "y_eq_thermo_tray" in diag:
        yv = diag["y_eq_thermo_tray"]
    elif hasattr(col, "y0"):
        yv = np.asarray(getattr(col, "y0"), dtype=float).reshape((N, Nc))
    else:
        yv = np.zeros((N, Nc), dtype=float)
    MV = u["MV_tot_tray"] if "MV_tot_tray" in u else np.zeros(N, dtype=float)

    P_spec = np.asarray(getattr(col, "P_psia", np.full(N, np.nan, dtype=float)), dtype=float).reshape((N,))

    Z_raw = diag.get("Z_tray", np.full(N, np.nan, dtype=float))
    Z = np.asarray(Z_raw, dtype=float).reshape((N,))
    Z = np.where(~np.isfinite(Z) | (Z <= 0.0), 1.0, Z)

    T = _tray_temperature_F(col, layout, y, include_temperature)
    # Condenser/distillate temperature is stage-1 tray temperature.
    T_distillate = float(T[0])
    Q_cond_calc_BTUph = np.nan
    if "Q_cond_calc_BTUph" in diag:
        try:
            Q_cond_calc_BTUph = float(np.asarray(diag["Q_cond_calc_BTUph"], dtype=float).reshape((-1,))[0])
        except Exception:
            Q_cond_calc_BTUph = np.nan
    Q_cond_used_BTUph = np.nan
    if "Q_cond_used_BTUph" in diag:
        try:
            Q_cond_used_BTUph = float(np.asarray(diag["Q_cond_used_BTUph"], dtype=float).reshape((-1,))[0])
        except Exception:
            Q_cond_used_BTUph = np.nan
    Q_reb_used_BTUph = np.nan
    if "Q_reb_used_BTUph" in diag:
        try:
            Q_reb_used_BTUph = float(np.asarray(diag["Q_reb_used_BTUph"], dtype=float).reshape((-1,))[0])
        except Exception:
            Q_reb_used_BTUph = np.nan
    Q_cond_cmd_BTUph = np.nan
    if "Q_cond_cmd_BTUph" in diag:
        try:
            Q_cond_cmd_BTUph = float(np.asarray(diag["Q_cond_cmd_BTUph"], dtype=float).reshape((-1,))[0])
        except Exception:
            Q_cond_cmd_BTUph = np.nan
    P_top_drum_psia = np.nan
    if "P_top_drum_psia" in diag:
        try:
            P_top_drum_psia = float(np.asarray(diag["P_top_drum_psia"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_drum_psia = np.nan
    P_top_drum_psia_raw = np.nan
    if "P_top_drum_psia_raw" in diag:
        try:
            P_top_drum_psia_raw = float(np.asarray(diag["P_top_drum_psia_raw"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_drum_psia_raw = np.nan
    Z_top_drum_vapor = np.nan
    if "Z_top_drum_vapor" in diag:
        try:
            Z_top_drum_vapor = float(np.asarray(diag["Z_top_drum_vapor"], dtype=float).reshape((-1,))[0])
        except Exception:
            Z_top_drum_vapor = np.nan
    MV_top_drum_lbmol = np.nan
    if "MV_top_drum_lbmol" in diag:
        try:
            MV_top_drum_lbmol = float(np.asarray(diag["MV_top_drum_lbmol"], dtype=float).reshape((-1,))[0])
        except Exception:
            MV_top_drum_lbmol = np.nan
    V_condensed_in_lbmolph = np.nan
    if "V_condensed_in_lbmolph" in diag:
        try:
            V_condensed_in_lbmolph = float(np.asarray(diag["V_condensed_in_lbmolph"], dtype=float).reshape((-1,))[0])
        except Exception:
            V_condensed_in_lbmolph = np.nan
    V_to_top_drum_lbmolph = np.nan
    if "V_to_top_drum_lbmolph" in diag:
        try:
            V_to_top_drum_lbmolph = float(np.asarray(diag["V_to_top_drum_lbmolph"], dtype=float).reshape((-1,))[0])
        except Exception:
            V_to_top_drum_lbmolph = np.nan
    dP_stage2_to_top_drum_psia = np.nan
    if "dP_stage2_to_top_drum_psia" in diag:
        try:
            dP_stage2_to_top_drum_psia = float(
                np.asarray(diag["dP_stage2_to_top_drum_psia"], dtype=float).reshape((-1,))[0]
            )
        except Exception:
            dP_stage2_to_top_drum_psia = np.nan
    V_to_top_drum_pressure_gate_scale = np.nan
    if "V_to_top_drum_pressure_gate_scale" in diag:
        try:
            V_to_top_drum_pressure_gate_scale = float(
                np.asarray(diag["V_to_top_drum_pressure_gate_scale"], dtype=float).reshape((-1,))[0]
            )
        except Exception:
            V_to_top_drum_pressure_gate_scale = np.nan
    V_to_top_drum_blocked_lbmolph = np.nan
    if "V_to_top_drum_blocked_lbmolph" in diag:
        try:
            V_to_top_drum_blocked_lbmolph = float(
                np.asarray(diag["V_to_top_drum_blocked_lbmolph"], dtype=float).reshape((-1,))[0]
            )
        except Exception:
            V_to_top_drum_blocked_lbmolph = np.nan
    V_condensed_top_lbmolph = np.nan
    if "V_condensed_top_lbmolph" in diag:
        try:
            V_condensed_top_lbmolph = float(np.asarray(diag["V_condensed_top_lbmolph"], dtype=float).reshape((-1,))[0])
        except Exception:
            V_condensed_top_lbmolph = np.nan
    V_psv_top_lbmolph = np.nan
    if "V_psv_top_lbmolph" in diag:
        try:
            V_psv_top_lbmolph = float(np.asarray(diag["V_psv_top_lbmolph"], dtype=float).reshape((-1,))[0])
        except Exception:
            V_psv_top_lbmolph = np.nan
    PSV_open_flag = np.nan
    if "PSV_open_flag" in diag:
        try:
            PSV_open_flag = float(np.asarray(diag["PSV_open_flag"], dtype=float).reshape((-1,))[0])
        except Exception:
            PSV_open_flag = np.nan
    PSV_setpoint_psia = np.nan
    if "PSV_setpoint_psia" in diag:
        try:
            PSV_setpoint_psia = float(np.asarray(diag["PSV_setpoint_psia"], dtype=float).reshape((-1,))[0])
        except Exception:
            PSV_setpoint_psia = np.nan
    PSV_pv_psia = np.nan
    if "PSV_pv_psia" in diag:
        try:
            PSV_pv_psia = float(np.asarray(diag["PSV_pv_psia"], dtype=float).reshape((-1,))[0])
        except Exception:
            PSV_pv_psia = np.nan
    V_top_drum_vapor_ft3 = np.nan
    if "V_top_drum_vapor_ft3" in diag:
        try:
            V_top_drum_vapor_ft3 = float(np.asarray(diag["V_top_drum_vapor_ft3"], dtype=float).reshape((-1,))[0])
        except Exception:
            V_top_drum_vapor_ft3 = np.nan
    V_top_drum_liquid_ft3 = np.nan
    if "V_top_drum_liquid_ft3" in diag:
        try:
            V_top_drum_liquid_ft3 = float(np.asarray(diag["V_top_drum_liquid_ft3"], dtype=float).reshape((-1,))[0])
        except Exception:
            V_top_drum_liquid_ft3 = np.nan
    rho_top_drum_liq_lbmol_ft3 = np.nan
    if "rho_top_drum_liq_lbmol_ft3" in diag:
        try:
            rho_top_drum_liq_lbmol_ft3 = float(np.asarray(diag["rho_top_drum_liq_lbmol_ft3"], dtype=float).reshape((-1,))[0])
        except Exception:
            rho_top_drum_liq_lbmol_ft3 = np.nan
    Q_reb_cmd_BTUph = np.nan
    if "Q_reb_cmd_BTUph" in diag:
        try:
            Q_reb_cmd_BTUph = float(np.asarray(diag["Q_reb_cmd_BTUph"], dtype=float).reshape((-1,))[0])
        except Exception:
            Q_reb_cmd_BTUph = np.nan
    P_top_anchor_cmd_psia = np.nan
    if "P_top_anchor_cmd_psia" in diag:
        try:
            P_top_anchor_cmd_psia = float(np.asarray(diag["P_top_anchor_cmd_psia"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_anchor_cmd_psia = np.nan
    P_top_ctrl_pv_raw_psia = np.nan
    if "P_top_ctrl_pv_raw_psia" in diag:
        try:
            P_top_ctrl_pv_raw_psia = float(np.asarray(diag["P_top_ctrl_pv_raw_psia"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_pv_raw_psia = np.nan
    P_top_ctrl_pv_filt_psia = np.nan
    if "P_top_ctrl_pv_filt_psia" in diag:
        try:
            P_top_ctrl_pv_filt_psia = float(np.asarray(diag["P_top_ctrl_pv_filt_psia"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_pv_filt_psia = np.nan
    P_top_ctrl_gain_scale = np.nan
    if "P_top_ctrl_gain_scale" in diag:
        try:
            P_top_ctrl_gain_scale = float(np.asarray(diag["P_top_ctrl_gain_scale"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_gain_scale = np.nan
    P_top_ctrl_energy_resid_abs_BTUps = np.nan
    if "P_top_ctrl_energy_resid_abs_BTUps" in diag:
        try:
            P_top_ctrl_energy_resid_abs_BTUps = float(
                np.asarray(diag["P_top_ctrl_energy_resid_abs_BTUps"], dtype=float).reshape((-1,))[0]
            )
        except Exception:
            P_top_ctrl_energy_resid_abs_BTUps = np.nan
    P_top_ctrl_err_psia = np.nan
    if "P_top_ctrl_err_psia" in diag:
        try:
            P_top_ctrl_err_psia = float(np.asarray(diag["P_top_ctrl_err_psia"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_err_psia = np.nan
    P_top_ctrl_p_term = np.nan
    if "P_top_ctrl_p_term" in diag:
        try:
            P_top_ctrl_p_term = float(np.asarray(diag["P_top_ctrl_p_term"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_p_term = np.nan
    P_top_ctrl_i_term = np.nan
    if "P_top_ctrl_i_term" in diag:
        try:
            P_top_ctrl_i_term = float(np.asarray(diag["P_top_ctrl_i_term"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_i_term = np.nan
    P_top_ctrl_u_unclamped_BTUph = np.nan
    if "P_top_ctrl_u_unclamped_BTUph" in diag:
        try:
            P_top_ctrl_u_unclamped_BTUph = float(
                np.asarray(diag["P_top_ctrl_u_unclamped_BTUph"], dtype=float).reshape((-1,))[0]
            )
        except Exception:
            P_top_ctrl_u_unclamped_BTUph = np.nan
    P_top_ctrl_sat_hi = np.nan
    if "P_top_ctrl_sat_hi" in diag:
        try:
            P_top_ctrl_sat_hi = float(np.asarray(diag["P_top_ctrl_sat_hi"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_sat_hi = np.nan
    P_top_ctrl_sat_lo = np.nan
    if "P_top_ctrl_sat_lo" in diag:
        try:
            P_top_ctrl_sat_lo = float(np.asarray(diag["P_top_ctrl_sat_lo"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_sat_lo = np.nan
    P_top_ctrl_allow_int = np.nan
    if "P_top_ctrl_allow_int" in diag:
        try:
            P_top_ctrl_allow_int = float(np.asarray(diag["P_top_ctrl_allow_int"], dtype=float).reshape((-1,))[0])
        except Exception:
            P_top_ctrl_allow_int = np.nan
    xD_comp_sp = np.nan
    if "xD_comp_sp" in diag:
        try:
            xD_comp_sp = float(np.asarray(diag["xD_comp_sp"], dtype=float).reshape((-1,))[0])
        except Exception:
            xD_comp_sp = np.nan
    xD_comp_pv = np.nan
    if "xD_comp_pv" in diag:
        try:
            xD_comp_pv = float(np.asarray(diag["xD_comp_pv"], dtype=float).reshape((-1,))[0])
        except Exception:
            xD_comp_pv = np.nan
    xD_comp_err = np.nan
    if "xD_comp_err" in diag:
        try:
            xD_comp_err = float(np.asarray(diag["xD_comp_err"], dtype=float).reshape((-1,))[0])
        except Exception:
            xD_comp_err = np.nan
    RR_comp_cmd = np.nan
    if "RR_comp_cmd" in diag:
        try:
            RR_comp_cmd = float(np.asarray(diag["RR_comp_cmd"], dtype=float).reshape((-1,))[0])
        except Exception:
            RR_comp_cmd = np.nan
    Reflux_cmd_lbmolph = np.nan
    if "Reflux_cmd_lbmolph" in diag:
        try:
            Reflux_cmd_lbmolph = float(np.asarray(diag["Reflux_cmd_lbmolph"], dtype=float).reshape((-1,))[0])
        except Exception:
            Reflux_cmd_lbmolph = np.nan
    Reflux_cmd_unclamped_lbmolph = np.nan
    if "Reflux_cmd_unclamped_lbmolph" in diag:
        try:
            Reflux_cmd_unclamped_lbmolph = float(np.asarray(diag["Reflux_cmd_unclamped_lbmolph"], dtype=float).reshape((-1,))[0])
        except Exception:
            Reflux_cmd_unclamped_lbmolph = np.nan
    Reflux_cmd_active_max_lbmolph = np.nan
    if "Reflux_cmd_active_max_lbmolph" in diag:
        try:
            Reflux_cmd_active_max_lbmolph = float(np.asarray(diag["Reflux_cmd_active_max_lbmolph"], dtype=float).reshape((-1,))[0])
        except Exception:
            Reflux_cmd_active_max_lbmolph = np.nan
    Reflux_cap_active_flag = np.nan
    if "Reflux_cap_active_flag" in diag:
        try:
            Reflux_cap_active_flag = float(np.asarray(diag["Reflux_cap_active_flag"], dtype=float).reshape((-1,))[0])
        except Exception:
            Reflux_cap_active_flag = np.nan
    xB_comp_sp = np.nan
    if "xB_comp_sp" in diag:
        try:
            xB_comp_sp = float(np.asarray(diag["xB_comp_sp"], dtype=float).reshape((-1,))[0])
        except Exception:
            xB_comp_sp = np.nan
    xB_comp_pv = np.nan
    if "xB_comp_pv" in diag:
        try:
            xB_comp_pv = float(np.asarray(diag["xB_comp_pv"], dtype=float).reshape((-1,))[0])
        except Exception:
            xB_comp_pv = np.nan
    Boilup_cmd_lbmolph = np.nan
    if "Boilup_cmd_lbmolph" in diag:
        try:
            Boilup_cmd_lbmolph = float(np.asarray(diag["Boilup_cmd_lbmolph"], dtype=float).reshape((-1,))[0])
        except Exception:
            Boilup_cmd_lbmolph = np.nan
    M_total_lbmol = np.nan
    if "M_total_lbmol" in diag:
        try:
            M_total_lbmol = float(np.asarray(diag["M_total_lbmol"], dtype=float).reshape((-1,))[0])
        except Exception:
            M_total_lbmol = np.nan
    dM_total_dt_lbmolph = np.nan
    if "dM_total_dt_lbmolph" in diag:
        try:
            dM_total_dt_lbmolph = float(np.asarray(diag["dM_total_dt_lbmolph"], dtype=float).reshape((-1,))[0])
        except Exception:
            dM_total_dt_lbmolph = np.nan
    net_F_minus_D_minus_B_lbmolph = np.nan
    if "net_F_minus_D_minus_B_lbmolph" in diag:
        try:
            net_F_minus_D_minus_B_lbmolph = float(np.asarray(diag["net_F_minus_D_minus_B_lbmolph"], dtype=float).reshape((-1,))[0])
        except Exception:
            net_F_minus_D_minus_B_lbmolph = np.nan
    global_mass_closure_error_lbmolph = np.nan
    if "global_mass_closure_error_lbmolph" in diag:
        try:
            global_mass_closure_error_lbmolph = float(np.asarray(diag["global_mass_closure_error_lbmolph"], dtype=float).reshape((-1,))[0])
        except Exception:
            global_mass_closure_error_lbmolph = np.nan
    global_mass_closure_cum_lbmol = np.nan
    if "global_mass_closure_cum_lbmol" in diag:
        try:
            global_mass_closure_cum_lbmol = float(np.asarray(diag["global_mass_closure_cum_lbmol"], dtype=float).reshape((-1,))[0])
        except Exception:
            global_mass_closure_cum_lbmol = np.nan
    stage_mass_resid_sum_lbmolps = np.nan
    if "stage_mass_resid_sum_lbmolps" in diag:
        try:
            stage_mass_resid_sum_lbmolps = float(np.asarray(diag["stage_mass_resid_sum_lbmolps"], dtype=float).reshape((-1,))[0])
        except Exception:
            stage_mass_resid_sum_lbmolps = np.nan
    pv_inner_iter_count = np.nan
    if "pv_inner_iter_count" in diag:
        try:
            pv_inner_iter_count = float(np.asarray(diag["pv_inner_iter_count"], dtype=float).reshape((-1,))[0])
        except Exception:
            pv_inner_iter_count = np.nan
    pv_inner_converged = np.nan
    if "pv_inner_converged" in diag:
        try:
            pv_inner_converged = float(np.asarray(diag["pv_inner_converged"], dtype=float).reshape((-1,))[0])
        except Exception:
            pv_inner_converged = np.nan
    pv_inner_dp_max_psia = np.nan
    if "pv_inner_dp_max_psia" in diag:
        try:
            pv_inner_dp_max_psia = float(np.asarray(diag["pv_inner_dp_max_psia"], dtype=float).reshape((-1,))[0])
        except Exception:
            pv_inner_dp_max_psia = np.nan
    pv_inner_dv_max_lbmolph = np.nan
    if "pv_inner_dv_max_lbmolph" in diag:
        try:
            pv_inner_dv_max_lbmolph = float(np.asarray(diag["pv_inner_dv_max_lbmolph"], dtype=float).reshape((-1,))[0])
        except Exception:
            pv_inner_dv_max_lbmolph = np.nan
    K_state_over_K_thermo_max_abs = np.nan
    K_state_minus_K_thermo_max_abs = np.nan
    if "K_state_over_K_thermo_tray" in diag:
        try:
            k_ratio = np.asarray(diag["K_state_over_K_thermo_tray"], dtype=float).reshape((N, Nc))
            k_ratio_f = k_ratio[np.isfinite(k_ratio)]
            if k_ratio_f.size > 0:
                K_state_over_K_thermo_max_abs = float(np.max(np.abs(k_ratio_f)))
        except Exception:
            K_state_over_K_thermo_max_abs = np.nan
    if "K_state_minus_K_thermo_tray" in diag:
        try:
            k_delta = np.asarray(diag["K_state_minus_K_thermo_tray"], dtype=float).reshape((N, Nc))
            k_delta_f = k_delta[np.isfinite(k_delta)]
            if k_delta_f.size > 0:
                K_state_minus_K_thermo_max_abs = float(np.max(np.abs(k_delta_f)))
        except Exception:
            K_state_minus_K_thermo_max_abs = np.nan
    steady_state_enabled = _mapping_scalar(diag, "steady_state_enabled")
    steady_state_flag = _mapping_scalar(diag, "steady_state_flag")
    steady_state_score = _mapping_scalar(diag, "steady_state_score")
    steady_state_active_criteria = _mapping_scalar(diag, "steady_state_active_criteria")
    ss_max_rel_state_rate_per_s = _mapping_scalar(diag, "ss_max_rel_state_rate_per_s")
    ss_rel_state_rate_stage_1based = _mapping_scalar(diag, "ss_rel_state_rate_stage_1based")
    ss_rel_state_rate_component_1based = _mapping_scalar(diag, "ss_rel_state_rate_component_1based")
    ss_rel_state_rate_state_key = ""
    if "ss_rel_state_rate_state_key" in diag:
        try:
            ss_rel_state_rate_state_key = str(np.asarray(diag["ss_rel_state_rate_state_key"], dtype=object).reshape((-1,))[0])
        except Exception:
            ss_rel_state_rate_state_key = ""
    ss_max_kpi_slope_per_s = _mapping_scalar(diag, "ss_max_kpi_slope_per_s")
    ss_max_mv_rate_per_s = _mapping_scalar(diag, "ss_max_mv_rate_per_s")
    ss_max_temp_rate_F_per_s = _mapping_scalar(diag, "ss_max_temp_rate_F_per_s")
    ss_max_sp_error = _mapping_scalar(diag, "ss_max_sp_error")
    ss_window_samples = _mapping_scalar(diag, "ss_window_samples")
    ss_window_sec = _mapping_scalar(diag, "ss_window_sec")
    ss_min_time_sec = _mapping_scalar(diag, "ss_min_time_sec")
    ss_tol_rel_state_rate_per_s = _mapping_scalar(diag, "ss_tol_rel_state_rate_per_s")
    ss_tol_kpi_slope_per_s = _mapping_scalar(diag, "ss_tol_kpi_slope_per_s")
    ss_tol_mv_rate_per_s = _mapping_scalar(diag, "ss_tol_mv_rate_per_s")
    ss_tol_temp_rate_F_per_s = _mapping_scalar(diag, "ss_tol_temp_rate_F_per_s")
    ss_tol_sp_error = _mapping_scalar(diag, "ss_tol_sp_error")
    ss_require_sp = _mapping_scalar(diag, "ss_require_sp")
    ss_rel_state_rate_component_name = ""
    if np.isfinite(ss_rel_state_rate_component_1based):
        try:
            cidx = int(round(float(ss_rel_state_rate_component_1based))) - 1
            if 0 <= cidx < len(col.components_excel):
                ss_rel_state_rate_component_name = str(col.components_excel[cidx])
        except Exception:
            ss_rel_state_rate_component_name = ""
    debug_freeze_tray_vapor_derivatives_active = _mapping_scalar(
        diag, "debug_freeze_tray_vapor_derivatives_active"
    )
    debug_max_orig_dmVdt = _mapping_scalar(diag, "debug_max_orig_dmVdt")
    debug_max_orig_dmVdt_rel_per_s = _mapping_scalar(diag, "debug_max_orig_dmVdt_rel_per_s")
    debug_worst_v_stage = _mapping_scalar(diag, "debug_worst_v_stage")
    debug_worst_v_comp = _mapping_scalar(diag, "debug_worst_v_comp")
    debug_worst_v_rel_stage = _mapping_scalar(diag, "debug_worst_v_rel_stage")
    debug_worst_v_rel_comp = _mapping_scalar(diag, "debug_worst_v_rel_comp")
    debug_total_v_cancellation_lbmolps = _mapping_scalar(diag, "debug_total_v_cancellation_lbmolps")
    debug_net_orig_dmVdt_lbmolps = _mapping_scalar(diag, "debug_net_orig_dmVdt_lbmolps")
    total_reflux_mode_active = _mapping_scalar(diag, "total_reflux_mode_active")
    total_reflux_actual_lbmolps = _mapping_scalar(diag, "total_reflux_actual_lbmolps")
    total_reflux_kickstart_lbmolps = _mapping_scalar(diag, "total_reflux_kickstart_lbmolps")
    total_reflux_used_lbmolps = _mapping_scalar(diag, "total_reflux_used_lbmolps")
    total_reflux_startup_factor = _mapping_scalar(diag, "total_reflux_startup_factor")
    total_reflux_reflux_startup_factor = _mapping_scalar(diag, "total_reflux_reflux_startup_factor")
    total_reflux_boundary_external_scale = _mapping_scalar(diag, "total_reflux_boundary_external_scale")
    total_reflux_boundary_closed_fraction = _mapping_scalar(diag, "total_reflux_boundary_closed_fraction")
    eq_relax_tau_effective_sec = _mapping_scalar(diag, "eq_relax_tau_effective_sec")
    startup_total_reflux_washout_active = _mapping_scalar(diag, "startup_total_reflux_washout_active")
    startup_sequence_phase_id = _mapping_scalar(diag, "startup_sequence_phase_id")
    startup_sequence_liquid_alpha_cmd = _mapping_scalar(diag, "startup_sequence_liquid_alpha_cmd")
    startup_sequence_vapor_homotopy_enabled = _mapping_scalar(diag, "startup_sequence_vapor_homotopy_enabled")
    startup_sequence_vapor_beta_cmd = _mapping_scalar(diag, "startup_sequence_vapor_beta_cmd")
    startup_sequence_vapor_guard_phase_id = _mapping_scalar(diag, "startup_sequence_vapor_guard_phase_id")
    startup_sequence_vapor_rel_rate_gate_per_s = _mapping_scalar(diag, "startup_sequence_vapor_rel_rate_gate_per_s")
    startup_sequence_rel_rate_prev_per_s = _mapping_scalar(diag, "startup_sequence_rel_rate_prev_per_s")
    vflow_homotopy_active = _mapping_scalar(diag, "vflow_homotopy_active")
    vflow_homotopy_beta = _mapping_scalar(diag, "vflow_homotopy_beta")

    integ = dict(integrator_info) if isinstance(integrator_info, dict) else {}
    integrator_requested_mode = str(integ.get("requested_mode", "")).strip()
    integrator_used_mode = str(integ.get("used_mode", "")).strip()
    integrator_fallback_reason = str(integ.get("fallback_reason", "")).strip()
    if not integrator_requested_mode:
        integrator_requested_mode = ""
    if not integrator_used_mode:
        integrator_used_mode = ""
    if not integrator_fallback_reason:
        integrator_fallback_reason = ""
    integrator_fallback_used = float("nan")
    if "fallback_used" in integ:
        integrator_fallback_used = 1.0 if bool(integ.get("fallback_used", False)) else 0.0
    integrator_nfev = _mapping_scalar(integ, "nfev")
    ida_iter_max = _mapping_scalar(integ, "ida_iter_max")
    ida_iter_mean = _mapping_scalar(integ, "ida_iter_mean")
    ida_converged = _mapping_scalar(integ, "ida_converged")
    ida_last_err = _mapping_scalar(integ, "ida_last_err")
    ida_alg_p_inf_psia = _mapping_scalar(integ, "ida_alg_p_inf_psia")
    ida_alg_v_inf_lbmolph = _mapping_scalar(integ, "ida_alg_v_inf_lbmolph")
    ida_alg_weighted = _mapping_scalar(integ, "ida_alg_weighted")
    ida_alg_converged = _mapping_scalar(integ, "ida_alg_converged")
    ida_resid_energy_btups = _mapping_scalar(integ, "ida_resid_energy_btups")

    if "P_psia_diag" in diag:
        P_diag = np.asarray(diag["P_psia_diag"], dtype=float).reshape((N,))
    else:
        P_diag = _pressure_diag_psia(col, volume_model, T, MV, Z)

    if N >= 1 and np.isfinite(P_spec[0]):
        P_diag[0] = float(P_spec[0])

    p_ctrl_idx = 0
    P_top_meas = float(P_diag[p_ctrl_idx])
    has_ctrl_pv = False
    if "P_top_ctrl_pv_psia" in diag:
        try:
            p_ctrl = float(np.asarray(diag["P_top_ctrl_pv_psia"], dtype=float).reshape((-1,))[0])
            if np.isfinite(p_ctrl) and p_ctrl > 0.0:
                P_top_meas = float(p_ctrl)
                has_ctrl_pv = True
        except Exception:
            pass
    has_top_drum_pv = False
    if (not has_ctrl_pv) and "P_top_drum_psia" in diag:
        try:
            p_top_drum = float(np.asarray(diag["P_top_drum_psia"], dtype=float).reshape((-1,))[0])
            if np.isfinite(p_top_drum) and p_top_drum > 0.0:
                P_top_meas = float(p_top_drum)
                has_top_drum_pv = True
        except Exception:
            pass
    if (not has_ctrl_pv) and (not has_top_drum_pv) and "P_psia_hyd" in diag:
        try:
            p_h = np.asarray(diag["P_psia_hyd"], dtype=float).reshape((N,))
            if np.isfinite(float(p_h[p_ctrl_idx])) and float(p_h[p_ctrl_idx]) > 0.0:
                P_top_meas = float(p_h[p_ctrl_idx])
        except Exception:
            pass

    top_L_total = float(np.sum(u["top_L"])) if (layout.include_top and "top_L" in u) else None
    top_x = _normalized_component_holdup_comp(u.get("top_L"), Nc) if layout.include_top else None
    top_L_cond_in_lbmolph = _mapping_scalar(diag, "top_L_cond_in_lbmolph")
    top_L_reflux_out_lbmolph = _mapping_scalar(diag, "top_L_reflux_out_lbmolph")
    top_L_distillate_out_lbmolph = _mapping_scalar(diag, "top_L_distillate_out_lbmolph")
    top_L_net_lbmolph = _mapping_scalar(diag, "top_L_net_lbmolph")
    top_L_net_worst_component_1based = _mapping_scalar(diag, "top_L_net_worst_component_1based")
    top_L_net_worst_lbmolph = _mapping_scalar(diag, "top_L_net_worst_lbmolph")
    top_L_net_worst_abs_lbmolph = _mapping_scalar(diag, "top_L_net_worst_abs_lbmolph")
    bottom_L_total = float(np.sum(u["bottom_L"])) if (layout.include_bottom and "bottom_L" in u) else None
    bottom_sump_x = _normalized_component_holdup_comp(u.get("bottom_L"), Nc) if layout.include_bottom else None
    bottom_stage_x = np.asarray(x[-1, :], dtype=float).reshape((Nc,))
    bottom_x = bottom_sump_x if bottom_sump_x is not None else bottom_stage_x
    bottom_x_source = "sump" if bottom_sump_x is not None else "bottom-stage_missing-sump"
    bottom_stage_sump_tv = _composition_tv_distance(bottom_stage_x, bottom_sump_x)

    T_sump = None
    if layout.include_bottom and "bottom_T_f" in u:
        try:
            T_sump = float(u["bottom_T_f"][0])
        except Exception:
            T_sump = None

    P_bot_meas = np.nan
    if "P_psia_hyd" in diag:
        try:
            p_h = np.asarray(diag["P_psia_hyd"], dtype=float).reshape((N,))
            if np.isfinite(float(p_h[-1])) and float(p_h[-1]) > 0.0:
                P_bot_meas = float(p_h[-1])
        except Exception:
            P_bot_meas = np.nan
    if not np.isfinite(P_bot_meas):
        try:
            if np.isfinite(float(P_diag[-1])) and float(P_diag[-1]) > 0.0:
                P_bot_meas = float(P_diag[-1])
        except Exception:
            P_bot_meas = np.nan

    out: Dict[str, Any] = {
        "wall_clock_iso": wall_clock_iso,
        "wall_elapsed_s": float(wall_elapsed_s),
        "time_s": float(t_s),
        "P_top_psia": float(P_top_meas) if np.isfinite(P_top_meas) else (float(P_spec[0]) if np.isfinite(P_spec[0]) else float(P_diag[0])),
        "P_top_psia_spec": float(P_spec[0]) if np.isfinite(P_spec[0]) else np.nan,
        "P_top_ctrl_pv_psia": float(P_top_meas),
        "P_bot_psia": float(P_bot_meas) if np.isfinite(P_bot_meas) else (float(P_spec[-1]) if np.isfinite(P_spec[-1]) else float(P_diag[-1])),
        "P_bot_psia_spec": float(P_spec[-1]) if np.isfinite(P_spec[-1]) else np.nan,
        "T_Distillate_F": float(T_distillate) if T_distillate is not None else np.nan,
        "Q_cond_calc_BTUph": float(Q_cond_calc_BTUph) if np.isfinite(Q_cond_calc_BTUph) else np.nan,
        "Q_cond_used_BTUph": float(Q_cond_used_BTUph) if np.isfinite(Q_cond_used_BTUph) else np.nan,
        "Q_reb_used_BTUph": float(Q_reb_used_BTUph) if np.isfinite(Q_reb_used_BTUph) else np.nan,
        "boilup_from_duty_lbmolph": float(_mapping_scalar(diag, "boilup_from_duty_lbmolph"))
        if np.isfinite(_mapping_scalar(diag, "boilup_from_duty_lbmolph"))
        else np.nan,
        "boilup_realized_lbmolph": float(_mapping_scalar(diag, "boilup_realized_lbmolph"))
        if np.isfinite(_mapping_scalar(diag, "boilup_realized_lbmolph"))
        else np.nan,
        "reboiler_latent_heat_BTU_per_lbmol": float(_mapping_scalar(diag, "reboiler_latent_heat_BTU_per_lbmol"))
        if np.isfinite(_mapping_scalar(diag, "reboiler_latent_heat_BTU_per_lbmol"))
        else np.nan,
        "reboiler_temperature_F": float(_mapping_scalar(diag, "reboiler_temperature_F"))
        if np.isfinite(_mapping_scalar(diag, "reboiler_temperature_F"))
        else np.nan,
        "reboiler_mode_duty_active": float(_mapping_scalar(diag, "reboiler_mode_duty_active"))
        if np.isfinite(_mapping_scalar(diag, "reboiler_mode_duty_active"))
        else np.nan,
        "reboiler_neighbor_stage_1based": float(_mapping_scalar(diag, "reboiler_neighbor_stage_1based"))
        if np.isfinite(_mapping_scalar(diag, "reboiler_neighbor_stage_1based"))
        else np.nan,
        "reboiler_neighbor_vflow_calc_lbmolph": float(_mapping_scalar(diag, "reboiler_neighbor_vflow_calc_lbmolph"))
        if np.isfinite(_mapping_scalar(diag, "reboiler_neighbor_vflow_calc_lbmolph"))
        else np.nan,
        "reboiler_neighbor_vflow_used_lbmolph": float(_mapping_scalar(diag, "reboiler_neighbor_vflow_used_lbmolph"))
        if np.isfinite(_mapping_scalar(diag, "reboiler_neighbor_vflow_used_lbmolph"))
        else np.nan,
        "reboiler_neighbor_vflow_limit_hi_lbmolph": float(_mapping_scalar(diag, "reboiler_neighbor_vflow_limit_hi_lbmolph"))
        if np.isfinite(_mapping_scalar(diag, "reboiler_neighbor_vflow_limit_hi_lbmolph"))
        else np.nan,
        "reboiler_neighbor_vflow_limit_lo_lbmolph": float(_mapping_scalar(diag, "reboiler_neighbor_vflow_limit_lo_lbmolph"))
        if np.isfinite(_mapping_scalar(diag, "reboiler_neighbor_vflow_limit_lo_lbmolph"))
        else np.nan,
        "reboiler_neighbor_vflow_clamped_flag": float(_mapping_scalar(diag, "reboiler_neighbor_vflow_clamped_flag"))
        if np.isfinite(_mapping_scalar(diag, "reboiler_neighbor_vflow_clamped_flag"))
        else np.nan,
        "Q_cond_cmd_BTUph": float(Q_cond_cmd_BTUph) if np.isfinite(Q_cond_cmd_BTUph) else np.nan,
        "P_top_drum_psia": float(P_top_drum_psia) if np.isfinite(P_top_drum_psia) else np.nan,
        "P_top_drum_psia_raw": float(P_top_drum_psia_raw) if np.isfinite(P_top_drum_psia_raw) else np.nan,
        "Z_top_drum_vapor": float(Z_top_drum_vapor) if np.isfinite(Z_top_drum_vapor) else np.nan,
        "MV_top_drum_lbmol": float(MV_top_drum_lbmol) if np.isfinite(MV_top_drum_lbmol) else np.nan,
        "V_condensed_in_lbmolph": float(V_condensed_in_lbmolph) if np.isfinite(V_condensed_in_lbmolph) else np.nan,
        "V_to_top_drum_lbmolph": float(V_to_top_drum_lbmolph) if np.isfinite(V_to_top_drum_lbmolph) else np.nan,
        "top_L_cond_in_lbmolph": (
            float(top_L_cond_in_lbmolph) if np.isfinite(top_L_cond_in_lbmolph) else np.nan
        ),
        "top_L_reflux_out_lbmolph": (
            float(top_L_reflux_out_lbmolph) if np.isfinite(top_L_reflux_out_lbmolph) else np.nan
        ),
        "top_L_distillate_out_lbmolph": (
            float(top_L_distillate_out_lbmolph) if np.isfinite(top_L_distillate_out_lbmolph) else np.nan
        ),
        "top_L_net_lbmolph": float(top_L_net_lbmolph) if np.isfinite(top_L_net_lbmolph) else np.nan,
        "top_L_net_worst_component_1based": (
            float(top_L_net_worst_component_1based)
            if np.isfinite(top_L_net_worst_component_1based)
            else np.nan
        ),
        "top_L_net_worst_lbmolph": (
            float(top_L_net_worst_lbmolph) if np.isfinite(top_L_net_worst_lbmolph) else np.nan
        ),
        "top_L_net_worst_abs_lbmolph": (
            float(top_L_net_worst_abs_lbmolph)
            if np.isfinite(top_L_net_worst_abs_lbmolph)
            else np.nan
        ),
        "total_reflux_mode_active": (
            float(total_reflux_mode_active) if np.isfinite(total_reflux_mode_active) else np.nan
        ),
        "total_reflux_actual_lbmolph": (
            float(total_reflux_actual_lbmolps) * 3600.0 if np.isfinite(total_reflux_actual_lbmolps) else np.nan
        ),
        "total_reflux_kickstart_lbmolph": (
            float(total_reflux_kickstart_lbmolps) * 3600.0
            if np.isfinite(total_reflux_kickstart_lbmolps)
            else np.nan
        ),
        "total_reflux_used_lbmolph": (
            float(total_reflux_used_lbmolps) * 3600.0 if np.isfinite(total_reflux_used_lbmolps) else np.nan
        ),
        "total_reflux_startup_factor": (
            float(total_reflux_startup_factor) if np.isfinite(total_reflux_startup_factor) else np.nan
        ),
        "total_reflux_reflux_startup_factor": (
            float(total_reflux_reflux_startup_factor)
            if np.isfinite(total_reflux_reflux_startup_factor)
            else np.nan
        ),
        "total_reflux_boundary_external_scale": (
            float(total_reflux_boundary_external_scale)
            if np.isfinite(total_reflux_boundary_external_scale)
            else np.nan
        ),
        "total_reflux_boundary_closed_fraction": (
            float(total_reflux_boundary_closed_fraction)
            if np.isfinite(total_reflux_boundary_closed_fraction)
            else np.nan
        ),
        "eq_relax_tau_effective_sec": (
            float(eq_relax_tau_effective_sec) if np.isfinite(eq_relax_tau_effective_sec) else np.nan
        ),
        "startup_total_reflux_washout_active": (
            float(startup_total_reflux_washout_active)
            if np.isfinite(startup_total_reflux_washout_active)
            else np.nan
        ),
        "startup_sequence_phase_id": (
            float(startup_sequence_phase_id) if np.isfinite(startup_sequence_phase_id) else np.nan
        ),
        "startup_sequence_liquid_alpha_cmd": (
            float(startup_sequence_liquid_alpha_cmd)
            if np.isfinite(startup_sequence_liquid_alpha_cmd)
            else np.nan
        ),
        "startup_sequence_vapor_homotopy_enabled": (
            float(startup_sequence_vapor_homotopy_enabled)
            if np.isfinite(startup_sequence_vapor_homotopy_enabled)
            else np.nan
        ),
        "startup_sequence_vapor_beta_cmd": (
            float(startup_sequence_vapor_beta_cmd)
            if np.isfinite(startup_sequence_vapor_beta_cmd)
            else np.nan
        ),
        "startup_sequence_vapor_guard_phase_id": (
            float(startup_sequence_vapor_guard_phase_id)
            if np.isfinite(startup_sequence_vapor_guard_phase_id)
            else np.nan
        ),
        "startup_sequence_vapor_rel_rate_gate_per_s": (
            float(startup_sequence_vapor_rel_rate_gate_per_s)
            if np.isfinite(startup_sequence_vapor_rel_rate_gate_per_s)
            else np.nan
        ),
        "startup_sequence_rel_rate_prev_per_s": (
            float(startup_sequence_rel_rate_prev_per_s)
            if np.isfinite(startup_sequence_rel_rate_prev_per_s)
            else np.nan
        ),
        "vflow_homotopy_active": (
            float(vflow_homotopy_active) if np.isfinite(vflow_homotopy_active) else np.nan
        ),
        "vflow_homotopy_beta": (
            float(vflow_homotopy_beta) if np.isfinite(vflow_homotopy_beta) else np.nan
        ),
        "dP_stage2_to_top_drum_psia": (
            float(dP_stage2_to_top_drum_psia) if np.isfinite(dP_stage2_to_top_drum_psia) else np.nan
        ),
        "V_to_top_drum_pressure_gate_scale": (
            float(V_to_top_drum_pressure_gate_scale) if np.isfinite(V_to_top_drum_pressure_gate_scale) else np.nan
        ),
        "V_to_top_drum_blocked_lbmolph": (
            float(V_to_top_drum_blocked_lbmolph) if np.isfinite(V_to_top_drum_blocked_lbmolph) else np.nan
        ),
        "V_condensed_top_lbmolph": float(V_condensed_top_lbmolph) if np.isfinite(V_condensed_top_lbmolph) else np.nan,
        "V_psv_top_lbmolph": float(V_psv_top_lbmolph) if np.isfinite(V_psv_top_lbmolph) else np.nan,
        "PSV_open_flag": float(PSV_open_flag) if np.isfinite(PSV_open_flag) else np.nan,
        "PSV_setpoint_psia": float(PSV_setpoint_psia) if np.isfinite(PSV_setpoint_psia) else np.nan,
        "PSV_pv_psia": float(PSV_pv_psia) if np.isfinite(PSV_pv_psia) else np.nan,
        "V_top_drum_vapor_ft3": float(V_top_drum_vapor_ft3) if np.isfinite(V_top_drum_vapor_ft3) else np.nan,
        "V_top_drum_liquid_ft3": float(V_top_drum_liquid_ft3) if np.isfinite(V_top_drum_liquid_ft3) else np.nan,
        "rho_top_drum_liq_lbmol_ft3": float(rho_top_drum_liq_lbmol_ft3) if np.isfinite(rho_top_drum_liq_lbmol_ft3) else np.nan,
        "Q_reb_cmd_BTUph": float(Q_reb_cmd_BTUph) if np.isfinite(Q_reb_cmd_BTUph) else np.nan,
        "P_top_anchor_cmd_psia": float(P_top_anchor_cmd_psia) if np.isfinite(P_top_anchor_cmd_psia) else np.nan,
        "P_top_ctrl_pv_raw_psia": float(P_top_ctrl_pv_raw_psia) if np.isfinite(P_top_ctrl_pv_raw_psia) else np.nan,
        "P_top_ctrl_pv_filt_psia": float(P_top_ctrl_pv_filt_psia) if np.isfinite(P_top_ctrl_pv_filt_psia) else np.nan,
            "P_top_ctrl_gain_scale": float(P_top_ctrl_gain_scale) if np.isfinite(P_top_ctrl_gain_scale) else np.nan,
            "P_top_ctrl_energy_resid_abs_BTUps": (
                float(P_top_ctrl_energy_resid_abs_BTUps) if np.isfinite(P_top_ctrl_energy_resid_abs_BTUps) else np.nan
            ),
        "P_top_ctrl_err_psia": float(P_top_ctrl_err_psia) if np.isfinite(P_top_ctrl_err_psia) else np.nan,
        "P_top_ctrl_p_term": float(P_top_ctrl_p_term) if np.isfinite(P_top_ctrl_p_term) else np.nan,
        "P_top_ctrl_i_term": float(P_top_ctrl_i_term) if np.isfinite(P_top_ctrl_i_term) else np.nan,
        "P_top_ctrl_u_unclamped_BTUph": (
            float(P_top_ctrl_u_unclamped_BTUph) if np.isfinite(P_top_ctrl_u_unclamped_BTUph) else np.nan
        ),
        "P_top_ctrl_sat_hi": float(P_top_ctrl_sat_hi) if np.isfinite(P_top_ctrl_sat_hi) else np.nan,
        "P_top_ctrl_sat_lo": float(P_top_ctrl_sat_lo) if np.isfinite(P_top_ctrl_sat_lo) else np.nan,
        "P_top_ctrl_allow_int": float(P_top_ctrl_allow_int) if np.isfinite(P_top_ctrl_allow_int) else np.nan,
        "Top_level_ctrl_pv": float(_mapping_scalar(diag, "Top_level_ctrl_pv")),
        "Top_level_ctrl_sp": float(_mapping_scalar(diag, "Top_level_ctrl_sp")),
        "Top_level_ctrl_err": float(_mapping_scalar(diag, "Top_level_ctrl_err")),
        "Top_level_ctrl_p_term": float(_mapping_scalar(diag, "Top_level_ctrl_p_term")),
        "Top_level_ctrl_i_term": float(_mapping_scalar(diag, "Top_level_ctrl_i_term")),
        "Top_level_ctrl_pi_term_lbmolph": float(_mapping_scalar(diag, "Top_level_ctrl_pi_term_lbmolph")),
        "Top_level_ctrl_ff_term_lbmolph": float(_mapping_scalar(diag, "Top_level_ctrl_ff_term_lbmolph")),
        "Top_level_ctrl_u_unclamped_lbmolph": float(_mapping_scalar(diag, "Top_level_ctrl_u_unclamped_lbmolph")),
        "Top_level_ctrl_sat_hi": float(_mapping_scalar(diag, "Top_level_ctrl_sat_hi")),
        "Top_level_ctrl_sat_lo": float(_mapping_scalar(diag, "Top_level_ctrl_sat_lo")),
        "Top_level_ctrl_allow_int": float(_mapping_scalar(diag, "Top_level_ctrl_allow_int")),
        "Bottom_level_ctrl_pv": float(_mapping_scalar(diag, "Bottom_level_ctrl_pv")),
        "Bottom_level_ctrl_sp": float(_mapping_scalar(diag, "Bottom_level_ctrl_sp")),
        "Bottom_level_ctrl_err": float(_mapping_scalar(diag, "Bottom_level_ctrl_err")),
        "Bottom_level_ctrl_p_term": float(_mapping_scalar(diag, "Bottom_level_ctrl_p_term")),
        "Bottom_level_ctrl_i_term": float(_mapping_scalar(diag, "Bottom_level_ctrl_i_term")),
        "Bottom_level_ctrl_u_unclamped_lbmolph": float(_mapping_scalar(diag, "Bottom_level_ctrl_u_unclamped_lbmolph")),
        "Bottom_level_ctrl_sat_hi": float(_mapping_scalar(diag, "Bottom_level_ctrl_sat_hi")),
        "Bottom_level_ctrl_sat_lo": float(_mapping_scalar(diag, "Bottom_level_ctrl_sat_lo")),
        "Bottom_level_ctrl_allow_int": float(_mapping_scalar(diag, "Bottom_level_ctrl_allow_int")),
            "xD_comp_sp": float(xD_comp_sp) if np.isfinite(xD_comp_sp) else np.nan,
        "xD_comp_pv": float(xD_comp_pv) if np.isfinite(xD_comp_pv) else np.nan,
        "xD_comp_err": float(xD_comp_err) if np.isfinite(xD_comp_err) else np.nan,
        "RR_comp_cmd": float(RR_comp_cmd) if np.isfinite(RR_comp_cmd) else np.nan,
        "Reflux_cmd_lbmolph": float(Reflux_cmd_lbmolph) if np.isfinite(Reflux_cmd_lbmolph) else np.nan,
        "Reflux_cmd_unclamped_lbmolph": (
            float(Reflux_cmd_unclamped_lbmolph) if np.isfinite(Reflux_cmd_unclamped_lbmolph) else np.nan
        ),
        "Reflux_cmd_active_max_lbmolph": (
            float(Reflux_cmd_active_max_lbmolph) if np.isfinite(Reflux_cmd_active_max_lbmolph) else np.nan
        ),
        "Reflux_cap_active_flag": float(Reflux_cap_active_flag) if np.isfinite(Reflux_cap_active_flag) else np.nan,
        "xB_comp_sp": float(xB_comp_sp) if np.isfinite(xB_comp_sp) else np.nan,
        "xB_comp_pv": float(xB_comp_pv) if np.isfinite(xB_comp_pv) else np.nan,
        "Boilup_cmd_lbmolph": float(Boilup_cmd_lbmolph) if np.isfinite(Boilup_cmd_lbmolph) else np.nan,
        "M_total_lbmol": float(M_total_lbmol) if np.isfinite(M_total_lbmol) else np.nan,
        "dM_total_dt_lbmolph": float(dM_total_dt_lbmolph) if np.isfinite(dM_total_dt_lbmolph) else np.nan,
        "net_F_minus_D_minus_B_lbmolph": float(net_F_minus_D_minus_B_lbmolph) if np.isfinite(net_F_minus_D_minus_B_lbmolph) else np.nan,
        "global_mass_closure_error_lbmolph": float(global_mass_closure_error_lbmolph) if np.isfinite(global_mass_closure_error_lbmolph) else np.nan,
        "global_mass_closure_cum_lbmol": float(global_mass_closure_cum_lbmol) if np.isfinite(global_mass_closure_cum_lbmol) else np.nan,
        "stage_mass_resid_sum_lbmolps": float(stage_mass_resid_sum_lbmolps) if np.isfinite(stage_mass_resid_sum_lbmolps) else np.nan,
        "pv_inner_iter_count": float(pv_inner_iter_count) if np.isfinite(pv_inner_iter_count) else np.nan,
        "pv_inner_converged": float(pv_inner_converged) if np.isfinite(pv_inner_converged) else np.nan,
        "pv_inner_dp_max_psia": float(pv_inner_dp_max_psia) if np.isfinite(pv_inner_dp_max_psia) else np.nan,
        "pv_inner_dv_max_lbmolph": float(pv_inner_dv_max_lbmolph) if np.isfinite(pv_inner_dv_max_lbmolph) else np.nan,
        "K_state_over_K_thermo_max_abs": (
            float(K_state_over_K_thermo_max_abs) if np.isfinite(K_state_over_K_thermo_max_abs) else np.nan
        ),
        "K_state_minus_K_thermo_max_abs": (
            float(K_state_minus_K_thermo_max_abs) if np.isfinite(K_state_minus_K_thermo_max_abs) else np.nan
        ),
        "steady_state_enabled": float(steady_state_enabled) if np.isfinite(steady_state_enabled) else np.nan,
        "steady_state_flag": float(steady_state_flag) if np.isfinite(steady_state_flag) else np.nan,
        "steady_state_score": float(steady_state_score) if np.isfinite(steady_state_score) else np.nan,
        "steady_state_active_criteria": (
            float(steady_state_active_criteria) if np.isfinite(steady_state_active_criteria) else np.nan
        ),
        "ss_max_rel_state_rate_per_s": (
            float(ss_max_rel_state_rate_per_s) if np.isfinite(ss_max_rel_state_rate_per_s) else np.nan
        ),
        "ss_rel_state_rate_state_key": str(ss_rel_state_rate_state_key),
        "ss_rel_state_rate_stage_1based": (
            float(ss_rel_state_rate_stage_1based) if np.isfinite(ss_rel_state_rate_stage_1based) else np.nan
        ),
        "ss_rel_state_rate_component_1based": (
            float(ss_rel_state_rate_component_1based)
            if np.isfinite(ss_rel_state_rate_component_1based)
            else np.nan
        ),
        "ss_rel_state_rate_component_name": str(ss_rel_state_rate_component_name),
        "ss_max_kpi_slope_per_s": float(ss_max_kpi_slope_per_s) if np.isfinite(ss_max_kpi_slope_per_s) else np.nan,
        "ss_max_mv_rate_per_s": float(ss_max_mv_rate_per_s) if np.isfinite(ss_max_mv_rate_per_s) else np.nan,
        "ss_max_temp_rate_F_per_s": (
            float(ss_max_temp_rate_F_per_s) if np.isfinite(ss_max_temp_rate_F_per_s) else np.nan
        ),
        "ss_max_sp_error": float(ss_max_sp_error) if np.isfinite(ss_max_sp_error) else np.nan,
        "ss_window_samples": float(ss_window_samples) if np.isfinite(ss_window_samples) else np.nan,
        "ss_window_sec": float(ss_window_sec) if np.isfinite(ss_window_sec) else np.nan,
        "ss_min_time_sec": float(ss_min_time_sec) if np.isfinite(ss_min_time_sec) else np.nan,
        "ss_tol_rel_state_rate_per_s": (
            float(ss_tol_rel_state_rate_per_s) if np.isfinite(ss_tol_rel_state_rate_per_s) else np.nan
        ),
        "ss_tol_kpi_slope_per_s": (
            float(ss_tol_kpi_slope_per_s) if np.isfinite(ss_tol_kpi_slope_per_s) else np.nan
        ),
        "ss_tol_mv_rate_per_s": float(ss_tol_mv_rate_per_s) if np.isfinite(ss_tol_mv_rate_per_s) else np.nan,
        "ss_tol_temp_rate_F_per_s": (
            float(ss_tol_temp_rate_F_per_s) if np.isfinite(ss_tol_temp_rate_F_per_s) else np.nan
        ),
        "ss_tol_sp_error": float(ss_tol_sp_error) if np.isfinite(ss_tol_sp_error) else np.nan,
        "ss_require_sp": float(ss_require_sp) if np.isfinite(ss_require_sp) else np.nan,
        "debug_freeze_tray_vapor_derivatives_active": (
            float(debug_freeze_tray_vapor_derivatives_active)
            if np.isfinite(debug_freeze_tray_vapor_derivatives_active)
            else np.nan
        ),
        "debug_max_orig_dmVdt": (
            float(debug_max_orig_dmVdt) if np.isfinite(debug_max_orig_dmVdt) else np.nan
        ),
        "debug_max_orig_dmVdt_rel_per_s": (
            float(debug_max_orig_dmVdt_rel_per_s)
            if np.isfinite(debug_max_orig_dmVdt_rel_per_s)
            else np.nan
        ),
        "debug_worst_v_stage": (
            float(debug_worst_v_stage) if np.isfinite(debug_worst_v_stage) else np.nan
        ),
        "debug_worst_v_comp": (
            float(debug_worst_v_comp) if np.isfinite(debug_worst_v_comp) else np.nan
        ),
        "debug_worst_v_rel_stage": (
            float(debug_worst_v_rel_stage) if np.isfinite(debug_worst_v_rel_stage) else np.nan
        ),
        "debug_worst_v_rel_comp": (
            float(debug_worst_v_rel_comp) if np.isfinite(debug_worst_v_rel_comp) else np.nan
        ),
        "debug_total_v_cancellation_lbmolps": (
            float(debug_total_v_cancellation_lbmolps)
            if np.isfinite(debug_total_v_cancellation_lbmolps)
            else np.nan
        ),
        "debug_net_orig_dmVdt_lbmolps": (
            float(debug_net_orig_dmVdt_lbmolps)
            if np.isfinite(debug_net_orig_dmVdt_lbmolps)
            else np.nan
        ),
        "integrator_requested_mode": str(integrator_requested_mode),
        "integrator_used_mode": str(integrator_used_mode),
        "integrator_fallback_used": (
            float(integrator_fallback_used) if np.isfinite(integrator_fallback_used) else np.nan
        ),
        "integrator_fallback_reason": str(integrator_fallback_reason),
        "integrator_nfev": float(integrator_nfev) if np.isfinite(integrator_nfev) else np.nan,
        "ida_iter_max": float(ida_iter_max) if np.isfinite(ida_iter_max) else np.nan,
        "ida_iter_mean": float(ida_iter_mean) if np.isfinite(ida_iter_mean) else np.nan,
        "ida_converged": float(ida_converged) if np.isfinite(ida_converged) else np.nan,
        "ida_last_err": float(ida_last_err) if np.isfinite(ida_last_err) else np.nan,
        "ida_alg_p_inf_psia": float(ida_alg_p_inf_psia) if np.isfinite(ida_alg_p_inf_psia) else np.nan,
        "ida_alg_v_inf_lbmolph": float(ida_alg_v_inf_lbmolph) if np.isfinite(ida_alg_v_inf_lbmolph) else np.nan,
        "ida_alg_weighted": float(ida_alg_weighted) if np.isfinite(ida_alg_weighted) else np.nan,
        "ida_alg_converged": float(ida_alg_converged) if np.isfinite(ida_alg_converged) else np.nan,
        "ida_resid_energy_btups": float(ida_resid_energy_btups) if np.isfinite(ida_resid_energy_btups) else np.nan,
        # New: overall stream flow scalars
        "F_lbmolph": float(feed_tag.flow_lbmolph) if feed_tag.flow_lbmolph is not None else np.nan,
        "D_lbmolph": float(dist_tag.flow_lbmolph) if dist_tag.flow_lbmolph is not None else np.nan,
        "B_lbmolph": float(bots_tag.flow_lbmolph) if bots_tag.flow_lbmolph is not None else np.nan,
        "Distillate_L_lbmol": float(top_L_total) if top_L_total is not None else np.nan,
        "Bottoms_L_lbmol": float(bottom_L_total) if bottom_L_total is not None else np.nan,
        "T_sump_F": float(T_sump) if T_sump is not None else np.nan,
        "Bottoms_x_source": str(bottom_x_source),
        "Bottoms_stage_sump_tv_distance": float(bottom_stage_sump_tv) if np.isfinite(bottom_stage_sump_tv) else np.nan,
    }

    def _diag_component_vector(name: str) -> np.ndarray:
        if name not in diag:
            return np.full((Nc,), np.nan, dtype=float)
        try:
            arr = np.asarray(diag[name], dtype=float).reshape((Nc,))
            return arr.copy()
        except Exception:
            return np.full((Nc,), np.nan, dtype=float)

    top_L_cond_in_comp = _diag_component_vector("top_L_cond_in_lbmolph_comp")
    top_L_reflux_out_comp = _diag_component_vector("top_L_reflux_out_lbmolph_comp")
    top_L_distillate_out_comp = _diag_component_vector("top_L_distillate_out_lbmolph_comp")
    top_L_net_comp = _diag_component_vector("top_L_net_lbmolph_comp")
    top_L_cond_x_delta = _diag_component_vector("top_L_cond_x_minus_drum_x")

    for k in range(Nc):
        label = comp_labels[k]
        out[f"x_Distillate_{label}"] = float(x[0, k])
        out[f"y_Distillate_{label}"] = float(yv[0, k])
        out[f"x_Bottoms_{label}"] = float(x[-1, k])
        out[f"y_Bottoms_{label}"] = float(yv[-1, k])
        if top_x is not None:
            out[f"Distillate_x_{label}"] = float(top_x[k])
        if bottom_x is not None:
            out[f"Bottoms_x_{label}"] = float(bottom_x[k])
        if bottom_sump_x is not None:
            out[f"Bottoms_sump_x_{label}"] = float(bottom_sump_x[k])
        out[f"top_L_cond_in_{label}_lbmolph"] = float(top_L_cond_in_comp[k])
        out[f"top_L_reflux_out_{label}_lbmolph"] = float(top_L_reflux_out_comp[k])
        out[f"top_L_distillate_out_{label}_lbmolph"] = float(top_L_distillate_out_comp[k])
        out[f"top_L_net_{label}_lbmolph"] = float(top_L_net_comp[k])
        out[f"top_L_cond_x_minus_drum_x_{label}"] = float(top_L_cond_x_delta[k])

    return out


# -------------------------
# Runner
# -------------------------


def run_smoke_simulation(cfg: RunnerConfig) -> Dict[str, Any]:
    _configure_progress_streams()
    startup_logs_dir = Path(cfg.logs_dir)
    if not startup_logs_dir.is_absolute():
        startup_logs_dir = Path.cwd() / startup_logs_dir
    startup_trace_tag = _clean_optional_text(getattr(cfg, "run_name", None)) or _timestamp_tag()
    startup_trace_path = startup_logs_dir / f"startup_trace_{startup_trace_tag}.log"
    _set_progress_trace_path(startup_trace_path)
    milestone_t0 = time.perf_counter()

    def _milestone(label: str) -> None:
        wall = time.perf_counter() - milestone_t0
        clock = _dt.datetime.now().isoformat(timespec="seconds")
        _emit_progress(f"[Milestone] {label}  wall={wall:8.2f} s  clock={clock}")

    def _runtime_trace(message: str, *, echo: bool = False) -> None:
        _emit_progress(f"[RuntimeTrace] {message}", echo=echo)

    _emit_progress(f"[Init] Startup trace active: {startup_trace_path}")
    _milestone("start")
    try:
        case = load_case_from_excel(cfg.excel_path)
        _milestone("loaded case from Excel")
    except Exception as exc:
        _emit_progress("[Validation] FAIL  errors=1  warnings=0")
        _emit_progress(f"[Validation][Error] Failed to load Excel case: {exc}")
        raise

    try:
        col = build_column_spec_from_case(case)
        _milestone("built column spec")
    except Exception as exc:
        _emit_progress("[Validation] FAIL  errors=1  warnings=0")
        _emit_progress(f"[Validation][Error] Failed to build column spec: {exc}")
        raise

    validation = validate_loaded_case(case, col)
    print_validation_report(validation)
    _milestone("validated Excel input")
    if not validation.ok:
        raise ValueError("Excel input validation failed. See [Validation][Error] lines above.")

    dt = float(cfg.dt_sec) if cfg.dt_sec is not None else float(col.sim.dt_sec)
    if dt <= 0.0:
        raise ValueError("dt_sec must be > 0")

    log_every = int(cfg.log_every_n_steps) if cfg.log_every_n_steps is not None else int(col.sim.log_every_n_steps)
    if log_every <= 0:
        raise ValueError("log_every_n_steps must be > 0")
    parity_runtime_thermo_defer_visible_steps = _resolve_parity_runtime_thermo_defer_visible_steps(
        cfg,
        log_every_n_steps=int(log_every),
    )

    thermo_every = int(cfg.thermo_every_n_steps)
    if thermo_every <= 0:
        thermo_every = 1

    layout = StateVectorLayout(
        n_stages=col.n_stages,
        n_components=col.n_components,
        include_top=bool(cfg.include_boundary_states),
        include_bottom=bool(cfg.include_boundary_states),
        include_vapor=bool(cfg.include_vapor_states),
        include_temperature=bool(cfg.include_temperature),
        include_energy=bool(cfg.include_energy),
    )
    _milestone("built state vector layout")

    _emit_progress("[Init] Building runner inputs and thermo provider")
    base_inputs, thermo_provider = build_inputs_for_runner(case, col, cfg)
    _milestone("built inputs and thermo provider")
    startup_timing_sec: Dict[str, Any] = dict(getattr(base_inputs, "startup_build_timing_sec", {}) or {})
    startup_build_info: Dict[str, Any] = dict(getattr(base_inputs, "startup_build_info", {}) or {})
    _thermo_providers_to_reset: list[Any] = []
    for _prov in (
        thermo_provider,
        getattr(base_inputs, "thermo_provider", None),
        getattr(base_inputs, "equilibrium_relaxation_thermo_provider", None),
    ):
        if _prov is None:
            continue
        if any(_prov is existing for existing in _thermo_providers_to_reset):
            continue
        _thermo_providers_to_reset.append(_prov)
    for _prov in _thermo_providers_to_reset:
        reset_fn = getattr(_prov, "reset_call_counters", None)
        if callable(reset_fn):
            try:
                reset_fn()
            except Exception:
                pass
    runtime_mode = _normalize_runtime_mode(getattr(cfg, "runtime_mode", None), default="legacy")
    integrator_mode = _normalize_integrator_mode(getattr(cfg, "integrator", None), default="explicit-euler")
    eff_ida = _effective_hydraulic_ida_profile(
        cfg,
        runtime_mode=str(runtime_mode),
        integrator_mode=str(integrator_mode),
    )
    ida_max_iter_eff = int(eff_ida.get("ida_max_iter", getattr(cfg, "ida_max_iter", 8)))
    dae_pilot_enabled_eff = bool(
        eff_ida.get("dae_pilot_enabled", getattr(cfg, "enable_dae_pilot_algebraic_solve", False))
    )
    dae_pilot_max_iter_eff = int(eff_ida.get("dae_pilot_max_iter", getattr(cfg, "dae_pilot_max_iter", 3)))
    dae_pilot_p_tol_eff = eff_ida.get("dae_pilot_p_tol_psia", getattr(cfg, "dae_pilot_p_tol_psia", None))
    dae_pilot_v_tol_eff = eff_ida.get("dae_pilot_v_tol_lbmolph", getattr(cfg, "dae_pilot_v_tol_lbmolph", None))
    ida_defaults_applied = list(eff_ida.get("defaults_applied", []))

    startup_sequence_enabled = bool(cfg.enable_startup_hydraulic_sequence)
    if runtime_mode in ("parity", "calibration", "total-reflux"):
        if startup_sequence_enabled:
            print(
                f"[Init] runtime_mode={runtime_mode} disables startup hydraulic sequencing; using direct mode behavior."
            )
        startup_sequence_enabled = False
    if runtime_mode in ("parity", "calibration", "hydraulic", "total-reflux"):
        print(f"[Init] Runtime mode active: {runtime_mode}")
    if ida_defaults_applied:
        print("[Init] Applied hydraulic+ida defaults: " + ", ".join(str(x) for x in ida_defaults_applied))
    if integrator_mode != "explicit-euler":
        rhs_cap_txt = (
            str(int(cfg.integrator_max_rhs_evals_per_step))
            if cfg.integrator_max_rhs_evals_per_step is not None
            else "off"
        )
        substep_txt = (
            f"{float(cfg.integrator_substep_sec):.3g}s"
            if cfg.integrator_substep_sec is not None
            else "off"
        )
        wall_cap_txt = (
            f"{float(cfg.integrator_step_wall_limit_sec):.3g}s"
            if cfg.integrator_step_wall_limit_sec is not None
            else "off"
        )
        smooth_cfg = getattr(cfg, "stiff_vflow_smooth_clamp_lbmolph", None)
        if smooth_cfg is None:
            smooth_txt = "auto"
        else:
            try:
                smooth_try = float(smooth_cfg)
            except Exception:
                smooth_try = 0.0
            if np.isfinite(smooth_try) and smooth_try > 0.0:
                smooth_txt = f"{smooth_try:.3g} lbmol/h"
            else:
                smooth_txt = "off"
        if integrator_mode in ("bdf", "radau") and _solve_ivp is None:
            print(
                f"[Warn] integrator={integrator_mode} requested but SciPy is unavailable; "
                "falling back to explicit-euler."
            )
        elif integrator_mode == "ida":
            ida_dae_p_txt = "off"
            ida_dae_v_txt = "off"
            try:
                p_tol_try = float(dae_pilot_p_tol_eff) if dae_pilot_p_tol_eff is not None else np.nan
            except Exception:
                p_tol_try = np.nan
            try:
                v_tol_try = float(dae_pilot_v_tol_eff) if dae_pilot_v_tol_eff is not None else np.nan
            except Exception:
                v_tol_try = np.nan
            if np.isfinite(p_tol_try):
                ida_dae_p_txt = f"{p_tol_try:.3g}"
            if np.isfinite(v_tol_try):
                ida_dae_v_txt = f"{v_tol_try:.3g}"
            print(
                "[Init] Integrator active: "
                f"{integrator_mode}  rtol={float(cfg.integrator_rtol):.3g}  "
                f"atol={float(cfg.integrator_atol):.3g}  "
                f"substep={substep_txt}  "
                f"max_rhs={rhs_cap_txt}  wall_cap={wall_cap_txt}  "
                f"ida_max_iter={int(ida_max_iter_eff)}  "
                f"ida_relax={float(getattr(cfg, 'ida_relax', 1.0)):.3g}  "
                f"dae_pilot={'on' if bool(dae_pilot_enabled_eff) else 'off'}  "
                f"dae_p_tol={ida_dae_p_txt}  dae_v_tol={ida_dae_v_txt}  "
                f"stiff_vflow_smooth={smooth_txt}"
            )
        else:
            print(
                "[Init] Integrator active: "
                f"{integrator_mode}  rtol={float(cfg.integrator_rtol):.3g}  "
                f"atol={float(cfg.integrator_atol):.3g}  "
                f"substep={substep_txt}  "
                f"max_rhs={rhs_cap_txt}  wall_cap={wall_cap_txt}  "
                f"stiff_vflow_smooth={smooth_txt}"
            )
    if (
        str(base_inputs.pressure_model).strip().lower() == "hydraulic"
        and str(base_inputs.vapor_flow_model).strip().lower() == "energy"
        and str(base_inputs.condenser_duty_mode).strip().lower() == "total-condense"
    ):
        print(
            "[Warn] hydraulic+energy with condenser-duty-mode=total-condense can be stiff; "
            "consider condenser-duty-mode=specified while approaching steady state."
        )
    if startup_sequence_enabled:
        print(
            "[Init] Startup hydraulic sequencing enabled  "
            f"energy_on={float(cfg.startup_sequence_energy_on_sec):.3g}s  "
            f"liquid_on={float(cfg.startup_sequence_liquid_on_sec):.3g}s  "
            f"liquid_ramp={float(cfg.startup_sequence_liquid_ramp_sec):.3g}s  "
            f"resid_gate={float(cfg.startup_sequence_mass_resid_gate_lbmolph) if cfg.startup_sequence_mass_resid_gate_lbmolph is not None else float('nan'):.3g} lbmol/h"
        )
        if bool(cfg.enable_startup_vapor_homotopy):
            v_on = cfg.startup_sequence_vapor_on_sec
            if v_on is None:
                v_on = float(cfg.startup_sequence_liquid_on_sec) + float(cfg.startup_sequence_liquid_ramp_sec)
            print(
                "[Init] Startup vapor homotopy enabled  "
                f"profile_hold={float(cfg.startup_sequence_profile_hold_sec):.3g}s  "
                f"vapor_on={float(v_on):.3g}s  "
                f"vapor_ramp={float(cfg.startup_sequence_vapor_ramp_sec):.3g}s  "
                f"rel_gate={float(cfg.startup_sequence_vapor_rel_rate_gate_per_s) if cfg.startup_sequence_vapor_rel_rate_gate_per_s is not None else float('nan'):.3g}/s"
            )

    last_Z_tray: Optional[np.ndarray] = None
    last_y_eq: Optional[np.ndarray] = None
    last_P_diag: Optional[np.ndarray] = None
    last_P_hyd: Optional[np.ndarray] = None
    last_V_out: Optional[np.ndarray] = None
    last_dT_tray: Optional[np.ndarray] = None
    last_rhoL: Optional[np.ndarray] = None
    last_tray_thermo_packet: Optional[TrayThermoPacket] = None
    last_K_tray: Optional[np.ndarray] = None
    last_HL: Optional[np.ndarray] = None
    last_HV: Optional[np.ndarray] = None
    last_energy_resid_tray: Optional[np.ndarray] = None
    last_phase_energy_damping_min: Optional[np.ndarray] = None
    last_tray_temp_pressure_slope: Optional[np.ndarray] = None
    last_tray_bubble_target_F: Optional[np.ndarray] = None
    last_Zfac: Optional[np.ndarray] = None
    last_z_overall: Optional[np.ndarray] = None
    last_diag: Optional[Dict[str, np.ndarray]] = None
    last_reb_T: Optional[float] = None
    last_reb_x: Optional[np.ndarray] = None
    last_reb_y: Optional[np.ndarray] = None
    last_reb_beta: Optional[float] = None
    last_condenser_duty_packet: Optional[CondenserDutyPacket] = None
    last_feed_stage_flash_packet: Optional[FeedStageFlashPacket] = None
    last_bottom_sump_cp_packet: Optional[BottomSumpCpPacket] = None
    startup_seeded_condenser_duty_packet = False
    last_T_tray: Optional[np.ndarray] = None
    last_top_drum_pressure_T: Optional[float] = None
    last_mass_resid_max_lbmolph: Optional[float] = None
    last_mass_resid_lbmolph_per_stage: Optional[np.ndarray] = None
    restart_memory = getattr(col, "memory_state", None) or {}
    try:
        p_prev_restart = restart_memory.get("P_tray_prev_psia")
        if p_prev_restart is not None:
            last_P_diag = np.asarray(p_prev_restart, dtype=float).reshape((col.n_stages,)).copy()
            last_P_hyd = np.asarray(p_prev_restart, dtype=float).reshape((col.n_stages,)).copy()
    except Exception:
        last_P_diag = None
        last_P_hyd = None
    try:
        t_prev_restart = restart_memory.get("T_tray_prev_F")
        if t_prev_restart is not None:
            last_T_tray = np.asarray(t_prev_restart, dtype=float).reshape((col.n_stages,)).copy()
    except Exception:
        last_T_tray = None
    seq_liquid_alpha_base = _clip_unit(base_inputs.liquid_hydraulic_override_alpha, default=1.0)
    if not bool(base_inputs.enable_liquid_hydraulic_override):
        seq_liquid_alpha_base = 0.0
    seq_liquid_alpha_state = np.full((col.n_stages,), float(seq_liquid_alpha_base), dtype=float)
    seq_liquid_guard_phase = "disabled" if seq_liquid_alpha_base <= 0.0 else "hold"
    seq_vapor_beta_state = 0.0 if bool(cfg.enable_startup_vapor_homotopy) else 1.0
    seq_vapor_guard_phase = "disabled" if not bool(cfg.enable_startup_vapor_homotopy) else "profile_hold"
    last_ss_rel_state_rate_per_s: Optional[float] = None
    explicit_runtime_restart = bool(_has_explicit_runtime_restart_state(col))
    startup_seed_loaded = False
    startup_seed_cache_path = _resolve_startup_seed_cache_path(cfg)
    startup_seed_cache_info: Dict[str, Any] = {
        "enabled": bool(getattr(cfg, "enable_startup_seed_cache", False)),
        "path": str(startup_seed_cache_path) if startup_seed_cache_path is not None else "",
        "loaded": False,
        "saved": False,
        "reason": "disabled",
    }

    # Initial conditions from ColumnSpec
    y = layout.pack_y0(col)
    if (
        bool(getattr(cfg, "enable_startup_seed_cache", False))
        and (not explicit_runtime_restart)
        and startup_seed_cache_path is not None
        and (not bool(getattr(cfg, "refresh_startup_seed_cache", False)))
    ):
        startup_seed_load_t0 = time.perf_counter()
        loaded_seed, startup_seed_cache_info = _load_startup_seed_cache(
            path=startup_seed_cache_path,
            cfg=cfg,
            col=col,
            layout=layout,
            base_inputs=base_inputs,
        )
        startup_timing_sec["startup_seed_cache_load"] = float(time.perf_counter() - startup_seed_load_t0)
        if loaded_seed is not None:
            y = np.asarray(loaded_seed["y"], dtype=float).reshape((layout.n_states(),)).copy()
            last_T_tray = loaded_seed.get("last_T_tray")
            last_P_diag = loaded_seed.get("last_P_diag")
            last_P_hyd = loaded_seed.get("last_P_hyd")
            last_K_tray = loaded_seed.get("last_K_tray")
            last_HL = loaded_seed.get("last_HL")
            last_HV = loaded_seed.get("last_HV")
            last_Zfac = loaded_seed.get("last_Zfac")
            last_z_overall = loaded_seed.get("last_z_overall")
            last_tray_bubble_target_F = loaded_seed.get("last_tray_bubble_target_F")
            last_tray_thermo_packet = loaded_seed.get("last_tray_thermo_packet")
            last_condenser_duty_packet = loaded_seed.get("last_condenser_duty_packet")
            last_feed_stage_flash_packet = loaded_seed.get("last_feed_stage_flash_packet")
            last_bottom_sump_cp_packet = loaded_seed.get("last_bottom_sump_cp_packet")
            last_reb_T = loaded_seed.get("last_reb_T")
            last_reb_x = loaded_seed.get("last_reb_x")
            last_reb_y = loaded_seed.get("last_reb_y")
            last_reb_beta = loaded_seed.get("last_reb_beta")
            startup_seeded_condenser_duty_packet = bool(
                loaded_seed.get("startup_seeded_condenser_duty_packet", False)
            )
            startup_seed_loaded = True
            _emit_progress(
                "[Init] Loaded startup seed cache  "
                f"path={startup_seed_cache_info.get('path', '')}  "
                "skipping fresh-startup vapor reseed and heavy conditioning"
            )
    if not bool(cfg.use_excel_vapor_holdup) and (not explicit_runtime_restart) and (not startup_seed_loaded):
        y = _clear_initial_tray_vapor_holdup(y, layout)
    _milestone("packed initial state")

    # Make MV consistent with P_spec at t=0 (uses ideal-gas Z seed at startup).
    init_inputs = base_inputs
    startup_thermo_diag: Dict[str, Any] = {}
    if (not explicit_runtime_restart) and (not startup_seed_loaded):
        vapor_holdup_init_t0 = time.perf_counter()
        _emit_progress("[Init] Starting vapor holdup initialization from startup pressure specification")
        y, startup_thermo_diag = _initialize_vapor_holdup_from_spec_pressure(
            col=col,
            layout=layout,
            y=y,
            inputs=init_inputs,
            include_temperature=bool(cfg.include_temperature),
            return_diag=True,
            preserve_tray_vapor_holdup=bool(cfg.use_excel_vapor_holdup),
        )
        startup_timing_sec["vapor_holdup_initialization"] = float(time.perf_counter() - vapor_holdup_init_t0)
        if "startup_vapor_holdup_refresh_wall_sec" in startup_thermo_diag:
            try:
                startup_timing_sec["startup_tray_refresh"] = float(
                    startup_thermo_diag["startup_vapor_holdup_refresh_wall_sec"]
                )
            except Exception:
                pass
        if "startup_vapor_holdup_refresh_source" in startup_thermo_diag:
            startup_timing_sec["startup_tray_refresh_source"] = str(
                startup_thermo_diag["startup_vapor_holdup_refresh_source"]
            )
        if "startup_vapor_holdup_refresh_batch_used" in startup_thermo_diag:
            startup_timing_sec["startup_tray_refresh_batch_used"] = bool(
                startup_thermo_diag["startup_vapor_holdup_refresh_batch_used"]
            )
        _emit_progress("[Init] Completed vapor holdup initialization from startup pressure specification")
        try:
            T_seed = _tray_temperature_F(col, layout, y, bool(cfg.include_temperature))
        except Exception:
            T_seed = None
        P_seed = None
        for key in ("P_psia_hyd", "P_psia_diag"):
            if key in startup_thermo_diag:
                try:
                    P_seed = np.asarray(startup_thermo_diag[key], dtype=float).reshape((col.n_stages,))
                    break
                except Exception:
                    P_seed = None
        if P_seed is None:
            try:
                P_seed = np.asarray(getattr(col, "P_psia"), dtype=float).reshape((col.n_stages,))
            except Exception:
                P_seed = None
        startup_packet = _tray_thermo_packet_from_diag(
            startup_thermo_diag,
            n_stages=col.n_stages,
            n_components=col.n_components,
            T_tray_F=T_seed,
            P_tray_psia=P_seed,
        )
        if startup_packet is not None:
            try:
                u_seed = layout.unpack(y)
                tray_L_seed = np.asarray(u_seed["tray_L"], dtype=float).reshape((col.n_stages, col.n_components))
                tray_V_seed = np.asarray(u_seed["tray_V"], dtype=float).reshape((col.n_stages, col.n_components))
                x_seed = np.asarray(u_seed["x_tray"], dtype=float).reshape((col.n_stages, col.n_components))
                z_seed = np.zeros((col.n_stages, col.n_components), dtype=float)
                for i in range(col.n_stages):
                    z_i = tray_L_seed[i, :] + tray_V_seed[i, :]
                    s_i = float(np.sum(z_i))
                    if (not np.isfinite(s_i)) or s_i <= 1e-300:
                        z_i = x_seed[i, :]
                        s_i = float(np.sum(z_i))
                    z_seed[i, :] = z_i / max(s_i, 1e-300)
                startup_packet.z_overall_tray = z_seed
            except Exception:
                pass
            last_tray_thermo_packet = startup_packet
            last_K_tray = startup_packet.K_tray.copy()
            last_HL = startup_packet.HL.copy()
            last_HV = startup_packet.HV.copy()
            last_Zfac = startup_packet.Zfac_tray.copy()
            last_z_overall = startup_packet.z_overall.copy()
            if startup_packet.T_state is not None:
                last_T_tray = np.asarray(startup_packet.T_state, dtype=float).reshape((col.n_stages,)).copy()
            if startup_packet.P_state is not None:
                p_seed_arr = np.asarray(startup_packet.P_state, dtype=float).reshape((col.n_stages,)).copy()
                last_P_diag = p_seed_arr.copy()
                last_P_hyd = p_seed_arr.copy()
            seeded_condenser_packet = _seed_startup_condenser_duty_packet(
                col=col,
                layout=layout,
                y=y,
                startup_packet=startup_packet,
                condenser_duty_mode=getattr(base_inputs, "condenser_duty_mode", None),
            )
            if seeded_condenser_packet is not None:
                last_condenser_duty_packet = seeded_condenser_packet
                startup_seeded_condenser_duty_packet = True
    thermo_init_info = {
        "attempted": False,
        "success": False,
        "n_iter": 0,
        "max_dx": np.nan,
        "max_dy": np.nan,
        "eq_phase_change_init_lbmolps": np.nan,
        "eq_phase_change_final_lbmolps": np.nan,
    }
    hydraulic_energy_init_info = {
        "attempted": False,
        "success": False,
        "n_iter": 0,
        "objective_init": np.nan,
        "objective_final": np.nan,
        "mass_resid_init_lbmolph": np.nan,
        "mass_resid_final_lbmolph": np.nan,
        "energy_resid_init_btups": np.nan,
        "energy_resid_final_btups": np.nan,
        "alg_p_init_psia": np.nan,
        "alg_p_final_psia": np.nan,
        "alg_v_init_lbmolph": np.nan,
        "alg_v_final_lbmolph": np.nan,
    }
    startup_flags = _resolve_startup_execution_flags(cfg)
    if bool(startup_flags.get("fast_startup", False)) and (not explicit_runtime_restart) and (not startup_seed_loaded):
        _emit_progress(
            "[Init] Fast startup enabled  "
            "skipping startup thermo conditioning, hydraulic-energy startup consistency, "
            "and top-drum startup steadying"
        )
    restart_reentry_info = {
        "attempted": False,
        "success": False,
        "thermo": {"attempted": False, "success": False, "n_iter": 0},
        "top_drum": {"attempted": False, "success": False, "n_iter": 0, "hit_wall_limit": False},
    }
    if (
        bool(startup_flags.get("enable_startup_thermo_conditioning", False))
        and (not explicit_runtime_restart)
        and (not startup_seed_loaded)
    ):
        _emit_progress("[Init] Entering startup thermo conditioning")
        y, thermo_init_info = _initialize_thermo_consistent_state(
            col=col,
            layout=layout,
            y=y,
            inputs=init_inputs,
            include_temperature=bool(cfg.include_temperature),
            max_iter=int(startup_flags.get("startup_thermo_conditioning_iters", cfg.startup_thermo_conditioning_iters)),
            relaxation=float(
                startup_flags.get(
                    "startup_thermo_conditioning_relaxation",
                    cfg.startup_thermo_conditioning_relaxation,
                )
            ),
            preserve_tray_vapor_holdup=bool(cfg.use_excel_vapor_holdup),
        )
        if bool(thermo_init_info.get("attempted", False)):
            eq0 = float(thermo_init_info.get("eq_phase_change_init_lbmolps", np.nan))
            eqf = float(thermo_init_info.get("eq_phase_change_final_lbmolps", np.nan))
            mdx = float(thermo_init_info.get("max_dx", np.nan))
            mdy = float(thermo_init_info.get("max_dy", np.nan))
            nit = int(thermo_init_info.get("n_iter", 0))
            _emit_progress(
                "[Init] Thermo startup conditioning  "
                f"eq_phase: {eq0:+.6g}->{eqf:+.6g} lbmol/s  "
                f"max_dx={mdx:.3g}  max_dy={mdy:.3g}  "
                f"iters={nit}  success={bool(thermo_init_info.get('success', False))}"
            )
    if (
        bool(startup_flags.get("enable_startup_hydraulic_energy_consistency", False))
        and (not explicit_runtime_restart)
        and (not startup_seed_loaded)
    ):
        _emit_progress("[Init] Entering hydraulic-energy startup consistency")
        y, hydraulic_energy_init_info = _initialize_hydraulic_energy_consistent_state(
            col=col,
            layout=layout,
            y=y,
            inputs=init_inputs,
            include_temperature=bool(cfg.include_temperature),
            max_iter=int(cfg.startup_hydraulic_energy_consistency_iters),
            pseudo_dt_sec=float(cfg.startup_hydraulic_energy_consistency_dt_sec),
            mass_tol_lbmolph=cfg.startup_hydraulic_energy_consistency_mass_tol_lbmolph,
            energy_tol_btups=cfg.startup_hydraulic_energy_consistency_energy_tol_btups,
            dae_max_iter=int(max(3, getattr(cfg, "dae_pilot_max_iter", 3))),
            dae_p_tol_psia=getattr(cfg, "dae_pilot_p_tol_psia", 0.05),
            dae_v_tol_lbmolph=getattr(cfg, "dae_pilot_v_tol_lbmolph", 25.0),
            dae_jac_rel_step=float(getattr(cfg, "dae_pilot_jac_rel_step", 1.0e-6)),
            dae_line_search_max=int(getattr(cfg, "dae_pilot_line_search_max", 4)),
        )
        if bool(hydraulic_energy_init_info.get("attempted", False)):
            print(
                "[Init] Hydraulic-energy startup consistency  "
                f"obj: {float(hydraulic_energy_init_info.get('objective_init', np.nan)):.3g}"
                f"->{float(hydraulic_energy_init_info.get('objective_final', np.nan)):.3g}  "
                f"mass: {float(hydraulic_energy_init_info.get('mass_resid_init_lbmolph', np.nan)):.3g}"
                f"->{float(hydraulic_energy_init_info.get('mass_resid_final_lbmolph', np.nan)):.3g} lbmol/h  "
                f"energy: {float(hydraulic_energy_init_info.get('energy_resid_init_btups', np.nan)):.3g}"
                f"->{float(hydraulic_energy_init_info.get('energy_resid_final_btups', np.nan)):.3g} BTU/s  "
                f"iters={int(hydraulic_energy_init_info.get('n_iter', 0))}  "
                f"success={bool(hydraulic_energy_init_info.get('success', False))}"
            )
    if _has_explicit_top_boundary_restart_state(col):
        top_drum_init_info = {
            "attempted": False,
            "success": True,
            "skipped_explicit_restart_state": True,
            "n_iter": 0,
            "pressure_coupled": False,
            "d_top_L_init_lbmolps": np.nan,
            "d_top_V_init_lbmolps": np.nan,
            "d_top_L_final_lbmolps": np.nan,
            "d_top_V_final_lbmolps": np.nan,
        }
    elif startup_seed_loaded:
        top_drum_init_info = {
            "attempted": False,
            "success": True,
            "skipped_startup_seed_cache": True,
            "n_iter": 0,
            "pressure_coupled": False,
            "d_top_L_init_lbmolps": np.nan,
            "d_top_V_init_lbmolps": np.nan,
            "d_top_L_final_lbmolps": np.nan,
            "d_top_V_final_lbmolps": np.nan,
        }
    elif not bool(startup_flags.get("enable_top_drum_startup_steadying", True)):
        top_drum_init_info = {
            "attempted": False,
            "success": True,
            "skipped_fast_startup": True,
            "n_iter": 0,
            "pressure_coupled": False,
            "d_top_L_init_lbmolps": np.nan,
            "d_top_V_init_lbmolps": np.nan,
            "d_top_L_final_lbmolps": np.nan,
            "d_top_V_final_lbmolps": np.nan,
        }
    else:
        y, top_drum_init_info = _initialize_top_drum_dynamic_steady(
            col=col,
            layout=layout,
            y=y,
            inputs=init_inputs,
            max_iter=int(startup_flags.get("top_drum_steady_max_iter", 6)),
            tol_lbmolps=float(startup_flags.get("top_drum_steady_tol_lbmolps", 1.0e-6)),
            wall_limit_sec=startup_flags.get("top_drum_steady_wall_limit_sec"),
        )
    if bool(top_drum_init_info.get("attempted", False)):
        dL0 = float(top_drum_init_info.get("d_top_L_init_lbmolps", np.nan))
        dV0 = float(top_drum_init_info.get("d_top_V_init_lbmolps", np.nan))
        dLf = float(top_drum_init_info.get("d_top_L_final_lbmolps", np.nan))
        dVf = float(top_drum_init_info.get("d_top_V_final_lbmolps", np.nan))
        iters = int(top_drum_init_info.get("n_iter", 0))
        print(
            "[Init] Top-drum startup steadying  "
            f"dL: {dL0:+.6g}->{dLf:+.6g} lbmol/s  "
            f"dV: {dV0:+.6g}->{dVf:+.6g} lbmol/s  "
            f"iters={iters}  success={bool(top_drum_init_info.get('success', False))}"
            f"  wall_cap_hit={bool(top_drum_init_info.get('hit_wall_limit', False))}"
        )
    elif bool(top_drum_init_info.get("skipped_explicit_restart_state", False)):
        print("[Init] Top-drum startup steadying skipped  explicit boundary restart state detected")
    elif bool(top_drum_init_info.get("skipped_startup_seed_cache", False)):
        print("[Init] Top-drum startup steadying skipped  startup seed cache loaded")
    elif bool(top_drum_init_info.get("skipped_fast_startup", False)):
        print("[Init] Top-drum startup steadying skipped  fast startup enabled")
    if explicit_runtime_restart:
        if bool(startup_flags.get("enable_restart_reentry_settling", True)):
            y, restart_reentry_info = _initialize_restart_reentry_settling(
                col=col,
                layout=layout,
                y=y,
                inputs=init_inputs,
                include_temperature=bool(cfg.include_temperature),
                thermo_max_iter=int(startup_flags.get("restart_reentry_thermo_conditioning_iters", 1)),
                thermo_relaxation=float(startup_flags.get("restart_reentry_thermo_conditioning_relaxation", 1.0)),
                top_drum_max_iter=int(startup_flags.get("restart_reentry_top_drum_max_iter", 2)),
                top_drum_tol_lbmolps=float(startup_flags.get("restart_reentry_top_drum_tol_lbmolps", 1.0e-4)),
                top_drum_wall_limit_sec=startup_flags.get("restart_reentry_top_drum_wall_limit_sec"),
                preserve_tray_vapor_holdup=True,
            )
            if bool(restart_reentry_info.get("attempted", False)):
                thermo_part = restart_reentry_info.get("thermo", {})
                top_part = restart_reentry_info.get("top_drum", {})
                print(
                    "[Init] Restart re-entry settling  "
                    f"thermo_iters={int(thermo_part.get('n_iter', 0))}  "
                    f"top_drum_iters={int(top_part.get('n_iter', 0))}  "
                    f"success={bool(restart_reentry_info.get('success', False))}  "
                    f"wall_cap_hit={bool(top_part.get('hit_wall_limit', False))}"
                )
        print("[Init] Explicit runtime restart state detected  skipped fresh-startup vapor reseed and heavy conditioning")
    _milestone("initialized vapor holdup from spec pressure")
    if (
        (not startup_seed_loaded or last_tray_bubble_target_F is None)
        and str(base_inputs.pressure_model).strip().lower() == "hydraulic"
        and str(base_inputs.vapor_flow_model).strip().lower() == "energy"
        and str(base_inputs.hydraulic_energy_temperature_mode).strip().lower() == "bubble-point-follower"
        and base_inputs.thermo_provider is not None
        and bool(cfg.include_temperature)
    ):
        try:
            p_init = np.asarray(getattr(col, "P_psia", np.full(col.n_stages, 200.0, dtype=float)), dtype=float).reshape((col.n_stages,))
            last_tray_bubble_target_F = _refresh_tray_bubble_targets_F(
                col=col,
                layout=layout,
                y=y,
                thermo_provider=base_inputs.thermo_provider,
                P_tray_psia=p_init,
            )
            if last_tray_bubble_target_F is not None and np.any(np.isfinite(last_tray_bubble_target_F)):
                _milestone("initialized cached tray bubble-point targets")
        except Exception:
            last_tray_bubble_target_F = None

    if (not startup_seed_loaded) and base_inputs.thermo_provider is not None:
        try:
            feed_seed_pressure = (
                np.asarray(last_P_hyd, dtype=float).copy()
                if last_P_hyd is not None
                else (
                    np.asarray(last_P_diag, dtype=float).copy()
                    if last_P_diag is not None
                    else np.asarray(getattr(col, "P_psia", np.full(col.n_stages, 200.0, dtype=float)), dtype=float).copy()
                )
            )
            last_feed_stage_flash_packet = _seed_startup_feed_stage_flash_packet(
                col=col,
                thermo_provider=base_inputs.thermo_provider,
                P_tray_psia=feed_seed_pressure,
            )
        except Exception:
            last_feed_stage_flash_packet = None

    if (
        bool(getattr(cfg, "enable_startup_seed_cache", False))
        and (not explicit_runtime_restart)
        and startup_seed_cache_path is not None
        and ((not startup_seed_loaded) or bool(getattr(cfg, "refresh_startup_seed_cache", False)))
    ):
        startup_seed_write_t0 = time.perf_counter()
        startup_seed_cache_save = _write_startup_seed_cache(
            path=startup_seed_cache_path,
            cfg=cfg,
            col=col,
            layout=layout,
            base_inputs=base_inputs,
            y=y,
            last_T_tray=last_T_tray,
            last_P_diag=last_P_diag,
            last_P_hyd=last_P_hyd,
            last_K_tray=last_K_tray,
            last_HL=last_HL,
            last_HV=last_HV,
            last_Zfac=last_Zfac,
            last_z_overall=last_z_overall,
            last_tray_bubble_target_F=last_tray_bubble_target_F,
            last_tray_thermo_packet=last_tray_thermo_packet,
            last_condenser_duty_packet=last_condenser_duty_packet,
            last_feed_stage_flash_packet=last_feed_stage_flash_packet,
            last_bottom_sump_cp_packet=last_bottom_sump_cp_packet,
            last_reb_T=last_reb_T,
            last_reb_x=last_reb_x,
            last_reb_y=last_reb_y,
            last_reb_beta=last_reb_beta,
            startup_seeded_condenser_duty_packet=startup_seeded_condenser_duty_packet,
        )
        startup_timing_sec["startup_seed_cache_write"] = float(time.perf_counter() - startup_seed_write_t0)
        startup_seed_cache_info.update(startup_seed_cache_save)
        if bool(startup_seed_cache_info.get("saved", False)):
            _emit_progress(
                "[Init] Wrote startup seed cache  "
                f"path={startup_seed_cache_info.get('path', '')}"
            )

    # Resolve streams for logging placement
    feed_tag, dist_tag, bots_tag = _resolve_logging_streams(case, col)
    _milestone("resolved logging stream placement")
    feed_step_time_sec = cfg.feed_step_time_sec
    feed_step_scale = cfg.feed_step_scale
    if feed_step_time_sec is not None:
        try:
            feed_step_time_sec = float(feed_step_time_sec)
        except Exception:
            feed_step_time_sec = None
    if feed_step_scale is not None:
        try:
            feed_step_scale = float(feed_step_scale)
        except Exception:
            feed_step_scale = None
    feed_step_enabled = (
        feed_step_time_sec is not None
        and np.isfinite(float(feed_step_time_sec))
        and float(feed_step_time_sec) >= 0.0
        and feed_step_scale is not None
        and np.isfinite(float(feed_step_scale))
        and float(feed_step_scale) > 0.0
        and abs(float(feed_step_scale) - 1.0) > 1e-12
    )
    base_feed_stream = None
    base_feed_total_lbmolph: Optional[float] = None
    base_feed_component_flows: Optional[Dict[str, float]] = None
    if feed_step_enabled:
        try:
            base_feed_stream = (getattr(col, "streams", {}) or {}).get("Feed")
        except Exception:
            base_feed_stream = None
        if base_feed_stream is not None:
            try:
                base_feed_total_lbmolph = float(base_feed_stream.total_molar_flow_lbmolph)
            except Exception:
                base_feed_total_lbmolph = None
            try:
                comp_raw = getattr(base_feed_stream, "component_molar_flows_lbmolph", None)
                if isinstance(comp_raw, dict):
                    base_feed_component_flows = {str(k): float(v) for k, v in comp_raw.items()}
            except Exception:
                base_feed_component_flows = None
        if base_feed_total_lbmolph is None and feed_tag.flow_lbmolph is not None:
            try:
                base_feed_total_lbmolph = float(feed_tag.flow_lbmolph)
            except Exception:
                base_feed_total_lbmolph = None
        if base_feed_stream is None or base_feed_total_lbmolph is None:
            feed_step_enabled = False
    level_control_enabled, top_level_ctrl, bot_level_ctrl, top_level_sp, bot_level_sp = _build_level_controllers(
        col=col,
        cfg=cfg,
        layout=layout,
        y0=y,
        dist_tag=dist_tag,
        bots_tag=bots_tag,
    )
    top_level_pv_mode_spec = None
    bottom_level_pv_mode_spec = None
    if not bool(getattr(cfg, "ignore_workbook_level_pv_mode", False)):
        top_level_pv_mode_spec = _spec_get(
            getattr(col, "specs_raw", None) or {},
            "Top Level PV Mode",
            "Top Drum PV Mode",
            "Reflux Drum PV Mode",
        )
        bottom_level_pv_mode_spec = _spec_get(
            getattr(col, "specs_raw", None) or {},
            "Bottom Level PV Mode",
            "Bottom Sump PV Mode",
        )
    top_level_pv_mode = str(
        top_level_pv_mode_spec
        if top_level_pv_mode_spec is not None and str(getattr(cfg, "top_level_pv_mode", "molar-holdup") or "molar-holdup").strip().lower() == "molar-holdup"
        else (getattr(cfg, "top_level_pv_mode", "molar-holdup") or "molar-holdup")
    ).strip().lower()
    if top_level_pv_mode not in {"molar-holdup", "true-level"}:
        top_level_pv_mode = "molar-holdup"
    bottom_level_pv_mode = str(
        bottom_level_pv_mode_spec
        if bottom_level_pv_mode_spec is not None and str(getattr(cfg, "bottom_level_pv_mode", "molar-holdup") or "molar-holdup").strip().lower() == "molar-holdup"
        else (getattr(cfg, "bottom_level_pv_mode", "molar-holdup") or "molar-holdup")
    ).strip().lower()
    if bottom_level_pv_mode not in {"molar-holdup", "true-level"}:
        bottom_level_pv_mode = "molar-holdup"
    top_level_rho_ref_lbmol_ft3: Optional[float] = None
    top_level_scale_lbmol_per_frac: Optional[float] = None
    top_level_sp_frac: Optional[float] = None
    bottom_level_rho_ref_lbmol_ft3: Optional[float] = None
    bottom_level_scale_lbmol_per_frac: Optional[float] = None
    bottom_level_sp_frac: Optional[float] = None
    if (
        level_control_enabled
        and top_level_pv_mode == "true-level"
        and base_inputs.top_drum_total_volume_ft3 is not None
    ):
        try:
            _top_v0, top_level_rho_ref_lbmol_ft3, _top_frac0 = _estimate_top_drum_liquid_volume_ft3(
                col=col,
                layout=layout,
                y_vec=y,
                thermo_provider=base_inputs.thermo_provider,
                top_drum_total_volume_ft3=base_inputs.top_drum_total_volume_ft3,
                p_top_psia=(float(np.asarray(col.P_psia, dtype=float).reshape((col.n_stages,))[0]) if hasattr(col, "P_psia") else None),
            )
        except Exception:
            top_level_rho_ref_lbmol_ft3 = None
        if top_level_rho_ref_lbmol_ft3 is None or (not np.isfinite(float(top_level_rho_ref_lbmol_ft3))) or float(top_level_rho_ref_lbmol_ft3) <= 1e-12:
            top_level_pv_mode = "molar-holdup"
        else:
            try:
                top_level_scale_lbmol_per_frac = float(top_level_rho_ref_lbmol_ft3) * float(base_inputs.top_drum_total_volume_ft3)
            except Exception:
                top_level_scale_lbmol_per_frac = None
            top_level_sp_frac = cfg.top_level_sp_frac
            if top_level_sp_frac is None:
                top_level_sp_frac = _spec_float(
                    getattr(col, "specs_raw", None) or {},
                    "Top Drum Liquid Fraction (-)",
                    "Top Drum Liquid Volume Fraction",
                    "Top Drum Liquid Fraction",
                    "Top Accumulator Liquid Volume Fraction",
                    "Top Accumulator Liquid Fraction",
                    "Reflux Drum Liquid Volume Fraction",
                    "Reflux Drum Liquid Fraction",
                    "Distillate Drum Liquid Volume Fraction",
                    "Distillate Drum Liquid Fraction",
                    "Top Drum Fill Fraction",
                    "Reflux Drum Fill Fraction",
                    "Distillate Drum Fill Fraction",
                )
            if top_level_sp_frac is not None and top_level_sp_frac > 1.0 and top_level_sp_frac <= 100.0:
                top_level_sp_frac = float(top_level_sp_frac) / 100.0
            if top_level_sp_frac is None:
                try:
                    _top_v0, _top_rho0, _top_frac0 = _estimate_top_drum_liquid_volume_ft3(
                        col=col,
                        layout=layout,
                        y_vec=y,
                        thermo_provider=base_inputs.thermo_provider,
                        top_drum_total_volume_ft3=base_inputs.top_drum_total_volume_ft3,
                        p_top_psia=(float(np.asarray(col.P_psia, dtype=float).reshape((col.n_stages,))[0]) if hasattr(col, "P_psia") else None),
                    )
                    top_level_sp_frac = _top_frac0
                except Exception:
                    top_level_sp_frac = None
            if (
                top_level_sp_frac is None
                or (not np.isfinite(float(top_level_sp_frac)))
                or float(top_level_sp_frac) < 0.0
                or float(top_level_sp_frac) > 1.0
            ):
                top_level_pv_mode = "molar-holdup"
            else:
                top_level_sp = float(top_level_sp_frac)
                if top_level_ctrl is not None:
                    try:
                        vfrac_ref = _horizontal_cylinder_volume_fraction_from_height_fraction(float(top_level_sp_frac))
                        d_h = 1.0e-4
                        h_lo = max(float(top_level_sp_frac) - d_h, 0.0)
                        h_hi = min(float(top_level_sp_frac) + d_h, 1.0)
                        if h_hi > h_lo:
                            v_lo = _horizontal_cylinder_volume_fraction_from_height_fraction(h_lo)
                            v_hi = _horizontal_cylinder_volume_fraction_from_height_fraction(h_hi)
                            dv_dh = (float(v_hi) - float(v_lo)) / max(float(h_hi - h_lo), 1.0e-12)
                        else:
                            dv_dh = np.nan
                    except Exception:
                        dv_dh = np.nan
                    gain_scale = None
                    if (
                        top_level_scale_lbmol_per_frac is not None
                        and np.isfinite(float(top_level_scale_lbmol_per_frac))
                        and float(top_level_scale_lbmol_per_frac) > 1e-12
                        and np.isfinite(float(dv_dh))
                        and float(dv_dh) > 1e-12
                    ):
                        gain_scale = float(top_level_scale_lbmol_per_frac) * float(dv_dh)
                    if gain_scale is not None and np.isfinite(float(gain_scale)) and float(gain_scale) > 1e-12:
                        top_level_ctrl.kc = float(top_level_ctrl.kc) * float(gain_scale)
    if (
        level_control_enabled
        and bottom_level_pv_mode == "true-level"
        and base_inputs.bottom_sump_total_volume_ft3 is not None
    ):
        try:
            _bot_v0, bottom_level_rho_ref_lbmol_ft3, _bot_frac0 = _estimate_bottom_sump_liquid_volume_ft3(
                col=col,
                layout=layout,
                y_vec=y,
                thermo_provider=base_inputs.thermo_provider,
                bottom_sump_total_volume_ft3=base_inputs.bottom_sump_total_volume_ft3,
                p_bottom_psia=(float(np.asarray(col.P_psia, dtype=float).reshape((col.n_stages,))[-1]) if hasattr(col, "P_psia") else None),
            )
        except Exception:
            bottom_level_rho_ref_lbmol_ft3 = None
        if (
            bottom_level_rho_ref_lbmol_ft3 is None
            or (not np.isfinite(float(bottom_level_rho_ref_lbmol_ft3)))
            or float(bottom_level_rho_ref_lbmol_ft3) <= 1e-12
        ):
            bottom_level_pv_mode = "molar-holdup"
        else:
            try:
                bottom_level_scale_lbmol_per_frac = float(bottom_level_rho_ref_lbmol_ft3) * float(base_inputs.bottom_sump_total_volume_ft3)
            except Exception:
                bottom_level_scale_lbmol_per_frac = None
            bottom_level_sp_frac = cfg.bottom_level_sp_frac
            if bottom_level_sp_frac is None:
                bottom_level_sp_frac = _spec_float(
                    getattr(col, "specs_raw", None) or {},
                    "Bottom Sump Liquid Fraction (-)",
                    "Bottom Sump Liquid Volume Fraction",
                    "Bottom Sump Liquid Fraction",
                    "Bottom Sump Fill Fraction",
                    "Bottom Liquid Fraction (-)",
                    "Bottom Liquid Volume Fraction",
                    "Bottom Liquid Fraction",
                    "Bottom Fill Fraction",
                )
            if bottom_level_sp_frac is not None and bottom_level_sp_frac > 1.0 and bottom_level_sp_frac <= 100.0:
                bottom_level_sp_frac = float(bottom_level_sp_frac) / 100.0
            if bottom_level_sp_frac is None and bot_level_sp is not None and bottom_level_scale_lbmol_per_frac is not None:
                try:
                    vfrac_guess = float(bot_level_sp) / max(float(bottom_level_scale_lbmol_per_frac), 1.0e-12)
                    vfrac_guess = float(np.clip(vfrac_guess, 0.0, 1.0))
                    bottom_level_sp_frac = float(vfrac_guess)
                except Exception:
                    bottom_level_sp_frac = None
            if bottom_level_sp_frac is None:
                try:
                    _bot_v0, _bot_rho0, _bot_frac0 = _estimate_bottom_sump_liquid_volume_ft3(
                        col=col,
                        layout=layout,
                        y_vec=y,
                        thermo_provider=base_inputs.thermo_provider,
                        bottom_sump_total_volume_ft3=base_inputs.bottom_sump_total_volume_ft3,
                        p_bottom_psia=(float(np.asarray(col.P_psia, dtype=float).reshape((col.n_stages,))[-1]) if hasattr(col, "P_psia") else None),
                    )
                    if _bot_frac0 is not None and np.isfinite(float(_bot_frac0)):
                        bottom_level_sp_frac = float(_bot_frac0)
                except Exception:
                    bottom_level_sp_frac = None
            if (
                bottom_level_sp_frac is None
                or (not np.isfinite(float(bottom_level_sp_frac)))
                or float(bottom_level_sp_frac) < 0.0
                or float(bottom_level_sp_frac) > 1.0
            ):
                bottom_level_pv_mode = "molar-holdup"
            else:
                bot_level_sp = float(bottom_level_sp_frac)
                if bot_level_ctrl is not None:
                    dv_dh = 1.0
                    gain_scale = None
                    if (
                        bottom_level_scale_lbmol_per_frac is not None
                        and np.isfinite(float(bottom_level_scale_lbmol_per_frac))
                        and float(bottom_level_scale_lbmol_per_frac) > 1e-12
                        and np.isfinite(float(dv_dh))
                        and float(dv_dh) > 1e-12
                    ):
                        gain_scale = float(bottom_level_scale_lbmol_per_frac) * float(dv_dh)
                    if gain_scale is not None and np.isfinite(float(gain_scale)) and float(gain_scale) > 1e-12:
                        bot_level_ctrl.kc = float(bot_level_ctrl.kc) * float(gain_scale)
    last_valid_top_level_pv: Optional[float] = None
    last_valid_bottom_level_pv: Optional[float] = None
    if level_control_enabled:
        try:
            u_level0 = layout.unpack(np.asarray(y, dtype=float))
        except Exception:
            u_level0 = {}
        if top_level_pv_mode == "true-level":
            try:
                p_top0 = float(np.asarray(col.P_psia, dtype=float).reshape((col.n_stages,))[0]) if hasattr(col, "P_psia") else None
            except Exception:
                p_top0 = None
            try:
                _top_v0, _top_rho0, _top_frac0 = _estimate_top_drum_liquid_volume_ft3(
                    col=col,
                    layout=layout,
                    y_vec=y,
                    thermo_provider=base_inputs.thermo_provider,
                    top_drum_total_volume_ft3=base_inputs.top_drum_total_volume_ft3,
                    p_top_psia=p_top0,
                )
                if _top_frac0 is not None and np.isfinite(float(_top_frac0)):
                    last_valid_top_level_pv = float(
                        _horizontal_cylinder_height_fraction_from_volume_fraction(float(_top_frac0))
                    )
            except Exception:
                last_valid_top_level_pv = None
        if bottom_level_pv_mode == "true-level":
            try:
                p_bot0 = float(np.asarray(col.P_psia, dtype=float).reshape((col.n_stages,))[-1]) if hasattr(col, "P_psia") else None
            except Exception:
                p_bot0 = None
            try:
                _bot_v0, _bot_rho0, _bot_frac0 = _estimate_bottom_sump_liquid_volume_ft3(
                    col=col,
                    layout=layout,
                    y_vec=y,
                    thermo_provider=base_inputs.thermo_provider,
                    bottom_sump_total_volume_ft3=base_inputs.bottom_sump_total_volume_ft3,
                    p_bottom_psia=p_bot0,
                )
                if _bot_frac0 is not None and np.isfinite(float(_bot_frac0)):
                    last_valid_bottom_level_pv = float(np.clip(float(_bot_frac0), 0.0, 1.0))
            except Exception:
                last_valid_bottom_level_pv = None
    if level_control_enabled:
        if top_level_pv_mode == "true-level" and top_level_sp_frac is not None:
            print(
                "[Control] Level control enabled  "
                f"top_SP={100.0*float(top_level_sp_frac):.2f}% full  "
                + (
                    f"bottom_SP={100.0*float(bottom_level_sp_frac):.2f}% full  "
                    if bottom_level_pv_mode == "true-level" and bottom_level_sp_frac is not None
                    else f"bottom_SP={float(bot_level_sp):.3f} lbmol  "
                )
                + f"top_PV_mode={top_level_pv_mode}  bottom_PV_mode={bottom_level_pv_mode}"
            )
        else:
            print(
                "[Control] Level control enabled  "
                + (
                    f"top_SP={float(top_level_sp):.3f} lbmol  "
                    if top_level_sp is not None
                    else "top_SP=nan  "
                )
                + (
                    f"bottom_SP={100.0*float(bottom_level_sp_frac):.2f}% full  "
                    if bottom_level_pv_mode == "true-level" and bottom_level_sp_frac is not None
                    else f"bottom_SP={float(bot_level_sp):.3f} lbmol  "
                )
                + f"top_PV_mode={top_level_pv_mode}  bottom_PV_mode={bottom_level_pv_mode}"
            )

    init_pack_top_drum_vapor_info: Dict[str, Any] = {
        "enabled": bool(getattr(cfg, "init_pack_top_drum_vapor_to_pressure", False)),
        "applied": False,
        "target_pressure_psia": np.nan,
        "raw_pressure_initial_psia": np.nan,
        "raw_pressure_final_psia": np.nan,
        "top_vapor_initial_lbmol": np.nan,
        "top_vapor_final_lbmol": np.nan,
        "scale_factor": np.nan,
        "reason": "disabled",
    }
    if bool(getattr(cfg, "init_pack_top_drum_vapor_to_pressure", False)):
        p_pack = getattr(cfg, "init_top_drum_vapor_pressure_psia", None)
        if p_pack is None:
            p_pack = getattr(cfg, "top_pressure_sp_psia", None)
        if p_pack is None:
            try:
                p_pack = float(np.asarray(getattr(col, "P_psia"), dtype=float).reshape((-1,))[0])
            except Exception:
                p_pack = None
        y, init_pack_top_drum_vapor_info = _pack_top_drum_vapor_to_pressure(
            y=y,
            col=col,
            layout=layout,
            inputs=base_inputs,
            target_pressure_psia=p_pack,
        )
        if bool(init_pack_top_drum_vapor_info.get("applied", False)):
            print(
                "[Init] Packed top-drum vapor inventory to pressure  "
                f"P_raw={float(init_pack_top_drum_vapor_info.get('raw_pressure_initial_psia', np.nan)):.3f}"
                f"->{float(init_pack_top_drum_vapor_info.get('raw_pressure_final_psia', np.nan)):.3f} psia  "
                f"target={float(init_pack_top_drum_vapor_info.get('target_pressure_psia', np.nan)):.3f} psia  "
                f"scale={float(init_pack_top_drum_vapor_info.get('scale_factor', np.nan)):.4g}"
            )
        else:
            print(
                "[Init] Top-drum vapor inventory pressure pack skipped: "
                f"{init_pack_top_drum_vapor_info.get('reason', '')}"
            )

    init_match_condenser_duty_info: Dict[str, Any] = {
        "enabled": bool(getattr(cfg, "init_match_condenser_duty", False)),
        "applied": False,
        "q_calc_BTUph": np.nan,
        "top_pressure_sp_psia": np.nan,
        "reason": "disabled",
    }
    if bool(getattr(cfg, "init_match_condenser_duty", False)):
        p_match = cfg.top_pressure_sp_psia
        if p_match is None:
            try:
                p_match = float(np.asarray(getattr(col, "P_psia"), dtype=float).reshape((-1,))[0])
            except Exception:
                p_match = None
        try:
            p_match_f = float(p_match) if p_match is not None else np.nan
        except Exception:
            p_match_f = np.nan
        init_match_condenser_duty_info["top_pressure_sp_psia"] = (
            float(p_match_f) if np.isfinite(float(p_match_f)) else np.nan
        )
        if (not np.isfinite(float(p_match_f))) or float(p_match_f) <= 0.0:
            init_match_condenser_duty_info["reason"] = "invalid_top_pressure_sp"
            print("[Init] Condenser duty match skipped: invalid top pressure setpoint")
        elif base_inputs.thermo_provider is None:
            init_match_condenser_duty_info["reason"] = "missing_thermo_provider"
            print("[Init] Condenser duty match skipped: missing thermo provider")
        else:
            try:
                match_inputs = replace(
                    base_inputs,
                    condenser_duty_mode="total-condense",
                    condenser_duty_btu_per_h=None,
                    condenser_duty_trim_btu_per_h=None,
                    pressure_top_anchor_psia=float(p_match_f),
                    condenser_duty_prev=None,
                )
                _dydt_match, diag_match = column_rhs(
                    0.0,
                    np.asarray(y, dtype=float),
                    col,
                    layout,
                    match_inputs,
                )
                q_match = _mapping_scalar(diag_match, "Q_cond_calc_BTUph")
                if not np.isfinite(float(q_match)):
                    q_match = _mapping_scalar(diag_match, "Q_cond_used_BTUph")
                if np.isfinite(float(q_match)) and float(q_match) < 0.0:
                    base_inputs = replace(
                        base_inputs,
                        condenser_duty_btu_per_h=float(q_match),
                        condenser_duty_trim_btu_per_h=0.0,
                    )
                    init_match_condenser_duty_info["applied"] = True
                    init_match_condenser_duty_info["q_calc_BTUph"] = float(q_match)
                    init_match_condenser_duty_info["reason"] = "applied"
                    print(
                        "[Init] Condenser duty matched from t=0 total-condenser solve  "
                        f"Q={float(q_match):.6g} BTU/h at P_top={float(p_match_f):.3f} psia"
                    )
                else:
                    init_match_condenser_duty_info["reason"] = "no_finite_negative_duty"
                    print("[Init] Condenser duty match skipped: no finite heat-removal duty from RHS")
            except Exception as exc:
                init_match_condenser_duty_info["reason"] = f"exception:{type(exc).__name__}"
                print(f"[Init] Condenser duty match skipped: {type(exc).__name__}: {exc}")

    init_align_top_liquid_info: Dict[str, Any] = {
        "enabled": bool(getattr(cfg, "init_align_top_liquid_to_condensate", False)),
        "applied": False,
        "top_L_total_lbmol": np.nan,
        "condensate_total_lbmolph": np.nan,
        "max_composition_delta": np.nan,
        "reason": "disabled",
    }
    if bool(getattr(cfg, "init_align_top_liquid_to_condensate", False)):
        try:
            sl_init = layout.slices()
        except Exception:
            sl_init = {}
        if (not getattr(layout, "include_top", False)) or "top_L" not in sl_init:
            init_align_top_liquid_info["reason"] = "missing_top_L_state"
            print("[Init] Top liquid/condensate alignment skipped: missing top_L state")
        else:
            try:
                align_inputs = replace(base_inputs, condenser_duty_prev=None)
                _dydt_align, diag_align = column_rhs(
                    0.0,
                    np.asarray(y, dtype=float),
                    col,
                    layout,
                    align_inputs,
                )
                cond_comp = np.asarray(
                    diag_align.get("top_L_cond_in_lbmolph_comp", []),
                    dtype=float,
                ).reshape((-1,))
                if cond_comp.size != int(col.n_components):
                    init_align_top_liquid_info["reason"] = "missing_condensate_component_diag"
                    print("[Init] Top liquid/condensate alignment skipped: missing condensate split")
                else:
                    cond_total = float(np.sum(np.where(np.isfinite(cond_comp), cond_comp, 0.0)))
                    top_L_old = np.asarray(y[sl_init["top_L"]], dtype=float).reshape((-1,))
                    top_total = float(np.sum(np.where(np.isfinite(top_L_old), top_L_old, 0.0)))
                    if (not np.isfinite(cond_total)) or cond_total <= 0.0:
                        init_align_top_liquid_info["reason"] = "nonpositive_condensate_flow"
                        print("[Init] Top liquid/condensate alignment skipped: nonpositive condensate flow")
                    elif (not np.isfinite(top_total)) or top_total <= 0.0:
                        init_align_top_liquid_info["reason"] = "nonpositive_top_L"
                        print("[Init] Top liquid/condensate alignment skipped: nonpositive top_L")
                    else:
                        x_old = _normalize_comp(np.where(np.isfinite(top_L_old), top_L_old, 0.0))
                        x_cond = _normalize_comp(np.where(np.isfinite(cond_comp), cond_comp, 0.0))
                        y = np.asarray(y, dtype=float).reshape((-1,)).copy()
                        y[sl_init["top_L"]] = float(top_total) * np.asarray(x_cond, dtype=float)
                        init_align_top_liquid_info["applied"] = True
                        init_align_top_liquid_info["top_L_total_lbmol"] = float(top_total)
                        init_align_top_liquid_info["condensate_total_lbmolph"] = float(cond_total)
                        init_align_top_liquid_info["max_composition_delta"] = float(
                            np.max(np.abs(np.asarray(x_cond, dtype=float) - np.asarray(x_old, dtype=float)))
                        )
                        init_align_top_liquid_info["reason"] = "applied"
                        print(
                            "[Init] Aligned top liquid composition to live condenser condensate  "
                            f"M_top_L={float(top_total):.6g} lbmol  "
                            f"condensate={float(cond_total):.6g} lbmol/h  "
                            f"max_dx={float(init_align_top_liquid_info['max_composition_delta']):.6g}"
                        )
            except Exception as exc:
                init_align_top_liquid_info["reason"] = f"exception:{type(exc).__name__}"
                print(f"[Init] Top liquid/condensate alignment skipped: {type(exc).__name__}: {exc}")

    (
        pressure_control_enabled,
        top_pressure_ctrl,
        top_pressure_sp,
        pressure_control_mv,
        pressure_mode_note,
    ) = _build_pressure_controller(
        col=col,
        cfg=cfg,
    )
    top_pressure_resid_ref_btups = cfg.top_pressure_resid_ref_btups
    top_pressure_resid_ref_auto = False
    if top_pressure_resid_ref_btups is None:
        specs = getattr(col, "specs_raw", None) or {}
        top_pressure_resid_ref_btups = _spec_float(
            specs,
            "Top Pressure Resid Ref (BTU/s)",
            "Top Pressure Resid Ref (Btu/s)",
            "Top Pressure Residual Ref (BTU/s)",
            "Top Pressure Residual Reference (BTU/s)",
        )
    if top_pressure_resid_ref_btups is None:
        if (
            str(base_inputs.pressure_model).strip().lower() == "hydraulic"
            and str(base_inputs.vapor_flow_model).strip().lower() == "energy"
        ):
            top_pressure_resid_ref_btups = 250.0
            top_pressure_resid_ref_auto = True
    if top_pressure_resid_ref_btups is not None:
        try:
            top_pressure_resid_ref_btups = float(top_pressure_resid_ref_btups)
            if (not np.isfinite(top_pressure_resid_ref_btups)) or top_pressure_resid_ref_btups <= 0.0:
                top_pressure_resid_ref_btups = None
                top_pressure_resid_ref_auto = False
        except Exception:
            top_pressure_resid_ref_btups = None
            top_pressure_resid_ref_auto = False
    if pressure_control_enabled and top_pressure_ctrl is not None and top_pressure_sp is not None:
        if pressure_mode_note:
            print(f"[Warn] {pressure_mode_note}")
        pv_tau_txt = "off"
        if cfg.top_pressure_pv_filter_tau_sec is not None and np.isfinite(float(cfg.top_pressure_pv_filter_tau_sec)):
            pv_tau_txt = f"{float(cfg.top_pressure_pv_filter_tau_sec):.3g}s"
        mv_slew_txt = "off"
        if cfg.top_pressure_mv_slew_limit_per_s is not None and np.isfinite(float(cfg.top_pressure_mv_slew_limit_per_s)):
            mv_slew_txt = f"{float(cfg.top_pressure_mv_slew_limit_per_s):.3g}/s"
        resid_ref_txt = "off"
        if top_pressure_resid_ref_btups is not None and np.isfinite(float(top_pressure_resid_ref_btups)):
            resid_ref_txt = f"{float(top_pressure_resid_ref_btups):.3g} BTU/s"
            if top_pressure_resid_ref_auto:
                resid_ref_txt += " (auto)"
        print(
            "[Control] Pressure control enabled  "
            f"MV={str(pressure_control_mv)}  "
            f"top_P_SP={float(top_pressure_sp):.3f} psia  "
            f"Kc={float(top_pressure_ctrl.kc):.3g}  Ti={float(top_pressure_ctrl.ti_sec):.3g} s  "
            f"PV_tau={pv_tau_txt}  "
            f"MV_slew={mv_slew_txt}  "
            f"resid_ref={resid_ref_txt}  "
            f"resid_min_gain={float(cfg.top_pressure_resid_min_gain):.3g}"
        )
    comp_control_enabled, dist_comp_ctrl, dist_comp_sp, dist_comp_idx, dist_comp_name = _build_distillate_composition_controller(
        col=col,
        cfg=cfg,
        boundary=base_inputs.boundary,
        dist_tag=dist_tag,
    )
    if comp_control_enabled and dist_comp_ctrl is not None and dist_comp_sp is not None and dist_comp_idx is not None:
        print(
            "[Control] Distillate composition control enabled  "
            f"component={str(dist_comp_name)}  "
            f"x_SP={float(dist_comp_sp):.6f}  "
            f"Kc={float(dist_comp_ctrl.kc):.3g}  Ti={float(dist_comp_ctrl.ti_sec):.3g} s  "
            f"MV=reflux_lbmolph  limits=({float(dist_comp_ctrl.out_min):.3f}, {float(dist_comp_ctrl.out_max):.3f})"
        )
        if not bool(getattr(cfg, "enable_reflux_feasibility_cap", True)):
            print("[Control] Distillate reflux feasibility cap disabled")
    (
        bot_comp_control_enabled,
        bot_comp_mv_mode,
        bot_comp_ctrl,
        bot_comp_sp,
        bot_comp_idx,
        bot_comp_name,
    ) = _build_bottoms_composition_controller(
        col=col,
        cfg=cfg,
        boundary=base_inputs.boundary,
    )
    if bot_comp_control_enabled and bot_comp_ctrl is not None and bot_comp_sp is not None and bot_comp_idx is not None:
        mv_label = "boilup_lbmolph" if str(bot_comp_mv_mode) != "reboiler-duty" else "reboiler_duty_btuph"
        print(
            "[Control] Bottoms composition control enabled  "
            f"component={str(bot_comp_name)}  "
            f"x_SP={float(bot_comp_sp):.6f}  "
            f"Kc={float(bot_comp_ctrl.kc):.3g}  Ti={float(bot_comp_ctrl.ti_sec):.3g} s  "
            f"MV={mv_label}  limits=({float(bot_comp_ctrl.out_min):.3f}, {float(bot_comp_ctrl.out_max):.3f})"
        )

    last_top_pressure_pv_psia: Optional[float] = None
    last_top_pressure_pv_filt_psia: Optional[float] = None
    last_top_energy_resid_abs_btups: Optional[float] = None
    last_pressure_mv_cmd: Optional[float] = None
    try:
        p_ctrl_idx = 0
        p0 = float(np.asarray(getattr(col, "P_psia"), dtype=float).reshape((-1,))[p_ctrl_idx])
        if np.isfinite(p0) and p0 > 0.0:
            last_top_pressure_pv_psia = p0
            last_top_pressure_pv_filt_psia = p0
    except Exception:
        pass
    if pressure_control_enabled and top_pressure_ctrl is not None:
        try:
            last_pressure_mv_cmd = float(top_pressure_ctrl.bias)
        except Exception:
            last_pressure_mv_cmd = None
        pv_filt_restart = _get_controller_restart_value(col, "top_pressure_pv_filt_psia")
        if pv_filt_restart is not None and np.isfinite(float(pv_filt_restart)) and float(pv_filt_restart) > 0.0:
            last_top_pressure_pv_filt_psia = float(pv_filt_restart)
            last_top_pressure_pv_psia = float(pv_filt_restart)
        mv_cmd_restart = _get_controller_restart_value(col, "top_pressure_mv_cmd_btuph")
        if mv_cmd_restart is not None and np.isfinite(float(mv_cmd_restart)):
            last_pressure_mv_cmd = float(mv_cmd_restart)
    top_pressure_resid_restart = _get_controller_restart_value(col, "top_pressure_resid_abs_btups")
    if top_pressure_resid_restart is not None and np.isfinite(float(top_pressure_resid_restart)):
        last_top_energy_resid_abs_btups = float(top_pressure_resid_restart)
    top_pT_restart = _get_controller_restart_value(col, "top_drum_pressure_T_prev_F")
    if top_pT_restart is not None and np.isfinite(float(top_pT_restart)):
        last_top_drum_pressure_T = float(top_pT_restart)
    restart_distillate_cmd_lbmolph = _get_controller_restart_value(col, "distillate_cmd_lbmolph")
    restart_bottoms_cmd_lbmolph = _get_controller_restart_value(col, "bottoms_cmd_lbmolph")
    restart_reflux_cmd_lbmolph = _get_controller_restart_value(col, "reflux_cmd_lbmolph")
    restart_boilup_cmd_lbmolph = _get_controller_restart_value(col, "boilup_cmd_lbmolph")

    tag = _timestamp_tag()
    run_name = _clean_optional_text(getattr(cfg, "run_name", None))
    run_description = _clean_optional_text(getattr(cfg, "run_description", None))
    logs_dir = Path(cfg.logs_dir)
    if not logs_dir.is_absolute():
        logs_dir = Path.cwd() / logs_dir
    if cfg.write_logs:
        _ensure_dir(logs_dir)
        _milestone(f"ensured logs directory: {logs_dir}")

    profile_path = logs_dir / f"column_profile_{tag}.csv"
    summary_path = logs_dir / f"column_summary_{tag}.csv"
    metadata_path = logs_dir / f"run_metadata_{tag}.json"
    runtime_control_path = logs_dir / "runtime_control.json"

    profile_file = None
    summary_file = None

    start_perf = time.perf_counter()
    start_wall_dt = _dt.datetime.now()
    restart_hidden_warmup_steps = 0
    restart_hidden_warmup_time_sec = 0.0
    if explicit_runtime_restart and bool(getattr(cfg, "enable_restart_reentry_dynamic_warmup", True)):
        try:
            restart_hidden_warmup_time_sec = max(float(getattr(cfg, "restart_reentry_dynamic_warmup_sec", 5.0)), 0.0)
        except Exception:
            restart_hidden_warmup_time_sec = 5.0
        if restart_hidden_warmup_time_sec > 0.0 and float(dt) > 0.0:
            restart_hidden_warmup_steps = max(1, int(round(restart_hidden_warmup_time_sec / float(dt))))
            restart_hidden_warmup_time_sec = float(restart_hidden_warmup_steps) * float(dt)
            print(
                "[Init] Hidden restart dynamic warmup enabled  "
                f"sim_time={restart_hidden_warmup_time_sec:.2f} s  "
                f"steps={int(restart_hidden_warmup_steps)}"
            )
    t_s = -float(restart_hidden_warmup_time_sec)
    global_mass_closure_cum_lbmol = 0.0
    integrator_fallback_count = 0
    last_step_integrator_info: Dict[str, Any] = {"requested_mode": str(integrator_mode)}
    steady_state_status_last: Dict[str, float] = {
        "steady_state_flag": np.nan,
        "steady_state_score": np.nan,
        "ss_max_rel_state_rate_per_s": np.nan,
        "ss_max_kpi_slope_per_s": np.nan,
        "ss_max_mv_rate_per_s": np.nan,
        "ss_max_temp_rate_F_per_s": np.nan,
        "ss_max_sp_error": np.nan,
    }

    ss_enabled = bool(getattr(cfg, "enable_steady_state_detection", True))
    try:
        ss_window_sec = float(getattr(cfg, "steady_state_window_sec", 30.0))
    except Exception:
        ss_window_sec = 30.0
    if (not np.isfinite(ss_window_sec)) or ss_window_sec <= 0.0:
        ss_window_sec = max(float(dt), 1e-9)
    try:
        ss_min_time_sec = float(getattr(cfg, "steady_state_min_time_sec", 60.0))
    except Exception:
        ss_min_time_sec = 60.0
    if (not np.isfinite(ss_min_time_sec)) or ss_min_time_sec < 0.0:
        ss_min_time_sec = 0.0
    try:
        ss_rate_floor_lbmol = float(getattr(cfg, "steady_state_rate_denom_floor_lbmol", 1.0))
    except Exception:
        ss_rate_floor_lbmol = 1.0
    if (not np.isfinite(ss_rate_floor_lbmol)) or ss_rate_floor_lbmol < 0.0:
        ss_rate_floor_lbmol = 1.0

    def _tol_or_none(v: Any) -> Optional[float]:
        try:
            x = float(v)
        except Exception:
            return None
        if (not np.isfinite(x)) or x <= 0.0:
            return None
        return float(x)

    ss_tol_rel = _tol_or_none(getattr(cfg, "steady_state_rel_state_rate_tol_per_s", None))
    ss_tol_kpi = _tol_or_none(getattr(cfg, "steady_state_kpi_slope_tol_per_s", None))
    ss_tol_mv = _tol_or_none(getattr(cfg, "steady_state_mv_rate_tol_per_s", None))
    ss_tol_temp = _tol_or_none(getattr(cfg, "steady_state_temp_rate_tol_F_per_s", None))
    ss_tol_sp = _tol_or_none(getattr(cfg, "steady_state_sp_error_tol", None))
    ss_require_sp = bool(getattr(cfg, "steady_state_require_sp", False))

    ss_hist: deque = deque()
    ss_prev_t_s: Optional[float] = None
    ss_prev_y: Optional[np.ndarray] = None
    if ss_enabled:
        print(
            "[Init] Steady-state detector enabled  "
            f"window={float(ss_window_sec):.3g}s  "
            f"min_time={float(ss_min_time_sec):.3g}s  "
            f"tol_rel={float(ss_tol_rel) if ss_tol_rel is not None else float('nan'):.3g}/s  "
            f"tol_kpi={float(ss_tol_kpi) if ss_tol_kpi is not None else float('nan'):.3g}/s  "
            f"tol_mv={float(ss_tol_mv) if ss_tol_mv is not None else float('nan'):.3g}/s  "
            f"tol_T={float(ss_tol_temp) if ss_tol_temp is not None else float('nan'):.3g}F/s  "
            f"tol_sp={float(ss_tol_sp) if ss_tol_sp is not None else float('nan'):.3g}  "
            f"require_sp={'on' if ss_require_sp else 'off'}"
        )

    metadata_doc: Dict[str, Any] = {
        "run_id": str(tag),
        "run_name": run_name or "",
        "run_description": run_description or "",
        "status": "starting",
        "pid": int(os.getpid()),
        "started_at_local": start_wall_dt.strftime("%Y-%m-%d %H:%M:%S"),
        "excel_path": str(Path(cfg.excel_path).resolve()),
        "logs_dir": str(logs_dir),
        "runtime_mode": str(runtime_mode),
        "thermo_mode": str(getattr(cfg, "thermo_mode", "")),
        "clapeyron_model": str(getattr(cfg, "clapeyron_model", "")),
        "clapeyron_pr_parameter_source": str(
            getattr(cfg, "clapeyron_pr_parameter_source", "default") or "default"
        ),
        "thermo_every_n_steps": int(getattr(cfg, "thermo_every_n_steps", 1)),
        "n_steps": int(cfg.n_steps),
        "dt_sec": float(dt),
        "profile_csv": str(profile_path) if cfg.write_logs else "",
        "summary_csv": str(summary_path) if cfg.write_logs else "",
        "runtime_control_json": str(runtime_control_path) if cfg.write_logs else "",
        "startup_trace_log": str(startup_trace_path),
        "startup_timing_sec": dict(startup_timing_sec),
        "startup_build_info": dict(startup_build_info),
        "init_pack_top_drum_vapor": dict(init_pack_top_drum_vapor_info),
        "init_match_condenser_duty": dict(init_match_condenser_duty_info),
        "init_align_top_liquid_to_condensate": dict(init_align_top_liquid_info),
        "startup_seed_cache": dict(startup_seed_cache_info),
        "thermo_call_counters": {},
    }
    if cfg.write_logs:
        _write_json_atomic(metadata_path, metadata_doc)

    run_status = "running"
    run_error: Optional[str] = None
    run_error_traceback: str = ""
    failure_phase = "startup"
    failure_step: Optional[int] = None
    failure_time_s = float(max(float(t_s), 0.0))

    try:
        if cfg.write_logs:
            metadata_doc["status"] = "running"
            _write_json_atomic(metadata_path, metadata_doc)
            profile_file = profile_path.open("w", newline="", encoding="utf-8")
            summary_file = summary_path.open("w", newline="", encoding="utf-8")
            _milestone("opened log files")

        profile_writer = None
        summary_writer = None
        profile_header_written = False
        summary_header_written = False
        initial_snapshot_written = False

        def _write_log_snapshot(
            *,
            t_snapshot_s: float,
            col_snapshot: ColumnSpec,
            y_snapshot: np.ndarray,
            diag_snapshot: Dict[str, np.ndarray],
            feed_tag_snapshot: StreamTag,
            dist_tag_snapshot: StreamTag,
            bots_tag_snapshot: StreamTag,
            wall_clock_iso: str,
            wall_elapsed_s: float,
        ) -> None:
            nonlocal profile_writer, summary_writer, profile_header_written, summary_header_written
            if (not cfg.write_logs) or profile_file is None or summary_file is None:
                return

            prow = _profile_rows(
                t_snapshot_s,
                case,
                col_snapshot,
                layout,
                y_snapshot,
                diag_snapshot,
                include_temperature=cfg.include_temperature,
                volume_model=base_inputs.volume_model,
                wall_clock_iso=wall_clock_iso,
                wall_elapsed_s=wall_elapsed_s,
                feed_tag=feed_tag_snapshot,
                dist_tag=dist_tag_snapshot,
                bots_tag=bots_tag_snapshot,
            )
            srow = _summary_row(
                t_snapshot_s,
                case,
                col_snapshot,
                layout,
                y_snapshot,
                diag_snapshot,
                include_temperature=cfg.include_temperature,
                volume_model=base_inputs.volume_model,
                wall_clock_iso=wall_clock_iso,
                wall_elapsed_s=wall_elapsed_s,
                feed_tag=feed_tag_snapshot,
                dist_tag=dist_tag_snapshot,
                bots_tag=bots_tag_snapshot,
                integrator_info=last_step_integrator_info,
            )

            if (not profile_header_written) and prow:
                profile_writer = csv.DictWriter(profile_file, fieldnames=list(prow[0].keys()))
                profile_writer.writeheader()
                profile_header_written = True
            for r in prow:
                profile_writer.writerow(r)
            profile_file.write("\n")
            profile_file.flush()

            if not summary_header_written:
                summary_writer = csv.DictWriter(summary_file, fieldnames=list(srow.keys()))
                summary_writer.writeheader()
                summary_header_written = True
            summary_writer.writerow(srow)
            summary_file.flush()

        # Visual separator between startup milestones and runtime progress output.
        print()

        if cfg.write_logs and restart_hidden_warmup_steps <= 0:
            _write_log_snapshot(
                t_snapshot_s=0.0,
                col_snapshot=col,
                y_snapshot=y,
                diag_snapshot={},
                feed_tag_snapshot=feed_tag,
                dist_tag_snapshot=dist_tag,
                bots_tag_snapshot=bots_tag,
                wall_clock_iso=_dt.datetime.now().isoformat(timespec="seconds"),
                wall_elapsed_s=float(time.perf_counter() - start_perf),
            )
            initial_snapshot_written = True
            _runtime_trace("initial snapshot written at t=0.00 s")

        total_steps = int(cfg.n_steps) + int(restart_hidden_warmup_steps)
        step = 0
        while step <= int(total_steps):
            visible_step = int(step) - int(restart_hidden_warmup_steps)
            visible_logging_phase = bool(visible_step >= 0)
            t_visible_s = float(max(float(t_s), 0.0))
            runtime_trace_active = bool(
                visible_logging_phase
                and visible_step <= max(1, int(parity_runtime_thermo_defer_visible_steps))
            )
            failure_step = int(visible_step)
            failure_time_s = float(t_visible_s)
            failure_phase = "integration" if visible_logging_phase else "restart_hidden_warmup"
            if visible_logging_phase and visible_step == 0 and restart_hidden_warmup_steps > 0:
                ss_hist.clear()
                ss_prev_t_s = None
                ss_prev_y = None
                steady_state_status_last = {
                    "steady_state_flag": np.nan,
                    "steady_state_score": np.nan,
                    "ss_max_rel_state_rate_per_s": np.nan,
                    "ss_max_kpi_slope_per_s": np.nan,
                    "ss_max_mv_rate_per_s": np.nan,
                    "ss_max_temp_rate_F_per_s": np.nan,
                    "ss_max_sp_error": np.nan,
                }
            step_col = col
            step_boundary = base_inputs.boundary
            step_feed_tag = feed_tag
            step_dist_tag = dist_tag
            step_bots_tag = bots_tag
            if feed_step_enabled and float(t_s) >= float(feed_step_time_sec):
                try:
                    step_feed_flow_lbmolph = float(base_feed_total_lbmolph) * float(feed_step_scale)
                    step_feed_comp = None
                    if base_feed_component_flows is not None:
                        step_feed_comp = {
                            str(k): float(v) * float(feed_step_scale)
                            for k, v in base_feed_component_flows.items()
                        }
                    step_feed_stream = replace(
                        base_feed_stream,
                        total_molar_flow_lbmolph=float(step_feed_flow_lbmolph),
                        component_molar_flows_lbmolph=step_feed_comp,
                    )
                    step_streams = dict(getattr(col, "streams", {}) or {})
                    step_streams["Feed"] = step_feed_stream
                    step_col = replace(col, streams=step_streams)
                    step_feed_tag = StreamTag(
                        name=feed_tag.name,
                        flow_lbmolph=float(step_feed_flow_lbmolph),
                        stage_1based=feed_tag.stage_1based,
                    )
                except Exception:
                    step_col = col
                    step_feed_tag = feed_tag
            step_condenser_duty_mode = str(base_inputs.condenser_duty_mode)
            step_condenser_duty_btu_per_h = (
                float(base_inputs.condenser_duty_btu_per_h)
                if base_inputs.condenser_duty_btu_per_h is not None
                else None
            )
            step_condenser_duty_trim_btu_per_h: Optional[float] = (
                float(base_inputs.condenser_duty_trim_btu_per_h)
                if base_inputs.condenser_duty_trim_btu_per_h is not None
                else None
            )
            step_condenser_duty_cmd_btu_per_h: Optional[float] = None
            step_reboiler_mode = str(base_inputs.reboiler_mode)
            step_reboiler_duty_btu_per_h: Optional[float] = (
                float(base_inputs.reboiler_duty_btu_per_h)
                if base_inputs.reboiler_duty_btu_per_h is not None
                else None
            )
            step_reboiler_duty_trim_btu_per_h: Optional[float] = (
                float(base_inputs.reboiler_duty_trim_btu_per_h)
                if base_inputs.reboiler_duty_trim_btu_per_h is not None
                else None
            )
            step_reboiler_duty_cmd_btu_per_h: Optional[float] = None
            step_pressure_top_anchor_psia: Optional[float] = (
                float(base_inputs.pressure_top_anchor_psia)
                if base_inputs.pressure_top_anchor_psia is not None
                else None
            )
            step_pressure_top_anchor_cmd_psia: Optional[float] = None
            step_distillate_comp_pv: Optional[float] = None
            step_distillate_comp_sp: Optional[float] = None
            step_reflux_ratio_cmd: Optional[float] = None
            step_reflux_cmd_lbmolph: Optional[float] = None
            step_distillate_comp_error: Optional[float] = None
            step_distillate_comp_u_unclamped: Optional[float] = None
            step_distillate_comp_out_max: Optional[float] = None
            step_distillate_comp_cap_active: Optional[float] = None
            step_bottoms_comp_pv: Optional[float] = None
            step_bottoms_comp_sp: Optional[float] = None
            step_boilup_cmd_lbmolph: Optional[float] = None
            step_pressure_ctrl_pv_raw_psia: Optional[float] = None
            step_pressure_ctrl_pv_filt_psia: Optional[float] = None
            step_pressure_ctrl_gain_scale: Optional[float] = None
            step_pressure_ctrl_energy_resid_abs_btups: Optional[float] = None
            step_pressure_ctrl_error_psia: Optional[float] = None
            step_pressure_ctrl_p_term: Optional[float] = None
            step_pressure_ctrl_i_term: Optional[float] = None
            step_pressure_ctrl_u_unclamped_btuph: Optional[float] = None
            step_pressure_ctrl_sat_hi: Optional[float] = None
            step_pressure_ctrl_sat_lo: Optional[float] = None
            step_pressure_ctrl_allow_int: Optional[float] = None
            step_top_level_ctrl_pv: Optional[float] = None
            step_top_level_ctrl_sp: Optional[float] = None
            step_top_level_ctrl_err: Optional[float] = None
            step_top_level_ctrl_p_term: Optional[float] = None
            step_top_level_ctrl_i_term: Optional[float] = None
            step_top_level_ctrl_pi_term_lbmolph: Optional[float] = None
            step_top_level_ctrl_ff_term_lbmolph: Optional[float] = None
            step_top_level_ctrl_u_unclamped_lbmolph: Optional[float] = None
            step_top_level_ctrl_sat_hi: Optional[float] = None
            step_top_level_ctrl_sat_lo: Optional[float] = None
            step_top_level_ctrl_allow_int: Optional[float] = None
            step_bottom_level_ctrl_pv: Optional[float] = None
            step_bottom_level_ctrl_sp: Optional[float] = None
            step_bottom_level_ctrl_err: Optional[float] = None
            step_bottom_level_ctrl_p_term: Optional[float] = None
            step_bottom_level_ctrl_i_term: Optional[float] = None
            step_bottom_level_ctrl_u_unclamped_lbmolph: Optional[float] = None
            step_bottom_level_ctrl_sat_hi: Optional[float] = None
            step_bottom_level_ctrl_sat_lo: Optional[float] = None
            step_bottom_level_ctrl_allow_int: Optional[float] = None
            controllers_active = step > 0
            if (
                level_control_enabled
                and top_level_ctrl is not None
                and bot_level_ctrl is not None
                and top_level_sp is not None
                and bot_level_sp is not None
            ):
                u_now = layout.unpack(y)
                top_level_pv = float(np.sum(np.asarray(u_now.get("top_L", []), dtype=float)))
                if (
                    top_level_pv_mode == "true-level"
                    and top_level_rho_ref_lbmol_ft3 is not None
                    and base_inputs.top_drum_total_volume_ft3 is not None
                    and base_inputs.thermo_provider is not None
                ):
                    try:
                        p_top_for_level = None
                        if last_P_hyd is not None:
                            p_top_for_level = float(np.asarray(last_P_hyd, dtype=float).reshape((col.n_stages,))[0])
                        elif last_P_diag is not None:
                            p_top_for_level = float(np.asarray(last_P_diag, dtype=float).reshape((col.n_stages,))[0])
                        elif hasattr(col, "P_psia"):
                            p_top_for_level = float(np.asarray(col.P_psia, dtype=float).reshape((col.n_stages,))[0])
                        top_liq_vol_now_ft3, _top_rho_now, _top_level_frac_now = _estimate_top_drum_liquid_volume_ft3(
                            col=col,
                            layout=layout,
                            y_vec=y,
                            thermo_provider=base_inputs.thermo_provider,
                            top_drum_total_volume_ft3=base_inputs.top_drum_total_volume_ft3,
                            p_top_psia=p_top_for_level,
                        )
                        if _top_level_frac_now is not None and np.isfinite(float(_top_level_frac_now)):
                            top_level_pv = float(
                                _horizontal_cylinder_height_fraction_from_volume_fraction(
                                    float(_top_level_frac_now)
                                )
                            )
                            last_valid_top_level_pv = float(top_level_pv)
                    except Exception:
                        pass
                    if (
                        (not np.isfinite(float(top_level_pv)))
                        or float(top_level_pv) < 0.0
                        or float(top_level_pv) > 1.0
                    ) and last_valid_top_level_pv is not None and np.isfinite(float(last_valid_top_level_pv)):
                        top_level_pv = float(np.clip(float(last_valid_top_level_pv), 0.0, 1.0))
                bot_level_pv = float(np.sum(np.asarray(u_now.get("bottom_L", []), dtype=float)))
                if (
                    bottom_level_pv_mode == "true-level"
                    and base_inputs.bottom_sump_total_volume_ft3 is not None
                    and base_inputs.thermo_provider is not None
                ):
                    try:
                        p_bottom_for_level = None
                        if last_P_hyd is not None:
                            p_bottom_for_level = float(np.asarray(last_P_hyd, dtype=float).reshape((col.n_stages,))[-1])
                        elif last_P_diag is not None:
                            p_bottom_for_level = float(np.asarray(last_P_diag, dtype=float).reshape((col.n_stages,))[-1])
                        elif hasattr(col, "P_psia"):
                            p_bottom_for_level = float(np.asarray(col.P_psia, dtype=float).reshape((col.n_stages,))[-1])
                        bottom_liq_vol_now_ft3, _bot_rho_now, _bot_level_frac_now = _estimate_bottom_sump_liquid_volume_ft3(
                            col=col,
                            layout=layout,
                            y_vec=y,
                            thermo_provider=base_inputs.thermo_provider,
                            bottom_sump_total_volume_ft3=base_inputs.bottom_sump_total_volume_ft3,
                            p_bottom_psia=p_bottom_for_level,
                        )
                        if _bot_level_frac_now is not None and np.isfinite(float(_bot_level_frac_now)):
                            bot_level_pv = float(np.clip(float(_bot_level_frac_now), 0.0, 1.0))
                            last_valid_bottom_level_pv = float(bot_level_pv)
                    except Exception:
                        pass
                    if (
                        (not np.isfinite(float(bot_level_pv)))
                        or float(bot_level_pv) < 0.0
                        or float(bot_level_pv) > 1.0
                    ) and last_valid_bottom_level_pv is not None and np.isfinite(float(last_valid_bottom_level_pv)):
                        bot_level_pv = float(np.clip(float(last_valid_bottom_level_pv), 0.0, 1.0))
                step_top_level_ctrl_pv = float(top_level_pv)
                step_top_level_ctrl_sp = float(top_level_sp)
                step_bottom_level_ctrl_pv = float(bot_level_pv)
                step_bottom_level_ctrl_sp = float(bot_level_sp)
                if controllers_active:
                    top_level_ctrl_debug: Dict[str, float] = {}
                    bot_level_ctrl_debug: Dict[str, float] = {}
                    dist_cmd_pi = _pi_update(
                        top_level_ctrl,
                        pv=top_level_pv,
                        sp=float(top_level_sp),
                        dt_sec=float(dt),
                        debug=top_level_ctrl_debug,
                    )
                    bot_cmd = _pi_update(
                        bot_level_ctrl,
                        pv=bot_level_pv,
                        sp=float(bot_level_sp),
                        dt_sec=float(dt),
                        debug=bot_level_ctrl_debug,
                    )
                    step_top_level_ctrl_err = float(top_level_ctrl_debug.get("error", np.nan))
                    step_top_level_ctrl_p_term = float(top_level_ctrl_debug.get("p_term", np.nan))
                    step_top_level_ctrl_i_term = float(top_level_ctrl_debug.get("i_term", np.nan))
                    step_top_level_ctrl_pi_term_lbmolph = float(dist_cmd_pi)
                    step_top_level_ctrl_u_unclamped_lbmolph = float(top_level_ctrl_debug.get("u_unclamped", np.nan))
                    step_top_level_ctrl_sat_hi = float(top_level_ctrl_debug.get("sat_hi", np.nan))
                    step_top_level_ctrl_sat_lo = float(top_level_ctrl_debug.get("sat_lo", np.nan))
                    step_top_level_ctrl_allow_int = float(top_level_ctrl_debug.get("allow_int", np.nan))
                    step_bottom_level_ctrl_err = float(bot_level_ctrl_debug.get("error", np.nan))
                    step_bottom_level_ctrl_p_term = float(bot_level_ctrl_debug.get("p_term", np.nan))
                    step_bottom_level_ctrl_i_term = float(bot_level_ctrl_debug.get("i_term", np.nan))
                    step_bottom_level_ctrl_u_unclamped_lbmolph = float(bot_level_ctrl_debug.get("u_unclamped", np.nan))
                    step_bottom_level_ctrl_sat_hi = float(bot_level_ctrl_debug.get("sat_hi", np.nan))
                    step_bottom_level_ctrl_sat_lo = float(bot_level_ctrl_debug.get("sat_lo", np.nan))
                    step_bottom_level_ctrl_allow_int = float(bot_level_ctrl_debug.get("allow_int", np.nan))
                    dist_cmd = float(dist_cmd_pi)
                    if bool(getattr(cfg, "enable_top_level_feedforward", False)):
                        try:
                            ff_gain = float(getattr(cfg, "top_level_feedforward_gain", 0.0) or 0.0)
                        except Exception:
                            ff_gain = 0.0
                        if np.isfinite(ff_gain) and abs(ff_gain) > 1e-12 and base_feed_total_lbmolph is not None:
                            try:
                                step_feed_flow_now = float(step_feed_tag.flow_lbmolph)
                                feed_delta_lbmolph = float(step_feed_flow_now) - float(base_feed_total_lbmolph)
                                step_top_level_ctrl_ff_term_lbmolph = float(ff_gain) * float(feed_delta_lbmolph)
                                dist_cmd = float(dist_cmd) + float(step_top_level_ctrl_ff_term_lbmolph)
                            except Exception:
                                step_top_level_ctrl_ff_term_lbmolph = 0.0
                else:
                    if (
                        explicit_runtime_restart
                        and step == 0
                        and restart_distillate_cmd_lbmolph is not None
                        and np.isfinite(float(restart_distillate_cmd_lbmolph))
                    ):
                        dist_cmd = float(restart_distillate_cmd_lbmolph)
                    elif step_boundary.distillate_lbmolph is not None:
                        dist_cmd = float(step_boundary.distillate_lbmolph)
                    else:
                        dist_cmd = float(dist_tag.flow_lbmolph)
                    if (
                        explicit_runtime_restart
                        and step == 0
                        and restart_bottoms_cmd_lbmolph is not None
                        and np.isfinite(float(restart_bottoms_cmd_lbmolph))
                    ):
                        bot_cmd = float(restart_bottoms_cmd_lbmolph)
                    elif step_boundary.bottoms_lbmolph is not None:
                        bot_cmd = float(step_boundary.bottoms_lbmolph)
                    else:
                        bot_cmd = float(bots_tag.flow_lbmolph)
                step_boundary = BoundaryFlows(
                    reflux_lbmolph=base_inputs.boundary.reflux_lbmolph,
                    boilup_lbmolph=base_inputs.boundary.boilup_lbmolph,
                    distillate_lbmolph=float(dist_cmd),
                    bottoms_lbmolph=float(bot_cmd),
                )
                step_dist_tag = StreamTag(
                    name=dist_tag.name,
                    flow_lbmolph=float(dist_cmd),
                    stage_1based=dist_tag.stage_1based,
                )
                step_bots_tag = StreamTag(
                    name=bots_tag.name,
                    flow_lbmolph=float(bot_cmd),
                    stage_1based=bots_tag.stage_1based,
                )
            if (
                comp_control_enabled
                and dist_comp_ctrl is not None
                and dist_comp_sp is not None
                and dist_comp_idx is not None
            ):
                u_now = layout.unpack(y)
                top_liq = np.asarray(u_now.get("top_L", []), dtype=float).reshape((-1,))
                xD_pv = np.nan
                top_total = np.nan
                if top_liq.size == int(col.n_components):
                    top_total = float(np.sum(top_liq))
                    if np.isfinite(top_total) and top_total > 0.0:
                        xD_vec = top_liq / max(top_total, 1e-300)
                        xD_pv = float(xD_vec[int(dist_comp_idx)])

                # Reflux feasibility clamp:
                # keep reflux bounded by condenser liquid generation plus a limited
                # drawdown of reflux-drum inventory so composition control cannot
                # starve top holdup and force distillate draw to zero.
                d_cmd_lbmolph = np.nan
                if step_boundary.distillate_lbmolph is not None:
                    d_cmd_lbmolph = float(step_boundary.distillate_lbmolph)
                elif step_dist_tag.flow_lbmolph is not None:
                    d_cmd_lbmolph = float(step_dist_tag.flow_lbmolph)
                elif dist_tag.flow_lbmolph is not None:
                    d_cmd_lbmolph = float(dist_tag.flow_lbmolph)
                if (not np.isfinite(d_cmd_lbmolph)) or d_cmd_lbmolph < 0.0:
                    d_cmd_lbmolph = 0.0

                vin0_est_lbmolph = np.nan
                try:
                    if last_V_out is not None:
                        vprev = np.asarray(last_V_out, dtype=float).reshape((-1,))
                        if vprev.size > 1 and np.isfinite(float(vprev[1])):
                            vin0_est_lbmolph = float(vprev[1])
                        elif vprev.size > 0 and np.isfinite(float(vprev[0])):
                            vin0_est_lbmolph = float(vprev[0])
                except Exception:
                    vin0_est_lbmolph = np.nan
                if (not np.isfinite(vin0_est_lbmolph)) and int(getattr(col, "n_stages", 0)) > 1:
                    try:
                        vprof = np.asarray(getattr(col, "V_lbmolph"), dtype=float).reshape((-1,))
                        if vprof.size > 1 and np.isfinite(float(vprof[1])):
                            vin0_est_lbmolph = float(vprof[1])
                    except Exception:
                        vin0_est_lbmolph = np.nan
                if (not np.isfinite(vin0_est_lbmolph)) or vin0_est_lbmolph < 0.0:
                    vin0_est_lbmolph = 0.0

                reflux_max_feasible = float(dist_comp_ctrl.out_max)
                if bool(getattr(cfg, "enable_reflux_feasibility_cap", True)):
                    if np.isfinite(top_total):
                        # Keep top inventory moving toward its level setpoint so the
                        # composition loop cannot pin operation at D=0 while draining
                        # (or holding low) reflux-drum holdup.
                        recover_tau_sec = 120.0
                        if top_level_ctrl is not None and np.isfinite(float(top_level_ctrl.ti_sec)):
                            recover_tau_sec = max(float(top_level_ctrl.ti_sec), 30.0)
                        desired_dM_top_lbmolph = _desired_inventory_recovery_rate_lbmolph(
                            total_lbmol=float(top_total),
                            pv=float(step_top_level_ctrl_pv) if step_top_level_ctrl_pv is not None and np.isfinite(float(step_top_level_ctrl_pv)) else None,
                            sp=float(top_level_sp) if top_level_sp is not None and np.isfinite(float(top_level_sp)) else None,
                            pv_mode=str(top_level_pv_mode),
                            lbmol_per_volume_fraction_scale=(
                                float(top_level_scale_lbmol_per_frac)
                                if top_level_scale_lbmol_per_frac is not None and np.isfinite(float(top_level_scale_lbmol_per_frac))
                                else None
                            ),
                            recover_tau_sec=float(recover_tau_sec),
                        )
                        sustainable_lbmolph = float(vin0_est_lbmolph) - float(d_cmd_lbmolph)
                        reflux_max_feasible = min(
                            float(dist_comp_ctrl.out_max),
                            max(0.0, sustainable_lbmolph - desired_dM_top_lbmolph),
                        )
                    reflux_max_feasible = max(float(dist_comp_ctrl.out_min), float(reflux_max_feasible))

                if controllers_active:
                    reflux_ctrl_debug: Dict[str, float] = {}
                    reflux_cmd = _pi_update(
                        dist_comp_ctrl,
                        pv=float(xD_pv),
                        sp=float(dist_comp_sp),
                        dt_sec=float(dt),
                        out_max=float(reflux_max_feasible),
                        debug=reflux_ctrl_debug,
                    )
                else:
                    reflux_ctrl_debug = {}
                    if (
                        explicit_runtime_restart
                        and step == 0
                        and restart_reflux_cmd_lbmolph is not None
                        and np.isfinite(float(restart_reflux_cmd_lbmolph))
                    ):
                        reflux_cmd = float(restart_reflux_cmd_lbmolph)
                    elif step_boundary.reflux_lbmolph is not None:
                        reflux_cmd = float(step_boundary.reflux_lbmolph)
                    else:
                        reflux_cmd = float(reflux_tag.flow_lbmolph)
                step_boundary = BoundaryFlows(
                    reflux_lbmolph=float(reflux_cmd),
                    boilup_lbmolph=step_boundary.boilup_lbmolph,
                    distillate_lbmolph=step_boundary.distillate_lbmolph,
                    bottoms_lbmolph=step_boundary.bottoms_lbmolph,
                )
                step_reflux_cmd_lbmolph = float(reflux_cmd)
                step_distillate_comp_error = float(reflux_ctrl_debug.get("error", np.nan))
                step_distillate_comp_u_unclamped = float(reflux_ctrl_debug.get("u_unclamped", np.nan))
                step_distillate_comp_out_max = float(reflux_max_feasible)
                step_distillate_comp_cap_active = (
                    1.0
                    if np.isfinite(float(reflux_max_feasible))
                    and float(reflux_max_feasible) < float(dist_comp_ctrl.out_max) - 1e-9
                    else 0.0
                )
                d_for_ratio = np.nan
                if step_boundary.distillate_lbmolph is not None:
                    d_for_ratio = float(step_boundary.distillate_lbmolph)
                elif step_dist_tag.flow_lbmolph is not None:
                    d_for_ratio = float(step_dist_tag.flow_lbmolph)
                elif dist_tag.flow_lbmolph is not None:
                    d_for_ratio = float(dist_tag.flow_lbmolph)
                step_distillate_comp_pv = float(xD_pv) if np.isfinite(xD_pv) else np.nan
                step_distillate_comp_sp = float(dist_comp_sp)
                if np.isfinite(d_for_ratio) and d_for_ratio > 0.0:
                    step_reflux_ratio_cmd = float(reflux_cmd) / float(d_for_ratio)
            if (
                bot_comp_control_enabled
                and bot_comp_ctrl is not None
                and bot_comp_sp is not None
                and bot_comp_idx is not None
            ):
                u_now = layout.unpack(y)
                bot_liq = np.asarray(u_now.get("bottom_L", []), dtype=float).reshape((-1,))
                xB_pv = np.nan
                if bot_liq.size == int(col.n_components):
                    bot_total = float(np.sum(bot_liq))
                    if np.isfinite(bot_total) and bot_total > 0.0:
                        xB_vec = bot_liq / max(bot_total, 1e-300)
                        xB_pv = float(xB_vec[int(bot_comp_idx)])
                if controllers_active:
                    boilup_cmd = _pi_update(
                        bot_comp_ctrl,
                        pv=float(xB_pv),
                        sp=float(bot_comp_sp),
                        dt_sec=float(dt),
                    )
                elif str(bot_comp_mv_mode) == "reboiler-duty":
                    if step_reboiler_duty_btu_per_h is not None:
                        boilup_cmd = float(step_reboiler_duty_btu_per_h)
                    else:
                        boilup_cmd = float(bot_comp_ctrl.bias)
                else:
                    if (
                        explicit_runtime_restart
                        and step == 0
                        and restart_boilup_cmd_lbmolph is not None
                        and np.isfinite(float(restart_boilup_cmd_lbmolph))
                    ):
                        boilup_cmd = float(restart_boilup_cmd_lbmolph)
                    elif step_boundary.boilup_lbmolph is not None:
                        boilup_cmd = float(step_boundary.boilup_lbmolph)
                    else:
                        boilup_cmd = float(boilup_tag.flow_lbmolph)
                if str(bot_comp_mv_mode) == "reboiler-duty":
                    step_reboiler_mode = "duty"
                    step_reboiler_duty_btu_per_h = float(boilup_cmd)
                    step_reboiler_duty_cmd_btu_per_h = float(boilup_cmd)
                    # In duty mode, clear the explicit boilup boundary so RHS can
                    # realize boilup from the current reboiler duty and latent heat.
                    step_boundary = BoundaryFlows(
                        reflux_lbmolph=step_boundary.reflux_lbmolph,
                        boilup_lbmolph=None,
                        distillate_lbmolph=step_boundary.distillate_lbmolph,
                        bottoms_lbmolph=step_boundary.bottoms_lbmolph,
                    )
                else:
                    step_boundary = BoundaryFlows(
                        reflux_lbmolph=step_boundary.reflux_lbmolph,
                        boilup_lbmolph=float(boilup_cmd),
                        distillate_lbmolph=step_boundary.distillate_lbmolph,
                        bottoms_lbmolph=step_boundary.bottoms_lbmolph,
                    )
                    step_boilup_cmd_lbmolph = float(boilup_cmd)
                step_bottoms_comp_pv = float(xB_pv) if np.isfinite(xB_pv) else np.nan
                step_bottoms_comp_sp = float(bot_comp_sp)

            if pressure_control_enabled and top_pressure_ctrl is not None and top_pressure_sp is not None:
                p_ctrl_idx = 0
                pv_raw = None
                if last_top_pressure_pv_psia is not None:
                    try:
                        pv_try = float(last_top_pressure_pv_psia)
                        if np.isfinite(pv_try) and pv_try > 0.0:
                            pv_raw = pv_try
                    except Exception:
                        pv_raw = None
                if pv_raw is None:
                    try:
                        pv_raw = float(np.asarray(getattr(col, "P_psia"), dtype=float).reshape((-1,))[p_ctrl_idx])
                    except Exception:
                        pv_raw = float(top_pressure_sp)
                step_pressure_ctrl_pv_raw_psia = float(pv_raw)

                pv = float(pv_raw)
                pv_tau = cfg.top_pressure_pv_filter_tau_sec
                if (
                    pv_tau is not None
                    and np.isfinite(float(pv_tau))
                    and float(pv_tau) > 0.0
                ):
                    if last_top_pressure_pv_filt_psia is None or (not np.isfinite(float(last_top_pressure_pv_filt_psia))):
                        pv = float(pv_raw)
                    else:
                        alpha = float(dt) / (float(pv_tau) + float(dt))
                        alpha = float(np.clip(alpha, 0.0, 1.0))
                        pv = float(last_top_pressure_pv_filt_psia) + alpha * (float(pv_raw) - float(last_top_pressure_pv_filt_psia))
                last_top_pressure_pv_filt_psia = float(pv)
                step_pressure_ctrl_pv_filt_psia = float(pv)

                gain_scale = _pressure_resid_gain_scale(
                    resid_abs_btups=last_top_energy_resid_abs_btups,
                    resid_ref_btups=top_pressure_resid_ref_btups,
                    min_gain=float(cfg.top_pressure_resid_min_gain),
                )
                step_pressure_ctrl_gain_scale = float(gain_scale)
                if last_top_energy_resid_abs_btups is not None and np.isfinite(float(last_top_energy_resid_abs_btups)):
                    step_pressure_ctrl_energy_resid_abs_btups = float(last_top_energy_resid_abs_btups)

                if controllers_active:
                    pressure_ctrl_debug: Dict[str, float] = {}
                    if gain_scale < (1.0 - 1e-12):
                        damped_ctrl = PIController(
                            kc=float(top_pressure_ctrl.kc) * float(gain_scale),
                            ti_sec=float(top_pressure_ctrl.ti_sec),
                            bias=float(top_pressure_ctrl.bias),
                            out_min=float(top_pressure_ctrl.out_min),
                            out_max=float(top_pressure_ctrl.out_max),
                            integ=float(top_pressure_ctrl.integ),
                        )
                        q_cmd = _pi_update(
                            damped_ctrl,
                            pv=float(pv),
                            sp=float(top_pressure_sp),
                            dt_sec=float(dt),
                            debug=pressure_ctrl_debug,
                        )
                        # Keep integrator continuity in the primary controller state.
                        top_pressure_ctrl.integ = float(damped_ctrl.integ)
                    else:
                        q_cmd = _pi_update(
                            top_pressure_ctrl,
                            pv=float(pv),
                            sp=float(top_pressure_sp),
                            dt_sec=float(dt),
                            debug=pressure_ctrl_debug,
                        )
                    step_pressure_ctrl_error_psia = float(pressure_ctrl_debug.get("error", np.nan))
                    step_pressure_ctrl_p_term = float(pressure_ctrl_debug.get("p_term", np.nan))
                    step_pressure_ctrl_i_term = float(pressure_ctrl_debug.get("i_term", np.nan))
                    step_pressure_ctrl_u_unclamped_btuph = float(pressure_ctrl_debug.get("u_unclamped", np.nan))
                    step_pressure_ctrl_sat_hi = float(pressure_ctrl_debug.get("sat_hi", np.nan))
                    step_pressure_ctrl_sat_lo = float(pressure_ctrl_debug.get("sat_lo", np.nan))
                    step_pressure_ctrl_allow_int = float(pressure_ctrl_debug.get("allow_int", np.nan))
                elif str(pressure_control_mv) == "top-anchor":
                    if step_pressure_top_anchor_psia is not None:
                        q_cmd = float(step_pressure_top_anchor_psia)
                    else:
                        q_cmd = float(top_pressure_ctrl.bias)
                elif (
                    explicit_runtime_restart
                    and step == 0
                    and last_pressure_mv_cmd is not None
                    and np.isfinite(float(last_pressure_mv_cmd))
                ):
                    q_cmd = float(last_pressure_mv_cmd)
                elif step_condenser_duty_btu_per_h is not None:
                    q_cmd = float(step_condenser_duty_btu_per_h)
                else:
                    q_cmd = float(top_pressure_ctrl.bias)

                if controllers_active:
                    q_cmd = _apply_slew_limit(
                        cmd=float(q_cmd),
                        prev_cmd=last_pressure_mv_cmd,
                        rate_limit_per_s=cfg.top_pressure_mv_slew_limit_per_s,
                        dt_sec=float(dt),
                    )
                last_pressure_mv_cmd = float(q_cmd)

                if str(pressure_control_mv) == "top-anchor":
                    step_pressure_top_anchor_cmd_psia = float(q_cmd)
                    step_pressure_top_anchor_psia = float(q_cmd)
                else:
                    step_condenser_duty_cmd_btu_per_h = float(q_cmd)
                    if controllers_active and str(step_condenser_duty_mode).strip().lower() == "total-condense":
                        # Keep total-condense closure and apply PI as a duty trim.
                        ctrl_trim = float(q_cmd) - float(top_pressure_ctrl.bias)
                        base_trim = float(step_condenser_duty_trim_btu_per_h) if step_condenser_duty_trim_btu_per_h is not None else 0.0
                        step_condenser_duty_trim_btu_per_h = base_trim + ctrl_trim
                    else:
                        step_condenser_duty_mode = "specified"
                        step_condenser_duty_btu_per_h = float(q_cmd)

            pressure_model_step, vapor_flow_model_step, seq_liquid_alpha_scalar, seq_phase = (
                _resolve_startup_hydraulic_sequence_step(
                    t_s=float(t_s),
                    dt_sec=float(dt),
                    base_inputs=base_inputs,
                    enable_sequence=bool(startup_sequence_enabled),
                    energy_on_sec=max(
                        float(cfg.startup_sequence_energy_on_sec),
                        float(cfg.startup_sequence_profile_hold_sec),
                    ),
                    liquid_on_sec=max(
                        float(cfg.startup_sequence_liquid_on_sec),
                        float(cfg.startup_sequence_profile_hold_sec),
                    ),
                    liquid_ramp_sec=float(cfg.startup_sequence_liquid_ramp_sec),
                    liquid_resid_gate_lbmolph=cfg.startup_sequence_mass_resid_gate_lbmolph,
                    liquid_backoff_sec=cfg.startup_sequence_liquid_backoff_sec,
                    liquid_alpha_state=float(np.nanmean(seq_liquid_alpha_state)),
                    last_mass_resid_max_lbmolph=last_mass_resid_max_lbmolph,
                )
            )
            seq_liquid_alpha_state = np.full((col.n_stages,), float(seq_liquid_alpha_scalar), dtype=float)
            if bool(cfg.enable_startup_vapor_homotopy) and bool(startup_sequence_enabled):
                seq_vapor_beta_state, seq_vapor_guard_phase = _resolve_startup_vapor_homotopy_beta(
                    t_s=float(t_s),
                    dt_sec=float(dt),
                    base_inputs=base_inputs,
                    enable_homotopy=True,
                    liquid_on_sec=max(
                        float(cfg.startup_sequence_liquid_on_sec),
                        float(cfg.startup_sequence_profile_hold_sec),
                    ),
                    liquid_ramp_sec=float(cfg.startup_sequence_liquid_ramp_sec),
                    vapor_on_sec=cfg.startup_sequence_vapor_on_sec,
                    vapor_ramp_sec=float(cfg.startup_sequence_vapor_ramp_sec),
                    vapor_rel_rate_gate_per_s=cfg.startup_sequence_vapor_rel_rate_gate_per_s,
                    vapor_backoff_sec=cfg.startup_sequence_vapor_backoff_sec,
                    beta_state=float(seq_vapor_beta_state),
                    last_rel_state_rate_per_s=last_ss_rel_state_rate_per_s,
                )
                if str(base_inputs.vapor_flow_model or "").strip().lower() in ("energy", "conductance"):
                    vapor_on_eff = cfg.startup_sequence_vapor_on_sec
                    if vapor_on_eff is None:
                        vapor_on_eff = float(cfg.startup_sequence_liquid_on_sec) + float(cfg.startup_sequence_liquid_ramp_sec)
                    if float(t_s) < float(vapor_on_eff):
                        vapor_flow_model_step = "profile"
                    else:
                        vapor_flow_model_step = str(base_inputs.vapor_flow_model or vapor_flow_model_step).strip().lower()
                        seq_phase = "vapor_homotopy_ramp"
                else:
                    seq_vapor_beta_state = 1.0
                    seq_vapor_guard_phase = "disabled"
            else:
                seq_vapor_beta_state = 1.0
                seq_vapor_guard_phase = "disabled"
            if (
                (not startup_sequence_enabled)
                and runtime_mode == "hydraulic"
                and bool(base_inputs.enable_liquid_hydraulic_override)
            ):
                seq_liquid_alpha_state, seq_liquid_guard_phase = _resolve_residual_guarded_liquid_hydraulic_alpha_per_stage(
                    dt_sec=float(dt),
                    base_inputs=base_inputs,
                    liquid_resid_gate_lbmolph=cfg.startup_sequence_mass_resid_gate_lbmolph,
                    liquid_recover_sec=max(float(cfg.startup_sequence_liquid_ramp_sec), 1.0),
                    liquid_alpha_state=seq_liquid_alpha_state,
                    last_mass_resid_lbmolph_per_stage=last_mass_resid_lbmolph_per_stage,
                )
            else:
                seq_liquid_guard_phase = "sequence" if startup_sequence_enabled else "disabled"

            startup_total_reflux_washout_active = _startup_total_reflux_washout_active(
                float(t_s),
                cfg,
            )
            step_runtime_mode = (
                "total-reflux"
                if bool(startup_total_reflux_washout_active)
                else str(base_inputs.runtime_mode)
            )

            do_thermo, thermo_execution_reason = _resolve_runtime_thermo_execution_plan(
                step=int(step),
                thermo_every=int(thermo_every),
                col=col,
                layout=layout,
                y=y,
                include_temperature=bool(cfg.include_temperature),
                pressure_model=str(getattr(base_inputs, "pressure_model", "") or ""),
                base_inputs=base_inputs,
                last_T_tray=last_T_tray,
                last_P_hyd=last_P_hyd,
                last_P_diag=last_P_diag,
                last_Zfac=last_Zfac,
                last_z_overall=last_z_overall,
                last_K_tray=last_K_tray,
            )
            defer_runtime_thermo = bool(
                runtime_mode == "parity"
                and visible_logging_phase
                and 0 <= int(visible_step) < int(parity_runtime_thermo_defer_visible_steps)
            )
            if defer_runtime_thermo and do_thermo:
                do_thermo = False
                thermo_execution_reason = "deferred"
                _runtime_trace(
                    "step="
                    f"{int(visible_step)} deferring live thermo "
                    f"until visible_step>={int(parity_runtime_thermo_defer_visible_steps)}"
                )
            runtime_equilibrium_relaxation = bool(base_inputs.equilibrium_relaxation) and (not defer_runtime_thermo)
            (
                startup_packet_mainflash_reuse,
                thermo_refresh_dT_step,
                thermo_refresh_dP_step,
                thermo_refresh_dx_step,
            ) = _resolve_step0_startup_packet_reuse_thresholds(
                startup_seed_loaded=bool(startup_seed_loaded),
                runtime_mode=str(runtime_mode),
                step=int(step),
                last_tray_thermo_packet=last_tray_thermo_packet,
                last_T_tray=last_T_tray,
                last_P_hyd=last_P_hyd,
                last_P_diag=last_P_diag,
                last_z_overall=last_z_overall,
                base_inputs=base_inputs,
            )
            (
                thermo_packet_phase_reuse_dx_step,
                thermo_packet_vapor_reuse_dx_step,
                thermo_packet_phase_reuse_dT_step,
                thermo_packet_phase_reuse_dP_step,
            ) = _resolve_step0_startup_packet_phase_reuse_settings(
                startup_seed_loaded=bool(startup_seed_loaded),
                runtime_mode=str(runtime_mode),
                step=int(step),
                last_tray_thermo_packet=last_tray_thermo_packet,
                last_T_tray=last_T_tray,
                last_P_hyd=last_P_hyd,
                last_P_diag=last_P_diag,
                last_z_overall=last_z_overall,
                base_inputs=base_inputs,
            )
            if do_thermo:
                use_top_drum_pressure_as_anchor = bool(
                    pressure_control_enabled
                    and str(pressure_control_mv).strip().lower() == "condenser-duty"
                )
                inputs = ColumnInputs(
                    boundary=step_boundary,
                    volume_model=base_inputs.volume_model,
                    runtime_mode=str(step_runtime_mode),
                    thermo=base_inputs.thermo,
                    thermo_provider=base_inputs.thermo_provider,
                    enable_legacy_temperature_state=base_inputs.enable_legacy_temperature_state,
                    compute_thermo_diag=base_inputs.compute_thermo_diag,
                    equilibrium_relaxation=runtime_equilibrium_relaxation,
                    equilibrium_relaxation_mode=base_inputs.equilibrium_relaxation_mode,
                    equilibrium_tau_ramp_initial_sec=base_inputs.equilibrium_tau_ramp_initial_sec,
                    equilibrium_tau_ramp_final_sec=base_inputs.equilibrium_tau_ramp_final_sec,
                    equilibrium_tau_ramp_decay_sec=base_inputs.equilibrium_tau_ramp_decay_sec,
                    equilibrium_phase_holdup_guard_lbmol=base_inputs.equilibrium_phase_holdup_guard_lbmol,
                    equilibrium_energy_damping_gain=base_inputs.equilibrium_energy_damping_gain,
                    hydraulic_energy_temperature_damping=base_inputs.hydraulic_energy_temperature_damping,
                    hydraulic_energy_temperature_mode=base_inputs.hydraulic_energy_temperature_mode,
                    hydraulic_energy_temperature_follow_tau_sec=base_inputs.hydraulic_energy_temperature_follow_tau_sec,
                    hydraulic_energy_temperature_resid_frac=base_inputs.hydraulic_energy_temperature_resid_frac,
                    hydraulic_energy_temperature_pressure_slope_F_per_psi=(
                        base_inputs.hydraulic_energy_temperature_pressure_slope_F_per_psi
                    ),
                    tray_temp_pressure_slope_prev_F_per_psi=last_tray_temp_pressure_slope,
                    tray_bubble_target_prev_F=last_tray_bubble_target_F,
                    energy_balance_resid_prev_BTUps_tray=last_energy_resid_tray,
                    phase_energy_damping_min_prev_tray=last_phase_energy_damping_min,
                    tau_eq_sec=base_inputs.tau_eq_sec,
                    condenser_alpha=base_inputs.condenser_alpha,
                    clamp_alpha=base_inputs.clamp_alpha,
                    condenser_duty_mode=str(step_condenser_duty_mode),
                    condenser_duty_btu_per_h=(
                        float(step_condenser_duty_btu_per_h)
                        if step_condenser_duty_btu_per_h is not None
                        else None
                    ),
                    condenser_duty_trim_btu_per_h=(
                        float(step_condenser_duty_trim_btu_per_h)
                        if step_condenser_duty_trim_btu_per_h is not None
                        else None
                    ),
                    condenser_duty_partial_condense_if_limited=bool(
                        _allow_coupled_total_condenser_partial_condense(
                            cfg=cfg,
                            pressure_control_mv=str(pressure_control_mv),
                            condenser_duty_mode=str(step_condenser_duty_mode),
                        )
                    ),
                    enable_live_total_condenser_duty=base_inputs.enable_live_total_condenser_duty,
                    condenser_duty_prev=last_condenser_duty_packet,
                    condenser_duty_reuse_dT_F=base_inputs.condenser_duty_reuse_dT_F,
                    condenser_duty_reuse_dP_psia=base_inputs.condenser_duty_reuse_dP_psia,
                    condenser_duty_reuse_dx=base_inputs.condenser_duty_reuse_dx,
                    condenser_duty_reuse_dV_rel=base_inputs.condenser_duty_reuse_dV_rel,
                    reboiler_mode=str(step_reboiler_mode),
                    reboiler_duty_btu_per_h=(
                        float(step_reboiler_duty_btu_per_h)
                        if step_reboiler_duty_btu_per_h is not None
                        else None
                    ),
                    reboiler_duty_trim_btu_per_h=(
                        float(step_reboiler_duty_trim_btu_per_h)
                        if step_reboiler_duty_trim_btu_per_h is not None
                        else None
                    ),
                    reboiler_equilibrium=base_inputs.reboiler_equilibrium,
                    pressure_model=str(pressure_model_step),
                    pressure_top_anchor_psia=(
                        float(step_pressure_top_anchor_psia)
                        if step_pressure_top_anchor_psia is not None
                        else None
                    ),
                    hydraulic_use_top_drum_pressure_as_anchor=use_top_drum_pressure_as_anchor,
                    condenser_pressure_drop_psi=base_inputs.condenser_pressure_drop_psi,
                    top_drum_vapor_volume_ft3=base_inputs.top_drum_vapor_volume_ft3,
                    top_drum_extra_vapor_volume_ft3=base_inputs.top_drum_extra_vapor_volume_ft3,
                    top_drum_total_volume_ft3=base_inputs.top_drum_total_volume_ft3,
                    enforce_top_drum_pressure_gate=base_inputs.enforce_top_drum_pressure_gate,
                    top_drum_pressure_gate_soft_psi=base_inputs.top_drum_pressure_gate_soft_psi,
                    enforce_top_pressure_ordering=base_inputs.enforce_top_pressure_ordering,
                    top_pressure_ordering_margin_psi=base_inputs.top_pressure_ordering_margin_psi,
                    enable_top_drum_psv=base_inputs.enable_top_drum_psv,
                    top_drum_psv_setpoint_psia=base_inputs.top_drum_psv_setpoint_psia,
                    top_drum_psv_gain_lbmolps_per_psi=base_inputs.top_drum_psv_gain_lbmolps_per_psi,
                    top_drum_psv_max_vent_lbmolps=base_inputs.top_drum_psv_max_vent_lbmolps,
                    vapor_flow_model=str(vapor_flow_model_step),
                    vapor_flow_homotopy_beta=(
                        float(seq_vapor_beta_state)
                        if bool(cfg.enable_startup_vapor_homotopy)
                        and str(vapor_flow_model_step).strip().lower() in ("energy", "conductance")
                        else None
                    ),
                    debug_freeze_tray_vapor_derivatives=(
                        base_inputs.debug_freeze_tray_vapor_derivatives
                    ),
                    debug_override_reflux_composition=(
                        base_inputs.debug_override_reflux_composition
                    ),
                    debug_clamp_top_drum_pressure_psia=(
                        base_inputs.debug_clamp_top_drum_pressure_psia
                    ),
                    debug_clamp_top_drum_pressure_duration_sec=(
                        base_inputs.debug_clamp_top_drum_pressure_duration_sec
                    ),
                    total_reflux_startup_ramp_tau_sec=(
                        base_inputs.total_reflux_startup_ramp_tau_sec
                    ),
                    total_reflux_startup_min_ramp_fraction=(
                        base_inputs.total_reflux_startup_min_ramp_fraction
                    ),
                    total_reflux_scale_reflux_with_startup_factor=(
                        base_inputs.total_reflux_scale_reflux_with_startup_factor
                    ),
                    total_reflux_boundary_ramp_duration_sec=(
                        base_inputs.total_reflux_boundary_ramp_duration_sec
                    ),
                    dry_tray_K=base_inputs.dry_tray_K,
                    vapor_holdup_relaxation_sec=base_inputs.vapor_holdup_relaxation_sec,
                    hydraulic_pressure_relaxation_sec=base_inputs.hydraulic_pressure_relaxation_sec,
                    top_drum_pressure_temperature_relaxation_sec=(
                        base_inputs.top_drum_pressure_temperature_relaxation_sec
                    ),
                    top_drum_pressure_T_prev_F=last_top_drum_pressure_T,
                    vapor_flow_relaxation_sec=base_inputs.vapor_flow_relaxation_sec,
                    conductance_vflow_nominal_hi_ratio=(
                        base_inputs.conductance_vflow_nominal_hi_ratio
                    ),
                    reboiler_neighbor_vflow_hi_ratio=base_inputs.reboiler_neighbor_vflow_hi_ratio,
                    reboiler_neighbor_vflow_lo_ratio=base_inputs.reboiler_neighbor_vflow_lo_ratio,
                    thermo_refresh_dT_F=thermo_refresh_dT_step,
                    thermo_refresh_dP_psia=thermo_refresh_dP_step,
                    thermo_refresh_dx=thermo_refresh_dx_step,
                    thermo_packet_phase_reuse_dx=thermo_packet_phase_reuse_dx_step,
                    thermo_packet_vapor_reuse_dx=thermo_packet_vapor_reuse_dx_step,
                    thermo_packet_phase_reuse_dT_F=thermo_packet_phase_reuse_dT_step,
                    thermo_packet_phase_reuse_dP_psia=thermo_packet_phase_reuse_dP_step,
                    feed_stage_flash_prev=last_feed_stage_flash_packet,
                    bottom_sump_cp_prev=last_bottom_sump_cp_packet,
                    enable_liquid_hydraulic_override=base_inputs.enable_liquid_hydraulic_override,
                    liquid_hydraulic_override_alpha=float(np.nanmean(seq_liquid_alpha_state)),
                    liquid_hydraulic_override_alpha_per_stage=np.asarray(
                        seq_liquid_alpha_state,
                        dtype=float,
                    ).reshape((col.n_stages,)).copy(),
                    liquid_hydraulic_model=base_inputs.liquid_hydraulic_model,
                    liquid_hydraulic_htc_sec=base_inputs.liquid_hydraulic_htc_sec,
                    component_mw_lbm_per_lbmol=base_inputs.component_mw_lbm_per_lbmol,
                    progress_hook=base_inputs.progress_hook,
                    P_tray_prev=last_P_hyd if last_P_hyd is not None else last_P_diag,
                    V_out_prev_lbmolph=last_V_out,
                    dT_tray_target_F_per_s=last_dT_tray,
                    T_tray_prev_F=last_T_tray,
                    Z_overall_prev=last_z_overall,
                    rhoL_tray_lbmol_ft3=last_rhoL,
                    tray_thermo_prev=last_tray_thermo_packet,
                    K_tray_prev=last_K_tray,
                    HL_prev=last_HL,
                    HV_prev=last_HV,
                    Zfac_prev=last_Zfac,
                    reb_T_prev=last_reb_T,
                    reb_x_prev=last_reb_x,
                    reb_y_prev=last_reb_y,
                    reb_beta_prev=last_reb_beta,
                )
            else:
                # Skip thermo calls on intermediate steps
                use_top_drum_pressure_as_anchor = bool(
                    pressure_control_enabled
                    and str(pressure_control_mv).strip().lower() == "condenser-duty"
                )
                inputs = ColumnInputs(
                    boundary=step_boundary,
                    volume_model=base_inputs.volume_model,
                    runtime_mode=str(step_runtime_mode),
                    thermo=None,
                    thermo_provider=None,
                    enable_legacy_temperature_state=base_inputs.enable_legacy_temperature_state,
                    compute_thermo_diag=False,
                    equilibrium_relaxation=runtime_equilibrium_relaxation,
                    equilibrium_relaxation_mode=base_inputs.equilibrium_relaxation_mode,
                    equilibrium_tau_ramp_initial_sec=base_inputs.equilibrium_tau_ramp_initial_sec,
                    equilibrium_tau_ramp_final_sec=base_inputs.equilibrium_tau_ramp_final_sec,
                    equilibrium_tau_ramp_decay_sec=base_inputs.equilibrium_tau_ramp_decay_sec,
                    equilibrium_phase_holdup_guard_lbmol=base_inputs.equilibrium_phase_holdup_guard_lbmol,
                    equilibrium_energy_damping_gain=base_inputs.equilibrium_energy_damping_gain,
                    hydraulic_energy_temperature_damping=base_inputs.hydraulic_energy_temperature_damping,
                    hydraulic_energy_temperature_mode=base_inputs.hydraulic_energy_temperature_mode,
                    hydraulic_energy_temperature_follow_tau_sec=base_inputs.hydraulic_energy_temperature_follow_tau_sec,
                    hydraulic_energy_temperature_resid_frac=base_inputs.hydraulic_energy_temperature_resid_frac,
                    hydraulic_energy_temperature_pressure_slope_F_per_psi=(
                        base_inputs.hydraulic_energy_temperature_pressure_slope_F_per_psi
                    ),
                    tray_temp_pressure_slope_prev_F_per_psi=last_tray_temp_pressure_slope,
                    tray_bubble_target_prev_F=last_tray_bubble_target_F,
                    energy_balance_resid_prev_BTUps_tray=last_energy_resid_tray,
                    phase_energy_damping_min_prev_tray=last_phase_energy_damping_min,
                    tau_eq_sec=base_inputs.tau_eq_sec,
                    condenser_duty_mode=str(step_condenser_duty_mode),
                    condenser_duty_btu_per_h=(
                        float(step_condenser_duty_btu_per_h)
                        if step_condenser_duty_btu_per_h is not None
                        else None
                    ),
                    condenser_duty_trim_btu_per_h=(
                        float(step_condenser_duty_trim_btu_per_h)
                        if step_condenser_duty_trim_btu_per_h is not None
                        else None
                    ),
                    condenser_duty_partial_condense_if_limited=bool(
                        _allow_coupled_total_condenser_partial_condense(
                            cfg=cfg,
                            pressure_control_mv=str(pressure_control_mv),
                            condenser_duty_mode=str(step_condenser_duty_mode),
                        )
                    ),
                    enable_live_total_condenser_duty=base_inputs.enable_live_total_condenser_duty,
                    condenser_duty_prev=last_condenser_duty_packet,
                    condenser_duty_reuse_dT_F=base_inputs.condenser_duty_reuse_dT_F,
                    condenser_duty_reuse_dP_psia=base_inputs.condenser_duty_reuse_dP_psia,
                    condenser_duty_reuse_dx=base_inputs.condenser_duty_reuse_dx,
                    condenser_duty_reuse_dV_rel=base_inputs.condenser_duty_reuse_dV_rel,
                    reboiler_mode=str(step_reboiler_mode),
                    reboiler_duty_btu_per_h=(
                        float(step_reboiler_duty_btu_per_h)
                        if step_reboiler_duty_btu_per_h is not None
                        else None
                    ),
                    reboiler_duty_trim_btu_per_h=(
                        float(step_reboiler_duty_trim_btu_per_h)
                        if step_reboiler_duty_trim_btu_per_h is not None
                        else None
                    ),
                    pressure_model=str(pressure_model_step),
                    pressure_top_anchor_psia=(
                        float(step_pressure_top_anchor_psia)
                        if step_pressure_top_anchor_psia is not None
                        else None
                    ),
                    hydraulic_use_top_drum_pressure_as_anchor=use_top_drum_pressure_as_anchor,
                    condenser_pressure_drop_psi=base_inputs.condenser_pressure_drop_psi,
                    top_drum_vapor_volume_ft3=base_inputs.top_drum_vapor_volume_ft3,
                    top_drum_extra_vapor_volume_ft3=base_inputs.top_drum_extra_vapor_volume_ft3,
                    top_drum_total_volume_ft3=base_inputs.top_drum_total_volume_ft3,
                    enforce_top_drum_pressure_gate=base_inputs.enforce_top_drum_pressure_gate,
                    top_drum_pressure_gate_soft_psi=base_inputs.top_drum_pressure_gate_soft_psi,
                    enforce_top_pressure_ordering=base_inputs.enforce_top_pressure_ordering,
                    top_pressure_ordering_margin_psi=base_inputs.top_pressure_ordering_margin_psi,
                    enable_top_drum_psv=base_inputs.enable_top_drum_psv,
                    top_drum_psv_setpoint_psia=base_inputs.top_drum_psv_setpoint_psia,
                    top_drum_psv_gain_lbmolps_per_psi=base_inputs.top_drum_psv_gain_lbmolps_per_psi,
                    top_drum_psv_max_vent_lbmolps=base_inputs.top_drum_psv_max_vent_lbmolps,
                    # Do not run energy-based V closure without live thermo refresh.
                    # Pressure-conductance closure does not require fresh thermo.
                    vapor_flow_model=(
                        "profile"
                        if str(vapor_flow_model_step).strip().lower() == "energy"
                        else str(vapor_flow_model_step)
                    ),
                    vapor_flow_homotopy_beta=(
                        float(seq_vapor_beta_state)
                        if bool(cfg.enable_startup_vapor_homotopy)
                        and str(vapor_flow_model_step).strip().lower() == "conductance"
                        else None
                    ),
                    debug_freeze_tray_vapor_derivatives=(
                        base_inputs.debug_freeze_tray_vapor_derivatives
                    ),
                    debug_override_reflux_composition=(
                        base_inputs.debug_override_reflux_composition
                    ),
                    debug_clamp_top_drum_pressure_psia=(
                        base_inputs.debug_clamp_top_drum_pressure_psia
                    ),
                    debug_clamp_top_drum_pressure_duration_sec=(
                        base_inputs.debug_clamp_top_drum_pressure_duration_sec
                    ),
                    total_reflux_startup_ramp_tau_sec=(
                        base_inputs.total_reflux_startup_ramp_tau_sec
                    ),
                    total_reflux_startup_min_ramp_fraction=(
                        base_inputs.total_reflux_startup_min_ramp_fraction
                    ),
                    total_reflux_scale_reflux_with_startup_factor=(
                        base_inputs.total_reflux_scale_reflux_with_startup_factor
                    ),
                    total_reflux_boundary_ramp_duration_sec=(
                        base_inputs.total_reflux_boundary_ramp_duration_sec
                    ),
                    dry_tray_K=base_inputs.dry_tray_K,
                    vapor_holdup_relaxation_sec=base_inputs.vapor_holdup_relaxation_sec,
                    hydraulic_pressure_relaxation_sec=base_inputs.hydraulic_pressure_relaxation_sec,
                    top_drum_pressure_temperature_relaxation_sec=(
                        base_inputs.top_drum_pressure_temperature_relaxation_sec
                    ),
                    top_drum_pressure_T_prev_F=last_top_drum_pressure_T,
                    vapor_flow_relaxation_sec=base_inputs.vapor_flow_relaxation_sec,
                    conductance_vflow_nominal_hi_ratio=(
                        base_inputs.conductance_vflow_nominal_hi_ratio
                    ),
                    reboiler_neighbor_vflow_hi_ratio=base_inputs.reboiler_neighbor_vflow_hi_ratio,
                    reboiler_neighbor_vflow_lo_ratio=base_inputs.reboiler_neighbor_vflow_lo_ratio,
                    thermo_refresh_dT_F=thermo_refresh_dT_step,
                    thermo_refresh_dP_psia=thermo_refresh_dP_step,
                    thermo_refresh_dx=thermo_refresh_dx_step,
                    feed_stage_flash_prev=last_feed_stage_flash_packet,
                    bottom_sump_cp_prev=last_bottom_sump_cp_packet,
                    enable_liquid_hydraulic_override=base_inputs.enable_liquid_hydraulic_override,
                    liquid_hydraulic_override_alpha=float(np.nanmean(seq_liquid_alpha_state)),
                    liquid_hydraulic_override_alpha_per_stage=np.asarray(
                        seq_liquid_alpha_state,
                        dtype=float,
                    ).reshape((col.n_stages,)).copy(),
                    liquid_hydraulic_model=base_inputs.liquid_hydraulic_model,
                    liquid_hydraulic_htc_sec=base_inputs.liquid_hydraulic_htc_sec,
                    component_mw_lbm_per_lbmol=base_inputs.component_mw_lbm_per_lbmol,
                    progress_hook=base_inputs.progress_hook,
                    P_tray_prev=last_P_hyd if last_P_hyd is not None else last_P_diag,
                    V_out_prev_lbmolph=last_V_out,
                    dT_tray_target_F_per_s=last_dT_tray,
                    T_tray_prev_F=last_T_tray,
                    Z_overall_prev=last_z_overall,
                    rhoL_tray_lbmol_ft3=last_rhoL,
                    tray_thermo_prev=last_tray_thermo_packet,
                    K_tray_prev=last_K_tray,
                    HL_prev=last_HL,
                    HV_prev=last_HV,
                    Zfac_prev=last_Zfac,
                    reb_T_prev=last_reb_T,
                    reb_x_prev=last_reb_x,
                    reb_y_prev=last_reb_y,
                    reb_beta_prev=last_reb_beta,
                )

            if runtime_trace_active:
                inputs = replace(
                    inputs,
                    trace_stage_thermo=True,
                    thermo_stage_trace_label=f"runtime_step_{int(visible_step)}:outer_rhs",
                )
            if (
                startup_seeded_condenser_duty_packet
                and int(visible_step) == 0
                and last_condenser_duty_packet is not None
            ):
                inputs = replace(
                    inputs,
                    condenser_duty_reuse_dT_F=max(
                        float(getattr(inputs, "condenser_duty_reuse_dT_F", 0.0) or 0.0),
                        25.0,
                    ),
                    condenser_duty_reuse_dP_psia=max(
                        float(getattr(inputs, "condenser_duty_reuse_dP_psia", 0.0) or 0.0),
                        25.0,
                    ),
                    condenser_duty_reuse_dx=max(
                        float(getattr(inputs, "condenser_duty_reuse_dx", 0.0) or 0.0),
                        0.25,
                    ),
                    condenser_duty_reuse_dV_rel=max(
                        float(getattr(inputs, "condenser_duty_reuse_dV_rel", 0.0) or 0.0),
                        1.0,
                    ),
                )
            _runtime_trace(
                    "step="
                    f"{int(visible_step)} setup "
                    f"t={float(t_visible_s):.2f}s "
                    f"do_thermo={int(bool(do_thermo))} "
                    f"thermo_reason={str(thermo_execution_reason).strip() or 'n/a'} "
                    f"legacy_temp={int(bool(getattr(inputs, 'enable_legacy_temperature_state', True)))} "
                    f"live_cond={int(bool(getattr(inputs, 'enable_live_total_condenser_duty', True)))} "
                    f"reb_eq={int(bool(getattr(inputs, 'reboiler_equilibrium', True)))} "
                    f"pressure_model={str(getattr(inputs, 'pressure_model', '')).strip() or 'n/a'} "
                    f"vapor_model={str(getattr(inputs, 'vapor_flow_model', '')).strip() or 'n/a'} "
                    f"integrator={str(integrator_mode).strip()}"
                )

            dae_pilot_enabled = bool(dae_pilot_enabled_eff)
            pressure_model_eval = str(getattr(inputs, "pressure_model", "")).strip().lower()
            vapor_model_eval = str(getattr(inputs, "vapor_flow_model", "")).strip().lower()
            dae_pilot_active = bool(
                dae_pilot_enabled
                and pressure_model_eval == "hydraulic"
                and vapor_model_eval in ("energy", "conductance")
            )
            inputs_rhs = inputs
            smooth_cfg_lbmolph = getattr(cfg, "stiff_vflow_smooth_clamp_lbmolph", None)
            smooth_eps_lbmolps = 0.0
            if smooth_cfg_lbmolph is None:
                if (
                    str(integrator_mode).strip().lower() != "explicit-euler"
                    and pressure_model_eval == "hydraulic"
                    and vapor_model_eval in ("energy", "conductance")
                ):
                    # Small smoothing width to remove hard derivative kinks
                    # at vapor-flow clamps for stiff substep Jacobians.
                    smooth_eps_lbmolps = 5.0 / 3600.0
            else:
                try:
                    smooth_try_lbmolph = float(smooth_cfg_lbmolph)
                except Exception:
                    smooth_try_lbmolph = 0.0
                if np.isfinite(smooth_try_lbmolph) and smooth_try_lbmolph > 0.0:
                    smooth_eps_lbmolps = float(smooth_try_lbmolph) / 3600.0
            if smooth_eps_lbmolps > 0.0:
                try:
                    inputs_rhs = replace(
                        inputs_rhs,
                        vflow_smooth_clamp_epsilon_lbmolps=float(smooth_eps_lbmolps),
                    )
                except Exception:
                    pass
            pv_inner_iters = int(getattr(cfg, "pv_inner_max_iter", 1))
            if pressure_model_eval != "hydraulic":
                pv_inner_iters = 1
            if vapor_model_eval not in ("energy", "conductance"):
                pv_inner_iters = 1
            dae_outer_once_for_stiff = bool(
                dae_pilot_active and str(integrator_mode).strip().lower() in ("bdf", "radau")
            )

            def _eval_pv_rhs(
                t_eval_s: float,
                y_eval: np.ndarray,
                eval_inputs: ColumnInputs,
            ) -> Tuple[np.ndarray, Dict[str, np.ndarray]]:
                trace_label = str(getattr(eval_inputs, "thermo_stage_trace_label", "") or "").strip()
                provider = getattr(eval_inputs, "thermo_provider", None)
                if trace_label:
                    _runtime_trace(f"{trace_label} enter t={float(t_eval_s):.6g}s")
                try:
                    if trace_label and hasattr(provider, "set_debug_trace_context"):
                        provider.set_debug_trace_context(trace_label)
                    dydt_eval, diag_eval = _column_rhs_with_inner_pv_coupling(
                        t_s=float(t_eval_s),
                        y=np.asarray(y_eval, dtype=float),
                        col=step_col,
                        layout=layout,
                        inputs=eval_inputs,
                        max_iter=int(pv_inner_iters),
                        p_tol_psia=getattr(cfg, "pv_inner_p_tol_psia", None),
                        v_tol_lbmolph=getattr(cfg, "pv_inner_v_tol_lbmolph", None),
                    )
                finally:
                    if trace_label and hasattr(provider, "set_debug_trace_context"):
                        try:
                            provider.set_debug_trace_context("")
                        except Exception:
                            pass
                if trace_label:
                    _runtime_trace(f"{trace_label} return")
                return np.asarray(dydt_eval, dtype=float).reshape((-1,)), dict(diag_eval)

            def _make_step_rhs(eval_inputs: ColumnInputs):
                if dae_outer_once_for_stiff:
                    def _eval_step_rhs(t_eval_s: float, y_eval: np.ndarray) -> Tuple[np.ndarray, Dict[str, np.ndarray]]:
                        trace_label = str(getattr(eval_inputs, "thermo_stage_trace_label", "") or "").strip()
                        provider = getattr(eval_inputs, "thermo_provider", None)
                        if trace_label:
                            _runtime_trace(f"{trace_label} enter t={float(t_eval_s):.6g}s")
                        try:
                            if trace_label and hasattr(provider, "set_debug_trace_context"):
                                provider.set_debug_trace_context(trace_label)
                            dydt_eval, diag_eval = column_rhs(
                                float(t_eval_s),
                                np.asarray(y_eval, dtype=float),
                                step_col,
                                layout,
                                inputs=eval_inputs,
                            )
                        finally:
                            if trace_label and hasattr(provider, "set_debug_trace_context"):
                                try:
                                    provider.set_debug_trace_context("")
                                except Exception:
                                    pass
                        if trace_label:
                            _runtime_trace(f"{trace_label} return")
                        dydt_eval = np.asarray(dydt_eval, dtype=float).reshape((-1,))
                        diag_eval = dict(diag_eval)
                        diag_eval["dae_pilot_enabled"] = np.array([1.0], dtype=float)
                        diag_eval["dae_pilot_iter_count"] = np.array([0.0], dtype=float)
                        diag_eval["dae_pilot_converged"] = np.array([0.0], dtype=float)
                        diag_eval["dae_pilot_failed"] = np.array([0.0], dtype=float)
                        diag_eval["dae_pilot_stiff_outer_only"] = np.array([1.0], dtype=float)
                        return dydt_eval, diag_eval
                else:
                    def _eval_step_rhs(t_eval_s: float, y_eval: np.ndarray) -> Tuple[np.ndarray, Dict[str, np.ndarray]]:
                        if dae_pilot_active:
                            dydt_eval, diag_eval = _solve_dae_pilot_algebraic(
                                t_s=float(t_eval_s),
                                y=np.asarray(y_eval, dtype=float),
                                col=step_col,
                                layout=layout,
                                inputs=eval_inputs,
                                max_iter=int(dae_pilot_max_iter_eff),
                                p_tol_psia=dae_pilot_p_tol_eff,
                                v_tol_lbmolph=dae_pilot_v_tol_eff,
                                jac_rel_step=float(getattr(cfg, "dae_pilot_jac_rel_step", 1.0e-6)),
                                line_search_max=int(getattr(cfg, "dae_pilot_line_search_max", 4)),
                            )
                            return np.asarray(dydt_eval, dtype=float).reshape((-1,)), dict(diag_eval)

                        dydt_eval, diag_eval = _eval_pv_rhs(
                            float(t_eval_s),
                            np.asarray(y_eval, dtype=float),
                            eval_inputs,
                        )
                        if dae_pilot_enabled:
                            diag_eval["dae_pilot_enabled"] = np.array([1.0], dtype=float)
                            diag_eval["dae_pilot_iter_count"] = np.array([0.0], dtype=float)
                            diag_eval["dae_pilot_converged"] = np.array([0.0], dtype=float)
                            diag_eval["dae_pilot_failed"] = np.array([0.0], dtype=float)
                        return dydt_eval, diag_eval
                return _eval_step_rhs

            if dae_outer_once_for_stiff:
                # Run full DAE algebraic Newton once per outer step for diagnostics
                # and for a better initial algebraic seed.
                dydt, diag = _solve_dae_pilot_algebraic(
                    t_s=float(t_s),
                    y=np.asarray(y, dtype=float),
                    col=step_col,
                    layout=layout,
                    inputs=inputs_rhs,
                    max_iter=int(dae_pilot_max_iter_eff),
                    p_tol_psia=dae_pilot_p_tol_eff,
                    v_tol_lbmolph=dae_pilot_v_tol_eff,
                    jac_rel_step=float(getattr(cfg, "dae_pilot_jac_rel_step", 1.0e-6)),
                    line_search_max=int(getattr(cfg, "dae_pilot_line_search_max", 4)),
                )
                diag["dae_pilot_stiff_outer_only"] = np.array([1.0], dtype=float)

                # Feed solved algebraic quantities into stiff substeps.
                try:
                    rhs_updates: Dict[str, Any] = {}
                    p_seed = _diag_stage_vector(diag, "P_psia_hyd", col.n_stages, positive_only=True)
                    if p_seed is None:
                        p_seed = _diag_stage_vector(diag, "P_psia_diag", col.n_stages, positive_only=True)
                    if p_seed is not None:
                        rhs_updates["P_tray_prev"] = np.asarray(p_seed, dtype=float).copy()
                    v_seed = _diag_stage_vector(diag, "V_out_lbmolph", col.n_stages, positive_only=False)
                    if v_seed is not None:
                        rhs_updates["V_out_prev_lbmolph"] = np.asarray(v_seed, dtype=float).copy()
                    z_seed = _diag_stage_vector(diag, "Z_tray", col.n_stages, positive_only=True)
                    if z_seed is not None:
                        rhs_updates["Zfac_prev"] = np.asarray(z_seed, dtype=float).copy()
                    if "T_top_drum_pressure_used_F" in diag:
                        try:
                            t_top_used = float(np.asarray(diag["T_top_drum_pressure_used_F"], dtype=float).reshape((-1,))[0])
                            if np.isfinite(t_top_used):
                                rhs_updates["top_drum_pressure_T_prev_F"] = float(t_top_used)
                        except Exception:
                            pass
                    if rhs_updates:
                        inputs_rhs = replace(inputs_rhs, **rhs_updates)
                except Exception:
                    pass
            inputs_rhs_fallback = inputs_rhs
            if runtime_trace_active:
                inputs_rhs_fallback = replace(
                    inputs_rhs_fallback,
                    trace_stage_thermo=True,
                    thermo_stage_trace_label=f"runtime_step_{int(visible_step)}:outer_rhs",
                )
            _eval_step_rhs = _make_step_rhs(inputs_rhs_fallback)
            if not dae_outer_once_for_stiff:
                dydt, diag = _eval_step_rhs(float(t_s), y)

            inputs_rhs_integrator = inputs_rhs_fallback
            if integrator_mode != "explicit-euler" and do_thermo:
                try:
                    # Freeze thermo flashes only for the attempted stiff substep
                    # evaluations; explicit fallback must still use the live
                    # outer-step RHS.
                    inputs_rhs_integrator = replace(
                        inputs_rhs_integrator,
                        thermo=None,
                        thermo_provider=None,
                        compute_thermo_diag=False,
                        equilibrium_relaxation=False,
                    )
                except Exception:
                    inputs_rhs_integrator = inputs_rhs_fallback
            if runtime_trace_active:
                inputs_rhs_integrator = replace(
                    inputs_rhs_integrator,
                    trace_stage_thermo=True,
                    thermo_stage_trace_label=f"runtime_step_{int(visible_step)}:integrator_rhs",
                )
            _eval_step_rhs_integrator = _make_step_rhs(inputs_rhs_integrator)
            if (
                step_condenser_duty_cmd_btu_per_h is not None
                and np.isfinite(float(step_condenser_duty_cmd_btu_per_h))
            ):
                diag["Q_cond_cmd_BTUph"] = np.array([float(step_condenser_duty_cmd_btu_per_h)], dtype=float)
            elif step_condenser_duty_btu_per_h is not None and np.isfinite(float(step_condenser_duty_btu_per_h)):
                diag["Q_cond_cmd_BTUph"] = np.array([float(step_condenser_duty_btu_per_h)], dtype=float)
            if step_pressure_ctrl_pv_raw_psia is not None and np.isfinite(float(step_pressure_ctrl_pv_raw_psia)):
                diag["P_top_ctrl_pv_raw_psia"] = np.array([float(step_pressure_ctrl_pv_raw_psia)], dtype=float)
            if step_pressure_ctrl_pv_filt_psia is not None and np.isfinite(float(step_pressure_ctrl_pv_filt_psia)):
                diag["P_top_ctrl_pv_filt_psia"] = np.array([float(step_pressure_ctrl_pv_filt_psia)], dtype=float)
            if step_pressure_ctrl_gain_scale is not None and np.isfinite(float(step_pressure_ctrl_gain_scale)):
                diag["P_top_ctrl_gain_scale"] = np.array([float(step_pressure_ctrl_gain_scale)], dtype=float)
            if (
                step_pressure_ctrl_energy_resid_abs_btups is not None
                and np.isfinite(float(step_pressure_ctrl_energy_resid_abs_btups))
            ):
                diag["P_top_ctrl_energy_resid_abs_BTUps"] = np.array(
                    [float(step_pressure_ctrl_energy_resid_abs_btups)],
                    dtype=float,
                )
            if step_pressure_ctrl_error_psia is not None and np.isfinite(float(step_pressure_ctrl_error_psia)):
                diag["P_top_ctrl_err_psia"] = np.array([float(step_pressure_ctrl_error_psia)], dtype=float)
            if step_pressure_ctrl_p_term is not None and np.isfinite(float(step_pressure_ctrl_p_term)):
                diag["P_top_ctrl_p_term"] = np.array([float(step_pressure_ctrl_p_term)], dtype=float)
            if step_pressure_ctrl_i_term is not None and np.isfinite(float(step_pressure_ctrl_i_term)):
                diag["P_top_ctrl_i_term"] = np.array([float(step_pressure_ctrl_i_term)], dtype=float)
            if (
                step_pressure_ctrl_u_unclamped_btuph is not None
                and np.isfinite(float(step_pressure_ctrl_u_unclamped_btuph))
            ):
                diag["P_top_ctrl_u_unclamped_BTUph"] = np.array(
                    [float(step_pressure_ctrl_u_unclamped_btuph)],
                    dtype=float,
                )
            if step_pressure_ctrl_sat_hi is not None and np.isfinite(float(step_pressure_ctrl_sat_hi)):
                diag["P_top_ctrl_sat_hi"] = np.array([float(step_pressure_ctrl_sat_hi)], dtype=float)
            if step_pressure_ctrl_sat_lo is not None and np.isfinite(float(step_pressure_ctrl_sat_lo)):
                diag["P_top_ctrl_sat_lo"] = np.array([float(step_pressure_ctrl_sat_lo)], dtype=float)
            if step_pressure_ctrl_allow_int is not None and np.isfinite(float(step_pressure_ctrl_allow_int)):
                diag["P_top_ctrl_allow_int"] = np.array([float(step_pressure_ctrl_allow_int)], dtype=float)
            if step_top_level_ctrl_pv is not None and np.isfinite(float(step_top_level_ctrl_pv)):
                diag["Top_level_ctrl_pv"] = np.array([float(step_top_level_ctrl_pv)], dtype=float)
            if step_top_level_ctrl_sp is not None and np.isfinite(float(step_top_level_ctrl_sp)):
                diag["Top_level_ctrl_sp"] = np.array([float(step_top_level_ctrl_sp)], dtype=float)
            if step_top_level_ctrl_err is not None and np.isfinite(float(step_top_level_ctrl_err)):
                diag["Top_level_ctrl_err"] = np.array([float(step_top_level_ctrl_err)], dtype=float)
            if step_top_level_ctrl_p_term is not None and np.isfinite(float(step_top_level_ctrl_p_term)):
                diag["Top_level_ctrl_p_term"] = np.array([float(step_top_level_ctrl_p_term)], dtype=float)
            if step_top_level_ctrl_i_term is not None and np.isfinite(float(step_top_level_ctrl_i_term)):
                diag["Top_level_ctrl_i_term"] = np.array([float(step_top_level_ctrl_i_term)], dtype=float)
            if step_top_level_ctrl_pi_term_lbmolph is not None and np.isfinite(float(step_top_level_ctrl_pi_term_lbmolph)):
                diag["Top_level_ctrl_pi_term_lbmolph"] = np.array([float(step_top_level_ctrl_pi_term_lbmolph)], dtype=float)
            if step_top_level_ctrl_ff_term_lbmolph is not None and np.isfinite(float(step_top_level_ctrl_ff_term_lbmolph)):
                diag["Top_level_ctrl_ff_term_lbmolph"] = np.array([float(step_top_level_ctrl_ff_term_lbmolph)], dtype=float)
            if (
                step_top_level_ctrl_u_unclamped_lbmolph is not None
                and np.isfinite(float(step_top_level_ctrl_u_unclamped_lbmolph))
            ):
                diag["Top_level_ctrl_u_unclamped_lbmolph"] = np.array(
                    [float(step_top_level_ctrl_u_unclamped_lbmolph)], dtype=float
                )
            if step_top_level_ctrl_sat_hi is not None and np.isfinite(float(step_top_level_ctrl_sat_hi)):
                diag["Top_level_ctrl_sat_hi"] = np.array([float(step_top_level_ctrl_sat_hi)], dtype=float)
            if step_top_level_ctrl_sat_lo is not None and np.isfinite(float(step_top_level_ctrl_sat_lo)):
                diag["Top_level_ctrl_sat_lo"] = np.array([float(step_top_level_ctrl_sat_lo)], dtype=float)
            if step_top_level_ctrl_allow_int is not None and np.isfinite(float(step_top_level_ctrl_allow_int)):
                diag["Top_level_ctrl_allow_int"] = np.array([float(step_top_level_ctrl_allow_int)], dtype=float)
            if step_bottom_level_ctrl_pv is not None and np.isfinite(float(step_bottom_level_ctrl_pv)):
                diag["Bottom_level_ctrl_pv"] = np.array([float(step_bottom_level_ctrl_pv)], dtype=float)
            if step_bottom_level_ctrl_sp is not None and np.isfinite(float(step_bottom_level_ctrl_sp)):
                diag["Bottom_level_ctrl_sp"] = np.array([float(step_bottom_level_ctrl_sp)], dtype=float)
            if step_bottom_level_ctrl_err is not None and np.isfinite(float(step_bottom_level_ctrl_err)):
                diag["Bottom_level_ctrl_err"] = np.array([float(step_bottom_level_ctrl_err)], dtype=float)
            if step_bottom_level_ctrl_p_term is not None and np.isfinite(float(step_bottom_level_ctrl_p_term)):
                diag["Bottom_level_ctrl_p_term"] = np.array([float(step_bottom_level_ctrl_p_term)], dtype=float)
            if step_bottom_level_ctrl_i_term is not None and np.isfinite(float(step_bottom_level_ctrl_i_term)):
                diag["Bottom_level_ctrl_i_term"] = np.array([float(step_bottom_level_ctrl_i_term)], dtype=float)
            if (
                step_bottom_level_ctrl_u_unclamped_lbmolph is not None
                and np.isfinite(float(step_bottom_level_ctrl_u_unclamped_lbmolph))
            ):
                diag["Bottom_level_ctrl_u_unclamped_lbmolph"] = np.array(
                    [float(step_bottom_level_ctrl_u_unclamped_lbmolph)], dtype=float
                )
            if step_bottom_level_ctrl_sat_hi is not None and np.isfinite(float(step_bottom_level_ctrl_sat_hi)):
                diag["Bottom_level_ctrl_sat_hi"] = np.array([float(step_bottom_level_ctrl_sat_hi)], dtype=float)
            if step_bottom_level_ctrl_sat_lo is not None and np.isfinite(float(step_bottom_level_ctrl_sat_lo)):
                diag["Bottom_level_ctrl_sat_lo"] = np.array([float(step_bottom_level_ctrl_sat_lo)], dtype=float)
            if step_bottom_level_ctrl_allow_int is not None and np.isfinite(float(step_bottom_level_ctrl_allow_int)):
                diag["Bottom_level_ctrl_allow_int"] = np.array([float(step_bottom_level_ctrl_allow_int)], dtype=float)
            if (
                step_pressure_top_anchor_cmd_psia is not None
                and np.isfinite(float(step_pressure_top_anchor_cmd_psia))
            ):
                diag["P_top_anchor_cmd_psia"] = np.array([float(step_pressure_top_anchor_cmd_psia)], dtype=float)
            if step_distillate_comp_sp is not None and np.isfinite(float(step_distillate_comp_sp)):
                diag["xD_comp_sp"] = np.array([float(step_distillate_comp_sp)], dtype=float)
            if step_distillate_comp_pv is not None and np.isfinite(float(step_distillate_comp_pv)):
                diag["xD_comp_pv"] = np.array([float(step_distillate_comp_pv)], dtype=float)
            if step_distillate_comp_error is not None and np.isfinite(float(step_distillate_comp_error)):
                diag["xD_comp_err"] = np.array([float(step_distillate_comp_error)], dtype=float)
            if step_distillate_comp_u_unclamped is not None and np.isfinite(float(step_distillate_comp_u_unclamped)):
                diag["Reflux_cmd_unclamped_lbmolph"] = np.array([float(step_distillate_comp_u_unclamped)], dtype=float)
            if step_distillate_comp_out_max is not None and np.isfinite(float(step_distillate_comp_out_max)):
                diag["Reflux_cmd_active_max_lbmolph"] = np.array([float(step_distillate_comp_out_max)], dtype=float)
            if step_distillate_comp_cap_active is not None and np.isfinite(float(step_distillate_comp_cap_active)):
                diag["Reflux_cap_active_flag"] = np.array([float(step_distillate_comp_cap_active)], dtype=float)
            if step_reflux_ratio_cmd is not None and np.isfinite(float(step_reflux_ratio_cmd)):
                diag["RR_comp_cmd"] = np.array([float(step_reflux_ratio_cmd)], dtype=float)
            if step_reflux_cmd_lbmolph is not None and np.isfinite(float(step_reflux_cmd_lbmolph)):
                diag["Reflux_cmd_lbmolph"] = np.array([float(step_reflux_cmd_lbmolph)], dtype=float)
            if step_bottoms_comp_sp is not None and np.isfinite(float(step_bottoms_comp_sp)):
                diag["xB_comp_sp"] = np.array([float(step_bottoms_comp_sp)], dtype=float)
            if step_bottoms_comp_pv is not None and np.isfinite(float(step_bottoms_comp_pv)):
                diag["xB_comp_pv"] = np.array([float(step_bottoms_comp_pv)], dtype=float)
            if step_boilup_cmd_lbmolph is not None and np.isfinite(float(step_boilup_cmd_lbmolph)):
                diag["Boilup_cmd_lbmolph"] = np.array([float(step_boilup_cmd_lbmolph)], dtype=float)
            if (
                step_reboiler_duty_cmd_btu_per_h is not None
                and np.isfinite(float(step_reboiler_duty_cmd_btu_per_h))
            ):
                diag["Q_reb_cmd_BTUph"] = np.array([float(step_reboiler_duty_cmd_btu_per_h)], dtype=float)
            seq_phase_id = {
                "base": 0.0,
                "pressure_only": 1.0,
                "pressure_energy": 2.0,
                "pressure_energy_liquid_ramp": 3.0,
                "profile_hold": 4.0,
                "vapor_homotopy_ramp": 5.0,
            }.get(str(seq_phase), 0.0)
            diag["startup_sequence_enabled"] = np.array(
                [1.0 if bool(startup_sequence_enabled) else 0.0],
                dtype=float,
            )
            diag["startup_total_reflux_washout_active"] = np.array(
                [1.0 if bool(startup_total_reflux_washout_active) else 0.0],
                dtype=float,
            )
            diag["startup_sequence_phase_id"] = np.array([float(seq_phase_id)], dtype=float)
            diag["startup_sequence_pressure_hydraulic"] = np.array(
                [1.0 if str(pressure_model_step).strip().lower() == "hydraulic" else 0.0],
                dtype=float,
            )
            diag["startup_sequence_vapor_energy"] = np.array(
                [1.0 if str(vapor_flow_model_step).strip().lower() == "energy" else 0.0],
                dtype=float,
            )
            diag["startup_sequence_liquid_alpha_cmd"] = np.array(
                [float(np.nanmean(np.asarray(seq_liquid_alpha_state, dtype=float)))],
                dtype=float,
            )
            diag["startup_sequence_liquid_alpha_cmd_per_stage"] = np.asarray(
                seq_liquid_alpha_state,
                dtype=float,
            ).reshape((col.n_stages,)).copy()
            diag["startup_sequence_vapor_homotopy_enabled"] = np.array(
                [1.0 if bool(cfg.enable_startup_vapor_homotopy) else 0.0],
                dtype=float,
            )
            diag["startup_sequence_vapor_beta_cmd"] = np.array(
                [float(seq_vapor_beta_state)],
                dtype=float,
            )
            diag["startup_sequence_vapor_guard_phase_id"] = np.array(
                [
                    {
                        "disabled": 0.0,
                        "profile_hold": 1.0,
                        "ramp": 2.0,
                        "guard_hold": 3.0,
                        "backoff": 4.0,
                        "hold": 5.0,
                        "complete": 6.0,
                    }.get(str(seq_vapor_guard_phase), 0.0)
                ],
                dtype=float,
            )
            diag["startup_sequence_vapor_rel_rate_gate_per_s"] = np.array(
                [
                    float(cfg.startup_sequence_vapor_rel_rate_gate_per_s)
                    if cfg.startup_sequence_vapor_rel_rate_gate_per_s is not None
                    and np.isfinite(float(cfg.startup_sequence_vapor_rel_rate_gate_per_s))
                    else np.nan
                ],
                dtype=float,
            )
            diag["startup_sequence_rel_rate_prev_per_s"] = np.array(
                [
                    float(last_ss_rel_state_rate_per_s)
                    if last_ss_rel_state_rate_per_s is not None
                    and np.isfinite(float(last_ss_rel_state_rate_per_s))
                    else np.nan
                ],
                dtype=float,
            )
            diag["liquid_hydraulic_guard_phase_id"] = np.array(
                [
                    {
                        "disabled": 0.0,
                        "hold": 1.0,
                        "recover": 2.0,
                        "backoff": 3.0,
                        "sequence": 4.0,
                        "no-gate": 5.0,
                    }.get(str(seq_liquid_guard_phase), 0.0)
                ],
                dtype=float,
            )
            gate_val = cfg.startup_sequence_mass_resid_gate_lbmolph
            diag["startup_sequence_mass_resid_gate_lbmolph"] = np.array(
                [
                    float(gate_val)
                    if gate_val is not None and np.isfinite(float(gate_val))
                    else np.nan
                ],
                dtype=float,
            )
            diag["startup_sequence_mass_resid_max_prev_lbmolph"] = np.array(
                [
                    float(last_mass_resid_max_lbmolph)
                    if last_mass_resid_max_lbmolph is not None and np.isfinite(float(last_mass_resid_max_lbmolph))
                    else np.nan
                ],
                dtype=float,
            )
            try:
                mr_now = np.asarray(diag.get("mass_balance_resid_lbmolps_tray"), dtype=float).reshape((-1,))
                if mr_now.size > 0 and np.any(np.isfinite(mr_now)):
                    last_mass_resid_max_lbmolph = float(np.nanmax(np.abs(mr_now)) * 3600.0)
                    last_mass_resid_lbmolph_per_stage = np.abs(np.asarray(mr_now, dtype=float)) * 3600.0
                    diag["startup_sequence_mass_resid_max_now_lbmolph"] = np.array(
                        [float(last_mass_resid_max_lbmolph)],
                        dtype=float,
                    )
                else:
                    last_mass_resid_max_lbmolph = None
                    last_mass_resid_lbmolph_per_stage = None
            except Exception:
                last_mass_resid_max_lbmolph = None
                last_mass_resid_lbmolph_per_stage = None

            # Global mass-closure diagnostics (diagnostic-only; no correction).
            try:
                m_total_lbmol = _total_inventory_lbmol(layout, y)
                dM_total_dt_lbmolps = _total_inventory_rate_lbmolps(layout, dydt)
                dM_total_dt_lbmolph = float(dM_total_dt_lbmolps) * 3600.0

                F_flow = float(feed_tag.flow_lbmolph) if feed_tag.flow_lbmolph is not None else np.nan
                D_flow = np.nan
                B_flow = np.nan
                if step_boundary.distillate_lbmolph is not None:
                    D_flow = float(step_boundary.distillate_lbmolph)
                elif step_dist_tag.flow_lbmolph is not None:
                    D_flow = float(step_dist_tag.flow_lbmolph)
                if step_boundary.bottoms_lbmolph is not None:
                    B_flow = float(step_boundary.bottoms_lbmolph)
                elif step_bots_tag.flow_lbmolph is not None:
                    B_flow = float(step_bots_tag.flow_lbmolph)

                net_in_minus_out_lbmolph = np.nan
                if np.isfinite(F_flow) and np.isfinite(D_flow) and np.isfinite(B_flow):
                    net_in_minus_out_lbmolph = float(F_flow - D_flow - B_flow)

                closure_error_lbmolph = np.nan
                if np.isfinite(dM_total_dt_lbmolph) and np.isfinite(net_in_minus_out_lbmolph):
                    closure_error_lbmolph = float(dM_total_dt_lbmolph - net_in_minus_out_lbmolph)
                    if step > 0 and np.isfinite(closure_error_lbmolph):
                        global_mass_closure_cum_lbmol += float(closure_error_lbmolph) * float(dt) / 3600.0

                stage_mass_resid_sum_lbmolps = np.nan
                if "mass_balance_resid_lbmolps_tray" in diag:
                    try:
                        mr = np.asarray(diag["mass_balance_resid_lbmolps_tray"], dtype=float).reshape((-1,))
                        if mr.size > 0:
                            stage_mass_resid_sum_lbmolps = float(np.nansum(mr))
                    except Exception:
                        stage_mass_resid_sum_lbmolps = np.nan

                diag["M_total_lbmol"] = np.array([float(m_total_lbmol)], dtype=float)
                diag["dM_total_dt_lbmolps"] = np.array([float(dM_total_dt_lbmolps)], dtype=float)
                diag["dM_total_dt_lbmolph"] = np.array([float(dM_total_dt_lbmolph)], dtype=float)
                diag["net_F_minus_D_minus_B_lbmolph"] = np.array([float(net_in_minus_out_lbmolph)], dtype=float)
                diag["global_mass_closure_error_lbmolph"] = np.array([float(closure_error_lbmolph)], dtype=float)
                diag["global_mass_closure_cum_lbmol"] = np.array([float(global_mass_closure_cum_lbmol)], dtype=float)
                diag["stage_mass_resid_sum_lbmolps"] = np.array([float(stage_mass_resid_sum_lbmolps)], dtype=float)
            except Exception:
                pass

            # Pressure-controller PV source.
            # For top-anchor MV, use tray-top hydraulic pressure as the primary PV
            # so the controller acts on the pressure state it actually manipulates.
            p_ctrl_idx = 0
            p_top_pv = None
            pressure_mv_mode = str(pressure_control_mv or "").strip().lower()
            if pressure_mv_mode == "top-anchor":
                if "P_psia_hyd" in diag:
                    try:
                        p_h = np.asarray(diag["P_psia_hyd"], dtype=float).reshape((col.n_stages,))
                        if np.isfinite(float(p_h[p_ctrl_idx])) and float(p_h[p_ctrl_idx]) > 0.0:
                            p_top_pv = float(p_h[p_ctrl_idx])
                    except Exception:
                        p_top_pv = None
                if p_top_pv is None and "P_psia_diag" in diag:
                    try:
                        p_d = np.asarray(diag["P_psia_diag"], dtype=float).reshape((col.n_stages,))
                        if np.isfinite(float(p_d[p_ctrl_idx])) and float(p_d[p_ctrl_idx]) > 0.0:
                            p_top_pv = float(p_d[p_ctrl_idx])
                    except Exception:
                        p_top_pv = None
                if p_top_pv is None and "P_top_drum_psia" in diag:
                    try:
                        p_top = float(np.asarray(diag["P_top_drum_psia"], dtype=float).reshape((-1,))[0])
                        if np.isfinite(p_top) and p_top > 0.0:
                            p_top_pv = float(p_top)
                    except Exception:
                        p_top_pv = None
            else:
                if "P_top_drum_psia" in diag:
                    try:
                        p_top = float(np.asarray(diag["P_top_drum_psia"], dtype=float).reshape((-1,))[0])
                        if np.isfinite(p_top) and p_top > 0.0:
                            p_top_pv = float(p_top)
                    except Exception:
                        p_top_pv = None
                if p_top_pv is None and "P_psia_hyd" in diag:
                    try:
                        p_h = np.asarray(diag["P_psia_hyd"], dtype=float).reshape((col.n_stages,))
                        if np.isfinite(float(p_h[p_ctrl_idx])) and float(p_h[p_ctrl_idx]) > 0.0:
                            p_top_pv = float(p_h[p_ctrl_idx])
                    except Exception:
                        p_top_pv = None
                if p_top_pv is None and "P_psia_diag" in diag:
                    try:
                        p_d = np.asarray(diag["P_psia_diag"], dtype=float).reshape((col.n_stages,))
                        if np.isfinite(float(p_d[p_ctrl_idx])) and float(p_d[p_ctrl_idx]) > 0.0:
                            p_top_pv = float(p_d[p_ctrl_idx])
                    except Exception:
                        p_top_pv = None
            if p_top_pv is not None:
                last_top_pressure_pv_psia = float(p_top_pv)
                diag["P_top_ctrl_pv_psia"] = np.array([float(p_top_pv)], dtype=float)
            if "T_top_drum_pressure_used_F" in diag:
                try:
                    t_top_p = float(np.asarray(diag["T_top_drum_pressure_used_F"], dtype=float).reshape((-1,))[0])
                    if np.isfinite(t_top_p):
                        last_top_drum_pressure_T = float(t_top_p)
                except Exception:
                    pass
            last_top_energy_resid_abs_btups = None
            if "energy_balance_resid_BTUps_tray" in diag:
                try:
                    er = np.asarray(diag["energy_balance_resid_BTUps_tray"], dtype=float).reshape((col.n_stages,))
                    if er.size > 0:
                        finite_idx = np.where(np.isfinite(er))[0]
                        if finite_idx.size > 0:
                            last_top_energy_resid_abs_btups = abs(float(er[int(finite_idx[0])]))
                except Exception:
                    last_top_energy_resid_abs_btups = None

            # Steady-state detector (diagnostic-only).
            if ss_enabled and visible_logging_phase:
                max_rel_state_rate = np.nan
                max_rel_state_detail: Dict[str, Any] = {
                    "max_rel_rate_per_s": np.nan,
                    "state_key": "",
                    "stage_1based": np.nan,
                    "component_1based": np.nan,
                }
                dt_ss = np.nan
                if ss_prev_t_s is not None and ss_prev_y is not None:
                    try:
                        dt_ss = float(t_s) - float(ss_prev_t_s)
                    except Exception:
                        dt_ss = np.nan
                if np.isfinite(dt_ss) and float(dt_ss) > 0.0 and ss_prev_y is not None:
                    max_rel_state_detail = _max_rel_inventory_fd_rate_detail_per_s(
                        layout,
                        ss_prev_y,
                        y,
                        dt_sec=float(dt_ss),
                        denom_floor_lbmol=float(ss_rate_floor_lbmol),
                    )
                    max_rel_state_rate = float(max_rel_state_detail.get("max_rel_rate_per_s", np.nan))
                # Startup fallback for very first sample when finite-difference rate is unavailable.
                if not np.isfinite(max_rel_state_rate):
                    max_rel_state_detail = _max_rel_inventory_rate_detail_per_s(
                        layout,
                        y,
                        dydt,
                        denom_floor_lbmol=float(ss_rate_floor_lbmol),
                    )
                    max_rel_state_rate = float(max_rel_state_detail.get("max_rel_rate_per_s", np.nan))
                max_temp_rate = np.nan
                if np.isfinite(dt_ss) and float(dt_ss) > 0.0 and ss_prev_y is not None:
                    max_temp_rate = _max_abs_temperature_fd_rate_per_s(
                        layout,
                        ss_prev_y,
                        y,
                        dt_sec=float(dt_ss),
                    )
                # Startup fallback when FD estimate is unavailable.
                if not np.isfinite(max_temp_rate) and "dT_tray_F_per_s" in diag:
                    try:
                        dT_vec = np.asarray(diag["dT_tray_F_per_s"], dtype=float).reshape((-1,))
                        dT_f = np.abs(dT_vec[np.isfinite(dT_vec)])
                        if dT_f.size > 0:
                            max_temp_rate = float(np.max(dT_f))
                    except Exception:
                        max_temp_rate = np.nan

                u_now = layout.unpack(y)
                xD_now = np.nan
                if dist_comp_idx is not None and "top_L" in u_now:
                    try:
                        top_L = np.asarray(u_now["top_L"], dtype=float).reshape((-1,))
                        top_tot = float(np.sum(top_L))
                        if np.isfinite(top_tot) and top_tot > 1e-300:
                            xD_now = float(top_L[int(dist_comp_idx)] / top_tot)
                    except Exception:
                        xD_now = np.nan
                xB_now = np.nan
                if bot_comp_idx is not None and "bottom_L" in u_now:
                    try:
                        bot_L = np.asarray(u_now["bottom_L"], dtype=float).reshape((-1,))
                        bot_tot = float(np.sum(bot_L))
                        if np.isfinite(bot_tot) and bot_tot > 1e-300:
                            xB_now = float(bot_L[int(bot_comp_idx)] / bot_tot)
                    except Exception:
                        xB_now = np.nan

                p_top_now = np.nan
                if p_top_pv is not None and np.isfinite(float(p_top_pv)):
                    p_top_now = float(p_top_pv)
                elif "P_top_drum_psia" in diag:
                    p_top_now = _mapping_scalar(diag, "P_top_drum_psia")
                elif "P_psia_hyd" in diag:
                    try:
                        p_h = np.asarray(diag["P_psia_hyd"], dtype=float).reshape((col.n_stages,))
                        p_top_now = float(p_h[0])
                    except Exception:
                        p_top_now = np.nan

                reflux_now = np.nan
                if step_boundary.reflux_lbmolph is not None:
                    reflux_now = float(step_boundary.reflux_lbmolph)
                boilup_now = np.nan
                if step_boundary.boilup_lbmolph is not None:
                    boilup_now = float(step_boundary.boilup_lbmolph)

                ss_hist.append(
                    {
                        "t": float(t_visible_s),
                        "xD": float(xD_now),
                        "xB": float(xB_now),
                        "P_top": float(p_top_now),
                        "reflux": float(reflux_now),
                        "boilup": float(boilup_now),
                    }
                )
                while ss_hist and (float(t_visible_s) - float(ss_hist[0]["t"]) > float(ss_window_sec)):
                    ss_hist.popleft()

                hist_t = [float(h["t"]) for h in ss_hist]
                xD_slope = _linear_trend_slope_per_s(hist_t, [float(h["xD"]) for h in ss_hist])
                xB_slope = _linear_trend_slope_per_s(hist_t, [float(h["xB"]) for h in ss_hist])
                P_top_slope = _linear_trend_slope_per_s(hist_t, [float(h["P_top"]) for h in ss_hist])
                reflux_rate = _linear_trend_slope_per_s(hist_t, [float(h["reflux"]) for h in ss_hist])
                boilup_rate = _linear_trend_slope_per_s(hist_t, [float(h["boilup"]) for h in ss_hist])

                kpi_slopes = []
                for v in (xD_slope, xB_slope):
                    if np.isfinite(v):
                        kpi_slopes.append(abs(float(v)))
                max_kpi_slope = float(np.max(kpi_slopes)) if kpi_slopes else np.nan

                mv_rates = []
                for v in (reflux_rate, boilup_rate):
                    if np.isfinite(v):
                        mv_rates.append(abs(float(v)))
                max_mv_rate = float(np.max(mv_rates)) if mv_rates else np.nan

                sp_errs = []
                if np.isfinite(xD_now) and step_distillate_comp_sp is not None and np.isfinite(float(step_distillate_comp_sp)):
                    sp_errs.append(abs(float(xD_now) - float(step_distillate_comp_sp)))
                if np.isfinite(xB_now) and step_bottoms_comp_sp is not None and np.isfinite(float(step_bottoms_comp_sp)):
                    sp_errs.append(abs(float(xB_now) - float(step_bottoms_comp_sp)))
                max_sp_err = float(np.max(sp_errs)) if sp_errs else np.nan

                ss_active_criteria = 0
                ss_pass = True
                ss_ratios: List[float] = []

                def _apply_criterion(metric: float, tol: Optional[float], *, require_metric: bool = True) -> None:
                    nonlocal ss_active_criteria, ss_pass, ss_ratios
                    if tol is None:
                        return
                    if require_metric and (not np.isfinite(metric)):
                        ss_active_criteria += 1
                        ss_pass = False
                        return
                    if not np.isfinite(metric):
                        return
                    ss_active_criteria += 1
                    ss_ratios.append(float(metric) / float(tol))
                    if float(metric) > float(tol):
                        ss_pass = False

                _apply_criterion(float(max_rel_state_rate), ss_tol_rel, require_metric=True)
                _apply_criterion(float(max_temp_rate), ss_tol_temp, require_metric=True)
                _apply_criterion(float(max_kpi_slope), ss_tol_kpi, require_metric=False)
                _apply_criterion(float(max_mv_rate), ss_tol_mv, require_metric=False)
                if ss_require_sp:
                    _apply_criterion(float(max_sp_err), ss_tol_sp, require_metric=True)

                ss_score = float(np.max(ss_ratios)) if ss_ratios else np.nan
                if float(t_visible_s) < float(ss_min_time_sec):
                    ss_flag = 0.0
                elif ss_active_criteria <= 0:
                    ss_flag = np.nan
                else:
                    ss_flag = 1.0 if ss_pass else 0.0

                diag["steady_state_enabled"] = np.array([1.0], dtype=float)
                diag["steady_state_flag"] = np.array([float(ss_flag)], dtype=float)
                diag["steady_state_score"] = np.array([float(ss_score)], dtype=float)
                diag["steady_state_active_criteria"] = np.array([float(ss_active_criteria)], dtype=float)
                diag["ss_max_rel_state_rate_per_s"] = np.array([float(max_rel_state_rate)], dtype=float)
                diag["ss_rel_state_rate_stage_1based"] = np.array(
                    [float(max_rel_state_detail.get("stage_1based", np.nan))],
                    dtype=float,
                )
                diag["ss_rel_state_rate_component_1based"] = np.array(
                    [float(max_rel_state_detail.get("component_1based", np.nan))],
                    dtype=float,
                )
                diag["ss_rel_state_rate_state_key"] = np.array(
                    [str(max_rel_state_detail.get("state_key", "") or "")],
                    dtype=object,
                )
                diag["ss_max_kpi_slope_per_s"] = np.array([float(max_kpi_slope)], dtype=float)
                diag["ss_max_mv_rate_per_s"] = np.array([float(max_mv_rate)], dtype=float)
                diag["ss_max_temp_rate_F_per_s"] = np.array([float(max_temp_rate)], dtype=float)
                diag["ss_max_sp_error"] = np.array([float(max_sp_err)], dtype=float)
                diag["ss_window_samples"] = np.array([float(len(ss_hist))], dtype=float)
                diag["ss_window_sec"] = np.array([float(ss_window_sec)], dtype=float)
                diag["ss_min_time_sec"] = np.array([float(ss_min_time_sec)], dtype=float)
                diag["ss_tol_rel_state_rate_per_s"] = np.array(
                    [float(ss_tol_rel) if ss_tol_rel is not None else np.nan],
                    dtype=float,
                )
                diag["ss_tol_kpi_slope_per_s"] = np.array(
                    [float(ss_tol_kpi) if ss_tol_kpi is not None else np.nan],
                    dtype=float,
                )
                diag["ss_tol_mv_rate_per_s"] = np.array(
                    [float(ss_tol_mv) if ss_tol_mv is not None else np.nan],
                    dtype=float,
                )
                diag["ss_tol_temp_rate_F_per_s"] = np.array(
                    [float(ss_tol_temp) if ss_tol_temp is not None else np.nan],
                    dtype=float,
                )
                diag["ss_tol_sp_error"] = np.array(
                    [float(ss_tol_sp) if ss_tol_sp is not None else np.nan],
                    dtype=float,
                )
                diag["ss_require_sp"] = np.array([1.0 if bool(ss_require_sp) else 0.0], dtype=float)
                diag["ss_xD_slope_per_s"] = np.array([float(xD_slope)], dtype=float)
                diag["ss_xB_slope_per_s"] = np.array([float(xB_slope)], dtype=float)
                diag["ss_P_top_slope_per_s"] = np.array([float(P_top_slope)], dtype=float)
                diag["ss_reflux_rate_per_s"] = np.array([float(reflux_rate)], dtype=float)
                diag["ss_boilup_rate_per_s"] = np.array([float(boilup_rate)], dtype=float)

                steady_state_status_last = {
                    "steady_state_flag": float(ss_flag),
                    "steady_state_score": float(ss_score),
                    "ss_max_rel_state_rate_per_s": float(max_rel_state_rate),
                    "ss_rel_state_rate_stage_1based": float(max_rel_state_detail.get("stage_1based", np.nan)),
                    "ss_rel_state_rate_component_1based": float(max_rel_state_detail.get("component_1based", np.nan)),
                    "ss_max_kpi_slope_per_s": float(max_kpi_slope),
                    "ss_max_mv_rate_per_s": float(max_mv_rate),
                    "ss_max_temp_rate_F_per_s": float(max_temp_rate),
                    "ss_max_sp_error": float(max_sp_err),
                }
                if np.isfinite(float(max_rel_state_rate)):
                    last_ss_rel_state_rate_per_s = float(max_rel_state_rate)
                ss_prev_t_s = float(t_visible_s)
                ss_prev_y = np.asarray(y, dtype=float).copy()
            else:
                diag["steady_state_enabled"] = np.array([0.0], dtype=float)
                diag["steady_state_flag"] = np.array([np.nan], dtype=float)
                diag["steady_state_score"] = np.array([np.nan], dtype=float)

            # Cache and carry forward thermo diagnostics so intermediate log rows don't show NaNs
            if do_thermo:
                p_prev_for_slope = None
                t_prev_for_slope = None
                try:
                    if last_P_hyd is not None:
                        p_prev_for_slope = np.asarray(last_P_hyd, dtype=float).copy()
                    elif last_P_diag is not None:
                        p_prev_for_slope = np.asarray(last_P_diag, dtype=float).copy()
                except Exception:
                    p_prev_for_slope = None
                try:
                    if last_T_tray is not None:
                        t_prev_for_slope = np.asarray(last_T_tray, dtype=float).copy()
                except Exception:
                    t_prev_for_slope = None
                if "Z_tray" in diag:
                    last_Z_tray = np.asarray(diag["Z_tray"], dtype=float).copy()
                if "y_eq_tray" in diag:
                    last_y_eq = np.asarray(diag["y_eq_tray"], dtype=float).copy()
                if "P_psia_diag" in diag:
                    last_P_diag = np.asarray(diag["P_psia_diag"], dtype=float).copy()
                if "P_psia_hyd" in diag:
                    try:
                        last_P_hyd = np.asarray(diag["P_psia_hyd"], dtype=float).copy()
                    except Exception:
                        pass
                if "V_out_lbmolph" in diag:
                    try:
                        last_V_out = np.asarray(diag["V_out_lbmolph"], dtype=float).copy()
                    except Exception:
                        pass
                if "dT_tray_F_per_s" in diag:
                    try:
                        last_dT_tray = np.asarray(diag["dT_tray_F_per_s"], dtype=float).copy()
                    except Exception:
                        pass
                if "rhoL_tray_lbmol_ft3" in diag:
                    last_rhoL = np.asarray(diag["rhoL_tray_lbmol_ft3"], dtype=float).copy()
                current_T_tray_for_packet = None
                try:
                    current_T_tray_for_packet = _tray_temperature_F(
                        col,
                        layout,
                        y,
                        include_temperature=bool(cfg.include_temperature),
                    ).copy()
                except Exception:
                    current_T_tray_for_packet = None
                current_P_tray_for_packet = None
                try:
                    if last_P_hyd is not None:
                        current_P_tray_for_packet = np.asarray(last_P_hyd, dtype=float).copy()
                    elif last_P_diag is not None:
                        current_P_tray_for_packet = np.asarray(last_P_diag, dtype=float).copy()
                except Exception:
                    current_P_tray_for_packet = None
                current_tray_thermo_packet = _tray_thermo_packet_from_diag(
                    diag,
                    n_stages=col.n_stages,
                    n_components=col.n_components,
                    T_tray_F=current_T_tray_for_packet,
                    P_tray_psia=current_P_tray_for_packet,
                )
                if current_tray_thermo_packet is not None:
                    last_tray_thermo_packet = current_tray_thermo_packet
                    last_K_tray = current_tray_thermo_packet.K_tray.copy()
                    last_HL = current_tray_thermo_packet.HL.copy()
                    last_HV = current_tray_thermo_packet.HV.copy()
                    last_Zfac = current_tray_thermo_packet.Zfac_tray.copy()
                    last_z_overall = current_tray_thermo_packet.z_overall.copy()
                else:
                    if "K_tray" in diag:
                        last_K_tray = np.asarray(diag["K_tray"], dtype=float).copy()
                    if "HL_BTU_lbmol_tray" in diag:
                        last_HL = np.asarray(diag["HL_BTU_lbmol_tray"], dtype=float).copy()
                    if "HV_BTU_lbmol_tray" in diag:
                        last_HV = np.asarray(diag["HV_BTU_lbmol_tray"], dtype=float).copy()
                if "energy_balance_resid_BTUps_tray" in diag:
                    try:
                        last_energy_resid_tray = np.asarray(diag["energy_balance_resid_BTUps_tray"], dtype=float).copy()
                    except Exception:
                        pass
                if "eq_phase_energy_damping_tray" in diag:
                    try:
                        current_energy_damping = np.asarray(diag["eq_phase_energy_damping_tray"], dtype=float).copy()
                        current_energy_damping = np.where(
                            np.isfinite(current_energy_damping),
                            np.clip(current_energy_damping, 0.25, 1.0),
                            1.0,
                        )
                        if last_phase_energy_damping_min is None:
                            last_phase_energy_damping_min = current_energy_damping
                        else:
                            last_phase_energy_damping_min = np.minimum(
                                np.asarray(last_phase_energy_damping_min, dtype=float),
                                current_energy_damping,
                            )
                    except Exception:
                        pass
                if current_tray_thermo_packet is None:
                    if "Z_tray" in diag:
                        last_Zfac = np.asarray(diag["Z_tray"], dtype=float).copy()
                    if "z_overall_tray" in diag:
                        try:
                            last_z_overall = np.asarray(diag["z_overall_tray"], dtype=float).copy()
                        except Exception:
                            pass
                current_condenser_duty_packet = _condenser_duty_packet_from_diag(
                    diag,
                    n_components=col.n_components,
                )
                if current_condenser_duty_packet is not None:
                    last_condenser_duty_packet = current_condenser_duty_packet
                    startup_seeded_condenser_duty_packet = False
                current_bottom_sump_cp_packet = _bottom_sump_cp_packet_from_diag(
                    diag,
                    n_components=col.n_components,
                )
                if current_bottom_sump_cp_packet is not None:
                    last_bottom_sump_cp_packet = current_bottom_sump_cp_packet
                if "reb_T_F" in diag:
                    try:
                        last_reb_T = float(np.asarray(diag["reb_T_F"]).reshape((-1,))[0])
                    except Exception:
                        pass
                if "reb_beta" in diag:
                    try:
                        last_reb_beta = float(np.asarray(diag["reb_beta"]).reshape((-1,))[0])
                    except Exception:
                        pass
                if "reb_x" in diag:
                    try:
                        last_reb_x = np.asarray(diag["reb_x"], dtype=float).reshape((col.n_components,)).copy()
                    except Exception:
                        pass
                if "reb_y" in diag:
                    try:
                        last_reb_y = np.asarray(diag["reb_y"], dtype=float).reshape((col.n_components,)).copy()
                    except Exception:
                        pass
                if current_T_tray_for_packet is not None:
                    last_T_tray = current_T_tray_for_packet
                if (
                    str(base_inputs.pressure_model).strip().lower() == "hydraulic"
                    and str(base_inputs.vapor_flow_model).strip().lower() == "energy"
                    and str(base_inputs.hydraulic_energy_temperature_mode).strip().lower()
                    == "pressure-correction-follower"
                    and bool(cfg.include_temperature)
                    and last_T_tray is not None
                ):
                    try:
                        p_curr_for_slope = (
                            np.asarray(last_P_hyd, dtype=float).copy()
                            if last_P_hyd is not None
                            else (
                                np.asarray(last_P_diag, dtype=float).copy()
                                if last_P_diag is not None
                                else None
                            )
                        )
                    except Exception:
                        p_curr_for_slope = None
                    try:
                        last_tray_temp_pressure_slope = _update_tray_temp_pressure_slope_F_per_psi(
                            prev_slope_F_per_psi=last_tray_temp_pressure_slope,
                            prev_T_F=t_prev_for_slope,
                            curr_T_F=last_T_tray,
                            prev_P_psia=p_prev_for_slope,
                            curr_P_psia=p_curr_for_slope,
                            default_slope_F_per_psi=float(
                                base_inputs.hydraulic_energy_temperature_pressure_slope_F_per_psi
                            ),
                        )
                    except Exception:
                        pass
                if (
                    str(base_inputs.pressure_model).strip().lower() == "hydraulic"
                    and str(base_inputs.vapor_flow_model).strip().lower() == "energy"
                    and str(base_inputs.hydraulic_energy_temperature_mode).strip().lower()
                    == "bubble-point-follower"
                    and base_inputs.thermo_provider is not None
                    and bool(cfg.include_temperature)
                ):
                    refresh_steps = int(
                        getattr(cfg, "hydraulic_energy_temperature_target_refresh_steps", None) or 20
                    )
                    if refresh_steps <= 0:
                        refresh_steps = 1
                    should_refresh_targets = (
                        last_tray_bubble_target_F is None
                        or (step <= 1)
                        or ((step % refresh_steps) == 0)
                    )
                    if should_refresh_targets:
                        try:
                            p_target = (
                                np.asarray(last_P_hyd, dtype=float).copy()
                                if last_P_hyd is not None
                                else (
                                    np.asarray(last_P_diag, dtype=float).copy()
                                    if last_P_diag is not None
                                    else np.asarray(getattr(col, "P_psia", np.full(col.n_stages, 200.0, dtype=float)), dtype=float).reshape((col.n_stages,))
                                )
                            )
                            refreshed_targets = _refresh_tray_bubble_targets_F(
                                col=col,
                                layout=layout,
                                y=y,
                                thermo_provider=base_inputs.thermo_provider,
                                P_tray_psia=p_target,
                            )
                            if refreshed_targets is not None and np.any(np.isfinite(refreshed_targets)):
                                last_tray_bubble_target_F = np.asarray(refreshed_targets, dtype=float).copy()
                        except Exception:
                            pass
                last_diag = diag
            else:
                if last_Z_tray is not None:
                    diag["Z_tray"] = last_Z_tray
                if last_y_eq is not None:
                    diag["y_eq_tray"] = last_y_eq
                if last_P_diag is not None:
                    diag["P_psia_diag"] = last_P_diag
                if last_P_hyd is not None:
                    diag["P_psia_hyd"] = last_P_hyd
                if last_rhoL is not None:
                    diag["rhoL_tray_lbmol_ft3"] = last_rhoL
                if last_energy_resid_tray is not None:
                    diag["energy_balance_resid_BTUps_tray"] = last_energy_resid_tray

            # Log / print at cadence
            if visible_logging_phase and ((visible_step % log_every) == 0):
                if not (initial_snapshot_written and visible_step == 0):
                    failure_phase = "logging"
                    wall_elapsed_s = time.perf_counter() - start_perf
                    wall_clock_iso = _dt.datetime.now().isoformat(timespec="seconds")
                    sim_per_wall = (t_visible_s / wall_elapsed_s) if wall_elapsed_s > 1e-12 else float("inf")

                    progress_msg = (
                        f"[Progress] step={visible_step:6d}  sim_t={t_visible_s:10.2f} s  wall={wall_elapsed_s:10.2f} s  "
                        f"sim/wall={sim_per_wall:8.3f}"
                    )
                    if ss_enabled:
                        ss_flag = _mapping_scalar(diag, "steady_state_flag")
                        ss_score = _mapping_scalar(diag, "steady_state_score")
                        ss_rel = _mapping_scalar(diag, "ss_max_rel_state_rate_per_s")
                        ss_kpi = _mapping_scalar(diag, "ss_max_kpi_slope_per_s")
                        ss_mv = _mapping_scalar(diag, "ss_max_mv_rate_per_s")
                        ss_sp = _mapping_scalar(diag, "ss_max_sp_error")
                        if np.isfinite(ss_flag):
                            ss_flag_txt = str(int(round(float(ss_flag))))
                        else:
                            ss_flag_txt = "?"
                        progress_msg += f"  SS={ss_flag_txt}"
                        if np.isfinite(ss_score):
                            progress_msg += f"  score={float(ss_score):.2f}"
                        if np.isfinite(ss_rel):
                            progress_msg += f"  rel={float(ss_rel):.3g}/s"
                        if np.isfinite(ss_kpi):
                            progress_msg += f"  kpi={float(ss_kpi):.3g}/s"
                        if np.isfinite(ss_mv):
                            progress_msg += f"  mv={float(ss_mv):.3g}/s"
                        if np.isfinite(ss_sp):
                            progress_msg += f"  sp={float(ss_sp):.3g}"

                    _emit_progress(progress_msg)

                    if cfg.write_logs:
                        _write_log_snapshot(
                            t_snapshot_s=t_visible_s,
                            col_snapshot=step_col,
                            y_snapshot=y,
                            diag_snapshot=diag,
                            feed_tag_snapshot=step_feed_tag,
                            dist_tag_snapshot=step_dist_tag,
                            bots_tag_snapshot=step_bots_tag,
                            wall_clock_iso=wall_clock_iso,
                            wall_elapsed_s=wall_elapsed_s,
                        )

            control_doc = _read_json_if_exists(runtime_control_path)
            try:
                requested_visible_steps = int(control_doc.get("requested_total_steps", cfg.n_steps))
            except Exception:
                requested_visible_steps = int(cfg.n_steps)
            requested_total_steps = int(max(requested_visible_steps, int(cfg.n_steps))) + int(restart_hidden_warmup_steps)
            if requested_total_steps > int(total_steps):
                total_steps = int(requested_total_steps)
                if cfg.write_logs:
                    metadata_doc["n_steps"] = int(requested_visible_steps)
                    _write_json_atomic(metadata_path, metadata_doc)
                print(
                    "[Control] Extended simulation horizon  "
                    f"n_steps={int(requested_visible_steps)}  "
                    f"sim_time={float(requested_visible_steps) * float(dt):.2f} s"
                )

            if step == int(total_steps):
                break

            y_before_step = np.asarray(y, dtype=float).copy()
            step_wall_t0 = time.perf_counter()
            if runtime_trace_active:
                _runtime_trace(
                    "step="
                    f"{int(visible_step)} integrate start "
                    f"mode={str(integrator_mode).strip()} "
                    f"t={float(t_visible_s):.2f}s"
                )
            if str(integrator_mode).strip().lower() == "ida":
                y, step_integrator_info = _integrate_one_step_ida(
                    t_s=float(t_s),
                    y=y,
                    dt_sec=float(dt),
                    rhs_eval=_eval_step_rhs,
                    layout=layout,
                    thermo_provider=thermo_provider,
                    substep_sec=getattr(cfg, "integrator_substep_sec", None),
                    max_iter=int(ida_max_iter_eff),
                    relax=getattr(cfg, "ida_relax", None),
                    rtol=getattr(cfg, "integrator_rtol", None),
                    atol=getattr(cfg, "integrator_atol", None),
                    max_rhs_evals_per_step=getattr(cfg, "integrator_max_rhs_evals_per_step", None),
                    step_wall_limit_sec=getattr(cfg, "integrator_step_wall_limit_sec", None),
                    alg_p_tol_psia=dae_pilot_p_tol_eff,
                    alg_v_tol_lbmolph=dae_pilot_v_tol_eff,
                )
            elif str(integrator_mode).strip().lower() == "explicit-euler":
                y, step_integrator_info = _advance_explicit_euler_step(
                    y=y,
                    dydt=dydt,
                    dt_sec=float(dt),
                    layout=layout,
                    thermo_provider=thermo_provider,
                    requested_mode=str(integrator_mode).strip(),
                )
            else:
                y, step_integrator_info = _integrate_one_step(
                    t_s=float(t_s),
                    y=y,
                    dt_sec=float(dt),
                    rhs_eval=_eval_step_rhs_integrator,
                    rhs_eval_fallback=_eval_step_rhs,
                    layout=layout,
                    thermo_provider=thermo_provider,
                    integrator_mode=integrator_mode,
                    rtol=getattr(cfg, "integrator_rtol", None),
                    atol=getattr(cfg, "integrator_atol", None),
                    max_step_sec=getattr(cfg, "integrator_max_step_sec", None),
                    substep_sec=getattr(cfg, "integrator_substep_sec", None),
                    max_rhs_evals_per_step=getattr(cfg, "integrator_max_rhs_evals_per_step", None),
                    step_wall_limit_sec=getattr(cfg, "integrator_step_wall_limit_sec", None),
                )
            if bool(step_integrator_info.get("fallback_used", False)) and bool(dae_outer_once_for_stiff):
                # If stiff solve falls back, preserve DAE-outer-step consistency by
                # advancing with the outer DAE-consistent derivative.
                y = y_before_step + float(dt) * np.asarray(dydt, dtype=float).reshape((-1,))
                y = _clamp_nonnegative_holdups(y, layout)
                y = _clip_temperature_states_to_provider_bounds(y, layout, thermo_provider)
                step_integrator_info["fallback_reason"] = (
                    str(step_integrator_info.get("fallback_reason", "")).strip()
                    + " | advanced with outer DAE dydt"
                ).strip(" |")

            last_step_integrator_info = dict(step_integrator_info) if isinstance(step_integrator_info, dict) else {}
            if bool(step_integrator_info.get("fallback_used", False)):
                integrator_fallback_count += 1
                if integrator_fallback_count <= 3 or (integrator_fallback_count % 20) == 0:
                    msg = str(step_integrator_info.get("fallback_reason", "")).strip()
                    if not msg:
                        msg = "unknown reason"
                    print(
                        "[Warn] Integrator fallback "
                        f"({integrator_mode}->explicit-euler) at step={int(visible_step)} "
                        f"t={float(t_visible_s):.2f}s: {msg}"
                    )
            step_wall_elapsed_s = float(time.perf_counter() - step_wall_t0)
            if runtime_trace_active or step_wall_elapsed_s >= 5.0:
                _runtime_trace(
                    "step="
                    f"{int(visible_step)} integrate done "
                    f"wall={step_wall_elapsed_s:.2f}s "
                    f"fallback={int(bool(step_integrator_info.get('fallback_used', False)))} "
                    f"nfev={float(step_integrator_info.get('nfev', np.nan)) if np.isfinite(float(step_integrator_info.get('nfev', np.nan))) else float('nan'):.0f}"
                )
            t_s += dt
            step += 1

    except Exception as exc:
        run_status = "failed"
        run_error_traceback = traceback.format_exc()
        run_error = (
            f"{exc} "
            f"[phase={failure_phase} step={failure_step} t={float(failure_time_s):.6g}s]"
        )
        print(f"[Error] {run_error}")
        if run_error_traceback:
            print(run_error_traceback)
        raise
    finally:
        if profile_file is not None:
            profile_file.close()
        if summary_file is not None:
            summary_file.close()
        if cfg.write_logs:
            end_wall_dt = _dt.datetime.now()
            thermo_call_counters = _snapshot_thermo_call_counters(
                thermo_provider,
                getattr(base_inputs, "thermo_provider", None),
                getattr(base_inputs, "equilibrium_relaxation_thermo_provider", None),
            )
            metadata_doc.update(
                {
                    "status": run_status,
                    "ended_at_local": end_wall_dt.strftime("%Y-%m-%d %H:%M:%S"),
                    "elapsed_wall_sec": float(time.perf_counter() - start_perf),
                    "final_time_s": float(max(float(t_s), 0.0)),
                    "profile_csv": str(profile_path),
                    "summary_csv": str(summary_path),
                    "error": run_error or "",
                    "error_traceback": run_error_traceback,
                    "startup_timing_sec": dict(startup_timing_sec),
                    "startup_seed_cache": dict(startup_seed_cache_info),
                    "thermo_call_counters": thermo_call_counters,
                }
            )
            _write_json_atomic(metadata_path, metadata_doc)

    if ss_enabled:
        ssf = float(steady_state_status_last.get("steady_state_flag", np.nan))
        verdict = "UNKNOWN"
        if np.isfinite(ssf):
            verdict = "PASS" if int(round(ssf)) == 1 else "FAIL"
        sscore = float(steady_state_status_last.get("steady_state_score", np.nan))
        srel = float(steady_state_status_last.get("ss_max_rel_state_rate_per_s", np.nan))
        skpi = float(steady_state_status_last.get("ss_max_kpi_slope_per_s", np.nan))
        smv = float(steady_state_status_last.get("ss_max_mv_rate_per_s", np.nan))
        ssp = float(steady_state_status_last.get("ss_max_sp_error", np.nan))
        stemp = float(steady_state_status_last.get("ss_max_temp_rate_F_per_s", np.nan))
        print(
            "[SteadyState] "
            f"{verdict}  score={sscore:.3g}  rel={srel:.3g}/s  "
            f"kpi={skpi:.3g}/s  mv={smv:.3g}/s  dT={stemp:.3g}F/s  sp={ssp:.3g}"
        )

    controller_state_final: Dict[str, float] = {}
    if top_level_ctrl is not None and np.isfinite(float(top_level_ctrl.integ)):
        controller_state_final["top_level_integ"] = float(top_level_ctrl.integ)
    if bot_level_ctrl is not None and np.isfinite(float(bot_level_ctrl.integ)):
        controller_state_final["bottom_level_integ"] = float(bot_level_ctrl.integ)
    if top_pressure_ctrl is not None and np.isfinite(float(top_pressure_ctrl.integ)):
        controller_state_final["top_pressure_integ"] = float(top_pressure_ctrl.integ)
    if dist_comp_ctrl is not None and np.isfinite(float(dist_comp_ctrl.integ)):
        controller_state_final["distillate_comp_integ"] = float(dist_comp_ctrl.integ)
    if bot_comp_ctrl is not None and np.isfinite(float(bot_comp_ctrl.integ)):
        controller_state_final["bottoms_comp_integ"] = float(bot_comp_ctrl.integ)
    if last_top_pressure_pv_filt_psia is not None and np.isfinite(float(last_top_pressure_pv_filt_psia)):
        controller_state_final["top_pressure_pv_filt_psia"] = float(last_top_pressure_pv_filt_psia)
    if last_pressure_mv_cmd is not None and np.isfinite(float(last_pressure_mv_cmd)):
        controller_state_final["top_pressure_mv_cmd_btuph"] = float(last_pressure_mv_cmd)
    if last_top_energy_resid_abs_btups is not None and np.isfinite(float(last_top_energy_resid_abs_btups)):
        controller_state_final["top_pressure_resid_abs_btups"] = float(last_top_energy_resid_abs_btups)
    if last_top_drum_pressure_T is not None and np.isfinite(float(last_top_drum_pressure_T)):
        controller_state_final["top_drum_pressure_T_prev_F"] = float(last_top_drum_pressure_T)
    try:
        if step_dist_tag.flow_lbmolph is not None and np.isfinite(float(step_dist_tag.flow_lbmolph)):
            controller_state_final["distillate_cmd_lbmolph"] = float(step_dist_tag.flow_lbmolph)
    except Exception:
        pass
    try:
        if step_bots_tag.flow_lbmolph is not None and np.isfinite(float(step_bots_tag.flow_lbmolph)):
            controller_state_final["bottoms_cmd_lbmolph"] = float(step_bots_tag.flow_lbmolph)
    except Exception:
        pass
    try:
        if step_reflux_cmd_lbmolph is not None and np.isfinite(float(step_reflux_cmd_lbmolph)):
            controller_state_final["reflux_cmd_lbmolph"] = float(step_reflux_cmd_lbmolph)
    except Exception:
        pass
    try:
        if step_boilup_cmd_lbmolph is not None and np.isfinite(float(step_boilup_cmd_lbmolph)):
            controller_state_final["boilup_cmd_lbmolph"] = float(step_boilup_cmd_lbmolph)
    except Exception:
        pass

    restart_workbook_path: Optional[str] = None
    restart_export_error = ""
    try:
        _ensure_dir(logs_dir)
        restart_stem = Path(str(cfg.excel_path)).stem
        if "__restart_" in restart_stem:
            restart_stem = restart_stem.split("__restart_", 1)[0]
        max_stem_len = 80
        if len(restart_stem) > max_stem_len:
            restart_stem = restart_stem[:max_stem_len].rstrip("._- ")
        restart_name = f"{restart_stem}__restart_{tag}.xlsx"
        restart_path = logs_dir / restart_name
        restart_workbook_path = write_restart_workbook_from_run_result(
            run_result={
                "run_id": str(tag),
                "excel_path": str(Path(cfg.excel_path).resolve()),
                "final_state": y,
                "layout": layout,
                "column": col,
                "last_diag": last_diag,
                "controller_state_final": controller_state_final,
            },
            output_excel_path=str(restart_path),
        )
        print(f"[Output] Wrote restart workbook: {restart_workbook_path}")
    except Exception as exc:
        restart_export_error = str(exc)
        print(f"[Warn] Failed to write restart workbook: {exc}")

    if (
        bool(getattr(cfg, "enable_startup_seed_cache", False))
        and (not explicit_runtime_restart)
        and startup_seed_cache_path is not None
    ):
        try:
            feed_seed_pressure = (
                np.asarray(last_P_hyd, dtype=float).copy()
                if last_P_hyd is not None
                else (
                    np.asarray(last_P_diag, dtype=float).copy()
                    if last_P_diag is not None
                    else np.asarray(getattr(col, "P_psia", np.full(col.n_stages, 200.0, dtype=float)), dtype=float).copy()
                )
            )
            refreshed_feed_stage_flash_packet = last_feed_stage_flash_packet
            feed_seed_state = _startup_feed_stage_flash_state(
                col=col,
                P_tray_psia=feed_seed_pressure,
            )
            if feed_seed_state is not None:
                feed_stage0, feed_z, feed_T, feed_P, _feed_Ft = feed_seed_state
                compatible_feed_packet, _dT_feed, _dP_feed, _dx_feed = _compatible_feed_stage_flash_packet(
                    packet=last_feed_stage_flash_packet,
                    stage0=int(feed_stage0),
                    T_feed_F=float(feed_T),
                    P_feed_psia=float(feed_P),
                    z_feed=feed_z,
                    n_components=int(col.n_components),
                    max_abs_dT_F=float(getattr(base_inputs, "feed_stage_flash_reuse_dT_F", 0.5) or 0.5),
                    max_abs_dP_psia=float(getattr(base_inputs, "feed_stage_flash_reuse_dP_psia", 2.5) or 2.5),
                    max_abs_dx=float(getattr(base_inputs, "feed_stage_flash_reuse_dx", 1.0e-6) or 1.0e-6),
                )
                if (
                    compatible_feed_packet is None
                    or compatible_feed_packet.hL_BTU_lbmol is None
                    or compatible_feed_packet.hV_BTU_lbmol is None
                ):
                    refreshed_feed_stage_flash_packet = _seed_startup_feed_stage_flash_packet(
                        col=col,
                        thermo_provider=base_inputs.thermo_provider,
                        P_tray_psia=feed_seed_pressure,
                    )
                else:
                    refreshed_feed_stage_flash_packet = compatible_feed_packet
            if refreshed_feed_stage_flash_packet is not None:
                last_feed_stage_flash_packet = refreshed_feed_stage_flash_packet
        except Exception:
            pass
        startup_seed_refresh_t0 = time.perf_counter()
        startup_seed_cache_refresh = _write_startup_seed_cache(
            path=startup_seed_cache_path,
            cfg=cfg,
            col=col,
            layout=layout,
            base_inputs=base_inputs,
            y=y,
            last_T_tray=last_T_tray,
            last_P_diag=last_P_diag,
            last_P_hyd=last_P_hyd,
            last_K_tray=last_K_tray,
            last_HL=last_HL,
            last_HV=last_HV,
            last_Zfac=last_Zfac,
            last_z_overall=last_z_overall,
            last_tray_bubble_target_F=last_tray_bubble_target_F,
            last_tray_thermo_packet=last_tray_thermo_packet,
            last_condenser_duty_packet=last_condenser_duty_packet,
            last_feed_stage_flash_packet=last_feed_stage_flash_packet,
            last_bottom_sump_cp_packet=last_bottom_sump_cp_packet,
            last_reb_T=last_reb_T,
            last_reb_x=last_reb_x,
            last_reb_y=last_reb_y,
            last_reb_beta=last_reb_beta,
            startup_seeded_condenser_duty_packet=startup_seeded_condenser_duty_packet,
        )
        startup_timing_sec["startup_seed_cache_refresh"] = float(time.perf_counter() - startup_seed_refresh_t0)
        startup_seed_cache_info["saved"] = bool(
            startup_seed_cache_info.get("saved", False) or startup_seed_cache_refresh.get("saved", False)
        )
        if bool(startup_seed_cache_refresh.get("saved", False)):
            prior_reason = str(startup_seed_cache_info.get("reason", "") or "").strip()
            startup_seed_cache_info["reason"] = (
                "loaded+saved"
                if prior_reason == "loaded"
                else str(startup_seed_cache_refresh.get("reason", prior_reason or "saved"))
            )
        elif startup_seed_cache_refresh.get("reason"):
            startup_seed_cache_info["refresh_reason"] = str(startup_seed_cache_refresh.get("reason"))

    run_status = "completed"
    if cfg.write_logs:
        end_wall_dt = _dt.datetime.now()
        thermo_call_counters = _snapshot_thermo_call_counters(
            thermo_provider,
            getattr(base_inputs, "thermo_provider", None),
            getattr(base_inputs, "equilibrium_relaxation_thermo_provider", None),
        )
        metadata_doc.update(
            {
                "status": run_status,
                "ended_at_local": end_wall_dt.strftime("%Y-%m-%d %H:%M:%S"),
                "elapsed_wall_sec": float(time.perf_counter() - start_perf),
                "final_time_s": float(t_s),
                "profile_csv": str(profile_path),
                "summary_csv": str(summary_path),
                "restart_workbook": str(restart_workbook_path or ""),
                "restart_export_error": restart_export_error,
                "error": "",
                "startup_timing_sec": dict(startup_timing_sec),
                "init_pack_top_drum_vapor": dict(init_pack_top_drum_vapor_info),
                "init_match_condenser_duty": dict(init_match_condenser_duty_info),
                "startup_seed_cache": dict(startup_seed_cache_info),
                "thermo_call_counters": thermo_call_counters,
            }
        )
        _write_json_atomic(metadata_path, metadata_doc)

    return {
        "run_id": str(tag),
        "run_name": run_name,
        "run_description": run_description,
        "excel_path": str(Path(cfg.excel_path).resolve()),
        "logs_dir": str(logs_dir),
        "run_metadata_json": str(metadata_path) if cfg.write_logs else None,
        "startup_trace_log": str(startup_trace_path),
        "thermo_call_counters": _snapshot_thermo_call_counters(
            thermo_provider,
            getattr(base_inputs, "thermo_provider", None),
            getattr(base_inputs, "equilibrium_relaxation_thermo_provider", None),
        ),
        "runtime_mode": str(runtime_mode),
        "integrator_mode": str(integrator_mode),
        "effective_ida_max_iter": int(ida_max_iter_eff),
        "effective_dae_pilot_enabled": bool(dae_pilot_enabled_eff),
        "effective_dae_pilot_max_iter": int(dae_pilot_max_iter_eff),
        "effective_dae_pilot_p_tol_psia": (
            float(_as_float(dae_pilot_p_tol_eff))
            if _as_float(dae_pilot_p_tol_eff) is not None
            else np.nan
        ),
        "effective_dae_pilot_v_tol_lbmolph": (
            float(_as_float(dae_pilot_v_tol_eff))
            if _as_float(dae_pilot_v_tol_eff) is not None
            else np.nan
        ),
        "effective_hydraulic_ida_defaults_applied": list(ida_defaults_applied),
        "integrator_fallback_count": int(integrator_fallback_count),
        "profile_csv": str(profile_path) if cfg.write_logs else None,
        "summary_csv": str(summary_path) if cfg.write_logs else None,
        "restart_workbook": str(restart_workbook_path) if restart_workbook_path else None,
        "validation": {
            "ok": bool(validation.ok),
            "errors": list(validation.errors),
            "warnings": list(validation.warnings),
        },
        "final_time_s": float(max(float(t_s), 0.0)),
        "final_state": y,
        "thermo_provider": thermo_provider,
        "layout": layout,
        "inputs": base_inputs,
        "column": col,
        "last_diag": last_diag,
        "controller_state_final": controller_state_final,
        "steady_state_status_final": dict(steady_state_status_last),
        "startup_thermo_init_info": thermo_init_info,
        "startup_hydraulic_energy_init_info": hydraulic_energy_init_info,
        "startup_top_drum_init_info": top_drum_init_info,
        "startup_seed_cache_info": dict(startup_seed_cache_info),
    }


def write_restart_workbook_from_run_result(
    *,
    run_result: Dict[str, Any],
    output_excel_path: str,
    template_excel_path: Optional[str] = None,
) -> str:
    """Write a restart workbook from a completed run result.

    This persists:
    - Initial Conditions tray profile/state
    - Boundary State sheet for top/bottom liquid and vapor holdups
    - Energy State sheet for tray EL/EV
    - Controller State sheet for controller integrals / filtered PV memory
    """
    from openpyxl import load_workbook

    if not isinstance(run_result, dict):
        raise ValueError("run_result must be a dict returned by run_smoke_simulation().")

    src_excel = template_excel_path or run_result.get("excel_path")
    if not src_excel:
        raise ValueError("template_excel_path or run_result['excel_path'] is required.")

    y = np.asarray(run_result.get("final_state"), dtype=float).reshape((-1,))
    layout = run_result.get("layout")
    col = run_result.get("column")
    last_diag = run_result.get("last_diag") or {}
    if layout is None or col is None or y.size <= 0:
        raise ValueError("run_result is missing final_state/layout/column.")

    u = layout.unpack(y)
    N = int(col.n_stages)
    Nc = int(col.n_components)

    x = np.asarray(last_diag["x_tray"] if "x_tray" in last_diag else u["x_tray"], dtype=float).reshape((N, Nc))
    if "y_tray" in last_diag:
        yv = np.asarray(last_diag["y_tray"], dtype=float).reshape((N, Nc))
    elif "y_eq_thermo_tray" in last_diag:
        yv = np.asarray(last_diag["y_eq_thermo_tray"], dtype=float).reshape((N, Nc))
    elif "y_tray" in u:
        yv = np.asarray(u["y_tray"], dtype=float).reshape((N, Nc))
    else:
        yv = np.asarray(getattr(col, "y0"), dtype=float).reshape((N, Nc))
    T = np.asarray(u["tray_T_f"], dtype=float).reshape((N,))
    P = np.asarray(last_diag.get("P_psia_hyd", getattr(col, "P_psia", np.full(N, np.nan, dtype=float))), dtype=float).reshape((N,))
    ML = np.asarray(u["ML_tot_tray"], dtype=float).reshape((N,))
    MV = (
        np.asarray(u["MV_tot_tray"], dtype=float).reshape((N,))
        if "MV_tot_tray" in u
        else np.zeros(N, dtype=float)
    )
    L_out = np.asarray(last_diag.get("L_out_used_lbmolph", getattr(col, "L_lbmolph", np.full(N, np.nan, dtype=float))), dtype=float).reshape((N,))
    if "V_out_lbmolph" in last_diag:
        V_out = np.asarray(last_diag["V_out_lbmolph"], dtype=float).reshape((N,))
    elif "vflow_energy_used_lbmolph" in last_diag:
        V_out = np.asarray(last_diag["vflow_energy_used_lbmolph"], dtype=float).reshape((N,))
    else:
        V_out = np.asarray(getattr(col, "V_lbmolph", np.full(N, np.nan, dtype=float)), dtype=float).reshape((N,))

    wb = load_workbook(str(src_excel))
    ws = wb["Initial Conditions"]
    headers = [c.value for c in ws[1]]
    col_idx = {str(name): i + 1 for i, name in enumerate(headers) if name is not None}

    required_cols = [
        "Stage",
        "Temperature (F)",
        "Pressure (psia)",
        "Vapor Flow (lbmol/h)",
        "Liquid Flow (lbmol/h)",
        "Liquid Holdup (lbmol)",
    ]
    for name in required_cols:
        if name not in col_idx:
            raise ValueError(f"Initial Conditions missing required column '{name}'.")

    # Ensure stage-1 vapor composition is always normalized for validator purposes.
    yv_write = np.asarray(yv, dtype=float).copy()
    for i in range(N):
        row_sum = float(np.nansum(yv_write[i, :]))
        if np.isfinite(row_sum) and row_sum > 1.0e-12:
            yv_write[i, :] = yv_write[i, :] / row_sum
            continue
        src = min(max(i + 1, 1), N - 1) if N > 1 else i
        fallback = np.asarray(yv_write[src, :], dtype=float)
        fsum = float(np.nansum(fallback))
        if np.isfinite(fsum) and fsum > 1.0e-12:
            yv_write[i, :] = fallback / fsum
        else:
            yv_write[i, :] = np.ones((Nc,), dtype=float) / float(Nc)

    for i in range(N):
        r = i + 2
        ws.cell(r, col_idx["Stage"]).value = i + 1
        ws.cell(r, col_idx["Temperature (F)"]).value = float(T[i]) if np.isfinite(T[i]) else None
        ws.cell(r, col_idx["Pressure (psia)"]).value = float(P[i]) if np.isfinite(P[i]) else None
        ws.cell(r, col_idx["Vapor Flow (lbmol/h)"]).value = float(V_out[i]) if np.isfinite(V_out[i]) else 0.0
        ws.cell(r, col_idx["Liquid Flow (lbmol/h)"]).value = float(L_out[i]) if np.isfinite(L_out[i]) else 0.0
        ws.cell(r, col_idx["Liquid Holdup (lbmol)"]).value = float(ML[i]) if np.isfinite(ML[i]) else 0.0
        if "Vapor Holdup (lbmol)" in col_idx:
            ws.cell(r, col_idx["Vapor Holdup (lbmol)"]).value = float(MV[i]) if np.isfinite(MV[i]) else 0.0
        for k in range(Nc):
            vc = f"Vapor Composition Component {k+1}"
            lc = f"Liquid Composition Component {k+1}"
            if vc in col_idx:
                ws.cell(r, col_idx[vc]).value = float(yv_write[i, k])
            if lc in col_idx:
                ws.cell(r, col_idx[lc]).value = float(x[i, k])

    if "Boundary State" in wb.sheetnames:
        del wb["Boundary State"]
    wsb = wb.create_sheet("Boundary State")
    wsb.cell(1, 1).value = "State"
    comp_labels = list(getattr(col, "components_excel", [f"Component {i+1}" for i in range(Nc)]))
    for k in range(Nc):
        wsb.cell(1, k + 2).value = f"Component {k+1} ({comp_labels[k]})"

    boundary_rows = [
        ("top_L", np.asarray(u.get("top_L", np.zeros((Nc,), dtype=float)), dtype=float).reshape((Nc,))),
        ("top_V", np.asarray(u.get("top_V", np.zeros((Nc,), dtype=float)), dtype=float).reshape((Nc,))),
        ("bottom_L", np.asarray(u.get("bottom_L", np.zeros((Nc,), dtype=float)), dtype=float).reshape((Nc,))),
        ("bottom_V", np.asarray(u.get("bottom_V", np.zeros((Nc,), dtype=float)), dtype=float).reshape((Nc,))),
    ]
    for r, (label, vals) in enumerate(boundary_rows, start=2):
        wsb.cell(r, 1).value = label
        for k in range(Nc):
            wsb.cell(r, k + 2).value = float(vals[k]) if np.isfinite(vals[k]) else 0.0

    if "Energy State" in wb.sheetnames:
        del wb["Energy State"]
    wse = wb.create_sheet("Energy State")
    wse.cell(1, 1).value = "Stage"
    wse.cell(1, 2).value = "Tray EL (BTU)"
    wse.cell(1, 3).value = "Tray EV (BTU)"
    EL = np.asarray(u.get("tray_EL_BTU", np.full((N,), np.nan, dtype=float)), dtype=float).reshape((N,))
    EV = np.asarray(u.get("tray_EV_BTU", np.full((N,), np.nan, dtype=float)), dtype=float).reshape((N,))
    for i in range(N):
        wse.cell(i + 2, 1).value = i + 1
        wse.cell(i + 2, 2).value = float(EL[i]) if np.isfinite(EL[i]) else 0.0
        wse.cell(i + 2, 3).value = float(EV[i]) if np.isfinite(EV[i]) else 0.0

    if "Dynamic Memory" in wb.sheetnames:
        del wb["Dynamic Memory"]
    wsm = wb.create_sheet("Dynamic Memory")
    wsm.cell(1, 1).value = "Stage"
    wsm.cell(1, 2).value = "Prev Tray Pressure (psia)"
    wsm.cell(1, 3).value = "Prev Tray Temperature (F)"
    p_prev = np.asarray(
        last_diag.get("P_psia_hyd", last_diag.get("P_psia_diag", getattr(col, "P_psia", np.full(N, np.nan, dtype=float)))),
        dtype=float,
    ).reshape((N,))
    t_prev = np.asarray(T, dtype=float).reshape((N,))
    for i in range(N):
        wsm.cell(i + 2, 1).value = i + 1
        wsm.cell(i + 2, 2).value = float(p_prev[i]) if np.isfinite(p_prev[i]) else None
        wsm.cell(i + 2, 3).value = float(t_prev[i]) if np.isfinite(t_prev[i]) else None

    if "Controller State" in wb.sheetnames:
        del wb["Controller State"]
    wsc = wb.create_sheet("Controller State")
    wsc.cell(1, 1).value = "Controller"
    wsc.cell(1, 2).value = "Value"
    ctrl_state = dict(run_result.get("controller_state_final") or {})
    ctrl_rows = [
        ("top_level_integ", ctrl_state.get("top_level_integ")),
        ("bottom_level_integ", ctrl_state.get("bottom_level_integ")),
        ("top_pressure_integ", ctrl_state.get("top_pressure_integ")),
        ("top_pressure_pv_filt_psia", ctrl_state.get("top_pressure_pv_filt_psia")),
        ("top_pressure_mv_cmd_btuph", ctrl_state.get("top_pressure_mv_cmd_btuph")),
        ("top_pressure_resid_abs_btups", ctrl_state.get("top_pressure_resid_abs_btups")),
        ("top_drum_pressure_T_prev_F", ctrl_state.get("top_drum_pressure_T_prev_F")),
        ("distillate_cmd_lbmolph", ctrl_state.get("distillate_cmd_lbmolph")),
        ("bottoms_cmd_lbmolph", ctrl_state.get("bottoms_cmd_lbmolph")),
        ("reflux_cmd_lbmolph", ctrl_state.get("reflux_cmd_lbmolph")),
        ("boilup_cmd_lbmolph", ctrl_state.get("boilup_cmd_lbmolph")),
        ("distillate_comp_integ", ctrl_state.get("distillate_comp_integ")),
        ("bottoms_comp_integ", ctrl_state.get("bottoms_comp_integ")),
    ]
    rr = 2
    for label, val in ctrl_rows:
        if val is None or (not np.isfinite(float(val))):
            continue
        wsc.cell(rr, 1).value = label
        wsc.cell(rr, 2).value = float(val)
        rr += 1

    out_path = Path(output_excel_path).expanduser().resolve()
    wb.save(out_path)
    return str(out_path)


# -------------------------
# CLI
# -------------------------


def main(argv: Optional[List[str]] = None) -> int:
    import argparse

    _configure_progress_streams()

    p = argparse.ArgumentParser(description="Dynamic distillation smoke-test runner")

    p.add_argument("--excel", dest="excel_path", default="distillation_column_template.xlsx")
    p.add_argument("--run-name", dest="run_name", default=None, help="Optional human-friendly run label.")
    p.add_argument(
        "--run-description",
        dest="run_description",
        default=None,
        help="Optional longer run description stored with run metadata.",
    )
    p.add_argument(
        "--runtime-mode",
        dest="runtime_mode",
        choices=["legacy", "parity", "calibration", "hydraulic", "total-reflux"],
        default="parity",
        help=(
            "Runner behavior mode: "
            "parity=Pressure(spec)+Vapor(profile)+LiquidHydraulics(off), "
            "calibration=same closures as parity with explicit parity-check intent, "
            "hydraulic=Pressure(hydraulic)+Vapor(energy)+LiquidHydraulics(off unless explicitly enabled), "
            "total-reflux=hydraulic startup with feed and product draws suppressed and overhead condensate returned as reflux, "
            "legacy=use existing spec-driven behavior."
        ),
    )

    # Backward-compatible steps flag + alias
    p.add_argument("--n-steps", dest="n_steps", type=int, default=600)
    p.add_argument("--steps", dest="n_steps", type=int, default=None)

    p.add_argument("--dt", dest="dt_sec", type=float, default=None)
    p.add_argument("--log-every", dest="log_every_n_steps", type=int, default=None)
    p.add_argument(
        "--integrator",
        dest="integrator",
        choices=["explicit-euler", "bdf", "radau", "ida"],
        default="explicit-euler",
        help=(
            "Time integrator mode. explicit-euler keeps legacy behavior. "
            "bdf/radau use SciPy solve_ivp stiff methods and fall back to explicit-euler "
            "if SciPy is unavailable or a step solve fails. "
            "ida uses an implicit fixed-point DAE pilot stepper."
        ),
    )
    p.add_argument(
        "--integrator-rtol",
        dest="integrator_rtol",
        type=float,
        default=1.0e-3,
        help="Relative tolerance for stiff integrators and IDA fixed-point scaling.",
    )
    p.add_argument(
        "--integrator-atol",
        dest="integrator_atol",
        type=float,
        default=1.0e-6,
        help="Absolute tolerance for stiff integrators and IDA fixed-point scaling.",
    )
    p.add_argument(
        "--integrator-max-step-sec",
        dest="integrator_max_step_sec",
        type=float,
        default=None,
        help="Optional max internal substep size (s) for stiff integrators.",
    )
    p.add_argument(
        "--integrator-substep-sec",
        dest="integrator_substep_sec",
        type=float,
        default=None,
        help=(
            "Split each outer dt into fixed stiff substeps of this size (s). "
            "Useful when stiff solvers struggle on the full outer step."
        ),
    )
    p.add_argument(
        "--integrator-max-rhs-evals-per-step",
        dest="integrator_max_rhs_evals_per_step",
        type=int,
        default=24,
        help=(
            "Maximum RHS evaluations allowed inside one stiff integrator step. "
            "If exceeded, that step falls back to explicit Euler."
        ),
    )
    p.add_argument(
        "--integrator-step-wall-limit-sec",
        dest="integrator_step_wall_limit_sec",
        type=float,
        default=15.0,
        help=(
            "Per-step wall-time cap for stiff integrators. "
            "If exceeded, that step falls back to explicit Euler."
        ),
    )
    p.add_argument(
        "--ida-max-iter",
        dest="ida_max_iter",
        type=int,
        default=8,
        help="Maximum fixed-point iterations per IDA substep.",
    )
    p.add_argument(
        "--ida-relax",
        dest="ida_relax",
        type=float,
        default=1.0,
        help="Relaxation factor (0,1] for IDA fixed-point updates.",
    )

    # Backward-compatible temperature/energy flags
    p.add_argument("--no-temperature", dest="include_temperature", action="store_false")
    p.add_argument("--no-temp", dest="include_temperature", action="store_false")
    p.add_argument("--include-energy", dest="include_energy", action="store_true")
    p.add_argument("--energy", dest="include_energy", action="store_true")
    p.add_argument(
        "--disable-boundary-states",
        dest="include_boundary_states",
        action="store_false",
        help=(
            "Do not add separate reflux-drum/top and bottoms-sump boundary states. "
            "Use only for source-topology validation cases whose condenser and "
            "reboiler are already included in the stage set."
        ),
    )
    p.set_defaults(include_boundary_states=True)
    p.add_argument(
        "--disable-vapor-states",
        dest="include_vapor_states",
        action="store_false",
        help=(
            "Do not integrate tray vapor holdup/composition states. Vapor "
            "composition is treated as algebraic from liquid equilibrium; use "
            "only for constant-molar-overflow validation sources with no vapor holdup."
        ),
    )
    p.set_defaults(include_vapor_states=True)

    # Backward-compatible equilibrium flags
    p.add_argument("--no-equilibrium", dest="enable_equilibrium_relaxation", action="store_false")
    p.add_argument("--no-eq", dest="enable_equilibrium_relaxation", action="store_false")
    p.add_argument(
        "--equilibrium-relaxation-mode",
        "--eq-mode",
        dest="equilibrium_relaxation_mode",
        choices=["auto", "phase-holdup", "composition-only"],
        default="auto",
        help=(
            "Equilibrium-relaxation transfer target: "
            "phase-holdup=legacy flash phase split, "
            "composition-only=relax vapor composition at fixed MV."
        ),
    )
    p.add_argument(
        "--equilibrium-tau-sec",
        dest="equilibrium_tau_sec",
        type=float,
        default=None,
        help="Override tray phase-equilibrium relaxation time constant (sec).",
    )
    p.add_argument(
        "--equilibrium-tau-ramp-initial-sec",
        dest="equilibrium_tau_ramp_initial_sec",
        type=float,
        default=None,
        help="Initial effective equilibrium relaxation tau for an exponential runtime tau ramp.",
    )
    p.add_argument(
        "--equilibrium-tau-ramp-final-sec",
        dest="equilibrium_tau_ramp_final_sec",
        type=float,
        default=None,
        help="Final effective equilibrium relaxation tau for an exponential runtime tau ramp.",
    )
    p.add_argument(
        "--equilibrium-tau-ramp-decay-sec",
        dest="equilibrium_tau_ramp_decay_sec",
        type=float,
        default=None,
        help="Decay time constant for the runtime equilibrium tau ramp.",
    )
    p.add_argument(
        "--equilibrium-phase-holdup-guard-lbmol",
        dest="equilibrium_phase_holdup_guard_lbmol",
        type=float,
        default=None,
        help=(
            "Optional guard mass (lbmol) that softens phase-holdup relaxation "
            "toward the flash-predicted vapor amount."
        ),
    )
    p.add_argument(
        "--equilibrium-energy-damping-gain",
        dest="equilibrium_energy_damping_gain",
        type=float,
        default=None,
        help=(
            "Optional gain for previous-step energy-residual damping of "
            "phase-holdup relaxation. Zero disables this path."
        ),
    )
    p.add_argument(
        "--hydraulic-energy-temperature-damping",
        dest="hydraulic_energy_temperature_damping",
        type=float,
        default=None,
        help=(
            "Optional scalar damping on tray dT/dt in hydraulic+energy mode. "
            "1.0 keeps current behavior; smaller values soften the temperature path."
        ),
    )
    p.add_argument(
        "--hydraulic-energy-temperature-mode",
        dest="hydraulic_energy_temperature_mode",
        choices=[
            "legacy",
            "bubble-point-follower",
            "pressure-correction-follower",
            "enthalpy-state-follower",
        ],
        default=None,
        help="Tray temperature handling in hydraulic+energy mode.",
    )
    p.add_argument(
        "--hydraulic-energy-temperature-follow-tau-sec",
        dest="hydraulic_energy_temperature_follow_tau_sec",
        type=float,
        default=None,
        help="Relaxation time for hydraulic+energy bubble-point follower mode.",
    )
    p.add_argument(
        "--hydraulic-energy-temperature-resid-frac",
        dest="hydraulic_energy_temperature_resid_frac",
        type=float,
        default=None,
        help="Small residual dE/C fraction retained in bubble-point follower mode.",
    )
    p.add_argument(
        "--hydraulic-energy-temperature-pressure-slope-f-per-psi",
        dest="hydraulic_energy_temperature_pressure_slope_F_per_psi",
        type=float,
        default=None,
        help=(
            "Pressure-correction follower slope in F/psi for hydraulic+energy "
            "temperature mode."
        ),
    )
    p.add_argument(
        "--hydraulic-energy-temperature-target-refresh-steps",
        dest="hydraulic_energy_temperature_target_refresh_steps",
        type=int,
        default=None,
        help="Refresh cadence in outer time steps for cached bubble-point targets.",
    )

    p.add_argument(
        "--thermo",
        dest="thermo_mode",
        choices=[
            "stub",
            "relative-volatility",
            "simple-rv",
            "constant-alpha",
            "clapeyron",
            "dwsim",
            "dwsim-unifac",
            "dwsim-nrtl",
            "dwsim-uniquac",
            "dwsim-raoult",
            "dwsim-srk",
            "table",
            "table-pool",
        ],
        default="table-pool",
    )
    p.add_argument(
        "--clapeyron-model",
        dest="clapeyron_model",
        default="PR",
        help="Clapeyron.jl model constructor name used when --thermo clapeyron, e.g. PR, SRK, PCSAFT.",
    )
    p.add_argument(
        "--clapeyron-ideal-model",
        dest="clapeyron_ideal_model",
        default=None,
        help="Optional Clapeyron ideal-model constructor name, e.g. BasicIdeal or WalkerIdeal.",
    )
    p.add_argument(
        "--clapeyron-pr-parameter-source",
        dest="clapeyron_pr_parameter_source",
        choices=["default", "dwsim"],
        default="default",
        help=(
            "Parameter source for Clapeyron PR. 'default' uses Clapeyron's database; "
            "'dwsim' injects DWSIM PR Tc/Pc/MW/acentric-factor/kij values."
        ),
    )
    p.add_argument(
        "--dwsim-property-package",
        dest="dwsim_property_package",
        default="pr",
        help="DWSIM property package key for thermo_mode=dwsim. Examples: pr, srk, unifac, nrtl, uniquac, raoult.",
    )

    # Thermo throttling
    p.add_argument("--thermo-every", dest="thermo_every_n_steps", type=int, default=1)
    p.add_argument(
        "--disable-thermo-cadence-guardrails",
        dest="disable_thermo_cadence_guardrails",
        action="store_true",
        help="When --thermo-every > 1 in hydraulic runs, disable the default dT/dP/dx safety-refresh guardrails.",
    )
    p.add_argument("--thermo-refresh-dt", dest="thermo_refresh_dT_F", type=float, default=None)
    p.add_argument("--thermo-refresh-dp", dest="thermo_refresh_dP_psia", type=float, default=None)
    p.add_argument("--thermo-refresh-dx", dest="thermo_refresh_dx", type=float, default=None)
    p.add_argument(
        "--equilibrium-relaxation-live-pr",
        dest="equilibrium_relaxation_live_pr",
        action="store_true",
        default=False,
        help=(
            "Use live PR thermo only for the equilibrium-relaxation flash target "
            "while keeping the primary thermo mode unchanged."
        ),
    )
    p.add_argument(
        "--flash-feed-at-stage-conditions",
        dest="flash_feed_at_stage_conditions",
        action="store_true",
        default=None,
    )
    p.add_argument(
        "--no-flash-feed-at-stage-conditions",
        dest="flash_feed_at_stage_conditions",
        action="store_false",
    )
    p.add_argument("--thermo-table", dest="thermo_table_path", default=r"cache/thermo_table.json")
    p.add_argument(
        "--thermo-top-sat-table",
        dest="thermo_top_saturation_table_path",
        default=None,
        help="Optional local top-end PR-backed saturation table used by table thermo for condenser bubble-point lookup.",
    )
    p.add_argument(
        "--thermo-upper-section-table",
        dest="thermo_upper_section_table_path",
        default=None,
        help="Optional local upper-section PR-backed flash table overlay used by table thermo for the top stages.",
    )
    p.add_argument(
        "--thermo-upper-section-stage-count",
        dest="thermo_upper_section_stage_count",
        type=int,
        default=5,
        help="Number of top stages that use --thermo-upper-section-table when provided.",
    )
    p.add_argument(
        "--thermo-table-anchor-blend-count",
        dest="thermo_table_n_anchor_blend",
        type=int,
        default=3,
        help="Number of nearest composition anchors to blend in table/table-pool thermo modes.",
    )
    p.add_argument(
        "--thermo-table-anchor-blend-power",
        dest="thermo_table_anchor_blend_power",
        type=float,
        default=2.0,
        help="Inverse-distance power for composition-anchor blending in table/table-pool thermo modes.",
    )
    p.add_argument("--thermo-pool-workers", dest="thermo_pool_workers", type=int, default=2)
    p.add_argument("--thermo-pool-chunk-size", dest="thermo_pool_chunk_size", type=int, default=4)
    p.add_argument("--thermo-pool-timeout-sec", dest="thermo_pool_task_timeout_sec", type=float, default=None)
    p.add_argument("--reb-neighbor-vflow-hi-ratio", dest="reboiler_neighbor_vflow_hi_ratio", type=float, default=None)
    p.add_argument("--reb-neighbor-vflow-lo-ratio", dest="reboiler_neighbor_vflow_lo_ratio", type=float, default=None)
    p.add_argument("--use-excel-vapor-holdup", dest="use_excel_vapor_holdup", action="store_true")
    p.add_argument("--vapor-holdup-relaxation-sec", dest="vapor_holdup_relaxation_sec", type=float, default=None)
    p.add_argument(
        "--debug-freeze-tray-vapor-derivatives",
        dest="debug_freeze_tray_vapor_derivatives",
        action="store_true",
        help=(
            "Diagnostic: zero tray_V derivatives after RHS transport assembly and "
            "log the original tray vapor derivative defect."
        ),
    )
    p.add_argument(
        "--debug-override-reflux-composition",
        dest="debug_override_reflux_composition",
        action="store_true",
        help=(
            "Diagnostic: force the reflux composition entering stage 2 to match "
            "the condensed top vapor composition for reflux-loop closure isolation."
        ),
    )
    p.add_argument(
        "--debug-clamp-top-drum-pressure-psia",
        dest="debug_clamp_top_drum_pressure_psia",
        type=float,
        default=None,
        help=(
            "Diagnostic/initialization settle: override computed top-drum pressure "
            "used by hydraulic boundary logic. Restart states must be audited without this clamp."
        ),
    )
    p.add_argument(
        "--debug-clamp-top-drum-pressure-duration-sec",
        dest="debug_clamp_top_drum_pressure_duration_sec",
        type=float,
        default=None,
        help="Optional duration for --debug-clamp-top-drum-pressure-psia; after this time the clamp releases.",
    )
    p.add_argument(
        "--startup-total-reflux-washout-sec",
        dest="startup_total_reflux_washout_sec",
        type=float,
        default=None,
        help=(
            "For non-total-reflux startup runs, temporarily evaluate the RHS in "
            "total-reflux mode for this many seconds so the reflux drum is washed "
            "by live condenser condensate before product draws/feed boundaries resume."
        ),
    )
    p.add_argument(
        "--total-reflux-startup-ramp-tau-sec",
        dest="total_reflux_startup_ramp_tau_sec",
        type=float,
        default=None,
        help="In total-reflux mode, first-order ramp time constant for active boilup/reboiler duty.",
    )
    p.add_argument(
        "--total-reflux-startup-min-ramp-fraction",
        dest="total_reflux_startup_min_ramp_fraction",
        type=float,
        default=0.0,
        help="Minimum active boilup/duty fraction during the total-reflux startup ramp.",
    )
    p.add_argument(
        "--total-reflux-scale-reflux-with-startup-factor",
        dest="total_reflux_scale_reflux_with_startup_factor",
        action="store_true",
        help="In total-reflux mode, scale returned reflux by the startup ramp factor so reflux and boilup rise together.",
    )
    p.add_argument(
        "--total-reflux-boundary-ramp-duration-sec",
        dest="total_reflux_boundary_ramp_duration_sec",
        type=float,
        default=None,
        help=(
            "In total-reflux mode, start from active workbook feed/product boundaries "
            "and linearly seal them to zero over this duration. Omitted means immediate total reflux."
        ),
    )
    p.add_argument(
        "--hydraulic-pressure-relaxation-sec",
        dest="hydraulic_pressure_relaxation_sec",
        type=float,
        default=None,
    )
    p.add_argument(
        "--top-drum-pressure-temperature-relaxation-sec",
        dest="top_drum_pressure_temperature_relaxation_sec",
        type=float,
        default=None,
    )
    p.add_argument("--vapor-flow-relaxation-sec", dest="vapor_flow_relaxation_sec", type=float, default=None)
    p.add_argument(
        "--conductance-vflow-nominal-hi-ratio",
        dest="conductance_vflow_nominal_hi_ratio",
        type=float,
        default=None,
        help=(
            "Conductance-mode clamp: max internal vapor outflow as ratio of "
            "nominal profile V (e.g., 1.5)."
        ),
    )
    p.add_argument(
        "--stiff-vflow-smooth-clamp-lbmolph",
        dest="stiff_vflow_smooth_clamp_lbmolph",
        type=float,
        default=None,
        help=(
            "Optional smooth-clamp width (lbmol/h) for hydraulic vapor-flow limits "
            "during stiff RHS evaluation. "
            "None uses an automatic small value in stiff hydraulic mode; <=0 disables."
        ),
    )
    p.add_argument(
        "--pv-inner-max-iter",
        dest="pv_inner_max_iter",
        type=int,
        default=1,
        help=(
            "Inner fixed-point iterations per timestep for pressure-vapor coupling "
            "(applied only when pressure=hydraulic and vapor-flow=energy/conductance)."
        ),
    )
    p.add_argument(
        "--pv-inner-p-tol-psia",
        dest="pv_inner_p_tol_psia",
        type=float,
        default=0.05,
        help="Convergence tolerance for inner pressure iteration (psia).",
    )
    p.add_argument(
        "--pv-inner-v-tol-lbmolph",
        dest="pv_inner_v_tol_lbmolph",
        type=float,
        default=25.0,
        help="Convergence tolerance for inner vapor-flow iteration (lbmol/h).",
    )
    p.add_argument(
        "--enable-dae-pilot-algebraic-solve",
        dest="enable_dae_pilot_algebraic_solve",
        action="store_true",
        help=(
            "Enable pilot algebraic Newton solve for z=[P_tray, V_out] per timestep "
            "when pressure=hydraulic and vapor-flow model is energy/conductance."
        ),
    )
    p.add_argument(
        "--dae-pilot-max-iter",
        dest="dae_pilot_max_iter",
        type=int,
        default=3,
        help="Maximum Newton iterations for pilot algebraic solve.",
    )
    p.add_argument(
        "--dae-pilot-p-tol-psia",
        dest="dae_pilot_p_tol_psia",
        type=float,
        default=0.05,
        help="Pressure algebraic residual tolerance (psia).",
    )
    p.add_argument(
        "--dae-pilot-v-tol-lbmolph",
        dest="dae_pilot_v_tol_lbmolph",
        type=float,
        default=25.0,
        help="Vapor-flow algebraic residual tolerance (lbmol/h).",
    )
    p.add_argument(
        "--dae-pilot-jac-rel-step",
        dest="dae_pilot_jac_rel_step",
        type=float,
        default=1.0e-6,
        help="Relative finite-difference step for pilot Jacobian.",
    )
    p.add_argument(
        "--dae-pilot-line-search-max",
        dest="dae_pilot_line_search_max",
        type=int,
        default=4,
        help="Maximum backtracking line-search trials per Newton update.",
    )
    p.add_argument(
        "--fast-startup",
        dest="fast_startup",
        action="store_true",
        help=(
            "Reduce expensive pre-integration startup conditioning by skipping "
            "startup thermo conditioning, hydraulic-energy startup consistency, "
            "and top-drum startup steadying."
        ),
    )
    p.add_argument(
        "--enable-primary-thermo-startup-prewarm",
        dest="enable_primary_thermo_startup_prewarm",
        action="store_true",
        help=(
            "Opt in to an explicit primary-backend startup prewarm pass. "
            "Useful for profiling or warm-kernel experiments; disabled by default "
            "because it can relocate startup cost without improving end-to-end wall time."
        ),
    )
    p.add_argument(
        "--enable-startup-seed-cache",
        dest="enable_startup_seed_cache",
        action="store_true",
        help=(
            "Cache the post-startup state and thermo packets to JSON so later runs "
            "can skip vapor-holdup initialization and other fresh-startup work."
        ),
    )
    p.add_argument(
        "--refresh-startup-seed-cache",
        dest="refresh_startup_seed_cache",
        action="store_true",
        help="Force a fresh startup pass and overwrite the startup seed cache.",
    )
    p.add_argument(
        "--startup-seed-cache-path",
        dest="startup_seed_cache_path",
        default=None,
        help="Optional path for the startup seed cache JSON file.",
    )
    p.add_argument(
        "--disable-startup-thermo-conditioning",
        dest="enable_startup_thermo_conditioning",
        action="store_false",
        help="Disable startup thermo-consistent state conditioning.",
    )
    p.add_argument(
        "--startup-thermo-conditioning-iters",
        dest="startup_thermo_conditioning_iters",
        type=int,
        default=2,
    )
    p.add_argument(
        "--startup-thermo-conditioning-relax",
        dest="startup_thermo_conditioning_relaxation",
        type=float,
        default=1.0,
    )
    p.add_argument(
        "--disable-restart-reentry-settling",
        dest="enable_restart_reentry_settling",
        action="store_false",
        help=(
            "For explicit restart/boundary-state workbooks, skip the hidden "
            "restart re-entry conditioning pass. Use this when the workbook "
            "already contains a deliberately reconciled initial state."
        ),
    )
    p.add_argument(
        "--enable-startup-hydraulic-energy-consistency",
        dest="enable_startup_hydraulic_energy_consistency",
        action="store_true",
        help=(
            "Run a bounded t=0 consistency relaxation for hydraulic+energy cases "
            "using the pilot algebraic solve plus short pseudo-time steps."
        ),
    )
    p.add_argument(
        "--startup-hydraulic-energy-consistency-iters",
        dest="startup_hydraulic_energy_consistency_iters",
        type=int,
        default=6,
    )
    p.add_argument(
        "--startup-hydraulic-energy-consistency-dt-sec",
        dest="startup_hydraulic_energy_consistency_dt_sec",
        type=float,
        default=0.5,
    )
    p.add_argument(
        "--startup-hydraulic-energy-consistency-mass-tol-lbmolph",
        dest="startup_hydraulic_energy_consistency_mass_tol_lbmolph",
        type=float,
        default=5.0,
    )
    p.add_argument(
        "--startup-hydraulic-energy-consistency-energy-tol-btups",
        dest="startup_hydraulic_energy_consistency_energy_tol_btups",
        type=float,
        default=1000.0,
    )
    p.add_argument(
        "--enable-liquid-hydraulic-override",
        dest="enable_liquid_hydraulic_override",
        action="store_true",
        help="Force-enable internal liquid hydraulic downflow override.",
    )
    p.add_argument(
        "--disable-liquid-hydraulic-override",
        dest="enable_liquid_hydraulic_override",
        action="store_false",
        help="Disable internal liquid hydraulic downflow override (profile-only internal L).",
    )
    p.set_defaults(enable_liquid_hydraulic_override=None)
    p.add_argument(
        "--liquid-hydraulic-override-alpha",
        dest="liquid_hydraulic_override_alpha",
        type=float,
        default=None,
        help="Blend for internal liquid hydraulics override: 0=profile, 1=full hydraulic.",
    )
    p.add_argument(
        "--liquid-hydraulic-model",
        dest="liquid_hydraulic_model",
        choices=["francis", "linear-holdup", "skogestad-linear"],
        default=None,
        help="Internal liquid hydraulic closure. linear-holdup is intended for source-topology validation.",
    )
    p.add_argument(
        "--liquid-hydraulic-htc-sec",
        dest="liquid_hydraulic_htc_sec",
        type=float,
        default=None,
        help="Optional hydraulic time constant (s) reserved for non-Francis liquid closures.",
    )
    p.add_argument(
        "--enable-startup-hydraulic-sequence",
        dest="enable_startup_hydraulic_sequence",
        action="store_true",
        help=(
            "Startup sequence: pressure-only first, then energy vapor closure, "
            "then residual-gated liquid-hydraulic ramp."
        ),
    )
    p.add_argument(
        "--startup-sequence-energy-on-sec",
        dest="startup_sequence_energy_on_sec",
        type=float,
        default=30.0,
        help="Sequence time (s) to allow vapor_flow_model=energy.",
    )
    p.add_argument(
        "--startup-sequence-liquid-on-sec",
        dest="startup_sequence_liquid_on_sec",
        type=float,
        default=120.0,
        help="Sequence time (s) to begin liquid-hydraulic override ramp.",
    )
    p.add_argument(
        "--startup-sequence-liquid-ramp-sec",
        dest="startup_sequence_liquid_ramp_sec",
        type=float,
        default=180.0,
        help="Ramp timescale (s) for liquid-hydraulic override alpha.",
    )
    p.add_argument(
        "--startup-sequence-mass-resid-gate-lbmolph",
        dest="startup_sequence_mass_resid_gate_lbmolph",
        type=float,
        default=250.0,
        help="If max tray mass residual exceeds this, liquid-hydraulic alpha is paused/backed off.",
    )
    p.add_argument(
        "--startup-sequence-liquid-backoff-sec",
        dest="startup_sequence_liquid_backoff_sec",
        type=float,
        default=None,
        help="Optional backoff timescale (s) when residual-gate is exceeded.",
    )
    p.add_argument(
        "--enable-startup-vapor-homotopy",
        dest="enable_startup_vapor_homotopy",
        action="store_true",
        help=(
            "Startup sequence extension: after liquid hydraulics ramp, blend vapor traffic "
            "smoothly from the profile flow to the configured dynamic vapor-flow closure."
        ),
    )
    p.add_argument(
        "--startup-sequence-profile-hold-sec",
        dest="startup_sequence_profile_hold_sec",
        type=float,
        default=0.0,
        help="Minimum initial hold time (s) with both liquid and vapor traffic on profile values.",
    )
    p.add_argument(
        "--startup-sequence-vapor-on-sec",
        dest="startup_sequence_vapor_on_sec",
        type=float,
        default=None,
        help=(
            "Time (s) to begin vapor-flow homotopy. Defaults to liquid_on + liquid_ramp "
            "when omitted."
        ),
    )
    p.add_argument(
        "--startup-sequence-vapor-ramp-sec",
        dest="startup_sequence_vapor_ramp_sec",
        type=float,
        default=60.0,
        help="Cosine-ramp duration (s) for vapor-flow homotopy beta.",
    )
    p.add_argument(
        "--startup-sequence-vapor-rel-rate-gate-per-s",
        dest="startup_sequence_vapor_rel_rate_gate_per_s",
        type=float,
        default=1.0e-2,
        help="Pause vapor beta ramp when previous max relative inventory rate exceeds this [1/s].",
    )
    p.add_argument(
        "--startup-sequence-vapor-backoff-sec",
        dest="startup_sequence_vapor_backoff_sec",
        type=float,
        default=None,
        help="Optional beta backoff timescale (s) when the vapor residual gate is exceeded.",
    )
    p.add_argument(
        "--disable-steady-state-detection",
        dest="enable_steady_state_detection",
        action="store_false",
        help="Disable runtime steady-state detector diagnostics.",
    )
    p.add_argument("--steady-state-window-sec", dest="steady_state_window_sec", type=float, default=30.0)
    p.add_argument("--steady-state-min-time-sec", dest="steady_state_min_time_sec", type=float, default=60.0)
    p.add_argument(
        "--steady-state-rel-rate-tol-per-s",
        dest="steady_state_rel_state_rate_tol_per_s",
        type=float,
        default=3.0e-3,
        help="Tolerance on max relative inventory rate |dM/dt|/(|M|+floor) [1/s].",
    )
    p.add_argument(
        "--steady-state-kpi-slope-tol-per-s",
        dest="steady_state_kpi_slope_tol_per_s",
        type=float,
        default=1.0e-4,
        help="Tolerance on KPI slope magnitude (distillate/bottoms composition) [1/s].",
    )
    p.add_argument(
        "--steady-state-mv-rate-tol-per-s",
        dest="steady_state_mv_rate_tol_per_s",
        type=float,
        default=20.0,
        help="Tolerance on manipulated-variable rate (reflux/boilup) [lbmol/h/s].",
    )
    p.add_argument(
        "--steady-state-temp-rate-tol-fps",
        dest="steady_state_temp_rate_tol_F_per_s",
        type=float,
        default=0.15,
        help="Tolerance on max tray temperature rate [F/s].",
    )
    p.add_argument(
        "--steady-state-sp-error-tol",
        dest="steady_state_sp_error_tol",
        type=float,
        default=0.02,
        help="Tolerance on max composition setpoint error (mole fraction).",
    )
    p.add_argument(
        "--steady-state-require-sp",
        dest="steady_state_require_sp",
        action="store_true",
        help="Require composition setpoint error criterion for SS=1.",
    )
    p.add_argument(
        "--steady-state-rate-denom-floor-lbmol",
        dest="steady_state_rate_denom_floor_lbmol",
        type=float,
        default=1.0,
        help="Denominator floor used in relative inventory-rate metric (lbmol).",
    )

    # Boundary overrides
    p.add_argument("--reflux", dest="reflux_lbmolph", type=float, default=None)
    p.add_argument("--boilup", dest="boilup_lbmolph", type=float, default=None)
    p.add_argument("--feed-step-time", dest="feed_step_time_sec", type=float, default=None)
    p.add_argument("--feed-step-scale", dest="feed_step_scale", type=float, default=None)
    p.add_argument(
        "--condenser-duty-mode",
        dest="condenser_duty_mode",
        choices=["total-condense", "specified"],
        default="total-condense",
    )
    p.add_argument("--condenser-duty-btuph", dest="condenser_duty_btu_per_h", type=float, default=None)
    p.add_argument(
        "--condenser-duty-trim-btuph",
        dest="condenser_duty_trim_btu_per_h",
        type=float,
        default=None,
    )
    p.add_argument(
        "--init-pack-top-drum-vapor-to-pressure",
        dest="init_pack_top_drum_vapor_to_pressure",
        action="store_true",
        help=(
            "Initialization diagnostic: scale explicit top-drum vapor inventory "
            "so the raw top-drum pressure starts at the target pressure."
        ),
    )
    p.add_argument(
        "--init-top-drum-vapor-pressure-psia",
        dest="init_top_drum_vapor_pressure_psia",
        type=float,
        default=None,
        help=(
            "Pressure target for --init-pack-top-drum-vapor-to-pressure; "
            "defaults to --top-pressure-sp, then the workbook stage-1 pressure."
        ),
    )
    p.add_argument(
        "--init-match-condenser-duty",
        dest="init_match_condenser_duty",
        action="store_true",
        help=(
            "Initialization diagnostic: evaluate the live total-condenser duty "
            "requirement at t=0 and use it as the initial condenser-duty bias."
        ),
    )
    p.add_argument(
        "--init-align-top-liquid-to-condensate",
        dest="init_align_top_liquid_to_condensate",
        action="store_true",
        help=(
            "Initialization diagnostic: preserve reflux-drum liquid holdup but "
            "replace its component split with the live condenser condensate composition."
        ),
    )
    p.add_argument("--enable-level-control", dest="enable_level_control", action="store_true")
    p.add_argument(
        "--top-level-pv-mode",
        dest="top_level_pv_mode",
        choices=["molar-holdup", "true-level"],
        default="molar-holdup",
    )
    p.add_argument(
        "--ignore-workbook-level-pv-mode",
        dest="ignore_workbook_level_pv_mode",
        action="store_true",
        help="Use CLI level-controller PV modes even when the workbook specifies Top/Bottom Level PV Mode.",
    )
    p.add_argument("--top-level-sp", dest="top_level_sp_lbmol", type=float, default=None)
    p.add_argument("--top-level-sp-frac", dest="top_level_sp_frac", type=float, default=None)
    p.add_argument(
        "--bottom-level-pv-mode",
        dest="bottom_level_pv_mode",
        choices=["molar-holdup", "true-level"],
        default="molar-holdup",
    )
    p.add_argument("--bottom-level-sp", dest="bottom_level_sp_lbmol", type=float, default=None)
    p.add_argument("--bottom-level-sp-frac", dest="bottom_level_sp_frac", type=float, default=None)
    p.add_argument("--top-level-kc", dest="top_level_kc", type=float, default=None)
    p.add_argument("--top-level-ti", dest="top_level_ti_sec", type=float, default=None)
    p.add_argument("--enable-top-level-feedforward", dest="enable_top_level_feedforward", action="store_true")
    p.add_argument("--top-level-feedforward-gain", dest="top_level_feedforward_gain", type=float, default=None)
    p.add_argument("--bottom-level-kc", dest="bottom_level_kc", type=float, default=None)
    p.add_argument("--bottom-level-ti", dest="bottom_level_ti_sec", type=float, default=None)
    p.add_argument("--enable-pressure-control", dest="enable_pressure_control", action="store_true")
    p.add_argument(
        "--pressure-control-mv",
        dest="pressure_control_mv",
        choices=["auto", "condenser-duty", "top-anchor"],
        default="auto",
    )
    p.add_argument(
        "--allow-coupled-pressure-duty",
        dest="allow_coupled_pressure_duty",
        action="store_true",
        help=(
            "Allow pressure-control-mv=condenser-duty to remain coupled with "
            "condenser-duty-mode=total-condense. Default behavior auto-switches to top-anchor."
        ),
    )
    p.add_argument(
        "--disable-coupled-total-condenser-partial-condense",
        dest="enable_coupled_total_condenser_partial_condense",
        action="store_false",
        help=(
            "When pressure-control-mv=condenser-duty is explicitly coupled with "
            "condenser-duty-mode=total-condense, keep the older strict "
            "full-condensation material split instead of allowing duty-limited "
            "partial condensation to the top drum."
        ),
    )
    p.set_defaults(enable_coupled_total_condenser_partial_condense=True)
    p.add_argument("--top-pressure-sp", dest="top_pressure_sp_psia", type=float, default=None)
    p.add_argument("--top-pressure-kc", dest="top_pressure_kc", type=float, default=None)
    p.add_argument("--top-pressure-ti", dest="top_pressure_ti_sec", type=float, default=None)
    p.add_argument("--top-pressure-pv-filter-tau-sec", dest="top_pressure_pv_filter_tau_sec", type=float, default=None)
    p.add_argument("--top-pressure-mv-slew-limit-per-s", dest="top_pressure_mv_slew_limit_per_s", type=float, default=None)
    p.add_argument("--top-pressure-resid-ref-btups", dest="top_pressure_resid_ref_btups", type=float, default=None)
    p.add_argument("--top-pressure-resid-min-gain", dest="top_pressure_resid_min_gain", type=float, default=0.25)
    p.add_argument("--top-pressure-anchor-min", dest="top_pressure_anchor_min_psia", type=float, default=None)
    p.add_argument("--top-pressure-anchor-max", dest="top_pressure_anchor_max_psia", type=float, default=None)
    p.add_argument("--condenser-duty-min-btuph", dest="condenser_duty_min_btu_per_h", type=float, default=None)
    p.add_argument("--condenser-duty-max-btuph", dest="condenser_duty_max_btu_per_h", type=float, default=None)
    p.add_argument("--condenser-pressure-drop-psi", dest="condenser_pressure_drop_psi", type=float, default=None)
    p.add_argument("--top-drum-vapor-volume-ft3", dest="top_drum_vapor_volume_ft3", type=float, default=None)
    p.add_argument("--top-drum-total-volume-ft3", dest="top_drum_total_volume_ft3", type=float, default=None)
    p.add_argument("--bottom-sump-total-volume-ft3", dest="bottom_sump_total_volume_ft3", type=float, default=None)
    p.add_argument(
        "--disable-top-drum-pressure-gate",
        dest="enforce_top_drum_pressure_gate",
        action="store_false",
        help="Disable pressure-direction gating on stage-2 to top-drum vapor slip.",
    )
    p.add_argument(
        "--top-drum-pressure-gate-soft-psi",
        dest="top_drum_pressure_gate_soft_psi",
        type=float,
        default=None,
        help="Soft transition width (psi) for top-drum pressure gate; <=0 gives hard gating.",
    )
    p.add_argument(
        "--disable-top-pressure-ordering",
        dest="enforce_top_pressure_ordering",
        action="store_false",
        help="Disable top-end pressure ordering enforcement (stage-1 pressure >= top-drum pressure).",
    )
    p.add_argument(
        "--top-pressure-ordering-margin-psi",
        dest="top_pressure_ordering_margin_psi",
        type=float,
        default=0.0,
        help="Minimum margin (psi) enforced for stage-1 pressure above top-drum pressure.",
    )
    p.add_argument("--enable-top-psv", dest="enable_top_psv", action="store_true")
    p.add_argument("--top-psv-sp", dest="top_psv_setpoint_psia", type=float, default=None)
    p.add_argument(
        "--top-psv-gain-lbmolps-psi",
        dest="top_psv_gain_lbmolps_per_psi",
        type=float,
        default=None,
    )
    p.add_argument("--top-psv-max-lbmolps", dest="top_psv_max_vent_lbmolps", type=float, default=None)
    p.add_argument("--enable-distillate-composition-control", dest="enable_distillate_composition_control", action="store_true")
    p.add_argument("--distillate-comp-component", dest="distillate_composition_component", default="C4")
    p.add_argument("--distillate-comp-sp", dest="distillate_composition_sp_molfrac", type=float, default=None)
    p.add_argument("--distillate-comp-kc", dest="distillate_composition_kc", type=float, default=None)
    p.add_argument("--distillate-comp-ti", dest="distillate_composition_ti_sec", type=float, default=None)
    p.add_argument("--reflux-cmd-min", dest="reflux_cmd_min_lbmolph", type=float, default=None)
    p.add_argument("--reflux-cmd-max", dest="reflux_cmd_max_lbmolph", type=float, default=None)
    p.add_argument(
        "--disable-reflux-feasibility-cap",
        dest="enable_reflux_feasibility_cap",
        action="store_false",
        help="Disable reflux max-feasibility cap in distillate composition control.",
    )
    p.add_argument("--enable-bottoms-composition-control", dest="enable_bottoms_composition_control", action="store_true")
    p.add_argument("--bottoms-comp-component", dest="bottoms_composition_component", default="C5")
    p.add_argument("--bottoms-comp-sp", dest="bottoms_composition_sp_molfrac", type=float, default=None)
    p.add_argument("--bottoms-comp-kc", dest="bottoms_composition_kc", type=float, default=None)
    p.add_argument("--bottoms-comp-ti", dest="bottoms_composition_ti_sec", type=float, default=None)
    p.add_argument(
        "--bottoms-comp-mv",
        dest="bottoms_composition_mv",
        choices=["boilup", "reboiler-duty"],
        default="boilup",
    )
    p.add_argument("--boilup-cmd-min", dest="boilup_cmd_min_lbmolph", type=float, default=None)
    p.add_argument("--boilup-cmd-max", dest="boilup_cmd_max_lbmolph", type=float, default=None)
    p.add_argument("--reboiler-duty-cmd-min-btuph", dest="reboiler_duty_cmd_min_btu_per_h", type=float, default=None)
    p.add_argument("--reboiler-duty-cmd-max-btuph", dest="reboiler_duty_cmd_max_btu_per_h", type=float, default=None)
    p.add_argument("--reboiler-duty-btuph", dest="reboiler_duty_btu_per_h", type=float, default=None)
    # Backward-compatible aliases. These are converted to flow clamps using
    # current distillate flow if reflux-cmd min/max are not provided.
    p.add_argument("--reflux-ratio-min", dest="reflux_ratio_min", type=float, default=None)
    p.add_argument("--reflux-ratio-max", dest="reflux_ratio_max", type=float, default=None)

    p.add_argument("--logs-dir", dest="logs_dir", default="logs")
    p.add_argument("--no-write-logs", dest="write_logs", action="store_false")
    p.add_argument("--no-logs", dest="write_logs", action="store_false")
    p.add_argument(
        "--allow-repeat-command",
        dest="allow_repeat_command",
        action="store_true",
        help=(
            "Allow running even when this exact CLI command already exists in "
            "docs/experiment_ledger.csv."
        ),
    )

    raw_argv: List[str] = list(argv) if argv is not None else list(sys.argv[1:])
    args = p.parse_args(raw_argv)

    if args.n_steps is None:
        args.n_steps = 600

    module_name = "dynamic_distillation.dynamic_run_scaffold_v1"
    project_root = Path(__file__).resolve().parents[2]
    ledger_csv = project_root / "docs" / "experiment_ledger.csv"
    candidate_cmd = compose_cli_command(module_name, raw_argv)
    candidate_identity = compose_cli_command_identity(module_name, raw_argv)
    cmd_matches = find_exact_command_matches(
        ledger_csv_path=ledger_csv,
        module_name=module_name,
        argv=raw_argv,
    )
    if cmd_matches and (not bool(args.allow_repeat_command)):
        n_ok = sum(1 for m in cmd_matches if str(m.status).lower() == "ok")
        n_not_ok = len(cmd_matches) - n_ok
        print("[Abort] Exact command already exists in experiment ledger.")
        print(f"Candidate: {candidate_cmd}")
        print(f"Identity:  {candidate_identity}")
        print(f"Matches: total={len(cmd_matches)}, ok={n_ok}, not_ok={n_not_ok}")
        print("Recent matching runs:")
        for m in cmd_matches[:8]:
            p_txt = m.P_top_pv_psia_final if m.P_top_pv_psia_final else "NA"
            xd_txt = m.xD_pv_final if m.xD_pv_final else "NA"
            xb_txt = m.xB_pv_final if m.xB_pv_final else "NA"
            print(
                f"  - run_id={m.run_id} status={m.status} t_final={m.t_final_s} "
                f"P_top={p_txt} xD={xd_txt} xB={xb_txt} src={m.command_source}"
            )
        print("Pass --allow-repeat-command to run this command again intentionally.")
        return 2
    if cmd_matches and bool(args.allow_repeat_command):
        print(f"[Warn] Repeating known command (matches={len(cmd_matches)}).")

    cfg = RunnerConfig(
        excel_path=str(args.excel_path),
        n_steps=int(args.n_steps),
        dt_sec=args.dt_sec,
        log_every_n_steps=args.log_every_n_steps,
        runtime_mode=str(args.runtime_mode),
        integrator=str(args.integrator),
        integrator_rtol=float(args.integrator_rtol),
        integrator_atol=float(args.integrator_atol),
        integrator_max_step_sec=args.integrator_max_step_sec,
        integrator_substep_sec=args.integrator_substep_sec,
        integrator_max_rhs_evals_per_step=args.integrator_max_rhs_evals_per_step,
        integrator_step_wall_limit_sec=args.integrator_step_wall_limit_sec,
        ida_max_iter=args.ida_max_iter,
        ida_relax=args.ida_relax,
        include_temperature=bool(args.include_temperature),
        include_energy=bool(args.include_energy),
        include_boundary_states=bool(args.include_boundary_states),
        include_vapor_states=bool(args.include_vapor_states),
        enable_equilibrium_relaxation=bool(args.enable_equilibrium_relaxation),
        equilibrium_relaxation_mode=str(args.equilibrium_relaxation_mode),
        equilibrium_tau_sec=args.equilibrium_tau_sec,
        equilibrium_tau_ramp_initial_sec=args.equilibrium_tau_ramp_initial_sec,
        equilibrium_tau_ramp_final_sec=args.equilibrium_tau_ramp_final_sec,
        equilibrium_tau_ramp_decay_sec=args.equilibrium_tau_ramp_decay_sec,
        equilibrium_phase_holdup_guard_lbmol=args.equilibrium_phase_holdup_guard_lbmol,
        equilibrium_energy_damping_gain=args.equilibrium_energy_damping_gain,
        hydraulic_energy_temperature_damping=args.hydraulic_energy_temperature_damping,
        hydraulic_energy_temperature_mode=args.hydraulic_energy_temperature_mode,
        hydraulic_energy_temperature_follow_tau_sec=args.hydraulic_energy_temperature_follow_tau_sec,
        hydraulic_energy_temperature_resid_frac=args.hydraulic_energy_temperature_resid_frac,
        hydraulic_energy_temperature_pressure_slope_F_per_psi=(
            args.hydraulic_energy_temperature_pressure_slope_F_per_psi
        ),
        hydraulic_energy_temperature_target_refresh_steps=args.hydraulic_energy_temperature_target_refresh_steps,
        thermo_mode=str(args.thermo_mode),
        clapeyron_model=str(args.clapeyron_model),
        clapeyron_ideal_model=args.clapeyron_ideal_model,
        clapeyron_pr_parameter_source=str(args.clapeyron_pr_parameter_source),
        dwsim_property_package=str(args.dwsim_property_package),
        thermo_every_n_steps=int(args.thermo_every_n_steps),
        enable_thermo_cadence_guardrails=(not bool(args.disable_thermo_cadence_guardrails)),
        thermo_refresh_dT_F=args.thermo_refresh_dT_F,
        thermo_refresh_dP_psia=args.thermo_refresh_dP_psia,
        thermo_refresh_dx=args.thermo_refresh_dx,
        equilibrium_relaxation_live_pr=bool(args.equilibrium_relaxation_live_pr),
        thermo_table_path=args.thermo_table_path,
        thermo_top_saturation_table_path=args.thermo_top_saturation_table_path,
        thermo_upper_section_table_path=args.thermo_upper_section_table_path,
        thermo_upper_section_stage_count=int(args.thermo_upper_section_stage_count),
        thermo_table_n_anchor_blend=int(args.thermo_table_n_anchor_blend),
        thermo_table_anchor_blend_power=float(args.thermo_table_anchor_blend_power),
        thermo_pool_workers=args.thermo_pool_workers,
        thermo_pool_chunk_size=args.thermo_pool_chunk_size,
        thermo_pool_task_timeout_sec=args.thermo_pool_task_timeout_sec,
        reboiler_neighbor_vflow_hi_ratio=args.reboiler_neighbor_vflow_hi_ratio,
        reboiler_neighbor_vflow_lo_ratio=args.reboiler_neighbor_vflow_lo_ratio,
        vapor_holdup_relaxation_sec=args.vapor_holdup_relaxation_sec,
        debug_freeze_tray_vapor_derivatives=bool(args.debug_freeze_tray_vapor_derivatives),
        debug_override_reflux_composition=bool(args.debug_override_reflux_composition),
        debug_clamp_top_drum_pressure_psia=args.debug_clamp_top_drum_pressure_psia,
        debug_clamp_top_drum_pressure_duration_sec=args.debug_clamp_top_drum_pressure_duration_sec,
        startup_total_reflux_washout_sec=args.startup_total_reflux_washout_sec,
        total_reflux_startup_ramp_tau_sec=args.total_reflux_startup_ramp_tau_sec,
        total_reflux_startup_min_ramp_fraction=args.total_reflux_startup_min_ramp_fraction,
        total_reflux_scale_reflux_with_startup_factor=bool(args.total_reflux_scale_reflux_with_startup_factor),
        total_reflux_boundary_ramp_duration_sec=args.total_reflux_boundary_ramp_duration_sec,
        hydraulic_pressure_relaxation_sec=args.hydraulic_pressure_relaxation_sec,
        top_drum_pressure_temperature_relaxation_sec=args.top_drum_pressure_temperature_relaxation_sec,
        vapor_flow_relaxation_sec=args.vapor_flow_relaxation_sec,
        conductance_vflow_nominal_hi_ratio=args.conductance_vflow_nominal_hi_ratio,
        stiff_vflow_smooth_clamp_lbmolph=args.stiff_vflow_smooth_clamp_lbmolph,
        pv_inner_max_iter=int(args.pv_inner_max_iter),
        pv_inner_p_tol_psia=args.pv_inner_p_tol_psia,
        pv_inner_v_tol_lbmolph=args.pv_inner_v_tol_lbmolph,
        enable_dae_pilot_algebraic_solve=bool(args.enable_dae_pilot_algebraic_solve),
        dae_pilot_max_iter=int(args.dae_pilot_max_iter),
        dae_pilot_p_tol_psia=args.dae_pilot_p_tol_psia,
        dae_pilot_v_tol_lbmolph=args.dae_pilot_v_tol_lbmolph,
        dae_pilot_jac_rel_step=float(args.dae_pilot_jac_rel_step),
        dae_pilot_line_search_max=int(args.dae_pilot_line_search_max),
        enable_liquid_hydraulic_override=args.enable_liquid_hydraulic_override,
        liquid_hydraulic_override_alpha=args.liquid_hydraulic_override_alpha,
        liquid_hydraulic_model=args.liquid_hydraulic_model,
        liquid_hydraulic_htc_sec=args.liquid_hydraulic_htc_sec,
        reflux_lbmolph=args.reflux_lbmolph,
        boilup_lbmolph=args.boilup_lbmolph,
        feed_step_time_sec=args.feed_step_time_sec,
        feed_step_scale=args.feed_step_scale,
        condenser_duty_mode=str(args.condenser_duty_mode),
        condenser_duty_btu_per_h=args.condenser_duty_btu_per_h,
        condenser_duty_trim_btu_per_h=args.condenser_duty_trim_btu_per_h,
        init_pack_top_drum_vapor_to_pressure=bool(args.init_pack_top_drum_vapor_to_pressure),
        init_top_drum_vapor_pressure_psia=args.init_top_drum_vapor_pressure_psia,
        init_match_condenser_duty=bool(args.init_match_condenser_duty),
        init_align_top_liquid_to_condensate=bool(args.init_align_top_liquid_to_condensate),
        enable_level_control=bool(args.enable_level_control),
        top_level_pv_mode=str(args.top_level_pv_mode),
        ignore_workbook_level_pv_mode=bool(args.ignore_workbook_level_pv_mode),
        top_level_sp_lbmol=args.top_level_sp_lbmol,
        top_level_sp_frac=args.top_level_sp_frac,
        bottom_level_pv_mode=str(args.bottom_level_pv_mode),
        bottom_level_sp_lbmol=args.bottom_level_sp_lbmol,
        bottom_level_sp_frac=args.bottom_level_sp_frac,
        top_level_kc=args.top_level_kc,
        top_level_ti_sec=args.top_level_ti_sec,
        enable_top_level_feedforward=bool(args.enable_top_level_feedforward),
        top_level_feedforward_gain=args.top_level_feedforward_gain,
        bottom_level_kc=args.bottom_level_kc,
        bottom_level_ti_sec=args.bottom_level_ti_sec,
        enable_pressure_control=bool(args.enable_pressure_control),
        pressure_control_mv=str(args.pressure_control_mv),
        allow_coupled_pressure_duty=bool(args.allow_coupled_pressure_duty),
        enable_coupled_total_condenser_partial_condense=bool(
            args.enable_coupled_total_condenser_partial_condense
        ),
        top_pressure_sp_psia=args.top_pressure_sp_psia,
        top_pressure_kc=args.top_pressure_kc,
        top_pressure_ti_sec=args.top_pressure_ti_sec,
        top_pressure_pv_filter_tau_sec=args.top_pressure_pv_filter_tau_sec,
        top_pressure_mv_slew_limit_per_s=args.top_pressure_mv_slew_limit_per_s,
        top_pressure_resid_ref_btups=args.top_pressure_resid_ref_btups,
        top_pressure_resid_min_gain=float(args.top_pressure_resid_min_gain),
        top_pressure_anchor_min_psia=args.top_pressure_anchor_min_psia,
        top_pressure_anchor_max_psia=args.top_pressure_anchor_max_psia,
        condenser_duty_min_btu_per_h=args.condenser_duty_min_btu_per_h,
        condenser_duty_max_btu_per_h=args.condenser_duty_max_btu_per_h,
        condenser_pressure_drop_psi=args.condenser_pressure_drop_psi,
        top_drum_vapor_volume_ft3=args.top_drum_vapor_volume_ft3,
        top_drum_total_volume_ft3=args.top_drum_total_volume_ft3,
        bottom_sump_total_volume_ft3=args.bottom_sump_total_volume_ft3,
        enforce_top_drum_pressure_gate=bool(args.enforce_top_drum_pressure_gate),
        top_drum_pressure_gate_soft_psi=args.top_drum_pressure_gate_soft_psi,
        enforce_top_pressure_ordering=bool(args.enforce_top_pressure_ordering),
        top_pressure_ordering_margin_psi=float(args.top_pressure_ordering_margin_psi),
        enable_top_psv=bool(args.enable_top_psv),
        top_psv_setpoint_psia=args.top_psv_setpoint_psia,
        top_psv_gain_lbmolps_per_psi=args.top_psv_gain_lbmolps_per_psi,
        top_psv_max_vent_lbmolps=args.top_psv_max_vent_lbmolps,
        enable_distillate_composition_control=bool(args.enable_distillate_composition_control),
        distillate_composition_component=str(args.distillate_composition_component),
        distillate_composition_sp_molfrac=args.distillate_composition_sp_molfrac,
        distillate_composition_kc=args.distillate_composition_kc,
        distillate_composition_ti_sec=args.distillate_composition_ti_sec,
        reflux_cmd_min_lbmolph=args.reflux_cmd_min_lbmolph,
        reflux_cmd_max_lbmolph=args.reflux_cmd_max_lbmolph,
        enable_reflux_feasibility_cap=bool(args.enable_reflux_feasibility_cap),
        reflux_ratio_min=args.reflux_ratio_min,
        reflux_ratio_max=args.reflux_ratio_max,
        enable_bottoms_composition_control=bool(args.enable_bottoms_composition_control),
        bottoms_composition_component=str(args.bottoms_composition_component),
        bottoms_composition_sp_molfrac=args.bottoms_composition_sp_molfrac,
        bottoms_composition_kc=args.bottoms_composition_kc,
        bottoms_composition_ti_sec=args.bottoms_composition_ti_sec,
        bottoms_composition_mv=str(args.bottoms_composition_mv),
        boilup_cmd_min_lbmolph=args.boilup_cmd_min_lbmolph,
        boilup_cmd_max_lbmolph=args.boilup_cmd_max_lbmolph,
        reboiler_duty_cmd_min_btu_per_h=args.reboiler_duty_cmd_min_btu_per_h,
        reboiler_duty_cmd_max_btu_per_h=args.reboiler_duty_cmd_max_btu_per_h,
        reboiler_duty_btu_per_h=args.reboiler_duty_btu_per_h,
        logs_dir=str(args.logs_dir),
        run_name=args.run_name,
        run_description=args.run_description,
        write_logs=bool(args.write_logs),
        use_excel_vapor_holdup=bool(args.use_excel_vapor_holdup),
        fast_startup=bool(args.fast_startup),
        enable_startup_seed_cache=bool(args.enable_startup_seed_cache),
        refresh_startup_seed_cache=bool(args.refresh_startup_seed_cache),
        startup_seed_cache_path=args.startup_seed_cache_path,
        enable_primary_thermo_startup_prewarm=bool(args.enable_primary_thermo_startup_prewarm),
        enable_startup_thermo_conditioning=bool(args.enable_startup_thermo_conditioning),
        startup_thermo_conditioning_iters=int(args.startup_thermo_conditioning_iters),
        startup_thermo_conditioning_relaxation=float(args.startup_thermo_conditioning_relaxation),
        enable_restart_reentry_settling=bool(args.enable_restart_reentry_settling),
        enable_startup_hydraulic_energy_consistency=bool(args.enable_startup_hydraulic_energy_consistency),
        startup_hydraulic_energy_consistency_iters=int(args.startup_hydraulic_energy_consistency_iters),
        startup_hydraulic_energy_consistency_dt_sec=float(args.startup_hydraulic_energy_consistency_dt_sec),
        startup_hydraulic_energy_consistency_mass_tol_lbmolph=args.startup_hydraulic_energy_consistency_mass_tol_lbmolph,
        startup_hydraulic_energy_consistency_energy_tol_btups=args.startup_hydraulic_energy_consistency_energy_tol_btups,
        enable_startup_hydraulic_sequence=bool(args.enable_startup_hydraulic_sequence),
        startup_sequence_energy_on_sec=float(args.startup_sequence_energy_on_sec),
        startup_sequence_liquid_on_sec=float(args.startup_sequence_liquid_on_sec),
        startup_sequence_liquid_ramp_sec=float(args.startup_sequence_liquid_ramp_sec),
        startup_sequence_mass_resid_gate_lbmolph=args.startup_sequence_mass_resid_gate_lbmolph,
        startup_sequence_liquid_backoff_sec=args.startup_sequence_liquid_backoff_sec,
        enable_startup_vapor_homotopy=bool(args.enable_startup_vapor_homotopy),
        startup_sequence_profile_hold_sec=float(args.startup_sequence_profile_hold_sec),
        startup_sequence_vapor_on_sec=args.startup_sequence_vapor_on_sec,
        startup_sequence_vapor_ramp_sec=float(args.startup_sequence_vapor_ramp_sec),
        startup_sequence_vapor_rel_rate_gate_per_s=args.startup_sequence_vapor_rel_rate_gate_per_s,
        startup_sequence_vapor_backoff_sec=args.startup_sequence_vapor_backoff_sec,
        enable_steady_state_detection=bool(args.enable_steady_state_detection),
        steady_state_window_sec=float(args.steady_state_window_sec),
        steady_state_min_time_sec=float(args.steady_state_min_time_sec),
        steady_state_rel_state_rate_tol_per_s=args.steady_state_rel_state_rate_tol_per_s,
        steady_state_kpi_slope_tol_per_s=args.steady_state_kpi_slope_tol_per_s,
        steady_state_mv_rate_tol_per_s=args.steady_state_mv_rate_tol_per_s,
        steady_state_temp_rate_tol_F_per_s=args.steady_state_temp_rate_tol_F_per_s,
        steady_state_sp_error_tol=args.steady_state_sp_error_tol,
        steady_state_require_sp=bool(args.steady_state_require_sp),
        steady_state_rate_denom_floor_lbmol=float(args.steady_state_rate_denom_floor_lbmol),
    )

    out = run_smoke_simulation(cfg)
    try:
        fallback_n = int(out.get("integrator_fallback_count", 0))
    except Exception:
        fallback_n = 0
    if fallback_n > 0:
        print(f"[Info] Integrator fallback count (to explicit-euler): {fallback_n}")

    # Auto-register exact CLI command and refresh experiment ledger after each
    # run that produces log files.
    if out.get("summary_csv"):
        try:
            append_run_registry_entry(
                logs_dir=Path(str(out.get("logs_dir", cfg.logs_dir))),
                module_name=module_name,
                argv=raw_argv,
                summary_csv_path=out.get("summary_csv"),
                profile_csv_path=out.get("profile_csv"),
                run_name=cfg.run_name,
                run_description=cfg.run_description,
                metadata_json_path=out.get("run_metadata_json"),
            )
            rebuild_experiment_ledger(project_root=project_root)
        except Exception as exc:
            print(f"[Warn] Failed to update experiment ledger: {exc}")

    if out.get("profile_csv"):
        print(f"Wrote: {out['profile_csv']}")
    if out.get("summary_csv"):
        print(f"Wrote: {out['summary_csv']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
