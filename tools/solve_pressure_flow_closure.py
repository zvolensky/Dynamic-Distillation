#!/usr/bin/env python
"""
Solve a formal pressure/flow closure seed for a column workbook.

This tool wraps the existing DAE pilot residual z=[P_tray, V_out] and writes a
new workbook with the closed pressure and vapor-flow profile. It is intended as
a diagnostic initializer, not as a validation acceptance shortcut.
"""

from __future__ import annotations

import argparse
from dataclasses import replace
import json
from pathlib import Path
import shutil
import subprocess
import sys
from typing import Any, Dict, Optional

import numpy as np
from openpyxl import load_workbook
from scipy.optimize import least_squares, root

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_PATH = PROJECT_ROOT / "src"
if str(SRC_PATH) not in sys.path:
    sys.path.insert(0, str(SRC_PATH))

from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case  # noqa: E402
from dynamic_distillation.dae_pilot_v1 import (  # noqa: E402
    DaePilotLayout,
    default_algebraic_seed,
    evaluate_pilot_residual,
    inf_norm,
)
from dynamic_distillation.dynamic_run_scaffold_v1 import (  # noqa: E402
    RunnerConfig,
    _clear_initial_tray_vapor_holdup,
    _initialize_vapor_holdup_from_spec_pressure,
    _vapor_volume_ft3_per_stage,
    build_inputs_for_runner,
)
from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel  # noqa: E402
from dynamic_distillation.state_vector_layout_v1 import StateVectorLayout  # noqa: E402


def _resolve(raw: str | Path) -> Path:
    p = Path(str(raw))
    if p.is_absolute():
        return p.resolve()
    return (PROJECT_ROOT / p).resolve()


def _norm_label(value: Any) -> str:
    return "".join(ch for ch in str(value).strip().lower() if ch.isalnum())


def _find_header(ws: Any, header: str) -> tuple[int, int]:
    target = str(header).strip().lower()
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value).strip().lower() == target:
                return int(cell.row), int(cell.column)
    raise ValueError(f"Could not find header {header!r} in sheet {ws.title!r}")


def _update_spec_value(wb: Any, label: str, value: Any) -> None:
    if "Specifications" not in wb.sheetnames:
        return
    ws = wb["Specifications"]
    target = _norm_label(label)
    for r in range(1, ws.max_row + 1):
        if _norm_label(ws.cell(r, 1).value) == target:
            ws.cell(r, 2).value = value
            return
    row = ws.max_row + 1
    ws.cell(row, 1).value = label
    ws.cell(row, 2).value = value


def _build_case_runtime(args: argparse.Namespace) -> tuple[Any, StateVectorLayout, Any, np.ndarray]:
    excel_path = _resolve(args.input)
    case = load_case_from_excel(str(excel_path))
    col = build_column_spec_from_case(case)
    layout = StateVectorLayout(
        n_stages=int(col.n_stages),
        n_components=int(col.n_components),
        include_top=bool(args.include_boundary_states),
        include_bottom=bool(args.include_boundary_states),
        include_vapor=True,
        include_temperature=bool(args.include_temperature),
        include_energy=bool(args.include_energy),
    )
    cfg = RunnerConfig(
        excel_path=str(excel_path),
        runtime_mode=str(args.runtime_mode),
        include_temperature=bool(args.include_temperature),
        include_energy=bool(args.include_energy),
        include_boundary_states=bool(args.include_boundary_states),
        include_vapor_states=True,
        thermo_mode=str(args.thermo),
        clapeyron_model=str(args.clapeyron_model),
        condenser_duty_mode=str(args.condenser_duty_mode),
        condenser_duty_btu_per_h=args.condenser_duty_btuph,
        enable_equilibrium_relaxation=not bool(args.no_equilibrium),
        flash_feed_at_stage_conditions=(False if bool(args.no_flash_feed_at_stage_conditions) else None),
        vapor_holdup_relaxation_sec=float(args.vapor_holdup_relaxation_sec),
        write_logs=False,
    )
    inputs, provider = build_inputs_for_runner(case, col, cfg)
    inputs = replace(
        inputs,
        pressure_model=str(args.pressure_model),
        vapor_flow_model=str(args.vapor_flow_model),
        feed_stage_flash_prev=None,
        feed_stage_flash_reuse_dT_F=0.0,
        feed_stage_flash_reuse_dP_psia=0.0,
        feed_stage_flash_reuse_dx=0.0,
    )
    y0 = layout.pack_y0(col)
    if not bool(args.use_excel_vapor_holdup):
        y0 = _clear_initial_tray_vapor_holdup(y0, layout)
    y0 = _initialize_vapor_holdup_from_spec_pressure(
        col=col,
        layout=layout,
        y=y0,
        inputs=inputs,
        include_temperature=bool(args.include_temperature),
        preserve_tray_vapor_holdup=bool(args.use_excel_vapor_holdup),
    )
    # Keep provider alive by attaching it to inputs; return path mirrors other
    # tools and lets Python clean up after process exit.
    _ = provider
    return col, layout, inputs, np.asarray(y0, dtype=float)


def _write_pf_workbook(
    *,
    input_path: Path,
    output_path: Path,
    pressure_psia: np.ndarray,
    vapor_flow_lbmolph: np.ndarray,
    vapor_holdup_lbmol: Optional[np.ndarray],
    tray_ev_btu: Optional[np.ndarray],
    pressure_model: str,
    vapor_flow_model: str,
) -> None:
    if input_path.resolve() != output_path.resolve():
        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(input_path, output_path)
    wb = load_workbook(output_path)
    ws = wb["Initial Conditions"]
    stage_hdr_row, stage_col = _find_header(ws, "Stage")
    _row, pressure_col = _find_header(ws, "Pressure (psia)")
    _row, vapor_flow_col = _find_header(ws, "Vapor Flow (lbmol/h)")
    vapor_holdup_col = None
    if vapor_holdup_lbmol is not None:
        try:
            _row, vapor_holdup_col = _find_header(ws, "Vapor Holdup (lbmol)")
        except Exception:
            vapor_holdup_col = None
    stage_to_row: Dict[int, int] = {}
    for row in range(stage_hdr_row + 1, ws.max_row + 1):
        val = ws.cell(row, stage_col).value
        if val is None:
            continue
        try:
            stage_to_row[int(float(val))] = int(row)
        except Exception:
            pass
    for i in range(int(pressure_psia.size)):
        row = stage_to_row.get(i + 1)
        if row is None:
            continue
        ws.cell(row, pressure_col).value = float(pressure_psia[i])
        ws.cell(row, vapor_flow_col).value = float(vapor_flow_lbmolph[i])
        if vapor_holdup_col is not None and vapor_holdup_lbmol is not None:
            ws.cell(row, vapor_holdup_col).value = float(vapor_holdup_lbmol[i])
    _update_spec_value(wb, "Pressure Model", str(pressure_model))
    _update_spec_value(wb, "Vapor Flow Model", str(vapor_flow_model))
    if tray_ev_btu is not None and "Energy State" in wb.sheetnames:
        ws_energy = wb["Energy State"]
        try:
            energy_stage_hdr_row, energy_stage_col = _find_header(ws_energy, "Stage")
            _row, tray_ev_col = _find_header(ws_energy, "Tray EV (BTU)")
            energy_stage_to_row: Dict[int, int] = {}
            for row in range(energy_stage_hdr_row + 1, ws_energy.max_row + 1):
                val = ws_energy.cell(row, energy_stage_col).value
                if val is None:
                    continue
                try:
                    energy_stage_to_row[int(float(val))] = int(row)
                except Exception:
                    pass
            for i in range(int(tray_ev_btu.size)):
                row = energy_stage_to_row.get(i + 1)
                if row is not None:
                    ws_energy.cell(row, tray_ev_col).value = float(tray_ev_btu[i])
        except Exception:
            pass
    wb.save(output_path)


def _sync_vapor_state_from_pressure(
    *,
    col: Any,
    layout: StateVectorLayout,
    inputs: Any,
    y_state: np.ndarray,
    pressure_psia: np.ndarray,
    diag: Dict[str, Any],
) -> tuple[np.ndarray, Optional[np.ndarray], Dict[str, float]]:
    n = int(col.n_stages)
    nc = int(col.n_components)
    u = layout.unpack(y_state)
    if "tray_T_f" in u:
        tray_t_f = np.asarray(u["tray_T_f"], dtype=float).reshape((n,))
    else:
        tray_t_f = np.asarray(getattr(col, "T_f", np.full(n, 100.0)), dtype=float).reshape((n,))
    tray_t_r = np.where(np.isfinite(tray_t_f + 459.67) & ((tray_t_f + 459.67) > 1.0e-6), tray_t_f + 459.67, 559.67)

    vapor_volume = _vapor_volume_ft3_per_stage(inputs.volume_model, n)
    z = np.asarray(diag.get("Z_tray", np.ones(n, dtype=float)), dtype=float).reshape((n,))
    z = np.where(np.isfinite(z) & (z > 0.0), z, 1.0)
    p = np.asarray(pressure_psia, dtype=float).reshape((n,))
    p = np.where(np.isfinite(p) & (p > 0.0), p, np.asarray(getattr(col, "P_psia", np.full(n, 200.0)), dtype=float).reshape((n,)))

    gas_r = 10.7316  # psia*ft3/(lbmol*R)
    mv = p * vapor_volume / np.maximum(z * gas_r * tray_t_r, 1.0e-300)
    mv = np.where(np.isfinite(mv) & (mv >= 0.0), mv, 0.0)
    # Preserve the total-condenser convention used by startup vapor seeding.
    if n > 0:
        mv[0] = 0.0

    ev = None
    if bool(getattr(layout, "include_energy", False)):
        hv = None
        if "HV_BTU_lbmol_tray" in diag:
            try:
                hv = np.asarray(diag["HV_BTU_lbmol_tray"], dtype=float).reshape((n,))
            except Exception:
                hv = None
        if hv is None:
            try:
                tray_ev_old = np.asarray(u.get("tray_EV_BTU", np.zeros(n)), dtype=float).reshape((n,))
                mv_old = np.sum(np.asarray(u.get("tray_V", np.zeros((n, nc))), dtype=float).reshape((n, nc)), axis=1)
                hv = tray_ev_old / np.maximum(mv_old, float(layout.epsilon_lbmol))
            except Exception:
                hv = np.zeros(n, dtype=float)
        hv = np.where(np.isfinite(hv), hv, 0.0)
        ev = mv * hv
        ev = np.where(np.isfinite(ev), ev, 0.0)
        ev[mv <= float(layout.epsilon_lbmol)] = 0.0

    old_mv = np.sum(np.asarray(u.get("tray_V", np.zeros((n, nc))), dtype=float).reshape((n, nc)), axis=1)
    stats = {
        "max_abs_delta_vapor_holdup_lbmol": float(np.max(np.abs(mv - old_mv))),
        "max_rel_delta_vapor_holdup": float(np.max(np.abs(mv - old_mv) / np.maximum(np.abs(old_mv), 1.0))),
        "min_synced_vapor_holdup_lbmol": float(np.min(mv)),
        "max_synced_vapor_holdup_lbmol": float(np.max(mv)),
    }
    return mv, ev, stats


def main() -> int:
    ap = argparse.ArgumentParser(description="Solve t=0 pressure/flow algebraic closure for a workbook.")
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--solver", choices=["least-squares", "root"], default="least-squares")
    ap.add_argument("--max-nfev", type=int, default=20)
    ap.add_argument("--pressure-scale-psia", type=float, default=10.0)
    ap.add_argument("--vapor-flow-scale-lbmolph", type=float, default=1000.0)
    ap.add_argument("--max-pressure-delta-psia", type=float, default=25.0)
    ap.add_argument("--max-vapor-log-delta", type=float, default=0.25)
    ap.add_argument(
        "--no-sync-vapor-state",
        dest="sync_vapor_state",
        action="store_false",
        help="Do not back-calculate Vapor Holdup and Tray EV from the solved pressure profile.",
    )
    ap.set_defaults(sync_vapor_state=True)
    ap.add_argument("--pressure-model", default="hydraulic")
    ap.add_argument("--vapor-flow-model", default="energy")
    ap.add_argument("--runtime-mode", default="hydraulic")
    ap.add_argument("--thermo", default="clapeyron")
    ap.add_argument("--clapeyron-model", default="PR")
    ap.add_argument("--condenser-duty-mode", default="total-condense")
    ap.add_argument("--condenser-duty-btuph", type=float, default=None)
    ap.add_argument("--no-equilibrium", action="store_true")
    ap.add_argument("--no-flash-feed-at-stage-conditions", action="store_true")
    ap.add_argument("--vapor-holdup-relaxation-sec", type=float, default=0.0)
    ap.add_argument("--use-excel-vapor-holdup", action="store_true")
    ap.add_argument("--no-temperature", dest="include_temperature", action="store_false")
    ap.add_argument("--include-energy", dest="include_energy", action="store_true")
    ap.set_defaults(include_temperature=True, include_energy=False)
    ap.add_argument("--disable-boundary-states", dest="include_boundary_states", action="store_false")
    ap.set_defaults(include_boundary_states=True)
    ap.add_argument("--audit-output-dir", default=None)
    args = ap.parse_args()

    input_path = _resolve(args.input)
    output_path = _resolve(args.output)
    col, layout, inputs, y0 = _build_case_runtime(args)
    n = int(col.n_stages)
    z_layout = DaePilotLayout(n_stages=n)

    p0 = np.asarray(getattr(col, "P_psia", np.full(n, 200.0)), dtype=float).reshape((n,))
    v0 = np.asarray(getattr(col, "V_lbmolph", np.zeros(n)), dtype=float).reshape((n,))
    z0 = default_algebraic_seed(n_stages=n, p_fallback_psia=p0, v_fallback_lbmolph=v0)
    p_seed, v_seed = z_layout.split(z0)
    v_seed = np.where(np.isfinite(v_seed) & (v_seed >= 0.0), v_seed, 0.0)
    z0 = z_layout.join(p_seed, v_seed)

    p_scale = max(abs(float(args.pressure_scale_psia)), 1.0e-12)
    v_scale = max(abs(float(args.vapor_flow_scale_lbmolph)), 1.0e-12)

    def encode(z_abs: np.ndarray) -> np.ndarray:
        p, v = z_layout.split(z_abs)
        return np.concatenate([(p - p_seed) / p_scale, np.log(np.maximum(v, 1.0e-9) / np.maximum(v_seed, 1.0e-9))])

    def decode(x: np.ndarray) -> np.ndarray:
        x_arr = np.asarray(x, dtype=float).reshape((2 * n,))
        p = p_seed + p_scale * x_arr[:n]
        v = np.maximum(v_seed, 1.0e-9) * np.exp(np.clip(x_arr[n:], -20.0, 20.0))
        v[0] = 0.0
        v[-1] = v_seed[-1]
        return z_layout.join(p, v)

    def residual_scaled(x: np.ndarray) -> np.ndarray:
        z_abs = decode(x)
        rr = evaluate_pilot_residual(
            t_s=0.0,
            y=y0,
            ydot=np.zeros_like(y0),
            z=z_abs,
            col=col,
            layout=layout,
            inputs=inputs,
        )
        r_p = np.asarray(rr.alg_pressure, dtype=float).reshape((n,)) / p_scale
        r_v = np.asarray(rr.alg_vapor, dtype=float).reshape((n,)) / v_scale
        return np.concatenate([r_p, r_v], axis=0)

    x0 = encode(z0)
    if str(args.solver) == "root":
        result = root(residual_scaled, x0, method="hybr", options={"maxfev": max(int(args.max_nfev), 1)})
        x_best = np.asarray(result.x, dtype=float)
        success = bool(result.success)
        message = str(result.message)
        eval_count = int(getattr(result, "nfev", -1))
    else:
        p_delta = abs(float(args.max_pressure_delta_psia)) / p_scale
        v_delta = abs(float(args.max_vapor_log_delta))
        lb = np.concatenate([np.full(n, -p_delta), np.full(n, -v_delta)], axis=0)
        ub = np.concatenate([np.full(n, p_delta), np.full(n, v_delta)], axis=0)
        # Boundary vapor flows are fixed by topology/specs. SciPy requires
        # strict inequality, so use a numerically tiny box around zero.
        fixed_eps = 1.0e-15
        lb[n] = -fixed_eps
        ub[n] = fixed_eps
        lb[(2 * n) - 1] = -fixed_eps
        ub[(2 * n) - 1] = fixed_eps
        result = least_squares(
            residual_scaled,
            x0,
            bounds=(lb, ub),
            max_nfev=max(int(args.max_nfev), 1),
            x_scale="jac",
            verbose=1,
        )
        x_best = np.asarray(result.x, dtype=float)
        success = bool(result.success)
        message = str(result.message)
        eval_count = int(getattr(result, "nfev", -1))

    z_best = decode(x_best)
    p_best, v_best = z_layout.split(z_best)
    rr_best = evaluate_pilot_residual(
        t_s=0.0,
        y=y0,
        ydot=np.zeros_like(y0),
        z=z_best,
        col=col,
        layout=layout,
        inputs=inputs,
    )
    rr_sync = rr_best
    vapor_holdup_sync = None
    tray_ev_sync = None
    vapor_sync_stats: Dict[str, float] = {}
    if bool(args.sync_vapor_state):
        sync_inputs = replace(inputs, compute_thermo_diag=True)
        rr_sync = evaluate_pilot_residual(
            t_s=0.0,
            y=y0,
            ydot=np.zeros_like(y0),
            z=z_best,
            col=col,
            layout=layout,
            inputs=sync_inputs,
        )
        vapor_holdup_sync, tray_ev_sync, vapor_sync_stats = _sync_vapor_state_from_pressure(
            col=col,
            layout=layout,
            inputs=inputs,
            y_state=y0,
            pressure_psia=p_best,
            diag=rr_sync.diag,
        )

    _write_pf_workbook(
        input_path=input_path,
        output_path=output_path,
        pressure_psia=p_best,
        vapor_flow_lbmolph=v_best,
        vapor_holdup_lbmol=vapor_holdup_sync,
        tray_ev_btu=tray_ev_sync,
        pressure_model=str(args.pressure_model),
        vapor_flow_model=str(args.vapor_flow_model),
    )

    audit_dir = args.audit_output_dir
    if audit_dir is None:
        audit_dir = str(output_path.with_suffix("")) + "_pf_audit"
    audit_dir_path = _resolve(audit_dir)
    audit_cmd = [
        sys.executable,
        str(PROJECT_ROOT / "tools" / "column_initialization_residual_audit.py"),
        "--excel",
        str(output_path),
        "--thermo",
        str(args.thermo),
        "--clapeyron-model",
        str(args.clapeyron_model),
        "--runtime-mode",
        str(args.runtime_mode),
        "--condenser-duty-mode",
        str(args.condenser_duty_mode),
        "--vapor-holdup-relaxation-sec",
        str(float(args.vapor_holdup_relaxation_sec)),
        "--output-dir",
        str(audit_dir_path),
    ]
    if args.condenser_duty_btuph is not None:
        audit_cmd.extend(["--condenser-duty-btuph", str(float(args.condenser_duty_btuph))])
    if bool(args.no_equilibrium):
        audit_cmd.append("--no-equilibrium")
    if bool(args.no_flash_feed_at_stage_conditions):
        audit_cmd.append("--no-flash-feed-at-stage-conditions")
    if bool(args.include_energy):
        audit_cmd.append("--include-energy")
    if not bool(args.include_temperature):
        audit_cmd.append("--no-temperature")
    if bool(args.use_excel_vapor_holdup):
        audit_cmd.append("--use-excel-vapor-holdup")
    subprocess.run(audit_cmd, cwd=str(PROJECT_ROOT), check=True)

    summary = {
        "input": str(input_path),
        "output": str(output_path),
        "solver": str(args.solver),
        "success": success,
        "message": message,
        "eval_count": eval_count,
        "pressure_model": str(args.pressure_model),
        "vapor_flow_model": str(args.vapor_flow_model),
        "alg_pressure_inf_psia": inf_norm(rr_best.alg_pressure),
        "alg_vapor_inf_lbmolph": inf_norm(rr_best.alg_vapor),
        "alg_full_inf_scaled": inf_norm(residual_scaled(x_best)),
        "sync_vapor_state": bool(args.sync_vapor_state),
        **vapor_sync_stats,
        "max_abs_delta_pressure_psia": float(np.max(np.abs(p_best - p_seed))),
        "max_rel_delta_vapor_flow": float(np.max(np.abs(v_best - v_seed) / np.maximum(np.abs(v_seed), 1.0))),
        "audit_dir": str(audit_dir_path),
    }
    summary_path = output_path.with_suffix(".pf_closure_summary.json")
    summary_path.write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")

    print(f"Wrote: {output_path}")
    print(f"Summary: {summary_path}")
    print(f"Audit: {audit_dir_path}")
    print(
        "PF residuals: "
        f"P_inf={summary['alg_pressure_inf_psia']:.6g} psia, "
        f"V_inf={summary['alg_vapor_inf_lbmolph']:.6g} lbmol/h"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
