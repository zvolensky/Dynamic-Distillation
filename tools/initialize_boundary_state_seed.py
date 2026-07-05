from __future__ import annotations

import argparse
from math import pi
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import load_workbook


R_PSIA_FT3_PER_LBMOL_R = 10.731577089016


def _norm_label(value: Any) -> str:
    return " ".join(str(value or "").strip().split()).lower()


def _norm_token(value: Any) -> str:
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())


def _spec_value(ws, label: str, default: float | str | None = None) -> Any:
    target = _norm_label(label)
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        if _norm_label(row[0]) != target:
            continue
        for value in reversed(row[1:]):
            if value is None:
                continue
            if isinstance(value, str) and not value.strip():
                continue
            return value
    return default


def _float_spec(ws, label: str, default: float | None = None) -> float | None:
    value = _spec_value(ws, label, default)
    if value is None:
        return default
    try:
        return float(value)
    except Exception:
        return default


def _component_names(ws) -> list[str]:
    n_comp = int(float(_spec_value(ws, "Number of Components")))
    for row in ws.iter_rows(values_only=True):
        if row and _norm_label(row[0]) == "component name":
            names = [str(v).strip() for v in row[1 : 1 + n_comp] if v is not None and str(v).strip()]
            if len(names) == n_comp:
                return names
    return [f"Component {i}" for i in range(1, n_comp + 1)]


def _stream_column(ws, stream_name: str) -> int:
    target = _norm_token(stream_name)
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 3)):
        for cell in row:
            if _norm_token(cell.value) == target:
                return int(cell.column)
    raise ValueError(f"Could not find stream column {stream_name!r} on Streams sheet.")


def _stream_total_and_mole_flows(ws, stream_name: str, components: list[str]) -> tuple[float, np.ndarray]:
    col = _stream_column(ws, stream_name)
    total = None
    flows = np.full((len(components),), np.nan, dtype=float)
    comp_rows = {_norm_token(name): i for i, name in enumerate(components)}
    for row in ws.iter_rows():
        label = row[0].value if row else None
        label_norm = _norm_label(label)
        token = _norm_token(label)
        if "total molar flow" in label_norm:
            total = float(row[col - 1].value)
        elif token in comp_rows:
            value = row[col - 1].value
            flows[comp_rows[token]] = float(value) if value is not None else np.nan
    if total is None or not np.isfinite(total) or total <= 0.0:
        if np.all(np.isfinite(flows)):
            total = float(np.sum(flows))
        else:
            raise ValueError(f"Could not read total molar flow for stream {stream_name!r}.")
    if not np.all(np.isfinite(flows)) or float(np.sum(flows)) <= 0.0:
        raise ValueError(f"Could not read component mole flows for stream {stream_name!r}.")
    return float(total), flows


def _stream_row_float(ws, stream_name: str, row_label: str) -> float:
    col = _stream_column(ws, stream_name)
    target = _norm_label(row_label)
    for row in ws.iter_rows():
        label = row[0].value if row else None
        if _norm_label(label) == target:
            return float(row[col - 1].value)
    raise ValueError(f"Could not read {row_label!r} for stream {stream_name!r}.")


def _normalized(values: np.ndarray, fallback: np.ndarray | None = None) -> np.ndarray:
    arr = np.clip(np.asarray(values, dtype=float), 0.0, None)
    total = float(np.sum(arr))
    if total > 0.0 and np.isfinite(total):
        return arr / total
    if fallback is not None:
        fb = np.clip(np.asarray(fallback, dtype=float), 0.0, None)
        fb_total = float(np.sum(fb))
        if fb_total > 0.0 and np.isfinite(fb_total):
            return fb / fb_total
    raise ValueError("Cannot normalize a non-positive composition vector.")


def _initial_condition_composition(ws, stage: int, phase: str, n_comp: int) -> np.ndarray:
    headers = [str(c.value or "").strip() for c in ws[1]]
    stage_col = headers.index("Stage") + 1
    if phase.lower().startswith("vap"):
        prefix = "Vapor Composition Component "
    else:
        prefix = "Liquid Composition Component "
    comp_cols = [headers.index(f"{prefix}{i}") + 1 for i in range(1, n_comp + 1)]
    for row in ws.iter_rows(min_row=2):
        if int(float(row[stage_col - 1].value)) == int(stage):
            return _normalized(np.array([row[c - 1].value for c in comp_cols], dtype=float))
    raise ValueError(f"Could not find stage {stage} on Initial Conditions sheet.")


def _top_vapor_volume_ft3(ws, override: float | None) -> float:
    if override is not None:
        return float(override)
    explicit = _float_spec(ws, "Top Drum Vapor Volume (ft3)")
    if explicit is not None and explicit > 0.0:
        return explicit
    diameter = _float_spec(ws, "Distillate Drum Diameter (ft)")
    length = _float_spec(ws, "Distillate Drum Length (ft)")
    if diameter is not None and length is not None and diameter > 0.0 and length > 0.0:
        frac = _float_spec(ws, "Top Drum Vapor Fraction (-)", None)
        if frac is None:
            liquid_frac = _float_spec(ws, "Top Level SP Frac", 0.5) or 0.5
            frac = max(0.0, min(1.0, 1.0 - liquid_frac))
        return pi / 4.0 * diameter * diameter * length * frac
    line_vol = _float_spec(ws, "Overhead Vapor Line Volume (ft3)", 0.0) or 0.0
    condenser_vol = _float_spec(ws, "Condenser Vapor Volume (ft3)", 0.0) or 0.0
    total = line_vol + condenser_vol
    if total > 0.0:
        return total
    raise ValueError("Could not determine top vapor volume; pass --top-vapor-volume-ft3.")


def _write_boundary_state(
    wb,
    components: list[str],
    top_l: np.ndarray,
    top_v: np.ndarray,
    bottom_l: np.ndarray,
    bottom_v: np.ndarray,
) -> None:
    if "Boundary State" in wb.sheetnames:
        del wb["Boundary State"]
    ws = wb.create_sheet("Boundary State")
    rows = [
        ["State", *components],
        ["top_L", *map(float, top_l)],
        ["top_V", *map(float, top_v)],
        ["bottom_L", *map(float, bottom_l)],
        ["bottom_V", *map(float, bottom_v)],
    ]
    for r, row in enumerate(rows, start=1):
        for c, value in enumerate(row, start=1):
            ws.cell(r, c).value = value


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Add a Boundary State sheet seeded from product streams and target drum vapor pressure."
    )
    parser.add_argument("--input", required=True, help="Input Excel workbook.")
    parser.add_argument("--output", required=True, help="Output Excel workbook.")
    parser.add_argument("--top-pressure-psia", type=float, default=None)
    parser.add_argument("--top-temperature-f", type=float, default=None)
    parser.add_argument("--top-vapor-volume-ft3", type=float, default=None)
    parser.add_argument("--top-vapor-z", type=float, default=1.0)
    parser.add_argument("--bottom-vapor-total-lbmol", type=float, default=1.0e-8)
    args = parser.parse_args()

    in_path = Path(args.input)
    out_path = Path(args.output)
    wb = load_workbook(in_path)
    if "Specifications" not in wb.sheetnames or "Streams" not in wb.sheetnames or "Initial Conditions" not in wb.sheetnames:
        raise ValueError("Workbook must contain Specifications, Streams, and Initial Conditions sheets.")

    specs = wb["Specifications"]
    streams = wb["Streams"]
    initials = wb["Initial Conditions"]
    components = _component_names(specs)
    n_comp = len(components)

    _, dist_flows = _stream_total_and_mole_flows(streams, "Distillate", components)
    _, bottom_flows = _stream_total_and_mole_flows(streams, "Bottom", components)
    top_x = _normalized(dist_flows)
    bottom_x = _normalized(bottom_flows)
    top_y = _initial_condition_composition(initials, stage=1, phase="vapor", n_comp=n_comp)
    bottom_y = _initial_condition_composition(initials, stage=int(float(_spec_value(specs, "Number of Stages"))), phase="vapor", n_comp=n_comp)

    top_l_total = _float_spec(specs, "Top Accumulator Holdup (lbmol)")
    bottom_l_total = _float_spec(specs, "Bottom Holdup (lbmol)")
    if top_l_total is None or top_l_total <= 0.0:
        raise ValueError("Missing positive Top Accumulator Holdup (lbmol).")
    if bottom_l_total is None or bottom_l_total <= 0.0:
        raise ValueError("Missing positive Bottom Holdup (lbmol).")

    top_pressure = args.top_pressure_psia
    if top_pressure is None:
        top_pressure = _stream_row_float(streams, "Distillate", "Pressure (psia)")
    top_temp_f = args.top_temperature_f
    if top_temp_f is None:
        top_temp_f = _stream_row_float(streams, "Distillate", "Temperature (F)")
    top_volume = _top_vapor_volume_ft3(specs, args.top_vapor_volume_ft3)
    top_v_total = float(top_pressure) * float(top_volume) / (
        max(float(args.top_vapor_z), 1.0e-12) * R_PSIA_FT3_PER_LBMOL_R * (float(top_temp_f) + 459.67)
    )

    top_l = float(top_l_total) * top_x
    top_v = top_v_total * top_y
    bottom_l = float(bottom_l_total) * bottom_x
    bottom_v = max(float(args.bottom_vapor_total_lbmol), 0.0) * bottom_y
    _write_boundary_state(wb, components, top_l, top_v, bottom_l, bottom_v)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Wrote {out_path}")
    print(f"top_L total lbmol: {float(np.sum(top_l)):.9g}")
    print(f"top_V total lbmol: {float(np.sum(top_v)):.9g}")
    print(f"top vapor pressure psia: {float(top_pressure):.9g}")
    print(f"top vapor temperature F: {float(top_temp_f):.9g}")
    print(f"top vapor volume ft3: {float(top_volume):.9g}")
    print(f"bottom_L total lbmol: {float(np.sum(bottom_l)):.9g}")
    print(f"bottom_V total lbmol: {float(np.sum(bottom_v)):.9g}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
