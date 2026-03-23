from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Dict, Iterable

import pandas as pd
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = REPO_ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from dynamic_distillation.compound_registry_v1 import canonicalize_to_dwsim_id
from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel


def _norm_label(value: object) -> str:
    return "".join(ch for ch in str(value).strip().lower() if ch.isalnum())


def _find_spec_row(ws, labels: Iterable[str]) -> int | None:
    targets = {_norm_label(x) for x in labels}
    for row in range(1, ws.max_row + 1):
        key = _norm_label(ws.cell(row=row, column=1).value)
        if key in targets:
            return row
    return None


def _find_stream_column(ws, name: str) -> int | None:
    target = _norm_label(name)
    for col in range(1, ws.max_column + 1):
        key = _norm_label(ws.cell(row=1, column=col).value)
        if key == target:
            return col
    return None


def _find_stream_row(ws, label: str) -> int | None:
    target = _norm_label(label)
    for row in range(1, ws.max_row + 1):
        key = _norm_label(ws.cell(row=row, column=1).value)
        if key == target:
            return row
    return None


def _profile_component_column_map(columns: Iterable[str], prefix: str) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for col in columns:
        if not col.startswith(prefix):
            continue
        suffix = col[len(prefix) :]
        norm = _norm_label(suffix)
        out[norm] = col
        if norm.startswith("n") and len(norm) > 1:
            out[norm[1:]] = col
    return out


def _component_profile_columns(components_excel: list[str], profile_columns: Iterable[str], prefix: str) -> list[str]:
    available = _profile_component_column_map(profile_columns, prefix)
    out: list[str] = []
    for comp in components_excel:
        canon = canonicalize_to_dwsim_id(comp)
        key = _norm_label(canon)
        candidates = [key]
        if key.startswith("n") and len(key) > 1:
            candidates.append(key[1:])
        raw = _norm_label(comp)
        candidates.append(raw)
        if raw.startswith("n") and len(raw) > 1:
            candidates.append(raw[1:])
        chosen = None
        for cand in candidates:
            if cand in available:
                chosen = available[cand]
                break
        if chosen is None:
            raise KeyError(f"Could not map component '{comp}' / '{canon}' to profile columns with prefix '{prefix}'")
        out.append(chosen)
    return out


def export_restart_workbook(
    *,
    base_excel: Path,
    summary_csv: Path,
    profile_csv: Path,
    output_excel: Path,
) -> None:
    case = load_case_from_excel(str(base_excel))
    summary = pd.read_csv(summary_csv)
    if summary.empty:
        raise ValueError(f"Summary CSV is empty: {summary_csv}")
    profile = pd.read_csv(profile_csv)
    if profile.empty:
        raise ValueError(f"Profile CSV is empty: {profile_csv}")

    final_summary = summary.iloc[-1]
    final_time = float(profile["time_s"].max())
    final_profile = profile[profile["time_s"] == final_time].copy()
    stage_profile = final_profile[final_profile["node_type"] == "stage"].copy()
    stage_profile = stage_profile.sort_values("stage").reset_index(drop=True)
    if len(stage_profile) != len(case.initial_conditions):
        raise ValueError(
            f"Stage profile row count {len(stage_profile)} does not match workbook stages {len(case.initial_conditions)}"
        )

    x_cols = _component_profile_columns(case.components, stage_profile.columns, "x_n_")
    y_cols = _component_profile_columns(case.components, stage_profile.columns, "y_n_")
    base_ic = case.initial_conditions.copy()

    wb = load_workbook(base_excel)

    ws_ic = wb["Initial Conditions"]
    headers = {str(ws_ic.cell(row=1, column=col).value): col for col in range(1, ws_ic.max_column + 1)}
    if "Vapor Holdup (lbmol)" not in headers:
        new_col = ws_ic.max_column + 1
        ws_ic.cell(row=1, column=new_col).value = "Vapor Holdup (lbmol)"
        headers["Vapor Holdup (lbmol)"] = new_col

    required_headers = [
        "Stage",
        "Temperature (F)",
        "Pressure (psia)",
        "Vapor Flow (lbmol/h)",
        "Liquid Flow (lbmol/h)",
        "Liquid Holdup (lbmol)",
        "Vapor Holdup (lbmol)",
    ]
    for header in required_headers:
        if header not in headers:
            raise KeyError(f"Workbook Initial Conditions sheet missing expected header: {header}")

    for i, row in stage_profile.iterrows():
        excel_row = i + 2
        base_row = base_ic.iloc[i]
        ws_ic.cell(row=excel_row, column=headers["Stage"]).value = int(row["stage"])
        ws_ic.cell(row=excel_row, column=headers["Temperature (F)"]).value = float(row["T_F"])
        ws_ic.cell(row=excel_row, column=headers["Pressure (psia)"]).value = float(row["P_psia_hyd"])
        ws_ic.cell(row=excel_row, column=headers["Vapor Flow (lbmol/h)"]).value = float(row["V_out_lbmolph"])
        ws_ic.cell(row=excel_row, column=headers["Liquid Flow (lbmol/h)"]).value = float(row["L_out_used_lbmolph"])
        ws_ic.cell(row=excel_row, column=headers["Liquid Holdup (lbmol)"]).value = float(row["ML_lbmol"])
        ws_ic.cell(row=excel_row, column=headers["Vapor Holdup (lbmol)"]).value = float(row["MV_lbmol"])
        x_vals = [float(row[x_col]) for x_col in x_cols]
        if not pd.Series(x_vals).replace([pd.NA], 0.0).notna().all() or sum(x_vals) <= 0.0:
            x_vals = [
                float(base_row[f"Liquid Composition Component {comp_idx}"])
                for comp_idx in range(1, len(x_cols) + 1)
            ]
        y_vals = [float(row[y_col]) for y_col in y_cols]
        if not pd.Series(y_vals).replace([pd.NA], 0.0).notna().all() or sum(y_vals) <= 0.0:
            y_vals = [
                float(base_row[f"Vapor Composition Component {comp_idx}"])
                for comp_idx in range(1, len(y_cols) + 1)
            ]
        for comp_idx, x_val in enumerate(x_vals, start=1):
            ws_ic.cell(row=excel_row, column=headers[f"Liquid Composition Component {comp_idx}"]).value = x_val
        for comp_idx, y_val in enumerate(y_vals, start=1):
            ws_ic.cell(row=excel_row, column=headers[f"Vapor Composition Component {comp_idx}"]).value = y_val

    ws_specs = wb["Specifications"]
    top_row = _find_spec_row(ws_specs, ["Top Accumulator Holdup (lbmol)", "Top Drum Holdup (lbmol)"])
    bottom_row = _find_spec_row(ws_specs, ["Bottom Holdup (lbmol)", "Bottom Sump Holdup (lbmol)"])
    if top_row is not None:
        ws_specs.cell(row=top_row, column=2).value = float(final_summary["Distillate_L_lbmol"])
    if bottom_row is not None:
        ws_specs.cell(row=bottom_row, column=2).value = float(final_summary["Bottoms_L_lbmol"])

    if "Streams" in wb.sheetnames:
        ws_streams = wb["Streams"]
        dist_col = _find_stream_column(ws_streams, "Distillate")
        bot_col = _find_stream_column(ws_streams, "Bottom")
        if dist_col is not None and bot_col is not None:
            row_stage = _find_stream_row(ws_streams, "Stage")
            row_press = _find_stream_row(ws_streams, "Pressure (psia)")
            row_temp = _find_stream_row(ws_streams, "Temperature (F)")
            row_total = _find_stream_row(ws_streams, "Total molar flow (lbmol/h)")
            if row_stage is not None:
                ws_streams.cell(row=row_stage, column=dist_col).value = 1
                ws_streams.cell(row=row_stage, column=bot_col).value = int(stage_profile["stage"].max())
            if row_press is not None:
                ws_streams.cell(row=row_press, column=dist_col).value = float(final_summary["P_top_drum_psia"])
                ws_streams.cell(row=row_press, column=bot_col).value = float(final_summary["P_bot_psia"])
            if row_temp is not None:
                ws_streams.cell(row=row_temp, column=dist_col).value = float(final_summary["T_Distillate_F"])
                ws_streams.cell(row=row_temp, column=bot_col).value = float(final_summary["T_sump_F"])
            if row_total is not None:
                ws_streams.cell(row=row_total, column=dist_col).value = float(final_summary["D_lbmolph"])
                ws_streams.cell(row=row_total, column=bot_col).value = float(final_summary["B_lbmolph"])

            dist_total = float(final_summary["D_lbmolph"])
            bot_total = float(final_summary["B_lbmolph"])
            for comp_name, dist_frac_col, bot_frac_col in zip(case.components, x_cols, x_cols):
                row_comp = _find_stream_row(ws_streams, comp_name)
                if row_comp is None:
                    continue
                ws_streams.cell(row=row_comp, column=dist_col).value = dist_total * float(final_summary[dist_frac_col.replace("x_n_", "Distillate_x_n_")])
                ws_streams.cell(row=row_comp, column=bot_col).value = bot_total * float(final_summary[bot_frac_col.replace("x_n_", "Bottoms_x_n_")])

    output_excel.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_excel)


def main() -> None:
    parser = argparse.ArgumentParser(description="Create a restart workbook from final summary/profile CSVs.")
    parser.add_argument("--base-excel", required=True, help="Template workbook to copy and update.")
    parser.add_argument("--summary-csv", required=True, help="Run summary CSV path.")
    parser.add_argument("--profile-csv", required=True, help="Run profile CSV path.")
    parser.add_argument("--output-excel", required=True, help="Output workbook path.")
    args = parser.parse_args()

    export_restart_workbook(
        base_excel=Path(args.base_excel),
        summary_csv=Path(args.summary_csv),
        profile_csv=Path(args.profile_csv),
        output_excel=Path(args.output_excel),
    )
    print(args.output_excel)


if __name__ == "__main__":
    main()
