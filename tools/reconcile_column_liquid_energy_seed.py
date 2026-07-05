#!/usr/bin/env python
"""
Bounded liquid-holdup / temperature initializer for active liquid hydraulics.

This tool evaluates the live column RHS at t=0 with active liquid hydraulics,
varies selected tray liquid holdup scale factors and tray temperature deltas,
and writes a new Excel seed workbook.  It is intentionally narrower than the
general composition initializer: the first objective targets total liquid
material and liquid-energy closure only.
"""

from __future__ import annotations

import argparse
from dataclasses import replace
import json
from pathlib import Path
import shutil
import sys
import time
from typing import Any, Dict, List, Optional

import numpy as np
from openpyxl import load_workbook
from scipy.optimize import least_squares
from scipy.sparse import lil_matrix

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_PATH = PROJECT_ROOT / "src"
if str(SRC_PATH) not in sys.path:
    sys.path.insert(0, str(SRC_PATH))

from dynamic_distillation.column_rhs_v1 import BoundaryFlows, column_rhs  # noqa: E402
from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case  # noqa: E402
from dynamic_distillation.dynamic_run_scaffold_v1 import (  # noqa: E402
    RunnerConfig,
    _clear_initial_tray_vapor_holdup,
    _initialize_vapor_holdup_from_spec_pressure,
    _mapping_scalar,
    _pack_top_drum_vapor_to_pressure,
    build_inputs_for_runner,
)
from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel  # noqa: E402
from dynamic_distillation.state_vector_layout_v1 import StateVectorLayout  # noqa: E402


def _resolve(raw: str | Path) -> Path:
    p = Path(str(raw))
    if p.is_absolute():
        return p.resolve()
    return (PROJECT_ROOT / p).resolve()


def _selected_stages(raw: str, n: int) -> List[int]:
    key = str(raw).strip().lower()
    if key in ("all", "*"):
        return list(range(n))
    out: List[int] = []
    for part in key.split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            a, b = [int(x) for x in part.split("-", 1)]
            out.extend(range(a - 1, b))
        else:
            out.append(int(part) - 1)
    return sorted({i for i in out if 0 <= i < n})


def _find_header(ws: Any, header: str) -> tuple[int, int]:
    target = str(header).strip().lower()
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value).strip().lower() == target:
                return int(cell.row), int(cell.column)
    raise ValueError(f"Could not find header {header!r} in sheet {ws.title!r}")


def _stage_row(ws: Any, stage: int) -> int:
    row0, col_stage = _find_header(ws, "Stage")
    for r in range(row0 + 1, ws.max_row + 1):
        try:
            if int(ws.cell(r, col_stage).value) == int(stage):
                return int(r)
        except Exception:
            continue
    raise ValueError(f"Could not find stage {stage} in sheet {ws.title!r}")


def _write_spec_value(ws: Any, label: str, value: float) -> bool:
    target = str(label).strip().lower()
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value).strip().lower() == target:
                ws.cell(int(cell.row), int(cell.column) + 1).value = float(value)
                return True
    return False


def _write_boundary_vector(ws: Any, state_name: str, values: np.ndarray) -> bool:
    try:
        row0, col_state = _find_header(ws, "State")
    except Exception:
        row0, col_state = 1, 1
    target = str(state_name).strip().lower()
    state_row = None
    for r in range(row0 + 1, ws.max_row + 1):
        if str(ws.cell(r, col_state).value).strip().lower() == target:
            state_row = int(r)
            break
    if state_row is None:
        state_row = int(ws.max_row) + 1
        ws.cell(state_row, col_state).value = str(state_name)
    vals = np.asarray(values, dtype=float).reshape((-1,))
    for j, val in enumerate(vals, start=1):
        ws.cell(state_row, col_state + j).value = float(val)
    return True


def _stream_total(col: Any, key: str) -> float:
    def norm(s: Any) -> str:
        return "".join(ch for ch in str(s).strip().lower() if ch.isalnum())

    streams = getattr(col, "streams", {}) or {}
    stream = None
    for name, candidate in streams.items():
        n = norm(name)
        if key == "distillate" and ("distillate" in n or n.startswith("top")):
            stream = candidate
            break
        if key == "bottoms" and "bottom" in n:
            stream = candidate
            break
    if stream is None:
        return 0.0
    total = getattr(stream, "total_molar_flow_lbmolph", None)
    if total is not None:
        try:
            val = float(total)
            if np.isfinite(val):
                return val
        except Exception:
            pass
    flows = getattr(stream, "component_molar_flows_lbmolph", None)
    if isinstance(flows, dict):
        return float(sum(float(v or 0.0) for v in flows.values()))
    return 0.0


def _build_runtime(args: argparse.Namespace) -> tuple[Any, StateVectorLayout, Any, np.ndarray]:
    excel_path = _resolve(args.input)
    case = load_case_from_excel(str(excel_path))
    col = build_column_spec_from_case(case)
    layout = StateVectorLayout(
        n_stages=int(col.n_stages),
        n_components=int(col.n_components),
        include_top=bool(args.include_boundary_states),
        include_bottom=bool(args.include_boundary_states),
        include_vapor=True,
        include_temperature=True,
        include_energy=True,
    )
    cfg = RunnerConfig(
        excel_path=str(excel_path),
        runtime_mode=str(args.runtime_mode),
        include_temperature=True,
        include_energy=True,
        include_boundary_states=bool(args.include_boundary_states),
        include_vapor_states=True,
        thermo_mode=str(args.thermo),
        clapeyron_model=str(args.clapeyron_model),
        thermo_table_path=str(args.thermo_table),
        condenser_duty_mode=str(args.condenser_duty_mode),
        condenser_duty_btu_per_h=args.condenser_duty_btuph,
        enable_equilibrium_relaxation=not bool(args.no_equilibrium),
        equilibrium_relaxation_mode=str(args.equilibrium_relaxation_mode),
        equilibrium_tau_sec=args.equilibrium_tau_sec,
        flash_feed_at_stage_conditions=(False if bool(args.no_flash_feed_at_stage_conditions) else None),
        vapor_holdup_relaxation_sec=float(args.vapor_holdup_relaxation_sec),
        write_logs=False,
    )
    inputs, _provider = build_inputs_for_runner(case, col, cfg)
    inputs = replace(
        inputs,
        feed_stage_flash_prev=None,
        feed_stage_flash_reuse_dT_F=0.0,
        feed_stage_flash_reuse_dP_psia=0.0,
        feed_stage_flash_reuse_dx=0.0,
    )
    if args.reflux_lbmolph is not None:
        b0 = inputs.boundary
        inputs = replace(
            inputs,
            boundary=BoundaryFlows(
                reflux_lbmolph=float(args.reflux_lbmolph),
                boilup_lbmolph=float(getattr(b0, "boilup_lbmolph", 0.0) or 0.0),
                distillate_lbmolph=float(_stream_total(col, "distillate")),
                bottoms_lbmolph=float(_stream_total(col, "bottoms")),
            ),
        )

    y0 = layout.pack_y0(col)
    if not bool(args.use_excel_vapor_holdup):
        y0 = _clear_initial_tray_vapor_holdup(y0, layout)
    y0 = _initialize_vapor_holdup_from_spec_pressure(
        col=col,
        layout=layout,
        y=y0,
        inputs=inputs,
        include_temperature=True,
        preserve_tray_vapor_holdup=bool(args.use_excel_vapor_holdup),
    )
    if bool(args.init_pack_top_drum_vapor_to_pressure):
        p_target = args.init_top_drum_vapor_pressure_psia
        if p_target is None:
            p_target = args.top_pressure_sp_psia
        y0, info = _pack_top_drum_vapor_to_pressure(
            y=np.asarray(y0, dtype=float),
            col=col,
            layout=layout,
            inputs=inputs,
            target_pressure_psia=p_target,
        )
        print(
            "top_vapor_pack "
            f"applied={bool(info.get('applied', False))} "
            f"reason={info.get('reason', '')} "
            f"P={info.get('raw_pressure_initial_psia', np.nan):.6g}->{info.get('raw_pressure_final_psia', np.nan):.6g}"
        )
    if bool(args.init_match_condenser_duty):
        p_match = args.top_pressure_sp_psia
        if p_match is None:
            p_match = float(np.asarray(getattr(col, "P_psia"), dtype=float).reshape((-1,))[0])
        match_inputs = replace(
            inputs,
            condenser_duty_mode="total-condense",
            condenser_duty_btu_per_h=None,
            condenser_duty_trim_btu_per_h=None,
            pressure_top_anchor_psia=float(p_match),
            condenser_duty_prev=None,
        )
        _dydt_match, diag_match = column_rhs(0.0, np.asarray(y0, dtype=float), col, layout, match_inputs)
        q_match = _mapping_scalar(diag_match, "Q_cond_calc_BTUph")
        if not np.isfinite(float(q_match)):
            q_match = _mapping_scalar(diag_match, "Q_cond_used_BTUph")
        if np.isfinite(float(q_match)) and float(q_match) < 0.0:
            inputs = replace(
                inputs,
                condenser_duty_btu_per_h=float(q_match),
                condenser_duty_trim_btu_per_h=0.0,
            )
            print(f"matched_condenser_duty_BTUph={float(q_match):.8g}")
        else:
            print("matched_condenser_duty skipped: no finite negative duty")
    inputs = replace(
        inputs,
        enable_liquid_hydraulic_override=True,
        liquid_hydraulic_model=str(args.liquid_hydraulic_model),
        liquid_hydraulic_override_alpha=float(args.liquid_hydraulic_override_alpha),
        liquid_hydraulic_override_alpha_per_stage=None,
    )
    return col, layout, inputs, np.asarray(y0, dtype=float)


def _liquid_h(col: Any, inputs: Any, T_f: float, P_psia: float, x: np.ndarray, fallback: float) -> float:
    thermo = getattr(inputs, "thermo_provider", None)
    if thermo is not None and hasattr(thermo, "h_liq_btu_per_lbmol"):
        try:
            val = float(thermo.h_liq_btu_per_lbmol(float(T_f), float(P_psia), np.asarray(x, dtype=float)))
            if np.isfinite(val):
                return val
        except Exception:
            pass
    return float(fallback)


def _liquid_cp(inputs: Any, T_f: float, P_psia: float, x: np.ndarray, fallback: float) -> float:
    thermo = getattr(inputs, "thermo_provider", None)
    if thermo is not None and hasattr(thermo, "cp_liq_btu_per_lbmolF"):
        try:
            val = float(thermo.cp_liq_btu_per_lbmolF(float(T_f), float(P_psia), np.asarray(x, dtype=float)))
            if np.isfinite(val) and val > 1.0e-9:
                return val
        except Exception:
            pass
    return float(max(fallback, 1.0e-6))


def _build_local_jacobian_sparsity(
    stages: List[int],
    *,
    include_temperature_rate: bool = False,
    include_top_boundary: bool = False,
) -> Any:
    """Build a conservative tri-diagonal stage-neighbor sparsity pattern.

    Variable order is [ML_scale for selected stages, T_delta for selected stages,
    optional top_V_scale, top_L_scale, condenser_duty_trim_units,
    reflux_trim_units].  Neighbor coupling is defined on adjacent selected-stage
    positions, not absolute tray numbers, so disjoint stage windows remain local
    to their selected neighbors.  Top-boundary scalar variables are conservatively
    dense against the physical residual rows and diagonal against their own
    regularization rows.
    """
    m = int(len(stages))
    n_stage_vars = 2 * m
    n_boundary_vars = 4 if include_top_boundary else 0
    n_vars = n_stage_vars + n_boundary_vars
    n_physical_blocks = 2 + (1 if include_temperature_rate else 0)
    n_top_rows = 2 if include_top_boundary else 0
    n_reg_rows = 2 * m + n_boundary_vars
    n_res = n_physical_blocks * m + n_top_rows + n_reg_rows
    sp = lil_matrix((n_res, n_vars), dtype=int)
    row = 0
    dense_boundary_cols = list(range(n_stage_vars, n_vars))

    def local_cols_for(pos: int) -> List[int]:
        local_cols: List[int] = []
        for neighbor in (-1, 0, 1):
            j = pos + neighbor
            if 0 <= j < m:
                local_cols.append(j)
                local_cols.append(m + j)
        local_cols.extend(dense_boundary_cols)
        return local_cols

    # Total liquid material residuals.
    for pos in range(m):
        sp[row + pos, local_cols_for(pos)] = 1
    row += m
    # Liquid-energy residuals.
    for pos in range(m):
        sp[row + pos, local_cols_for(pos)] = 1
    row += m
    # Direct temperature-rate residuals.
    if include_temperature_rate:
        for pos in range(m):
            sp[row + pos, local_cols_for(pos)] = 1
        row += m
    # Top boundary total-rate residuals.
    if include_top_boundary:
        sp[row, :] = 1
        sp[row + 1, :] = 1
        row += 2
    # Stage regularization terms are strictly diagonal.
    for pos in range(m):
        sp[row + pos, pos] = 1
    row += m
    for pos in range(m):
        sp[row + pos, m + pos] = 1
    row += m
    if include_top_boundary:
        for j in range(n_boundary_vars):
            sp[row + j, n_stage_vars + j] = 1
    return sp.tocsr()


def main() -> int:
    ap = argparse.ArgumentParser(description="Reconcile tray liquid holdup and temperature against active liquid hydraulics.")
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--stages", default="2-11", help="1-based stages to vary and target, e.g. 2-11.")
    ap.add_argument("--runtime-mode", default="hydraulic", choices=["legacy", "hydraulic", "parity", "calibration", "total-reflux"])
    ap.add_argument("--thermo", default="clapeyron")
    ap.add_argument("--clapeyron-model", default="PR")
    ap.add_argument("--thermo-table", default="cache/thermo_table.json")
    ap.add_argument("--include-boundary-states", action="store_true", default=True)
    ap.add_argument("--no-boundary-states", dest="include_boundary_states", action="store_false")
    ap.add_argument("--use-excel-vapor-holdup", action="store_true")
    ap.add_argument("--no-equilibrium", action="store_true")
    ap.add_argument("--equilibrium-relaxation-mode", default="composition-only")
    ap.add_argument("--equilibrium-tau-sec", type=float, default=0.5)
    ap.add_argument("--no-flash-feed-at-stage-conditions", action="store_true")
    ap.add_argument("--vapor-holdup-relaxation-sec", type=float, default=0.0)
    ap.add_argument("--condenser-duty-mode", default="total-condense")
    ap.add_argument("--condenser-duty-btuph", type=float, default=None)
    ap.add_argument("--init-pack-top-drum-vapor-to-pressure", action="store_true")
    ap.add_argument("--init-top-drum-vapor-pressure-psia", type=float, default=None)
    ap.add_argument("--init-match-condenser-duty", action="store_true")
    ap.add_argument("--top-pressure-sp", dest="top_pressure_sp_psia", type=float, default=None)
    ap.add_argument("--reflux", dest="reflux_lbmolph", type=float, default=None)
    ap.add_argument("--liquid-hydraulic-model", default="francis", choices=["francis", "linear-holdup", "skogestad-linear"])
    ap.add_argument("--liquid-hydraulic-override-alpha", type=float, default=1.0)
    ap.add_argument("--ml-lower", type=float, default=0.7)
    ap.add_argument("--ml-upper", type=float, default=1.3)
    ap.add_argument("--temperature-delta-lower-f", type=float, default=-10.0)
    ap.add_argument("--temperature-delta-upper-f", type=float, default=10.0)
    ap.add_argument("--stage2-ml-lower", type=float, default=None)
    ap.add_argument("--stage2-ml-upper", type=float, default=None)
    ap.add_argument("--stage2-temperature-delta-lower-f", type=float, default=None)
    ap.add_argument("--stage2-temperature-delta-upper-f", type=float, default=None)
    ap.add_argument(
        "--stage2-regularization-multiplier",
        type=float,
        default=10.0,
        help="Multiplier on ML/T regularization for selected stage 2.",
    )
    ap.add_argument(
        "--include-temperature-rate-residual",
        action="store_true",
        help="Add direct tray dT/dt residuals for selected stages.",
    )
    ap.add_argument(
        "--enable-top-boundary-dofs",
        action="store_true",
        help="Add scalar top boundary DOFs: top_V scale, top_L scale, condenser duty trim, reflux trim.",
    )
    ap.add_argument("--top-v-scale-lower", type=float, default=0.5)
    ap.add_argument("--top-v-scale-upper", type=float, default=2.0)
    ap.add_argument("--top-l-scale-lower", type=float, default=0.5)
    ap.add_argument("--top-l-scale-upper", type=float, default=2.0)
    ap.add_argument("--condenser-duty-trim-scale-btuph", type=float, default=1.0e6)
    ap.add_argument("--condenser-duty-trim-lower-units", type=float, default=-5.0)
    ap.add_argument("--condenser-duty-trim-upper-units", type=float, default=5.0)
    ap.add_argument("--reflux-trim-scale-lbmolph", type=float, default=100.0)
    ap.add_argument("--reflux-trim-lower-units", type=float, default=-5.0)
    ap.add_argument("--reflux-trim-upper-units", type=float, default=5.0)
    ap.add_argument("--lambda-mass", type=float, default=0.05)
    ap.add_argument("--lambda-temp", type=float, default=0.005)
    ap.add_argument("--lambda-top-v", type=float, default=0.05)
    ap.add_argument("--lambda-top-l", type=float, default=0.05)
    ap.add_argument("--lambda-condenser-duty-trim", type=float, default=0.02)
    ap.add_argument("--lambda-reflux-trim", type=float, default=0.02)
    ap.add_argument("--mass-residual-weight", type=float, default=1.0)
    ap.add_argument("--energy-residual-weight", type=float, default=1.0)
    ap.add_argument("--temperature-rate-residual-weight", type=float, default=1.0)
    ap.add_argument("--top-boundary-rate-weight", type=float, default=1.0)
    ap.add_argument("--cp-fallback-btu-per-lbmol-f", type=float, default=50.0)
    ap.add_argument(
        "--progress-every",
        type=int,
        default=1,
        help="Print optimizer timing/progress every N objective evaluations; use 0 to disable.",
    )
    ap.add_argument(
        "--progress-log",
        default=None,
        help="Optional path for flushed progress telemetry; defaults to OUTPUT.progress.log when omitted.",
    )
    ap.add_argument(
        "--evaluate-only",
        action="store_true",
        help="Evaluate the scaled t=0 residual and write a summary without running least_squares.",
    )
    ap.add_argument(
        "--no-jac-sparsity",
        dest="use_jac_sparsity",
        action="store_false",
        help="Disable the local neighbor finite-difference sparsity pattern.",
    )
    ap.set_defaults(use_jac_sparsity=True)
    ap.add_argument("--max-nfev", type=int, default=80)
    args = ap.parse_args()

    progress_log_path = _resolve(args.progress_log) if args.progress_log else _resolve(args.output).with_suffix(".progress.log")
    progress_log_path.parent.mkdir(parents=True, exist_ok=True)

    def log_progress(message: str) -> None:
        line = f"{time.strftime('%Y-%m-%dT%H:%M:%S')} {message}"
        print(line, flush=True)
        with progress_log_path.open("a", encoding="utf-8") as fh:
            fh.write(line + "\n")
            fh.flush()

    log_progress("build_runtime start")
    col, layout, inputs, y_base = _build_runtime(args)
    log_progress("build_runtime done")
    n = int(col.n_stages)
    nc = int(col.n_components)
    stages = _selected_stages(args.stages, n)
    if not stages:
        raise ValueError("No stages selected.")
    sl = layout.slices()
    u_base = layout.unpack(y_base)
    tray_l_base = np.asarray(u_base["tray_L"], dtype=float).reshape((n, nc))
    top_l_base = np.asarray(u_base.get("top_L", np.zeros(nc)), dtype=float).reshape((nc,))
    top_v_base = np.asarray(u_base.get("top_V", np.zeros(nc)), dtype=float).reshape((nc,))
    top_l_total_seed = max(float(np.sum(top_l_base)), 1.0e-12)
    top_v_total_seed = max(float(np.sum(top_v_base)), 1.0e-12)
    ml_seed = np.sum(tray_l_base, axis=1)
    x_seed = np.asarray(u_base["x_tray"], dtype=float).reshape((n, nc))
    t_seed = np.asarray(u_base["tray_T_f"], dtype=float).reshape((n,))
    p_seed = np.asarray(getattr(col, "P_psia"), dtype=float).reshape((n,))
    el_seed = np.asarray(u_base["tray_EL_BTU"], dtype=float).reshape((n,))
    h_seed = np.divide(el_seed, np.maximum(ml_seed, 1.0e-12))
    cp_seed = np.array(
        [
            _liquid_cp(
                inputs,
                float(t_seed[i]),
                float(p_seed[i]),
                x_seed[i, :],
                float(args.cp_fallback_btu_per_lbmol_f),
            )
            for i in range(n)
        ],
        dtype=float,
    )

    m = len(stages)
    include_top_boundary = bool(args.enable_top_boundary_dofs)
    n_stage_vars = 2 * m
    n_boundary_vars = 4 if include_top_boundary else 0
    nvar = n_stage_vars + n_boundary_vars
    z_parts = [np.ones(m, dtype=float), np.zeros(m, dtype=float)]
    lb_parts = [
        np.full(m, float(args.ml_lower), dtype=float),
        np.full(m, float(args.temperature_delta_lower_f), dtype=float),
    ]
    ub_parts = [
        np.full(m, float(args.ml_upper), dtype=float),
        np.full(m, float(args.temperature_delta_upper_f), dtype=float),
    ]
    for pos, stage0 in enumerate(stages):
        if int(stage0) == 1:
            if args.stage2_ml_lower is not None:
                lb_parts[0][pos] = float(args.stage2_ml_lower)
            if args.stage2_ml_upper is not None:
                ub_parts[0][pos] = float(args.stage2_ml_upper)
            if args.stage2_temperature_delta_lower_f is not None:
                lb_parts[1][pos] = float(args.stage2_temperature_delta_lower_f)
            if args.stage2_temperature_delta_upper_f is not None:
                ub_parts[1][pos] = float(args.stage2_temperature_delta_upper_f)
    if include_top_boundary:
        z_parts.append(np.array([1.0, 1.0, 0.0, 0.0], dtype=float))
        lb_parts.append(
            np.array(
                [
                    float(args.top_v_scale_lower),
                    float(args.top_l_scale_lower),
                    float(args.condenser_duty_trim_lower_units),
                    float(args.reflux_trim_lower_units),
                ],
                dtype=float,
            )
        )
        ub_parts.append(
            np.array(
                [
                    float(args.top_v_scale_upper),
                    float(args.top_l_scale_upper),
                    float(args.condenser_duty_trim_upper_units),
                    float(args.reflux_trim_upper_units),
                ],
                dtype=float,
            )
        )
    z0 = np.concatenate(z_parts)
    lb = np.concatenate(lb_parts)
    ub = np.concatenate(ub_parts)
    base_boundary = inputs.boundary
    base_reflux_lbmolph = float(getattr(base_boundary, "reflux_lbmolph", 0.0) or 0.0)
    base_q_cond_btu_per_h = getattr(inputs, "condenser_duty_btu_per_h", None)
    try:
        base_q_cond_btu_per_h = float(base_q_cond_btu_per_h)
    except Exception:
        base_q_cond_btu_per_h = float("nan")

    eval_count = 0
    best: Dict[str, Any] = {"norm": float("inf"), "z": z0.copy(), "max_abs": float("inf")}
    timing: Dict[str, float] = {
        "objective_wall_total_sec": 0.0,
        "rhs_wall_total_sec": 0.0,
        "make_state_wall_total_sec": 0.0,
        "last_objective_wall_sec": 0.0,
        "last_rhs_wall_sec": 0.0,
        "last_make_state_wall_sec": 0.0,
    }

    def _boundary_values(z: np.ndarray) -> Dict[str, float]:
        if not include_top_boundary:
            return {
                "top_v_scale": 1.0,
                "top_l_scale": 1.0,
                "q_trim_btu_per_h": 0.0,
                "reflux_trim_lbmolph": 0.0,
                "reflux_lbmolph": base_reflux_lbmolph,
                "condenser_duty_btu_per_h": float(base_q_cond_btu_per_h),
            }
        b = np.asarray(z[n_stage_vars:n_stage_vars + 4], dtype=float)
        q_trim = float(b[2]) * float(args.condenser_duty_trim_scale_btuph)
        reflux_trim = float(b[3]) * float(args.reflux_trim_scale_lbmolph)
        q_base = float(base_q_cond_btu_per_h) if np.isfinite(float(base_q_cond_btu_per_h)) else 0.0
        return {
            "top_v_scale": float(b[0]),
            "top_l_scale": float(b[1]),
            "q_trim_btu_per_h": float(q_trim),
            "reflux_trim_lbmolph": float(reflux_trim),
            "reflux_lbmolph": float(base_reflux_lbmolph + reflux_trim),
            "condenser_duty_btu_per_h": float(q_base + q_trim),
        }

    def make_inputs(z: np.ndarray) -> Any:
        vals = _boundary_values(z)
        if not include_top_boundary:
            return inputs
        b0 = inputs.boundary
        return replace(
            inputs,
            condenser_duty_btu_per_h=float(vals["condenser_duty_btu_per_h"]),
            condenser_duty_trim_btu_per_h=0.0,
            boundary=BoundaryFlows(
                reflux_lbmolph=float(vals["reflux_lbmolph"]),
                boilup_lbmolph=float(getattr(b0, "boilup_lbmolph", 0.0) or 0.0),
                distillate_lbmolph=float(getattr(b0, "distillate_lbmolph", 0.0) or 0.0),
                bottoms_lbmolph=float(getattr(b0, "bottoms_lbmolph", 0.0) or 0.0),
            ),
        )

    def make_state(z: np.ndarray) -> np.ndarray:
        t_make0 = time.perf_counter()
        y = np.asarray(y_base, dtype=float).copy()
        tray_l = tray_l_base.copy()
        t = t_seed.copy()
        el = el_seed.copy()
        s_ml = np.asarray(z[:m], dtype=float)
        d_t = np.asarray(z[m:n_stage_vars], dtype=float)
        for j, i in enumerate(stages):
            scale = float(s_ml[j])
            tray_l[i, :] = tray_l_base[i, :] * scale
            t[i] = float(t_seed[i]) + float(d_t[j])
            ml_i = float(np.sum(tray_l[i, :]))
            x_i = x_seed[i, :]
            h_i = _liquid_h(col, inputs, float(t[i]), float(p_seed[i]), x_i, float(h_seed[i]))
            el[i] = ml_i * h_i
        y[sl["tray_L"]] = tray_l.reshape(-1)
        y[sl["tray_T_f"]] = t.reshape(-1)
        y[sl["tray_EL_BTU"]] = el.reshape(-1)
        if include_top_boundary:
            vals = _boundary_values(z)
            if "top_L" in sl:
                y[sl["top_L"]] = top_l_base * float(vals["top_l_scale"])
            if "top_V" in sl:
                y[sl["top_V"]] = top_v_base * float(vals["top_v_scale"])
        timing["last_make_state_wall_sec"] = float(time.perf_counter() - t_make0)
        timing["make_state_wall_total_sec"] += float(timing["last_make_state_wall_sec"])
        return y

    def scaled_residual_from_state(y: np.ndarray, eval_inputs: Any) -> Dict[str, Any]:
        t_rhs0 = time.perf_counter()
        dydt, _diag = column_rhs(0.0, np.asarray(y, dtype=float), col, layout, eval_inputs)
        timing["last_rhs_wall_sec"] = float(time.perf_counter() - t_rhs0)
        timing["rhs_wall_total_sec"] += float(timing["last_rhs_wall_sec"])
        du = layout.unpack(np.asarray(dydt, dtype=float))
        dml = np.sum(np.asarray(du["tray_L"], dtype=float).reshape((n, nc)), axis=1)
        deldt = np.asarray(du["tray_EL_BTU"], dtype=float).reshape((n,))
        dtemp = np.asarray(du.get("tray_T_f", np.zeros(n)), dtype=float).reshape((n,))
        d_top_l = np.asarray(du.get("top_L", np.zeros(nc)), dtype=float).reshape((nc,))
        d_top_v = np.asarray(du.get("top_V", np.zeros(nc)), dtype=float).reshape((nc,))
        idx = np.asarray(stages, dtype=int)
        scaled_mass = dml[idx] / np.maximum(ml_seed[idx], 1.0e-12)
        scaled_energy = deldt[idx] / np.maximum(ml_seed[idx] * cp_seed[idx], 1.0e-12)
        scaled_top_l = float(np.sum(d_top_l) / top_l_total_seed)
        scaled_top_v = float(np.sum(d_top_v) / top_v_total_seed)
        return {
            "dml": dml,
            "deldt": deldt,
            "dtemp": dtemp,
            "scaled_mass": scaled_mass,
            "scaled_energy": scaled_energy,
            "scaled_top_l": scaled_top_l,
            "scaled_top_v": scaled_top_v,
            "dydt": dydt,
        }

    log_progress("initial scaled residual start")
    before = scaled_residual_from_state(y_base, inputs)
    log_progress(
        "initial scaled residual done "
        f"rhs_wall={float(timing['last_rhs_wall_sec']):.3f}s "
        f"max_mass={float(np.max(np.abs(before['scaled_mass']))):.8g} "
        f"max_energy={float(np.max(np.abs(before['scaled_energy']))):.8g}"
    )
    if bool(args.evaluate_only):
        in_path = _resolve(args.input)
        out_path = _resolve(args.output)
        summary = {
            "input": str(in_path),
            "output": str(out_path),
            "stages": [int(i + 1) for i in stages],
            "evaluate_only": True,
            "before_max_abs_scaled_mass_per_s": float(np.max(np.abs(before["scaled_mass"]))),
            "before_max_abs_scaled_energy_F_per_s": float(np.max(np.abs(before["scaled_energy"]))),
            "before_scaled_mass_per_s_by_stage": {
                str(i + 1): float(before["scaled_mass"][j]) for j, i in enumerate(stages)
            },
            "before_scaled_energy_F_per_s_by_stage": {
                str(i + 1): float(before["scaled_energy"][j]) for j, i in enumerate(stages)
            },
        }
        summary_path = out_path.with_suffix(".liquid_energy_summary.json")
        summary_path.parent.mkdir(parents=True, exist_ok=True)
        summary_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")
        log_progress(f"Wrote: {summary_path}")
        log_progress(
            "evaluate-only mass/energy max="
            f"{summary['before_max_abs_scaled_mass_per_s']:.8g}/"
            f"{summary['before_max_abs_scaled_energy_F_per_s']:.8g}"
        )
        return 0

    def objective(z: np.ndarray) -> np.ndarray:
        nonlocal eval_count, best
        eval_count += 1
        t_obj0 = time.perf_counter()
        y = make_state(z)
        rr = scaled_residual_from_state(y, make_inputs(z))
        s_ml = np.asarray(z[:m], dtype=float)
        d_t = np.asarray(z[m:n_stage_vars], dtype=float)
        stage_reg_mult = np.ones(m, dtype=float)
        for j, i in enumerate(stages):
            if int(i) == 1:
                stage_reg_mult[j] = max(float(args.stage2_regularization_multiplier), 0.0)
        parts = [
            float(args.mass_residual_weight) * np.asarray(rr["scaled_mass"], dtype=float),
            float(args.energy_residual_weight) * np.asarray(rr["scaled_energy"], dtype=float),
        ]
        if bool(args.include_temperature_rate_residual):
            idx = np.asarray(stages, dtype=int)
            parts.append(float(args.temperature_rate_residual_weight) * np.asarray(rr["dtemp"], dtype=float)[idx])
        if include_top_boundary:
            parts.append(
                float(args.top_boundary_rate_weight)
                * np.array([float(rr["scaled_top_v"]), float(rr["scaled_top_l"])], dtype=float)
            )
        parts.append(float(args.lambda_mass) * stage_reg_mult * (s_ml - 1.0))
        parts.append(float(args.lambda_temp) * stage_reg_mult * d_t)
        if include_top_boundary:
            b = np.asarray(z[n_stage_vars:n_stage_vars + 4], dtype=float)
            parts.append(
                np.array(
                    [
                        float(args.lambda_top_v) * (float(b[0]) - 1.0),
                        float(args.lambda_top_l) * (float(b[1]) - 1.0),
                        float(args.lambda_condenser_duty_trim) * float(b[2]),
                        float(args.lambda_reflux_trim) * float(b[3]),
                    ],
                    dtype=float,
                )
            )
        r = np.concatenate(parts)
        norm = float(np.linalg.norm(r))
        max_abs = float(np.max(np.abs(r))) if r.size else 0.0
        if norm < float(best["norm"]):
            best = {"norm": norm, "z": np.asarray(z, dtype=float).copy(), "max_abs": max_abs}
        obj_dt = float(time.perf_counter() - t_obj0)
        timing["last_objective_wall_sec"] = obj_dt
        timing["objective_wall_total_sec"] += obj_dt
        progress_every = max(int(args.progress_every), 0)
        if progress_every > 0 and (eval_count == 1 or eval_count % progress_every == 0):
            cost = float(np.sum(np.asarray(r, dtype=float) ** 2))
            log_progress(
                "[Tracker] "
                f"eval={eval_count} "
                f"obj_wall={obj_dt:.3f}s "
                f"rhs_wall={float(timing['last_rhs_wall_sec']):.3f}s "
                f"pack_wall={float(timing['last_make_state_wall_sec']):.3f}s "
                f"cost={cost:.6e} "
                f"norm={norm:.8g} "
                f"max_abs={max_abs:.8g} "
                f"best_norm={best['norm']:.8g}",
            )
        return r

    jac_sparsity = (
        _build_local_jacobian_sparsity(
            stages,
            include_temperature_rate=bool(args.include_temperature_rate_residual),
            include_top_boundary=include_top_boundary,
        )
        if bool(args.use_jac_sparsity)
        else None
    )
    n_residuals = int(
        (2 + (1 if bool(args.include_temperature_rate_residual) else 0)) * m
        + (2 if include_top_boundary else 0)
        + 2 * m
        + (4 if include_top_boundary else 0)
    )
    log_progress(
        "least_squares start "
        f"nvar={int(nvar)} nres={int(n_residuals)} "
        f"jac_sparsity={bool(args.use_jac_sparsity)} "
        f"nnz={int(jac_sparsity.nnz) if jac_sparsity is not None else 0}"
    )
    result = least_squares(
        objective,
        z0,
        bounds=(lb, ub),
        method="trf",
        jac_sparsity=jac_sparsity,
        x_scale="jac",
        max_nfev=max(int(args.max_nfev), 1),
        verbose=1,
    )
    log_progress(
        "least_squares done "
        f"success={bool(result.success)} nfev={int(result.nfev)} "
        f"eval_count={int(eval_count)} best_norm={float(best['norm']):.8g}"
    )
    z_best = np.asarray(best["z"], dtype=float)
    y_best = make_state(z_best)
    after_inputs = make_inputs(z_best)
    after = scaled_residual_from_state(y_best, after_inputs)
    u_best = layout.unpack(y_best)

    in_path = _resolve(args.input)
    out_path = _resolve(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    if in_path.resolve() != out_path.resolve():
        shutil.copy2(in_path, out_path)
    wb = load_workbook(out_path)
    if "Initial Conditions" not in wb.sheetnames:
        raise ValueError("Workbook is missing Initial Conditions sheet.")
    ws_ic = wb["Initial Conditions"]
    _row_t, col_t = _find_header(ws_ic, "Temperature (F)")
    _row_ml, col_ml = _find_header(ws_ic, "Liquid Holdup (lbmol)")
    ml_best = np.sum(np.asarray(u_best["tray_L"], dtype=float).reshape((n, nc)), axis=1)
    t_best = np.asarray(u_best["tray_T_f"], dtype=float).reshape((n,))
    for i in stages:
        r = _stage_row(ws_ic, i + 1)
        ws_ic.cell(r, col_ml).value = float(ml_best[i])
        ws_ic.cell(r, col_t).value = float(t_best[i])
    if "Energy State" in wb.sheetnames:
        ws_e = wb["Energy State"]
        _row_el, col_el = _find_header(ws_e, "Tray EL (BTU)")
        h_col = None
        try:
            _row_h, h_col = _find_header(ws_e, "hL solved (BTU/lbmol)")
        except Exception:
            h_col = None
        el_best = np.asarray(u_best["tray_EL_BTU"], dtype=float).reshape((n,))
        for i in stages:
            r = _stage_row(ws_e, i + 1)
            ws_e.cell(r, col_el).value = float(el_best[i])
            if h_col is not None:
                ws_e.cell(r, h_col).value = float(el_best[i] / max(float(ml_best[i]), 1.0e-12))
    boundary_vals = _boundary_values(z_best)
    if include_top_boundary and "Boundary State" in wb.sheetnames:
        ws_b = wb["Boundary State"]
        if "top_L" in u_best:
            _write_boundary_vector(ws_b, "top_L", np.asarray(u_best["top_L"], dtype=float))
        if "top_V" in u_best:
            _write_boundary_vector(ws_b, "top_V", np.asarray(u_best["top_V"], dtype=float))
    if include_top_boundary and "Specifications" in wb.sheetnames:
        _write_spec_value(wb["Specifications"], "Condenser Duty (Btu/h)", float(boundary_vals["condenser_duty_btu_per_h"]))
    wb.save(out_path)

    idx = np.asarray(stages, dtype=int)
    s_best = np.asarray(z_best[:m], dtype=float)
    dt_best = np.asarray(z_best[m:n_stage_vars], dtype=float)
    before_temp_rate = np.asarray(before["dtemp"], dtype=float)[idx]
    after_temp_rate = np.asarray(after["dtemp"], dtype=float)[idx]
    summary = {
        "input": str(in_path),
        "output": str(out_path),
        "stages": [int(i + 1) for i in stages],
        "include_temperature_rate_residual": bool(args.include_temperature_rate_residual),
        "enable_top_boundary_dofs": bool(include_top_boundary),
        "least_squares_success": bool(result.success),
        "least_squares_message": str(result.message),
        "nfev": int(result.nfev),
        "jac_sparsity_enabled": bool(args.use_jac_sparsity),
        "jac_sparsity_nnz": int(jac_sparsity.nnz) if jac_sparsity is not None else 0,
        "objective_eval_count": int(eval_count),
        "objective_wall_total_sec": float(timing["objective_wall_total_sec"]),
        "rhs_wall_total_sec": float(timing["rhs_wall_total_sec"]),
        "make_state_wall_total_sec": float(timing["make_state_wall_total_sec"]),
        "best_norm": float(best["norm"]),
        "before_max_abs_scaled_mass_per_s": float(np.max(np.abs(before["scaled_mass"]))),
        "before_max_abs_scaled_energy_F_per_s": float(np.max(np.abs(before["scaled_energy"]))),
        "after_max_abs_scaled_mass_per_s": float(np.max(np.abs(after["scaled_mass"]))),
        "after_max_abs_scaled_energy_F_per_s": float(np.max(np.abs(after["scaled_energy"]))),
        "before_max_abs_temperature_rate_F_per_s": float(np.max(np.abs(before_temp_rate))),
        "after_max_abs_temperature_rate_F_per_s": float(np.max(np.abs(after_temp_rate))),
        "before_scaled_top_v_per_s": float(before["scaled_top_v"]),
        "before_scaled_top_l_per_s": float(before["scaled_top_l"]),
        "after_scaled_top_v_per_s": float(after["scaled_top_v"]),
        "after_scaled_top_l_per_s": float(after["scaled_top_l"]),
        "max_abs_ML_scale_delta": float(np.max(np.abs(s_best - 1.0))),
        "max_abs_T_delta_F": float(np.max(np.abs(dt_best))),
        "ML_scale_by_stage": {str(i + 1): float(s_best[j]) for j, i in enumerate(stages)},
        "T_delta_F_by_stage": {str(i + 1): float(dt_best[j]) for j, i in enumerate(stages)},
        "top_boundary": boundary_vals,
    }
    if include_top_boundary:
        b = np.asarray(z_best[n_stage_vars:n_stage_vars + 4], dtype=float)
        summary["top_boundary_units"] = {
            "top_V_scale": float(b[0]),
            "top_L_scale": float(b[1]),
            "condenser_duty_trim_units": float(b[2]),
            "reflux_trim_units": float(b[3]),
            "base_top_V_total_lbmol": float(top_v_total_seed),
            "base_top_L_total_lbmol": float(top_l_total_seed),
            "optimized_top_V_total_lbmol": float(np.sum(np.asarray(u_best.get("top_V", []), dtype=float))),
            "optimized_top_L_total_lbmol": float(np.sum(np.asarray(u_best.get("top_L", []), dtype=float))),
            "base_reflux_lbmolph": float(base_reflux_lbmolph),
            "base_condenser_duty_btu_per_h": float(base_q_cond_btu_per_h),
            "condenser_duty_trim_scale_btu_per_h": float(args.condenser_duty_trim_scale_btuph),
            "reflux_trim_scale_lbmolph": float(args.reflux_trim_scale_lbmolph),
        }
    summary_path = out_path.with_suffix(".liquid_energy_summary.json")
    summary_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")
    print(f"Wrote: {out_path}")
    print(f"Wrote: {summary_path}")
    print(
        "before mass/energy max="
        f"{summary['before_max_abs_scaled_mass_per_s']:.8g}/"
        f"{summary['before_max_abs_scaled_energy_F_per_s']:.8g}; "
        "after="
        f"{summary['after_max_abs_scaled_mass_per_s']:.8g}/"
        f"{summary['after_max_abs_scaled_energy_F_per_s']:.8g}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
