#!/usr/bin/env python
"""Create a Gani workbook seed adjusted for explicit model boundary states.

The ChemSep/source topology treats the terminal reboiler as the last stage:
the last liquid-flow entry is the bottoms product.  The full dynamic model,
when explicit bottom-sump states are enabled, treats the last stage as draining
to a separate sump/reboiler boundary.  For that topology the last stage liquid
outflow must be the liquid entering the sump, i.e. bottoms plus boilup, not the
bottoms product alone.
"""

from __future__ import annotations

import argparse
from pathlib import Path
import sys

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]


def _resolve_path(raw: str) -> Path:
    p = Path(raw)
    if p.is_absolute():
        return p
    return PROJECT_ROOT / p


def _find_header(ws, header: str) -> tuple[int, int]:
    target = str(header).strip().lower()
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value).strip().lower() == target:
                return int(cell.row), int(cell.column)
    raise ValueError(f"Could not find header {header!r} in sheet {ws.title!r}")


def _append_note(wb, field: str, value: str) -> None:
    ws = wb["Notes"] if "Notes" in wb.sheetnames else wb.create_sheet("Notes")
    if ws.max_row < 1:
        ws.cell(1, 1).value = "Field"
        ws.cell(1, 2).value = "Value"
    row = int(ws.max_row) + 1
    ws.cell(row, 1).value = field
    ws.cell(row, 2).value = value


def main() -> int:
    ap = argparse.ArgumentParser(description="Create Gani explicit-boundary model-topology seed workbook.")
    ap.add_argument("--input", default="validation_gani_1986_debutanizer.xlsx")
    ap.add_argument("--output", default="validation_gani_1986_debutanizer_model_topology_seed.xlsx")
    ap.add_argument(
        "--bottom-liquid-source",
        choices=["previous-stage-liquid", "bottoms-plus-boilup"],
        default="previous-stage-liquid",
        help="How to set the last-stage liquid flow for explicit sump topology.",
    )
    args = ap.parse_args()

    in_path = _resolve_path(str(args.input))
    out_path = _resolve_path(str(args.output))
    if not in_path.exists():
        raise FileNotFoundError(in_path)

    wb = load_workbook(in_path)
    ws_ic = wb["Initial Conditions"]
    stage_hdr_row, stage_col = _find_header(ws_ic, "Stage")
    _, liquid_col = _find_header(ws_ic, "Liquid Flow (lbmol/h)")
    _, vapor_col = _find_header(ws_ic, "Vapor Flow (lbmol/h)")

    stage_rows: list[tuple[int, int]] = []
    for row in range(stage_hdr_row + 1, ws_ic.max_row + 1):
        val = ws_ic.cell(row, stage_col).value
        if val is None:
            continue
        try:
            stage_rows.append((int(val), row))
        except Exception:
            continue
    if len(stage_rows) < 2:
        raise ValueError("Need at least two stage rows in Initial Conditions.")

    stage_rows.sort(key=lambda item: item[0])
    last_stage, last_row = stage_rows[-1]
    prev_stage, prev_row = stage_rows[-2]

    old_last_liquid = float(ws_ic.cell(last_row, liquid_col).value)
    prev_liquid = float(ws_ic.cell(prev_row, liquid_col).value)
    boilup = float(ws_ic.cell(last_row, vapor_col).value)

    if str(args.bottom_liquid_source) == "bottoms-plus-boilup":
        ws_streams = wb["Streams"]
        bottom_total = None
        for row in range(1, ws_streams.max_row + 1):
            if str(ws_streams.cell(row, 1).value).strip().lower() == "total molar flow (lbmol/h)":
                bottom_total = float(ws_streams.cell(row, 4).value)
                break
        if bottom_total is None:
            raise ValueError("Could not find Bottom total molar flow in Streams.")
        new_last_liquid = bottom_total + boilup
    else:
        new_last_liquid = prev_liquid

    ws_ic.cell(last_row, liquid_col).value = float(new_last_liquid)

    _append_note(
        wb,
        "Model-topology bottom liquid conversion",
        (
            f"Set stage {last_stage} Liquid Flow from {old_last_liquid:.9g} to "
            f"{float(new_last_liquid):.9g} lbmol/h for explicit bottom-sump topology. "
            f"The ChemSep/source value is the bottoms product; the full model needs "
            f"liquid to sump approximately equal to bottoms plus boilup. Source basis: "
            f"{args.bottom_liquid_source}; previous stage {prev_stage} liquid={prev_liquid:.9g}, "
            f"boilup={boilup:.9g}."
        ),
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Wrote {out_path}")
    print(f"Stage {last_stage} Liquid Flow (lbmol/h): {old_last_liquid:.9g} -> {float(new_last_liquid):.9g}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
