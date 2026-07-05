#!/usr/bin/env python
"""
Smooth the feed-stage initialization cliff for total-reflux startup probes.

This is an opt-in workbook preconditioner. It edits a copy of the seed workbook
so the selected feed stage starts from a local average of its neighboring
stages rather than from a continuous-feed inflection point. The goal is to test
whether total-reflux startup can avoid the stage feed-bulge transient without
changing the model equations.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
import sys

import numpy as np
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]


def _resolve(path: str | Path) -> Path:
    p = Path(path)
    if p.is_absolute():
        return p.resolve()
    return (PROJECT_ROOT / p).resolve()


def _headers(ws) -> dict[str, int]:
    return {str(ws.cell(1, c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value is not None}


def _find_stage_row(ws, stage: int) -> int:
    for r in range(2, ws.max_row + 1):
        try:
            if int(ws.cell(r, 1).value) == int(stage):
                return r
        except Exception:
            continue
    raise ValueError(f"Stage {stage} not found in sheet {ws.title!r}")


def _float_cell(ws, row: int, col: int) -> float:
    val = ws.cell(row, col).value
    try:
        out = float(val)
    except Exception as exc:
        raise ValueError(f"Cannot parse numeric cell {ws.title}!R{row}C{col}: {val!r}") from exc
    if not np.isfinite(out):
        raise ValueError(f"Non-finite numeric cell {ws.title}!R{row}C{col}: {val!r}")
    return float(out)


def _normalize(values: np.ndarray) -> np.ndarray:
    arr = np.asarray(values, dtype=float).reshape((-1,))
    arr = np.clip(arr, 0.0, None)
    total = float(np.sum(arr))
    if total <= 0.0 or not np.isfinite(total):
        raise ValueError("Cannot normalize zero or non-finite composition vector")
    return arr / total


def _component_columns(headers: dict[str, int], prefix: str) -> list[int]:
    cols: list[tuple[int, int]] = []
    for name, col in headers.items():
        if name.startswith(prefix):
            try:
                idx = int(name.rsplit(" ", 1)[-1])
            except Exception:
                idx = len(cols) + 1
            cols.append((idx, col))
    return [col for _idx, col in sorted(cols)]


def precondition(
    *,
    excel: Path,
    output: Path,
    stage: int,
    smooth_temperature: bool,
    smooth_liquid_composition: bool,
    smooth_vapor_composition: bool,
    smooth_holdup: bool,
    smooth_energy: bool,
) -> dict:
    wb = load_workbook(excel)
    if "Initial Conditions" not in wb.sheetnames:
        raise ValueError("Workbook is missing 'Initial Conditions' sheet")
    ws = wb["Initial Conditions"]
    h = _headers(ws)

    row = _find_stage_row(ws, stage)
    row_prev = _find_stage_row(ws, stage - 1)
    row_next = _find_stage_row(ws, stage + 1)

    summary: dict[str, object] = {
        "input": str(excel),
        "output": str(output),
        "stage": int(stage),
        "actions": [],
    }

    if smooth_temperature:
        col_T = h.get("Temperature (F)")
        if col_T is None:
            raise ValueError("Initial Conditions missing 'Temperature (F)'")
        old = _float_cell(ws, row, col_T)
        new = 0.5 * (_float_cell(ws, row_prev, col_T) + _float_cell(ws, row_next, col_T))
        ws.cell(row, col_T).value = float(new)
        summary["actions"].append({"field": "Temperature (F)", "old": old, "new": float(new)})

    if smooth_holdup:
        col_ML = h.get("Liquid Holdup (lbmol)")
        if col_ML is None:
            raise ValueError("Initial Conditions missing 'Liquid Holdup (lbmol)'")
        old = _float_cell(ws, row, col_ML)
        new = 0.5 * (_float_cell(ws, row_prev, col_ML) + _float_cell(ws, row_next, col_ML))
        ws.cell(row, col_ML).value = float(new)
        summary["actions"].append({"field": "Liquid Holdup (lbmol)", "old": old, "new": float(new)})

    liq_cols = _component_columns(h, "Liquid Composition Component")
    vap_cols = _component_columns(h, "Vapor Composition Component")
    if smooth_liquid_composition:
        old = np.array([_float_cell(ws, row, c) for c in liq_cols], dtype=float)
        new = _normalize(
            0.5
            * (
                np.array([_float_cell(ws, row_prev, c) for c in liq_cols], dtype=float)
                + np.array([_float_cell(ws, row_next, c) for c in liq_cols], dtype=float)
            )
        )
        for c, val in zip(liq_cols, new):
            ws.cell(row, c).value = float(val)
        summary["actions"].append({"field": "liquid_composition", "old": old.tolist(), "new": new.tolist()})

    if smooth_vapor_composition:
        old = np.array([_float_cell(ws, row, c) for c in vap_cols], dtype=float)
        new = _normalize(
            0.5
            * (
                np.array([_float_cell(ws, row_prev, c) for c in vap_cols], dtype=float)
                + np.array([_float_cell(ws, row_next, c) for c in vap_cols], dtype=float)
            )
        )
        for c, val in zip(vap_cols, new):
            ws.cell(row, c).value = float(val)
        summary["actions"].append({"field": "vapor_composition", "old": old.tolist(), "new": new.tolist()})

    if smooth_energy and "Energy State" in wb.sheetnames:
        ews = wb["Energy State"]
        eh = _headers(ews)
        erow = _find_stage_row(ews, stage)
        erow_prev = _find_stage_row(ews, stage - 1)
        erow_next = _find_stage_row(ews, stage + 1)
        col_EL = eh.get("Tray EL (BTU)")
        col_hL = eh.get("hL solved (BTU/lbmol)")
        col_ML = h.get("Liquid Holdup (lbmol)")
        if col_EL is not None and col_hL is not None and col_ML is not None:
            old_el = _float_cell(ews, erow, col_EL)
            old_hl = _float_cell(ews, erow, col_hL)
            new_hl = 0.5 * (_float_cell(ews, erow_prev, col_hL) + _float_cell(ews, erow_next, col_hL))
            ml = _float_cell(ws, row, col_ML)
            new_el = float(ml) * float(new_hl)
            ews.cell(erow, col_hL).value = float(new_hl)
            ews.cell(erow, col_EL).value = float(new_el)
            summary["actions"].append(
                {
                    "field": "Tray EL/hL",
                    "old_EL": old_el,
                    "old_hL": old_hl,
                    "new_EL": float(new_el),
                    "new_hL": float(new_hl),
                }
            )

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return summary


def main() -> int:
    ap = argparse.ArgumentParser(description="Precondition a feed-stage seed for total-reflux startup.")
    ap.add_argument("--excel", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--summary", default=None)
    ap.add_argument("--stage", type=int, default=12)
    ap.add_argument("--no-smooth-temperature", dest="smooth_temperature", action="store_false")
    ap.add_argument("--no-smooth-liquid-composition", dest="smooth_liquid_composition", action="store_false")
    ap.add_argument("--smooth-vapor-composition", action="store_true")
    ap.add_argument("--smooth-holdup", dest="smooth_holdup", action="store_true")
    ap.add_argument("--smooth-energy", dest="smooth_energy", action="store_true")
    ap.set_defaults(
        smooth_temperature=True,
        smooth_liquid_composition=True,
        smooth_holdup=False,
        smooth_energy=False,
    )
    args = ap.parse_args()

    summary = precondition(
        excel=_resolve(args.excel),
        output=_resolve(args.output),
        stage=int(args.stage),
        smooth_temperature=bool(args.smooth_temperature),
        smooth_liquid_composition=bool(args.smooth_liquid_composition),
        smooth_vapor_composition=bool(args.smooth_vapor_composition),
        smooth_holdup=bool(args.smooth_holdup),
        smooth_energy=bool(args.smooth_energy),
    )
    text = json.dumps(summary, indent=2)
    print(text)
    if args.summary:
        out = _resolve(args.summary)
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(text + "\n", encoding="utf-8")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
