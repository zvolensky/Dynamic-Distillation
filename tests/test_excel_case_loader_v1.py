"""
test_excel_case_loader_v1.py

Dynamic Distillation - Excel Loader Tests

PURPOSE
-------
Verify that `excel_case_loader_v1` reads template inputs into expected case
structures and preserves recognized spec keys used downstream.

SCOPE
-----
- component canonicalization and base shape checks
- presence/parse behavior for optional thermo/geometry/drum keys

KEY DEPENDENCIES
----------------
- excel_case_loader_v1
"""

import tempfile
from pathlib import Path

from openpyxl import Workbook

from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel


def test_load_case_from_excel_smoke():
    p = Path("distillation_column_template.xlsx")
    if not p.exists():
        p = Path("distillation_column_template_20stage_chemsep_warmer_feed_seed_20260323.xlsx")
    c = load_case_from_excel(str(p))
    assert len(c.components) == 3
    assert c.component_ids_dwsim == ["Propane", "N-butane", "N-pentane"]
    assert "Number of Stages" in c.specs
    # Thermo refresh threshold is optional but should be parsed when present.
    assert "Thermo Refresh dT (F)" in c.specs
    # Reboiler-neighbor vapor guard ratios are optional but should be parsed when present.
    assert "Reboiler Neighbor Vapor Hi Ratio" in c.specs
    assert "Reboiler Neighbor Vapor Lo Ratio" in c.specs
    # Stage-level thermo refresh thresholds are optional but should be parsed when present.
    assert "Thermo Refresh dP (psia)" in c.specs
    assert "Thermo Refresh dX" in c.specs
    assert "Condenser Pressure Drop (psi)" in c.specs
    # Reflux-drum geometry/vapor-volume fields should always exist in specs map.
    assert "Top Drum Vapor Volume (ft3)" in c.specs
    assert "Top Drum Total Volume (ft3)" in c.specs
    assert "Top Drum Diameter (ft)" in c.specs
    assert "Top Drum Length (ft)" in c.specs
    assert "Top Drum Liquid Fraction (-)" in c.specs
    assert "Overhead Vapor Line Volume (ft3)" in c.specs
    assert "Condenser Vapor Volume (ft3)" in c.specs
    assert "Bottom Sump Total Volume (ft3)" in c.specs
    assert "Bottom Sump Diameter (ft)" in c.specs
    assert "Bottom Sump Height (ft)" in c.specs
    assert "Bottom Sump Liquid Fraction (-)" in c.specs
    if c.specs["Thermo Refresh dT (F)"] is not None:
        assert float(c.specs["Thermo Refresh dT (F)"]) > 0.0
    # Geometry table is optional. If present, loader should parse and normalize void fraction.
    assert "Geometry Sections" in c.specs
    secs = c.specs["Geometry Sections"]
    if secs is not None:
        assert isinstance(secs, list)
        assert len(secs) >= 1
        # The (newer) template uses 0.75 (fraction) and also 75 (percent); loader normalizes both.
        if len(secs) >= 2:
            assert abs(float(secs[0]["gas_void_frac"]) - 0.75) < 1e-12
            assert abs(float(secs[1]["gas_void_frac"]) - 0.75) < 1e-12
    assert c.initial_conditions is not None


def test_load_case_from_excel_boundary_state_sheet():
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "boundary_restart.xlsx"
        wb = Workbook()

        ws_specs = wb.active
        ws_specs.title = "Specifications"
        specs_rows = [
            ("Number of Stages", 2),
            ("Number of Components", 2),
            ("Simulation Length (min)", 1.0),
            ("Timestep (sec)", 1.0),
            ("Log Frequency (timesteps)", 1),
        ]
        for r, (k, v) in enumerate(specs_rows, start=1):
            ws_specs.cell(r, 1).value = k
            ws_specs.cell(r, 2).value = v

        ws_ic = wb.create_sheet("Initial Conditions")
        headers = [
            "Stage",
            "Temperature (F)",
            "Pressure (psia)",
            "Vapor Flow (lbmol/h)",
            "Liquid Flow (lbmol/h)",
            "Liquid Holdup (lbmol)",
            "Vapor Composition Component 1",
            "Vapor Composition Component 2",
            "Liquid Composition Component 1",
            "Liquid Composition Component 2",
        ]
        for c, h in enumerate(headers, start=1):
            ws_ic.cell(1, c).value = h
        rows = [
            (1, 100.0, 200.0, 10.0, 20.0, 5.0, 0.6, 0.4, 0.3, 0.7),
            (2, 120.0, 210.0, 11.0, 21.0, 6.0, 0.5, 0.5, 0.4, 0.6),
        ]
        for r, row in enumerate(rows, start=2):
            for c, v in enumerate(row, start=1):
                ws_ic.cell(r, c).value = v

        ws_comp = wb.create_sheet("Components")
        ws_comp.cell(1, 1).value = "Component Name"
        ws_comp.cell(2, 1).value = "Propane"
        ws_comp.cell(3, 1).value = "N-butane"

        ws_boundary = wb.create_sheet("Boundary State")
        ws_boundary.append(["State", "Component 1", "Component 2"])
        ws_boundary.append(["top_L", 10.0, 20.0])
        ws_boundary.append(["top_V", 1.0, 2.0])
        ws_boundary.append(["bottom_L", 30.0, 40.0])
        ws_boundary.append(["bottom_V", 3.0, 4.0])

        ws_energy = wb.create_sheet("Energy State")
        ws_energy.append(["Stage", "Tray EL (BTU)", "Tray EV (BTU)"])
        ws_energy.append([1, 101.0, 201.0])
        ws_energy.append([2, 102.0, 202.0])

        ws_ctrl = wb.create_sheet("Controller State")
        ws_ctrl.append(["Controller", "Value"])
        ws_ctrl.append(["top_level_integ", 1.5])
        ws_ctrl.append(["top_pressure_integ", -2.5])
        ws_ctrl.append(["top_pressure_pv_filt_psia", 221.25])
        ws_ctrl.append(["top_pressure_mv_cmd_btuph", -49640000.0])
        ws_ctrl.append(["top_pressure_resid_abs_btups", 922.55])
        ws_ctrl.append(["top_drum_pressure_T_prev_F", 115.75])
        ws_ctrl.append(["distillate_cmd_lbmolph", 2412.83])
        ws_ctrl.append(["bottoms_cmd_lbmolph", 4761.97])
        ws_ctrl.append(["reflux_cmd_lbmolph", 5945.41])
        ws_ctrl.append(["boilup_cmd_lbmolph", 8014.56])
        ws_ctrl.append(["distillate_comp_integ", 12.5])
        ws_ctrl.append(["bottoms_comp_integ", -7.25])

        ws_mem = wb.create_sheet("Dynamic Memory")
        ws_mem.append(["Stage", "Prev Tray Pressure (psia)", "Prev Tray Temperature (F)"])
        ws_mem.append([1, 200.0, 100.0])
        ws_mem.append([2, 210.0, 120.0])

        wb.save(p)

        c = load_case_from_excel(str(p))
        assert c.boundary_state["top_L"] == [10.0, 20.0]
        assert c.boundary_state["top_V"] == [1.0, 2.0]
        assert c.boundary_state["bottom_L"] == [30.0, 40.0]
        assert c.boundary_state["bottom_V"] == [3.0, 4.0]
        assert c.energy_state["tray_EL_BTU"] == [101.0, 102.0]
        assert c.energy_state["tray_EV_BTU"] == [201.0, 202.0]
        assert c.controller_state["top_level_integ"] == 1.5
        assert c.controller_state["top_pressure_integ"] == -2.5
        assert c.controller_state["top_pressure_pv_filt_psia"] == 221.25
        assert c.controller_state["top_pressure_mv_cmd_btuph"] == -49640000.0
        assert c.controller_state["top_pressure_resid_abs_btups"] == 922.55
        assert c.controller_state["top_drum_pressure_T_prev_F"] == 115.75
        assert c.controller_state["distillate_cmd_lbmolph"] == 2412.83
        assert c.controller_state["bottoms_cmd_lbmolph"] == 4761.97
        assert c.controller_state["reflux_cmd_lbmolph"] == 5945.41
        assert c.controller_state["boilup_cmd_lbmolph"] == 8014.56
        assert c.controller_state["distillate_comp_integ"] == 12.5
        assert c.controller_state["bottoms_comp_integ"] == -7.25
        assert c.memory_state["P_tray_prev_psia"] == [200.0, 210.0]
        assert c.memory_state["T_tray_prev_F"] == [100.0, 120.0]


def test_load_case_from_excel_preserves_control_and_eq_spec_rows():
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "controls.xlsx"
        wb = Workbook()

        ws_specs = wb.active
        ws_specs.title = "Specifications"
        specs_rows = [
            ("Number of Stages", 2),
            ("Number of Components", 2),
            ("Runtime Mode", "hydraulic"),
            ("Thermo Mode", "table"),
            ("Thermo Table", "cache/thermo_table.json"),
            ("Include Energy", True),
            ("Condenser Duty Mode", "specified"),
            ("Simulation Length (min)", 10.0),
            ("Timestep (sec)", 0.2),
            ("Log Frequency (timesteps)", 150),
            ("Stage time constant [tau] (sec)", 4.0),
            ("Equilibrium Relaxation Mode", "phase-holdup"),
            ("Equilibrium Tau (sec)", 4.0),
            ("Equilibrium Energy Damping Gain", 0.2),
            ("Equilibrium Relaxation Live PR", True),
            ("Hydraulic Energy Temperature Follow Tau (sec)", 0.5),
            ("Enable Level Control", True),
            ("Top Level PV Mode", "true-level"),
            ("Top Level SP Frac", 0.5),
            ("Top Level Kc", 0.5),
            ("Top Level Ti (sec)", 1200.0),
            ("Bottom Level PV Mode", "true-level"),
            ("Bottom Level SP (lbmol)", 794.0),
            ("Bottom Level SP Frac", 0.5),
            ("Bottom Level Kc", 8.0),
            ("Bottom Level Ti (sec)", 120.0),
            ("Enable Pressure Control", True),
            ("Pressure Control MV", "condenser-duty"),
            ("Top Pressure SP (psia)", 220.44),
            ("Top Pressure Kc", -150000.0),
            ("Top Pressure Ti (sec)", 120.0),
            ("Enable Distillate Composition Control", True),
            ("Distillate Composition Component", "C4"),
            ("Distillate Composition SP", 0.11),
            ("Distillate Composition Kc", 500.0),
            ("Distillate Composition Ti (sec)", 600.0),
            ("Distillate Composition Reflux Min (lbmol/h)", 2000.0),
            ("Distillate Composition Reflux Max (lbmol/h)", 10000.0),
        ]
        for r, (k, v) in enumerate(specs_rows, start=1):
            ws_specs.cell(r, 1).value = k
            ws_specs.cell(r, 2).value = v

        ws_ic = wb.create_sheet("Initial Conditions")
        headers = [
            "Stage",
            "Temperature (F)",
            "Pressure (psia)",
            "Vapor Flow (lbmol/h)",
            "Liquid Flow (lbmol/h)",
            "Liquid Holdup (lbmol)",
            "Vapor Composition Component 1",
            "Vapor Composition Component 2",
            "Liquid Composition Component 1",
            "Liquid Composition Component 2",
        ]
        for c, h in enumerate(headers, start=1):
            ws_ic.cell(1, c).value = h
        rows = [
            (1, 100.0, 200.0, 10.0, 20.0, 5.0, 0.6, 0.4, 0.3, 0.7),
            (2, 120.0, 210.0, 11.0, 21.0, 6.0, 0.5, 0.5, 0.4, 0.6),
        ]
        for r, row in enumerate(rows, start=2):
            for c, v in enumerate(row, start=1):
                ws_ic.cell(r, c).value = v

        ws_comp = wb.create_sheet("Components")
        ws_comp.cell(1, 1).value = "Component Name"
        ws_comp.cell(2, 1).value = "Propane"
        ws_comp.cell(3, 1).value = "N-butane"

        wb.save(p)

        c = load_case_from_excel(str(p))
        assert c.specs["Runtime Mode"] == "hydraulic"
        assert c.specs["Thermo Mode"] == "table"
        assert c.specs["Thermo Table"] == "cache/thermo_table.json"
        assert c.specs["Include Energy"] is True
        assert c.specs["Condenser Duty Mode"] == "specified"
        assert c.specs["Equilibrium Relaxation Mode"] == "phase-holdup"
        assert c.specs["Equilibrium Tau (sec)"] == 4.0
        assert c.specs["Equilibrium Energy Damping Gain"] == 0.2
        assert c.specs["Equilibrium Relaxation Live PR"] is True
        assert c.specs["Hydraulic Energy Temperature Follow Tau (sec)"] == 0.5
        assert c.specs["Enable Level Control"] is True
        assert c.specs["Top Level PV Mode"] == "true-level"
        assert c.specs["Top Level SP Frac"] == 0.5
        assert c.specs["Top Level Kc"] == 0.5
        assert c.specs["Top Level Ti (sec)"] == 1200.0
        assert c.specs["Bottom Level PV Mode"] == "true-level"
        assert c.specs["Bottom Level SP (lbmol)"] == 794.0
        assert c.specs["Bottom Level SP Frac"] == 0.5
        assert c.specs["Bottom Level Kc"] == 8.0
        assert c.specs["Bottom Level Ti (sec)"] == 120.0
        assert c.specs["Enable Pressure Control"] is True
        assert c.specs["Pressure Control MV"] == "condenser-duty"
        assert c.specs["Top Pressure SP (psia)"] == 220.44
        assert c.specs["Top Pressure Kc"] == -150000.0
        assert c.specs["Top Pressure Ti (sec)"] == 120.0
        assert c.specs["Enable Distillate Composition Control"] is True
        assert c.specs["Distillate Composition Component"] == "C4"
        assert c.specs["Distillate Composition SP"] == 0.11


def test_load_case_from_excel_accepts_bottom_level_holdup_alias():
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "bottom_level_holdup_alias.xlsx"
        wb = Workbook()

        ws_specs = wb.active
        ws_specs.title = "Specifications"
        specs_rows = [
            ("Number of Stages", 2),
            ("Number of Components", 2),
            ("Simulation Length (min)", 1.0),
            ("Timestep (sec)", 1.0),
            ("Log Frequency (timesteps)", 1),
            ("Bottom Level Holdup (lbmol)", 794.0),
        ]
        for r, (k, v) in enumerate(specs_rows, start=1):
            ws_specs.cell(r, 1).value = k
            ws_specs.cell(r, 2).value = v

        ws_ic = wb.create_sheet("Initial Conditions")
        headers = [
            "Stage",
            "Temperature (F)",
            "Pressure (psia)",
            "Vapor Flow (lbmol/h)",
            "Liquid Flow (lbmol/h)",
            "Liquid Holdup (lbmol)",
            "Vapor Composition Component 1",
            "Vapor Composition Component 2",
            "Liquid Composition Component 1",
            "Liquid Composition Component 2",
        ]
        for c, h in enumerate(headers, start=1):
            ws_ic.cell(1, c).value = h
        rows = [
            (1, 100.0, 200.0, 10.0, 20.0, 5.0, 0.6, 0.4, 0.3, 0.7),
            (2, 120.0, 210.0, 11.0, 21.0, 6.0, 0.5, 0.5, 0.4, 0.6),
        ]
        for r, row in enumerate(rows, start=2):
            for c, v in enumerate(row, start=1):
                ws_ic.cell(r, c).value = v

        ws_comp = wb.create_sheet("Components")
        ws_comp.cell(1, 1).value = "Component Name"
        ws_comp.cell(2, 1).value = "Propane"
        ws_comp.cell(3, 1).value = "N-butane"

        wb.save(p)

        c = load_case_from_excel(str(p))
        assert c.specs["Bottom Holdup (lbmol)"] == 794.0
