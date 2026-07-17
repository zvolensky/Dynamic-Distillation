#!/usr/bin/env python
"""Create a dynamic-model seed workbook from a ChemSep result workbook."""

from __future__ import annotations

import argparse
import math
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]


def _norm(value: Any) -> str:
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())


def _as_float(value: Any) -> float | None:
    try:
        out = float(value)
    except (TypeError, ValueError):
        return None
    return out if math.isfinite(out) else None


def _find_soffice() -> Path:
    found = shutil.which("soffice")
    candidates = [
        Path(found) if found else None,
        Path(r"C:\Program Files\LibreOffice\program\soffice.exe"),
        Path(r"C:\Program Files (x86)\LibreOffice\program\soffice.exe"),
    ]
    for candidate in candidates:
        if candidate is not None and candidate.exists():
            return candidate
    raise RuntimeError("LibreOffice soffice was not found; convert the ChemSep .xls to .xlsx first.")


def _convert_source_to_xlsx(source: Path, output_dir: Path) -> Path:
    if source.suffix.lower() == ".xlsx":
        return source
    if source.suffix.lower() != ".xls":
        raise ValueError(f"Unsupported ChemSep workbook extension: {source.suffix}")
    subprocess.run(
        [
            str(_find_soffice()),
            "--headless",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(output_dir),
            str(source),
        ],
        check=True,
        capture_output=True,
        text=True,
    )
    converted = output_dir / f"{source.stem}.xlsx"
    if not converted.exists():
        raise RuntimeError(f"LibreOffice did not create {converted}")
    return converted


def _find_row(ws, label: str, *, column: int = 1) -> int:
    target = _norm(label)
    for row in range(1, ws.max_row + 1):
        if _norm(ws.cell(row, column).value) == target:
            return row
    raise KeyError(f"Could not find {label!r} in {ws.title}")


def _find_header_column(ws, label: str) -> int:
    target = _norm(label)
    for column in range(1, ws.max_column + 1):
        if _norm(ws.cell(1, column).value) == target:
            return column
    raise KeyError(f"Could not find column {label!r} in {ws.title}")


def _set_spec(ws, label: str, value: Any) -> None:
    try:
        row = _find_row(ws, label)
    except KeyError:
        row = ws.max_row + 1
        ws.cell(row, 1).value = label
    ws.cell(row, 2).value = value


def _read_stage_profile(ws) -> list[dict[str, float]]:
    rows: list[dict[str, float]] = []
    previous_stage = 0
    for row in range(1, ws.max_row + 1):
        stage_value = _as_float(ws.cell(row, 2).value)
        temperature = _as_float(ws.cell(row, 3).value)
        pressure = _as_float(ws.cell(row, 4).value)
        if stage_value is None or temperature is None or pressure is None:
            if rows:
                break
            continue
        stage = int(round(stage_value))
        if stage <= previous_stage:
            break
        rows.append(
            {
                "stage": float(stage),
                "temperature_F": temperature,
                "pressure_psia": pressure,
                "liquid_flow_lbmolph": _as_float(ws.cell(row, 5).value) or math.nan,
                "vapor_flow_lbmolph": _as_float(ws.cell(row, 6).value) or math.nan,
            }
        )
        previous_stage = stage
    if not rows:
        raise ValueError("No stage rows found on ChemSep T_P_Flow profiles sheet")
    return rows


def _read_composition_profile(ws) -> tuple[list[str], dict[int, list[float]]]:
    header_row = None
    for row in range(1, ws.max_row + 1):
        if _norm(ws.cell(row, 2).value) == "stage":
            header_row = row
            break
    if header_row is None:
        raise ValueError(f"No stage header found on {ws.title}")

    components: list[str] = []
    column = 3
    while column <= ws.max_column:
        value = ws.cell(header_row, column).value
        if value is None or not str(value).strip():
            break
        components.append(str(value).strip())
        column += 1
    if not components:
        raise ValueError(f"No component headers found on {ws.title}")

    profile: dict[int, list[float]] = {}
    for row in range(header_row + 1, ws.max_row + 1):
        stage_value = _as_float(ws.cell(row, 2).value)
        if stage_value is None:
            if profile:
                break
            continue
        values = [_as_float(ws.cell(row, 3 + index).value) for index in range(len(components))]
        if any(value is None for value in values):
            break
        total = sum(float(value) for value in values if value is not None)
        if total <= 0.0:
            raise ValueError(f"Non-positive composition sum at stage {stage_value} on {ws.title}")
        profile[int(round(stage_value))] = [float(value) / total for value in values if value is not None]
    return components, profile


def _read_streams(ws) -> tuple[list[str], dict[str, dict[str, Any]]]:
    stream_header_row = None
    stream_columns: dict[str, int] = {}
    for row in range(1, ws.max_row + 1):
        if _norm(ws.cell(row, 2).value) == "stream":
            stream_header_row = row
            for column in range(3, ws.max_column + 1):
                name = str(ws.cell(row, column).value or "").strip()
                if name:
                    stream_columns[name] = column
            break
    if stream_header_row is None or not stream_columns:
        raise ValueError("No stream header found on ChemSep Streams sheet")

    row_by_label: dict[str, int] = {}
    for row in range(stream_header_row + 1, ws.max_row + 1):
        label = _norm(ws.cell(row, 2).value)
        if label:
            row_by_label.setdefault(label, row)

    components: list[str] = []
    mole_flow_row = row_by_label.get("moleflowslbmolh")
    if mole_flow_row is None:
        raise ValueError("No mole-flow block found on ChemSep Streams sheet")
    row = mole_flow_row + 1
    while row <= ws.max_row:
        name = str(ws.cell(row, 2).value or "").strip()
        if not name:
            break
        components.append(name)
        row += 1

    streams: dict[str, dict[str, Any]] = {}
    for name, column in stream_columns.items():
        component_flows = {
            component: float(ws.cell(mole_flow_row + 1 + index, column).value)
            for index, component in enumerate(components)
        }
        streams[name] = {
            "stage": int(round(float(ws.cell(row_by_label["stage"], column).value))),
            "vapor_fraction": float(ws.cell(row_by_label["vapourfraction"], column).value),
            "temperature_F": float(ws.cell(row_by_label["temperaturef"], column).value),
            "total_molar_flow_lbmolph": float(ws.cell(row_by_label["totalmolarflowlbmolh"], column).value),
            "component_flows_lbmolph": component_flows,
        }
    return components, streams


def _read_duties(ws) -> tuple[float, float]:
    values: dict[str, float] = {}
    for row in range(1, ws.max_row + 1):
        label = _norm(ws.cell(row, 2).value)
        if label == "qcondenser":
            values[label] = float(ws.cell(row, 4).value)
        elif label == "qreboiler":
            values[label] = float(ws.cell(row, 4).value)
    if "qcondenser" not in values or "qreboiler" not in values:
        raise ValueError("ChemSep condenser/reboiler duties were not found")
    return values["qcondenser"], values["qreboiler"]


def _template_components(wb) -> list[str]:
    ws = wb["Specifications"]
    row = _find_row(ws, "Component Name")
    components: list[str] = []
    column = 2
    while column <= ws.max_column:
        value = ws.cell(row, column).value
        if value is None or not str(value).strip():
            break
        components.append(str(value).strip())
        column += 1
    return components


def _component_order(source: list[str], target: list[str]) -> list[int]:
    source_index = {_norm(name): index for index, name in enumerate(source)}
    missing = [name for name in target if _norm(name) not in source_index]
    if missing:
        raise ValueError(f"ChemSep output is missing template components: {missing}")
    return [source_index[_norm(name)] for name in target]


def _linear_pressure(stage_index: int, stage_count: int, top: float, bottom: float) -> float:
    if stage_count <= 1:
        return float(top)
    return float(top) + (float(bottom) - float(top)) * float(stage_index) / float(stage_count - 1)


def create_seed(
    *,
    source: Path,
    template: Path,
    output: Path,
    top_pressure_psia: float,
    bottom_pressure_psia: float,
    overwrite: bool,
) -> None:
    if output.exists() and not overwrite:
        raise FileExistsError(f"Output already exists: {output}; pass --overwrite to replace it")
    if bottom_pressure_psia < top_pressure_psia:
        raise ValueError("Bottom pressure must be greater than or equal to top pressure")

    with tempfile.TemporaryDirectory(prefix="chemsep_seed_") as temp_dir:
        source_xlsx = _convert_source_to_xlsx(source, Path(temp_dir))
        source_wb = load_workbook(source_xlsx, data_only=True, read_only=True)
        stage_profile = _read_stage_profile(source_wb["T_P_Flow profiles"])
        x_components, liquid_profile = _read_composition_profile(
            source_wb["Liquid x composition profiles"]
        )
        y_components, vapor_profile = _read_composition_profile(
            source_wb["Vapour y composition profiles"]
        )
        stream_components, streams = _read_streams(source_wb["Streams"])
        condenser_duty, reboiler_duty = _read_duties(source_wb["Mass and Energy Balances"])

        output.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(template, output)
        target_wb = load_workbook(output)
        target_components = _template_components(target_wb)
        x_order = _component_order(x_components, target_components)
        y_order = _component_order(y_components, target_components)
        stream_order = _component_order(stream_components, target_components)

        stage_count = len(stage_profile)
        if int(target_wb["Specifications"]["B2"].value) != stage_count:
            raise ValueError("ChemSep stage count does not match the template")

        ws_initial = target_wb["Initial Conditions"]
        columns = {
            str(ws_initial.cell(1, column).value): column
            for column in range(1, ws_initial.max_column + 1)
        }
        stage_rows = {
            int(round(float(ws_initial.cell(row, columns["Stage"]).value))): row
            for row in range(2, ws_initial.max_row + 1)
            if _as_float(ws_initial.cell(row, columns["Stage"]).value) is not None
        }
        for stage_index, source_row in enumerate(stage_profile):
            stage = int(source_row["stage"])
            row = stage_rows[stage]
            pressure = _linear_pressure(
                stage_index,
                stage_count,
                top_pressure_psia,
                bottom_pressure_psia,
            )
            ws_initial.cell(row, columns["Temperature (F)"]).value = source_row["temperature_F"]
            ws_initial.cell(row, columns["Pressure (psia)"]).value = pressure
            liquid_flow = source_row["liquid_flow_lbmolph"]
            vapor_flow = source_row["vapor_flow_lbmolph"]
            if math.isfinite(liquid_flow):
                ws_initial.cell(row, columns["Liquid Flow (lbmol/h)"]).value = liquid_flow
            if math.isfinite(vapor_flow):
                ws_initial.cell(row, columns["Vapor Flow (lbmol/h)"]).value = vapor_flow
            for target_index, source_index in enumerate(y_order, start=1):
                label = f"Vapor Composition Component {target_index}"
                ws_initial.cell(row, columns[label]).value = vapor_profile[stage][source_index]
            for target_index, source_index in enumerate(x_order, start=1):
                label = f"Liquid Composition Component {target_index}"
                ws_initial.cell(row, columns[label]).value = liquid_profile[stage][source_index]

        ws_specs = target_wb["Specifications"]
        _set_spec(ws_specs, "Condenser Duty (Btu/h)", condenser_duty)
        _set_spec(ws_specs, "Reboiler Duty (Btu/h)", reboiler_duty)
        _set_spec(ws_specs, "Source ChemSep Workbook", str(source.resolve()))
        _set_spec(
            ws_specs,
            "Source Notes",
            "ChemSep Excess Enthalpy seed; pressure overridden linearly for dynamic initialization.",
        )
        condenser_drop = 0.0 if stage_count <= 1 else (
            bottom_pressure_psia - top_pressure_psia
        ) / float(stage_count - 1)
        _set_spec(ws_specs, "Condenser Pressure Drop (psi)", condenser_drop)

        ws_streams = target_wb["Streams"]
        target_stream_columns = {
            str(ws_streams.cell(1, column).value).strip(): column
            for column in range(2, ws_streams.max_column + 1)
            if ws_streams.cell(1, column).value
        }
        target_stream_rows = {
            _norm(ws_streams.cell(row, 1).value): row
            for row in range(1, ws_streams.max_row + 1)
            if ws_streams.cell(row, 1).value
        }
        for stream_name, stream in streams.items():
            matching_name = next(
                (name for name in target_stream_columns if _norm(name) == _norm(stream_name)),
                None,
            )
            if matching_name is None:
                continue
            column = target_stream_columns[matching_name]
            stage = int(stream["stage"])
            pressure = _linear_pressure(
                stage - 1,
                stage_count,
                top_pressure_psia,
                bottom_pressure_psia,
            )
            ws_streams.cell(target_stream_rows["stage"], column).value = stage
            ws_streams.cell(target_stream_rows["pressurepsia"], column).value = pressure
            ws_streams.cell(target_stream_rows["vapourfraction"], column).value = stream["vapor_fraction"]
            ws_streams.cell(target_stream_rows["temperaturef"], column).value = stream["temperature_F"]
            ws_streams.cell(target_stream_rows["totalmolarflowlbmolh"], column).value = stream[
                "total_molar_flow_lbmolph"
            ]
            mole_flow_header = target_stream_rows["moleflowslbmolh"]
            for target_index, source_index in enumerate(stream_order):
                component = stream_components[source_index]
                ws_streams.cell(mole_flow_header + 1 + target_index, column).value = stream[
                    "component_flows_lbmolph"
                ][component]

        if "Notes" not in target_wb.sheetnames:
            ws_notes = target_wb.create_sheet("Notes")
            ws_notes.append(["Field", "Value"])
        else:
            ws_notes = target_wb["Notes"]
        ws_notes.append(["ChemSep source", str(source.resolve())])
        ws_notes.append(["Pressure override", f"Linear {top_pressure_psia:g} to {bottom_pressure_psia:g} psia"])
        ws_notes.append(["Pressure caution", "ChemSep temperatures and compositions were solved at flat pressure."])
        ws_notes.append(["Holdup caution", "Liquid and vapor holdups were preserved from the template."])

        target_wb.save(output)
        target_wb.close()
        source_wb.close()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", required=True, help="ChemSep .xls or .xlsx result workbook")
    parser.add_argument("--template", required=True, help="Existing dynamic-model .xlsx workbook")
    parser.add_argument("--output", required=True, help="New dynamic-model .xlsx workbook")
    parser.add_argument("--top-pressure-psia", type=float, required=True)
    parser.add_argument("--bottom-pressure-psia", type=float, required=True)
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args()

    source = Path(args.source).resolve()
    template = Path(args.template).resolve()
    output = Path(args.output).resolve()
    create_seed(
        source=source,
        template=template,
        output=output,
        top_pressure_psia=float(args.top_pressure_psia),
        bottom_pressure_psia=float(args.bottom_pressure_psia),
        overwrite=bool(args.overwrite),
    )
    print(f"created: {output}")
    print(f"pressure_profile_psia={args.top_pressure_psia:g}..{args.bottom_pressure_psia:g}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
