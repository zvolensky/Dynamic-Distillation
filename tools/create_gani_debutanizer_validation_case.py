from __future__ import annotations

from math import pi
from pathlib import Path

from openpyxl import Workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUT_PATH = PROJECT_ROOT / "validation_gani_1986_debutanizer.xlsx"

KMOL_TO_LBMOL = 2.20462262185
KPA_TO_PSIA = 0.14503773773
CM_TO_FT = 0.03280839895
GCAL_PER_H_TO_BTU_PER_H = 1.0e6 * 3.968320719


def _k_to_f(value_k: float) -> float:
    return (float(value_k) - 273.15) * 9.0 / 5.0 + 32.0


def _kmolh_to_lbmolh(value_kmolh: float) -> float:
    return float(value_kmolh) * KMOL_TO_LBMOL


def _kpa_to_psia(value_kpa: float) -> float:
    return float(value_kpa) * KPA_TO_PSIA


def _normalize(values: list[float]) -> list[float]:
    total = sum(float(v) for v in values)
    if total <= 0.0:
        raise ValueError("Cannot normalize non-positive composition total.")
    return [float(v) / total for v in values]


def _linear_profile(top: list[float], bottom: list[float], frac: float) -> list[float]:
    return _normalize([t + (b - t) * frac for t, b in zip(top, bottom)])


def _horizontal_drum_from_volume(total_volume_ft3: float, length_to_diameter: float = 3.0) -> tuple[float, float]:
    diameter_ft = (4.0 * float(total_volume_ft3) / (pi * float(length_to_diameter))) ** (1.0 / 3.0)
    return diameter_ft, float(length_to_diameter) * diameter_ft


def _vertical_cylinder_from_volume(total_volume_ft3: float, height_to_diameter: float = 2.0) -> tuple[float, float]:
    diameter_ft = (4.0 * float(total_volume_ft3) / (pi * float(height_to_diameter))) ** (1.0 / 3.0)
    return diameter_ft, float(height_to_diameter) * diameter_ft


def _liquid_design_volume_ft3(
    flow_lbmolh: float,
    mixture_mw_lb_per_lbmol: float,
    liquid_density_lbft3: float,
    residence_min: float = 10.0,
    liquid_fraction: float = 0.5,
) -> tuple[float, float]:
    liquid_flow_ft3h = float(flow_lbmolh) * float(mixture_mw_lb_per_lbmol) / float(liquid_density_lbft3)
    liquid_volume_ft3 = liquid_flow_ft3h * float(residence_min) / 60.0
    return liquid_volume_ft3, liquid_volume_ft3 / float(liquid_fraction)


def build_workbook(path: Path = OUT_PATH) -> Path:
    n_stages = 28
    feed_stage = 23
    n_components = 6
    component_names = [
        "1,3-butadiene",
        "Isobutene",
        "N-pentane",
        "1-pentene",
        "1-hexene",
        "Benzene",
    ]

    reflux_lbmolh = _kmolh_to_lbmolh(429.8)
    distillate_lbmolh = _kmolh_to_lbmolh(236.86)
    bottoms_lbmolh = _kmolh_to_lbmolh(266.18)
    source_feed_liq_kmolh = 427.8
    source_feed_vap_kmolh = 68.5
    source_feed_total_kmolh = source_feed_liq_kmolh + source_feed_vap_kmolh
    reference_product_total_kmolh = 236.86 + 266.18
    feed_vapor_fraction = source_feed_vap_kmolh / source_feed_total_kmolh
    # Table 1b feed and Table 2a reference product rates are not exactly
    # mass-balanced. For a dynamic open-loop reference search, preserve the
    # published product rates and scale total feed to D+B.
    feed_total_lbmolh = _kmolh_to_lbmolh(reference_product_total_kmolh)
    feed_vap_lbmolh = feed_total_lbmolh * feed_vapor_fraction
    feed_liq_lbmolh = feed_total_lbmolh - feed_vap_lbmolh
    reboiler_duty_btuh = 2.760 * GCAL_PER_H_TO_BTU_PER_H

    feed_z = _normalize([0.23791, 0.30817, 0.09959, 0.13727, 0.08872, 0.12834])
    top_x = _normalize([0.41061, 0.58937, 0.0, 0.00001, 0.0, 0.0])
    bottom_x = _normalize([0.08856, 0.06511, 0.18754, 0.25586, 0.16546, 0.23927])

    top_temp_f = _k_to_f(320.10)
    bottom_temp_f = _k_to_f(365.40)
    top_pressure_psia = _kpa_to_psia(527.50)
    bottom_pressure_psia = _kpa_to_psia(555.30)

    # Density estimates from the Python thermo package's liquid-density
    # correlation at the Gani Table 2a reference product states. These are not
    # published Gani vessel data; they only support the 10 min design sizing.
    top_mw = 55.27871150291503
    bottom_mw = 72.40284026192853
    top_liq_density_lbft3 = 569.9024948119153 * 0.0624279605761
    bottom_liq_density_lbft3 = 603.1378496659061 * 0.0624279605761

    top_liq_vol_ft3, top_total_vol_ft3 = _liquid_design_volume_ft3(
        distillate_lbmolh, top_mw, top_liq_density_lbft3
    )
    bottom_liq_vol_ft3, bottom_total_vol_ft3 = _liquid_design_volume_ft3(
        bottoms_lbmolh, bottom_mw, bottom_liq_density_lbft3
    )
    top_drum_diam_ft, top_drum_len_ft = _horizontal_drum_from_volume(top_total_vol_ft3)
    bottom_sump_diam_ft, bottom_sump_height_ft = _vertical_cylinder_from_volume(bottom_total_vol_ft3)

    top_holdup_lbmol = distillate_lbmolh * (10.0 / 60.0)
    bottom_holdup_lbmol = bottoms_lbmolh * (10.0 / 60.0)

    column_diameter_ft = 5.97
    tray_spacing_ft = 50.0 * CM_TO_FT
    active_area_ft2 = 21.06
    column_cross_section_ft2 = pi / 4.0 * column_diameter_ft**2
    active_area_frac = active_area_ft2 / column_cross_section_ft2
    weir_height_in = 0.164 * 12.0
    weir_length_ft = 40.61 / 12.0

    wb = Workbook()

    ws = wb.active
    ws.title = "Specifications"
    specs = [
        ("Parameter", "Value"),
        ("Number of Stages", n_stages),
        ("Number of Components", n_components),
        ("Condenser Type", "Total"),
        ("Component Name", *component_names),
        ("Thermo Mode", "clapeyron"),
        ("Runtime Mode", "hydraulic"),
        ("Include Energy", True),
        ("Condenser Duty Mode", "specified"),
        ("Reboiler Duty (Btu/h)", reboiler_duty_btuh),
        ("Simulation Length (min)", 90.0),
        ("Timestep (sec)", 1.0),
        ("Log Frequency (timesteps)", 60),
        ("Top Accumulator Holdup (lbmol)", top_holdup_lbmol),
        ("Bottom Holdup (lbmol)", bottom_holdup_lbmol),
        ("Top Drum Total Volume (ft3)", top_total_vol_ft3),
        ("Top Drum Diameter (ft)", top_drum_diam_ft),
        ("Top Drum Length (ft)", top_drum_len_ft),
        ("Top Drum Liquid Fraction (-)", 0.5),
        ("Overhead Vapor Line Volume (ft3)", 0.0),
        ("Condenser Vapor Volume (ft3)", 0.0),
        ("Bottom Sump Total Volume (ft3)", bottom_total_vol_ft3),
        ("Bottom Sump Diameter (ft)", bottom_sump_diam_ft),
        ("Bottom Sump Height (ft)", bottom_sump_height_ft),
        ("Bottom Sump Liquid Fraction (-)", 0.5),
        ("Pressure Model", "hydraulic"),
        ("Vapor Flow Model", "energy"),
        ("Stage time constant [tau] (sec)", 10.0),
        ("Condenser Pressure Drop (psi)", 0.0),
        ("Equilibrium Relaxation Mode", "composition-only"),
        ("Equilibrium Tau (sec)", 5.0),
        ("Equilibrium Energy Damping Gain", 0.0),
        ("Equilibrium Relaxation Live PR", False),
        ("Vapor Holdup Relaxation (sec)", 0.0),
        ("Vapor Flow Relaxation (sec)", 5.0),
        ("Feed Source", "Gani, Ruiz, and Cameron 1986, Problem II industrial debutanizer"),
        ("Source Disturbance", "+5% reflux rate at constant reboiler duty after steady-state search"),
        ("Source Feed Liquid (kmol/h)", source_feed_liq_kmolh),
        ("Source Feed Vapor (kmol/h)", source_feed_vap_kmolh),
        ("Mass-Balanced Feed Total (lbmol/h)", feed_total_lbmolh),
        ("Residence Time for Derived Vessels (min)", 10.0),
        ("Derived Vessel Normal Liquid Fraction (-)", 0.5),
        (None, None),
        ("Start Stage", "End Stage", "Diameter (ft)", "Tray Spacing (ft)", "Gas Void Fraction", "Weir Height", "Weir Length", "Active Area"),
        (1, n_stages, column_diameter_ft, tray_spacing_ft, 0.80, weir_height_in, weir_length_ft, active_area_frac),
    ]
    for r, row in enumerate(specs, start=1):
        for c, value in enumerate(row, start=1):
            ws.cell(r, c).value = value

    ws = wb.create_sheet("Initial Conditions")
    headers = [
        "Stage",
        "Temperature (F)",
        "Pressure (psia)",
        "Vapor Flow (lbmol/h)",
        "Liquid Flow (lbmol/h)",
        "Liquid Holdup (lbmol)",
        "Vapor Holdup (lbmol)",
    ]
    headers += [f"Vapor Composition Component {i}" for i in range(1, n_components + 1)]
    headers += [f"Liquid Composition Component {i}" for i in range(1, n_components + 1)]
    for c, header in enumerate(headers, start=1):
        ws.cell(1, c).value = header

    vapor_above_feed = reflux_lbmolh + distillate_lbmolh
    vapor_below_feed = max(vapor_above_feed - feed_vap_lbmolh, 0.0)
    liquid_above_feed = reflux_lbmolh
    liquid_below_feed = reflux_lbmolh + feed_liq_lbmolh
    tray_liq_holdup = 5.0
    tray_vap_holdup = 0.05

    for stage in range(1, n_stages + 1):
        frac = (stage - 1) / max(n_stages - 1, 1)
        x = _linear_profile(top_x, bottom_x, frac)
        y = _linear_profile(top_x, bottom_x, max(frac - 0.05, 0.0))
        temp_f = top_temp_f + (bottom_temp_f - top_temp_f) * frac
        pressure_psia = top_pressure_psia + (bottom_pressure_psia - top_pressure_psia) * frac
        vapor_flow = 0.0 if stage == 1 else (vapor_above_feed if stage < feed_stage else vapor_below_feed)
        liquid_flow = liquid_above_feed if stage < feed_stage else liquid_below_feed
        row = [
            stage,
            temp_f,
            pressure_psia,
            vapor_flow,
            liquid_flow,
            tray_liq_holdup,
            0.0 if stage == 1 else tray_vap_holdup,
            *y,
            *x,
        ]
        for c, value in enumerate(row, start=1):
            ws.cell(stage + 1, c).value = value

    ws = wb.create_sheet("Streams")
    stream_rows = [
        ("Stream", "Feed", "Distillate", "Bottom"),
        ("Stage", feed_stage, 1, n_stages),
        ("Pressure (psia)", top_pressure_psia + (bottom_pressure_psia - top_pressure_psia) * ((feed_stage - 1) / (n_stages - 1)), top_pressure_psia, bottom_pressure_psia),
        ("Vapour fraction", feed_vap_lbmolh / feed_total_lbmolh, 0.0, 0.0),
        ("Temperature (F)", _k_to_f(338.0), top_temp_f, bottom_temp_f),
        ("Total molar flow (lbmol/h)", feed_total_lbmolh, distillate_lbmolh, bottoms_lbmolh),
        ("Mole flows (lbmol/h)", None, None, None),
    ]
    for name, z in zip(component_names, feed_z):
        stream_rows.append((name, z * feed_total_lbmolh, None, None))
    for r, row in enumerate(stream_rows, start=1):
        for c, value in enumerate(row, start=1):
            ws.cell(r, c).value = value

    ws = wb.create_sheet("Components")
    ws.cell(1, 1).value = "Component Name"
    for r, name in enumerate(component_names, start=2):
        ws.cell(r, 1).value = name

    ws = wb.create_sheet("Boundary State")
    boundary_rows = [
        ("State", *component_names),
        ("top_L", *[top_holdup_lbmol * z for z in top_x]),
        ("top_V", *([0.0] * n_components)),
        ("bottom_L", *[bottom_holdup_lbmol * z for z in bottom_x]),
        ("bottom_V", *([0.0] * n_components)),
    ]
    for r, row in enumerate(boundary_rows, start=1):
        for c, value in enumerate(row, start=1):
            ws.cell(r, c).value = value

    ws = wb.create_sheet("Notes")
    notes = [
        ("Field", "Value"),
        ("Source", "Gani, Ruiz, and Cameron 1986, A generalized model for distillation columns-I, Problem II"),
        ("Units", "Workbook values are British units: lbmol/h, Btu/h, deg F, psia, ft, ft3."),
        ("Thermo", "Original paper does not name the exact thermo package for Problem II; use --thermo clapeyron --clapeyron-model PR or SRK for probes."),
        ("Tray efficiency treatment", "Source lists 40 actual trays at 70% efficiency and 28 effective plates simulated; workbook uses 28 modeled stages."),
        ("Feed stage mapping", "Actual feed tray 33 scaled to effective stage 23."),
        ("Feed mass-balance adjustment", "Table 1b feed totals 496.3 kmol/h while Table 2a reference D+B totals 503.04 kmol/h. Workbook preserves Table 2a D/B and scales feed to D+B using the published feed vapor fraction."),
        ("Vessel sizing", "Top drum and bottom sump volumes are derived, not source-published."),
        ("Residence time", "10 min at reference distillate/bottoms product flow."),
        ("Normal liquid fraction", "50%."),
        ("Column active area", f"{active_area_ft2:.6g} ft2, written to geometry as fraction {active_area_frac:.6g}"),
        ("Weir height", f"0.164 ft = {weir_height_in:.6g} in; workbook geometry parser expects inches."),
        ("Top liquid density used", f"{top_liq_density_lbft3:.6g} lb/ft3"),
        ("Bottom liquid density used", f"{bottom_liq_density_lbft3:.6g} lb/ft3"),
        ("Top drum sizing model", "Horizontal cylinder, L/D = 3."),
        ("Bottom sump sizing model", "Vertical cylinder, H/D = 2."),
        ("Top liquid volume at 10 min", top_liq_vol_ft3),
        ("Top total volume", top_total_vol_ft3),
        ("Bottom liquid volume at 10 min", bottom_liq_vol_ft3),
        ("Bottom total volume", bottom_total_vol_ft3),
        ("Disturbance to test", "Increase reflux from 947.55 to 994.92 lbmol/h at constant reboiler duty after a steady-state search."),
    ]
    for r, row in enumerate(notes, start=1):
        for c, value in enumerate(row, start=1):
            ws.cell(r, c).value = value

    path = Path(path)
    wb.save(path)
    return path


if __name__ == "__main__":
    print(build_workbook())
