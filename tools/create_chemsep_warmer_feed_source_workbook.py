#!/usr/bin/env python
"""
Create a source-aligned C3/C4/C5 workbook from Depropanizer_warmer_feed.sep.

This preserves model-specific runtime settings, geometry, and holdups from an
existing workbook, while replacing the ChemSep-derived steady profiles,
streams, and duties with values parsed from the exact .sep result block.
"""

from __future__ import annotations

import argparse
from pathlib import Path
import sys

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from tools.audit_chemsep_warmer_feed_parity import (  # noqa: E402
    duty_to_btu_h,
    flow_to_lbmol_h,
    k_to_f,
    pa_to_psia,
    parse_sep,
)


def _resolve(path: str | Path) -> Path:
    p = Path(path)
    if p.is_absolute():
        return p
    return PROJECT_ROOT / p


def _set_spec(ws, key: str, value) -> None:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value == key:
            row[1].value = value
            return
    raise KeyError(f"specification row not found: {key}")


def _initial_headers(ws) -> dict[str, int]:
    return {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], start=1) if cell.value is not None}


def _stream_column(ws, name: str) -> int:
    target = name.strip().lower()
    for cell in ws[1]:
        if cell.value is not None and str(cell.value).strip().lower() == target:
            return int(cell.column)
    raise KeyError(f"stream column not found: {name}")


def _stream_row(ws, label: str) -> int:
    target = label.strip().lower()
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value is not None and str(row[0].value).strip().lower() == target:
            return int(row[0].row)
    raise KeyError(f"stream row not found: {label}")


def _component_stream_row(ws, component: str) -> int:
    target = component.strip().lower()
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value is not None and str(row[0].value).strip().lower() == target:
            return int(row[0].row)
    raise KeyError(f"component stream row not found: {component}")


def create_workbook(*, sep_path: Path, template_path: Path, output_path: Path) -> Path:
    sep = parse_sep(sep_path)
    wb = load_workbook(template_path)

    ws_spec = wb["Specifications"]
    ws_init = wb["Initial Conditions"]
    ws_streams = wb["Streams"]

    _set_spec(ws_spec, "Condenser Duty (Btu/h)", duty_to_btu_h(sep.condenser_duty_w))
    _set_spec(ws_spec, "Reboiler Duty (Btu/h)", duty_to_btu_h(sep.reboiler_duty_w))

    headers = _initial_headers(ws_init)
    for profile in sep.profiles:
        stage = int(profile["stage"])
        row_idx = stage + 1
        ws_init.cell(row=row_idx, column=headers["Temperature (F)"]).value = k_to_f(profile["temperature_k"])
        ws_init.cell(row=row_idx, column=headers["Pressure (psia)"]).value = pa_to_psia(profile["pressure_pa"])

        vapor_flow = flow_to_lbmol_h(profile["vapor_flow_basis"])
        liquid_flow = flow_to_lbmol_h(profile["liquid_flow_basis"])

        # Model topology mapping: total condenser vapor outlet is represented by
        # the top boundary/product stream, not a tray-1 vapor traffic term.
        if stage == 1:
            vapor_flow = 0.0
        # Model topology mapping: bottoms product leaves through the reboiler/
        # bottom boundary, not as a tray-20 downflow in the tray profile.
        if stage == sep.n_stages:
            liquid_flow = 0.0

        ws_init.cell(row=row_idx, column=headers["Vapor Flow (lbmol/h)"]).value = vapor_flow
        ws_init.cell(row=row_idx, column=headers["Liquid Flow (lbmol/h)"]).value = liquid_flow

        for comp_idx in range(len(sep.feed_component_flows_basis)):
            ws_init.cell(row=row_idx, column=headers[f"Vapor Composition Component {comp_idx + 1}"]).value = sep.vapor_x[stage - 1][comp_idx]
            ws_init.cell(row=row_idx, column=headers[f"Liquid Composition Component {comp_idx + 1}"]).value = sep.liquid_x[stage - 1][comp_idx]

    feed_col = _stream_column(ws_streams, "Feed")
    dist_col = _stream_column(ws_streams, "Distillate")
    bottom_col = _stream_column(ws_streams, "Bottom")

    # Stream scalar rows.
    ws_streams.cell(row=_stream_row(ws_streams, "Pressure (psia)"), column=feed_col).value = pa_to_psia(sep.feed_pressure_pa)
    ws_streams.cell(row=_stream_row(ws_streams, "Temperature (F)"), column=feed_col).value = k_to_f(sep.feed_temperature_k)
    ws_streams.cell(row=_stream_row(ws_streams, "Total molar flow (lbmol/h)"), column=feed_col).value = flow_to_lbmol_h(sum(sep.feed_component_flows_basis))

    ws_streams.cell(row=_stream_row(ws_streams, "Pressure (psia)"), column=dist_col).value = pa_to_psia(float(sep.top_product["pressure_pa"]))
    ws_streams.cell(row=_stream_row(ws_streams, "Temperature (F)"), column=dist_col).value = k_to_f(float(sep.top_product["temperature_k"]))
    ws_streams.cell(row=_stream_row(ws_streams, "Total molar flow (lbmol/h)"), column=dist_col).value = flow_to_lbmol_h(float(sep.top_product["flow_rate_basis"]))

    ws_streams.cell(row=_stream_row(ws_streams, "Pressure (psia)"), column=bottom_col).value = pa_to_psia(float(sep.bottom_product["pressure_pa"]))
    ws_streams.cell(row=_stream_row(ws_streams, "Temperature (F)"), column=bottom_col).value = k_to_f(float(sep.bottom_product["temperature_k"]))
    ws_streams.cell(row=_stream_row(ws_streams, "Total molar flow (lbmol/h)"), column=bottom_col).value = flow_to_lbmol_h(float(sep.bottom_product["flow_rate_basis"]))

    component_rows = {
        0: _component_stream_row(ws_streams, "n-Propane"),
        1: _component_stream_row(ws_streams, "n-Butane"),
        2: _component_stream_row(ws_streams, "N-Pentane"),
    }
    top_x = list(sep.top_product["x"])
    bottom_x = list(sep.bottom_product["x"])
    top_flow = float(sep.top_product["flow_rate_basis"])
    bottom_flow = float(sep.bottom_product["flow_rate_basis"])
    for comp_idx, basis_flow in enumerate(sep.feed_component_flows_basis):
        row = component_rows[comp_idx]
        ws_streams.cell(row=row, column=feed_col).value = flow_to_lbmol_h(basis_flow)
        ws_streams.cell(row=row, column=dist_col).value = flow_to_lbmol_h(top_flow * float(top_x[comp_idx]))
        ws_streams.cell(row=row, column=bottom_col).value = flow_to_lbmol_h(bottom_flow * float(bottom_x[comp_idx]))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Create an exact ChemSep warmer-feed source workbook.")
    parser.add_argument("--sep", default=r"d:\Users\Thoma\Documents\Depropanizer_warmer_feed.sep")
    parser.add_argument("--template", default="logs/c3c4_splitter_openloop_seed_20260526.xlsx")
    parser.add_argument("--output", default="logs/c3c4_depropanizer_chemsep_warmer_feed_pr76_source_20260531.xlsx")
    args = parser.parse_args()

    out = create_workbook(
        sep_path=_resolve(args.sep),
        template_path=_resolve(args.template),
        output_path=_resolve(args.output),
    )
    print(out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
