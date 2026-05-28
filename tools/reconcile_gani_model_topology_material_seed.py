#!/usr/bin/env python
"""Create a material-reconciled Gani full-topology PR seed workbook.

This pass is intentionally narrow.  It keeps the corrected explicit-sump liquid
flow profile and the current vapor-composition profile, then recomputes tray
liquid compositions stage-by-stage so component material balances close under
the full model topology with explicit top/bottom boundary states.
"""

from __future__ import annotations

import argparse
from pathlib import Path
import sys
from typing import Any

import numpy as np
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_PATH = PROJECT_ROOT / "src"
if str(SRC_PATH) not in sys.path:
    sys.path.insert(0, str(SRC_PATH))

from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case  # noqa: E402
from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel  # noqa: E402


def _resolve_path(raw: str) -> Path:
    p = Path(raw)
    if p.is_absolute():
        return p
    return PROJECT_ROOT / p


def _safe_norm(v: np.ndarray, fallback: np.ndarray | None = None) -> np.ndarray:
    arr = np.asarray(v, dtype=float).reshape((-1,))
    arr = np.where(np.isfinite(arr), arr, 0.0)
    arr = np.clip(arr, 0.0, None)
    s = float(np.sum(arr))
    if s > 1.0e-300:
        return arr / s
    if fallback is not None:
        return _safe_norm(fallback)
    return np.full(arr.size, 1.0 / float(arr.size), dtype=float)


def _find_header(ws: Any, header: str) -> tuple[int, int]:
    target = str(header).strip().lower()
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value).strip().lower() == target:
                return int(cell.row), int(cell.column)
    raise ValueError(f"Could not find header {header!r} in sheet {ws.title!r}")


def _component_columns(ws: Any, prefix: str, n_components: int) -> list[int]:
    cols: list[int] = []
    for k in range(1, int(n_components) + 1):
        _, col = _find_header(ws, f"{prefix} Component {k}")
        cols.append(col)
    return cols


def _append_note(wb: Any, field: str, value: str) -> None:
    ws = wb["Notes"] if "Notes" in wb.sheetnames else wb.create_sheet("Notes")
    if ws.max_row < 1:
        ws.cell(1, 1).value = "Field"
        ws.cell(1, 2).value = "Value"
    row = int(ws.max_row) + 1
    ws.cell(row, 1).value = field
    ws.cell(row, 2).value = value


def _stream_array(stream: Any, component_names: list[str], default: np.ndarray) -> np.ndarray:
    flows = getattr(stream, "component_molar_flows_lbmolph", None)
    if not flows:
        return np.asarray(default, dtype=float).reshape((-1,))
    vals = np.array([float(flows.get(c, 0.0)) for c in component_names], dtype=float)
    return vals


def main() -> int:
    ap = argparse.ArgumentParser(description="Reconcile Gani full-topology seed liquid compositions for material balance.")
    ap.add_argument("--input", default="validation_gani_1986_debutanizer_model_topology_seed.xlsx")
    ap.add_argument("--output", default="validation_gani_1986_debutanizer_model_topology_material_reconciled.xlsx")
    ap.add_argument("--clip-negative", action="store_true", help="Clip negative component values and renormalize if encountered.")
    ap.add_argument(
        "--reconcile-vapor-profile",
        action="store_true",
        help=(
            "Also recompute tray vapor compositions from bottom-up convective "
            "vapor balances before reconciling liquid compositions."
        ),
    )
    args = ap.parse_args()

    in_path = _resolve_path(str(args.input))
    out_path = _resolve_path(str(args.output))
    case = load_case_from_excel(str(in_path))
    col = build_column_spec_from_case(case)

    n = int(col.n_stages)
    nc = int(col.n_components)
    comps = [str(c) for c in col.components_excel]
    L = np.asarray(col.L_lbmolph, dtype=float).reshape((n,))
    V = np.asarray(col.V_lbmolph, dtype=float).reshape((n,)).copy()
    V[0] = 0.0
    y_old = np.asarray(col.y0, dtype=float).reshape((n, nc)).copy()
    y = y_old.copy()
    x_old = np.asarray(col.x0, dtype=float).reshape((n, nc)).copy()

    streams = getattr(col, "streams", {})
    feed = streams.get("Feed")
    dist = streams.get("Distillate")
    bottom = streams.get("Bottom")
    feed_stage0 = int(getattr(feed, "stage_1based", 1) or 1) - 1 if feed is not None else -1
    feed_total = float(getattr(feed, "total_molar_flow_lbmolph", 0.0) or 0.0) if feed is not None else 0.0
    feed_vf = float(getattr(feed, "vapor_fraction", 0.0) or 0.0) if feed is not None else 0.0
    z_feed = _safe_norm(_stream_array(feed, comps, np.ones(nc, dtype=float) / nc)) if feed is not None else np.zeros(nc)
    feed_L = feed_total * max(0.0, min(1.0, 1.0 - feed_vf)) * z_feed
    feed_V = feed_total * max(0.0, min(1.0, feed_vf)) * z_feed

    top_L0 = getattr(col, "top_L0_lbmol", None)
    if top_L0 is not None:
        x_top = _safe_norm(np.asarray(top_L0, dtype=float).reshape((nc,)), fallback=x_old[0])
    elif dist is not None:
        x_top = _safe_norm(_stream_array(dist, comps, x_old[0]), fallback=x_old[0])
    else:
        x_top = x_old[0].copy()

    x_new = x_old.copy()
    min_raw = 0.0
    max_abs_delta = 0.0
    max_abs_y_delta = 0.0
    max_balance_abs = 0.0

    if bool(args.reconcile_vapor_profile):
        # Stage N vapor composition is the reboiler vapor basis.  March upward
        # so each tray vapor state has dV ~= 0 under fixed profile vapor flows.
        y[-1] = _safe_norm(y[-1], fallback=y_old[-1])
        for i in range(n - 2, 0, -1):
            V_in = float(V[i + 1])
            rhs = V_in * y[i + 1]
            if feed_stage0 == i:
                rhs = rhs + feed_V
            denom = float(V[i])
            if denom <= 0.0:
                raise ValueError(f"Stage {i + 1} has nonpositive vapor flow; cannot reconcile vapor composition.")
            raw_y = rhs / denom
            if np.any(raw_y < -1.0e-10) and not bool(args.clip_negative):
                raise ValueError(
                    f"Stage {i + 1} vapor reconciliation produced negative component(s): {raw_y}. "
                    "Rerun with --clip-negative to create a bounded approximation."
                )
            y[i] = _safe_norm(raw_y, fallback=y_old[i])
        if n > 1:
            y[0] = y[1].copy()
        max_abs_y_delta = float(np.max(np.abs(y - y_old)))

    # In explicit-top topology, stage 1 is a condenser-transfer holdup.  Its
    # liquid composition should match the vapor being condensed from stage 2.
    if n > 1:
        x_new[0] = _safe_norm(y[1], fallback=x_old[0])
        x_top = x_new[0].copy()
        max_abs_delta = max(float(max_abs_delta), float(np.max(np.abs(x_new[0] - x_old[0]))))

    # Stage 1 is a condenser-transfer stage when explicit top states are used.
    # Reconcile stages 2..N with the same convective component balance used by
    # the RHS under profile-flow/no-equilibrium parity conditions.
    for i in range(1, n):
        L_in = float(L[i - 1])
        x_in = x_top if i == 1 else x_new[i - 1]
        V_in = float(V[i + 1]) if i < n - 1 else float(V[i])
        y_in = y[i + 1] if i < n - 1 else y[i]
        fL = feed_L if feed_stage0 == i else np.zeros(nc, dtype=float)
        fV = feed_V if feed_stage0 == i else np.zeros(nc, dtype=float)
        rhs = L_in * x_in + fL + V_in * y_in + fV - float(V[i]) * y[i]
        denom = float(L[i])
        if denom <= 0.0:
            raise ValueError(f"Stage {i + 1} has nonpositive liquid flow; cannot reconcile liquid composition.")
        raw = rhs / denom
        min_raw = min(float(min_raw), float(np.min(raw)))
        if np.any(raw < -1.0e-10) and not bool(args.clip_negative):
            raise ValueError(
                f"Stage {i + 1} reconciliation produced negative composition component(s): {raw}. "
                "Rerun with --clip-negative to create a bounded approximation."
            )
        x_new[i] = _safe_norm(raw, fallback=x_old[i])
        max_abs_delta = max(float(max_abs_delta), float(np.max(np.abs(x_new[i] - x_old[i]))))
        resid = L_in * x_in + fL + V_in * y_in + fV - float(V[i]) * y[i] - denom * x_new[i]
        max_balance_abs = max(float(max_balance_abs), float(np.max(np.abs(resid))))

    wb = load_workbook(in_path)
    ws_ic = wb["Initial Conditions"]
    stage_hdr_row, stage_col = _find_header(ws_ic, "Stage")
    x_cols = _component_columns(ws_ic, "Liquid Composition", nc)
    y_cols = _component_columns(ws_ic, "Vapor Composition", nc)

    stage_to_row: dict[int, int] = {}
    for row in range(stage_hdr_row + 1, ws_ic.max_row + 1):
        val = ws_ic.cell(row, stage_col).value
        if val is None:
            continue
        try:
            stage_to_row[int(float(val))] = row
        except Exception:
            continue
    for i in range(n):
        row = stage_to_row.get(i + 1)
        if row is None:
            raise ValueError(f"Missing Initial Conditions row for stage {i + 1}.")
        for k, col_idx in enumerate(x_cols):
            ws_ic.cell(row, col_idx).value = float(x_new[i, k])
        for k, col_idx in enumerate(y_cols):
            ws_ic.cell(row, col_idx).value = float(y[i, k])

    # Keep explicit boundary liquid compositions consistent with the reconciled
    # terminal transfer holdups.
    if "Boundary State" in wb.sheetnames:
        ws_b = wb["Boundary State"]
        state_col = 1
        top_row = None
        bottom_row = None
        for row in range(1, ws_b.max_row + 1):
            label = str(ws_b.cell(row, state_col).value).strip().lower()
            if label == "top_l":
                top_row = row
            elif label == "bottom_l":
                bottom_row = row
        if top_row is not None and getattr(col, "top_L0_lbmol", None) is not None:
            top_total = float(np.sum(np.asarray(col.top_L0_lbmol, dtype=float).reshape((nc,))))
            for k in range(nc):
                ws_b.cell(top_row, k + 2).value = float(top_total * x_new[0, k])
        if bottom_row is not None and getattr(col, "bottom_L0_lbmol", None) is not None:
            bottom_total = float(np.sum(np.asarray(col.bottom_L0_lbmol, dtype=float).reshape((nc,))))
            for k in range(nc):
                ws_b.cell(bottom_row, k + 2).value = float(bottom_total * x_new[-1, k])

    _append_note(
        wb,
        "Model-topology material reconciliation",
        (
            "Recomputed stage liquid compositions 2..N from fixed PR vapor profile, "
            "corrected liquid-flow profile, feed split, and explicit top/bottom boundary topology; "
            "stage 1/top liquid were aligned to incoming condensed vapor. "
            f"max_abs_liquid_composition_delta={max_abs_delta:.9g}; "
            f"max_abs_vapor_composition_delta={max_abs_y_delta:.9g}; "
            f"min_raw_component_before_clip={min_raw:.9g}; "
            f"max_component_balance_residual_lbmolph={max_balance_abs:.9g}."
        ),
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Wrote {out_path}")
    print(f"max_abs_liquid_composition_delta: {max_abs_delta:.9g}")
    print(f"max_abs_vapor_composition_delta: {max_abs_y_delta:.9g}")
    print(f"min_raw_component_before_clip: {min_raw:.9g}")
    print(f"max_component_balance_residual_lbmolph: {max_balance_abs:.9g}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
