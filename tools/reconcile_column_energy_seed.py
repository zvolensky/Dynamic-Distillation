#!/usr/bin/env python
"""
Write an Energy State sheet that is internally consistent with column_rhs energy transport.

This is a narrow initializer aid: it evaluates the current workbook once, uses
the model's own L/V flows, duties, feed enthalpy, and holdups, then solves a
linear least-squares problem for tray liquid/vapor specific enthalpies that
minimize dEL and dEV. The resulting EL/EV inventories are written to a copy of
the workbook for follow-up residual audits.
"""

from __future__ import annotations

import argparse
import json
import math
from dataclasses import replace
from pathlib import Path
import sys

import numpy as np
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_PATH = PROJECT_ROOT / "src"
if str(SRC_PATH) not in sys.path:
    sys.path.insert(0, str(SRC_PATH))

from dynamic_distillation.column_rhs_v1 import column_rhs  # noqa: E402
from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case  # noqa: E402
from dynamic_distillation.dynamic_run_scaffold_v1 import (  # noqa: E402
    RunnerConfig,
    _clear_initial_tray_vapor_holdup,
    _initialize_vapor_holdup_from_spec_pressure,
    build_inputs_for_runner,
)
from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel  # noqa: E402
from dynamic_distillation.state_vector_layout_v1 import StateVectorLayout  # noqa: E402


def _resolve(raw: str | Path) -> Path:
    p = Path(raw)
    if p.is_absolute():
        return p.resolve()
    return (PROJECT_ROOT / p).resolve()


def _as_vec(diag: dict, key: str, n: int, default: float = 0.0) -> np.ndarray:
    arr = np.asarray(diag.get(key, np.full(n, default)), dtype=float).reshape((-1,))
    if arr.size != n:
        return np.full(n, default, dtype=float)
    return np.where(np.isfinite(arr), arr, default)


def _build_context(
    *,
    excel_path: Path,
    thermo_mode: str,
    clapeyron_model: str,
    runtime_mode: str,
    scenario: str,
    include_boundary_states: bool,
    include_vapor_states: bool,
    use_excel_vapor_holdup: bool,
    flash_feed_at_stage_conditions: bool,
):
    case = load_case_from_excel(str(excel_path))
    cfg = RunnerConfig(
        excel_path=excel_path,
        thermo_mode=thermo_mode,
        clapeyron_model=clapeyron_model,
        runtime_mode=runtime_mode,
        include_temperature=True,
        include_energy=True,
        include_boundary_states=include_boundary_states,
        include_vapor_states=include_vapor_states,
        enable_equilibrium_relaxation=False,
        vapor_holdup_relaxation_sec=0.0,
        flash_feed_at_stage_conditions=flash_feed_at_stage_conditions,
        write_logs=False,
    )
    col = build_column_spec_from_case(case)
    inputs, _provider = build_inputs_for_runner(case, col, cfg)
    if scenario == "spec_profile_with_feed_flash":
        inputs = replace(inputs, pressure_model="spec", vapor_flow_model="profile", flash_feed_at_stage_conditions=True)
    elif scenario == "spec_profile_no_feed_flash":
        inputs = replace(inputs, pressure_model="spec", vapor_flow_model="profile", flash_feed_at_stage_conditions=False)
    layout = StateVectorLayout(
        col.n_stages,
        col.n_components,
        include_top=include_boundary_states,
        include_bottom=include_boundary_states,
        include_vapor=include_vapor_states,
        include_temperature=True,
        include_energy=True,
    )
    y0 = layout.pack_y0(col)
    if not use_excel_vapor_holdup:
        y0 = _clear_initial_tray_vapor_holdup(y0, layout)
    y0 = _initialize_vapor_holdup_from_spec_pressure(
        col=col,
        layout=layout,
        y=y0,
        inputs=inputs,
        include_temperature=True,
        preserve_tray_vapor_holdup=use_excel_vapor_holdup,
    )
    dydt, diag = column_rhs(0.0, y0, col, layout, inputs)
    u = layout.unpack(y0)
    return col, layout, inputs, y0, dydt, diag, u


def _solve_energy_specific_enthalpies(diag: dict, n: int, *, h_scale: float = 1.0e4) -> tuple[np.ndarray, np.ndarray, dict]:
    if "L_out_lbmolps_tray" in diag:
        L = _as_vec(diag, "L_out_lbmolps_tray", n)
    else:
        L = _as_vec(diag, "L_out_lbmolph", n) / 3600.0
    if "V_out_lbmolps_tray" in diag:
        V = _as_vec(diag, "V_out_lbmolps_tray", n)
    else:
        V = _as_vec(diag, "V_out_lbmolph", n) / 3600.0
    q_feed = _as_vec(diag, "Q_feed_BTUps_tray", n)
    q_cond = float(np.asarray(diag.get("Q_cond_used_BTUph", [0.0]), dtype=float).reshape((-1,))[0]) / 3600.0
    q_reb = float(np.asarray(diag.get("Q_reb_used_BTUph", [0.0]), dtype=float).reshape((-1,))[0]) / 3600.0

    # Solve in scaled enthalpy variables to improve conditioning.
    # Liquid equations: L[i-1] hL[i-1] - L[i] hL[i] + Q_i = 0
    a_l = []
    b_l = []
    for i in range(n):
        row = np.zeros(n, dtype=float)
        if i > 0:
            row[i - 1] += L[i - 1] * h_scale
        row[i] -= L[i] * h_scale
        source = q_feed[i]
        if i == 0:
            source += q_cond
        if i == n - 1:
            source += q_reb
        a_l.append(row)
        b_l.append(-source)

    # Vapor equations: V[i+1] hV[i+1] - V[i] hV[i] = 0.
    # Bottom-stage vapor residual is commonly zeroed elsewhere for no-holdup
    # reboiler mappings, so anchor hV at the bottom and fit the upstream chain.
    a_v = []
    b_v = []
    for i in range(max(n - 1, 0)):
        row = np.zeros(n, dtype=float)
        row[i] -= V[i] * h_scale
        row[i + 1] += V[i + 1] * h_scale
        a_v.append(row)
        b_v.append(0.0)
    anchor = np.zeros(n, dtype=float)
    anchor[-1] = 1.0
    a_v.append(anchor)
    b_v.append(0.0)

    hL_scaled, *_ = np.linalg.lstsq(np.vstack(a_l), np.asarray(b_l, dtype=float), rcond=None)
    hV_scaled, *_ = np.linalg.lstsq(np.vstack(a_v), np.asarray(b_v, dtype=float), rcond=None)
    hL = hL_scaled * h_scale
    hV = hV_scaled * h_scale

    meta = {
        "max_abs_hL_BTU_lbmol": float(np.max(np.abs(hL))) if hL.size else math.nan,
        "max_abs_hV_BTU_lbmol": float(np.max(np.abs(hV))) if hV.size else math.nan,
        "q_cond_BTUps": q_cond,
        "q_reb_BTUps": q_reb,
        "q_feed_BTUps": q_feed.tolist(),
    }
    return hL, hV, meta


def _write_energy_sheet(template: Path, output: Path, stages: np.ndarray, EL: np.ndarray, EV: np.ndarray, hL: np.ndarray, hV: np.ndarray) -> None:
    wb = load_workbook(template)
    if "Energy State" in wb.sheetnames:
        del wb["Energy State"]
    ws = wb.create_sheet("Energy State")
    ws.append(["Stage", "Tray EL (BTU)", "Tray EV (BTU)", "hL solved (BTU/lbmol)", "hV solved (BTU/lbmol)"])
    for i in range(len(stages)):
        ws.append([int(stages[i]), float(EL[i]), float(EV[i]), float(hL[i]), float(hV[i])])
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> int:
    ap = argparse.ArgumentParser(description="Reconcile workbook energy seed inventories.")
    ap.add_argument("--excel", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--summary", default=None)
    ap.add_argument("--thermo", dest="thermo_mode", choices=["table", "live", "dwsim", "clapeyron"], default="clapeyron")
    ap.add_argument("--clapeyron-model", default="PR")
    ap.add_argument("--runtime-mode", choices=["legacy", "parity", "calibration", "hydraulic"], default="parity")
    ap.add_argument("--scenario", choices=["default", "spec_profile_no_feed_flash", "spec_profile_with_feed_flash"], default="spec_profile_with_feed_flash")
    ap.add_argument("--disable-boundary-states", dest="include_boundary_states", action="store_false")
    ap.add_argument("--disable-vapor-states", dest="include_vapor_states", action="store_false")
    ap.add_argument("--use-excel-vapor-holdup", action="store_true")
    ap.set_defaults(include_boundary_states=True, include_vapor_states=True)
    args = ap.parse_args()

    excel = _resolve(args.excel)
    output = _resolve(args.output)
    col, _layout, _inputs, _y0, _dydt, diag, u = _build_context(
        excel_path=excel,
        thermo_mode=args.thermo_mode,
        clapeyron_model=args.clapeyron_model,
        runtime_mode=args.runtime_mode,
        scenario=args.scenario,
        include_boundary_states=bool(args.include_boundary_states),
        include_vapor_states=bool(args.include_vapor_states),
        use_excel_vapor_holdup=bool(args.use_excel_vapor_holdup),
        flash_feed_at_stage_conditions=("with_feed_flash" in args.scenario),
    )
    n = int(col.n_stages)
    hL, hV, meta = _solve_energy_specific_enthalpies(diag, n)
    ML = np.asarray(diag["ML_tot_tray"], dtype=float).reshape((n,))
    MV = np.asarray(diag["MV_tot_tray"], dtype=float).reshape((n,))
    EL = ML * hL
    EV = MV * hV
    stages = np.arange(1, n + 1, dtype=int)
    _write_energy_sheet(excel, output, stages, EL, EV, hL, hV)

    summary = {
        "input": str(excel),
        "output": str(output),
        "scenario": args.scenario,
        "runtime_mode": args.runtime_mode,
        "include_boundary_states": bool(args.include_boundary_states),
        "include_vapor_states": bool(args.include_vapor_states),
        "use_excel_vapor_holdup": bool(args.use_excel_vapor_holdup),
        **meta,
    }
    summary_path = _resolve(args.summary) if args.summary else output.with_suffix(".energy_seed_summary.json")
    summary_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")
    print(output)
    print(summary_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
