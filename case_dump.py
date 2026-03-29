#!/usr/bin/env python3
"""
case_dump.py

Dynamic Distillation - Excel Case Dump Export

PURPOSE
-------
Export workbook case data into a plain key/value dump used by Scilab scripts.
Includes tray profiles, composition rows, stream totals, geometry vectors,
and molecular-weight vectors derived from component labels.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook


def _as_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _norm(s: str) -> str:
    s = _as_str(s).lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _float_or_none(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = _as_str(v)
    if not s:
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def _int_or_none(v: Any) -> Optional[int]:
    f = _float_or_none(v)
    if f is None:
        return None
    return int(round(f))


def _csv(vals: Sequence[float]) -> str:
    return ",".join(f"{float(v):.12g}" for v in vals)


def _row_values(ws, r: int) -> List[str]:
    out: List[str] = []
    for c in range(1, ws.max_column + 1):
        out.append(_norm(_as_str(ws.cell(r, c).value)))
    return out


def _header_map(ws, r: int) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = _as_str(ws.cell(r, c).value)
        if v:
            m[_norm(v)] = c
    return m


def _find_header_row_containing(ws, required: Sequence[str], max_rows: int = 400) -> Optional[int]:
    req = [_norm(x) for x in required]
    for r in range(1, min(max_rows, ws.max_row) + 1):
        row = _row_values(ws, r)
        if all(x in row for x in req):
            return r
    return None


def _find_param_value(wb, param_name: str) -> Optional[float]:
    p_norm = _norm(param_name)
    for ws in wb.worksheets:
        hdr = _find_header_row_containing(ws, ["parameter", "value"])
        if hdr is None:
            continue
        hm = _header_map(ws, hdr)
        if "parameter" not in hm or "value" not in hm:
            continue
        c_p = hm["parameter"]
        c_v = hm["value"]
        for r in range(hdr + 1, min(hdr + 250, ws.max_row) + 1):
            p = _norm(_as_str(ws.cell(r, c_p).value))
            if not p:
                continue
            if p == p_norm:
                return _float_or_none(ws.cell(r, c_v).value)
    return None


def _find_component_names(wb, nc: int) -> List[str]:
    for ws in wb.worksheets:
        for r in range(1, min(350, ws.max_row) + 1):
            for c in range(1, ws.max_column + 1):
                if _norm(_as_str(ws.cell(r, c).value)) == "component name":
                    names: List[str] = []
                    for cc in range(c + 1, c + 1 + nc):
                        names.append(_as_str(ws.cell(r, cc).value))
                    if len(names) == nc and all(names):
                        return names
    return [f"Component{i+1}" for i in range(nc)]


def _mw_lookup(name: str) -> float:
    k = _norm(name)
    lookup = {
        "n-propane": 44.09562,
        "propane": 44.09562,
        "n-butane": 58.1222,
        "butane": 58.1222,
        "n-pentane": 72.14878,
        "pentane": 72.14878,
    }
    return lookup.get(k, 60.0)


def _find_stage_geometry_table(wb) -> Tuple[Any, int, Dict[str, int]]:
    required = [
        "start stage",
        "end stage",
        "diameter (ft)",
        "tray spacing (ft)",
        "gas void fraction",
        "weir height (in)",
        "weir length (ft)",
        "active area fraction",
    ]
    for ws in wb.worksheets:
        for r in range(1, min(450, ws.max_row) + 1):
            hm = _header_map(ws, r)
            if all(_norm(h) in hm for h in required):
                return ws, r, hm
    raise RuntimeError("Stage Geometry header row not found in any sheet.")


def _read_stage_geometry(ws, hdr_row: int, hm: Dict[str, int]) -> List[Tuple[int, int, float, float, float, float, float, float]]:
    rows: List[Tuple[int, int, float, float, float, float, float, float]] = []
    r = hdr_row + 1
    while r <= ws.max_row:
        v_start = ws.cell(r, hm[_norm("start stage")]).value
        v_end = ws.cell(r, hm[_norm("end stage")]).value
        if _as_str(v_start) == "" and _as_str(v_end) == "":
            break
        start = _int_or_none(v_start)
        end = _int_or_none(v_end)
        if start is None or end is None:
            break

        def gv(h: str) -> float:
            v = _float_or_none(ws.cell(r, hm[_norm(h)]).value)
            if v is None:
                raise RuntimeError(f"Stage Geometry row {r}: missing '{h}'")
            return float(v)

        rows.append(
            (
                start,
                end,
                gv("diameter (ft)"),
                gv("tray spacing (ft)"),
                gv("gas void fraction"),
                gv("weir height (in)"),
                gv("weir length (ft)"),
                gv("active area fraction"),
            )
        )
        r += 1

    if not rows:
        raise RuntimeError("Stage Geometry table found but no data rows read.")
    return rows


def _expand_geometry(N: int, geom_rows) -> Dict[str, List[float]]:
    diam = [0.0] * N
    spacing = [0.0] * N
    gasvoid = [0.0] * N
    weir_h_in = [0.0] * N
    weir_L = [0.0] * N
    aaf = [0.0] * N

    for (start, end, d, sp, gv, wh, wl, aa) in geom_rows:
        s0 = max(1, start)
        s1 = min(N, end)
        for s in range(s0, s1 + 1):
            i = s - 1
            diam[i] = d
            spacing[i] = sp
            gasvoid[i] = gv
            weir_h_in[i] = wh
            weir_L[i] = wl
            aaf[i] = aa

    return {
        "TRAY_DIAM_FT": diam,
        "TRAY_SPACING_FT": spacing,
        "GAS_VOID_FRAC": gasvoid,
        "WEIR_HEIGHT_IN": weir_h_in,
        "WEIR_LENGTH_FT": weir_L,
        "ACTIVE_AREA_FRACTION": aaf,
    }


def _find_stage_profile_table(wb, comps: List[str]) -> Tuple[Any, int, Dict[str, int], List[int]]:
    stage_keys = ["stage"]
    p_keys = ["p (psia)", "p_psia", "pressure (psia)", "pressure", "p"]
    t_keys = ["t (f)", "t_f", "temperature (f)", "temperature", "t0_f", "t0 (f)"]
    holdup_keys = ["ml0 (lbmol)", "ml0_lbmol", "holdup (lbmol)", "liquid holdup (lbmol)", "ml0"]

    comp_norm = [_norm(c) for c in comps]

    for ws in wb.worksheets:
        for r in range(1, min(600, ws.max_row) + 1):
            hm = _header_map(ws, r)

            def has_any(keys: List[str]) -> Optional[str]:
                for k in keys:
                    if _norm(k) in hm:
                        return _norm(k)
                return None

            k_stage = has_any(stage_keys)
            k_p = has_any(p_keys)
            k_t = has_any(t_keys)
            k_m = has_any(holdup_keys)

            if not (k_stage and k_p and k_t and k_m):
                continue

            comp_cols: List[int] = []
            for cn in comp_norm:
                if cn in hm:
                    comp_cols.append(hm[cn])

            if len(comp_cols) != len(comps):
                comp_cols = []
                for j in range(1, len(comps) + 1):
                    for alt in [f"x{j}", f"x_{j}", f"liq x{j}", f"liquid x{j}"]:
                        if _norm(alt) in hm:
                            comp_cols.append(hm[_norm(alt)])
                            break

            return ws, r, hm, comp_cols

    raise RuntimeError("Could not locate a Stage Profile table (Stage, Pressure, Temperature, Holdup).")


def _read_stage_profile(
    ws,
    hdr_row: int,
    hm: Dict[str, int],
    N: int,
    NC: int,
    comp_cols: List[int],
) -> Tuple[List[float], List[float], List[float], List[List[float]]]:
    def pick(keys: List[str]) -> int:
        for k in keys:
            kk = _norm(k)
            if kk in hm:
                return hm[kk]
        raise RuntimeError("Missing expected column in Stage Profile: one of " + ", ".join(keys))

    c_stage = pick(["stage"])
    c_p = pick(["p (psia)", "p_psia", "pressure (psia)", "pressure", "p"])
    c_t = pick(["t (f)", "t_f", "temperature (f)", "temperature", "t0_f", "t0 (f)"])
    c_m = pick(["ml0 (lbmol)", "ml0_lbmol", "holdup (lbmol)", "liquid holdup (lbmol)", "ml0"])

    P = [0.0] * N
    T = [0.0] * N
    ML0 = [0.0] * N
    X0 = [[0.0] * NC for _ in range(N)]

    for r in range(hdr_row + 1, min(hdr_row + 1 + 2000, ws.max_row) + 1):
        s = _int_or_none(ws.cell(r, c_stage).value)
        if s is None or s < 1 or s > N:
            continue

        pv = _float_or_none(ws.cell(r, c_p).value)
        tv = _float_or_none(ws.cell(r, c_t).value)
        mv = _float_or_none(ws.cell(r, c_m).value)

        if pv is None or tv is None or mv is None:
            raise RuntimeError(f"Stage Profile row {r}: missing P/T/ML0 for stage {s}")

        P[s - 1] = float(pv)
        T[s - 1] = float(tv)
        ML0[s - 1] = float(mv)

        if len(comp_cols) == NC:
            xs: List[float] = []
            for c in comp_cols:
                xv = _float_or_none(ws.cell(r, c).value)
                xs.append(0.0 if xv is None else float(xv))
            tot = sum(xs)
            if tot <= 0.0:
                raise RuntimeError(f"Stage Profile row {r}: composition row sums to 0 at stage {s}")
            X0[s - 1] = [v / tot for v in xs]

    if any(p == 0.0 for p in P):
        raise RuntimeError("Stage Profile: did not populate P for all stages 1..N")
    if any(t == 0.0 for t in T):
        raise RuntimeError("Stage Profile: did not populate T for all stages 1..N")
    if len(comp_cols) != NC:
        raise RuntimeError(
            "Stage Profile: could not locate NC composition columns for X0. "
            "Expected either component-name columns or x1/x2/... columns."
        )

    return P, T, ML0, X0


def _find_streams_table(wb, comps: List[str]) -> Tuple[Any, int, Dict[str, int], List[int]]:
    stream_keys = ["stream", "stream name", "name"]
    stage_keys = ["stage", "to stage", "tray", "feed stage"]
    flow_keys = ["flow (lbmol/hr)", "flow lbmol/hr", "lbmol/hr", "flow"]

    comp_norm = [_norm(c) for c in comps]

    for ws in wb.worksheets:
        for r in range(1, min(600, ws.max_row) + 1):
            hm = _header_map(ws, r)

            def any_key(keys: List[str]) -> Optional[str]:
                for k in keys:
                    if _norm(k) in hm:
                        return _norm(k)
                return None

            k_name = any_key(stream_keys)
            k_stage = any_key(stage_keys)
            k_flow = any_key(flow_keys)
            if not (k_name and k_stage and k_flow):
                continue

            comp_cols: List[int] = []
            for cn in comp_norm:
                if cn in hm:
                    comp_cols.append(hm[cn])

            if len(comp_cols) != len(comps):
                comp_cols = []
                for j in range(1, len(comps) + 1):
                    for alt in [f"z{j}", f"z_{j}", f"x{j}", f"x_{j}"]:
                        if _norm(alt) in hm:
                            comp_cols.append(hm[_norm(alt)])
                            break

            return ws, r, hm, comp_cols

    raise RuntimeError("Could not locate a Streams table (Stream Name, Stage, Flow).")


def _read_streams(ws, hdr_row: int, hm: Dict[str, int], comps: List[str], comp_cols: List[int]) -> Dict[str, Dict[str, Any]]:
    def pick(keys: List[str]) -> int:
        for k in keys:
            kk = _norm(k)
            if kk in hm:
                return hm[kk]
        raise RuntimeError("Streams table missing expected column: " + ", ".join(keys))

    c_name = pick(["stream", "stream name", "name"])
    c_stage = pick(["stage", "to stage", "tray", "feed stage"])
    c_flow = pick(["flow (lbmol/hr)", "flow lbmol/hr", "lbmol/hr", "flow"])

    out: Dict[str, Dict[str, Any]] = {}
    NC = len(comps)

    for r in range(hdr_row + 1, min(hdr_row + 1 + 400, ws.max_row) + 1):
        nm = _as_str(ws.cell(r, c_name).value)
        if not nm:
            continue
        st = _int_or_none(ws.cell(r, c_stage).value)
        fl = _float_or_none(ws.cell(r, c_flow).value)
        if st is None or fl is None:
            continue

        z = None
        if len(comp_cols) == NC:
            zs: List[float] = []
            for c in comp_cols:
                v = _float_or_none(ws.cell(r, c).value)
                zs.append(0.0 if v is None else float(v))
            tot = sum(zs)
            if tot > 0.0:
                z = [v / tot for v in zs]

        out[_norm(nm)] = {"name": nm, "stage": int(st), "flow": float(fl), "z": z}

    return out


def _pick_stream(streams: Dict[str, Dict[str, Any]], keywords: List[str]) -> Optional[Dict[str, Any]]:
    for k, rec in streams.items():
        for kw in keywords:
            if kw in k:
                return rec
    return None


def build_case_dump(excel_path: Path) -> List[str]:
    wb = load_workbook(excel_path, data_only=True)

    n = _find_param_value(wb, "Number of Stages")
    nc = _find_param_value(wb, "Number of Components")
    if n is None or nc is None:
        raise RuntimeError("Could not find 'Number of Stages' and/or 'Number of Components' in Excel template.")

    N = int(round(n))
    NC = int(round(nc))
    comps = _find_component_names(wb, NC)

    ws_geom, hdr_geom, hm_geom = _find_stage_geometry_table(wb)
    geom_rows = _read_stage_geometry(ws_geom, hdr_geom, hm_geom)
    geom_vecs = _expand_geometry(N, geom_rows)

    ws_prof, hdr_prof, hm_prof, comp_cols = _find_stage_profile_table(wb, comps)
    P_psia, T0_F, ML0_lbmol, X0 = _read_stage_profile(ws_prof, hdr_prof, hm_prof, N, NC, comp_cols)

    ws_str, hdr_str, hm_str, comp_cols_s = _find_streams_table(wb, comps)
    streams = _read_streams(ws_str, hdr_str, hm_str, comps, comp_cols_s)

    rec_D = _pick_stream(streams, ["distillate", "dist", "d "])
    rec_B = _pick_stream(streams, ["bottoms", "bottom", "btm", "b "])
    rec_F = _pick_stream(streams, ["feed", "f "])
    if rec_F is None:
        raise RuntimeError("Could not find FEED stream in Streams table (name containing 'feed').")
    if rec_D is None:
        raise RuntimeError("Could not find DISTILLATE stream in Streams table (name containing 'distillate').")
    if rec_B is None:
        raise RuntimeError("Could not find BOTTOMS stream in Streams table (name containing 'bottoms').")

    FEED_STAGE = int(rec_F["stage"])
    F_LBMOLPH = float(rec_F["flow"])
    D_LBMOLPH = float(rec_D["flow"])
    B_LBMOLPH = float(rec_B["flow"])

    zF = rec_F.get("z", None)
    if zF is None:
        zF = X0[FEED_STAGE - 1][:]

    MW_COMP = [_mw_lookup(nm) for nm in comps]
    MW_LIQ: List[float] = []
    for i in range(N):
        mw = 0.0
        for j in range(NC):
            mw += X0[i][j] * MW_COMP[j]
        MW_LIQ.append(mw)

    lines: List[str] = []
    lines.append(f"N={N}")
    lines.append(f"NC={NC}")
    lines.append(f"COMPONENTS_EXCEL={','.join(comps)}")
    lines.append(f"P_PSIA={_csv(P_psia)}")
    lines.append(f"T0_F={_csv(T0_F)}")
    lines.append(f"ML0_LBMOL={_csv(ML0_lbmol)}")

    for i in range(N):
        lines.append(f"X0ROW={_csv(X0[i])}")
    for i in range(N):
        lines.append(f"ZROW={_csv(X0[i])}")

    lines.append(f"FEED_STAGE={FEED_STAGE}")
    lines.append(f"F_LBMOLPH={F_LBMOLPH:.12g}")
    lines.append(f"D_LBMOLPH={D_LBMOLPH:.12g}")
    lines.append(f"B_LBMOLPH={B_LBMOLPH:.12g}")
    lines.append(f"ZF={_csv(zF)}")

    for k in [
        "TRAY_DIAM_FT",
        "TRAY_SPACING_FT",
        "GAS_VOID_FRAC",
        "WEIR_HEIGHT_IN",
        "WEIR_LENGTH_FT",
        "ACTIVE_AREA_FRACTION",
    ]:
        lines.append(f"{k}={_csv(geom_vecs[k])}")

    lines.append(f"MW_COMP={_csv(MW_COMP)}")
    lines.append(f"MW_LIQ={_csv(MW_LIQ)}")
    return lines


def main(argv: List[str]) -> int:
    if len(argv) != 3:
        print("Usage: python case_dump.py <excel_path> <out_path>", file=sys.stderr)
        return 2

    excel_path = Path(argv[1]).expanduser().resolve()
    out_path = Path(argv[2]).expanduser().resolve()
    if not excel_path.exists():
        print(f"ERROR: Excel file not found: {excel_path}", file=sys.stderr)
        return 2

    try:
        lines = build_case_dump(excel_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
        return 0
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
