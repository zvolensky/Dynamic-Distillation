from __future__ import annotations

from pathlib import Path
import ssl
from urllib.error import URLError
from urllib.request import urlopen

from openpyxl import Workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUT_PATH = PROJECT_ROOT / "validation_skogestad_column_a_relative_volatility.xlsx"
SOURCE_URL = "https://skoge.folk.ntnu.no/book/matlab_m/cola/cola.dat"


def _fetch_source_text() -> str:
    try:
        with urlopen(SOURCE_URL, timeout=20) as response:
            return response.read().decode("utf-8", errors="replace")
    except (ssl.SSLError, URLError):
        context = ssl._create_unverified_context()
        with urlopen(SOURCE_URL, timeout=20, context=context) as response:
            return response.read().decode("utf-8", errors="replace")


def _parse_cola_dat(text: str) -> list[dict[str, float]]:
    marker = "STAGE"
    if marker not in text:
        raise ValueError("Could not find Skogestad cola.dat stage profile marker.")
    tail = text.split(marker, 1)[1]
    tokens = tail.replace("d", "e").replace("D", "e").split()
    while tokens:
        try:
            float(tokens[0])
            break
        except ValueError:
            tokens.pop(0)
    if len(tokens) < 5 * 41:
        raise ValueError(f"Expected at least 41 stage rows in cola.dat, found {len(tokens) // 5}.")
    rows: list[dict[str, float]] = []
    for i in range(0, 5 * 41, 5):
        rows.append(
            {
                "source_stage": int(float(tokens[i])),
                "L_kmol_min": float(tokens[i + 1]),
                "V_kmol_min": float(tokens[i + 2]),
                "x_light": float(tokens[i + 3]),
                "y_light": float(tokens[i + 4]),
            }
        )
    return rows


def _y_from_x(x_light: float, alpha: float) -> float:
    return alpha * x_light / (1.0 + (alpha - 1.0) * x_light)


def build_workbook(path: Path = OUT_PATH) -> Path:
    include_boundary_state = False
    source_rows_bottom_to_top = _parse_cola_dat(_fetch_source_text())
    source_rows_top_to_bottom = sorted(
        source_rows_bottom_to_top,
        key=lambda row: row["source_stage"],
        reverse=True,
    )

    n_stages = len(source_rows_top_to_bottom)
    alpha = 1.5
    feed_source_stage = 21
    feed_stage_top_based = n_stages + 1 - feed_source_stage
    flow_scale = 60.0
    feed_flow = 1.0 * flow_scale
    dist_flow = 0.5 * flow_scale
    bottom_flow = 0.5 * flow_scale
    tray_holdup = 0.5
    # Skogestad Column A neglects vapor holdup. This implementation still
    # carries vapor composition states, so seed a small fixed internal vapor
    # inventory to preserve y_i without adding pressure-based MV dynamics.
    tray_vapor_holdup = 0.5

    wb = Workbook()
    ws = wb.active
    ws.title = "Specifications"
    specs = [
        ("Parameter", "Value"),
        ("Number of Stages", n_stages),
        ("Number of Components", 2),
        ("Condenser Type", "Total"),
        ("Component Name", "N-butane", "n-Pentane"),
        ("Thermo Mode", "relative-volatility"),
        ("Relative Volatility", alpha),
        ("Runtime Mode", "parity"),
        ("Include Energy", False),
        ("Simulation Length (min)", 5.0),
        ("Timestep (sec)", 0.2),
        ("Log Frequency (timesteps)", 300),
        ("Top Accumulator Holdup (lbmol)", tray_holdup),
        ("Bottom Holdup (lbmol)", tray_holdup),
        ("Pressure Model", "static"),
        ("Vapor Flow Model", "profile"),
        ("Stage time constant [tau] (sec)", 10.0),
        ("Vapor Holdup Relaxation (sec)", 0.0),
        ("Equilibrium Relaxation Mode", "composition-only"),
        ("Equilibrium Tau (sec)", 0.5),
        ("Source", SOURCE_URL),
        ("Source Stage Count Includes Reboiler And Total Condenser", True),
        ("Source Feed Stage Counted From Bottom", feed_source_stage),
        ("Source Flow Unit", "kmol/min; workbook flow values are scaled by 60"),
    ]
    for r, row in enumerate(specs, start=1):
        for c, value in enumerate(row, start=1):
            ws.cell(r, c).value = value

    ws = wb.create_sheet("Initial Conditions")
    headers = [
        "Stage",
        "Source Stage",
        "Temperature (F)",
        "Pressure (psia)",
        "Vapor Flow (lbmol/h)",
        "Liquid Flow (lbmol/h)",
        "Liquid Holdup (lbmol)",
        "Vapor Holdup (lbmol)",
        "Vapor Composition Component 1",
        "Vapor Composition Component 2",
        "Liquid Composition Component 1",
        "Liquid Composition Component 2",
    ]
    for c, header in enumerate(headers, start=1):
        ws.cell(1, c).value = header

    for stage_top, row in enumerate(source_rows_top_to_bottom, start=1):
        frac = (stage_top - 1) / max(n_stages - 1, 1)
        x_light = row["x_light"]
        y_light = row["y_light"]
        # The source table uses zero vapor composition on the total condenser row.
        if stage_top != 1:
            y_light = _y_from_x(x_light, alpha)
        temp_f = 180.0 + (230.0 - 180.0) * frac
        out = [
            stage_top,
            row["source_stage"],
            temp_f,
            14.7,
            0.0 if stage_top == 1 else row["V_kmol_min"] * flow_scale,
            row["L_kmol_min"] * flow_scale,
            tray_holdup,
            0.0 if stage_top == 1 else tray_vapor_holdup,
            y_light,
            1.0 - y_light,
            x_light,
            1.0 - x_light,
        ]
        for c, value in enumerate(out, start=1):
            ws.cell(stage_top + 1, c).value = value

    ws = wb.create_sheet("Streams")
    x_top = source_rows_top_to_bottom[0]["x_light"]
    x_bottom = source_rows_top_to_bottom[-1]["x_light"]
    rows = [
        ("Stream", "Feed", "Distillate", "Bottom"),
        ("Stage", feed_stage_top_based, 1, n_stages),
        ("Pressure (psia)", 14.7, 14.7, 14.7),
        ("Vapour fraction", 0.0, 0.0, 0.0),
        ("Temperature (F)", 205.0, 180.0, 230.0),
        ("Total molar flow (lbmol/h)", feed_flow, dist_flow, bottom_flow),
        ("Mole flows (lbmol/h)", None, None, None),
        # Skogestad colamod.m specifies product draws as D*xD and B*xB,
        # so only the feed uses fixed component rates. Leaving product
        # component rates blank makes the runner draw product composition from
        # the current condenser/reboiler states instead of pinning it to the
        # initial product split.
        ("N-butane", 0.5 * feed_flow, None, None),
        ("n-Pentane", 0.5 * feed_flow, None, None),
    ]
    for r, row in enumerate(rows, start=1):
        for c, value in enumerate(row, start=1):
            ws.cell(r, c).value = value

    ws = wb.create_sheet("Components")
    ws.cell(1, 1).value = "Component Name"
    ws.cell(2, 1).value = "N-butane"
    ws.cell(3, 1).value = "n-Pentane"

    if include_boundary_state:
        ws = wb.create_sheet("Boundary State")
        boundary_rows = [
            ("State", "N-butane", "n-Pentane"),
            ("top_L", tray_holdup * x_top, tray_holdup * (1.0 - x_top)),
            ("top_V", 0.0, 0.0),
            ("bottom_L", tray_holdup * x_bottom, tray_holdup * (1.0 - x_bottom)),
            ("bottom_V", 0.0, 0.0),
        ]
        for r, row in enumerate(boundary_rows, start=1):
            for c, value in enumerate(row, start=1):
                ws.cell(r, c).value = value

    path = Path(path)
    wb.save(path)
    return path


if __name__ == "__main__":
    print(build_workbook())
