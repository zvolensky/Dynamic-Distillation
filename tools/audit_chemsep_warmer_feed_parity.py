#!/usr/bin/env python
"""
Audit the C3/C4/C5 warmer-feed ChemSep .sep file against an Excel seed.

The goal is not to force an exact one-to-one topology match. ChemSep's total
condenser report has source-specific terminal-stage conventions, while the
dynamic model may intentionally map a total condenser to zero stage-1 vapor
traffic. The report therefore highlights exact numeric differences and calls
out likely topology mappings separately.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]

PA_TO_PSIA = 0.00014503773773020923
K_TO_F_OFFSET = -459.67
KMOL_S_TO_LBMOL_H = 2.2046226218487757 * 3600.0
W_TO_BTU_H = 3.41214163312794


def k_to_f(value: float) -> float:
    return value * 9.0 / 5.0 + K_TO_F_OFFSET


def pa_to_psia(value: float) -> float:
    return value * PA_TO_PSIA


def flow_to_lbmol_h(value: float) -> float:
    return value * KMOL_S_TO_LBMOL_H


def duty_to_btu_h(value: float) -> float:
    return value * W_TO_BTU_H


def _float(text: str) -> float:
    return float(text.replace("D", "E"))


@dataclass
class SepData:
    path: Path
    condenser_type: str
    reboiler_type: str
    n_stages: int
    feed_stage: int
    thermo: str
    reflux_ratio: float
    bottom_product_flow_basis: float
    feed_temperature_k: float
    feed_pressure_pa: float
    feed_component_flows_basis: list[float]
    condenser_pressure_pa: float
    top_pressure_pa: float
    bottom_pressure_pa: float
    profiles: list[dict[str, float]]
    vapor_x: list[list[float]]
    liquid_x: list[list[float]]
    condenser_duty_w: float
    reboiler_duty_w: float
    top_product: dict[str, float | list[float]]
    bottom_product: dict[str, float | list[float]]


def _section(lines: list[str], name: str) -> list[str]:
    marker = f"[{name}]"
    try:
        start = next(i for i, line in enumerate(lines) if line.strip() == marker) + 1
    except StopIteration:
        return []
    end = len(lines)
    for i in range(start, len(lines)):
        if lines[i].startswith("[") and lines[i].rstrip().endswith("]"):
            end = i
            break
    return lines[start:end]


def _span(lines: list[str], start_name: str, end_name: str) -> list[str]:
    start_marker = f"[{start_name}]"
    end_marker = f"[{end_name}]"
    try:
        start = next(i for i, line in enumerate(lines) if line.strip() == start_marker) + 1
    except StopIteration:
        return []
    try:
        end = next(i for i in range(start, len(lines)) if lines[i].strip() == end_marker)
    except StopIteration:
        end = len(lines)
    return lines[start:end]


def _find_first_number(section: Iterable[str], suffix: str) -> float:
    for line in section:
        stripped = line.strip()
        if stripped.endswith(suffix):
            return _float(stripped.split()[0])
    raise ValueError(f"could not find line ending with {suffix!r}")


def _parse_profiles(lines: list[str]) -> list[dict[str, float]]:
    rows: list[dict[str, float]] = []
    for line in lines:
        parts = line.split()
        if len(parts) < 6:
            continue
        if not parts[0].isdigit():
            continue
        rows.append(
            {
                "stage": int(parts[0]),
                "temperature_k": _float(parts[1]),
                "pressure_pa": _float(parts[2]),
                "vapor_flow_basis": _float(parts[3]),
                "liquid_flow_basis": _float(parts[4]),
                "duty_w": _float(parts[5]),
            }
        )
    return rows


def _parse_component_blocks(lines: list[str], marker: str, n_stages: int, n_components: int) -> list[list[float]]:
    try:
        start = next(i for i, line in enumerate(lines) if line.strip() == marker)
    except StopIteration as exc:
        raise ValueError(f"could not find composition marker {marker!r}") from exc

    values = [[math.nan for _ in range(n_components)] for _ in range(n_stages)]
    current_stages: list[int] = []
    i = start + 1
    while i < len(lines):
        line = lines[i]
        if line.startswith("[") and i > start + 1:
            break
        header_match = re.search(r"stages:\s*(\d+)\s+to\s+(\d+)", line)
        if header_match:
            first = int(header_match.group(1))
            last = int(header_match.group(2))
            current_stages = list(range(first, last + 1))
            i += 1
            continue
        parts = line.split()
        if current_stages and len(parts) == len(current_stages) + 1 and parts[0].isdigit():
            comp = int(parts[0]) - 1
            if 0 <= comp < n_components:
                for stage, text in zip(current_stages, parts[1:]):
                    values[stage - 1][comp] = _float(text)
        if line.strip().startswith("[Reaction rates]") or line.strip().startswith("[Slopes]"):
            break
        i += 1
    return values


def _parse_product(lines: list[str], name: str, n_components: int) -> dict[str, float | list[float]]:
    sec = _section(lines, name)
    out: dict[str, float | list[float]] = {}
    comp: list[float] = []
    for line in sec:
        stripped = line.strip()
        if stripped.endswith("Stage Number"):
            out["stage"] = _float(stripped.split()[0])
        elif stripped.endswith("Flow rate") and "flow_rate" not in out:
            out["flow_rate_basis"] = _float(stripped.split()[0])
        elif stripped.endswith("Temperature [K]"):
            out["temperature_k"] = _float(stripped.split()[0])
        elif stripped.endswith("Pressure [Pa]"):
            out["pressure_pa"] = _float(stripped.split()[0])
        elif "Mole fraction of component" in stripped and len(comp) < n_components:
            comp.append(_float(stripped.split()[0]))
    out["x"] = comp
    return out


def parse_sep(path: Path) -> SepData:
    lines = path.read_text(encoding="latin-1").splitlines()
    operation = _section(lines, "Operation")
    thermo = _section(lines, "Thermodynamics")
    pressures = _section(lines, "Pressures")
    feeds = _section(lines, "Feeds")
    condenser = _section(lines, "Condenser")
    reboiler = _section(lines, "Reboiler")
    results_block = _span(lines, "Profiles", "Condenser Heat Duty")
    profiles_section = _span(lines, "Profiles", "Enthalpies/Entropies")

    n_stages = int(_find_first_number(operation, "Stages"))
    feed_stage_line = next(line.strip() for line in operation if line.strip().startswith("F="))
    feed_stage = int(feed_stage_line.split("=", 1)[1])
    condenser_type = next(line.strip() for line in operation if "Condenser" in line)
    reboiler_type = next(line.strip() for line in operation if "Reboiler" in line)
    thermo_desc = "; ".join(line.strip() for line in thermo if "K model" in line or "Cubic EOS" in line)

    feed_component_flows: list[float] = []
    for line in feeds:
        stripped = line.strip()
        if "Component" in stripped and stripped.endswith("flow"):
            feed_component_flows.append(_float(stripped.split()[0]))

    profiles = _parse_profiles(profiles_section)
    vapor_x = _parse_component_blocks(results_block, "[Vapour phase compositions]", n_stages, len(feed_component_flows))
    liquid_x = _parse_component_blocks(results_block, "[Liquid phase compositions]", n_stages, len(feed_component_flows))

    return SepData(
        path=path,
        condenser_type=condenser_type,
        reboiler_type=reboiler_type,
        n_stages=n_stages,
        feed_stage=feed_stage,
        thermo=thermo_desc,
        reflux_ratio=_find_first_number(condenser, "Value Qcondenser"),
        bottom_product_flow_basis=_find_first_number(reboiler, "Value Qreboiler"),
        feed_temperature_k=_find_first_number(feeds, "Temperature"),
        feed_pressure_pa=_find_first_number(feeds, "Pressure"),
        feed_component_flows_basis=feed_component_flows,
        condenser_pressure_pa=_find_first_number(pressures, "Condenser pressure"),
        top_pressure_pa=_find_first_number(pressures, "Top pressure"),
        bottom_pressure_pa=_find_first_number(pressures, "Bottom pressure"),
        profiles=profiles,
        vapor_x=vapor_x,
        liquid_x=liquid_x,
        condenser_duty_w=_find_first_number(_section(lines, "Condenser Heat Duty"), "Duty"),
        reboiler_duty_w=_find_first_number(_section(lines, "Reboiler Heat Duty"), "Duty"),
        top_product=_parse_product(lines, "Top product", len(feed_component_flows)),
        bottom_product=_parse_product(lines, "Bottom product", len(feed_component_flows)),
    )


def _sheet_rows(path: Path, sheet: str) -> list[list[object]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _kv_rows(rows: list[list[object]]) -> dict[str, object]:
    out: dict[str, object] = {}
    for row in rows:
        if row and isinstance(row[0], str) and row[0].strip():
            out[row[0].strip()] = row[1] if len(row) > 1 else None
    return out


def _initial_conditions(path: Path) -> tuple[list[str], list[dict[str, float]]]:
    rows = _sheet_rows(path, "Initial Conditions")
    headers = [str(h).strip() for h in rows[0]]
    out: list[dict[str, float]] = []
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        item: dict[str, float] = {}
        for key, value in zip(headers, row):
            if value is None:
                continue
            try:
                item[key] = float(value)
            except Exception:
                pass
        out.append(item)
    return headers, out


def _stream_table(path: Path) -> dict[str, dict[str, float]]:
    rows = _sheet_rows(path, "Streams")
    header = rows[0]
    names = [str(x).strip() if x is not None else "" for x in header[1:]]
    result: dict[str, dict[str, float]] = {name: {} for name in names if name}
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        label = str(row[0]).strip()
        for name, value in zip(names, row[1:]):
            if not name or value is None:
                continue
            try:
                result[name][label] = float(value)
            except Exception:
                pass
    return result


def _add_delta(
    rows: list[dict[str, object]],
    category: str,
    item: str,
    sep_value: float,
    workbook_value: float,
    units: str = "",
    note: str = "",
) -> None:
    rows.append(
        {
            "category": category,
            "item": item,
            "sep_value": sep_value,
            "workbook_value": workbook_value,
            "delta": workbook_value - sep_value,
            "abs_delta": abs(workbook_value - sep_value),
            "units": units,
            "note": note,
        }
    )


def build_audit(sep: SepData, workbook: Path) -> tuple[list[dict[str, object]], dict[str, object]]:
    rows: list[dict[str, object]] = []
    specs = _kv_rows(_sheet_rows(workbook, "Specifications"))
    streams = _stream_table(workbook)
    _, initial = _initial_conditions(workbook)
    stage_by_number = {int(row["Stage"]): row for row in initial if "Stage" in row}

    _add_delta(rows, "spec", "number_of_stages", sep.n_stages, float(specs.get("Number of Stages", math.nan)))
    _add_delta(rows, "spec", "number_of_components", len(sep.feed_component_flows_basis), float(specs.get("Number of Components", math.nan)))
    _add_delta(rows, "duty", "condenser_duty", duty_to_btu_h(sep.condenser_duty_w), float(specs.get("Condenser Duty (Btu/h)", math.nan)), "Btu/h")
    _add_delta(rows, "duty", "reboiler_duty", duty_to_btu_h(sep.reboiler_duty_w), float(specs.get("Reboiler Duty (Btu/h)", math.nan)), "Btu/h")

    feed = streams.get("Feed", {})
    _add_delta(rows, "stream", "feed_temperature", k_to_f(sep.feed_temperature_k), feed.get("Temperature (F)", math.nan), "F")
    _add_delta(rows, "stream", "feed_pressure", pa_to_psia(sep.feed_pressure_pa), feed.get("Pressure (psia)", math.nan), "psia")
    _add_delta(rows, "stream", "feed_total_flow", flow_to_lbmol_h(sum(sep.feed_component_flows_basis)), feed.get("Total molar flow (lbmol/h)", math.nan), "lbmol/h")

    dist = streams.get("Distillate", {})
    bot = streams.get("Bottom", {})
    _add_delta(
        rows,
        "stream",
        "distillate_pressure_vs_chemsep_condenser",
        pa_to_psia(float(sep.top_product.get("pressure_pa", sep.condenser_pressure_pa))),
        dist.get("Pressure (psia)", math.nan),
        "psia",
        "ChemSep top product uses condenser pressure; workbook appears to use top-stage pressure.",
    )
    _add_delta(rows, "stream", "distillate_flow", flow_to_lbmol_h(float(sep.top_product.get("flow_rate_basis", math.nan))), dist.get("Total molar flow (lbmol/h)", math.nan), "lbmol/h")
    _add_delta(rows, "stream", "bottom_pressure", pa_to_psia(float(sep.bottom_product.get("pressure_pa", sep.bottom_pressure_pa))), bot.get("Pressure (psia)", math.nan), "psia")
    _add_delta(rows, "stream", "bottom_flow", flow_to_lbmol_h(float(sep.bottom_product.get("flow_rate_basis", math.nan))), bot.get("Total molar flow (lbmol/h)", math.nan), "lbmol/h")

    for profile in sep.profiles:
        stage = int(profile["stage"])
        wb_row = stage_by_number.get(stage)
        if not wb_row:
            continue
        _add_delta(rows, "profile", f"stage_{stage:02d}_temperature", k_to_f(profile["temperature_k"]), wb_row.get("Temperature (F)", math.nan), "F")
        _add_delta(rows, "profile", f"stage_{stage:02d}_pressure", pa_to_psia(profile["pressure_pa"]), wb_row.get("Pressure (psia)", math.nan), "psia")
        note = ""
        if stage == 1:
            note = "Likely intentional total-condenser topology mapping if workbook stage-1 vapor flow is zero."
        _add_delta(rows, "profile", f"stage_{stage:02d}_vapor_flow", flow_to_lbmol_h(profile["vapor_flow_basis"]), wb_row.get("Vapor Flow (lbmol/h)", math.nan), "lbmol/h", note)
        liquid_note = ""
        if stage == sep.n_stages:
            liquid_note = "Likely intentional reboiler/bottoms topology mapping if workbook bottom-stage liquid flow is zero."
        _add_delta(rows, "profile", f"stage_{stage:02d}_liquid_flow", flow_to_lbmol_h(profile["liquid_flow_basis"]), wb_row.get("Liquid Flow (lbmol/h)", math.nan), "lbmol/h", liquid_note)
        for comp_idx in range(len(sep.feed_component_flows_basis)):
            _add_delta(
                rows,
                "vapor_composition",
                f"stage_{stage:02d}_component_{comp_idx + 1}",
                sep.vapor_x[stage - 1][comp_idx],
                wb_row.get(f"Vapor Composition Component {comp_idx + 1}", math.nan),
                "mole fraction",
            )
            _add_delta(
                rows,
                "liquid_composition",
                f"stage_{stage:02d}_component_{comp_idx + 1}",
                sep.liquid_x[stage - 1][comp_idx],
                wb_row.get(f"Liquid Composition Component {comp_idx + 1}", math.nan),
                "mole fraction",
            )

    summary: dict[str, object] = {
        "sep_path": str(sep.path),
        "workbook": str(workbook),
        "condenser_type": sep.condenser_type,
        "reboiler_type": sep.reboiler_type,
        "thermo": sep.thermo,
        "feed_stage": sep.feed_stage,
        "flow_scale_assumption": "ChemSep basis values are treated as kmol/s and converted to lbmol/h.",
        "duty_scale_assumption": "ChemSep duties are treated as W and converted to Btu/h.",
        "max_abs_delta_by_category": {},
    }
    for category in sorted({str(row["category"]) for row in rows}):
        finite = [float(row["abs_delta"]) for row in rows if row["category"] == category and math.isfinite(float(row["abs_delta"]))]
        summary["max_abs_delta_by_category"][category] = max(finite) if finite else None
    return rows, summary


def write_report(rows: list[dict[str, object]], summary: dict[str, object], out_dir: Path) -> tuple[Path, Path, Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    csv_path = out_dir / "chemsep_warmer_feed_parity_audit.csv"
    json_path = out_dir / "chemsep_warmer_feed_parity_audit_summary.json"
    md_path = out_dir / "chemsep_warmer_feed_parity_audit.md"

    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["category", "item", "sep_value", "workbook_value", "delta", "abs_delta", "units", "note"])
        writer.writeheader()
        writer.writerows(rows)

    json_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    sorted_rows = sorted(rows, key=lambda row: float(row["abs_delta"]) if math.isfinite(float(row["abs_delta"])) else -1.0, reverse=True)
    top_rows = sorted_rows[:20]
    lines = [
        "# ChemSep Warmer-Feed Parity Audit",
        "",
        f"- ChemSep file: `{summary['sep_path']}`",
        f"- Workbook: `{summary['workbook']}`",
        f"- Condenser: {summary['condenser_type']}",
        f"- Reboiler: {summary['reboiler_type']}",
        f"- Thermo: {summary['thermo']}",
        f"- Feed stage: {summary['feed_stage']}",
        f"- Flow scaling: {summary['flow_scale_assumption']}",
        f"- Duty scaling: {summary['duty_scale_assumption']}",
        "",
        "## Max Absolute Delta By Category",
        "",
    ]
    for category, value in dict(summary["max_abs_delta_by_category"]).items():
        lines.append(f"- {category}: {value}")
    lines.extend(
        [
            "",
            "## Largest Differences",
            "",
            "| Category | Item | ChemSep | Workbook | Delta | Units | Note |",
            "|---|---:|---:|---:|---:|---|---|",
        ]
    )
    for row in top_rows:
        lines.append(
            "| {category} | {item} | {sep_value:.9g} | {workbook_value:.9g} | {delta:.9g} | {units} | {note} |".format(
                **row
            )
        )
    lines.extend(
        [
            "",
            "## Interpretation Notes",
            "",
            "- A large stage-1 vapor-flow difference is expected if the workbook intentionally maps the total condenser to zero vapor traffic in the dynamic model.",
            "- A large bottom-stage liquid-flow difference is expected if the workbook intentionally maps the reboiler/bottoms outlet outside the tray liquid-flow profile.",
            "- The distillate pressure difference is not just rounding: ChemSep reports a condenser/top-product pressure below the top tray pressure.",
            "- Composition differences are small but real; they likely reflect the workbook being generated from a rounded/exported ChemSep table rather than this exact `.sep` result block.",
        ]
    )
    md_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return csv_path, json_path, md_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Audit ChemSep warmer-feed .sep parity against an Excel seed.")
    parser.add_argument("--sep", default=r"d:\Users\Thoma\Documents\Depropanizer_warmer_feed.sep")
    parser.add_argument("--excel", default="logs/c3c4_splitter_openloop_seed_20260526.xlsx")
    parser.add_argument("--out-dir", default=None)
    args = parser.parse_args()

    sep_path = Path(args.sep)
    excel_path = Path(args.excel)
    if not excel_path.is_absolute():
        excel_path = PROJECT_ROOT / excel_path
    if args.out_dir:
        out_dir = Path(args.out_dir)
        if not out_dir.is_absolute():
            out_dir = PROJECT_ROOT / out_dir
    else:
        stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        out_dir = PROJECT_ROOT / "logs" / f"chemsep_warmer_feed_parity_audit_{stamp}"

    sep = parse_sep(sep_path)
    rows, summary = build_audit(sep, excel_path)
    csv_path, json_path, md_path = write_report(rows, summary, out_dir)

    print(f"report: {md_path}")
    print(f"csv: {csv_path}")
    print(f"summary: {json_path}")
    print(json.dumps(summary["max_abs_delta_by_category"], indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
