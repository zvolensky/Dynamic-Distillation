import importlib.util
from pathlib import Path
import sys

import numpy as np
from openpyxl import load_workbook
import pytest

from dynamic_distillation.core_v2.source_equation_dynamics_v1 import (
    SourceFeedSchedule,
    SourceIntegrationOptions,
    external_material_rates,
    integrate_source_trajectory,
    pack_source_state,
    unpack_source_state,
)
from dynamic_distillation.core_v2.source_equation_gate_v1 import (
    BinarySourceColumnSpec,
)


ROOT = Path(__file__).resolve().parents[1]


def _source_profile() -> np.ndarray:
    workbook = load_workbook(
        ROOT / "validation_skogestad_column_a_relative_volatility.xlsx",
        read_only=True,
        data_only=True,
    )
    try:
        sheet = workbook["Initial Conditions"]
        headers = {
            str(cell.value): index
            for index, cell in enumerate(next(sheet.iter_rows()), start=1)
        }
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            source_stage = row[headers["Source Stage"] - 1]
            if source_stage is None:
                continue
            rows.append(
                (
                    int(source_stage),
                    float(
                        row[
                            headers[
                                "Liquid Composition Component 1"
                            ]
                            - 1
                        ]
                    ),
                )
            )
    finally:
        workbook.close()
    return np.asarray([x for _, x in sorted(rows)], dtype=float)


def _reference_module():
    path = ROOT / "tools" / "compare_skogestad_dynamic_response.py"
    module_spec = importlib.util.spec_from_file_location(
        "_dd079_test_reference",
        path,
    )
    assert module_spec is not None
    assert module_spec.loader is not None
    module = importlib.util.module_from_spec(module_spec)
    sys.modules[module_spec.name] = module
    module_spec.loader.exec_module(module)
    return module


def _reference_rhs(reference):
    def rhs(time_min, augmented_state, spec):
        n_stages = int(spec.n_stages)
        packed = np.asarray(
            augmented_state[: 2 * n_stages],
            dtype=float,
        )
        x, _ = unpack_source_state(packed, n_stages=n_stages)
        case = reference.SourceCase(
            n_stages=n_stages,
            feed_stage_bottom_based=spec.feed_stage_from_bottom,
            alpha=spec.relative_volatility,
            taul_min=spec.liquid_hydraulic_tau_min,
            f0_kmol_min=spec.nominal_feed_kmol_min,
            qf0=spec.nominal_feed_liquid_fraction,
            l0_kmol_min=spec.nominal_rectifying_liquid_kmol_min,
            v0_kmol_min=spec.nominal_boilup_kmol_min,
            lambda_k2=spec.liquid_vapor_coupling,
            reflux_kmol_min=spec.reflux_kmol_min,
            boilup_kmol_min=spec.boilup_kmol_min,
            distillate_kmol_min=spec.distillate_kmol_min,
            bottoms_kmol_min=spec.bottoms_kmol_min,
            feed_kmol_min=spec.feed_kmol_min,
            zf=spec.feed_light_mole_fraction,
            qf=spec.feed_liquid_fraction,
        )
        physical = reference.colamod_rhs_min(
            time_min,
            packed,
            case,
        )
        external = external_material_rates(spec, x)
        return np.concatenate((physical, np.asarray(external)))

    return rhs


def _initial_state() -> tuple[BinarySourceColumnSpec, np.ndarray]:
    spec = BinarySourceColumnSpec()
    state = pack_source_state(
        _source_profile(),
        np.full(spec.n_stages, 0.5, dtype=float),
    )
    return spec, state


def test_dd079_source_state_pack_unpack_round_trip():
    spec, state = _initial_state()

    x, holdup = unpack_source_state(state, n_stages=spec.n_stages)

    np.testing.assert_array_equal(x, state[: spec.n_stages])
    np.testing.assert_array_equal(holdup, state[spec.n_stages :])
    np.testing.assert_array_equal(pack_source_state(x, holdup), state)


def test_dd079_nonzero_feed_event_is_segmented_at_exact_time():
    spec, state = _initial_state()
    times = np.arange(0.0, 11.0, 1.0)
    schedule = SourceFeedSchedule(
        step_time_min=5.0,
        feed_before_kmol_min=1.0,
        feed_after_kmol_min=1.01,
    )

    trajectory = integrate_source_trajectory(
        base_spec=spec,
        initial_packed_state=state,
        time_min=times,
        feed_schedule=schedule,
    )

    step_index = int(np.flatnonzero(times == 5.0)[0])
    assert (
        abs(trajectory.cumulative_external_total_kmol[step_index])
        < 1.0e-12
    )
    assert trajectory.cumulative_external_total_kmol[-1] == pytest.approx(
        0.05,
        rel=0.0,
        abs=2.0e-10,
    )
    assert trajectory.feed_step_time_min == 5.0


@pytest.mark.parametrize(
    ("case_name", "feed_schedule", "perturb"),
    (
        ("nominal", None, False),
        (
            "feed",
            SourceFeedSchedule(
                step_time_min=0.0,
                feed_before_kmol_min=1.0,
                feed_after_kmol_min=1.01,
            ),
            False,
        ),
        ("perturbed", None, True),
    ),
)
def test_dd079_v2_trajectory_matches_independent_reference(
    case_name,
    feed_schedule,
    perturb,
):
    del case_name
    spec, state = _initial_state()
    if perturb:
        indices = np.arange(spec.n_stages)
        x, holdup = unpack_source_state(
            state,
            n_stages=spec.n_stages,
        )
        state = pack_source_state(
            x + 1.0e-3 * np.sin(indices),
            holdup + 2.0e-3 * np.cos(indices),
        )
    times = np.arange(0.0, 31.0, 1.0)
    options = SourceIntegrationOptions(
        method="BDF",
        rtol=1.0e-10,
        atol=1.0e-12,
        max_step_min=1.0,
    )
    v2 = integrate_source_trajectory(
        base_spec=spec,
        initial_packed_state=state,
        time_min=times,
        options=options,
        feed_schedule=feed_schedule,
    )
    reference = integrate_source_trajectory(
        base_spec=spec,
        initial_packed_state=state,
        time_min=times,
        options=options,
        feed_schedule=feed_schedule,
        augmented_rhs=_reference_rhs(_reference_module()),
    )

    scale = np.maximum(np.abs(reference.packed_state), 1.0e-6)
    error = np.max(
        np.abs(v2.packed_state - reference.packed_state) / scale
    )
    assert error < 1.0e-9
    assert not v2.safeguard_activated


def test_dd079_integrated_total_and_light_material_conservation():
    spec, state = _initial_state()
    schedule = SourceFeedSchedule(
        step_time_min=0.0,
        feed_before_kmol_min=1.0,
        feed_after_kmol_min=1.01,
    )
    trajectory = integrate_source_trajectory(
        base_spec=spec,
        initial_packed_state=state,
        time_min=np.arange(0.0, 101.0, 1.0),
        feed_schedule=schedule,
    )
    holdup = trajectory.liquid_holdup_kmol
    x = trajectory.light_mole_fraction
    total_delta = np.sum(holdup, axis=1) - np.sum(holdup[0])
    light_inventory = np.sum(x * holdup, axis=1)
    light_delta = light_inventory - light_inventory[0]

    total_error = np.max(
        np.abs(
            total_delta
            - trajectory.cumulative_external_total_kmol
        )
    ) / np.sum(holdup[0])
    light_error = np.max(
        np.abs(
            light_delta
            - trajectory.cumulative_external_light_kmol
        )
    ) / light_inventory[0]
    assert total_error < 1.0e-10
    assert light_error < 1.0e-10


def test_dd079_bdf_and_radau_trajectories_are_converged():
    spec, state = _initial_state()
    schedule = SourceFeedSchedule(
        step_time_min=0.0,
        feed_before_kmol_min=1.0,
        feed_after_kmol_min=1.01,
    )
    times = np.arange(0.0, 31.0, 1.0)
    bdf = integrate_source_trajectory(
        base_spec=spec,
        initial_packed_state=state,
        time_min=times,
        options=SourceIntegrationOptions(
            method="BDF",
            rtol=1.0e-10,
            atol=1.0e-12,
            max_step_min=1.0,
        ),
        feed_schedule=schedule,
    )
    radau = integrate_source_trajectory(
        base_spec=spec,
        initial_packed_state=state,
        time_min=times,
        options=SourceIntegrationOptions(
            method="Radau",
            rtol=2.0e-11,
            atol=2.0e-13,
            max_step_min=0.5,
        ),
        feed_schedule=schedule,
    )

    scale = np.maximum(np.abs(radau.packed_state), 1.0e-6)
    difference = np.max(
        np.abs(bdf.packed_state - radau.packed_state) / scale
    )
    assert difference < 1.0e-7


def test_dd079_terminal_component_withdrawal_uses_live_composition():
    spec = BinarySourceColumnSpec(n_stages=5, feed_stage_from_bottom=3)
    x_first = np.asarray([0.1, 0.2, 0.3, 0.4, 0.8])
    x_second = np.asarray([0.12, 0.2, 0.3, 0.4, 0.75])

    _, light_first = external_material_rates(spec, x_first)
    _, light_second = external_material_rates(spec, x_second)

    expected_first = (
        spec.feed_kmol_min * spec.feed_light_mole_fraction
        - spec.distillate_kmol_min * x_first[-1]
        - spec.bottoms_kmol_min * x_first[0]
    )
    expected_second = (
        spec.feed_kmol_min * spec.feed_light_mole_fraction
        - spec.distillate_kmol_min * x_second[-1]
        - spec.bottoms_kmol_min * x_second[0]
    )
    assert light_first == pytest.approx(expected_first)
    assert light_second == pytest.approx(expected_second)
    assert light_first != light_second


def test_dd079_integration_rejects_invalid_initial_state_without_repair():
    spec, state = _initial_state()
    state[3] = 1.01

    with pytest.raises(ValueError, match="compositions"):
        integrate_source_trajectory(
            base_spec=spec,
            initial_packed_state=state,
            time_min=np.asarray([0.0, 1.0]),
        )
