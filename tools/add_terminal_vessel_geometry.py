#!/usr/bin/env python
"""Size terminal liquid vessels and write their geometry to a case workbook."""

from __future__ import annotations

import argparse
import math
import shutil
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_PATH = PROJECT_ROOT / "src"
if str(SRC_PATH) not in sys.path:
    sys.path.insert(0, str(SRC_PATH))

from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel  # noqa: E402
from dynamic_distillation.thermo_provider_v1 import ThermoProviderV1  # noqa: E402


@dataclass(frozen=True)
class HorizontalDrumGeometry:
    total_volume_ft3: float
    normal_liquid_volume_ft3: float
    diameter_ft: float
    tangent_length_ft: float
    overall_length_ft: float
    normal_holdup_lbmol: float


@dataclass(frozen=True)
class VerticalSumpGeometry:
    total_volume_ft3: float
    normal_liquid_volume_ft3: float
    diameter_ft: float
    total_height_ft: float
    required_height_ft: float
    normal_liquid_height_ft: float
    low_liquid_height_ft: float
    normal_holdup_lbmol: float
    usable_residence_time_min: float


def size_horizontal_hemispherical_drum(
    *,
    liquid_flow_lbmolph: float,
    liquid_density_lbmol_ft3: float,
    residence_time_min: float,
    normal_liquid_fraction: float,
    tangent_length_to_diameter: float,
) -> HorizontalDrumGeometry:
    """Size a horizontal cylinder with two hemispherical heads."""
    _require_positive("liquid flow", liquid_flow_lbmolph)
    _require_positive("liquid density", liquid_density_lbmol_ft3)
    _require_positive("residence time", residence_time_min)
    _require_fraction("normal liquid fraction", normal_liquid_fraction)
    _require_positive("tangent length-to-diameter ratio", tangent_length_to_diameter)

    holdup = liquid_flow_lbmolph * residence_time_min / 60.0
    liquid_volume = holdup / liquid_density_lbmol_ft3
    total_volume = liquid_volume / normal_liquid_fraction

    # Cylinder plus two hemispherical heads (equivalent to one sphere).
    volume_coefficient = math.pi * (tangent_length_to_diameter / 4.0 + 1.0 / 6.0)
    diameter = (total_volume / volume_coefficient) ** (1.0 / 3.0)
    tangent_length = tangent_length_to_diameter * diameter
    return HorizontalDrumGeometry(
        total_volume_ft3=total_volume,
        normal_liquid_volume_ft3=liquid_volume,
        diameter_ft=diameter,
        tangent_length_ft=tangent_length,
        overall_length_ft=tangent_length + diameter,
        normal_holdup_lbmol=holdup,
    )


def size_vertical_cylindrical_sump(
    *,
    liquid_flow_lbmolph: float,
    liquid_density_lbmol_ft3: float,
    residence_time_min: float,
    normal_liquid_fraction: float,
    diameter_ft: float,
    low_liquid_fraction: float = 0.0,
    minimum_height_ft: float = 0.0,
) -> VerticalSumpGeometry:
    """Size a vertical sump for usable inventory between normal and low level."""
    _require_positive("liquid flow", liquid_flow_lbmolph)
    _require_positive("liquid density", liquid_density_lbmol_ft3)
    _require_positive("residence time", residence_time_min)
    _require_fraction("normal liquid fraction", normal_liquid_fraction)
    _require_positive("sump diameter", diameter_ft)
    if not math.isfinite(float(low_liquid_fraction)) or not 0.0 <= float(low_liquid_fraction) < normal_liquid_fraction:
        raise ValueError("low liquid fraction must be at least zero and below normal liquid fraction")
    if not math.isfinite(float(minimum_height_ft)) or float(minimum_height_ft) < 0.0:
        raise ValueError("minimum sump height must be non-negative")

    required_usable_holdup = liquid_flow_lbmolph * residence_time_min / 60.0
    required_usable_volume = required_usable_holdup / liquid_density_lbmol_ft3
    area = math.pi * diameter_ft**2 / 4.0
    required_height = required_usable_volume / (
        area * (normal_liquid_fraction - low_liquid_fraction)
    )
    total_height = max(required_height, minimum_height_ft)
    total_volume = area * total_height
    liquid_volume = normal_liquid_fraction * total_volume
    holdup = liquid_volume * liquid_density_lbmol_ft3
    usable_volume = (normal_liquid_fraction - low_liquid_fraction) * total_volume
    usable_residence_time = (
        usable_volume * liquid_density_lbmol_ft3 / liquid_flow_lbmolph * 60.0
    )
    return VerticalSumpGeometry(
        total_volume_ft3=total_volume,
        normal_liquid_volume_ft3=liquid_volume,
        diameter_ft=diameter_ft,
        total_height_ft=total_height,
        required_height_ft=required_height,
        normal_liquid_height_ft=normal_liquid_fraction * total_height,
        low_liquid_height_ft=low_liquid_fraction * total_height,
        normal_holdup_lbmol=holdup,
        usable_residence_time_min=usable_residence_time,
    )


def _require_positive(label: str, value: float) -> None:
    if not math.isfinite(float(value)) or float(value) <= 0.0:
        raise ValueError(f"{label} must be positive")


def _require_fraction(label: str, value: float) -> None:
    if not math.isfinite(float(value)) or not 0.0 < float(value) < 1.0:
        raise ValueError(f"{label} must be between zero and one")


def _normalized(value: Any) -> str:
    return "".join(ch for ch in str(value or "").lower() if ch.isalnum())


def _find_product_stream(streams: Mapping[str, Mapping[str, Any]], aliases: set[str]) -> Mapping[str, Any]:
    for name, stream in streams.items():
        if _normalized(name) in aliases:
            return stream
    raise KeyError(f"Could not find product stream matching {sorted(aliases)}")


def _stream_composition(stream: Mapping[str, Any], components: list[str]) -> list[float]:
    component_flows = stream.get("Component Mole Flows (lbmol/h)", {})
    values = [float(component_flows.get(name, 0.0)) for name in components]
    total = sum(values)
    if total <= 0.0:
        raise ValueError("Product stream has no positive component flow breakdown")
    return [value / total for value in values]


def _set_spec(ws, label: str, value: Any) -> None:
    target = _normalized(label)
    for row in range(1, ws.max_row + 1):
        if _normalized(ws.cell(row, 1).value) == target:
            ws.cell(row, 2).value = value
            return
    ws.cell(ws.max_row + 1, 1).value = label
    ws.cell(ws.max_row, 2).value = value


def _append_note(ws, field: str, value: Any) -> None:
    ws.append([field, value])


def add_terminal_geometry(
    *,
    input_path: Path,
    output_path: Path,
    residence_time_min: float,
    initial_level_fraction: float,
    sump_low_level_fraction: float,
    minimum_sump_height_ft: float,
    drum_length_to_diameter: float,
    property_package: str,
    overwrite: bool,
) -> tuple[HorizontalDrumGeometry, VerticalSumpGeometry, float, float, float, float, float]:
    if output_path.exists() and not overwrite:
        raise FileExistsError(f"Output already exists: {output_path}; pass --overwrite to replace it")
    if input_path.resolve() == output_path.resolve():
        raise ValueError("Input and output workbooks must be different")

    case = load_case_from_excel(str(input_path))
    top_stream = _find_product_stream(case.streams, {"top", "distillate", "overhead", "producttop"})
    bottom_stream = _find_product_stream(case.streams, {"bottom", "bottoms", "bottomproduct", "productbottom"})
    geometry_sections = case.specs.get("Geometry Sections") or []
    if not geometry_sections:
        raise ValueError("Workbook has no tray diameter in Geometry Sections")

    reflux_lbmolph = float(case.initial_conditions.iloc[0]["Liquid Flow (lbmol/h)"])
    distillate_lbmolph = float(top_stream["Total Molar Flow (lbmol/h)"])
    bottoms_lbmolph = float(bottom_stream["Total Molar Flow (lbmol/h)"])
    tray_diameter_ft = float(geometry_sections[-1]["diameter_ft"])

    provider = ThermoProviderV1(
        component_names_excel=case.components,
        component_ids_dwsim=case.component_ids_dwsim,
        property_package=property_package,
        silence_backend_console=True,
    )
    top_x = _stream_composition(top_stream, case.components)
    bottom_x = _stream_composition(bottom_stream, case.components)
    top_density = provider.liquid_density_lbmol_ft3(
        float(top_stream["Temperature (F)"]),
        float(top_stream["Pressure (psia)"]),
        top_x,
    )
    bottom_density = provider.liquid_density_lbmol_ft3(
        float(bottom_stream["Temperature (F)"]),
        float(bottom_stream["Pressure (psia)"]),
        bottom_x,
    )
    if top_density is None or bottom_density is None:
        raise RuntimeError("Thermo provider did not return both product liquid densities")

    drum = size_horizontal_hemispherical_drum(
        liquid_flow_lbmolph=reflux_lbmolph + distillate_lbmolph,
        liquid_density_lbmol_ft3=float(top_density),
        residence_time_min=residence_time_min,
        normal_liquid_fraction=initial_level_fraction,
        tangent_length_to_diameter=drum_length_to_diameter,
    )
    sump = size_vertical_cylindrical_sump(
        liquid_flow_lbmolph=bottoms_lbmolph,
        liquid_density_lbmol_ft3=float(bottom_density),
        residence_time_min=residence_time_min,
        normal_liquid_fraction=initial_level_fraction,
        diameter_ft=tray_diameter_ft,
        low_liquid_fraction=sump_low_level_fraction,
        minimum_height_ft=minimum_sump_height_ft,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(input_path, output_path)
    workbook = load_workbook(output_path)
    specs = workbook["Specifications"]
    _set_spec(specs, "Top Accumulator Holdup (lbmol)", drum.normal_holdup_lbmol)
    _set_spec(specs, "Top Drum Total Volume (ft3)", drum.total_volume_ft3)
    _set_spec(specs, "Top Drum Diameter (ft)", drum.diameter_ft)
    _set_spec(specs, "Top Drum Length (ft)", drum.tangent_length_ft)
    _set_spec(specs, "Top Drum Liquid Fraction", initial_level_fraction)
    _set_spec(specs, "Bottom Holdup (lbmol)", sump.normal_holdup_lbmol)
    _set_spec(specs, "Bottom Sump Total Volume (ft3)", sump.total_volume_ft3)
    _set_spec(specs, "Bottom Sump Diameter (ft)", sump.diameter_ft)
    _set_spec(specs, "Bottom Sump Height (ft)", sump.total_height_ft)
    _set_spec(specs, "Bottom Sump Liquid Fraction", initial_level_fraction)

    notes = workbook["Notes"] if "Notes" in workbook.sheetnames else workbook.create_sheet("Notes")
    if notes.max_row == 1 and notes.cell(1, 1).value is None:
        notes.append(["Field", "Value"])
    _append_note(notes, "Terminal vessel sizing basis", f"{residence_time_min:g} min liquid residence at {initial_level_fraction:.1%} level")
    _append_note(notes, "Distillate drum liquid rate", f"Reflux + distillate = {reflux_lbmolph + distillate_lbmolph:.6g} lbmol/h")
    _append_note(notes, "Distillate drum shape", f"Horizontal cylinder, two hemispherical heads, tangent L/D = {drum_length_to_diameter:g}")
    _append_note(notes, "Distillate drum overall length", f"{drum.overall_length_ft:.6g} ft including heads")
    _append_note(
        notes,
        "Bottom sump sizing basis",
        f"{residence_time_min:g} min usable drawdown from {initial_level_fraction:.1%} normal to "
        f"{sump_low_level_fraction:.1%} low level at tray diameter = {tray_diameter_ft:.6g} ft",
    )
    _append_note(notes, "Bottom sump calculated minimum height", f"{sump.required_height_ft:.6g} ft")
    _append_note(notes, "Bottom sump selected height", f"{sump.total_height_ft:.6g} ft")
    _append_note(notes, "Bottom sump normal liquid height", f"{sump.normal_liquid_height_ft:.6g} ft")
    _append_note(notes, "Bottom sump low liquid height", f"{sump.low_liquid_height_ft:.6g} ft")
    _append_note(notes, "Bottom sump usable residence time", f"{sump.usable_residence_time_min:.6g} min")
    _append_note(notes, "Terminal liquid density source", f"DWSIM {property_package.upper()} at product temperature, pressure, and composition")
    _append_note(notes, "Distillate liquid density", f"{float(top_density):.9g} lbmol/ft3")
    _append_note(notes, "Bottoms liquid density", f"{float(bottom_density):.9g} lbmol/ft3")
    workbook.save(output_path)
    workbook.close()

    return (
        drum,
        sump,
        reflux_lbmolph,
        distillate_lbmolph,
        bottoms_lbmolph,
        float(top_density),
        float(bottom_density),
    )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, help="Existing dynamic-model .xlsx workbook")
    parser.add_argument("--output", required=True, help="New .xlsx workbook with terminal geometry")
    parser.add_argument("--residence-time-min", type=float, default=10.0)
    parser.add_argument("--initial-level-fraction", type=float, default=0.5)
    parser.add_argument("--sump-low-level-fraction", type=float, default=0.25)
    parser.add_argument("--minimum-sump-height-ft", type=float, default=0.0)
    parser.add_argument("--drum-length-to-diameter", type=float, default=3.0)
    parser.add_argument("--property-package", default="unifac")
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args()

    result = add_terminal_geometry(
        input_path=Path(args.input).resolve(),
        output_path=Path(args.output).resolve(),
        residence_time_min=float(args.residence_time_min),
        initial_level_fraction=float(args.initial_level_fraction),
        sump_low_level_fraction=float(args.sump_low_level_fraction),
        minimum_sump_height_ft=float(args.minimum_sump_height_ft),
        drum_length_to_diameter=float(args.drum_length_to_diameter),
        property_package=str(args.property_package),
        overwrite=bool(args.overwrite),
    )
    drum, sump, reflux, distillate, bottoms, top_density, bottom_density = result
    print(f"created: {Path(args.output).resolve()}")
    print(f"flows_lbmolph: reflux={reflux:.6f}, distillate={distillate:.6f}, bottoms={bottoms:.6f}")
    print(f"densities_lbmol_ft3: distillate={top_density:.9f}, bottoms={bottom_density:.9f}")
    print(
        "drum: "
        f"volume={drum.total_volume_ft3:.6f} ft3, ID={drum.diameter_ft:.6f} ft, "
        f"tangent_length={drum.tangent_length_ft:.6f} ft, overall_length={drum.overall_length_ft:.6f} ft"
    )
    print(
        "sump: "
        f"volume={sump.total_volume_ft3:.6f} ft3, diameter={sump.diameter_ft:.6f} ft, "
        f"required_height={sump.required_height_ft:.6f} ft, selected_height={sump.total_height_ft:.6f} ft, "
        f"normal_liquid_height={sump.normal_liquid_height_ft:.6f} ft, "
        f"usable_residence={sump.usable_residence_time_min:.6f} min"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
