from dataclasses import replace
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

from dynamic_distillation.core_v2.source_equation_gate_v1 import (
    BinarySourceColumnSpec,
    evaluate_binary_source_column,
)


def _source_profile() -> np.ndarray:
    path = (
        Path(__file__).resolve().parents[1]
        / "validation_skogestad_column_a_relative_volatility.xlsx"
    )
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["Initial Conditions"]
        headers = {
            str(cell.value): index
            for index, cell in enumerate(next(sheet.iter_rows()), start=1)
        }
        source_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[headers["Source Stage"] - 1] is None:
                continue
            source_rows.append(
                (
                    int(row[headers["Source Stage"] - 1]),
                    float(
                        row[
                            headers[
                                "Liquid Composition Component 1"
                            ]
                            - 1
                        ]
                    ),
                )
            )
    finally:
        workbook.close()
    return np.asarray(
        [x for _, x in sorted(source_rows)],
        dtype=float,
    )


def _independent_source_translation(
    spec: BinarySourceColumnSpec,
    x: np.ndarray,
    holdup: np.ndarray,
) -> np.ndarray:
    nt = spec.n_stages
    nf = spec.feed_stage_from_bottom
    y = (
        spec.relative_volatility
        * x[:-1]
        / (1.0 + (spec.relative_volatility - 1.0) * x[:-1])
    )
    vapor = np.full(nt - 1, spec.boilup_kmol_min, dtype=float)
    vapor[nf - 1 :] += (
        (1.0 - spec.feed_liquid_fraction) * spec.feed_kmol_min
    )
    liquid = np.zeros(nt + 1, dtype=float)
    for stage in range(2, nf + 1):
        index = stage - 1
        liquid[index] = (
            spec.nominal_stripping_liquid_kmol_min
            + (
                holdup[index]
                - spec.nominal_liquid_holdup_kmol
            )
            / spec.liquid_hydraulic_tau_min
            + spec.liquid_vapor_coupling
            * (vapor[index - 1] - spec.nominal_boilup_kmol_min)
        )
    for stage in range(nf + 1, nt):
        index = stage - 1
        liquid[index] = (
            spec.nominal_rectifying_liquid_kmol_min
            + (
                holdup[index]
                - spec.nominal_liquid_holdup_kmol
            )
            / spec.liquid_hydraulic_tau_min
            + spec.liquid_vapor_coupling
            * (
                vapor[index - 1]
                - spec.nominal_rectifying_vapor_kmol_min
            )
        )
    liquid[nt - 1] = spec.reflux_kmol_min

    d_holdup = np.zeros(nt, dtype=float)
    d_light = np.zeros(nt, dtype=float)
    for stage in range(2, nt):
        index = stage - 1
        d_holdup[index] = (
            liquid[index + 1]
            - liquid[index]
            + vapor[index - 1]
            - vapor[index]
        )
        d_light[index] = (
            liquid[index + 1] * x[index + 1]
            - liquid[index] * x[index]
            + vapor[index - 1] * y[index - 1]
            - vapor[index] * y[index]
        )

    feed_index = nf - 1
    d_holdup[feed_index] += spec.feed_kmol_min
    d_light[feed_index] += (
        spec.feed_kmol_min * spec.feed_light_mole_fraction
    )
    d_holdup[0] = (
        liquid[1] - vapor[0] - spec.bottoms_kmol_min
    )
    d_light[0] = (
        liquid[1] * x[1]
        - vapor[0] * y[0]
        - spec.bottoms_kmol_min * x[0]
    )
    d_holdup[-1] = (
        vapor[-1]
        - spec.reflux_kmol_min
        - spec.distillate_kmol_min
    )
    d_light[-1] = (
        vapor[-1] * y[-1]
        - spec.reflux_kmol_min * x[-1]
        - spec.distillate_kmol_min * x[-1]
    )
    d_x = (d_light - x * d_holdup) / holdup
    return np.concatenate((d_x, d_holdup))


def test_dd078_v2_source_equations_match_published_assembly_at_steady_state():
    spec = BinarySourceColumnSpec()
    x = _source_profile()
    holdup = np.full(spec.n_stages, 0.5, dtype=float)

    evaluation = evaluate_binary_source_column(
        spec,
        light_mole_fraction=x,
        liquid_holdup_kmol=holdup,
    )
    reference = _independent_source_translation(spec, x, holdup)

    np.testing.assert_allclose(
        evaluation.packed_state_rate,
        reference,
        rtol=0.0,
        atol=2.0e-15,
    )
    assert np.max(np.abs(evaluation.packed_state_rate)) < 1.0e-7


def test_dd078_v2_source_equations_match_feed_and_state_perturbations():
    base = BinarySourceColumnSpec()
    spec = replace(
        base,
        feed_kmol_min=1.01,
        feed_liquid_fraction=0.82,
        liquid_vapor_coupling=0.11,
    )
    x = _source_profile()
    x = x + 1.0e-3 * np.sin(np.arange(spec.n_stages))
    holdup = (
        np.full(spec.n_stages, 0.5, dtype=float)
        + 2.0e-3 * np.cos(np.arange(spec.n_stages))
    )

    evaluation = evaluate_binary_source_column(
        spec,
        light_mole_fraction=x,
        liquid_holdup_kmol=holdup,
    )
    reference = _independent_source_translation(spec, x, holdup)

    np.testing.assert_allclose(
        evaluation.packed_state_rate,
        reference,
        rtol=2.0e-15,
        atol=2.0e-14,
    )


def test_dd078_source_assembly_is_generic_in_stage_and_feed_location():
    spec = BinarySourceColumnSpec(
        n_stages=7,
        feed_stage_from_bottom=4,
    )
    x = np.linspace(0.03, 0.97, spec.n_stages)
    holdup = np.linspace(0.45, 0.55, spec.n_stages)

    evaluation = evaluate_binary_source_column(
        spec,
        light_mole_fraction=x,
        liquid_holdup_kmol=holdup,
    )
    reference = _independent_source_translation(spec, x, holdup)

    assert evaluation.packed_state_rate.shape == (14,)
    np.testing.assert_allclose(
        evaluation.packed_state_rate,
        reference,
        rtol=2.0e-15,
        atol=2.0e-14,
    )


def test_dd078_source_assembly_closes_global_material_balances():
    spec = replace(
        BinarySourceColumnSpec(),
        feed_kmol_min=1.01,
    )
    x = _source_profile()
    holdup = np.full(spec.n_stages, 0.5, dtype=float)

    evaluation = evaluate_binary_source_column(
        spec,
        light_mole_fraction=x,
        liquid_holdup_kmol=holdup,
    )

    expected_total = (
        spec.feed_kmol_min
        - spec.distillate_kmol_min
        - spec.bottoms_kmol_min
    )
    expected_light = (
        spec.feed_kmol_min * spec.feed_light_mole_fraction
        - spec.distillate_kmol_min * x[-1]
        - spec.bottoms_kmol_min * x[0]
    )
    assert (
        abs(
            np.sum(evaluation.total_holdup_rate_kmol_min)
            - expected_total
        )
        < 1.0e-14
    )
    assert (
        abs(
            np.sum(evaluation.light_inventory_rate_kmol_min)
            - expected_light
        )
        < 1.0e-14
    )


def test_dd078_source_assembly_rejects_nonphysical_state_without_projection():
    spec = BinarySourceColumnSpec(n_stages=5, feed_stage_from_bottom=3)
    x = np.linspace(0.1, 0.9, spec.n_stages)
    x[2] = 1.01

    try:
        evaluate_binary_source_column(
            spec,
            light_mole_fraction=x,
            liquid_holdup_kmol=np.full(spec.n_stages, 0.5),
        )
    except ValueError as exc:
        assert "compositions" in str(exc)
    else:
        raise AssertionError("nonphysical composition was silently accepted")
