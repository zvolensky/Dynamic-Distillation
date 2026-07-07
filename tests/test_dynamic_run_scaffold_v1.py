"""
test_dynamic_run_scaffold_v1.py

Dynamic Distillation - Runner/CLI Unit Tests

PURPOSE
-------
Validate core runner scaffolding behavior in `dynamic_run_scaffold_v1`
without requiring full external thermo backends.

SCOPE
-----
- RunnerConfig behaviors and startup-state helpers
- controller helper utilities and guardrail logic
- summary/log-field population and selective runtime branches

KEY DEPENDENCIES
----------------
- dynamic_run_scaffold_v1 runner helpers
- column_rhs_v1 integration points
- numpy/pytest fixtures
"""


from __future__ import annotations

import json
from pathlib import Path

import numpy as np
import pytest
from openpyxl import Workbook, load_workbook

import dynamic_distillation.column_rhs_v1 as rhs_module
import dynamic_distillation.dynamic_run_scaffold_v1 as runmod
from dynamic_distillation.dynamic_run_scaffold_v1 import (
    PIController,
    _advance_explicit_euler_step,
    _autocalibrate_francis_hydraulic_c_factors_from_seed,
    _max_abs_temperature_fd_rate_per_s,
    _max_rel_inventory_fd_rate_detail_per_s,
    _max_rel_inventory_rate_detail_per_s,
    _effective_hydraulic_ida_profile,
    _integrate_one_step,
    _integrate_one_step_ida,
    _normalize_integrator_mode,
    _normalize_runtime_mode,
    _column_rhs_with_inner_pv_coupling,
    _solve_dae_pilot_algebraic,
    _apply_slew_limit,
    RunnerConfig,
    _clear_initial_tray_vapor_holdup,
    _component_index_by_name,
    _clip_temperature_states_to_provider_bounds,
    _initialize_hydraulic_energy_consistent_state,
    _initialize_restart_reentry_settling,
    _initialize_thermo_consistent_state,
    _initialize_vapor_holdup_from_spec_pressure,
    _initialize_top_drum_dynamic_steady,
    _refresh_tray_bubble_targets_F,
    _pi_update,
    _pressure_resid_gain_scale,
    _update_tray_temp_pressure_slope_F_per_psi,
    _resolve_parity_runtime_thermo_defer_visible_steps,
    _resolve_runtime_thermo_execution_plan,
    _resolve_residual_guarded_liquid_hydraulic_alpha,
    _resolve_residual_guarded_liquid_hydraulic_alpha_per_stage,
    _resolve_step0_startup_packet_phase_reuse_settings,
    _resolve_step0_startup_packet_reuse_thresholds,
    _resolve_startup_hydraulic_sequence_step,
    _resolve_startup_execution_flags,
    _sync_algebraic_tray_temperature_state,
    build_inputs_for_runner,
    load_native_checkpoint_initial_state,
    run_smoke_simulation,
    read_native_checkpoint,
    write_native_checkpoint_from_run_result,
    write_restart_workbook_from_run_result,
)
from dynamic_distillation.column_rhs_v1 import ColumnInputs, column_rhs
from dynamic_distillation.column_spec_builder_v1 import (
    ColumnGeometry,
    ColumnGeometrySection,
    ColumnSpec,
    HeatDuties,
    SimulationSettings,
    StreamSpecNormalized,
)
from dynamic_distillation.state_vector_layout_v1 import StateVectorLayout
from dynamic_distillation.stage_hydraulics_francis_v1 import compute_francis_weir_liquid_outflow


def test_write_restart_workbook_from_run_result_writes_boundary_state_sheet(tmp_path: Path):
    template = tmp_path / "template.xlsx"
    wb = Workbook()
    ws_specs = wb.active
    ws_specs.title = "Specifications"
    specs_rows = [
        ("Number of Stages", 2),
        ("Number of Components", 2),
        ("Simulation Length (min)", 1.0),
        ("Timestep (sec)", 1.0),
        ("Log Frequency (timesteps)", 1),
    ]
    for r, (k, v) in enumerate(specs_rows, start=1):
        ws_specs.cell(r, 1).value = k
        ws_specs.cell(r, 2).value = v

    ws_ic = wb.create_sheet("Initial Conditions")
    headers = [
        "Stage",
        "Temperature (F)",
        "Pressure (psia)",
        "Vapor Flow (lbmol/h)",
        "Liquid Flow (lbmol/h)",
        "Liquid Holdup (lbmol)",
        "Vapor Holdup (lbmol)",
        "Vapor Composition Component 1",
        "Vapor Composition Component 2",
        "Liquid Composition Component 1",
        "Liquid Composition Component 2",
    ]
    for c, h in enumerate(headers, start=1):
        ws_ic.cell(1, c).value = h
    wb.save(template)

    col = ColumnSpec(
        excel_path=str(template),
        components_excel=["Propane", "N-butane"],
        components_dwsim=["Propane", "N-butane"],
        n_components=2,
        n_stages=2,
        stage_1based=np.array([1, 2], dtype=int),
        sim=SimulationSettings(dt_sec=1.0, t_final_sec=10.0, log_every_n_steps=1),
        duties=HeatDuties(condenser_type="Total", q_cond_btu_per_h=0.0, q_reb_btu_per_h=0.0),
        specs_raw={
            "Number of Stages": 2,
            "Number of Components": 2,
            "Simulation Length (min)": 1.0,
            "Timestep (sec)": 1.0,
            "Log Frequency (timesteps)": 1,
        },
        T_f=np.array([100.0, 120.0], dtype=float),
        P_psia=np.array([200.0, 210.0], dtype=float),
        V_lbmolph=np.array([10.0, 20.0], dtype=float),
        L_lbmolph=np.array([30.0, 40.0], dtype=float),
        M_L_lbmol=np.array([5.0, 6.0], dtype=float),
        M_V_lbmol=np.array([1.0, 2.0], dtype=float),
        y0=np.array([[0.6, 0.4], [0.55, 0.45]], dtype=float),
        x0=np.array([[0.7, 0.3], [0.65, 0.35]], dtype=float),
        streams={},
    )
    layout = StateVectorLayout(
        n_stages=2,
        n_components=2,
        include_top=True,
        include_bottom=True,
        include_vapor=True,
        include_temperature=True,
        include_energy=True,
    )
    y = layout.pack_y0(col)
    sl = layout.slices()
    y[sl["tray_T_f"]] = np.array([111.0, 222.0], dtype=float)
    y[sl["top_L"]] = np.array([10.0, 20.0], dtype=float)
    y[sl["top_V"]] = np.array([1.0, 2.0], dtype=float)
    y[sl["bottom_L"]] = np.array([30.0, 40.0], dtype=float)
    y[sl["bottom_V"]] = np.array([3.0, 4.0], dtype=float)
    y[sl["tray_EL_BTU"]] = np.array([101.0, 102.0], dtype=float)
    y[sl["tray_EV_BTU"]] = np.array([201.0, 202.0], dtype=float)

    out_path = tmp_path / "restart.xlsx"
    write_restart_workbook_from_run_result(
        run_result={
            "excel_path": str(template),
            "final_state": y,
            "layout": layout,
            "column": col,
            "controller_state_final": {
                "top_level_integ": 1.5,
                "top_pressure_integ": -2.5,
                "top_pressure_pv_filt_psia": 221.25,
                "top_pressure_mv_cmd_btuph": -49640000.0,
                "top_pressure_resid_abs_btups": 922.55,
                "top_drum_pressure_T_prev_F": 115.75,
                "distillate_cmd_lbmolph": 2412.83,
                "bottoms_cmd_lbmolph": 4761.97,
                "reflux_cmd_lbmolph": 5945.41,
                "boilup_cmd_lbmolph": 8014.56,
                "distillate_comp_integ": 12.5,
                "bottoms_comp_integ": -7.25,
            },
            "last_diag": {
                "x_tray": col.x0,
                "y_tray": col.y0,
                "P_psia_hyd": col.P_psia,
                "L_out_used_lbmolph": col.L_lbmolph,
                "V_out_lbmolph": col.V_lbmolph,
            },
        },
        output_excel_path=str(out_path),
    )

    wb2 = load_workbook(out_path, data_only=True)
    assert "Boundary State" in wb2.sheetnames
    assert "Energy State" in wb2.sheetnames
    assert "Dynamic Memory" in wb2.sheetnames
    assert "Controller State" in wb2.sheetnames
    wsb = wb2["Boundary State"]
    assert wsb.cell(2, 1).value == "top_L"
    assert wsb.cell(2, 2).value == pytest.approx(10.0)
    assert wsb.cell(3, 1).value == "top_V"
    assert wsb.cell(3, 2).value == pytest.approx(1.0)
    assert wsb.cell(4, 1).value == "bottom_L"
    assert wsb.cell(4, 2).value == pytest.approx(30.0)
    assert wsb.cell(5, 1).value == "bottom_V"
    assert wsb.cell(5, 2).value == pytest.approx(3.0)
    wse = wb2["Energy State"]
    assert wse.cell(2, 1).value == 1
    assert wse.cell(2, 2).value == pytest.approx(101.0)
    assert wse.cell(2, 3).value == pytest.approx(201.0)
    assert wse.cell(3, 1).value == 2
    assert wse.cell(3, 2).value == pytest.approx(102.0)
    assert wse.cell(3, 3).value == pytest.approx(202.0)
    wsm = wb2["Dynamic Memory"]
    assert wsm.cell(2, 1).value == 1
    assert wsm.cell(2, 2).value == pytest.approx(200.0)
    assert wsm.cell(2, 3).value == pytest.approx(111.0)
    assert wsm.cell(3, 1).value == 2
    assert wsm.cell(3, 2).value == pytest.approx(210.0)
    assert wsm.cell(3, 3).value == pytest.approx(222.0)
    wsc = wb2["Controller State"]
    ctrl_map = {
        wsc.cell(r, 1).value: wsc.cell(r, 2).value
        for r in range(2, wsc.max_row + 1)
        if wsc.cell(r, 1).value is not None
    }
    assert ctrl_map["top_level_integ"] == pytest.approx(1.5)
    assert ctrl_map["top_pressure_integ"] == pytest.approx(-2.5)
    assert ctrl_map["top_pressure_pv_filt_psia"] == pytest.approx(221.25)
    assert ctrl_map["top_pressure_mv_cmd_btuph"] == pytest.approx(-49640000.0)
    assert ctrl_map["top_pressure_resid_abs_btups"] == pytest.approx(922.55)
    assert ctrl_map["top_drum_pressure_T_prev_F"] == pytest.approx(115.75)
    assert ctrl_map["distillate_cmd_lbmolph"] == pytest.approx(2412.83)
    assert ctrl_map["bottoms_cmd_lbmolph"] == pytest.approx(4761.97)
    assert ctrl_map["reflux_cmd_lbmolph"] == pytest.approx(5945.41)
    assert ctrl_map["boilup_cmd_lbmolph"] == pytest.approx(8014.56)
    assert ctrl_map["distillate_comp_integ"] == pytest.approx(12.5)
    assert ctrl_map["bottoms_comp_integ"] == pytest.approx(-7.25)

    checkpoint_path = tmp_path / "checkpoint.npz"
    write_native_checkpoint_from_run_result(
        run_result={
            "run_id": "test-run",
            "excel_path": str(template),
            "final_time_s": 12.5,
            "final_state": y,
            "layout": layout,
            "column": col,
            "controller_state_final": {"top_level_integ": 1.5},
            "steady_state_status_final": {"steady_state_flag": 0.0},
            "startup_seed_cache_info": {"loaded": False},
            "last_diag": {
                "x_tray": col.x0,
                "P_psia_hyd": col.P_psia,
            },
        },
        output_checkpoint_path=str(checkpoint_path),
    )
    checkpoint = read_native_checkpoint(checkpoint_path)
    assert checkpoint["metadata"]["schema"] == "dynamic_distillation.native_checkpoint.v1"
    assert checkpoint["metadata"]["run_id"] == "test-run"
    assert checkpoint["metadata"]["final_time_s"] == pytest.approx(12.5)
    assert checkpoint["metadata"]["layout"]["n_stages"] == 2
    assert checkpoint["metadata"]["layout"]["include_bottom"] is True
    assert checkpoint["metadata"]["controller_state_final"]["top_level_integ"] == pytest.approx(1.5)
    assert checkpoint["arrays"]["final_state"].tolist() == pytest.approx(y.tolist())
    assert checkpoint["arrays"]["diag__x_tray"].shape == (2, 2)
    assert checkpoint["arrays"]["diag__P_psia_hyd"].tolist() == pytest.approx([200.0, 210.0])


def test_fast_startup_skips_expensive_startup_passes():
    cfg = RunnerConfig(
        excel_path="dummy.xlsx",
        fast_startup=True,
        enable_startup_thermo_conditioning=True,
        enable_startup_hydraulic_energy_consistency=True,
    )

    flags = _resolve_startup_execution_flags(cfg)

    assert flags["fast_startup"] is True
    assert flags["enable_startup_thermo_conditioning"] is False
    assert flags["startup_thermo_conditioning_iters"] == 1
    assert flags["startup_thermo_conditioning_relaxation"] == pytest.approx(1.0)
    assert flags["enable_startup_hydraulic_energy_consistency"] is False
    assert flags["enable_top_drum_startup_steadying"] is False
    assert flags["top_drum_steady_max_iter"] == 2
    assert flags["top_drum_steady_tol_lbmolps"] == pytest.approx(1.0e-4)
    assert flags["top_drum_steady_wall_limit_sec"] == pytest.approx(30.0)


def test_load_native_checkpoint_initial_state_restores_state_and_memory(tmp_path: Path):
    col = ColumnSpec(
        excel_path=str(tmp_path / "case.xlsx"),
        components_excel=["A", "B"],
        components_dwsim=["A", "B"],
        n_components=2,
        n_stages=3,
        stage_1based=np.array([1, 2, 3], dtype=int),
        sim=SimulationSettings(dt_sec=1.0, t_final_sec=10.0, log_every_n_steps=1),
        duties=HeatDuties(condenser_type="Total", q_cond_btu_per_h=0.0, q_reb_btu_per_h=0.0),
        specs_raw={},
        T_f=np.array([100.0, 110.0, 120.0], dtype=float),
        P_psia=np.array([200.0, 205.0, 210.0], dtype=float),
        V_lbmolph=np.array([10.0, 11.0, 12.0], dtype=float),
        L_lbmolph=np.array([20.0, 21.0, 22.0], dtype=float),
        M_L_lbmol=np.array([5.0, 6.0, 7.0], dtype=float),
        M_V_lbmol=np.array([1.0, 1.1, 1.2], dtype=float),
        y0=np.array([[0.6, 0.4], [0.55, 0.45], [0.5, 0.5]], dtype=float),
        x0=np.array([[0.7, 0.3], [0.65, 0.35], [0.6, 0.4]], dtype=float),
        streams={},
    )
    layout = StateVectorLayout(
        n_stages=3,
        n_components=2,
        include_top=True,
        include_bottom=True,
        include_vapor=True,
        include_temperature=True,
        include_energy=True,
    )
    y = layout.pack_y0(col)
    y = y + np.linspace(0.0, 1.0, y.size)
    condenser_packet = runmod.CondenserDutyPacket(
        q_calc_BTUph=-12345.0,
        T_bubble_F=101.25,
        mode="total-condense",
        V_vapor_in_lbmolps=2.5,
        T_vapor_in_F=120.0,
        P_vapor_in_psia=205.0,
        P_condenser_psia=200.0,
        y_vapor_in=np.array([0.6, 0.4], dtype=float),
        hL_cond_BTU_lbmol=-10.5,
    )
    checkpoint_path = tmp_path / "checkpoint.npz"
    write_native_checkpoint_from_run_result(
        run_result={
            "run_id": "source-run",
            "excel_path": str(tmp_path / "case.xlsx"),
            "final_time_s": 42.0,
            "final_state": y,
            "layout": layout,
            "column": col,
            "last_diag": {
                "P_psia_hyd": np.array([201.0, 206.0, 211.0], dtype=float),
                "tray_T_f": np.array([101.0, 111.0, 121.0], dtype=float),
            },
            "last_condenser_duty_packet": condenser_packet,
        },
        output_checkpoint_path=str(checkpoint_path),
    )

    y_loaded, info, memory = load_native_checkpoint_initial_state(path=checkpoint_path, layout=layout, col=col)

    assert y_loaded.tolist() == pytest.approx(y.tolist())
    assert info["loaded"] is True
    assert info["source_run_id"] == "source-run"
    assert info["source_final_time_s"] == pytest.approx(42.0)
    assert memory["last_P_hyd"].tolist() == pytest.approx([201.0, 206.0, 211.0])
    assert memory["last_P_diag"].tolist() == pytest.approx([201.0, 206.0, 211.0])
    assert memory["last_T_tray"].tolist() == pytest.approx([101.0, 111.0, 121.0])
    assert memory["last_condenser_duty_packet"].q_calc_BTUph == pytest.approx(-12345.0)
    assert memory["last_condenser_duty_packet"].T_bubble_F == pytest.approx(101.25)
    assert memory["last_condenser_duty_packet"].y_vapor_in.tolist() == pytest.approx([0.6, 0.4])

    incompatible_layout = StateVectorLayout(
        n_stages=3,
        n_components=2,
        include_top=True,
        include_bottom=True,
        include_vapor=False,
        include_temperature=True,
        include_energy=True,
    )
    with pytest.raises(ValueError, match="layout is incompatible"):
        load_native_checkpoint_initial_state(path=checkpoint_path, layout=incompatible_layout, col=col)


def test_load_native_checkpoint_initial_state_seeds_temperature_memory_from_state(tmp_path: Path):
    col = ColumnSpec(
        excel_path=str(tmp_path / "case.xlsx"),
        components_excel=["A", "B"],
        components_dwsim=["A", "B"],
        n_components=2,
        n_stages=2,
        stage_1based=np.array([1, 2], dtype=int),
        sim=SimulationSettings(dt_sec=1.0, t_final_sec=10.0, log_every_n_steps=1),
        duties=HeatDuties(condenser_type="Total", q_cond_btu_per_h=0.0, q_reb_btu_per_h=0.0),
        specs_raw={},
        T_f=np.array([100.0, 120.0], dtype=float),
        P_psia=np.array([200.0, 210.0], dtype=float),
        V_lbmolph=np.array([10.0, 20.0], dtype=float),
        L_lbmolph=np.array([30.0, 40.0], dtype=float),
        M_L_lbmol=np.array([5.0, 6.0], dtype=float),
        M_V_lbmol=np.array([1.0, 2.0], dtype=float),
        y0=np.array([[0.6, 0.4], [0.55, 0.45]], dtype=float),
        x0=np.array([[0.7, 0.3], [0.65, 0.35]], dtype=float),
        streams={},
    )
    layout = StateVectorLayout(
        n_stages=2,
        n_components=2,
        include_top=True,
        include_bottom=True,
        include_vapor=True,
        include_temperature=True,
        include_energy=True,
    )
    y = layout.pack_y0(col)
    sl = layout.slices()
    y[sl["tray_T_f"]] = np.array([88.0, 144.0], dtype=float)
    checkpoint_path = tmp_path / "checkpoint_no_temperature_diag.npz"
    write_native_checkpoint_from_run_result(
        run_result={
            "run_id": "source-run",
            "excel_path": str(tmp_path / "case.xlsx"),
            "final_time_s": 12.0,
            "final_state": y,
            "layout": layout,
            "column": col,
            "last_diag": {
                "P_psia_hyd": np.array([201.0, 211.0], dtype=float),
            },
        },
        output_checkpoint_path=str(checkpoint_path),
    )

    _y_loaded, info, memory = load_native_checkpoint_initial_state(path=checkpoint_path, layout=layout, col=col)

    assert info["restored_memory_keys"] == ["last_P_diag", "last_P_hyd", "last_T_tray"]
    assert memory["last_T_tray"].tolist() == pytest.approx([88.0, 144.0])


def test_default_startup_execution_flags_preserve_existing_behavior():
    cfg = RunnerConfig(
        excel_path="dummy.xlsx",
        fast_startup=False,
        enable_startup_thermo_conditioning=True,
        enable_startup_hydraulic_energy_consistency=False,
    )

    flags = _resolve_startup_execution_flags(cfg)

    assert flags["fast_startup"] is False
    assert flags["enable_startup_thermo_conditioning"] is True
    assert flags["startup_thermo_conditioning_iters"] == 2
    assert flags["startup_thermo_conditioning_relaxation"] == pytest.approx(1.0)
    assert flags["enable_startup_hydraulic_energy_consistency"] is False
    assert flags["enable_top_drum_startup_steadying"] is True
    assert flags["top_drum_steady_max_iter"] == 6
    assert flags["top_drum_steady_tol_lbmolps"] == pytest.approx(1.0e-6)
    assert flags["top_drum_steady_wall_limit_sec"] is None
    assert flags["enable_restart_reentry_settling"] is True
    assert flags["restart_reentry_thermo_conditioning_iters"] == 1
    assert flags["restart_reentry_top_drum_max_iter"] == 2
    assert flags["restart_reentry_top_drum_wall_limit_sec"] == pytest.approx(10.0)


def test_parity_startup_execution_flags_skip_top_drum_steadying():
    cfg = RunnerConfig(
        excel_path="dummy.xlsx",
        runtime_mode="parity",
        fast_startup=True,
        enable_startup_thermo_conditioning=True,
    )

    flags = _resolve_startup_execution_flags(cfg)

    assert flags["enable_startup_thermo_conditioning"] is False
    assert flags["enable_top_drum_startup_steadying"] is False


def test_parity_startup_execution_flags_skip_thermo_conditioning():
    cfg = RunnerConfig(
        excel_path="dummy.xlsx",
        runtime_mode="parity",
        fast_startup=True,
        enable_startup_thermo_conditioning=True,
    )

    flags = _resolve_startup_execution_flags(cfg)

    assert flags["enable_startup_thermo_conditioning"] is False
    assert flags["enable_top_drum_startup_steadying"] is False


def test_disabling_startup_thermo_conditioning_also_skips_top_drum_steadying():
    cfg = RunnerConfig(
        excel_path="dummy.xlsx",
        runtime_mode="hydraulic",
        fast_startup=True,
        enable_startup_thermo_conditioning=False,
    )

    flags = _resolve_startup_execution_flags(cfg)

    assert flags["enable_startup_thermo_conditioning"] is False
    assert flags["enable_top_drum_startup_steadying"] is False


def test_resolve_parity_runtime_thermo_defer_visible_steps_uses_log_boundary():
    parity_cfg = RunnerConfig(excel_path="dummy.xlsx", runtime_mode="parity")
    legacy_cfg = RunnerConfig(excel_path="dummy.xlsx", runtime_mode="legacy")

    assert _resolve_parity_runtime_thermo_defer_visible_steps(parity_cfg, log_every_n_steps=5) == 5
    assert _resolve_parity_runtime_thermo_defer_visible_steps(parity_cfg, log_every_n_steps=1) == 1
    assert _resolve_parity_runtime_thermo_defer_visible_steps(legacy_cfg, log_every_n_steps=5) == 0


def test_initialize_restart_reentry_settling_runs_bounded_hidden_pass(monkeypatch: pytest.MonkeyPatch):
    calls: dict[str, object] = {"thermo_calls": []}

    def _fake_thermo(
        *,
        col,
        layout,
        y,
        inputs,
        include_temperature,
        max_iter,
        relaxation,
        preserve_tray_vapor_holdup=False,
    ):
        thermo_calls = calls.setdefault("thermo_calls", [])
        assert isinstance(thermo_calls, list)
        thermo_calls.append((include_temperature, max_iter, relaxation, preserve_tray_vapor_holdup))
        info = {
            "attempted": True,
            "success": True,
            "n_iter": int(max_iter),
            "max_dx": 0.0,
            "max_dy": 0.0,
            "eq_phase_change_init_lbmolps": 1.0,
            "eq_phase_change_final_lbmolps": 0.5,
        }
        return np.asarray(y, dtype=float) + 1.0, info

    def _fake_top_drum(*, col, layout, y, inputs, max_iter, tol_lbmolps, wall_limit_sec):
        calls["top_drum"] = (max_iter, tol_lbmolps, wall_limit_sec)
        info = {
            "attempted": True,
            "success": True,
            "n_iter": int(max_iter),
            "pressure_coupled": True,
            "hit_wall_limit": False,
            "d_top_L_init_lbmolps": 1.0,
            "d_top_V_init_lbmolps": 1.0,
            "d_top_L_final_lbmolps": 0.0,
            "d_top_V_final_lbmolps": 0.0,
        }
        return np.asarray(y, dtype=float) + 2.0, info

    monkeypatch.setattr(runmod, "_initialize_thermo_consistent_state", _fake_thermo)
    monkeypatch.setattr(runmod, "_initialize_top_drum_dynamic_steady", _fake_top_drum)

    y0 = np.array([1.0, 2.0], dtype=float)
    y_out, info = _initialize_restart_reentry_settling(
        col=object(),
        layout=object(),
        y=y0,
        inputs=object(),
        include_temperature=True,
        thermo_max_iter=1,
        thermo_relaxation=0.75,
        top_drum_max_iter=2,
        top_drum_tol_lbmolps=1.0e-4,
        top_drum_wall_limit_sec=12.0,
    )

    assert np.allclose(y_out, y0 + 3.0)
    thermo_calls = calls["thermo_calls"]
    assert isinstance(thermo_calls, list)
    assert len(thermo_calls) == 1
    assert thermo_calls[0][0] is True
    assert thermo_calls[0][1] == 1
    assert thermo_calls[0][2] == pytest.approx(0.75)
    assert thermo_calls[0][3] is True
    assert calls["top_drum"][0] == 2
    assert calls["top_drum"][1] == pytest.approx(1.0e-4)
    assert calls["top_drum"][2] == pytest.approx(12.0)
    assert info["attempted"] is True
    assert info["success"] is True
    assert info["thermo"]["n_iter"] == 1
    assert info["top_drum"]["n_iter"] == 2


def test_inventory_rate_detail_reports_tray_stage_and_component():
    layout = StateVectorLayout(
        n_stages=2,
        n_components=3,
        include_top=True,
        include_bottom=True,
        include_vapor=True,
        include_temperature=False,
        include_energy=False,
    )
    sl = layout.slices()
    y = np.zeros(layout.n_states(), dtype=float)
    dydt = np.zeros_like(y)
    tray_l = np.array([[10.0, 8.0, 6.0], [5.0, 4.0, 3.0]], dtype=float)
    y[sl["tray_L"]] = tray_l.ravel(order="C")
    tray_v = np.array([[2.0, 1.0, 0.5], [1.0, 0.5, 0.25]], dtype=float)
    y[sl["tray_V"]] = tray_v.ravel(order="C")
    dydt[sl["tray_V"]] = np.array([[0.0, 0.0, 0.0], [0.0, 0.0, 5.0]], dtype=float).ravel(order="C")

    detail = _max_rel_inventory_rate_detail_per_s(layout, y, dydt, denom_floor_lbmol=1.0)

    assert float(detail["max_rel_rate_per_s"]) == pytest.approx(4.0)
    assert detail["state_key"] == "tray_V"
    assert float(detail["stage_1based"]) == pytest.approx(2.0)
    assert float(detail["component_1based"]) == pytest.approx(3.0)


def test_sync_algebraic_tray_temperature_state_projects_state_to_diag_values():
    layout = StateVectorLayout(
        n_stages=3,
        n_components=2,
        include_top=False,
        include_bottom=False,
        include_vapor=True,
        include_temperature=True,
        include_energy=False,
    )
    sl = layout.slices()
    y = np.zeros(layout.n_states(), dtype=float)
    y[sl["tray_T_f"]] = np.array([100.0, 150.0, 200.0], dtype=float)
    diag = {
        "T_enthalpy_algebraic_F_tray": np.array([110.0, np.nan, 210.0], dtype=float),
    }

    y2 = _sync_algebraic_tray_temperature_state(
        y,
        layout,
        diag,
        thermo_provider=None,
    )

    assert np.allclose(
        y2[sl["tray_T_f"]],
        np.array([110.0, 150.0, 210.0], dtype=float),
        atol=1e-12,
    )


def test_sync_algebraic_tray_temperature_state_uses_bubble_target_when_present():
    layout = StateVectorLayout(
        n_stages=2,
        n_components=2,
        include_top=False,
        include_bottom=False,
        include_vapor=True,
        include_temperature=True,
        include_energy=False,
    )
    sl = layout.slices()
    y = np.zeros(layout.n_states(), dtype=float)
    y[sl["tray_T_f"]] = np.array([120.0, 140.0], dtype=float)
    diag = {
        "T_bubble_target_F_tray": np.array([121.5, 139.25], dtype=float),
    }

    y2 = _sync_algebraic_tray_temperature_state(
        y,
        layout,
        diag,
        thermo_provider=None,
    )

    assert np.allclose(
        y2[sl["tray_T_f"]],
        np.array([121.5, 139.25], dtype=float),
        atol=1e-12,
    )


def test_francis_hydraulics_uses_holdup_area_for_liquid_head():
    ml = np.array([0.0, 20.0, 0.0], dtype=float)
    rho = np.array([10.0, 10.0, 10.0], dtype=float)
    active_area = np.array([1.0, 1.0, 1.0], dtype=float)
    holdup_area = np.array([1.0, 2.0, 1.0], dtype=float)
    weir_h = np.zeros(3, dtype=float)
    weir_L = np.ones(3, dtype=float)

    base = compute_francis_weir_liquid_outflow(
        ML_lbmol=ml,
        rhoL_lbmol_ft3=rho,
        active_area_ft2=active_area,
        weir_height_in=weir_h,
        weir_length_ft=weir_L,
    )
    widened = compute_francis_weir_liquid_outflow(
        ML_lbmol=ml,
        rhoL_lbmol_ft3=rho,
        active_area_ft2=active_area,
        holdup_area_ft2=holdup_area,
        weir_height_in=weir_h,
        weir_length_ft=weir_L,
    )

    assert widened.h_ow[1] < base.h_ow[1]
    assert widened.ML_lbmolph[1] < base.ML_lbmolph[1]


def test_inventory_fd_rate_detail_reports_boundary_component():
    layout = StateVectorLayout(
        n_stages=3,
        n_components=2,
        include_top=True,
        include_bottom=True,
        include_vapor=True,
        include_temperature=False,
        include_energy=False,
    )
    sl = layout.slices()
    y0 = np.zeros(layout.n_states(), dtype=float)
    y1 = np.zeros(layout.n_states(), dtype=float)
    y0[sl["top_L"]] = np.array([100.0, 20.0], dtype=float)
    y1[sl["top_L"]] = np.array([100.0, 10.0], dtype=float)

    detail = _max_rel_inventory_fd_rate_detail_per_s(
        layout,
        y0,
        y1,
        dt_sec=10.0,
        denom_floor_lbmol=1.0,
    )

    assert float(detail["max_rel_rate_per_s"]) == pytest.approx(1.0 / 11.0)
    assert detail["state_key"] == "top_L"
    assert float(detail["stage_1based"]) == pytest.approx(0.0)
    assert float(detail["component_1based"]) == pytest.approx(2.0)


def test_steady_state_temperature_fd_rate_ignores_boundary_temperatures():
    layout = StateVectorLayout(
        n_stages=2,
        n_components=2,
        include_top=True,
        include_bottom=True,
        include_vapor=True,
        include_temperature=True,
        include_energy=False,
    )
    sl = layout.slices()
    y0 = np.zeros(layout.n_states(), dtype=float)
    y1 = np.zeros(layout.n_states(), dtype=float)

    # Keep tray temperatures fixed while perturbing boundary temperatures only.
    y0[sl["tray_T_f"]] = np.array([120.0, 140.0], dtype=float)
    y1[sl["tray_T_f"]] = np.array([120.0, 140.0], dtype=float)
    y0[sl["bottom_T_f"]] = np.array([180.0], dtype=float)
    y1[sl["bottom_T_f"]] = np.array([210.0], dtype=float)

    max_rate = _max_abs_temperature_fd_rate_per_s(layout, y0, y1, dt_sec=10.0)
    assert float(max_rate) == pytest.approx(0.0)


def test_smoke_runner_builds_and_rhs_runs(tmp_path: Path):
    # Use the project template if present; otherwise skip.
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        # In CI or partial checkouts, the template may not exist.
        return

    cfg = RunnerConfig(
        excel_path=str(excel),
        n_steps=2,
        dt_sec=0.1,
        log_every_n_steps=1,
        include_temperature=True,
        include_energy=False,
        enable_equilibrium_relaxation=True,
        thermo_mode="stub",
        logs_dir=str(tmp_path),
        write_logs=False,
    )

    out = run_smoke_simulation(cfg)

    assert out["profile_csv"] is None
    assert out["summary_csv"] is None

    layout = out["layout"]
    col = out["column"]
    inputs = out["inputs"]
    y_final = np.asarray(out["final_state"], dtype=float)

    assert y_final.size == layout.n_states()
    assert abs(out["final_time_s"] - cfg.n_steps * cfg.dt_sec) < 1e-12

    # One RHS call should produce equilibrium diagnostics when enabled
    _dydt, diag = column_rhs(out["final_time_s"], y_final, col, layout, inputs=inputs)
    assert "y_eq_tray" in diag


def test_smoke_runner_writes_startup_trace(tmp_path: Path):
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    cfg = RunnerConfig(
        excel_path=str(excel),
        n_steps=0,
        dt_sec=0.1,
        log_every_n_steps=1,
        include_temperature=True,
        include_energy=False,
        enable_equilibrium_relaxation=True,
        thermo_mode="stub",
        logs_dir=str(tmp_path),
        write_logs=False,
        run_name="startup_trace_smoke",
    )

    out = run_smoke_simulation(cfg)

    trace_path = Path(str(out["startup_trace_log"]))
    assert trace_path.exists()
    text = trace_path.read_text(encoding="utf-8")
    assert "[Milestone] start" in text
    assert "[Init] Building runner inputs and thermo provider" in text


def test_smoke_runner_writes_runtime_trace_markers(tmp_path: Path):
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    cfg = RunnerConfig(
        excel_path=str(excel),
        n_steps=1,
        dt_sec=0.1,
        log_every_n_steps=1,
        include_temperature=True,
        include_energy=False,
        enable_equilibrium_relaxation=True,
        thermo_mode="stub",
        logs_dir=str(tmp_path),
        write_logs=False,
        run_name="runtime_trace_smoke",
    )

    out = run_smoke_simulation(cfg)

    trace_path = Path(str(out["startup_trace_log"]))
    assert trace_path.exists()
    text = trace_path.read_text(encoding="utf-8")
    assert "[RuntimeTrace] step=0 setup" in text
    assert "[RuntimeTrace] step=0 integrate start" in text
    assert "[Progress] step=" in text


def test_runner_initializes_vapor_holdup_to_match_spec_pressure(tmp_path: Path):
    """At t=0, MV should be initialized so PV diagnostic pressure ~ P_spec.

    We skip condenser (stage index 0) because MV is intentionally zero there.
    """

    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    cfg = RunnerConfig(
        excel_path=str(excel),
        n_steps=0,  # only run initialization + one RHS evaluation
        dt_sec=0.1,
        log_every_n_steps=1,
        include_temperature=True,
        include_energy=False,
        enable_equilibrium_relaxation=True,
        thermo_mode="stub",
        logs_dir=str(tmp_path),
        write_logs=False,
    )

    out = run_smoke_simulation(cfg)

    col = out["column"]
    layout = out["layout"]
    inputs = out["inputs"]
    y0 = np.asarray(out["final_state"], dtype=float)

    # If the case lacks an operating pressure profile, nothing to check.
    if not hasattr(col, "P_psia"):
        return

    P_spec = np.asarray(col.P_psia, dtype=float).reshape((col.n_stages,))
    if not np.all(np.isfinite(P_spec[1:])):
        # If spec pressure is missing on trays, skip.
        return

    # Re-evaluate RHS at t=0 with thermo diagnostics to get P_psia_diag.
    _dydt, diag = column_rhs(0.0, y0, col, layout, inputs=inputs)
    assert "P_psia_diag" in diag

    P_diag = np.asarray(diag["P_psia_diag"], dtype=float).reshape((col.n_stages,))

    # Condenser stage has MV=0 by design.
    u = layout.unpack(y0)
    assert float(u["MV_tot_tray"][0]) < 1e-12

    # On trays, PV diagnostic pressure should match spec within a small tolerance.
    # The stub provider uses Z=1.0; initialization uses the same R and temperature basis.
    max_abs = float(np.max(np.abs(P_diag[1:] - P_spec[1:])))
    assert max_abs < 1e-2


def test_initialize_vapor_holdup_rescales_ev_when_energy_states_enabled():
    class TinyCol:
        n_stages = 2
        n_components = 2
        M_L_lbmol = np.array([5.0, 5.0], dtype=float)
        M_V_lbmol = np.array([0.0, 0.0], dtype=float)
        x0 = np.array([[0.8, 0.2], [0.3, 0.7]], dtype=float)
        y0 = np.array([[0.9, 0.1], [0.4, 0.6]], dtype=float)
        T_f = np.array([100.0, 120.0], dtype=float)
        P_psia = np.array([200.0, 210.0], dtype=float)
        streams = {}

    col = TinyCol()
    layout = StateVectorLayout(
        n_stages=2,
        n_components=2,
        include_top=False,
        include_bottom=False,
        include_vapor=True,
        include_temperature=False,
        include_energy=True,
    )
    y0 = layout.pack_y0(col)
    y0 = _clear_initial_tray_vapor_holdup(y0, layout)

    u0 = layout.unpack(y0)
    hV_prev = np.asarray(u0["tray_EV_BTU"], dtype=float) / np.maximum(
        np.asarray(u0["MV_tot_tray"], dtype=float),
        float(layout.epsilon_lbmol),
    )

    y1 = _initialize_vapor_holdup_from_spec_pressure(
        col=col,
        layout=layout,
        y=y0,
        inputs=ColumnInputs(Zfac_prev=np.ones(col.n_stages, dtype=float)),
        include_temperature=False,
    )
    u1 = layout.unpack(y1)
    mv1 = np.asarray(u1["MV_tot_tray"], dtype=float)
    ev1 = np.asarray(u1["tray_EV_BTU"], dtype=float)
    hV_new = ev1 / np.maximum(mv1, float(layout.epsilon_lbmol))

    assert float(mv1[1]) > 0.0
    assert abs(float(hV_new[1]) - float(hV_prev[1])) < 1e-9
    assert abs(float(ev1[0])) < 1e-12


def test_initialize_vapor_holdup_can_preserve_excel_tray_vapor_holdup():
    class TinyCol:
        n_stages = 2
        n_components = 2
        M_L_lbmol = np.array([5.0, 5.0], dtype=float)
        M_V_lbmol = np.array([0.2, 0.5], dtype=float)
        x0 = np.array([[0.8, 0.2], [0.3, 0.7]], dtype=float)
        y0 = np.array([[0.9, 0.1], [0.4, 0.6]], dtype=float)
        T_f = np.array([100.0, 120.0], dtype=float)
        P_psia = np.array([200.0, 210.0], dtype=float)
        streams = {}

    col = TinyCol()
    layout = StateVectorLayout(
        n_stages=2,
        n_components=2,
        include_top=False,
        include_bottom=False,
        include_vapor=True,
        include_temperature=False,
        include_energy=False,
    )
    y0 = layout.pack_y0(col)
    u0 = layout.unpack(y0)

    y1 = _initialize_vapor_holdup_from_spec_pressure(
        col=col,
        layout=layout,
        y=y0,
        inputs=ColumnInputs(Zfac_prev=np.ones(col.n_stages, dtype=float)),
        include_temperature=False,
        preserve_tray_vapor_holdup=True,
    )
    u1 = layout.unpack(y1)

    assert np.allclose(u1["tray_V"], u0["tray_V"])
    assert np.allclose(u1["MV_tot_tray"], col.M_V_lbmol)


def test_initialize_vapor_holdup_can_return_startup_thermo_diag(monkeypatch):
    class TinyCol:
        n_stages = 2
        n_components = 2
        M_L_lbmol = np.array([5.0, 5.0], dtype=float)
        M_V_lbmol = np.array([0.0, 0.0], dtype=float)
        x0 = np.array([[0.8, 0.2], [0.3, 0.7]], dtype=float)
        y0 = np.array([[0.9, 0.1], [0.4, 0.6]], dtype=float)
        T_f = np.array([100.0, 120.0], dtype=float)
        P_psia = np.array([200.0, 210.0], dtype=float)
        streams = {}

    col = TinyCol()
    layout = StateVectorLayout(
        n_stages=2,
        n_components=2,
        include_top=False,
        include_bottom=False,
        include_vapor=True,
        include_temperature=False,
        include_energy=False,
    )
    y0 = layout.pack_y0(col)
    y0 = _clear_initial_tray_vapor_holdup(y0, layout)

    seen = {}

    def fake_column_rhs(t_s, y_s, col_s, layout_s, inputs):
        seen["compute_thermo_diag"] = inputs.compute_thermo_diag
        return np.zeros_like(y_s), {
            "Z_tray": np.ones(col_s.n_stages, dtype=float),
            "z_overall_tray": np.array([[0.8, 0.2], [0.3, 0.7]], dtype=float),
            "K_tray": np.array([[1.1, 0.9], [0.6, 1.4]], dtype=float),
            "HL_BTU_lbmol_tray": np.array([-100.0, -80.0], dtype=float),
            "HV_BTU_lbmol_tray": np.array([120.0, 140.0], dtype=float),
            "x_eq_thermo_tray": np.array([[0.82, 0.18], [0.35, 0.65]], dtype=float),
            "y_eq_thermo_tray": np.array([[0.76, 0.24], [0.28, 0.72]], dtype=float),
        }

    monkeypatch.setattr(runmod, "column_rhs", fake_column_rhs)

    y1, diag = _initialize_vapor_holdup_from_spec_pressure(
        col=col,
        layout=layout,
        y=y0,
        inputs=ColumnInputs(),
        include_temperature=False,
        return_diag=True,
    )

    assert isinstance(diag, dict)
    assert seen["compute_thermo_diag"] is True
    assert "K_tray" in diag
    assert np.isfinite(np.sum(y1))


def test_initialize_vapor_holdup_prefers_direct_tray_refresh_when_provider_available(monkeypatch):
    class TinyCol:
        n_stages = 2
        n_components = 2
        M_L_lbmol = np.array([0.0, 0.0], dtype=float)
        M_V_lbmol = np.array([0.0, 0.0], dtype=float)
        x0 = np.array([[0.8, 0.2], [0.3, 0.7]], dtype=float)
        y0 = np.array([[0.9, 0.1], [0.4, 0.6]], dtype=float)
        T_f = np.array([100.0, 120.0], dtype=float)
        P_psia = np.array([200.0, 210.0], dtype=float)
        streams = {}

    class FakeProvider:
        def __init__(self):
            self.batch_calls = []

        def flash_TP_full_batch(self, T_rows_F, P_rows_psia, z_rows):
            self.batch_calls.append((list(T_rows_F), list(P_rows_psia), [list(z) for z in z_rows]))
            out = []
            for z in z_rows:
                z_arr = np.asarray(z, dtype=float).reshape((-1,))
                z_arr = z_arr / max(float(np.sum(z_arr)), 1.0e-300)
                out.append((z_arr, z_arr, np.ones_like(z_arr), -100.0, 100.0, 0.75))
            return out

    col = TinyCol()
    layout = StateVectorLayout(
        n_stages=2,
        n_components=2,
        include_top=False,
        include_bottom=False,
        include_vapor=True,
        include_temperature=False,
        include_energy=False,
    )
    y0 = layout.pack_y0(col)
    y0 = _clear_initial_tray_vapor_holdup(y0, layout)
    provider = FakeProvider()

    def _boom(*args, **kwargs):
        _ = (args, kwargs)
        raise AssertionError("column_rhs should not be used when direct startup tray refresh is available")

    monkeypatch.setattr(runmod, "column_rhs", _boom)

    y1, diag = _initialize_vapor_holdup_from_spec_pressure(
        col=col,
        layout=layout,
        y=y0,
        inputs=ColumnInputs(thermo_provider=provider),
        include_temperature=False,
        return_diag=True,
    )

    assert len(provider.batch_calls) == 1
    assert np.allclose(np.asarray(provider.batch_calls[0][2], dtype=float), col.x0)
    assert diag["startup_vapor_holdup_refresh_source"] == "direct-tray-refresh"
    assert bool(diag["startup_vapor_holdup_refresh_batch_used"]) is True
    assert np.allclose(np.asarray(diag["Z_tray"], dtype=float), np.full(col.n_stages, 0.75, dtype=float))
    assert np.isfinite(np.sum(y1))


def test_level_control_writes_dynamic_draws_to_summary_log(tmp_path: Path):
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    cfg = RunnerConfig(
        excel_path=str(excel),
        n_steps=1,
        dt_sec=0.1,
        log_every_n_steps=1,
        include_temperature=True,
        include_energy=False,
        enable_equilibrium_relaxation=True,
        thermo_mode="stub",
        logs_dir=str(tmp_path),
        write_logs=True,
        enable_level_control=True,
        top_level_sp_lbmol=300.0,
        bottom_level_sp_lbmol=700.0,
    )

    out = run_smoke_simulation(cfg)
    summary_csv = Path(str(out["summary_csv"]))
    assert summary_csv.exists()

    import csv

    with summary_csv.open("r", encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows, "summary log is empty"
    assert len(rows) >= 2

    r0 = rows[0]
    r1 = rows[1]
    d = float(r0["D_lbmolph"])
    b = float(r0["B_lbmolph"])
    assert np.isfinite(d) and np.isfinite(b)
    assert "Q_reb_used_BTUph" in r0
    assert np.isfinite(float(r0["Q_reb_used_BTUph"]))
    assert "M_total_lbmol" in r0
    assert "dM_total_dt_lbmolph" in r0
    assert "net_F_minus_D_minus_B_lbmolph" in r0
    assert "global_mass_closure_error_lbmolph" in r0
    assert "global_mass_closure_cum_lbmol" in r0
    assert np.isfinite(float(r0["M_total_lbmol"]))
    assert np.isfinite(float(r0["dM_total_dt_lbmolph"]))
    assert np.isfinite(float(r0["net_F_minus_D_minus_B_lbmolph"]))
    assert np.isfinite(float(r0["global_mass_closure_error_lbmolph"]))
    assert np.isfinite(float(r0["global_mass_closure_cum_lbmol"]))
    # Controller output is held at initialization row.
    assert d == pytest.approx(2380.99)
    assert b == pytest.approx(4761.98)
    # With setpoints far below initial inventories, both product draws should increase
    # on the first actionable timestep.
    assert float(r1["D_lbmolph"]) > 2380.99
    assert float(r1["B_lbmolph"]) > 4761.98


def test_runner_writes_run_metadata_json(tmp_path: Path):
    excel = Path("distillation_column_template_20stage_chemsep_warmer_feed_seed_20260323.xlsx")
    if not excel.exists():
        return

    cfg = RunnerConfig(
        excel_path=str(excel),
        n_steps=0,
        dt_sec=0.2,
        log_every_n_steps=1,
        include_temperature=True,
        include_energy=False,
        thermo_mode="stub",
        logs_dir=str(tmp_path),
        write_logs=True,
        run_name="UI metadata smoke",
        run_description="Ensure run metadata artifact is created.",
    )

    out = run_smoke_simulation(cfg)
    metadata_path = Path(str(out["run_metadata_json"]))
    assert metadata_path.exists()
    doc = json.loads(metadata_path.read_text(encoding="utf-8"))
    assert doc["run_name"] == "UI metadata smoke"
    assert doc["run_description"] == "Ensure run metadata artifact is created."
    assert doc["run_id"] == out["run_id"]
    assert doc["status"] == "completed"
    assert str(doc.get("restart_workbook", "")).endswith(".xlsx")
    assert Path(str(doc["restart_workbook"])).exists()
    assert Path(str(out["restart_workbook"])).exists()
    assert str(doc.get("native_checkpoint", "")).endswith(".npz")
    assert Path(str(doc["native_checkpoint"])).exists()
    assert Path(str(out["native_checkpoint"])).exists()
    checkpoint = read_native_checkpoint(str(out["native_checkpoint"]))
    assert checkpoint["metadata"]["run_id"] == out["run_id"]
    assert "final_state" in checkpoint["arrays"]


def test_snapshot_thermo_call_counters_merges_multiple_providers():
    class _Prov:
        def __init__(self, counters):
            self._counters = counters

        def get_call_counters(self):
            return self._counters

    merged = runmod._snapshot_thermo_call_counters(
        _Prov(
            {
                "main_tray_refresh": {"flash_requests": 3, "backend_flash_equivalents": 3, "wall_sec": 1.25},
                "temperature_state_cp_lookup": {"cp_requests": 2},
            }
        ),
        _Prov(
            {
                "main_tray_refresh": {"flash_requests": 1, "wall_sec": 0.75},
                "equilibrium_relaxation_flash": {"flash_requests": 4},
            }
        ),
    )

    assert merged["main_tray_refresh"]["flash_requests"] == 4
    assert merged["main_tray_refresh"]["backend_flash_equivalents"] == 3
    assert merged["main_tray_refresh"]["wall_sec"] == pytest.approx(2.0)
    assert merged["temperature_state_cp_lookup"]["cp_requests"] == 2
    assert merged["equilibrium_relaxation_flash"]["flash_requests"] == 4


def test_snapshot_thermo_call_counters_dedupes_repeated_provider_object():
    class _Prov:
        def __init__(self, counters):
            self._counters = counters

        def get_call_counters(self):
            return self._counters

    prov = _Prov(
        {
            "main_tray_refresh": {"flash_requests": 11, "backend_flash_equivalents": 11, "wall_sec": 12.5},
            "energy_vapor_flow_enthalpy_refresh": {"flash_requests": 9, "wall_sec": 8.75},
        }
    )

    merged = runmod._snapshot_thermo_call_counters(prov, prov, prov)

    assert merged["main_tray_refresh"]["flash_requests"] == 11
    assert merged["main_tray_refresh"]["backend_flash_equivalents"] == 11
    assert merged["main_tray_refresh"]["wall_sec"] == pytest.approx(12.5)
    assert merged["energy_vapor_flow_enthalpy_refresh"]["flash_requests"] == 9
    assert merged["energy_vapor_flow_enthalpy_refresh"]["wall_sec"] == pytest.approx(8.75)


def test_tray_thermo_packet_from_diag_extracts_packet():
    diag = {
        "z_overall_tray": np.array([[0.8, 0.2], [0.3, 0.7]], dtype=float),
        "K_tray": np.array([[1.1, 0.9], [0.6, 1.4]], dtype=float),
        "HL_BTU_lbmol_tray": np.array([-100.0, -80.0], dtype=float),
        "HV_BTU_lbmol_tray": np.array([120.0, 140.0], dtype=float),
        "Z_tray": np.array([0.98, 0.95], dtype=float),
        "cpL_BTU_lbmolF_tray": np.array([2.1, 2.4], dtype=float),
        "cpV_BTU_lbmolF_tray": np.array([3.1, 3.6], dtype=float),
        "x_eq_thermo_tray": np.array([[0.82, 0.18], [0.35, 0.65]], dtype=float),
        "y_eq_thermo_tray": np.array([[0.76, 0.24], [0.28, 0.72]], dtype=float),
    }

    packet = runmod._tray_thermo_packet_from_diag(
        diag,
        n_stages=2,
        n_components=2,
        T_tray_F=np.array([150.0, 190.0], dtype=float),
        P_tray_psia=np.array([14.7, 15.2], dtype=float),
    )

    assert packet is not None
    assert np.allclose(packet.z_overall, diag["z_overall_tray"])
    assert np.allclose(packet.K_tray, diag["K_tray"])
    assert np.allclose(packet.HL, diag["HL_BTU_lbmol_tray"])
    assert np.allclose(packet.HV, diag["HV_BTU_lbmol_tray"])
    assert np.allclose(packet.Zfac_tray, diag["Z_tray"])
    assert np.allclose(packet.cpL_tray, diag["cpL_BTU_lbmolF_tray"])
    assert np.allclose(packet.cpV_tray, diag["cpV_BTU_lbmolF_tray"])
    assert np.allclose(packet.x_eq, diag["x_eq_thermo_tray"])
    assert np.allclose(packet.y_eq, diag["y_eq_thermo_tray"])
    assert np.allclose(packet.T_state, np.array([150.0, 190.0], dtype=float))
    assert np.allclose(packet.P_state, np.array([14.7, 15.2], dtype=float))


def test_condenser_duty_packet_from_diag_extracts_packet():
    packet = runmod._condenser_duty_packet_from_diag(
        {
            "condenser_duty_cache_q_calc_BTUph": np.array([-1234.0], dtype=float),
            "condenser_duty_cache_T_bubble_F": np.array([111.0], dtype=float),
            "condenser_duty_cache_mode_total_condense": np.array([1.0], dtype=float),
            "condenser_duty_cache_V_vapor_in_lbmolps": np.array([1.2], dtype=float),
            "condenser_duty_cache_T_vapor_in_F": np.array([140.0], dtype=float),
            "condenser_duty_cache_P_vapor_in_psia": np.array([15.0], dtype=float),
            "condenser_duty_cache_P_condenser_psia": np.array([14.7], dtype=float),
            "condenser_duty_cache_y_vapor_in": np.array([0.2, 0.8], dtype=float),
        },
        n_components=2,
    )

    assert packet is not None
    assert packet.mode == "total-condense"
    assert abs(float(packet.q_calc_BTUph) + 1234.0) < 1e-12
    assert abs(float(packet.T_bubble_F) - 111.0) < 1e-12
    assert abs(float(packet.V_vapor_in_lbmolps) - 1.2) < 1e-12
    assert np.allclose(packet.y_vapor_in, np.array([0.2, 0.8], dtype=float))
    assert packet.hL_cond_BTU_lbmol is None


def test_seed_startup_condenser_duty_packet_uses_startup_tray_packet_and_profiles():
    col = ColumnSpec(
        excel_path="<unit-test>",
        components_excel=["A", "B"],
        components_dwsim=["A", "B"],
        n_components=2,
        n_stages=3,
        stage_1based=np.array([1, 2, 3], dtype=int),
        sim=SimulationSettings(dt_sec=1.0, t_final_sec=2.0, log_every_n_steps=1),
        duties=HeatDuties(condenser_type="Total", q_cond_btu_per_h=0.0, q_reb_btu_per_h=0.0),
        specs_raw={},
        T_f=np.array([120.0, 130.0, 140.0], dtype=float),
        P_psia=np.array([220.0, 221.0, 222.0], dtype=float),
        V_lbmolph=np.array([1000.0, 2000.0, 3000.0], dtype=float),
        L_lbmolph=np.array([4000.0, 5000.0, 6000.0], dtype=float),
        M_L_lbmol=np.array([10.0, 11.0, 12.0], dtype=float),
        M_V_lbmol=np.array([1.0, 2.0, 3.0], dtype=float),
        y0=np.array([[0.9, 0.1], [0.3, 0.7], [0.2, 0.8]], dtype=float),
        x0=np.array([[0.8, 0.2], [0.4, 0.6], [0.3, 0.7]], dtype=float),
        streams={},
    )
    layout = StateVectorLayout(
        n_stages=3,
        n_components=2,
        include_top=False,
        include_bottom=False,
        include_vapor=True,
        include_temperature=True,
        include_energy=False,
    )
    y = layout.pack_y0(col)
    startup_packet = runmod.TrayThermoPacket(
        z_overall_tray=np.array([[0.8, 0.2], [0.35, 0.65], [0.25, 0.75]], dtype=float),
        K_tray=np.ones((3, 2), dtype=float),
        HL_BTU_lbmol_tray=np.array([100.0, 110.0, 120.0], dtype=float),
        HV_BTU_lbmol_tray=np.array([200.0, 210.0, 220.0], dtype=float),
        Z_tray=np.array([0.95, 0.96, 0.97], dtype=float),
        T_tray_F=np.array([118.0, 128.0, 138.0], dtype=float),
        P_tray_psia=np.array([219.0, 220.0, 221.0], dtype=float),
    )

    packet = runmod._seed_startup_condenser_duty_packet(
        col=col,
        layout=layout,
        y=y,
        startup_packet=startup_packet,
        condenser_duty_mode="total-condense",
    )

    assert packet is not None
    assert packet.mode == "total-condense"
    assert packet.T_bubble_F == pytest.approx(118.0)
    assert packet.V_vapor_in_lbmolps == pytest.approx(2000.0 / 3600.0)
    assert packet.T_vapor_in_F == pytest.approx(128.0)
    assert packet.P_vapor_in_psia == pytest.approx(220.0)
    assert packet.P_condenser_psia == pytest.approx(219.0)
    assert np.allclose(packet.y_vapor_in, np.array([0.3, 0.7], dtype=float))
    assert packet.q_calc_BTUph == pytest.approx((2000.0 / 3600.0) * (100.0 - 210.0) * 3600.0)
    assert packet.hL_cond_BTU_lbmol == pytest.approx(100.0)


def test_startup_seed_cache_roundtrip_restores_state_and_packets(tmp_path: Path):
    col = ColumnSpec(
        excel_path=str(tmp_path / "case.xlsx"),
        components_excel=["A", "B"],
        components_dwsim=["A", "B"],
        n_components=2,
        n_stages=3,
        stage_1based=np.array([1, 2, 3], dtype=int),
        sim=SimulationSettings(dt_sec=1.0, t_final_sec=2.0, log_every_n_steps=1),
        duties=HeatDuties(condenser_type="Total", q_cond_btu_per_h=0.0, q_reb_btu_per_h=0.0),
        specs_raw={},
        T_f=np.array([120.0, 130.0, 140.0], dtype=float),
        P_psia=np.array([220.0, 221.0, 222.0], dtype=float),
        V_lbmolph=np.array([1000.0, 2000.0, 3000.0], dtype=float),
        L_lbmolph=np.array([4000.0, 5000.0, 6000.0], dtype=float),
        M_L_lbmol=np.array([10.0, 11.0, 12.0], dtype=float),
        M_V_lbmol=np.array([1.0, 2.0, 3.0], dtype=float),
        y0=np.array([[0.9, 0.1], [0.3, 0.7], [0.2, 0.8]], dtype=float),
        x0=np.array([[0.8, 0.2], [0.4, 0.6], [0.3, 0.7]], dtype=float),
        streams={},
    )
    layout = StateVectorLayout(
        n_stages=3,
        n_components=2,
        include_top=False,
        include_bottom=False,
        include_vapor=True,
        include_temperature=True,
        include_energy=False,
    )
    cfg = RunnerConfig(
        excel_path=str(tmp_path / "case.xlsx"),
        runtime_mode="hydraulic",
        thermo_mode="clapeyron",
        clapeyron_model="PR",
        include_temperature=True,
        include_energy=False,
        enable_startup_seed_cache=True,
        logs_dir=str(tmp_path / "logs"),
    )
    base_inputs = ColumnInputs(
        pressure_model="hydraulic",
        vapor_flow_model="energy",
        condenser_duty_mode="total-condense",
    )
    y = layout.pack_y0(col)
    tray_packet = runmod.TrayThermoPacket(
        z_overall_tray=np.array([[0.8, 0.2], [0.35, 0.65], [0.25, 0.75]], dtype=float),
        K_tray=np.array([[1.1, 0.9], [1.0, 1.0], [0.8, 1.2]], dtype=float),
        HL_BTU_lbmol_tray=np.array([100.0, 110.0, 120.0], dtype=float),
        HV_BTU_lbmol_tray=np.array([200.0, 210.0, 220.0], dtype=float),
        Z_tray=np.array([0.95, 0.96, 0.97], dtype=float),
        cpL_BTU_lbmolF_tray=np.array([2.0, 2.1, 2.2], dtype=float),
        cpV_BTU_lbmolF_tray=np.array([3.0, 3.1, 3.2], dtype=float),
        T_tray_F=np.array([118.0, 128.0, 138.0], dtype=float),
        P_tray_psia=np.array([219.0, 220.0, 221.0], dtype=float),
        x_equilibrium_tray=np.array([[0.82, 0.18], [0.34, 0.66], [0.20, 0.80]], dtype=float),
        y_equilibrium_tray=np.array([[0.76, 0.24], [0.30, 0.70], [0.18, 0.82]], dtype=float),
    )
    condenser_packet = runmod._seed_startup_condenser_duty_packet(
        col=col,
        layout=layout,
        y=y,
        startup_packet=tray_packet,
        condenser_duty_mode="total-condense",
    )
    feed_packet = runmod.FeedStageFlashPacket(
        stage0=1,
        T_feed_F=175.0,
        P_feed_psia=220.0,
        z_feed=np.array([0.4, 0.6], dtype=float),
        Fk_L_lbmolps=np.array([0.1, 0.2], dtype=float),
        Fk_V_lbmolps=np.array([0.3, 0.4], dtype=float),
        hL_BTU_lbmol=123.0,
        hV_BTU_lbmol=456.0,
    )
    bottom_sump_packet = runmod.BottomSumpCpPacket(
        T_sump_F=166.0,
        P_sump_psia=221.0,
        x_sump=np.array([0.22, 0.78], dtype=float),
        cpL_BTU_lbmolF=2.75,
    )
    path = runmod._resolve_startup_seed_cache_path(cfg)
    assert path is not None

    save_info = runmod._write_startup_seed_cache(
        path=path,
        cfg=cfg,
        col=col,
        layout=layout,
        base_inputs=base_inputs,
        y=y,
        last_T_tray=np.array([118.0, 128.0, 138.0], dtype=float),
        last_P_diag=np.array([219.0, 220.0, 221.0], dtype=float),
        last_P_hyd=np.array([219.0, 220.0, 221.0], dtype=float),
        last_K_tray=tray_packet.K_tray,
        last_HL=tray_packet.HL,
        last_HV=tray_packet.HV,
        last_Zfac=tray_packet.Zfac_tray,
        last_z_overall=tray_packet.z_overall,
        last_tray_bubble_target_F=np.array([118.0, 128.0, 138.0], dtype=float),
        last_tray_thermo_packet=tray_packet,
        last_condenser_duty_packet=condenser_packet,
        last_feed_stage_flash_packet=feed_packet,
        last_bottom_sump_cp_packet=bottom_sump_packet,
        last_reb_T=141.5,
        last_reb_x=np.array([0.25, 0.75], dtype=float),
        last_reb_y=np.array([0.4, 0.6], dtype=float),
        last_reb_beta=0.35,
        startup_seeded_condenser_duty_packet=True,
    )

    assert bool(save_info["saved"]) is True
    assert path.exists()

    loaded, load_info = runmod._load_startup_seed_cache(
        path=path,
        cfg=cfg,
        col=col,
        layout=layout,
        base_inputs=base_inputs,
    )

    assert loaded is not None
    assert bool(load_info["loaded"]) is True
    assert np.allclose(np.asarray(loaded["y"], dtype=float), y)
    assert np.allclose(np.asarray(loaded["last_T_tray"], dtype=float), np.array([118.0, 128.0, 138.0], dtype=float))
    assert np.allclose(np.asarray(loaded["last_K_tray"], dtype=float), tray_packet.K_tray)
    assert np.allclose(np.asarray(loaded["last_tray_thermo_packet"].HL, dtype=float), tray_packet.HL)
    assert loaded["last_condenser_duty_packet"] is not None
    assert loaded["last_condenser_duty_packet"].mode == "total-condense"
    assert loaded["last_condenser_duty_packet"].hL_cond_BTU_lbmol == pytest.approx(100.0)
    assert loaded["last_feed_stage_flash_packet"] is not None
    assert loaded["last_feed_stage_flash_packet"].hL_BTU_lbmol == pytest.approx(123.0)
    assert loaded["last_feed_stage_flash_packet"].hV_BTU_lbmol == pytest.approx(456.0)
    assert loaded["last_bottom_sump_cp_packet"] is not None
    assert loaded["last_bottom_sump_cp_packet"].cpL_BTU_lbmolF == pytest.approx(2.75)
    assert np.allclose(
        np.asarray(loaded["last_bottom_sump_cp_packet"].x_sump, dtype=float),
        np.array([0.22, 0.78], dtype=float),
    )
    assert loaded["last_reb_T"] == pytest.approx(141.5)
    assert np.allclose(np.asarray(loaded["last_reb_x"], dtype=float), np.array([0.25, 0.75], dtype=float))
    assert np.allclose(np.asarray(loaded["last_reb_y"], dtype=float), np.array([0.4, 0.6], dtype=float))
    assert loaded["last_reb_beta"] == pytest.approx(0.35)
    assert bool(loaded["startup_seeded_condenser_duty_packet"]) is True


def test_seed_startup_feed_stage_flash_packet_normalizes_component_names():
    col = ColumnSpec(
        excel_path="<unit-test>",
        components_excel=["n-Propane", "n-Butane", "n-Pentane"],
        components_dwsim=["Propane", "N-butane", "N-pentane"],
        n_components=3,
        n_stages=2,
        stage_1based=np.array([1, 2], dtype=int),
        sim=SimulationSettings(dt_sec=1.0, t_final_sec=1.0, log_every_n_steps=1),
        duties=HeatDuties(condenser_type="Total", q_cond_btu_per_h=0.0, q_reb_btu_per_h=0.0),
        specs_raw={},
        T_f=np.array([120.0, 140.0], dtype=float),
        P_psia=np.array([220.0, 225.0], dtype=float),
        V_lbmolph=np.array([1000.0, 1000.0], dtype=float),
        L_lbmolph=np.array([1000.0, 1000.0], dtype=float),
        M_L_lbmol=np.array([10.0, 10.0], dtype=float),
        M_V_lbmol=np.array([1.0, 1.0], dtype=float),
        y0=np.array([[0.7, 0.2, 0.1], [0.6, 0.25, 0.15]], dtype=float),
        x0=np.array([[0.65, 0.25, 0.10], [0.55, 0.30, 0.15]], dtype=float),
        streams={
            "Feed": StreamSpecNormalized(
                name="Feed",
                stage_1based=2,
                temperature_f=175.0,
                vapor_fraction=0.0,
                total_molar_flow_lbmolph=7142.98,
                component_molar_flows_lbmolph={
                    "n-Propane": 2380.99,
                    "n-Butane": 3968.32,
                    "N-Pentane": 793.664,
                },
            )
        },
    )

    class _FakeProvider:
        def flash_TP_full(self, T_f, P_psia, z):
            return type(
                "FlashResult",
                (),
                {
                    "K": np.array([1.2, 0.9, 0.6], dtype=float),
                    "HL_BTU_lbmol": 111.0,
                    "HV_BTU_lbmol": 222.0,
                },
            )()

    packet = runmod._seed_startup_feed_stage_flash_packet(
        col=col,
        thermo_provider=_FakeProvider(),
        P_tray_psia=np.array([220.0, 226.896], dtype=float),
    )

    assert packet is not None
    expected_fk = runmod._component_molar_flows_vector_lbmolps(
        {"n-Propane": 2380.99, "n-Butane": 3968.32, "N-Pentane": 793.664},
        np.asarray(col.components_excel, dtype=object),
    )
    expected_z = expected_fk / float(np.sum(expected_fk))
    assert np.allclose(packet.z_feed, expected_z)
    assert packet.z_feed[2] > 0.0
    assert packet.hL_BTU_lbmol == pytest.approx(111.0)
    assert packet.hV_BTU_lbmol == pytest.approx(222.0)


def test_resolve_step0_startup_packet_reuse_thresholds_relaxes_loaded_seed_defaults():
    base_inputs = ColumnInputs(
        boundary=rhs_module.BoundaryFlows(
            reflux_lbmolph=0.0,
            distillate_lbmolph=0.0,
            boilup_lbmolph=0.0,
            bottoms_lbmolph=0.0,
        )
    )
    packet = rhs_module.TrayThermoPacket(
        z_overall_tray=np.array([[1.0]], dtype=float),
        K_tray=np.array([[1.0]], dtype=float),
        HL_BTU_lbmol_tray=np.array([-100.0], dtype=float),
        HV_BTU_lbmol_tray=np.array([100.0], dtype=float),
        Z_tray=np.array([1.0], dtype=float),
    )

    reuse, dT, dP, dx = _resolve_step0_startup_packet_reuse_thresholds(
        startup_seed_loaded=True,
        runtime_mode="hydraulic",
        step=0,
        last_tray_thermo_packet=packet,
        last_T_tray=np.array([100.0], dtype=float),
        last_P_hyd=np.array([200.0], dtype=float),
        last_P_diag=None,
        last_z_overall=np.array([[1.0]], dtype=float),
        base_inputs=base_inputs,
    )

    assert reuse is True
    assert dT == pytest.approx(0.5)
    assert dP == pytest.approx(5.0)
    assert dx == pytest.approx(1.0e-5)


def test_resolve_step0_startup_packet_reuse_thresholds_keeps_legacy_defaults_without_loaded_seed():
    base_inputs = ColumnInputs(
        boundary=rhs_module.BoundaryFlows(
            reflux_lbmolph=0.0,
            distillate_lbmolph=0.0,
            boilup_lbmolph=0.0,
            bottoms_lbmolph=0.0,
        )
    )
    packet = rhs_module.TrayThermoPacket(
        z_overall_tray=np.array([[1.0]], dtype=float),
        K_tray=np.array([[1.0]], dtype=float),
        HL_BTU_lbmol_tray=np.array([-100.0], dtype=float),
        HV_BTU_lbmol_tray=np.array([100.0], dtype=float),
        Z_tray=np.array([1.0], dtype=float),
    )

    reuse, dT, dP, dx = _resolve_step0_startup_packet_reuse_thresholds(
        startup_seed_loaded=False,
        runtime_mode="hydraulic",
        step=0,
        last_tray_thermo_packet=packet,
        last_T_tray=np.array([100.0], dtype=float),
        last_P_hyd=np.array([200.0], dtype=float),
        last_P_diag=None,
        last_z_overall=np.array([[1.0]], dtype=float),
        base_inputs=base_inputs,
    )

    assert reuse is True
    assert dT == pytest.approx(1.0e-3)
    assert dP == pytest.approx(1.0e-3)
    assert dx == pytest.approx(1.0e-6)


def test_resolve_step0_startup_packet_phase_reuse_settings_relaxes_loaded_seed_vapor_dx():
    base_inputs = ColumnInputs(
        boundary=rhs_module.BoundaryFlows(
            reflux_lbmolph=0.0,
            distillate_lbmolph=0.0,
            boilup_lbmolph=0.0,
            bottoms_lbmolph=0.0,
        )
    )
    packet = rhs_module.TrayThermoPacket(
        z_overall_tray=np.array([[1.0]], dtype=float),
        K_tray=np.array([[1.0]], dtype=float),
        HL_BTU_lbmol_tray=np.array([-100.0], dtype=float),
        HV_BTU_lbmol_tray=np.array([100.0], dtype=float),
        Z_tray=np.array([1.0], dtype=float),
    )

    phase_dx, vapor_dx, phase_dT, phase_dP = _resolve_step0_startup_packet_phase_reuse_settings(
        startup_seed_loaded=True,
        runtime_mode="hydraulic",
        step=0,
        last_tray_thermo_packet=packet,
        last_T_tray=np.array([100.0], dtype=float),
        last_P_hyd=np.array([200.0], dtype=float),
        last_P_diag=None,
        last_z_overall=np.array([[1.0]], dtype=float),
        base_inputs=base_inputs,
    )

    assert phase_dx == pytest.approx(base_inputs.thermo_packet_phase_reuse_dx)
    assert vapor_dx == pytest.approx(0.25)
    assert phase_dT == pytest.approx(base_inputs.thermo_packet_phase_reuse_dT_F)
    assert phase_dP == pytest.approx(base_inputs.thermo_packet_phase_reuse_dP_psia)


def test_resolve_step0_startup_packet_phase_reuse_settings_keeps_base_without_loaded_seed():
    base_inputs = ColumnInputs(
        boundary=rhs_module.BoundaryFlows(
            reflux_lbmolph=0.0,
            distillate_lbmolph=0.0,
            boilup_lbmolph=0.0,
            bottoms_lbmolph=0.0,
        )
    )
    packet = rhs_module.TrayThermoPacket(
        z_overall_tray=np.array([[1.0]], dtype=float),
        K_tray=np.array([[1.0]], dtype=float),
        HL_BTU_lbmol_tray=np.array([-100.0], dtype=float),
        HV_BTU_lbmol_tray=np.array([100.0], dtype=float),
        Z_tray=np.array([1.0], dtype=float),
    )

    phase_dx, vapor_dx, phase_dT, phase_dP = _resolve_step0_startup_packet_phase_reuse_settings(
        startup_seed_loaded=False,
        runtime_mode="hydraulic",
        step=0,
        last_tray_thermo_packet=packet,
        last_T_tray=np.array([100.0], dtype=float),
        last_P_hyd=np.array([200.0], dtype=float),
        last_P_diag=None,
        last_z_overall=np.array([[1.0]], dtype=float),
        base_inputs=base_inputs,
    )

    assert phase_dx == pytest.approx(base_inputs.thermo_packet_phase_reuse_dx)
    assert vapor_dx == pytest.approx(base_inputs.thermo_packet_vapor_reuse_dx)
    assert phase_dT == pytest.approx(base_inputs.thermo_packet_phase_reuse_dT_F)
    assert phase_dP == pytest.approx(base_inputs.thermo_packet_phase_reuse_dP_psia)


def _make_runtime_thermo_execution_plan_fixture(
    *,
    thermo_refresh_dx: float | None = None,
    equilibrium_relaxation: bool = False,
):
    col = ColumnSpec(
        excel_path="<unit-test>",
        components_excel=["A", "B"],
        components_dwsim=["A", "B"],
        n_components=2,
        n_stages=2,
        stage_1based=np.array([1, 2], dtype=int),
        sim=SimulationSettings(dt_sec=1.0, t_final_sec=10.0, log_every_n_steps=1),
        duties=HeatDuties(condenser_type="Total", q_cond_btu_per_h=0.0, q_reb_btu_per_h=0.0),
        specs_raw={},
        T_f=np.array([120.0, 135.0], dtype=float),
        P_psia=np.array([210.0, 220.0], dtype=float),
        V_lbmolph=np.array([1200.0, 1300.0], dtype=float),
        L_lbmolph=np.array([2400.0, 2500.0], dtype=float),
        M_L_lbmol=np.array([12.0, 11.0], dtype=float),
        M_V_lbmol=np.array([2.0, 3.0], dtype=float),
        y0=np.array([[0.7, 0.3], [0.45, 0.55]], dtype=float),
        x0=np.array([[0.8, 0.2], [0.35, 0.65]], dtype=float),
        streams={},
    )
    layout = StateVectorLayout(
        n_stages=2,
        n_components=2,
        include_top=False,
        include_bottom=False,
        include_vapor=True,
        include_temperature=True,
        include_energy=False,
    )
    y = layout.pack_y0(col)
    base_inputs = ColumnInputs(
        boundary=rhs_module.BoundaryFlows(
            reflux_lbmolph=0.0,
            distillate_lbmolph=0.0,
            boilup_lbmolph=0.0,
            bottoms_lbmolph=0.0,
        ),
        pressure_model="hydraulic",
        thermo_refresh_dx=thermo_refresh_dx,
        equilibrium_relaxation=equilibrium_relaxation,
    )
    u = layout.unpack(y)
    z_now = np.zeros((col.n_stages, col.n_components), dtype=float)
    for i in range(col.n_stages):
        z_i = np.asarray(u["tray_L"][i, :] + u["tray_V"][i, :], dtype=float)
        z_now[i, :] = z_i / max(float(np.sum(z_i)), 1.0e-300)
    return col, layout, y, base_inputs, z_now


def test_resolve_runtime_thermo_execution_plan_runs_on_cadence_step():
    col, layout, y, base_inputs, z_now = _make_runtime_thermo_execution_plan_fixture()

    do_thermo, reason = _resolve_runtime_thermo_execution_plan(
        step=10,
        thermo_every=5,
        col=col,
        layout=layout,
        y=y,
        include_temperature=True,
        pressure_model="hydraulic",
        base_inputs=base_inputs,
        last_T_tray=np.array(col.T_f, dtype=float),
        last_P_hyd=np.array(col.P_psia, dtype=float),
        last_P_diag=None,
        last_Zfac=np.ones(col.n_stages, dtype=float),
        last_z_overall=z_now.copy(),
        last_K_tray=np.ones((col.n_stages, col.n_components), dtype=float),
    )

    assert do_thermo is True
    assert reason == "cadence"


def test_resolve_runtime_thermo_execution_plan_triggers_on_state_dx_guardrail():
    col, layout, y, base_inputs, z_now = _make_runtime_thermo_execution_plan_fixture(
        thermo_refresh_dx=1.0e-3,
    )
    last_z = z_now.copy()
    last_z[0, :] = np.array([0.75, 0.25], dtype=float)

    do_thermo, reason = _resolve_runtime_thermo_execution_plan(
        step=1,
        thermo_every=5,
        col=col,
        layout=layout,
        y=y,
        include_temperature=True,
        pressure_model="hydraulic",
        base_inputs=base_inputs,
        last_T_tray=np.array(col.T_f, dtype=float),
        last_P_hyd=np.array(col.P_psia, dtype=float),
        last_P_diag=None,
        last_Zfac=np.ones(col.n_stages, dtype=float),
        last_z_overall=last_z,
        last_K_tray=np.ones((col.n_stages, col.n_components), dtype=float),
    )

    assert do_thermo is True
    assert reason == "state_dx"


def test_resolve_runtime_thermo_execution_plan_holds_within_guardrails():
    col, layout, y, base_inputs, z_now = _make_runtime_thermo_execution_plan_fixture(
        thermo_refresh_dx=0.1,
    )

    do_thermo, reason = _resolve_runtime_thermo_execution_plan(
        step=1,
        thermo_every=5,
        col=col,
        layout=layout,
        y=y,
        include_temperature=True,
        pressure_model="hydraulic",
        base_inputs=base_inputs,
        last_T_tray=np.array(col.T_f, dtype=float),
        last_P_hyd=np.array(col.P_psia, dtype=float),
        last_P_diag=None,
        last_Zfac=np.ones(col.n_stages, dtype=float),
        last_z_overall=z_now.copy(),
        last_K_tray=np.ones((col.n_stages, col.n_components), dtype=float),
    )

    assert do_thermo is False
    assert reason == "hold"


def test_resolve_runtime_thermo_execution_plan_triggers_when_equilibrium_K_missing():
    col, layout, y, base_inputs, z_now = _make_runtime_thermo_execution_plan_fixture(
        equilibrium_relaxation=True,
    )

    do_thermo, reason = _resolve_runtime_thermo_execution_plan(
        step=1,
        thermo_every=5,
        col=col,
        layout=layout,
        y=y,
        include_temperature=True,
        pressure_model="hydraulic",
        base_inputs=base_inputs,
        last_T_tray=np.array(col.T_f, dtype=float),
        last_P_hyd=np.array(col.P_psia, dtype=float),
        last_P_diag=None,
        last_Zfac=np.ones(col.n_stages, dtype=float),
        last_z_overall=z_now.copy(),
        last_K_tray=None,
    )

    assert do_thermo is True
    assert reason == "missing_K"


def test_startup_seed_cache_signature_mismatch_is_rejected(tmp_path: Path):
    col = ColumnSpec(
        excel_path=str(tmp_path / "case.xlsx"),
        components_excel=["A", "B"],
        components_dwsim=["A", "B"],
        n_components=2,
        n_stages=2,
        stage_1based=np.array([1, 2], dtype=int),
        sim=SimulationSettings(dt_sec=1.0, t_final_sec=2.0, log_every_n_steps=1),
        duties=HeatDuties(condenser_type="Total", q_cond_btu_per_h=0.0, q_reb_btu_per_h=0.0),
        specs_raw={},
        T_f=np.array([120.0, 130.0], dtype=float),
        P_psia=np.array([220.0, 221.0], dtype=float),
        V_lbmolph=np.array([1000.0, 2000.0], dtype=float),
        L_lbmolph=np.array([4000.0, 5000.0], dtype=float),
        M_L_lbmol=np.array([10.0, 11.0], dtype=float),
        M_V_lbmol=np.array([1.0, 2.0], dtype=float),
        y0=np.array([[0.9, 0.1], [0.3, 0.7]], dtype=float),
        x0=np.array([[0.8, 0.2], [0.4, 0.6]], dtype=float),
        streams={},
    )
    layout = StateVectorLayout(
        n_stages=2,
        n_components=2,
        include_top=False,
        include_bottom=False,
        include_vapor=True,
        include_temperature=True,
        include_energy=False,
    )
    cfg_write = RunnerConfig(
        excel_path=str(tmp_path / "case.xlsx"),
        runtime_mode="hydraulic",
        thermo_mode="clapeyron",
        clapeyron_model="PR",
        enable_startup_seed_cache=True,
        logs_dir=str(tmp_path / "logs"),
    )
    cfg_read = RunnerConfig(
        excel_path=str(tmp_path / "case.xlsx"),
        runtime_mode="hydraulic",
        thermo_mode="clapeyron",
        clapeyron_model="SRK",
        enable_startup_seed_cache=True,
        logs_dir=str(tmp_path / "logs"),
    )
    base_inputs = ColumnInputs(
        pressure_model="hydraulic",
        vapor_flow_model="energy",
        condenser_duty_mode="total-condense",
    )
    y = layout.pack_y0(col)
    path = runmod._resolve_startup_seed_cache_path(cfg_write)
    assert path is not None
    save_info = runmod._write_startup_seed_cache(
        path=path,
        cfg=cfg_write,
        col=col,
        layout=layout,
        base_inputs=base_inputs,
        y=y,
        last_T_tray=None,
        last_P_diag=None,
        last_P_hyd=None,
        last_K_tray=None,
        last_HL=None,
        last_HV=None,
        last_Zfac=None,
        last_z_overall=None,
        last_tray_bubble_target_F=None,
        last_tray_thermo_packet=None,
        last_condenser_duty_packet=None,
        last_feed_stage_flash_packet=None,
        last_bottom_sump_cp_packet=None,
        last_reb_T=None,
        last_reb_x=None,
        last_reb_y=None,
        last_reb_beta=None,
        startup_seeded_condenser_duty_packet=False,
    )
    assert bool(save_info["saved"]) is True

    loaded, load_info = runmod._load_startup_seed_cache(
        path=path,
        cfg=cfg_read,
        col=col,
        layout=layout,
        base_inputs=base_inputs,
    )

    assert loaded is None
    assert "signature_mismatch" in str(load_info["reason"])


def test_runner_writes_and_reuses_startup_seed_cache(tmp_path: Path):
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    logs_dir = tmp_path / "seed_logs"
    cfg = RunnerConfig(
        excel_path=str(excel),
        n_steps=0,
        dt_sec=0.1,
        log_every_n_steps=1,
        include_temperature=True,
        include_energy=False,
        thermo_mode="stub",
        logs_dir=str(logs_dir),
        write_logs=True,
        enable_startup_seed_cache=True,
    )

    out1 = run_smoke_simulation(cfg)
    seed_info_1 = dict(out1.get("startup_seed_cache_info") or {})
    seed_path = Path(str(seed_info_1.get("path") or ""))
    assert seed_path.exists()
    assert bool(seed_info_1.get("saved", False)) is True
    assert bool(seed_info_1.get("loaded", False)) is False

    out2 = run_smoke_simulation(cfg)
    seed_info_2 = dict(out2.get("startup_seed_cache_info") or {})
    assert bool(seed_info_2.get("loaded", False)) is True
    metadata_path = Path(str(out2["run_metadata_json"]))
    doc = json.loads(metadata_path.read_text(encoding="utf-8"))
    assert bool(doc["startup_seed_cache"]["loaded"]) is True


def test_bottom_true_level_control_logs_fractional_pv_and_sp(tmp_path: Path):
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    cfg = RunnerConfig(
        excel_path=str(excel),
        n_steps=1,
        dt_sec=0.1,
        log_every_n_steps=1,
        include_temperature=True,
        include_energy=False,
        enable_equilibrium_relaxation=True,
        thermo_mode="stub",
        logs_dir=str(tmp_path),
        write_logs=True,
        enable_level_control=True,
        top_level_sp_lbmol=397.0,
        bottom_level_pv_mode="true-level",
        bottom_level_sp_frac=0.5,
        bottom_sump_total_volume_ft3=2500.0,
    )

    out = run_smoke_simulation(cfg)
    summary_csv = Path(str(out["summary_csv"]))
    assert summary_csv.exists()

    import csv

    with summary_csv.open("r", encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows, "summary log is empty"

    r0 = rows[0]
    assert "Bottom_level_ctrl_pv" in r0
    assert "Bottom_level_ctrl_sp" in r0
    sp = float(r0["Bottom_level_ctrl_sp"])
    pv = float(r0["Bottom_level_ctrl_pv"])
    assert np.isfinite(sp)
    assert np.isfinite(pv)
    assert sp == pytest.approx(0.5)
    assert 0.0 <= pv <= 1.0


def test_top_true_level_control_preserves_fractional_pv_when_runtime_estimate_fails(tmp_path: Path, monkeypatch):
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    import dynamic_distillation.dynamic_run_scaffold_v1 as scaffold

    call_counter = {"n": 0}

    def _fake_top_level_estimator(*args, **kwargs):
        call_counter["n"] += 1
        # First two calls cover controller setup + initial valid cached PV.
        if call_counter["n"] <= 2:
            return 2000.0, 0.65, 0.52
        # Later runtime estimate fails; controller should keep last valid true-level PV.
        return None, None, None

    monkeypatch.setattr(scaffold, "_estimate_top_drum_liquid_volume_ft3", _fake_top_level_estimator)

    cfg = RunnerConfig(
        excel_path=str(excel),
        n_steps=1,
        dt_sec=0.1,
        log_every_n_steps=1,
        include_temperature=True,
        include_energy=False,
        enable_equilibrium_relaxation=True,
        thermo_mode="stub",
        logs_dir=str(tmp_path),
        write_logs=True,
        enable_level_control=True,
        top_level_pv_mode="true-level",
        top_level_sp_frac=0.5,
        top_drum_total_volume_ft3=4000.0,
        bottom_level_sp_lbmol=794.0,
    )

    out = run_smoke_simulation(cfg)
    summary_csv = Path(str(out["summary_csv"]))
    assert summary_csv.exists()

    import csv

    with summary_csv.open("r", encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert len(rows) >= 2, "expected startup and first runtime rows"

    pv0 = float(rows[0]["Top_level_ctrl_pv"])
    pv1 = float(rows[1]["Top_level_ctrl_pv"])
    assert np.isfinite(pv0)
    assert np.isfinite(pv1)
    assert 0.0 <= pv0 <= 1.0
    assert 0.0 <= pv1 <= 1.0
    assert pv1 == pytest.approx(0.52)


def test_summary_row_prefers_logged_pressure_controller_pv():
    import dynamic_distillation.dynamic_run_scaffold_v1 as scaffold

    class TinyCol:
        n_stages = 2
        n_components = 2
        components_excel = ["C3", "C4"]
        T_f = np.array([120.0, 130.0], dtype=float)
        P_psia = np.array([220.0, 230.0], dtype=float)
        M_L_lbmol = np.array([5.0, 6.0], dtype=float)
        M_V_lbmol = np.array([1.0, 1.0], dtype=float)
        x0 = np.array([[0.80, 0.20], [0.30, 0.70]], dtype=float)
        y0 = np.array([[0.75, 0.25], [0.25, 0.75]], dtype=float)
        streams = {}

    col = TinyCol()
    layout = StateVectorLayout(
        n_stages=col.n_stages,
        n_components=col.n_components,
        include_top=False,
        include_bottom=False,
        include_vapor=True,
        include_temperature=False,
    )
    y = layout.pack_y0(col)

    diag = {
        "x_tray": np.asarray(col.x0, dtype=float).copy(),
        "y_tray": np.asarray(col.y0, dtype=float).copy(),
        "P_psia_diag": np.array([222.0, 231.0], dtype=float),
        "P_psia_hyd": np.array([224.0, 232.0], dtype=float),
        "P_top_drum_psia": np.array([226.0], dtype=float),
        "P_top_ctrl_pv_psia": np.array([221.0], dtype=float),
    }

    row = scaffold._summary_row(
        t_s=0.0,
        case=None,
        col=col,
        layout=layout,
        y=y,
        diag=diag,
        include_temperature=False,
        volume_model=scaffold.VolumeModel(default_vapor_volume_ft3=10.0),
        wall_clock_iso="2026-02-17T00:00:00",
        wall_elapsed_s=0.0,
        feed_tag=scaffold.StreamTag(name="Feed", flow_lbmolph=1000.0, stage_1based=2),
        dist_tag=scaffold.StreamTag(name="Distillate", flow_lbmolph=200.0, stage_1based=1),
        bots_tag=scaffold.StreamTag(name="Bottoms", flow_lbmolph=800.0, stage_1based=2),
    )

    assert float(row["P_top_ctrl_pv_psia"]) == pytest.approx(221.0)
    assert float(row["P_top_psia"]) == pytest.approx(221.0)
    assert float(row["P_bot_psia"]) == pytest.approx(232.0)
    assert float(row["P_top_psia"]) != pytest.approx(226.0)


def test_summary_row_prefers_logged_hydraulic_bottom_pressure():
    import dynamic_distillation.dynamic_run_scaffold_v1 as scaffold

    class TinyCol:
        n_stages = 2
        n_components = 2
        components_excel = ["C3", "C4"]
        T_f = np.array([120.0, 130.0], dtype=float)
        P_psia = np.array([220.0, 230.0], dtype=float)
        M_L_lbmol = np.array([5.0, 6.0], dtype=float)
        M_V_lbmol = np.array([1.0, 1.0], dtype=float)
        x0 = np.array([[0.80, 0.20], [0.30, 0.70]], dtype=float)
        y0 = np.array([[0.75, 0.25], [0.25, 0.75]], dtype=float)
        streams = {}

    col = TinyCol()
    layout = StateVectorLayout(
        n_stages=col.n_stages,
        n_components=col.n_components,
        include_top=False,
        include_bottom=False,
        include_vapor=True,
        include_temperature=False,
    )
    y = layout.pack_y0(col)

    diag = {
        "x_tray": np.asarray(col.x0, dtype=float).copy(),
        "y_tray": np.asarray(col.y0, dtype=float).copy(),
        "P_psia_diag": np.array([222.0, 231.0], dtype=float),
        "P_psia_hyd": np.array([224.0, 242.0], dtype=float),
    }

    row = scaffold._summary_row(
        t_s=0.0,
        case=None,
        col=col,
        layout=layout,
        y=y,
        diag=diag,
        include_temperature=False,
        volume_model=scaffold.VolumeModel(default_vapor_volume_ft3=10.0),
        wall_clock_iso="2026-02-17T00:00:00",
        wall_elapsed_s=0.0,
        feed_tag=scaffold.StreamTag(name="Feed", flow_lbmolph=1000.0, stage_1based=2),
        dist_tag=scaffold.StreamTag(name="Distillate", flow_lbmolph=200.0, stage_1based=1),
        bots_tag=scaffold.StreamTag(name="Bottoms", flow_lbmolph=800.0, stage_1based=2),
    )

    assert float(row["P_bot_psia"]) == pytest.approx(242.0)
    assert float(row["P_bot_psia_spec"]) == pytest.approx(230.0)


def test_summary_row_bottoms_product_composition_remains_sump_and_reports_stage_mismatch():
    import dynamic_distillation.dynamic_run_scaffold_v1 as scaffold

    class TinyCol:
        n_stages = 2
        n_components = 2
        components_excel = ["C3", "C4"]
        T_f = np.array([120.0, 130.0], dtype=float)
        P_psia = np.array([220.0, 230.0], dtype=float)
        M_L_lbmol = np.array([5.0, 6.0], dtype=float)
        M_V_lbmol = np.array([1.0, 1.0], dtype=float)
        x0 = np.array([[0.80, 0.20], [0.30, 0.70]], dtype=float)
        y0 = np.array([[0.75, 0.25], [0.25, 0.75]], dtype=float)
        streams = {}

    col = TinyCol()
    layout = StateVectorLayout(
        n_stages=col.n_stages,
        n_components=col.n_components,
        include_top=False,
        include_bottom=True,
        include_vapor=True,
        include_temperature=True,
    )
    y = layout.pack_y0(col)
    sl = layout.slices()
    y[sl["bottom_L"]] = np.array([100.0, 0.0], dtype=float)
    y[sl["bottom_T_f"]] = np.array([165.0], dtype=float)
    diag = {
        "x_tray": np.asarray(col.x0, dtype=float).copy(),
        "y_tray": np.asarray(col.y0, dtype=float).copy(),
        "P_psia_diag": np.array([222.0, 231.0], dtype=float),
        "P_psia_hyd": np.array([224.0, 232.0], dtype=float),
    }

    row = scaffold._summary_row(
        t_s=0.0,
        case=None,
        col=col,
        layout=layout,
        y=y,
        diag=diag,
        include_temperature=True,
        volume_model=scaffold.VolumeModel(default_vapor_volume_ft3=10.0),
        wall_clock_iso="2026-04-09T00:00:00",
        wall_elapsed_s=0.0,
        feed_tag=scaffold.StreamTag(name="Feed", flow_lbmolph=1000.0, stage_1based=2),
        dist_tag=scaffold.StreamTag(name="Distillate", flow_lbmolph=200.0, stage_1based=1),
        bots_tag=scaffold.StreamTag(name="Bottoms", flow_lbmolph=800.0, stage_1based=2),
    )

    assert float(row["x_Bottoms_C3"]) == pytest.approx(0.30)
    assert float(row["Bottoms_x_C3"]) == pytest.approx(1.0)
    assert float(row["Bottoms_x_C4"]) == pytest.approx(0.0)
    assert float(row["Bottoms_sump_x_C3"]) == pytest.approx(1.0)
    assert float(row["Bottoms_sump_x_C4"]) == pytest.approx(0.0)
    assert str(row["Bottoms_x_source"]) == "sump"
    assert float(row["Bottoms_stage_sump_tv_distance"]) == pytest.approx(0.70)


def test_desired_inventory_recovery_rate_uses_volume_fraction_in_true_level_mode():
    import dynamic_distillation.dynamic_run_scaffold_v1 as scaffold

    rate = scaffold._desired_inventory_recovery_rate_lbmolph(
        total_lbmol=472.0,
        pv=0.23858148859461004,
        sp=0.5,
        pv_mode="true-level",
        lbmol_per_volume_fraction_scale=2679.0,
        recover_tau_sec=120.0,
    )

    expected = 2679.0 * (
        scaffold._horizontal_cylinder_volume_fraction_from_height_fraction(0.5)
        - scaffold._horizontal_cylinder_volume_fraction_from_height_fraction(0.23858148859461004)
    ) * 3600.0 / 120.0
    assert float(rate) == pytest.approx(expected)
    assert float(rate) > 0.0


def test_desired_inventory_recovery_rate_uses_lbmol_delta_in_molar_holdup_mode():
    import dynamic_distillation.dynamic_run_scaffold_v1 as scaffold

    rate = scaffold._desired_inventory_recovery_rate_lbmolph(
        total_lbmol=450.0,
        pv=0.3,
        sp=500.0,
        pv_mode="molar-holdup",
        lbmol_per_volume_fraction_scale=9999.0,
        recover_tau_sec=100.0,
    )

    assert float(rate) == pytest.approx((500.0 - 450.0) * 3600.0 / 100.0)


def test_allow_coupled_total_condenser_partial_condense_defaults_on():
    import dynamic_distillation.dynamic_run_scaffold_v1 as scaffold

    cfg = scaffold.RunnerConfig(excel_path="case.xlsx")

    assert (
        scaffold._allow_coupled_total_condenser_partial_condense(
            cfg=cfg,
            pressure_control_mv="condenser-duty",
            condenser_duty_mode="total-condense",
        )
        is True
    )


def test_allow_coupled_total_condenser_partial_condense_can_be_disabled():
    import dynamic_distillation.dynamic_run_scaffold_v1 as scaffold

    cfg = scaffold.RunnerConfig(
        excel_path="case.xlsx",
        enable_coupled_total_condenser_partial_condense=False,
    )

    assert (
        scaffold._allow_coupled_total_condenser_partial_condense(
            cfg=cfg,
            pressure_control_mv="condenser-duty",
            condenser_duty_mode="total-condense",
        )
        is False
    )
    assert (
        scaffold._allow_coupled_total_condenser_partial_condense(
            cfg=cfg,
            pressure_control_mv="top-anchor",
            condenser_duty_mode="total-condense",
        )
        is False
    )


def test_summary_row_includes_integrator_diagnostics():
    import dynamic_distillation.dynamic_run_scaffold_v1 as scaffold

    class TinyCol:
        n_stages = 2
        n_components = 2
        components_excel = ["C3", "C4"]
        T_f = np.array([120.0, 130.0], dtype=float)
        P_psia = np.array([220.0, 230.0], dtype=float)
        M_L_lbmol = np.array([5.0, 6.0], dtype=float)
        M_V_lbmol = np.array([1.0, 1.0], dtype=float)
        x0 = np.array([[0.80, 0.20], [0.30, 0.70]], dtype=float)
        y0 = np.array([[0.75, 0.25], [0.25, 0.75]], dtype=float)
        streams = {}

    col = TinyCol()
    layout = StateVectorLayout(
        n_stages=col.n_stages,
        n_components=col.n_components,
        include_top=False,
        include_bottom=False,
        include_vapor=True,
        include_temperature=False,
    )
    y = layout.pack_y0(col)
    diag = {
        "x_tray": np.asarray(col.x0, dtype=float).copy(),
        "y_tray": np.asarray(col.y0, dtype=float).copy(),
        "P_psia_diag": np.array([222.0, 231.0], dtype=float),
    }
    integ = {
        "requested_mode": "ida",
        "used_mode": "ida",
        "fallback_used": False,
        "fallback_reason": "",
        "nfev": 9.0,
        "ida_iter_max": 3.0,
        "ida_iter_mean": 2.0,
        "ida_converged": 1.0,
        "ida_last_err": 0.2,
        "ida_alg_p_inf_psia": 0.01,
        "ida_alg_v_inf_lbmolph": 3.0,
        "ida_alg_weighted": 0.6,
        "ida_alg_converged": 1.0,
        "ida_resid_energy_btups": 125.0,
    }

    row = scaffold._summary_row(
        t_s=0.0,
        case=None,
        col=col,
        layout=layout,
        y=y,
        diag=diag,
        include_temperature=False,
        volume_model=scaffold.VolumeModel(default_vapor_volume_ft3=10.0),
        wall_clock_iso="2026-02-17T00:00:00",
        wall_elapsed_s=0.0,
        feed_tag=scaffold.StreamTag(name="Feed", flow_lbmolph=1000.0, stage_1based=2),
        dist_tag=scaffold.StreamTag(name="Distillate", flow_lbmolph=200.0, stage_1based=1),
        bots_tag=scaffold.StreamTag(name="Bottoms", flow_lbmolph=800.0, stage_1based=2),
        integrator_info=integ,
    )

    assert str(row["integrator_requested_mode"]) == "ida"
    assert str(row["integrator_used_mode"]) == "ida"
    assert float(row["integrator_fallback_used"]) == pytest.approx(0.0)
    assert float(row["integrator_nfev"]) == pytest.approx(9.0)
    assert float(row["ida_iter_max"]) == pytest.approx(3.0)
    assert float(row["ida_alg_weighted"]) == pytest.approx(0.6)


def test_summary_row_includes_reboiler_duty_diagnostics():
    import dynamic_distillation.dynamic_run_scaffold_v1 as scaffold

    class TinyCol:
        n_stages = 2
        n_components = 2
        components_excel = ["C3", "C4"]
        T_f = np.array([120.0, 130.0], dtype=float)
        P_psia = np.array([220.0, 230.0], dtype=float)
        M_L_lbmol = np.array([5.0, 6.0], dtype=float)
        M_V_lbmol = np.array([1.0, 1.0], dtype=float)
        x0 = np.array([[0.80, 0.20], [0.30, 0.70]], dtype=float)
        y0 = np.array([[0.75, 0.25], [0.25, 0.75]], dtype=float)
        streams = {}

    col = TinyCol()
    layout = StateVectorLayout(
        n_stages=col.n_stages,
        n_components=col.n_components,
        include_top=False,
        include_bottom=False,
        include_vapor=True,
        include_temperature=False,
    )
    y = layout.pack_y0(col)
    diag = {
        "x_tray": np.asarray(col.x0, dtype=float).copy(),
        "y_tray": np.asarray(col.y0, dtype=float).copy(),
        "P_psia_diag": np.array([222.0, 231.0], dtype=float),
        "boilup_from_duty_lbmolph": np.array([8123.4], dtype=float),
        "boilup_realized_lbmolph": np.array([8044.4], dtype=float),
        "reboiler_latent_heat_BTU_per_lbmol": np.array([7100.0], dtype=float),
        "reboiler_temperature_F": np.array([221.5], dtype=float),
        "reboiler_mode_duty_active": np.array([1.0], dtype=float),
        "reboiler_neighbor_stage_1based": np.array([1.0], dtype=float),
        "reboiler_neighbor_vflow_calc_lbmolph": np.array([7900.0], dtype=float),
        "reboiler_neighbor_vflow_used_lbmolph": np.array([7800.0], dtype=float),
        "reboiler_neighbor_vflow_limit_hi_lbmolph": np.array([8050.0], dtype=float),
        "reboiler_neighbor_vflow_limit_lo_lbmolph": np.array([7950.0], dtype=float),
        "reboiler_neighbor_vflow_clamped_flag": np.array([1.0], dtype=float),
    }

    row = scaffold._summary_row(
        t_s=0.0,
        case=None,
        col=col,
        layout=layout,
        y=y,
        diag=diag,
        include_temperature=False,
        volume_model=scaffold.VolumeModel(default_vapor_volume_ft3=10.0),
        wall_clock_iso="2026-03-31T00:00:00",
        wall_elapsed_s=0.0,
        feed_tag=scaffold.StreamTag(name="Feed", flow_lbmolph=1000.0, stage_1based=2),
        dist_tag=scaffold.StreamTag(name="Distillate", flow_lbmolph=200.0, stage_1based=1),
        bots_tag=scaffold.StreamTag(name="Bottoms", flow_lbmolph=800.0, stage_1based=2),
    )

    assert float(row["boilup_from_duty_lbmolph"]) == pytest.approx(8123.4)
    assert float(row["boilup_realized_lbmolph"]) == pytest.approx(8044.4)
    assert float(row["reboiler_latent_heat_BTU_per_lbmol"]) == pytest.approx(7100.0)
    assert float(row["reboiler_neighbor_vflow_used_lbmolph"]) == pytest.approx(7800.0)
    assert float(row["reboiler_neighbor_vflow_clamped_flag"]) == pytest.approx(1.0)


def test_summary_row_includes_steady_state_diagnostics():
    import dynamic_distillation.dynamic_run_scaffold_v1 as scaffold

    class TinyCol:
        n_stages = 2
        n_components = 2
        components_excel = ["C3", "C4"]
        T_f = np.array([120.0, 130.0], dtype=float)
        P_psia = np.array([220.0, 230.0], dtype=float)
        M_L_lbmol = np.array([5.0, 6.0], dtype=float)
        M_V_lbmol = np.array([1.0, 1.0], dtype=float)
        x0 = np.array([[0.80, 0.20], [0.30, 0.70]], dtype=float)
        y0 = np.array([[0.75, 0.25], [0.25, 0.75]], dtype=float)
        streams = {}

    col = TinyCol()
    layout = StateVectorLayout(
        n_stages=col.n_stages,
        n_components=col.n_components,
        include_top=False,
        include_bottom=False,
        include_vapor=True,
        include_temperature=False,
    )
    y = layout.pack_y0(col)
    diag = {
        "x_tray": np.asarray(col.x0, dtype=float).copy(),
        "y_tray": np.asarray(col.y0, dtype=float).copy(),
        "P_psia_diag": np.array([222.0, 231.0], dtype=float),
        "steady_state_enabled": np.array([1.0], dtype=float),
        "steady_state_flag": np.array([1.0], dtype=float),
        "steady_state_score": np.array([0.42], dtype=float),
        "steady_state_active_criteria": np.array([4.0], dtype=float),
        "ss_max_rel_state_rate_per_s": np.array([2.5e-3], dtype=float),
        "ss_max_kpi_slope_per_s": np.array([8.0e-5], dtype=float),
        "ss_max_mv_rate_per_s": np.array([9.0], dtype=float),
        "ss_max_temp_rate_F_per_s": np.array([0.08], dtype=float),
        "ss_max_sp_error": np.array([0.015], dtype=float),
        "ss_window_samples": np.array([31.0], dtype=float),
        "ss_window_sec": np.array([30.0], dtype=float),
        "ss_min_time_sec": np.array([60.0], dtype=float),
    }

    row = scaffold._summary_row(
        t_s=120.0,
        case=None,
        col=col,
        layout=layout,
        y=y,
        diag=diag,
        include_temperature=False,
        volume_model=scaffold.VolumeModel(default_vapor_volume_ft3=10.0),
        wall_clock_iso="2026-02-17T00:00:00",
        wall_elapsed_s=0.0,
        feed_tag=scaffold.StreamTag(name="Feed", flow_lbmolph=1000.0, stage_1based=2),
        dist_tag=scaffold.StreamTag(name="Distillate", flow_lbmolph=200.0, stage_1based=1),
        bots_tag=scaffold.StreamTag(name="Bottoms", flow_lbmolph=800.0, stage_1based=2),
    )

    assert float(row["steady_state_enabled"]) == pytest.approx(1.0)
    assert float(row["steady_state_flag"]) == pytest.approx(1.0)
    assert float(row["steady_state_score"]) == pytest.approx(0.42)
    assert float(row["ss_max_rel_state_rate_per_s"]) == pytest.approx(2.5e-3)
    assert float(row["ss_max_kpi_slope_per_s"]) == pytest.approx(8.0e-5)
    assert float(row["ss_max_mv_rate_per_s"]) == pytest.approx(9.0)
    assert float(row["ss_max_temp_rate_F_per_s"]) == pytest.approx(0.08)
    assert float(row["ss_max_sp_error"]) == pytest.approx(0.015)


def test_profile_rows_add_unit_rows_and_move_drum_sump_fields():
    import dynamic_distillation.dynamic_run_scaffold_v1 as scaffold

    class TinyCol:
        n_stages = 2
        n_components = 2
        components_excel = ["C3", "C4"]
        T_f = np.array([120.0, 130.0], dtype=float)
        P_psia = np.array([220.0, 230.0], dtype=float)
        M_L_lbmol = np.array([5.0, 6.0], dtype=float)
        M_V_lbmol = np.array([1.0, 1.0], dtype=float)
        x0 = np.array([[0.80, 0.20], [0.30, 0.70]], dtype=float)
        y0 = np.array([[0.75, 0.25], [0.25, 0.75]], dtype=float)
        streams = {}

    col = TinyCol()
    layout = StateVectorLayout(
        n_stages=col.n_stages,
        n_components=col.n_components,
        include_top=True,
        include_bottom=True,
        include_vapor=True,
        include_temperature=True,
    )
    y = layout.pack_y0(col)
    sl = layout.slices()
    y[sl["bottom_T_f"]] = np.array([165.0], dtype=float)

    diag = {
        "x_tray": np.asarray(col.x0, dtype=float).copy(),
        "y_tray": np.asarray(col.y0, dtype=float).copy(),
        "P_psia_hyd": np.array([220.0, 222.0], dtype=float),
        "L_out_hyd_lbmolph": np.array([100.0, 120.0], dtype=float),
        "V_out_lbmolph": np.array([1000.0, 1100.0], dtype=float),
        "Q_cond_cmd_BTUph": np.array([-5.1e7], dtype=float),
        "Q_cond_calc_BTUph": np.array([-5.0e7], dtype=float),
        "Q_cond_used_BTUph": np.array([-5.0e7], dtype=float),
        "P_top_drum_psia": np.array([219.0], dtype=float),
        "V_top_drum_vapor_ft3": np.array([500.0], dtype=float),
        "Q_reb_cmd_BTUph": np.array([8.2e7], dtype=float),
        "Q_reb_used_BTUph": np.array([8.1e7], dtype=float),
        "x_eq_tray": np.array([[0.78, 0.22], [0.28, 0.72]], dtype=float),
        "y_eq_tray": np.array([[0.74, 0.26], [0.24, 0.76]], dtype=float),
        "y_target_tray": np.array([[0.745, 0.255], [0.245, 0.755]], dtype=float),
        "xD_comp_sp": np.array([0.10], dtype=float),
        "xD_comp_pv": np.array([0.12], dtype=float),
        "Reflux_cmd_lbmolph": np.array([6000.0], dtype=float),
        "eq_flash_mv_total_lbmol_tray": np.array([1.2, 1.6], dtype=float),
        "eq_target_mv_total_lbmol_tray": np.array([1.3, 1.7], dtype=float),
        "eq_target_vapor_total_lbmol_tray": np.array([1.4, 1.8], dtype=float),
        "eq_target_vapor_delta_lbmol_tray": np.array([0.4, 0.8], dtype=float),
        "eq_target_vapor_fraction_tray": np.array([0.20, 0.257142857], dtype=float),
        "eq_current_vapor_fraction_tray": np.array([1.0 / 6.0, 1.0 / 7.0], dtype=float),
        "eq_phase_change_lbmolps_tray": np.array([0.20, 0.40], dtype=float),
        "eq_phase_rate_guard_scale_tray": np.array([1.0, 0.5], dtype=float),
        "eq_phase_rate_guard_limit_lbmolps_tray": np.array([np.nan, 0.4], dtype=float),
        "xB_comp_sp": np.array([0.30], dtype=float),
        "xB_comp_pv": np.array([0.28], dtype=float),
        "Boilup_cmd_lbmolph": np.array([12000.0], dtype=float),
        "dT_energy_raw_F_per_s_tray": np.array([0.1, 25.0], dtype=float),
        "tray_heat_capacity_BTU_per_F_tray": np.array([100.0, 5.0], dtype=float),
        "tray_effective_heat_capacity_BTU_per_F_tray": np.array([100.0, 25.0], dtype=float),
        "tray_temperature_guard_active_tray": np.array([0.0, 1.0], dtype=float),
        "tray_temperature_rate_limit_F_per_s_tray": np.array([np.nan, 10.0], dtype=float),
    }

    rows = scaffold._profile_rows(
        t_s=0.0,
        case=None,
        col=col,
        layout=layout,
        y=y,
        diag=diag,
        include_temperature=True,
        volume_model=scaffold.VolumeModel(default_vapor_volume_ft3=10.0),
        wall_clock_iso="2026-02-18T00:00:00",
        wall_elapsed_s=0.0,
        feed_tag=scaffold.StreamTag(name="Feed", flow_lbmolph=1000.0, stage_1based=2),
        dist_tag=scaffold.StreamTag(name="Distillate", flow_lbmolph=200.0, stage_1based=1),
        bots_tag=scaffold.StreamTag(name="Bottoms", flow_lbmolph=800.0, stage_1based=2),
    )

    assert len(rows) == 4
    assert str(rows[0]["node_type"]) == "distillate_drum"
    assert str(rows[-1]["node_type"]) == "bottoms_sump"
    stage_rows = [r for r in rows if str(r["node_type"]) == "stage"]
    drum_rows = [r for r in rows if str(r["node_type"]) == "distillate_drum"]
    sump_rows = [r for r in rows if str(r["node_type"]) == "bottoms_sump"]
    assert len(stage_rows) == 2
    assert len(drum_rows) == 1
    assert len(sump_rows) == 1

    stage1 = next(r for r in stage_rows if int(r["stage"]) == 1)
    stage2 = next(r for r in stage_rows if int(r["stage"]) == 2)
    drum = drum_rows[0]
    sump = sump_rows[0]

    assert int(drum["stage"]) == 0
    assert int(sump["stage"]) == 3
    assert float(stage1["P_psia_hyd"]) == pytest.approx(220.0)
    assert float(stage2["P_psia_hyd"]) == pytest.approx(222.0)

    assert np.isnan(float(stage1["P_top_drum_psia"]))
    assert float(drum["P_top_drum_psia"]) == pytest.approx(219.0)
    assert np.isnan(float(stage1["Distillate_L_lbmol"]))
    assert np.isfinite(float(drum["Distillate_L_lbmol"]))
    assert np.isnan(float(stage1["Distillate_x_C3"]))
    assert np.isfinite(float(drum["Distillate_x_C3"]))
    assert np.isnan(float(stage1["D_lbmolph"]))
    assert float(drum["D_lbmolph"]) == pytest.approx(200.0)

    assert np.isnan(float(stage2["Q_reb_cmd_BTUph"]))
    assert float(sump["Q_reb_cmd_BTUph"]) == pytest.approx(8.2e7)
    assert np.isnan(float(stage2["Bottoms_L_lbmol"]))
    assert np.isfinite(float(sump["Bottoms_L_lbmol"]))
    assert np.isnan(float(stage2["Bottoms_x_C4"]))
    assert np.isfinite(float(sump["Bottoms_x_C4"]))
    assert np.isnan(float(stage2["B_lbmolph"]))
    assert float(sump["B_lbmolph"]) == pytest.approx(800.0)
    assert float(stage1["eq_target_vapor_total_lbmol_tray"]) == pytest.approx(1.4)
    assert float(stage2["eq_target_vapor_delta_lbmol_tray"]) == pytest.approx(0.8)
    assert float(stage1["eq_current_vapor_fraction_tray"]) == pytest.approx(1.0 / 6.0)
    assert float(stage2["eq_phase_change_lbmolps_tray"]) == pytest.approx(0.40)
    assert float(stage2["eq_phase_rate_guard_scale_tray"]) == pytest.approx(0.5)
    assert float(stage2["eq_phase_rate_guard_limit_lbmolps_tray"]) == pytest.approx(0.4)
    assert float(stage1["eq_flash_mv_total_lbmol_tray"]) == pytest.approx(1.2)
    assert float(stage2["eq_target_mv_total_lbmol_tray"]) == pytest.approx(1.7)
    assert float(stage1["x_eq_C3"]) == pytest.approx(0.78)
    assert float(stage2["y_eq_C4"]) == pytest.approx(0.76)
    assert float(stage1["y_target_C4"]) == pytest.approx(0.255)
    assert float(stage2["tray_effective_heat_capacity_BTU_per_F"]) == pytest.approx(25.0)
    assert float(stage2["tray_temperature_guard_active_tray"]) == pytest.approx(1.0)
    assert float(stage2["tray_temperature_rate_limit_F_per_s_tray"]) == pytest.approx(10.0)


def test_clip_temperature_states_to_provider_bounds():
    layout = StateVectorLayout(
        n_stages=3,
        n_components=2,
        include_top=True,
        include_bottom=True,
        include_vapor=True,
        include_temperature=True,
    )

    class TinyCol:
        n_stages = 3
        n_components = 2
        T_f = np.array([100.0, 110.0, 120.0], dtype=float)
        P_psia = np.array([200.0, 205.0, 210.0], dtype=float)
        V_lbmolph = np.array([10.0, 10.0, 10.0], dtype=float)
        L_lbmolph = np.array([10.0, 10.0, 10.0], dtype=float)
        M_L_lbmol = np.array([5.0, 5.0, 5.0], dtype=float)
        M_V_lbmol = np.array([1.0, 1.0, 1.0], dtype=float)
        x0 = np.array([[0.5, 0.5], [0.5, 0.5], [0.5, 0.5]], dtype=float)
        y0 = np.array([[0.5, 0.5], [0.5, 0.5], [0.5, 0.5]], dtype=float)

    y = layout.pack_y0(TinyCol())
    sl = layout.slices()
    y[sl["tray_T_f"]] = np.array([70.0, 115.0, 170.0], dtype=float)
    y[sl["bottom_T_f"]] = np.array([180.0], dtype=float)

    class Provider:
        T_grid_F = np.array([95.0, 130.0, 150.0], dtype=float)

    y2 = _clip_temperature_states_to_provider_bounds(y, layout, Provider())
    assert np.allclose(y2[sl["tray_T_f"]], np.array([95.0, 115.0, 150.0], dtype=float))
    assert np.allclose(y2[sl["bottom_T_f"]], np.array([150.0], dtype=float))


def test_component_index_by_name_handles_c_number_alias():
    class TinyCol:
        n_components = 3
        components_excel = ["C3H8", "C4H10", "C5H12"]
        components_dwsim = ["Propane", "n-Butane", "n-Pentane"]

    assert _component_index_by_name(TinyCol(), "C4") == 1
    assert _component_index_by_name(TinyCol(), "c4h10") == 1
    assert _component_index_by_name(TinyCol(), "n-Butane") == 1


def test_pi_update_negative_gain_unwinds_low_saturation():
    ctrl = PIController(kc=-1.0, ti_sec=1.0, bias=0.0, out_min=-100.0, out_max=100.0, integ=500.0)

    # sat_lo with reverse-acting gain: negative error should unwind integrator.
    u = _pi_update(ctrl, pv=0.0, sp=10.0, dt_sec=1.0)
    assert u == pytest.approx(-100.0)
    assert ctrl.integ == pytest.approx(490.0)


def test_pi_update_negative_gain_blocks_windup_low_saturation():
    ctrl = PIController(kc=-1.0, ti_sec=1.0, bias=0.0, out_min=-100.0, out_max=100.0, integ=500.0)

    # sat_lo with reverse-acting gain: positive error pushes deeper into saturation.
    u = _pi_update(ctrl, pv=20.0, sp=10.0, dt_sec=1.0)
    assert u == pytest.approx(-100.0)
    assert ctrl.integ == pytest.approx(500.0)


def test_pressure_resid_gain_scale_maps_residual_to_bounded_gain():
    assert _pressure_resid_gain_scale(resid_abs_btups=None, resid_ref_btups=1000.0, min_gain=0.2) == pytest.approx(1.0)
    assert _pressure_resid_gain_scale(resid_abs_btups=0.0, resid_ref_btups=1000.0, min_gain=0.2) == pytest.approx(1.0)

    g_mid = _pressure_resid_gain_scale(resid_abs_btups=1000.0, resid_ref_btups=1000.0, min_gain=0.2)
    assert g_mid == pytest.approx(0.5)

    g_hi = _pressure_resid_gain_scale(resid_abs_btups=1.0e9, resid_ref_btups=1000.0, min_gain=0.2)
    assert g_hi == pytest.approx(0.2)


def test_apply_slew_limit_clamps_step_change():
    u = _apply_slew_limit(cmd=10.0, prev_cmd=0.0, rate_limit_per_s=2.0, dt_sec=1.0)
    assert u == pytest.approx(2.0)

    u2 = _apply_slew_limit(cmd=-10.0, prev_cmd=0.0, rate_limit_per_s=2.0, dt_sec=1.0)
    assert u2 == pytest.approx(-2.0)

    u3 = _apply_slew_limit(cmd=1.0, prev_cmd=0.0, rate_limit_per_s=2.0, dt_sec=1.0)
    assert u3 == pytest.approx(1.0)


def test_normalize_integrator_mode_aliases():
    assert _normalize_integrator_mode("explicit") == "explicit-euler"
    assert _normalize_integrator_mode("euler") == "explicit-euler"
    assert _normalize_integrator_mode("bdf") == "bdf"
    assert _normalize_integrator_mode("radau") == "radau"
    assert _normalize_integrator_mode("ida") == "ida"
    assert _normalize_integrator_mode("dae") == "ida"
    assert _normalize_integrator_mode("unknown") == "explicit-euler"


def test_effective_hydraulic_ida_profile_applies_tuned_defaults():
    cfg = RunnerConfig(excel_path="dummy.xlsx")
    eff = _effective_hydraulic_ida_profile(
        cfg,
        runtime_mode="hydraulic",
        integrator_mode="ida",
    )
    assert bool(eff["dae_pilot_enabled"]) is True
    assert int(eff["ida_max_iter"]) == 12
    assert float(eff["dae_pilot_v_tol_lbmolph"]) == pytest.approx(100.0)
    assert "enable_dae_pilot_algebraic_solve=True" in list(eff["defaults_applied"])
    assert "ida_max_iter=12" in list(eff["defaults_applied"])
    assert "dae_pilot_v_tol_lbmolph=100" in list(eff["defaults_applied"])


def test_runner_config_defaults_to_table_pool_with_standard_cache():
    cfg = RunnerConfig(excel_path="dummy.xlsx")
    assert cfg.thermo_mode == "table-pool"
    assert cfg.thermo_table_path == r"cache/thermo_table.json"
    assert cfg.thermo_pool_workers == 2
    assert cfg.thermo_pool_chunk_size == 4


def test_effective_hydraulic_ida_profile_preserves_explicit_overrides():
    cfg = RunnerConfig(
        excel_path="dummy.xlsx",
        ida_max_iter=16,
        enable_dae_pilot_algebraic_solve=True,
        dae_pilot_v_tol_lbmolph=60.0,
    )
    eff = _effective_hydraulic_ida_profile(
        cfg,
        runtime_mode="hydraulic",
        integrator_mode="ida",
    )
    assert bool(eff["dae_pilot_enabled"]) is True
    assert int(eff["ida_max_iter"]) == 16
    assert float(eff["dae_pilot_v_tol_lbmolph"]) == pytest.approx(60.0)
    assert list(eff["defaults_applied"]) == []


def test_integrate_one_step_explicit_linear_decay():
    layout = StateVectorLayout(
        n_stages=1,
        n_components=1,
        include_top=False,
        include_bottom=False,
        include_vapor=False,
        include_temperature=False,
    )
    y0 = np.array([2.0], dtype=float)

    def _rhs(_t, y):
        return -np.asarray(y, dtype=float), {}

    y1, info = _integrate_one_step(
        t_s=0.0,
        y=y0,
        dt_sec=0.1,
        rhs_eval=_rhs,
        rhs_eval_fallback=None,
        layout=layout,
        thermo_provider=None,
        integrator_mode="explicit-euler",
        rtol=1.0e-3,
        atol=1.0e-6,
        max_step_sec=None,
        substep_sec=None,
        max_rhs_evals_per_step=None,
        step_wall_limit_sec=None,
    )
    assert y1.shape == (1,)
    assert y1[0] == pytest.approx(1.8)
    assert str(info["used_mode"]) == "explicit-euler"
    assert bool(info["fallback_used"]) is False


def test_advance_explicit_euler_step_reuses_precomputed_rhs():
    layout = StateVectorLayout(
        n_stages=1,
        n_components=1,
        include_top=False,
        include_bottom=False,
        include_vapor=False,
        include_temperature=False,
    )
    y0 = np.array([2.0], dtype=float)
    dydt = np.array([-2.0], dtype=float)

    y1, info = _advance_explicit_euler_step(
        y=y0,
        dydt=dydt,
        dt_sec=0.1,
        layout=layout,
        thermo_provider=None,
        requested_mode="explicit-euler",
    )
    assert y1.shape == (1,)
    assert y1[0] == pytest.approx(1.8)
    assert str(info["used_mode"]) == "explicit-euler"
    assert bool(info["fallback_used"]) is False
    assert bool(info["used_precomputed_rhs"]) is True
    assert float(info["nfev"]) == pytest.approx(0.0)


def test_integrate_one_step_bdf_falls_back_without_scipy(monkeypatch):
    import dynamic_distillation.dynamic_run_scaffold_v1 as scaffold

    layout = StateVectorLayout(
        n_stages=1,
        n_components=1,
        include_top=False,
        include_bottom=False,
        include_vapor=False,
        include_temperature=False,
    )
    y0 = np.array([2.0], dtype=float)

    def _rhs(_t, y):
        return -np.asarray(y, dtype=float), {}

    monkeypatch.setattr(scaffold, "_solve_ivp", None)

    y1, info = _integrate_one_step(
        t_s=0.0,
        y=y0,
        dt_sec=0.1,
        rhs_eval=_rhs,
        rhs_eval_fallback=None,
        layout=layout,
        thermo_provider=None,
        integrator_mode="bdf",
        rtol=1.0e-3,
        atol=1.0e-6,
        max_step_sec=None,
        substep_sec=None,
        max_rhs_evals_per_step=None,
        step_wall_limit_sec=None,
    )
    assert y1[0] == pytest.approx(1.8)
    assert bool(info["fallback_used"]) is True
    assert str(info["used_mode"]) == "explicit-euler"


def test_integrate_one_step_bdf_fallback_can_use_separate_rhs(monkeypatch):
    import dynamic_distillation.dynamic_run_scaffold_v1 as scaffold

    layout = StateVectorLayout(
        n_stages=1,
        n_components=1,
        include_top=False,
        include_bottom=False,
        include_vapor=False,
        include_temperature=False,
    )
    y0 = np.array([2.0], dtype=float)

    def _rhs_stiff(_t, y):
        return -2.0 * np.asarray(y, dtype=float), {}

    def _rhs_fallback(_t, y):
        return -1.0 * np.asarray(y, dtype=float), {}

    monkeypatch.setattr(scaffold, "_solve_ivp", None)

    y1, info = _integrate_one_step(
        t_s=0.0,
        y=y0,
        dt_sec=0.1,
        rhs_eval=_rhs_stiff,
        rhs_eval_fallback=_rhs_fallback,
        layout=layout,
        thermo_provider=None,
        integrator_mode="bdf",
        rtol=1.0e-3,
        atol=1.0e-6,
        max_step_sec=None,
        substep_sec=None,
        max_rhs_evals_per_step=None,
        step_wall_limit_sec=None,
    )
    assert y1[0] == pytest.approx(1.8)
    assert bool(info["fallback_used"]) is True
    assert str(info["used_mode"]) == "explicit-euler"


def test_integrate_one_step_bdf_uses_solve_ivp_when_available(monkeypatch):
    import dynamic_distillation.dynamic_run_scaffold_v1 as scaffold

    layout = StateVectorLayout(
        n_stages=1,
        n_components=1,
        include_top=False,
        include_bottom=False,
        include_vapor=False,
        include_temperature=False,
    )
    y0 = np.array([2.0], dtype=float)

    class _Sol:
        success = True
        y = np.array([[1.234]], dtype=float)
        nfev = 7
        njev = 2
        nlu = 2
        status = 0
        message = "ok"

    def _rhs(_t, y):
        return -np.asarray(y, dtype=float), {}

    def _fake_solve_ivp(fun, t_span, y0, method, t_eval, rtol, atol, max_step, vectorized):
        assert method == "BDF"
        assert np.isfinite(float(rtol))
        assert np.isfinite(float(atol))
        _ = fun(float(t_eval[0]), np.asarray(y0, dtype=float))
        return _Sol()

    monkeypatch.setattr(scaffold, "_solve_ivp", _fake_solve_ivp)

    y1, info = _integrate_one_step(
        t_s=0.0,
        y=y0,
        dt_sec=0.1,
        rhs_eval=_rhs,
        rhs_eval_fallback=None,
        layout=layout,
        thermo_provider=None,
        integrator_mode="bdf",
        rtol=1.0e-3,
        atol=1.0e-6,
        max_step_sec=None,
        substep_sec=None,
        max_rhs_evals_per_step=None,
        step_wall_limit_sec=None,
    )
    assert y1[0] == pytest.approx(1.234)
    assert bool(info["fallback_used"]) is False
    assert str(info["used_mode"]) == "bdf"
    assert float(info["nfev"]) == pytest.approx(7.0)


def test_integrate_one_step_ida_fixed_point_converges():
    layout = StateVectorLayout(
        n_stages=1,
        n_components=1,
        include_top=False,
        include_bottom=False,
        include_vapor=False,
        include_temperature=False,
    )
    y0 = np.array([2.0], dtype=float)

    def _rhs(_t, y):
        # Stable linear system; implicit solve should converge quickly.
        return -np.asarray(y, dtype=float), {}

    y1, info = _integrate_one_step_ida(
        t_s=0.0,
        y=y0,
        dt_sec=0.1,
        rhs_eval=_rhs,
        layout=layout,
        thermo_provider=None,
        substep_sec=None,
        max_iter=8,
        relax=1.0,
        rtol=1.0e-6,
        atol=1.0e-9,
        max_rhs_evals_per_step=200,
        step_wall_limit_sec=5.0,
    )
    assert y1.shape == (1,)
    assert bool(info["fallback_used"]) is False
    assert str(info["used_mode"]) == "ida"
    assert float(info["ida_converged"]) == pytest.approx(1.0)
    # Implicit Euler exact value for y'=-y with dt=0.1 is y1 = y0 / (1+dt).
    assert y1[0] == pytest.approx(2.0 / 1.1, rel=1e-3)


def test_integrate_one_step_ida_waits_for_algebraic_residual():
    layout = StateVectorLayout(
        n_stages=1,
        n_components=1,
        include_top=False,
        include_bottom=False,
        include_vapor=False,
        include_temperature=False,
    )
    y0 = np.array([2.0], dtype=float)

    calls = {"n": 0}
    alg_p_series = [2.0, 2.0, 0.01]

    def _rhs(_t, y):
        idx = min(int(calls["n"]), len(alg_p_series) - 1)
        calls["n"] += 1
        dydt = np.zeros_like(np.asarray(y, dtype=float))
        diag = {"dae_pilot_alg_p_inf_psia": np.array([alg_p_series[idx]], dtype=float)}
        return dydt, diag

    y1, info = _integrate_one_step_ida(
        t_s=0.0,
        y=y0,
        dt_sec=0.1,
        rhs_eval=_rhs,
        layout=layout,
        thermo_provider=None,
        substep_sec=None,
        max_iter=4,
        relax=1.0,
        rtol=1.0e-6,
        atol=1.0e-9,
        max_rhs_evals_per_step=200,
        step_wall_limit_sec=5.0,
        alg_p_tol_psia=0.05,
        alg_v_tol_lbmolph=25.0,
    )
    assert bool(info["fallback_used"]) is False
    assert float(info["ida_converged"]) == pytest.approx(1.0)
    assert float(info["ida_iter_max"]) == pytest.approx(2.0)
    assert float(info["ida_alg_weighted"]) <= 1.0
    assert y1[0] == pytest.approx(2.0)


def test_startup_hydraulic_sequence_stages_pressure_then_energy_then_liquid_ramp():
    base = ColumnInputs(
        pressure_model="hydraulic",
        vapor_flow_model="energy",
        enable_liquid_hydraulic_override=True,
        liquid_hydraulic_override_alpha=1.0,
    )

    p0, v0, a0, ph0 = _resolve_startup_hydraulic_sequence_step(
        t_s=0.0,
        dt_sec=1.0,
        base_inputs=base,
        enable_sequence=True,
        energy_on_sec=10.0,
        liquid_on_sec=20.0,
        liquid_ramp_sec=40.0,
        liquid_resid_gate_lbmolph=250.0,
        liquid_backoff_sec=None,
        liquid_alpha_state=1.0,
        last_mass_resid_max_lbmolph=None,
    )
    assert p0 == "hydraulic"
    assert v0 == "profile"
    assert a0 == pytest.approx(0.0)
    assert ph0 == "pressure_only"

    p1, v1, a1, ph1 = _resolve_startup_hydraulic_sequence_step(
        t_s=12.0,
        dt_sec=1.0,
        base_inputs=base,
        enable_sequence=True,
        energy_on_sec=10.0,
        liquid_on_sec=20.0,
        liquid_ramp_sec=40.0,
        liquid_resid_gate_lbmolph=250.0,
        liquid_backoff_sec=None,
        liquid_alpha_state=a0,
        last_mass_resid_max_lbmolph=25.0,
    )
    assert p1 == "hydraulic"
    assert v1 == "energy"
    assert a1 == pytest.approx(0.0)
    assert ph1 == "pressure_energy"

    _p2, _v2, a2, ph2 = _resolve_startup_hydraulic_sequence_step(
        t_s=40.0,
        dt_sec=40.0,
        base_inputs=base,
        enable_sequence=True,
        energy_on_sec=10.0,
        liquid_on_sec=20.0,
        liquid_ramp_sec=40.0,
        liquid_resid_gate_lbmolph=250.0,
        liquid_backoff_sec=None,
        liquid_alpha_state=a1,
        last_mass_resid_max_lbmolph=25.0,
    )
    assert a2 == pytest.approx(0.5)
    assert ph2 == "pressure_energy_liquid_ramp"


def test_startup_hydraulic_sequence_backs_off_liquid_alpha_on_high_residual():
    base = ColumnInputs(
        pressure_model="hydraulic",
        vapor_flow_model="energy",
        enable_liquid_hydraulic_override=True,
        liquid_hydraulic_override_alpha=1.0,
    )
    _p, _v, a, _ph = _resolve_startup_hydraulic_sequence_step(
        t_s=90.0,
        dt_sec=10.0,
        base_inputs=base,
        enable_sequence=True,
        energy_on_sec=10.0,
        liquid_on_sec=20.0,
        liquid_ramp_sec=40.0,
        liquid_resid_gate_lbmolph=100.0,
        liquid_backoff_sec=20.0,
        liquid_alpha_state=0.8,
        last_mass_resid_max_lbmolph=250.0,
    )
    assert a == pytest.approx(0.3)


def test_startup_hydraulic_sequence_disabled_uses_base_modes():
    base = ColumnInputs(
        pressure_model="hydraulic",
        vapor_flow_model="energy",
        enable_liquid_hydraulic_override=True,
        liquid_hydraulic_override_alpha=0.75,
    )
    p, v, a, ph = _resolve_startup_hydraulic_sequence_step(
        t_s=0.0,
        dt_sec=1.0,
        base_inputs=base,
        enable_sequence=False,
        energy_on_sec=10.0,
        liquid_on_sec=20.0,
        liquid_ramp_sec=40.0,
        liquid_resid_gate_lbmolph=100.0,
        liquid_backoff_sec=None,
        liquid_alpha_state=0.0,
        last_mass_resid_max_lbmolph=None,
    )
    assert p == "hydraulic"
    assert v == "energy"
    assert a == pytest.approx(0.75)
    assert ph == "base"


def test_residual_guarded_liquid_hydraulic_alpha_backs_off_on_high_residual():
    base = ColumnInputs(
        pressure_model="hydraulic",
        vapor_flow_model="energy",
        enable_liquid_hydraulic_override=True,
        liquid_hydraulic_override_alpha=1.0,
    )
    a, phase = _resolve_residual_guarded_liquid_hydraulic_alpha(
        dt_sec=10.0,
        base_inputs=base,
        liquid_resid_gate_lbmolph=100.0,
        liquid_backoff_sec=20.0,
        liquid_recover_sec=40.0,
        liquid_alpha_state=0.8,
        last_mass_resid_max_lbmolph=250.0,
    )
    assert a == pytest.approx(0.4)
    assert phase == "backoff"


def test_residual_guarded_liquid_hydraulic_alpha_recovers_toward_alpha_max():
    base = ColumnInputs(
        pressure_model="hydraulic",
        vapor_flow_model="energy",
        enable_liquid_hydraulic_override=True,
        liquid_hydraulic_override_alpha=0.75,
    )
    a, phase = _resolve_residual_guarded_liquid_hydraulic_alpha(
        dt_sec=10.0,
        base_inputs=base,
        liquid_resid_gate_lbmolph=100.0,
        liquid_backoff_sec=20.0,
        liquid_recover_sec=40.0,
        liquid_alpha_state=0.25,
        last_mass_resid_max_lbmolph=25.0,
    )
    assert a == pytest.approx(0.4375)
    assert phase == "recover"


def test_residual_guarded_liquid_hydraulic_alpha_per_stage_only_backs_off_hot_trays():
    base = ColumnInputs(
        pressure_model="hydraulic",
        vapor_flow_model="energy",
        enable_liquid_hydraulic_override=True,
        liquid_hydraulic_override_alpha=1.0,
    )
    a, phase = _resolve_residual_guarded_liquid_hydraulic_alpha_per_stage(
        dt_sec=10.0,
        base_inputs=base,
        liquid_resid_gate_lbmolph=100.0,
        liquid_recover_sec=40.0,
        liquid_alpha_state=np.ones((5,), dtype=float),
        last_mass_resid_lbmolph_per_stage=np.array([20.0, 250.0, 50.0, 400.0, 10.0], dtype=float),
    )
    assert phase == "backoff"
    assert a.shape == (5,)
    assert a[0] == pytest.approx(1.0)
    assert a[1] == pytest.approx(0.4)
    assert a[2] == pytest.approx(1.0)
    assert a[3] == pytest.approx(0.25)
    assert a[4] == pytest.approx(1.0)


def test_startup_hydraulic_sequence_supports_conductance_vapor_mode():
    base = ColumnInputs(
        pressure_model="hydraulic",
        vapor_flow_model="conductance",
        enable_liquid_hydraulic_override=True,
        liquid_hydraulic_override_alpha=1.0,
    )
    p0, v0, a0, ph0 = _resolve_startup_hydraulic_sequence_step(
        t_s=0.0,
        dt_sec=1.0,
        base_inputs=base,
        enable_sequence=True,
        energy_on_sec=10.0,
        liquid_on_sec=20.0,
        liquid_ramp_sec=40.0,
        liquid_resid_gate_lbmolph=250.0,
        liquid_backoff_sec=None,
        liquid_alpha_state=1.0,
        last_mass_resid_max_lbmolph=None,
    )
    assert p0 == "hydraulic"
    assert v0 == "profile"
    assert a0 == pytest.approx(0.0)
    assert ph0 == "pressure_only"

    p1, v1, a1, ph1 = _resolve_startup_hydraulic_sequence_step(
        t_s=12.0,
        dt_sec=1.0,
        base_inputs=base,
        enable_sequence=True,
        energy_on_sec=10.0,
        liquid_on_sec=20.0,
        liquid_ramp_sec=40.0,
        liquid_resid_gate_lbmolph=250.0,
        liquid_backoff_sec=None,
        liquid_alpha_state=a0,
        last_mass_resid_max_lbmolph=25.0,
    )
    assert p1 == "hydraulic"
    assert v1 == "conductance"
    assert a1 == pytest.approx(0.0)
    assert ph1 == "pressure_energy"


def test_top_drum_dynamic_steady_initializer_with_mocked_rhs(monkeypatch):
    class TinyCol:
        n_stages = 2
        n_components = 2
        M_L_lbmol = np.array([10.0, 10.0], dtype=float)
        M_V_lbmol = np.array([1.0, 1.0], dtype=float)
        x0 = np.array([[0.8, 0.2], [0.4, 0.6]], dtype=float)
        y0 = np.array([[0.7, 0.3], [0.5, 0.5]], dtype=float)
        streams = {}

    col = TinyCol()
    layout = StateVectorLayout(
        n_stages=2,
        n_components=2,
        include_top=True,
        include_bottom=False,
        include_vapor=True,
        include_temperature=False,
        include_energy=False,
    )
    y0 = layout.pack_y0(col)
    sl = layout.slices()
    y0[sl["top_L"]] = np.array([0.8, 0.2], dtype=float) * 1.0
    y0[sl["top_V"]] = np.array([0.6, 0.4], dtype=float) * 9.0

    import dynamic_distillation.dynamic_run_scaffold_v1 as scaffold

    def _fake_rhs(_t, y_vec, _col, _layout, inputs=None):
        u = layout.unpack(np.asarray(y_vec, dtype=float))
        top_L_tot = float(np.sum(np.asarray(u["top_L"], dtype=float)))
        top_V_tot = float(np.sum(np.asarray(u["top_V"], dtype=float)))
        dydt = np.zeros(layout.n_states(), dtype=float)
        # Linear residual model with known root at top_L=5, top_V=3.
        dydt[sl["top_L"]] = np.array([0.7, 0.3], dtype=float) * (top_L_tot - 5.0)
        dydt[sl["top_V"]] = np.array([0.4, 0.6], dtype=float) * (top_V_tot - 3.0)
        return dydt, {}

    monkeypatch.setattr(scaffold, "column_rhs", _fake_rhs)

    y1, info = _initialize_top_drum_dynamic_steady(
        col=col,
        layout=layout,
        y=y0,
        inputs=ColumnInputs(),
        max_iter=6,
        tol_lbmolps=1e-9,
    )

    u1 = layout.unpack(y1)
    top_L_tot = float(np.sum(np.asarray(u1["top_L"], dtype=float)))
    top_V_tot = float(np.sum(np.asarray(u1["top_V"], dtype=float)))
    assert info["attempted"] is True
    assert top_L_tot == pytest.approx(5.0, abs=1e-6)
    assert top_V_tot == pytest.approx(3.0, abs=1e-6)
    assert abs(float(info["d_top_L_final_lbmolps"])) < 1e-7
    assert abs(float(info["d_top_V_final_lbmolps"])) < 1e-7


def test_top_drum_dynamic_steady_initializer_matches_runtime_pressure_model(monkeypatch):
    class TinyCol:
        n_stages = 2
        n_components = 2
        M_L_lbmol = np.array([1.0, 1.0], dtype=float)
        M_V_lbmol = np.array([1.0, 1.0], dtype=float)
        x0 = np.array([[0.8, 0.2], [0.5, 0.5]], dtype=float)
        y0 = np.array([[0.6, 0.4], [0.5, 0.5]], dtype=float)
        T_f = np.array([100.0, 110.0], dtype=float)
        P_psia = np.array([200.0, 205.0], dtype=float)
        streams = {}

    col = TinyCol()
    layout = StateVectorLayout(
        n_stages=2,
        n_components=2,
        include_top=True,
        include_bottom=False,
        include_vapor=True,
        include_temperature=False,
        include_energy=False,
    )
    y0 = layout.pack_y0(col)
    sl = layout.slices()
    y0[sl["top_L"]] = np.array([0.8, 0.2], dtype=float) * 1.0
    y0[sl["top_V"]] = np.array([0.6, 0.4], dtype=float) * 5.0

    import dynamic_distillation.dynamic_run_scaffold_v1 as scaffold

    def _fake_rhs(_t, _y_vec, _col, _layout, inputs=None):
        return np.zeros(layout.n_states(), dtype=float), {}

    def _fake_top_pressure(*, top_V, top_T_F, Z_top, top_vapor_volume_ft3, thermo_provider, y_top, P_seed_psia, return_details=False):
        mv = float(np.sum(np.asarray(top_V, dtype=float)))
        p = 20.0 * mv
        if return_details:
            return p, 0.5, mv
        return p

    monkeypatch.setattr(scaffold, "column_rhs", _fake_rhs)
    monkeypatch.setattr(scaffold, "_compute_top_drum_pressure_psia", _fake_top_pressure)

    y1, info = _initialize_top_drum_dynamic_steady(
        col=col,
        layout=layout,
        y=y0,
        inputs=ColumnInputs(top_drum_vapor_volume_ft3=10.0, thermo_provider=object()),
        max_iter=2,
        tol_lbmolps=1e-9,
    )

    u1 = layout.unpack(y1)
    top_V_tot = float(np.sum(np.asarray(u1["top_V"], dtype=float)))
    assert info["attempted"] is True
    assert info["pressure_coupled"] is True
    assert top_V_tot == pytest.approx(10.0, abs=1e-9)


def test_thermo_startup_conditioner_with_mocked_rhs(monkeypatch):
    class TinyCol:
        n_stages = 3
        n_components = 2
        M_L_lbmol = np.array([5.0, 6.0, 7.0], dtype=float)
        M_V_lbmol = np.array([1.0, 1.5, 2.0], dtype=float)
        x0 = np.array([[0.8, 0.2], [0.5, 0.5], [0.2, 0.8]], dtype=float)
        y0 = np.array([[0.7, 0.3], [0.4, 0.6], [0.1, 0.9]], dtype=float)
        T_f = np.array([100.0, 110.0, 120.0], dtype=float)
        P_psia = np.array([200.0, 205.0, 210.0], dtype=float)
        streams = {}

    col = TinyCol()
    layout = StateVectorLayout(
        n_stages=3,
        n_components=2,
        include_top=False,
        include_bottom=False,
        include_vapor=True,
        include_temperature=False,
        include_energy=True,
    )
    y0 = layout.pack_y0(col)
    sl = layout.slices()

    x_eq = np.array([[0.9, 0.1], [0.6, 0.4], [0.3, 0.7]], dtype=float)
    y_eq = np.array([[0.8, 0.2], [0.55, 0.45], [0.25, 0.75]], dtype=float)
    import dynamic_distillation.dynamic_run_scaffold_v1 as scaffold

    def _fake_rhs(_t, y_vec, _col, _layout, inputs=None):
        u = layout.unpack(np.asarray(y_vec, dtype=float))
        x_now = np.asarray(u["x_tray"], dtype=float).reshape((col.n_stages, col.n_components))
        y_now = np.asarray(u["y_tray"], dtype=float).reshape((col.n_stages, col.n_components))
        gap = np.sum(np.abs(x_now - x_eq), axis=1)
        HL_now = 10.0 + 100.0 * x_now[:, 0] + 10.0 * x_now[:, 1]
        HV_now = 50.0 + 200.0 * y_now[:, 0] + 20.0 * y_now[:, 1]
        dydt = np.zeros(layout.n_states(), dtype=float)
        diag = {
            "x_eq_tray": x_eq.copy(),
            "y_eq_tray": y_eq.copy(),
            "HL_BTU_lbmol_tray": HL_now,
            "HV_BTU_lbmol_tray": HV_now,
            "Z_tray": np.ones(col.n_stages, dtype=float),
            "eq_phase_change_lbmolps_tray": gap,
        }
        return dydt, diag

    monkeypatch.setattr(scaffold, "column_rhs", _fake_rhs)
    monkeypatch.setattr(
        scaffold,
        "_initialize_vapor_holdup_from_spec_pressure",
        lambda **kwargs: np.asarray(kwargs["y"], dtype=float).copy(),
    )

    y1, info = _initialize_thermo_consistent_state(
        col=col,
        layout=layout,
        y=y0,
        inputs=ColumnInputs(thermo_provider=object()),
        include_temperature=False,
        max_iter=1,
        relaxation=1.0,
    )

    u1 = layout.unpack(y1)
    x1 = np.asarray(u1["x_tray"], dtype=float).reshape((col.n_stages, col.n_components))
    y1_frac = np.asarray(u1["y_tray"], dtype=float).reshape((col.n_stages, col.n_components))
    assert info["attempted"] is True
    assert int(info["n_iter"]) == 1
    assert np.allclose(x1, x_eq, atol=1e-12)
    # Stage 1 vapor holdup is forced to zero by startup policy.
    assert np.allclose(y1_frac[1:, :], y_eq[1:, :], atol=1e-12)

    ml_tot = np.asarray(u1["ML_tot_tray"], dtype=float).reshape((col.n_stages,))
    mv_tot = np.asarray(u1["MV_tot_tray"], dtype=float).reshape((col.n_stages,))
    tray_EL = np.asarray(y1[sl["tray_EL_BTU"]], dtype=float).reshape((col.n_stages,))
    tray_EV = np.asarray(y1[sl["tray_EV_BTU"]], dtype=float).reshape((col.n_stages,))
    HL_expected = 10.0 + 100.0 * x_eq[:, 0] + 10.0 * x_eq[:, 1]
    HV_expected = 50.0 + 200.0 * y_eq[:, 0] + 20.0 * y_eq[:, 1]
    assert np.allclose(tray_EL, ml_tot * HL_expected, atol=1e-12)
    assert np.allclose(tray_EV, mv_tot * HV_expected, atol=1e-12)
    assert abs(float(info["eq_phase_change_final_lbmolps"])) < float(info["eq_phase_change_init_lbmolps"])


def test_thermo_startup_conditioner_preserves_boundary_liquid_compositions(monkeypatch):
    class TinyCol:
        n_stages = 3
        n_components = 2
        M_L_lbmol = np.array([5.0, 6.0, 7.0], dtype=float)
        M_V_lbmol = np.array([1.0, 1.5, 2.0], dtype=float)
        x0 = np.array([[0.8, 0.2], [0.5, 0.5], [0.2, 0.8]], dtype=float)
        y0 = np.array([[0.7, 0.3], [0.4, 0.6], [0.1, 0.9]], dtype=float)
        T_f = np.array([100.0, 110.0, 120.0], dtype=float)
        P_psia = np.array([200.0, 205.0, 210.0], dtype=float)
        streams = {}

    col = TinyCol()
    layout = StateVectorLayout(
        n_stages=3,
        n_components=2,
        include_top=True,
        include_bottom=True,
        include_vapor=True,
        include_temperature=False,
        include_energy=False,
    )
    y0 = layout.pack_y0(col)
    sl = layout.slices()
    y0 = np.asarray(y0, dtype=float).copy()
    y0[sl["top_L"]] = np.array([9.0, 1.0], dtype=float)
    y0[sl["bottom_L"]] = np.array([2.0, 8.0], dtype=float)

    x_eq = np.array([[0.1, 0.9], [0.6, 0.4], [0.9, 0.1]], dtype=float)
    y_eq = np.array([[0.2, 0.8], [0.55, 0.45], [0.75, 0.25]], dtype=float)

    import dynamic_distillation.dynamic_run_scaffold_v1 as scaffold

    def _fake_rhs(_t, y_vec, _col, _layout, inputs=None):
        dydt = np.zeros_like(np.asarray(y_vec, dtype=float))
        diag = {
            "x_eq_tray": x_eq.copy(),
            "y_eq_tray": y_eq.copy(),
            "Z_tray": np.ones(col.n_stages, dtype=float),
            "eq_phase_change_lbmolps_tray": np.ones(col.n_stages, dtype=float),
        }
        return dydt, diag

    monkeypatch.setattr(scaffold, "column_rhs", _fake_rhs)
    monkeypatch.setattr(
        scaffold,
        "_initialize_vapor_holdup_from_spec_pressure",
        lambda **kwargs: np.asarray(kwargs["y"], dtype=float).copy(),
    )

    y1, info = _initialize_thermo_consistent_state(
        col=col,
        layout=layout,
        y=y0,
        inputs=ColumnInputs(thermo_provider=object()),
        include_temperature=False,
        max_iter=1,
        relaxation=1.0,
    )

    u1 = layout.unpack(y1)
    top_L = np.asarray(u1["top_L"], dtype=float).reshape((-1,))
    bottom_L = np.asarray(u1["bottom_L"], dtype=float).reshape((-1,))
    assert info["attempted"] is True
    assert np.allclose(top_L, np.array([9.0, 1.0], dtype=float), atol=1e-12)
    assert np.allclose(bottom_L, np.array([2.0, 8.0], dtype=float), atol=1e-12)


def test_hydraulic_energy_startup_consistency_noops_for_non_hydraulic_energy_mode():
    class TinyCol:
        n_stages = 2
        n_components = 1
        M_L_lbmol = np.array([5.0, 6.0], dtype=float)
        M_V_lbmol = np.array([1.0, 1.0], dtype=float)
        x0 = np.array([[1.0], [1.0]], dtype=float)
        y0 = np.array([[1.0], [1.0]], dtype=float)
        T_f = np.array([100.0, 110.0], dtype=float)
        P_psia = np.array([200.0, 205.0], dtype=float)
        streams = {}

    col = TinyCol()
    layout = StateVectorLayout(
        n_stages=2,
        n_components=1,
        include_top=False,
        include_bottom=False,
        include_vapor=True,
        include_temperature=False,
        include_energy=True,
    )
    y0 = layout.pack_y0(col)

    y1, info = _initialize_hydraulic_energy_consistent_state(
        col=col,
        layout=layout,
        y=y0,
        inputs=ColumnInputs(pressure_model="spec", vapor_flow_model="profile"),
        include_temperature=False,
    )

    assert np.allclose(y1, y0)
    assert info["attempted"] is False


def test_hydraulic_energy_startup_consistency_relaxes_state_when_objective_improves(monkeypatch):
    import dynamic_distillation.dynamic_run_scaffold_v1 as scaffold

    class TinyCol:
        n_stages = 1
        n_components = 1
        M_L_lbmol = np.array([10.0], dtype=float)
        M_V_lbmol = np.array([2.0], dtype=float)
        x0 = np.array([[1.0]], dtype=float)
        y0 = np.array([[1.0]], dtype=float)
        T_f = np.array([100.0], dtype=float)
        P_psia = np.array([200.0], dtype=float)
        streams = {}

    col = TinyCol()
    layout = StateVectorLayout(
        n_stages=1,
        n_components=1,
        include_top=False,
        include_bottom=False,
        include_vapor=True,
        include_temperature=False,
        include_energy=False,
    )
    y0 = layout.pack_y0(col)
    sl = layout.slices()

    def _fake_solve_dae(*, y, **_kwargs):
        y_arr = np.asarray(y, dtype=float)
        ml = float(np.asarray(y_arr[sl["tray_L"]], dtype=float).reshape((-1,))[0])
        gap = ml - 4.0
        dydt = np.zeros_like(y_arr)
        dydt[sl["tray_L"]] = np.array([-gap], dtype=float)
        diag = {
            "mass_balance_resid_lbmolps_tray": np.array([abs(gap) / 3600.0], dtype=float),
            "resid_energy_btups": np.array([abs(gap) * 100.0], dtype=float),
            "dae_pilot_alg_p_inf_psia": np.array([abs(gap) * 0.1], dtype=float),
            "dae_pilot_alg_v_inf_lbmolph": np.array([abs(gap) * 10.0], dtype=float),
            "P_psia_hyd": np.array([200.0], dtype=float),
            "V_out_lbmolph": np.array([5000.0], dtype=float),
        }
        return dydt, diag

    monkeypatch.setattr(scaffold, "_solve_dae_pilot_algebraic", _fake_solve_dae)

    y1, info = _initialize_hydraulic_energy_consistent_state(
        col=col,
        layout=layout,
        y=y0,
        inputs=ColumnInputs(pressure_model="hydraulic", vapor_flow_model="energy"),
        include_temperature=False,
        max_iter=4,
        pseudo_dt_sec=0.5,
        mass_tol_lbmolph=5.0,
        energy_tol_btups=1000.0,
    )

    ml0 = float(np.asarray(y0[sl["tray_L"]], dtype=float).reshape((-1,))[0])
    ml1 = float(np.asarray(y1[sl["tray_L"]], dtype=float).reshape((-1,))[0])
    assert ml1 < ml0
    assert info["attempted"] is True
    assert info["success"] is True
    assert float(info["objective_final"]) < float(info["objective_init"])


def test_inner_pv_coupling_single_pass_when_max_iter_is_one(monkeypatch):
    import dynamic_distillation.dynamic_run_scaffold_v1 as scaffold

    calls = {"n": 0}

    def _fake_rhs(_t, y_vec, _col, _layout, inputs=None):
        calls["n"] += 1
        dydt = np.zeros_like(np.asarray(y_vec, dtype=float))
        diag = {
            "P_psia_hyd": np.array([220.0, 230.0], dtype=float),
            "V_out_lbmolph": np.array([5000.0, 5100.0], dtype=float),
        }
        return dydt, diag

    monkeypatch.setattr(scaffold, "column_rhs", _fake_rhs)

    class TinyCol:
        n_stages = 2

    layout = StateVectorLayout(
        n_stages=2,
        n_components=1,
        include_top=False,
        include_bottom=False,
        include_vapor=False,
    )
    y0 = np.zeros(layout.n_states(), dtype=float)

    dydt, diag = _column_rhs_with_inner_pv_coupling(
        t_s=0.0,
        y=y0,
        col=TinyCol(),
        layout=layout,
        inputs=ColumnInputs(
            P_tray_prev=np.array([210.0, 220.0], dtype=float),
            V_out_prev_lbmolph=np.array([4800.0, 4800.0], dtype=float),
        ),
        max_iter=1,
        p_tol_psia=0.05,
        v_tol_lbmolph=25.0,
    )

    assert calls["n"] == 1
    assert dydt.shape == y0.shape
    assert float(np.asarray(diag["pv_inner_iter_count"], dtype=float).reshape((-1,))[0]) == 1.0
    assert float(np.asarray(diag["pv_inner_converged"], dtype=float).reshape((-1,))[0]) == 0.0


def test_inner_pv_coupling_iterates_and_converges(monkeypatch):
    import dynamic_distillation.dynamic_run_scaffold_v1 as scaffold

    calls = {"n": 0}
    p_star = np.array([225.0, 235.0], dtype=float)
    v_star = np.array([5200.0, 5400.0], dtype=float)

    def _fake_rhs(_t, y_vec, _col, _layout, inputs=None):
        calls["n"] += 1
        p_prev = np.asarray(inputs.P_tray_prev, dtype=float).reshape((2,))
        v_prev = np.asarray(inputs.V_out_prev_lbmolph, dtype=float).reshape((2,))
        p_new = p_prev + 0.5 * (p_star - p_prev)
        v_new = v_prev + 0.5 * (v_star - v_prev)
        dydt = np.zeros_like(np.asarray(y_vec, dtype=float))
        diag = {
            "P_psia_hyd": p_new,
            "V_out_lbmolph": v_new,
            "Z_tray": np.ones(2, dtype=float),
        }
        return dydt, diag

    monkeypatch.setattr(scaffold, "column_rhs", _fake_rhs)

    class TinyCol:
        n_stages = 2

    layout = StateVectorLayout(
        n_stages=2,
        n_components=1,
        include_top=False,
        include_bottom=False,
        include_vapor=False,
    )
    y0 = np.zeros(layout.n_states(), dtype=float)

    _dydt, diag = _column_rhs_with_inner_pv_coupling(
        t_s=0.0,
        y=y0,
        col=TinyCol(),
        layout=layout,
        inputs=ColumnInputs(
            P_tray_prev=np.array([200.0, 210.0], dtype=float),
            V_out_prev_lbmolph=np.array([4800.0, 5000.0], dtype=float),
        ),
        max_iter=8,
        p_tol_psia=0.25,
        v_tol_lbmolph=10.0,
    )

    assert calls["n"] >= 2
    assert float(np.asarray(diag["pv_inner_converged"], dtype=float).reshape((-1,))[0]) == 1.0
    assert float(np.asarray(diag["pv_inner_iter_count"], dtype=float).reshape((-1,))[0]) <= 8.0
    assert float(np.asarray(diag["pv_inner_dp_max_psia"], dtype=float).reshape((-1,))[0]) <= 0.25
    assert float(np.asarray(diag["pv_inner_dv_max_lbmolph"], dtype=float).reshape((-1,))[0]) <= 10.0


def test_dae_pilot_algebraic_solver_converges_for_linear_map(monkeypatch):
    import dynamic_distillation.dynamic_run_scaffold_v1 as scaffold

    calls = {"n": 0}
    p_star = np.array([225.0, 235.0], dtype=float)
    v_star = np.array([5200.0, 5400.0], dtype=float)

    def _fake_rhs(_t, y_vec, _col, _layout, inputs=None):
        calls["n"] += 1
        p_prev = np.asarray(inputs.P_tray_prev, dtype=float).reshape((2,))
        v_prev = np.asarray(inputs.V_out_prev_lbmolph, dtype=float).reshape((2,))
        p_new = p_prev + 0.5 * (p_star - p_prev)
        v_new = v_prev + 0.5 * (v_star - v_prev)
        dydt = np.zeros_like(np.asarray(y_vec, dtype=float))
        diag = {
            "P_psia_hyd": p_new,
            "V_out_lbmolph": v_new,
        }
        return dydt, diag

    monkeypatch.setattr(scaffold, "column_rhs", _fake_rhs)

    class TinyCol:
        n_stages = 2

    layout = StateVectorLayout(
        n_stages=2,
        n_components=1,
        include_top=False,
        include_bottom=False,
        include_vapor=False,
    )
    y0 = np.zeros(layout.n_states(), dtype=float)

    _dydt, diag = _solve_dae_pilot_algebraic(
        t_s=0.0,
        y=y0,
        col=TinyCol(),
        layout=layout,
        inputs=ColumnInputs(
            P_tray_prev=np.array([200.0, 210.0], dtype=float),
            V_out_prev_lbmolph=np.array([4800.0, 5000.0], dtype=float),
        ),
        max_iter=4,
        p_tol_psia=1e-6,
        v_tol_lbmolph=1e-4,
        jac_rel_step=1e-7,
        line_search_max=3,
    )

    assert calls["n"] >= 2
    assert float(np.asarray(diag["dae_pilot_enabled"], dtype=float).reshape((-1,))[0]) == 1.0
    assert float(np.asarray(diag["dae_pilot_converged"], dtype=float).reshape((-1,))[0]) == 1.0
    assert float(np.asarray(diag["dae_pilot_failed"], dtype=float).reshape((-1,))[0]) == 0.0
    assert float(np.asarray(diag["dae_pilot_alg_p_inf_psia"], dtype=float).reshape((-1,))[0]) <= 1e-6
    assert float(np.asarray(diag["dae_pilot_alg_v_inf_lbmolph"], dtype=float).reshape((-1,))[0]) <= 1e-4


def test_dae_pilot_algebraic_solver_fallback_sets_failed_flag(monkeypatch):
    import dynamic_distillation.dynamic_run_scaffold_v1 as scaffold

    def _fake_rhs(_t, y_vec, _col, _layout, inputs=None):
        dydt = np.zeros_like(np.asarray(y_vec, dtype=float))
        diag = {
            "P_psia_hyd": np.asarray(inputs.P_tray_prev, dtype=float).reshape((2,)) + 1.0,
            "V_out_lbmolph": np.asarray(inputs.V_out_prev_lbmolph, dtype=float).reshape((2,)) + 10.0,
        }
        return dydt, diag

    def _raise_jac(*_args, **_kwargs):
        raise RuntimeError("fd jacobian failure")

    monkeypatch.setattr(scaffold, "column_rhs", _fake_rhs)
    monkeypatch.setattr(scaffold, "finite_difference_jacobian", _raise_jac)

    class TinyCol:
        n_stages = 2

    layout = StateVectorLayout(
        n_stages=2,
        n_components=1,
        include_top=False,
        include_bottom=False,
        include_vapor=False,
    )
    y0 = np.zeros(layout.n_states(), dtype=float)

    _dydt, diag = _solve_dae_pilot_algebraic(
        t_s=0.0,
        y=y0,
        col=TinyCol(),
        layout=layout,
        inputs=ColumnInputs(
            P_tray_prev=np.array([200.0, 210.0], dtype=float),
            V_out_prev_lbmolph=np.array([4800.0, 5000.0], dtype=float),
        ),
        max_iter=3,
        p_tol_psia=1e-12,
        v_tol_lbmolph=1e-12,
        jac_rel_step=1e-7,
        line_search_max=2,
    )

    assert float(np.asarray(diag["dae_pilot_enabled"], dtype=float).reshape((-1,))[0]) == 1.0
    assert float(np.asarray(diag["dae_pilot_failed"], dtype=float).reshape((-1,))[0]) == 1.0
    assert float(np.asarray(diag["dae_pilot_converged"], dtype=float).reshape((-1,))[0]) == 0.0


def test_stiff_integrator_uses_dae_outer_once_per_step(monkeypatch, tmp_path: Path):
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    import dynamic_distillation.dynamic_run_scaffold_v1 as scaffold

    calls = {"dae": 0, "pv": 0}
    smooth_eps_seen = []

    def _fake_dae(
        *,
        t_s,
        y,
        col,
        layout,
        inputs,
        max_iter,
        p_tol_psia,
        v_tol_lbmolph,
        jac_rel_step,
        line_search_max,
    ):
        calls["dae"] += 1
        smooth_eps_seen.append(float(getattr(inputs, "vflow_smooth_clamp_epsilon_lbmolps", 0.0) or 0.0))
        n = int(layout.n_states())
        ns = int(col.n_stages)
        dydt = np.zeros(n, dtype=float)
        diag = {
            "P_psia_hyd": np.full(ns, 220.0, dtype=float),
            "V_out_lbmolph": np.full(ns, 5000.0, dtype=float),
            "Z_tray": np.ones(ns, dtype=float),
            "dae_pilot_enabled": np.array([1.0], dtype=float),
            "dae_pilot_iter_count": np.array([1.0], dtype=float),
            "dae_pilot_converged": np.array([1.0], dtype=float),
            "dae_pilot_failed": np.array([0.0], dtype=float),
        }
        return dydt, diag

    def _fake_pv(*, t_s, y, col, layout, inputs, max_iter, p_tol_psia, v_tol_lbmolph):
        calls["pv"] += 1
        n = int(layout.n_states())
        ns = int(col.n_stages)
        dydt = np.zeros(n, dtype=float)
        diag = {
            "P_psia_hyd": np.full(ns, 220.0, dtype=float),
            "V_out_lbmolph": np.full(ns, 5000.0, dtype=float),
            "Z_tray": np.ones(ns, dtype=float),
        }
        return dydt, diag

    def _fake_integrate_one_step(
        *,
        t_s,
        y,
        dt_sec,
        rhs_eval,
        rhs_eval_fallback,
        layout,
        thermo_provider,
        integrator_mode,
        rtol,
        atol,
        max_step_sec,
        substep_sec,
        max_rhs_evals_per_step,
        step_wall_limit_sec,
    ):
        _d1, _g1 = rhs_eval(float(t_s), np.asarray(y, dtype=float))
        _d2, _g2 = rhs_eval(float(t_s) + 0.5 * float(dt_sec), np.asarray(y, dtype=float))
        return np.asarray(y, dtype=float), {"fallback_used": False}

    monkeypatch.setattr(scaffold, "_solve_dae_pilot_algebraic", _fake_dae)
    monkeypatch.setattr(scaffold, "_column_rhs_with_inner_pv_coupling", _fake_pv)
    monkeypatch.setattr(scaffold, "_integrate_one_step", _fake_integrate_one_step)

    cfg = RunnerConfig(
        excel_path=str(excel),
        n_steps=1,
        dt_sec=0.2,
        log_every_n_steps=1,
        runtime_mode="hydraulic",
        include_temperature=False,
        include_energy=False,
        thermo_mode="stub",
        write_logs=False,
        integrator="bdf",
        enable_dae_pilot_algebraic_solve=True,
        logs_dir=str(tmp_path),
    )
    out = run_smoke_simulation(cfg)

    # n_steps=1 -> outer-step evaluation at step 0 and step 1 only.
    assert calls["dae"] == 2
    # substep RHS inside fake integrator should not invoke the PV wrapper in this mode.
    assert calls["pv"] == 0
    assert len(smooth_eps_seen) == 2
    assert all(v > 0.0 for v in smooth_eps_seen)
    assert int(out.get("integrator_fallback_count", 0)) == 0


def test_distillate_composition_control_logs_command(tmp_path: Path):
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return
    thermo_table = Path("cache/thermo_table.json")
    if not thermo_table.exists():
        return

    cfg = RunnerConfig(
        excel_path=str(excel),
        n_steps=0,
        dt_sec=0.2,
        log_every_n_steps=1,
        include_temperature=True,
        include_energy=True,
        enable_equilibrium_relaxation=True,
        thermo_mode="table",
        thermo_table_path=str(thermo_table),
        logs_dir=str(tmp_path),
        write_logs=True,
        enable_level_control=True,
        top_level_sp_lbmol=397.0,
        bottom_level_sp_lbmol=794.0,
        enable_distillate_composition_control=True,
        distillate_composition_component="C4",
        distillate_composition_sp_molfrac=0.05,
    )

    out = run_smoke_simulation(cfg)
    summary_csv = Path(str(out["summary_csv"]))
    assert summary_csv.exists()

    import csv

    with summary_csv.open("r", encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows, "summary log is empty"

    r0 = rows[0]
    assert "xD_comp_sp" in r0
    assert "xD_comp_pv" in r0
    assert "RR_comp_cmd" in r0
    assert np.isfinite(float(r0["xD_comp_sp"]))
    assert np.isfinite(float(r0["RR_comp_cmd"]))


def test_build_inputs_reads_and_overrides_condenser_pressure_drop():
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)

    specs = dict(getattr(col, "specs_raw", {}) or {})
    specs["Condenser Pressure Drop (psi)"] = 2.0
    object.__setattr__(col, "specs_raw", specs)

    cfg = RunnerConfig(excel_path=str(excel), thermo_mode="stub")
    inputs, _ = build_inputs_for_runner(case, col, cfg)
    assert abs(float(inputs.condenser_pressure_drop_psi) - 2.0) < 1e-12

    cfg_override = RunnerConfig(excel_path=str(excel), thermo_mode="stub", condenser_pressure_drop_psi=1.25)
    inputs_override, _ = build_inputs_for_runner(case, col, cfg_override)
    assert abs(float(inputs_override.condenser_pressure_drop_psi) - 1.25) < 1e-12


def test_build_inputs_can_enable_selective_live_pr_for_equilibrium_relaxation(monkeypatch):
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case
    import dynamic_distillation.thermo_provider_v1 as thermo_provider_v1

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)

    class _FakeThermoProvider:
        def __init__(self, *args, **kwargs):
            self.args = args
            self.kwargs = kwargs

    monkeypatch.setattr(thermo_provider_v1, "ThermoProviderV1", _FakeThermoProvider)

    cfg = RunnerConfig(
        excel_path=str(excel),
        thermo_mode="stub",
        equilibrium_relaxation_live_pr=True,
    )
    inputs, _ = build_inputs_for_runner(case, col, cfg)
    assert inputs.equilibrium_relaxation_thermo_provider is not None


def test_build_inputs_can_enable_selective_live_pr_from_specs(monkeypatch):
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case
    import dynamic_distillation.thermo_provider_v1 as thermo_provider_v1

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)

    specs = dict(getattr(col, "specs_raw", {}) or {})
    specs["Equilibrium Relaxation Live PR"] = True
    object.__setattr__(col, "specs_raw", specs)

    class _FakeThermoProvider:
        def __init__(self, *args, **kwargs):
            self.args = args
            self.kwargs = kwargs

    monkeypatch.setattr(thermo_provider_v1, "ThermoProviderV1", _FakeThermoProvider)

    cfg = RunnerConfig(
        excel_path=str(excel),
        thermo_mode="stub",
    )
    inputs, _ = build_inputs_for_runner(case, col, cfg)
    assert inputs.equilibrium_relaxation_thermo_provider is not None


def test_build_inputs_can_select_dwsim_property_package(monkeypatch):
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case
    import dynamic_distillation.thermo_provider_v1 as thermo_provider_v1

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)

    calls = []

    class _FakeThermoProvider:
        def __init__(self, *args, **kwargs):
            calls.append(kwargs)

    monkeypatch.setattr(thermo_provider_v1, "ThermoProviderV1", _FakeThermoProvider)

    cfg = RunnerConfig(
        excel_path=str(excel),
        thermo_mode="dwsim",
        dwsim_property_package="unifac",
    )
    _inputs, _prov = build_inputs_for_runner(case, col, cfg)
    assert calls
    assert calls[0]["property_package"] == "unifac"


def test_build_inputs_dwsim_selective_live_pr_builds_separate_pr_provider(monkeypatch):
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case
    import dynamic_distillation.thermo_provider_v1 as thermo_provider_v1

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)

    calls = []

    class _FakeThermoProvider:
        def __init__(self, *args, **kwargs):
            calls.append(dict(kwargs))

    monkeypatch.setattr(thermo_provider_v1, "ThermoProviderV1", _FakeThermoProvider)

    cfg = RunnerConfig(
        excel_path=str(excel),
        thermo_mode="dwsim",
        dwsim_property_package="unifac",
        equilibrium_relaxation_live_pr=True,
    )
    inputs, _prov = build_inputs_for_runner(case, col, cfg)

    assert len(calls) >= 2
    assert calls[0]["property_package"] == "unifac"
    assert calls[1]["property_package"] == "pr"
    assert inputs.equilibrium_relaxation_thermo_provider is not None


def test_build_inputs_dwsim_pr_does_not_build_redundant_live_pr_override(monkeypatch):
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case
    import dynamic_distillation.thermo_provider_v1 as thermo_provider_v1

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)

    calls = []

    class _FakeThermoProvider:
        def __init__(self, *args, **kwargs):
            calls.append(dict(kwargs))

    monkeypatch.setattr(thermo_provider_v1, "ThermoProviderV1", _FakeThermoProvider)

    cfg = RunnerConfig(
        excel_path=str(excel),
        thermo_mode="dwsim",
        dwsim_property_package="pr",
        equilibrium_relaxation_live_pr=True,
    )
    inputs, _prov = build_inputs_for_runner(case, col, cfg)

    assert len(calls) == 1
    assert calls[0]["property_package"] == "pr"
    assert inputs.equilibrium_relaxation_thermo_provider is None


def test_run_smoke_explicit_euler_skips_integrator_rhs_recompute(monkeypatch, tmp_path: Path):
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    def _forbid_integrate_one_step(**kwargs):
        raise AssertionError("_integrate_one_step should not be called for explicit-euler runtime steps")

    monkeypatch.setattr(runmod, "_integrate_one_step", _forbid_integrate_one_step)

    cfg = RunnerConfig(
        excel_path=str(excel),
        n_steps=1,
        dt_sec=0.2,
        log_every_n_steps=1,
        runtime_mode="parity",
        include_temperature=False,
        include_energy=False,
        thermo_mode="stub",
        write_logs=False,
        integrator="explicit-euler",
        logs_dir=str(tmp_path),
    )

    out = run_smoke_simulation(cfg)
    assert float(out.get("final_time_s", 0.0)) == pytest.approx(0.2)


def test_build_inputs_dwsim_unifac_mode_alias_sets_unifac_package(monkeypatch):
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case
    import dynamic_distillation.thermo_provider_v1 as thermo_provider_v1

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)

    calls = []

    class _FakeThermoProvider:
        def __init__(self, *args, **kwargs):
            calls.append(kwargs)

    monkeypatch.setattr(thermo_provider_v1, "ThermoProviderV1", _FakeThermoProvider)

    cfg = RunnerConfig(
        excel_path=str(excel),
        thermo_mode="dwsim-unifac",
        dwsim_property_package="pr",
    )
    _inputs, _prov = build_inputs_for_runner(case, col, cfg)
    assert calls
    assert calls[0]["property_package"] == "unifac"


def test_build_inputs_reads_and_overrides_top_psv_settings():
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)

    specs = dict(getattr(col, "specs_raw", {}) or {})
    specs["Enable Top PSV"] = "Yes"
    specs["Top PSV SP (psia)"] = 245.0
    specs["Top PSV Gain (lbmol/s/psi)"] = 0.05
    specs["Top PSV Max Vent (lbmol/s)"] = 0.60
    object.__setattr__(col, "specs_raw", specs)

    cfg = RunnerConfig(excel_path=str(excel), thermo_mode="stub")
    inputs, _ = build_inputs_for_runner(case, col, cfg)
    assert bool(inputs.enable_top_drum_psv) is True
    assert abs(float(inputs.top_drum_psv_setpoint_psia) - 245.0) < 1e-12
    assert abs(float(inputs.top_drum_psv_gain_lbmolps_per_psi) - 0.05) < 1e-12
    assert abs(float(inputs.top_drum_psv_max_vent_lbmolps) - 0.60) < 1e-12

    cfg_override = RunnerConfig(
        excel_path=str(excel),
        thermo_mode="stub",
        enable_top_psv=True,
        top_psv_setpoint_psia=230.0,
        top_psv_gain_lbmolps_per_psi=0.20,
        top_psv_max_vent_lbmolps=0.90,
    )
    inputs_override, _ = build_inputs_for_runner(case, col, cfg_override)
    assert bool(inputs_override.enable_top_drum_psv) is True
    assert abs(float(inputs_override.top_drum_psv_setpoint_psia) - 230.0) < 1e-12
    assert abs(float(inputs_override.top_drum_psv_gain_lbmolps_per_psi) - 0.20) < 1e-12
    assert abs(float(inputs_override.top_drum_psv_max_vent_lbmolps) - 0.90) < 1e-12


def test_build_inputs_computes_top_drum_vapor_volume_from_geometry():
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)

    specs = dict(getattr(col, "specs_raw", {}) or {})
    specs["Top Drum Vapor Volume (ft3)"] = None
    specs["Top Drum Total Volume (ft3)"] = None
    specs["Top Drum Diameter (ft)"] = 10.0
    specs["Top Drum Length (ft)"] = 40.0
    specs["Top Drum Liquid Fraction (-)"] = 0.60
    specs["Overhead Vapor Line Volume (ft3)"] = None
    specs["Condenser Vapor Volume (ft3)"] = None
    object.__setattr__(col, "specs_raw", specs)

    cfg = RunnerConfig(excel_path=str(excel), thermo_mode="stub")
    inputs, _ = build_inputs_for_runner(case, col, cfg)

    total_vol = float(np.pi * 0.25 * 10.0 * 10.0 * 40.0)
    expected_vapor = total_vol * (1.0 - 0.60)
    assert inputs.top_drum_total_volume_ft3 is not None
    assert abs(float(inputs.top_drum_total_volume_ft3) - total_vol) < 1e-9
    assert inputs.top_drum_vapor_volume_ft3 is not None
    assert abs(float(inputs.top_drum_vapor_volume_ft3) - expected_vapor) < 1e-9


def test_build_inputs_prefers_explicit_top_holdup_over_liquid_fraction_for_startup():
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)

    specs = dict(getattr(col, "specs_raw", {}) or {})
    specs["Top Drum Vapor Volume (ft3)"] = None
    specs["Top Drum Total Volume (ft3)"] = None
    specs["Top Drum Diameter (ft)"] = 10.0
    specs["Top Drum Length (ft)"] = 40.0
    specs["Top Drum Liquid Fraction (-)"] = 0.60
    specs["Top Accumulator Holdup (lbmol)"] = 397.0
    specs["Overhead Vapor Line Volume (ft3)"] = None
    specs["Condenser Vapor Volume (ft3)"] = None
    object.__setattr__(col, "specs_raw", specs)

    cfg = RunnerConfig(excel_path=str(excel), thermo_mode="stub")
    inputs, _ = build_inputs_for_runner(case, col, cfg)

    total_vol = float(np.pi * 0.25 * 10.0 * 10.0 * 40.0)
    rho_top = float(col.rhoL_lbmol_ft3[0])
    expected_liq_vol = float(specs["Top Accumulator Holdup (lbmol)"]) / rho_top
    expected_vapor = total_vol - expected_liq_vol

    assert inputs.top_drum_total_volume_ft3 is not None
    assert abs(float(inputs.top_drum_total_volume_ft3) - total_vol) < 1e-9
    assert inputs.top_drum_vapor_volume_ft3 is not None
    assert abs(float(inputs.top_drum_vapor_volume_ft3) - expected_vapor) < 1e-9


def test_build_inputs_computes_bottom_sump_total_volume_from_diameter_and_height():
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)

    specs = dict(getattr(col, "specs_raw", {}) or {})
    specs["Bottom Sump Total Volume (ft3)"] = None
    specs["Bottom Sump Diameter (ft)"] = 18.0
    specs["Bottom Sump Height (ft)"] = 12.0
    object.__setattr__(col, "specs_raw", specs)

    cfg = RunnerConfig(excel_path=str(excel), thermo_mode="stub")
    inputs, _ = build_inputs_for_runner(case, col, cfg)

    total_vol = float(np.pi * 0.25 * 18.0 * 18.0 * 12.0)
    assert inputs.bottom_sump_total_volume_ft3 is not None
    assert abs(float(inputs.bottom_sump_total_volume_ft3) - total_vol) < 1e-9


def test_build_inputs_reads_equilibrium_tuning_from_specs():
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)

    specs = dict(getattr(col, "specs_raw", {}) or {})
    specs["Equilibrium Tau (sec)"] = 4.0
    specs["Equilibrium Energy Damping Gain"] = 0.2
    specs["Hydraulic Energy Temperature Follow Tau (sec)"] = 1.25
    object.__setattr__(col, "specs_raw", specs)

    cfg = RunnerConfig(excel_path=str(excel), thermo_mode="stub")
    inputs, _ = build_inputs_for_runner(case, col, cfg)

    assert inputs.tau_eq_sec == pytest.approx(4.0)
    assert inputs.equilibrium_energy_damping_gain == pytest.approx(0.2)
    assert inputs.hydraulic_energy_temperature_follow_tau_sec == pytest.approx(1.25)


def test_build_inputs_skips_primary_backend_prewarm_by_default(monkeypatch):
    class TinyCol:
        n_stages = 3
        n_components = 2
        components_excel = ["A", "B"]
        components_dwsim = ["A", "B"]
        specs_raw = {}
        geometry = None
        T_f = np.array([100.0, 110.0, 120.0], dtype=float)
        P_psia = np.array([200.0, 205.0, 210.0], dtype=float)
        V_lbmolph = np.array([10.0, 11.0, 12.0], dtype=float)
        L_lbmolph = np.array([20.0, 21.0, 22.0], dtype=float)
        M_L_lbmol = np.array([5.0, 6.0, 7.0], dtype=float)
        M_V_lbmol = np.array([1.0, 1.0, 1.0], dtype=float)
        y0 = np.array([[0.6, 0.4], [0.5, 0.5], [0.4, 0.6]], dtype=float)
        x0 = np.array([[0.8, 0.2], [0.5, 0.5], [0.3, 0.7]], dtype=float)
        top_L0_lbmol = None
        top_V0_lbmol = None
        bottom_L0_lbmol = None
        bottom_V0_lbmol = None
        tray_EL0_BTU = None
        tray_EV0_BTU = None
        controller_state = None
        memory_state = None
        streams = {}
        tau_eq_sec = 4.0

    class TinyCase:
        streams = {}

    class FakeProvider:
        def __init__(self):
            self.calls = []

        def warm_startup_kernels(self, *, density_state=None, flash_rows=None):
            self.calls.append((density_state, flash_rows))
            return {"density_ready": bool(density_state is not None), "flash_ready": bool(flash_rows)}

    from dynamic_distillation.thermo_backend_factory_v1 import ThermoBackendBuildResult
    from dynamic_distillation.thermo_backend_protocol_v1 import get_thermo_backend_capabilities

    provider = FakeProvider()
    monkeypatch.setattr(
        runmod,
        "build_primary_thermo_backend",
        lambda **kwargs: ThermoBackendBuildResult(
            provider=provider,
            thermo_mode="clapeyron",
            dwsim_property_package=None,
            capabilities=get_thermo_backend_capabilities(provider),
        ),
    )
    monkeypatch.setattr(runmod, "build_equilibrium_relaxation_pr_provider", lambda **kwargs: None)
    monkeypatch.setattr(runmod, "_autocalibrate_francis_hydraulic_c_factors_from_seed", lambda **kwargs: False)

    inputs, _ = build_inputs_for_runner(
        TinyCase(),
        TinyCol(),
        RunnerConfig(excel_path="<unit-test>", thermo_mode="clapeyron"),
    )

    assert provider.calls == []
    assert hasattr(inputs, "startup_build_timing_sec")
    assert "primary_backend_prewarm" not in inputs.startup_build_timing_sec
    assert hasattr(inputs, "startup_build_info")
    assert "primary_backend_prewarm" not in inputs.startup_build_info


def test_build_inputs_records_primary_backend_prewarm_timing_when_enabled(monkeypatch):
    class TinyCol:
        n_stages = 3
        n_components = 2
        components_excel = ["A", "B"]
        components_dwsim = ["A", "B"]
        specs_raw = {}
        geometry = None
        T_f = np.array([100.0, 110.0, 120.0], dtype=float)
        P_psia = np.array([200.0, 205.0, 210.0], dtype=float)
        V_lbmolph = np.array([10.0, 11.0, 12.0], dtype=float)
        L_lbmolph = np.array([20.0, 21.0, 22.0], dtype=float)
        M_L_lbmol = np.array([5.0, 6.0, 7.0], dtype=float)
        M_V_lbmol = np.array([1.0, 1.0, 1.0], dtype=float)
        y0 = np.array([[0.6, 0.4], [0.5, 0.5], [0.4, 0.6]], dtype=float)
        x0 = np.array([[0.8, 0.2], [0.5, 0.5], [0.3, 0.7]], dtype=float)
        top_L0_lbmol = None
        top_V0_lbmol = None
        bottom_L0_lbmol = None
        bottom_V0_lbmol = None
        tray_EL0_BTU = None
        tray_EV0_BTU = None
        controller_state = None
        memory_state = None
        streams = {}
        tau_eq_sec = 4.0

    class TinyCase:
        streams = {}

    class FakeProvider:
        def __init__(self):
            self.calls = []

        def warm_startup_kernels(self, *, density_state=None, flash_rows=None):
            self.calls.append((density_state, flash_rows))
            return {"density_ready": bool(density_state is not None), "flash_ready": bool(flash_rows)}

    from dynamic_distillation.thermo_backend_factory_v1 import ThermoBackendBuildResult
    from dynamic_distillation.thermo_backend_protocol_v1 import get_thermo_backend_capabilities

    provider = FakeProvider()
    monkeypatch.setattr(
        runmod,
        "build_primary_thermo_backend",
        lambda **kwargs: ThermoBackendBuildResult(
            provider=provider,
            thermo_mode="clapeyron",
            dwsim_property_package=None,
            capabilities=get_thermo_backend_capabilities(provider),
        ),
    )
    monkeypatch.setattr(runmod, "build_equilibrium_relaxation_pr_provider", lambda **kwargs: None)
    monkeypatch.setattr(runmod, "_autocalibrate_francis_hydraulic_c_factors_from_seed", lambda **kwargs: False)

    inputs, _ = build_inputs_for_runner(
        TinyCase(),
        TinyCol(),
        RunnerConfig(
            excel_path="<unit-test>",
            thermo_mode="clapeyron",
            enable_primary_thermo_startup_prewarm=True,
        ),
    )

    assert len(provider.calls) == 1
    density_state, flash_rows = provider.calls[0]
    assert density_state is not None
    assert flash_rows is not None
    assert len(flash_rows) == 3
    assert hasattr(inputs, "startup_build_timing_sec")
    assert float(inputs.startup_build_timing_sec["primary_backend_prewarm"]) >= 0.0
    assert hasattr(inputs, "startup_build_info")
    assert bool(inputs.startup_build_info["primary_backend_prewarm"]["executed"]) is True


def test_build_inputs_applies_default_thermo_cadence_guardrails_for_hydraulic_runs():
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)

    inputs, _ = build_inputs_for_runner(
        case,
        col,
        RunnerConfig(
            excel_path=str(excel),
            runtime_mode="hydraulic",
            thermo_mode="stub",
            thermo_every_n_steps=5,
            include_temperature=True,
        ),
    )

    assert inputs.thermo_refresh_dT_F == pytest.approx(1.0)
    assert inputs.thermo_refresh_dP_psia is None
    assert inputs.thermo_refresh_dx == pytest.approx(5.0e-3)


def test_build_inputs_can_disable_default_thermo_cadence_guardrails():
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)

    inputs, _ = build_inputs_for_runner(
        case,
        col,
        RunnerConfig(
            excel_path=str(excel),
            runtime_mode="hydraulic",
            thermo_mode="stub",
            thermo_every_n_steps=5,
            include_temperature=True,
            enable_thermo_cadence_guardrails=False,
        ),
    )

    assert inputs.thermo_refresh_dT_F is None
    assert inputs.thermo_refresh_dP_psia is None
    assert inputs.thermo_refresh_dx is None


def test_build_inputs_adds_overhead_and_condenser_vapor_capacitance():
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)

    specs = dict(getattr(col, "specs_raw", {}) or {})
    specs["Top Drum Vapor Volume (ft3)"] = None
    specs["Top Drum Total Volume (ft3)"] = None
    specs["Top Drum Diameter (ft)"] = 10.0
    specs["Top Drum Length (ft)"] = 40.0
    specs["Top Drum Liquid Fraction (-)"] = 0.60
    specs["Overhead Vapor Line Volume (ft3)"] = 56.0
    specs["Condenser Vapor Volume (ft3)"] = 100.0
    object.__setattr__(col, "specs_raw", specs)

    cfg = RunnerConfig(excel_path=str(excel), thermo_mode="stub")
    inputs, _ = build_inputs_for_runner(case, col, cfg)

    total_vol = float(np.pi * 0.25 * 10.0 * 10.0 * 40.0)
    expected_shell_vapor = total_vol * (1.0 - 0.60)
    adders = 56.0 + 100.0
    assert inputs.top_drum_total_volume_ft3 is not None
    assert abs(float(inputs.top_drum_total_volume_ft3) - total_vol) < 1e-9
    assert inputs.top_drum_vapor_volume_ft3 is not None
    assert abs(float(inputs.top_drum_vapor_volume_ft3) - expected_shell_vapor) < 1e-9
    assert inputs.top_drum_extra_vapor_volume_ft3 is not None
    assert abs(float(inputs.top_drum_extra_vapor_volume_ft3) - adders) < 1e-9


def test_build_inputs_infers_top_drum_total_volume_from_holdup_and_vapor_volume():
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)

    specs = dict(getattr(col, "specs_raw", {}) or {})
    specs["Top Drum Total Volume (ft3)"] = None
    specs["Top Drum Liquid Fraction (-)"] = None
    specs["Top Drum Vapor Volume (ft3)"] = 900.0
    specs["Overhead Vapor Line Volume (ft3)"] = None
    specs["Condenser Vapor Volume (ft3)"] = None
    specs["Top Accumulator Holdup (lbmol)"] = 397.0
    object.__setattr__(col, "specs_raw", specs)

    cfg = RunnerConfig(excel_path=str(excel), thermo_mode="stub")
    inputs, _ = build_inputs_for_runner(case, col, cfg)

    assert inputs.top_drum_vapor_volume_ft3 is not None
    assert abs(float(inputs.top_drum_vapor_volume_ft3) - 900.0) < 1e-12
    assert inputs.top_drum_total_volume_ft3 is not None
    assert float(inputs.top_drum_total_volume_ft3) > 900.0


def test_build_inputs_applies_hydraulic_energy_stability_defaults():
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)

    specs = dict(getattr(col, "specs_raw", {}) or {})
    specs["Pressure Model"] = "hydraulic"
    specs["Vapor Flow Model"] = "energy"
    specs["Vapor Holdup Relaxation (sec)"] = None
    specs["Reboiler Neighbor Vapor Hi Ratio"] = None
    specs["Reboiler Neighbor Vapor Lo Ratio"] = None
    specs["Thermo Refresh dT (F)"] = 5.0
    specs["Thermo Refresh dP (psia)"] = 2.0
    specs["Thermo Refresh dX"] = 0.05
    object.__setattr__(col, "specs_raw", specs)

    cfg = RunnerConfig(excel_path=str(excel), thermo_mode="stub")
    inputs, _ = build_inputs_for_runner(case, col, cfg)

    assert str(inputs.pressure_model).lower() == "hydraulic"
    assert str(inputs.vapor_flow_model).lower() == "energy"
    tau_ref = float(getattr(col, "tau_eq_sec", 10.0))
    assert abs(float(inputs.vapor_holdup_relaxation_sec) - tau_ref) < 1e-12
    assert abs(float(inputs.hydraulic_pressure_relaxation_sec) - tau_ref) < 1e-12
    assert abs(float(inputs.reboiler_neighbor_vflow_hi_ratio) - 1.20) < 1e-12
    assert abs(float(inputs.reboiler_neighbor_vflow_lo_ratio) - 0.80) < 1e-12
    assert inputs.thermo_refresh_dT_F is None
    assert inputs.thermo_refresh_dP_psia is None
    assert inputs.thermo_refresh_dx is None


def test_build_inputs_hydraulic_energy_stability_defaults_allow_cfg_overrides():
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)

    specs = dict(getattr(col, "specs_raw", {}) or {})
    specs["Pressure Model"] = "hydraulic"
    specs["Vapor Flow Model"] = "energy"
    object.__setattr__(col, "specs_raw", specs)

    cfg = RunnerConfig(
        excel_path=str(excel),
        thermo_mode="stub",
        vapor_holdup_relaxation_sec=6.0,
        hydraulic_pressure_relaxation_sec=3.0,
        reboiler_neighbor_vflow_hi_ratio=1.05,
        reboiler_neighbor_vflow_lo_ratio=0.95,
        thermo_refresh_dT_F=0.2,
        thermo_refresh_dP_psia=0.3,
        thermo_refresh_dx=1.0e-3,
        enforce_top_pressure_ordering=False,
        top_pressure_ordering_margin_psi=0.15,
    )
    inputs, _ = build_inputs_for_runner(case, col, cfg)

    assert abs(float(inputs.vapor_holdup_relaxation_sec) - 6.0) < 1e-12
    assert abs(float(inputs.hydraulic_pressure_relaxation_sec) - 3.0) < 1e-12
    assert abs(float(inputs.reboiler_neighbor_vflow_hi_ratio) - 1.05) < 1e-12
    assert abs(float(inputs.reboiler_neighbor_vflow_lo_ratio) - 0.95) < 1e-12
    assert abs(float(inputs.thermo_refresh_dT_F) - 0.2) < 1e-12
    assert abs(float(inputs.thermo_refresh_dP_psia) - 0.3) < 1e-12
    assert abs(float(inputs.thermo_refresh_dx) - 1.0e-3) < 1e-12
    assert bool(inputs.enforce_top_pressure_ordering) is False
    assert abs(float(inputs.top_pressure_ordering_margin_psi) - 0.15) < 1e-12


def test_runtime_mode_calibration_forces_parity_closures():
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)

    # Seed conflicting spec-side closures to verify runtime-mode override.
    specs = dict(getattr(col, "specs_raw", {}) or {})
    specs["Pressure Model"] = "hydraulic"
    specs["Vapor Flow Model"] = "energy"
    specs["Enable Liquid Hydraulic Override"] = True
    specs["Liquid Hydraulic Override Alpha"] = 1.0
    object.__setattr__(col, "specs_raw", specs)

    cfg = RunnerConfig(
        excel_path=str(excel),
        thermo_mode="stub",
        runtime_mode="calibration",
        enable_liquid_hydraulic_override=True,
        liquid_hydraulic_override_alpha=1.0,
    )
    inputs, _ = build_inputs_for_runner(case, col, cfg)

    assert str(inputs.pressure_model).lower() == "spec"
    assert str(inputs.vapor_flow_model).lower() == "profile"
    assert bool(inputs.enable_liquid_hydraulic_override) is False
    assert float(inputs.liquid_hydraulic_override_alpha) == pytest.approx(0.0)
    assert str(inputs.equilibrium_relaxation_mode).strip().lower() == "phase-holdup"


def test_runtime_mode_parity_disables_legacy_temperature_state():
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)

    cfg = RunnerConfig(
        excel_path=str(excel),
        thermo_mode="stub",
        runtime_mode="parity",
        include_temperature=True,
    )
    inputs, _ = build_inputs_for_runner(case, col, cfg)

    assert bool(inputs.enable_legacy_temperature_state) is False
    assert bool(inputs.enable_live_total_condenser_duty) is False
    assert bool(inputs.reboiler_equilibrium) is False
    assert str(inputs.reboiler_mode).strip().lower() == "specified"


def test_normalize_runtime_mode_accepts_calibration():
    assert _normalize_runtime_mode("calibration") == "calibration"
    assert _normalize_runtime_mode(" Calibration ") == "calibration"


def test_autocalibrate_francis_hydraulic_c_factors_from_seed_recovers_stage_targets():
    n = 4
    area = np.full(n, 10.0, dtype=float)
    weir_h = np.full(n, 1.0, dtype=float)
    weir_L = np.full(n, 2.0, dtype=float)
    geom = ColumnGeometry(
        sections=[
            ColumnGeometrySection(
                start_stage_1based=1,
                end_stage_1based=n,
                diameter_ft=3.0,
                tray_spacing_ft=1.0,
                gas_void_frac=0.5,
                weir_height_in=1.0,
                weir_length_ft=2.0,
                active_area_frac=1.0,
            )
        ],
        diameter_ft_per_stage=np.full(n, 3.0, dtype=float),
        tray_spacing_ft_per_stage=np.full(n, 1.0, dtype=float),
        gas_void_frac_per_stage=np.full(n, 0.5, dtype=float),
        area_ft2_per_stage=area,
        vapor_volume_ft3_per_stage=np.full(n, 5.0, dtype=float),
        weir_height_in_per_stage=weir_h,
        weir_length_ft_per_stage=weir_L,
        active_area_frac_per_stage=np.ones(n, dtype=float),
        active_area_ft2_per_stage=area,
        hydraulic_c_factor_per_stage=np.ones(n, dtype=float),
    )

    rhoL = np.array([1.0, 1.0, 1.0, 1.0], dtype=float)
    ML = np.array([0.0, 8.0, 10.0, 0.0], dtype=float)
    base = compute_francis_weir_liquid_outflow(
        ML_lbmol=ML,
        rhoL_lbmol_ft3=rhoL,
        active_area_ft2=area,
        weir_height_in=weir_h,
        weir_length_ft=weir_L,
        c_multiplier=None,
    )
    c_target = np.array([1.0, 0.5, 2.0, 1.0], dtype=float)
    L_profile = np.asarray(base.ML_lbmolph, dtype=float) * c_target

    col = ColumnSpec(
        excel_path="<unit-test>",
        components_excel=["A"],
        components_dwsim=["A"],
        n_components=1,
        n_stages=n,
        stage_1based=np.arange(1, n + 1, dtype=int),
        sim=SimulationSettings(dt_sec=1.0, t_final_sec=10.0, log_every_n_steps=1),
        duties=HeatDuties(condenser_type="Total", q_cond_btu_per_h=0.0, q_reb_btu_per_h=0.0),
        specs_raw={"Geometry Sections": [{"start_stage_1based": 1, "end_stage_1based": n}]},
        T_f=np.full(n, 100.0, dtype=float),
        P_psia=np.full(n, 200.0, dtype=float),
        V_lbmolph=np.zeros(n, dtype=float),
        L_lbmolph=L_profile,
        M_L_lbmol=ML,
        M_V_lbmol=np.zeros(n, dtype=float),
        y0=np.ones((n, 1), dtype=float),
        x0=np.ones((n, 1), dtype=float),
        streams={},
        geometry=geom,
    )

    class ConstRhoProvider:
        def liquid_density_lbmol_ft3(self, T_F, P_psia, x):
            return 1.0

    changed = _autocalibrate_francis_hydraulic_c_factors_from_seed(
        col=col,
        thermo_provider=ConstRhoProvider(),
    )

    assert changed is True
    got = np.asarray(col.geometry.hydraulic_c_factor_per_stage, dtype=float)
    assert got[1] == pytest.approx(0.5)
    assert got[2] == pytest.approx(2.0)


def test_autocalibrate_francis_hydraulic_c_factors_from_seed_respects_explicit_spec():
    n = 3
    geom = ColumnGeometry(
        sections=[
            ColumnGeometrySection(
                start_stage_1based=1,
                end_stage_1based=n,
                diameter_ft=3.0,
                tray_spacing_ft=1.0,
                gas_void_frac=0.5,
                weir_height_in=1.0,
                weir_length_ft=2.0,
                active_area_frac=1.0,
                hydraulic_c_factor=0.75,
            )
        ],
        diameter_ft_per_stage=np.full(n, 3.0, dtype=float),
        tray_spacing_ft_per_stage=np.full(n, 1.0, dtype=float),
        gas_void_frac_per_stage=np.full(n, 0.5, dtype=float),
        area_ft2_per_stage=np.full(n, 10.0, dtype=float),
        vapor_volume_ft3_per_stage=np.full(n, 5.0, dtype=float),
        weir_height_in_per_stage=np.full(n, 1.0, dtype=float),
        weir_length_ft_per_stage=np.full(n, 2.0, dtype=float),
        active_area_frac_per_stage=np.ones(n, dtype=float),
        active_area_ft2_per_stage=np.full(n, 10.0, dtype=float),
        hydraulic_c_factor_per_stage=np.full(n, 0.75, dtype=float),
    )
    col = ColumnSpec(
        excel_path="<unit-test>",
        components_excel=["A"],
        components_dwsim=["A"],
        n_components=1,
        n_stages=n,
        stage_1based=np.arange(1, n + 1, dtype=int),
        sim=SimulationSettings(dt_sec=1.0, t_final_sec=10.0, log_every_n_steps=1),
        duties=HeatDuties(condenser_type="Total", q_cond_btu_per_h=0.0, q_reb_btu_per_h=0.0),
        specs_raw={"Geometry Sections": [{"start_stage_1based": 1, "end_stage_1based": n, "hydraulic_c_factor": 0.75}]},
        T_f=np.full(n, 100.0, dtype=float),
        P_psia=np.full(n, 200.0, dtype=float),
        V_lbmolph=np.zeros(n, dtype=float),
        L_lbmolph=np.zeros(n, dtype=float),
        M_L_lbmol=np.array([0.0, 5.0, 0.0], dtype=float),
        M_V_lbmol=np.zeros(n, dtype=float),
        y0=np.ones((n, 1), dtype=float),
        x0=np.ones((n, 1), dtype=float),
        streams={},
        geometry=geom,
    )

    class ConstRhoProvider:
        def liquid_density_lbmol_ft3(self, T_F, P_psia, x):
            return 1.0

    changed = _autocalibrate_francis_hydraulic_c_factors_from_seed(
        col=col,
        thermo_provider=ConstRhoProvider(),
    )

    assert changed is False
    got = np.asarray(col.geometry.hydraulic_c_factor_per_stage, dtype=float)
    assert np.allclose(got, np.full(n, 0.75, dtype=float))


def test_build_inputs_runtime_hydraulic_defaults_equilibrium_mode_to_composition_only():
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)
    cfg = RunnerConfig(
        excel_path=str(excel),
        thermo_mode="stub",
        runtime_mode="hydraulic",
    )
    inputs, _ = build_inputs_for_runner(case, col, cfg)
    assert str(inputs.equilibrium_relaxation_mode).strip().lower() == "composition-only"
    assert bool(inputs.enable_liquid_hydraulic_override) is False
    assert float(inputs.liquid_hydraulic_override_alpha) == pytest.approx(0.0)
    assert inputs.vapor_holdup_relaxation_sec is None
    assert bool(inputs.flash_feed_at_stage_conditions) is False


def test_build_inputs_runtime_hydraulic_keeps_explicit_liquid_hydraulics_opt_out():
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)
    cfg = RunnerConfig(
        excel_path=str(excel),
        thermo_mode="stub",
        runtime_mode="hydraulic",
        enable_liquid_hydraulic_override=False,
    )
    inputs, _ = build_inputs_for_runner(case, col, cfg)
    assert bool(inputs.enable_liquid_hydraulic_override) is False
    assert float(inputs.liquid_hydraulic_override_alpha) == pytest.approx(0.0)


def test_refresh_tray_bubble_targets_uses_cached_state_and_provider(monkeypatch: pytest.MonkeyPatch):
    col = ColumnSpec(
        excel_path="<unit-test>",
        components_excel=["A", "B"],
        components_dwsim=["A", "B"],
        n_components=2,
        n_stages=2,
        stage_1based=np.array([1, 2], dtype=int),
        sim=SimulationSettings(dt_sec=1.0, t_final_sec=2.0, log_every_n_steps=1),
        duties=HeatDuties(condenser_type="Total", q_cond_btu_per_h=0.0, q_reb_btu_per_h=0.0),
        specs_raw={},
        T_f=np.array([110.0, 150.0], dtype=float),
        P_psia=np.array([220.0, 230.0], dtype=float),
        V_lbmolph=np.zeros(2, dtype=float),
        L_lbmolph=np.zeros(2, dtype=float),
        M_L_lbmol=np.ones(2, dtype=float),
        M_V_lbmol=np.ones(2, dtype=float),
        y0=np.array([[0.8, 0.2], [0.2, 0.8]], dtype=float),
        x0=np.array([[0.7, 0.3], [0.3, 0.7]], dtype=float),
        streams={},
        geometry=None,
    )
    layout = StateVectorLayout(
        n_stages=2,
        n_components=2,
        include_top=True,
        include_bottom=True,
        include_vapor=True,
        include_temperature=True,
        include_energy=False,
    )
    sl = layout.slices()
    y = np.zeros(layout.n_states(), dtype=float)
    y[sl["tray_L"]] = np.array([[7.0, 3.0], [3.0, 7.0]], dtype=float).ravel(order="C")
    y[sl["tray_V"]] = np.array([[1.0, 0.5], [0.5, 1.0]], dtype=float).ravel(order="C")
    y[sl["tray_T_f"]] = np.array([111.0, 149.0], dtype=float)

    calls = []

    def _fake_bubble(*, thermo_provider, P_psia, x, T_guess_F, **_kwargs):
        calls.append((float(P_psia), np.asarray(x, dtype=float).copy(), float(T_guess_F)))
        return float(T_guess_F + 10.0), object()

    monkeypatch.setattr("dynamic_distillation.dynamic_run_scaffold_v1._bubble_point_T_F", _fake_bubble)

    targets = _refresh_tray_bubble_targets_F(
        col=col,
        layout=layout,
        y=y,
        thermo_provider=object(),
        P_tray_psia=np.array([221.0, 229.0], dtype=float),
    )

    assert targets is not None
    assert np.allclose(targets, np.array([121.0, 159.0], dtype=float))
    assert len(calls) == 2
    assert calls[0][0] == pytest.approx(221.0)
    assert calls[1][0] == pytest.approx(229.0)
    assert calls[0][2] == pytest.approx(111.0)
    assert calls[1][2] == pytest.approx(149.0)


def test_build_inputs_runtime_hydraulic_keeps_explicit_liquid_hydraulics_opt_in():
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)
    cfg = RunnerConfig(
        excel_path=str(excel),
        thermo_mode="stub",
        runtime_mode="hydraulic",
        enable_liquid_hydraulic_override=True,
        liquid_hydraulic_override_alpha=0.75,
    )
    inputs, _ = build_inputs_for_runner(case, col, cfg)
    assert bool(inputs.enable_liquid_hydraulic_override) is True
    assert float(inputs.liquid_hydraulic_override_alpha) == pytest.approx(0.75)


def test_update_tray_temp_pressure_slope_uses_local_secant_and_fallback():
    out = _update_tray_temp_pressure_slope_F_per_psi(
        prev_slope_F_per_psi=np.array([1.5, 1.5, 1.5], dtype=float),
        prev_T_F=np.array([100.0, 150.0, 200.0], dtype=float),
        curr_T_F=np.array([101.0, 149.0, 201.0], dtype=float),
        prev_P_psia=np.array([200.0, 220.0, 240.0], dtype=float),
        curr_P_psia=np.array([200.5, 221.0, 240.01], dtype=float),
        default_slope_F_per_psi=1.5,
        dp_min_psia=0.05,
        blend_new=0.5,
    )

    assert out is not None
    assert out[0] == pytest.approx(1.75)
    assert out[1] == pytest.approx(0.25)
    assert out[2] == pytest.approx(1.5)


def test_build_inputs_runtime_hydraulic_keeps_explicit_vapor_holdup_relaxation():
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)
    cfg = RunnerConfig(
        excel_path=str(excel),
        thermo_mode="stub",
        runtime_mode="hydraulic",
        vapor_holdup_relaxation_sec=12.0,
    )
    inputs, _ = build_inputs_for_runner(case, col, cfg)
    assert float(inputs.vapor_holdup_relaxation_sec) == pytest.approx(12.0)


def test_build_inputs_runtime_hydraulic_keeps_explicit_feed_flash_opt_in():
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)
    cfg = RunnerConfig(
        excel_path=str(excel),
        thermo_mode="stub",
        runtime_mode="hydraulic",
        flash_feed_at_stage_conditions=True,
    )
    inputs, _ = build_inputs_for_runner(case, col, cfg)
    assert bool(inputs.flash_feed_at_stage_conditions) is True


def test_build_inputs_accepts_equilibrium_mode_override():
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)
    cfg = RunnerConfig(
        excel_path=str(excel),
        thermo_mode="stub",
        runtime_mode="hydraulic",
        equilibrium_relaxation_mode="phase-holdup",
    )
    inputs, _ = build_inputs_for_runner(case, col, cfg)
    assert str(inputs.equilibrium_relaxation_mode).strip().lower() == "phase-holdup"


def test_build_inputs_hydraulic_keeps_legacy_top_drum_pressure_gate_softness():
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)

    cfg = RunnerConfig(
        excel_path=str(excel),
        thermo_mode="stub",
        runtime_mode="hydraulic",
    )
    inputs, _ = build_inputs_for_runner(case, col, cfg)
    assert float(inputs.top_drum_pressure_gate_soft_psi) == pytest.approx(0.25)


def test_build_inputs_accepts_conductance_vapor_flow_model():
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)

    specs = dict(getattr(col, "specs_raw", {}) or {})
    specs["Pressure Model"] = "hydraulic"
    specs["Vapor Flow Model"] = "conductance"
    object.__setattr__(col, "specs_raw", specs)

    cfg = RunnerConfig(excel_path=str(excel), thermo_mode="stub")
    inputs, _ = build_inputs_for_runner(case, col, cfg)

    assert str(inputs.pressure_model).lower() == "hydraulic"
    assert str(inputs.vapor_flow_model).lower() == "conductance"


def test_build_inputs_conductance_nominal_hi_ratio_spec_and_cfg_override():
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel
    from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case

    case = load_case_from_excel(str(excel))
    col = build_column_spec_from_case(case)

    specs = dict(getattr(col, "specs_raw", {}) or {})
    specs["Pressure Model"] = "hydraulic"
    specs["Vapor Flow Model"] = "conductance"
    specs["Conductance Vapor Nominal Hi Ratio"] = 1.3
    object.__setattr__(col, "specs_raw", specs)

    cfg_from_spec = RunnerConfig(excel_path=str(excel), thermo_mode="stub")
    inputs_spec, _ = build_inputs_for_runner(case, col, cfg_from_spec)
    assert abs(float(inputs_spec.conductance_vflow_nominal_hi_ratio) - 1.3) < 1.0e-12

    cfg_override = RunnerConfig(
        excel_path=str(excel),
        thermo_mode="stub",
        conductance_vflow_nominal_hi_ratio=1.15,
    )
    inputs_cfg, _ = build_inputs_for_runner(case, col, cfg_override)
    assert abs(float(inputs_cfg.conductance_vflow_nominal_hi_ratio) - 1.15) < 1.0e-12


def test_bottoms_composition_control_logs_command(tmp_path: Path):
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    cfg = RunnerConfig(
        excel_path=str(excel),
        n_steps=0,
        dt_sec=0.2,
        log_every_n_steps=1,
        include_temperature=True,
        include_energy=False,
        enable_equilibrium_relaxation=True,
        thermo_mode="stub",
        logs_dir=str(tmp_path),
        write_logs=True,
        enable_bottoms_composition_control=True,
        bottoms_composition_component="C5",
        bottoms_composition_sp_molfrac=0.20,
    )

    out = run_smoke_simulation(cfg)
    summary_csv = Path(str(out["summary_csv"]))
    assert summary_csv.exists()

    import csv

    with summary_csv.open("r", encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows, "summary log is empty"

    r0 = rows[0]
    assert "xB_comp_sp" in r0
    assert "xB_comp_pv" in r0
    assert "Boilup_cmd_lbmolph" in r0
    assert np.isfinite(float(r0["xB_comp_sp"]))
    assert np.isfinite(float(r0["Boilup_cmd_lbmolph"]))


def test_run_writes_initial_log_snapshot_before_first_runtime_step(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    def _boom(*args, **kwargs):
        raise RuntimeError("forced step failure")

    monkeypatch.setattr(runmod, "_integrate_one_step", _boom)

    cfg = RunnerConfig(
        excel_path=str(excel),
        n_steps=1,
        dt_sec=0.2,
        log_every_n_steps=5,
        include_temperature=True,
        include_energy=False,
        enable_equilibrium_relaxation=True,
        thermo_mode="stub",
        logs_dir=str(tmp_path),
        write_logs=True,
    )

    with pytest.raises(RuntimeError, match="forced step failure"):
        run_smoke_simulation(cfg)

    summary_files = list(tmp_path.glob("column_summary_*.csv"))
    profile_files = list(tmp_path.glob("column_profile_*.csv"))
    assert summary_files, "summary CSV was not created"
    assert profile_files, "profile CSV was not created"

    import csv

    with summary_files[0].open("r", encoding="utf-8", newline="") as f:
        summary_rows = list(csv.DictReader(f))
    with profile_files[0].open("r", encoding="utf-8", newline="") as f:
        profile_rows = list(csv.DictReader(f))

    assert summary_rows, "summary CSV did not include the initial snapshot row"
    assert profile_rows, "profile CSV did not include the initial snapshot rows"


def test_parity_defers_live_thermo_on_first_visible_runtime_step(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    real_column_rhs = runmod.column_rhs
    captured: dict[str, object] = {}

    def _spy_column_rhs(t_s, y_vec, col, layout, inputs=None):
        label = str(getattr(inputs, "thermo_stage_trace_label", "") or "").strip()
        if label == "runtime_step_0:outer_rhs":
            captured["thermo_provider_is_none"] = getattr(inputs, "thermo_provider", None) is None
            captured["equilibrium_relaxation"] = bool(getattr(inputs, "equilibrium_relaxation", False))
            captured["progress_hook_callable"] = callable(getattr(inputs, "progress_hook", None))
            captured["legacy_temperature_state"] = bool(getattr(inputs, "enable_legacy_temperature_state", True))
        return real_column_rhs(t_s, y_vec, col, layout, inputs=inputs)

    monkeypatch.setattr(runmod, "column_rhs", _spy_column_rhs)

    cfg = RunnerConfig(
        excel_path=str(excel),
        runtime_mode="parity",
        n_steps=1,
        dt_sec=0.1,
        log_every_n_steps=5,
        include_temperature=True,
        include_energy=False,
        enable_equilibrium_relaxation=True,
        thermo_mode="stub",
        logs_dir=str(tmp_path),
        write_logs=False,
        fast_startup=True,
    )

    run_smoke_simulation(cfg)

    assert captured["thermo_provider_is_none"] is True
    assert captured["equilibrium_relaxation"] is False
    assert captured["progress_hook_callable"] is True
    assert captured["legacy_temperature_state"] is False


def test_parity_initial_snapshot_does_not_skip_step_integration(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    excel = Path("distillation_column_template.xlsx")
    if not excel.exists():
        return

    def _boom(*args, **kwargs):
        raise RuntimeError("parity step reached integrator")

    monkeypatch.setattr(runmod, "_integrate_one_step", _boom)

    cfg = RunnerConfig(
        excel_path=str(excel),
        runtime_mode="parity",
        n_steps=1,
        dt_sec=0.2,
        log_every_n_steps=5,
        include_temperature=True,
        include_energy=False,
        enable_equilibrium_relaxation=True,
        thermo_mode="stub",
        logs_dir=str(tmp_path),
        write_logs=True,
        fast_startup=True,
    )

    with pytest.raises(RuntimeError, match="parity step reached integrator"):
        run_smoke_simulation(cfg)
