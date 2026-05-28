#!/usr/bin/env python
"""
Create a copy of an Excel case with explicit tray vapor holdup.

The dynamic model can use pressure-derived tray vapor holdup, but validation
cases often need an independent vapor residence-time inventory so vapor
composition is not an effectively massless transport state. This helper adds
or updates `Vapor Holdup (lbmol)` and writes specs that preserve it.
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any, Optional

import numpy as np
from openpyxl import load_workbook


def _resolve_path(raw: str) -> Path:
    p = Path(str(raw))
    if p.is_absolute():
        return p.resolve()
    return (Path.cwd() / p).resolve()


def _cell_float(value: Any, default: float = 0.0) -> float:
    try:
        v = float(value)
    except Exception:
        return default
    return v if np.isfinite(v) else default


def _find_header(ws, header: str) -> Optional[int]:
    target = str(header).strip().lower()
    for cell in ws[1]:
        if str(cell.value).strip().lower() == target:
            return int(cell.column)
    return None


def _upsert_spec(ws, key: str, value: Any) -> None:
    key_norm = str(key).strip().lower()
    for row in range(1, ws.max_row + 1):
        raw = ws.cell(row, 1).value
        if str(raw).strip().lower() == key_norm:
            ws.cell(row, 2).value = value
            return
    row = ws.max_row + 1
    ws.cell(row, 1).value = key
    ws.cell(row, 2).value = value


def build_case(
    *,
    input_path: Path,
    output_path: Path,
    vapor_residence_sec: float,
    min_liquid_holdup_lbmol: float,
    max_vapor_to_liquid_frac: Optional[float],
    equilibrium_mode: str,
    vapor_holdup_relaxation_sec: float,
) -> Path:
    wb = load_workbook(input_path)
    if "Initial Conditions" not in wb.sheetnames:
        raise ValueError("Workbook is missing 'Initial Conditions' sheet")
    if "Specifications" not in wb.sheetnames:
        raise ValueError("Workbook is missing 'Specifications' sheet")

    ws = wb["Initial Conditions"]
    stage_col = _find_header(ws, "Stage")
    vapor_flow_col = _find_header(ws, "Vapor Flow (lbmol/h)")
    liquid_holdup_col = _find_header(ws, "Liquid Holdup (lbmol)")
    vapor_holdup_col = _find_header(ws, "Vapor Holdup (lbmol)")
    if stage_col is None or vapor_flow_col is None:
        raise ValueError("Initial Conditions must include Stage and Vapor Flow (lbmol/h)")
    if liquid_holdup_col is None:
        raise ValueError("Initial Conditions must include Liquid Holdup (lbmol)")
    if vapor_holdup_col is None:
        vapor_holdup_col = ws.max_column + 1
        ws.cell(1, vapor_holdup_col).value = "Vapor Holdup (lbmol)"

    tau = max(float(vapor_residence_sec), 0.0)
    min_liq = max(float(min_liquid_holdup_lbmol), 0.0)
    frac_cap = None
    if max_vapor_to_liquid_frac is not None:
        frac_cap_try = float(max_vapor_to_liquid_frac)
        if np.isfinite(frac_cap_try) and frac_cap_try > 0.0:
            frac_cap = frac_cap_try

    for row in range(2, ws.max_row + 1):
        stage = _cell_float(ws.cell(row, stage_col).value, default=np.nan)
        vapor_flow = max(_cell_float(ws.cell(row, vapor_flow_col).value, default=0.0), 0.0)
        liquid_holdup = max(_cell_float(ws.cell(row, liquid_holdup_col).value, default=0.0), 0.0)
        mv = vapor_flow * tau / 3600.0
        if liquid_holdup <= min_liq:
            mv = 0.0
        if stage <= 1.0:
            mv = 0.0
        if frac_cap is not None and liquid_holdup > 0.0:
            mv = min(mv, frac_cap * liquid_holdup)
        ws.cell(row, vapor_holdup_col).value = float(mv)

    specs = wb["Specifications"]
    _upsert_spec(specs, "Equilibrium Relaxation Mode", str(equilibrium_mode))
    _upsert_spec(specs, "Vapor Holdup Relaxation (sec)", float(vapor_holdup_relaxation_sec))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main() -> int:
    ap = argparse.ArgumentParser(description="Create an Excel case with explicit preserved tray vapor holdup.")
    ap.add_argument("--input", required=True, dest="input_path")
    ap.add_argument("--output", required=True, dest="output_path")
    ap.add_argument("--vapor-residence-sec", type=float, default=6.0)
    ap.add_argument("--min-liquid-holdup-lbmol", type=float, default=1.0)
    ap.add_argument("--max-vapor-to-liquid-frac", type=float, default=0.5)
    ap.add_argument("--equilibrium-mode", default="composition-only")
    ap.add_argument("--vapor-holdup-relaxation-sec", type=float, default=0.0)
    args = ap.parse_args()

    out = build_case(
        input_path=_resolve_path(args.input_path),
        output_path=_resolve_path(args.output_path),
        vapor_residence_sec=float(args.vapor_residence_sec),
        min_liquid_holdup_lbmol=float(args.min_liquid_holdup_lbmol),
        max_vapor_to_liquid_frac=float(args.max_vapor_to_liquid_frac),
        equilibrium_mode=str(args.equilibrium_mode),
        vapor_holdup_relaxation_sec=float(args.vapor_holdup_relaxation_sec),
    )
    print(out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
