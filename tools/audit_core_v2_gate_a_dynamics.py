#!/usr/bin/env python
"""Run the DD-079 property-free Gate A dynamic-response comparison."""

from __future__ import annotations

import argparse
from dataclasses import asdict, dataclass, replace
import csv
import importlib.util
import json
from pathlib import Path
import sys
from typing import Callable

import numpy as np
from openpyxl import load_workbook
from scipy.integrate import cumulative_trapezoid


ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from dynamic_distillation.core_v2.source_equation_dynamics_v1 import (
    SourceFeedSchedule,
    SourceIntegrationOptions,
    SourceTrajectory,
    external_material_rates,
    integrate_source_trajectory,
    pack_source_state,
    unpack_source_state,
)
from dynamic_distillation.core_v2.source_equation_gate_v1 import (
    BinarySourceColumnSpec,
    evaluate_binary_source_column,
)


@dataclass(frozen=True)
class DynamicCase:
    name: str
    spec: BinarySourceColumnSpec
    initial_packed_state: np.ndarray
    feed_schedule: SourceFeedSchedule | None


def _load_accepted_reference_module():
    path = ROOT / "tools" / "compare_skogestad_dynamic_response.py"
    module_spec = importlib.util.spec_from_file_location(
        "_dd079_accepted_skogestad_reference",
        path,
    )
    if module_spec is None or module_spec.loader is None:
        raise RuntimeError(f"could not load accepted reference at {path}")
    module = importlib.util.module_from_spec(module_spec)
    sys.modules[module_spec.name] = module
    module_spec.loader.exec_module(module)
    return module


def _load_source_profile(workbook_path: Path) -> np.ndarray:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook["Initial Conditions"]
        headers = {
            str(cell.value): index
            for index, cell in enumerate(next(sheet.iter_rows()), start=1)
        }
        rows: list[tuple[int, float]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            source_stage = row[headers["Source Stage"] - 1]
            if source_stage is None:
                continue
            rows.append(
                (
                    int(source_stage),
                    float(
                        row[
                            headers[
                                "Liquid Composition Component 1"
                            ]
                            - 1
                        ]
                    ),
                )
            )
    finally:
        workbook.close()
    return np.asarray([x for _, x in sorted(rows)], dtype=float)


def _reference_case(reference_module, spec: BinarySourceColumnSpec):
    return reference_module.SourceCase(
        n_stages=int(spec.n_stages),
        feed_stage_bottom_based=int(spec.feed_stage_from_bottom),
        alpha=float(spec.relative_volatility),
        taul_min=float(spec.liquid_hydraulic_tau_min),
        f0_kmol_min=float(spec.nominal_feed_kmol_min),
        qf0=float(spec.nominal_feed_liquid_fraction),
        l0_kmol_min=float(
            spec.nominal_rectifying_liquid_kmol_min
        ),
        v0_kmol_min=float(spec.nominal_boilup_kmol_min),
        lambda_k2=float(spec.liquid_vapor_coupling),
        reflux_kmol_min=float(spec.reflux_kmol_min),
        boilup_kmol_min=float(spec.boilup_kmol_min),
        distillate_kmol_min=float(spec.distillate_kmol_min),
        bottoms_kmol_min=float(spec.bottoms_kmol_min),
        feed_kmol_min=float(spec.feed_kmol_min),
        zf=float(spec.feed_light_mole_fraction),
        qf=float(spec.feed_liquid_fraction),
    )


def _reference_augmented_rhs(
    reference_module,
) -> Callable[[float, np.ndarray, BinarySourceColumnSpec], np.ndarray]:
    def rhs(
        time_min: float,
        augmented_state: np.ndarray,
        spec: BinarySourceColumnSpec,
    ) -> np.ndarray:
        n_stages = int(spec.n_stages)
        packed = np.asarray(
            augmented_state[: 2 * n_stages],
            dtype=float,
        )
        x, _ = unpack_source_state(packed, n_stages=n_stages)
        physical = np.asarray(
            reference_module.colamod_rhs_min(
                time_min,
                packed,
                _reference_case(reference_module, spec),
            ),
            dtype=float,
        )
        external_total, external_light = external_material_rates(
            spec,
            x,
        )
        return np.concatenate(
            (
                physical,
                np.asarray(
                    [external_total, external_light],
                    dtype=float,
                ),
            )
        )

    return rhs


def _normalized_max_difference(
    candidate: np.ndarray,
    reference: np.ndarray,
    *,
    floor: float = 1.0e-6,
) -> float:
    candidate_array = np.asarray(candidate, dtype=float)
    reference_array = np.asarray(reference, dtype=float)
    scale = np.maximum(np.abs(reference_array), float(floor))
    return float(
        np.max(np.abs(candidate_array - reference_array) / scale)
    )


def _spec_at_time(
    case: DynamicCase,
    time_min: float,
) -> BinarySourceColumnSpec:
    if case.feed_schedule is None:
        return case.spec
    return replace(
        case.spec,
        feed_kmol_min=case.feed_schedule.feed_at(time_min),
    )


def _conservation_metrics(
    case: DynamicCase,
    trajectory: SourceTrajectory,
) -> dict[str, float]:
    x = trajectory.light_mole_fraction
    holdup = trajectory.liquid_holdup_kmol
    total_inventory = np.sum(holdup, axis=1)
    light_inventory = np.sum(x * holdup, axis=1)
    total_delta = total_inventory - total_inventory[0]
    light_delta = light_inventory - light_inventory[0]
    total_scale = max(abs(float(total_inventory[0])), 1.0)
    light_scale = max(abs(float(light_inventory[0])), 1.0)

    external_total_rate = np.zeros(trajectory.time_min.size, dtype=float)
    external_light_rate = np.zeros(trajectory.time_min.size, dtype=float)
    differential_total_error = np.zeros(
        trajectory.time_min.size,
        dtype=float,
    )
    differential_light_error = np.zeros(
        trajectory.time_min.size,
        dtype=float,
    )
    for index, time_min in enumerate(trajectory.time_min):
        spec = _spec_at_time(case, float(time_min))
        external_total_rate[index], external_light_rate[index] = (
            external_material_rates(spec, x[index])
        )
        evaluation = evaluate_binary_source_column(
            spec,
            light_mole_fraction=x[index],
            liquid_holdup_kmol=holdup[index],
        )
        differential_total_error[index] = (
            np.sum(evaluation.total_holdup_rate_kmol_min)
            - external_total_rate[index]
        )
        differential_light_error[index] = (
            np.sum(evaluation.light_inventory_rate_kmol_min)
            - external_light_rate[index]
        )

    quadrature_total = cumulative_trapezoid(
        external_total_rate,
        trajectory.time_min,
        initial=0.0,
    )
    quadrature_light = cumulative_trapezoid(
        external_light_rate,
        trajectory.time_min,
        initial=0.0,
    )
    return {
        "differential_total_error_max_kmol_min": float(
            np.max(np.abs(differential_total_error))
        ),
        "differential_light_error_max_kmol_min": float(
            np.max(np.abs(differential_light_error))
        ),
        "solver_total_closure_normalized": float(
            np.max(
                np.abs(
                    total_delta
                    - trajectory.cumulative_external_total_kmol
                )
            )
            / total_scale
        ),
        "solver_light_closure_normalized": float(
            np.max(
                np.abs(
                    light_delta
                    - trajectory.cumulative_external_light_kmol
                )
            )
            / light_scale
        ),
        "output_grid_total_quadrature_error_normalized": float(
            np.max(
                np.abs(
                    quadrature_total
                    - trajectory.cumulative_external_total_kmol
                )
            )
            / total_scale
        ),
        "output_grid_light_quadrature_error_normalized": float(
            np.max(
                np.abs(
                    quadrature_light
                    - trajectory.cumulative_external_light_kmol
                )
            )
            / light_scale
        ),
        "initial_total_inventory_kmol": float(total_inventory[0]),
        "final_total_inventory_kmol": float(total_inventory[-1]),
        "initial_light_inventory_kmol": float(light_inventory[0]),
        "final_light_inventory_kmol": float(light_inventory[-1]),
    }


def _case_metrics(
    *,
    case: DynamicCase,
    time_min: np.ndarray,
    reference_rhs,
    primary_options: SourceIntegrationOptions,
    refinement_options: SourceIntegrationOptions,
) -> tuple[dict, SourceTrajectory, SourceTrajectory]:
    v2 = integrate_source_trajectory(
        base_spec=case.spec,
        initial_packed_state=case.initial_packed_state,
        time_min=time_min,
        options=primary_options,
        feed_schedule=case.feed_schedule,
    )
    reference = integrate_source_trajectory(
        base_spec=case.spec,
        initial_packed_state=case.initial_packed_state,
        time_min=time_min,
        options=primary_options,
        feed_schedule=case.feed_schedule,
        augmented_rhs=reference_rhs,
    )
    refined = integrate_source_trajectory(
        base_spec=case.spec,
        initial_packed_state=case.initial_packed_state,
        time_min=time_min,
        options=refinement_options,
        feed_schedule=case.feed_schedule,
    )
    n_stages = int(case.spec.n_stages)
    x_error = np.abs(
        v2.packed_state[:, :n_stages]
        - reference.packed_state[:, :n_stages]
    )
    holdup_error = np.abs(
        v2.packed_state[:, n_stages:]
        - reference.packed_state[:, n_stages:]
    )
    conservation = _conservation_metrics(case, v2)
    x = v2.light_mole_fraction
    holdup = v2.liquid_holdup_kmol
    metrics = {
        "trajectory_parity_normalized_max": (
            _normalized_max_difference(
                v2.packed_state,
                reference.packed_state,
            )
        ),
        "integrator_refinement_normalized_max": (
            _normalized_max_difference(
                v2.packed_state,
                refined.packed_state,
            )
        ),
        "max_abs_composition_difference": float(np.max(x_error)),
        "max_abs_holdup_difference_kmol": float(
            np.max(holdup_error)
        ),
        "minimum_composition": float(np.min(x)),
        "maximum_composition": float(np.max(x)),
        "minimum_holdup_kmol": float(np.min(holdup)),
        "physical_domain_valid": bool(
            np.all(x >= 0.0)
            and np.all(x <= 1.0)
            and np.all(holdup > 0.0)
        ),
        "v2_primary_nfev": int(v2.nfev),
        "reference_primary_nfev": int(reference.nfev),
        "v2_refinement_nfev": int(refined.nfev),
        "safeguard_activated": bool(
            v2.safeguard_activated
            or reference.safeguard_activated
            or refined.safeguard_activated
        ),
        "conservation": conservation,
    }
    return metrics, v2, reference


def _profile_rows(
    *,
    case: DynamicCase,
    v2: SourceTrajectory,
    reference: SourceTrajectory,
    stride: int,
) -> list[dict[str, float | str]]:
    rows: list[dict[str, float | str]] = []
    n_stages = int(case.spec.n_stages)
    feed_index = int(case.spec.feed_stage_from_bottom) - 1
    indices = list(range(0, v2.time_min.size, max(int(stride), 1)))
    if indices[-1] != v2.time_min.size - 1:
        indices.append(v2.time_min.size - 1)
    for index in indices:
        time = float(v2.time_min[index])
        x = v2.light_mole_fraction[index]
        holdup = v2.liquid_holdup_kmol[index]
        x_ref = reference.light_mole_fraction[index]
        holdup_ref = reference.liquid_holdup_kmol[index]
        spec = _spec_at_time(case, time)
        distillate_light = float(spec.distillate_kmol_min) * float(x[-1])
        bottoms_light = float(spec.bottoms_kmol_min) * float(x[0])
        rows.append(
            {
                "case": case.name,
                "time_min": time,
                "feed_kmol_min": float(spec.feed_kmol_min),
                "x_bottom": float(x[0]),
                "x_feed": float(x[feed_index]),
                "x_top": float(x[-1]),
                "M_bottom_kmol": float(holdup[0]),
                "M_feed_kmol": float(holdup[feed_index]),
                "M_top_kmol": float(holdup[-1]),
                "distillate_light_kmol_min": distillate_light,
                "bottoms_light_kmol_min": bottoms_light,
                "total_inventory_kmol": float(np.sum(holdup)),
                "light_inventory_kmol": float(np.sum(x * holdup)),
                "max_abs_x_v2_minus_reference": float(
                    np.max(np.abs(x - x_ref))
                ),
                "max_abs_M_v2_minus_reference_kmol": float(
                    np.max(np.abs(holdup - holdup_ref))
                ),
            }
        )
    return rows


def _render_markdown(report: dict) -> str:
    lines = [
        "# DD-079 Gate A Dynamic-Response Comparison",
        "",
        f"- Classification: `{report['classification']}`",
        f"- Decision: `{report['decision']}`",
        f"- Horizon/output interval: `{report['horizon_min']} / "
        f"{report['output_interval_min']} min`",
        f"- Primary integrator: `{report['primary_integration']['method']}`",
        f"- Refinement integrator: "
        f"`{report['refinement_integration']['method']}`",
        f"- Trajectory parity tolerance: "
        f"`{report['trajectory_parity_tolerance']}`",
        f"- Integration refinement tolerance: "
        f"`{report['integrator_refinement_tolerance']}`",
        f"- Conservation tolerance: "
        f"`{report['conservation_tolerance']}`",
        "",
        "## Cases",
        "",
        "| Case | V2/reference | BDF/refinement | Total closure | "
        "Light closure | Domain |",
        "|---|---:|---:|---:|---:|---:|",
    ]
    for name, case in report["cases"].items():
        conservation = case["conservation"]
        lines.append(
            f"| {name} | "
            f"{case['trajectory_parity_normalized_max']:.3e} | "
            f"{case['integrator_refinement_normalized_max']:.3e} | "
            f"{conservation['solver_total_closure_normalized']:.3e} | "
            f"{conservation['solver_light_closure_normalized']:.3e} | "
            f"{case['physical_domain_valid']} |"
        )
    lines.extend(
        (
            "",
            "## Interpretation",
            "",
            "All comparisons use the same initial state, output grid, exact "
            "feed-event segmentation, and tight solver tolerances. The "
            "independent reference calls the accepted direct Skogestad "
            "translation and does not call the v2 evaluator.",
            "",
            "The conservation gate uses external-balance accumulator states "
            "integrated by the solver. Separate trapezoidal values on the "
            "saved output grid are reported as quadrature diagnostics and "
            "are not confused with differential or solver closure.",
            "",
            "No clipping, projection, holdup floor, profile substitution, "
            "controller, DWSIM property, energy equation, mini8 equation, or "
            "historical trajectory was used.",
            "",
            "## Authorization",
            "",
            report["authorization"],
            "",
        )
    )
    return "\n".join(lines)


def run(
    *,
    workbook_path: Path,
    out_prefix: Path,
    horizon_min: float = 500.0,
    output_interval_min: float = 1.0,
    profile_stride: int = 10,
) -> dict:
    reference_module = _load_accepted_reference_module()
    reference_rhs = _reference_augmented_rhs(reference_module)
    x_source = _load_source_profile(workbook_path)
    base_spec = BinarySourceColumnSpec(n_stages=int(x_source.size))
    base_holdup = np.full(
        int(base_spec.n_stages),
        float(base_spec.nominal_liquid_holdup_kmol),
        dtype=float,
    )
    initial = pack_source_state(x_source, base_holdup)
    stage_index = np.arange(int(base_spec.n_stages), dtype=float)
    perturbed_initial = pack_source_state(
        x_source + 1.0e-3 * np.sin(stage_index),
        base_holdup + 2.0e-3 * np.cos(stage_index),
    )
    cases = (
        DynamicCase(
            name="nominal_profile_drift",
            spec=base_spec,
            initial_packed_state=initial,
            feed_schedule=None,
        ),
        DynamicCase(
            name="feed_plus_1_percent",
            spec=base_spec,
            initial_packed_state=initial,
            feed_schedule=SourceFeedSchedule(
                step_time_min=0.0,
                feed_before_kmol_min=1.0,
                feed_after_kmol_min=1.01,
            ),
        ),
        DynamicCase(
            name="bounded_perturbed_state",
            spec=base_spec,
            initial_packed_state=perturbed_initial,
            feed_schedule=None,
        ),
    )
    time_min = np.arange(
        0.0,
        float(horizon_min) + 0.5 * float(output_interval_min),
        float(output_interval_min),
        dtype=float,
    )
    primary = SourceIntegrationOptions(
        method="BDF",
        rtol=1.0e-10,
        atol=1.0e-12,
        max_step_min=1.0,
    )
    refinement = SourceIntegrationOptions(
        method="Radau",
        rtol=2.0e-11,
        atol=2.0e-13,
        max_step_min=0.5,
    )
    case_reports: dict[str, dict] = {}
    profile_rows: list[dict[str, float | str]] = []
    for case in cases:
        metrics, v2, reference = _case_metrics(
            case=case,
            time_min=time_min,
            reference_rhs=reference_rhs,
            primary_options=primary,
            refinement_options=refinement,
        )
        case_reports[case.name] = metrics
        profile_rows.extend(
            _profile_rows(
                case=case,
                v2=v2,
                reference=reference,
                stride=profile_stride,
            )
        )

    parity_tolerance = 1.0e-9
    refinement_tolerance = 1.0e-7
    conservation_tolerance = 1.0e-10
    parity_pass = all(
        case["trajectory_parity_normalized_max"] < parity_tolerance
        for case in case_reports.values()
    )
    refinement_pass = all(
        case["integrator_refinement_normalized_max"]
        < refinement_tolerance
        for case in case_reports.values()
    )
    conservation_pass = all(
        case["conservation"]["solver_total_closure_normalized"]
        < conservation_tolerance
        and case["conservation"]["solver_light_closure_normalized"]
        < conservation_tolerance
        and case["conservation"][
            "differential_total_error_max_kmol_min"
        ]
        < 1.0e-12
        and case["conservation"][
            "differential_light_error_max_kmol_min"
        ]
        < 1.0e-12
        for case in case_reports.values()
    )
    physical_pass = all(
        case["physical_domain_valid"]
        and not case["safeguard_activated"]
        for case in case_reports.values()
    )
    feed_case = case_reports["feed_plus_1_percent"]
    feed_schedule_pass = (
        cases[1].feed_schedule is not None
        and cases[1].feed_schedule.step_time_min == 0.0
        and cases[1].feed_schedule.feed_before_kmol_min == 1.0
        and cases[1].feed_schedule.feed_after_kmol_min == 1.01
    )
    feed_rows = [
        row
        for row in profile_rows
        if row["case"] == "feed_plus_1_percent"
    ]
    terminal_formula_pass = all(
        abs(
            float(row["distillate_light_kmol_min"])
            - float(base_spec.distillate_kmol_min)
            * float(row["x_top"])
        )
        < 1.0e-14
        and abs(
            float(row["bottoms_light_kmol_min"])
            - float(base_spec.bottoms_kmol_min)
            * float(row["x_bottom"])
        )
        < 1.0e-14
        for row in feed_rows
    )
    terminal_rate_changes = (
        np.ptp(
            [
                float(row["distillate_light_kmol_min"])
                for row in feed_rows
            ]
        )
        > 1.0e-6
        and np.ptp(
            [
                float(row["bottoms_light_kmol_min"])
                for row in feed_rows
            ]
        )
        > 1.0e-6
    )
    terminal_live_pass = bool(
        terminal_formula_pass
        and terminal_rate_changes
        and feed_case["physical_domain_valid"]
    )
    passed = (
        parity_pass
        and refinement_pass
        and conservation_pass
        and physical_pass
        and feed_schedule_pass
        and terminal_live_pass
    )
    try:
        workbook_label = str(workbook_path.resolve().relative_to(ROOT))
    except ValueError:
        workbook_label = str(workbook_path)
    profile_path = out_prefix.with_name(
        f"{out_prefix.name}_profiles"
    ).with_suffix(".csv")
    profile_path.parent.mkdir(parents=True, exist_ok=True)
    with profile_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=list(profile_rows[0]),
        )
        writer.writeheader()
        writer.writerows(profile_rows)

    report = {
        "schema_id": "dd079-core-v2-gate-a-dynamics-v1",
        "classification": (
            "dd079_gate_a_dynamics_passed"
            if passed
            else "dd079_gate_a_dynamics_failed"
        ),
        "decision": (
            "authorize_gate_b_one_volume_energy_property_closure"
            if passed
            else "stop_gate_a_and_fix_dynamic_wrapper"
        ),
        "authorization": (
            "Gate A is complete. Gate B may begin on one representative "
            "mini8 inventory volume with prescribed pressure and live DWSIM "
            "properties. The five-volume Gate C solve remains unauthorized."
            if passed
            else
            "Stop before properties or energy. Correct state mapping, event "
            "timing, conservation, or integration convergence."
        ),
        "source_workbook": workbook_label,
        "horizon_min": float(horizon_min),
        "output_interval_min": float(output_interval_min),
        "trajectory_parity_tolerance": parity_tolerance,
        "integrator_refinement_tolerance": refinement_tolerance,
        "conservation_tolerance": conservation_tolerance,
        "primary_integration": asdict(primary),
        "refinement_integration": asdict(refinement),
        "cases": case_reports,
        "trajectory_parity_pass": parity_pass,
        "integrator_refinement_pass": refinement_pass,
        "conservation_pass": conservation_pass,
        "saved_grid_quadrature_diagnostic_max_normalized": float(
            max(
                max(
                    case["conservation"][
                        "output_grid_total_quadrature_error_normalized"
                    ],
                    case["conservation"][
                        "output_grid_light_quadrature_error_normalized"
                    ],
                )
                for case in case_reports.values()
            )
        ),
        "saved_grid_quadrature_is_acceptance_gate": False,
        "physical_domain_pass": physical_pass,
        "feed_schedule_pass": feed_schedule_pass,
        "terminal_live_composition_withdrawal_pass": terminal_live_pass,
        "profile_csv": str(profile_path.relative_to(ROOT)),
        "dynamic_integration_attempted": True,
        "nonlinear_algebraic_solve_attempted": False,
        "live_property_evaluation_attempted": False,
        "energy_equations_used": False,
        "mini8_physics_used": False,
        "legacy_trajectory_used_as_reference": False,
        "clipping_or_projection_used": False,
    }
    out_prefix.with_suffix(".json").write_text(
        json.dumps(report, indent=2),
        encoding="utf-8",
    )
    out_prefix.with_suffix(".md").write_text(
        _render_markdown(report),
        encoding="utf-8",
    )
    return report


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook",
        type=Path,
        default=ROOT
        / "validation_skogestad_column_a_relative_volatility.xlsx",
    )
    parser.add_argument(
        "--out-prefix",
        type=Path,
        default=ROOT
        / "logs"
        / "dd079_core_v2_gate_a_dynamics_20260718",
    )
    parser.add_argument("--horizon-min", type=float, default=500.0)
    parser.add_argument(
        "--output-interval-min",
        type=float,
        default=1.0,
    )
    parser.add_argument("--profile-stride", type=int, default=10)
    return parser


if __name__ == "__main__":
    args = _parser().parse_args()
    result = run(
        workbook_path=args.workbook,
        out_prefix=args.out_prefix,
        horizon_min=args.horizon_min,
        output_interval_min=args.output_interval_min,
        profile_stride=args.profile_stride,
    )
    print(json.dumps(result, indent=2))
    raise SystemExit(
        0
        if result["classification"] == "dd079_gate_a_dynamics_passed"
        else 2
    )
