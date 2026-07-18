#!/usr/bin/env python
"""Run the DD-078 property-free Gate A source-equation parity audit."""

from __future__ import annotations

import argparse
from dataclasses import replace
import importlib.util
import json
from pathlib import Path
import sys

import numpy as np
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from dynamic_distillation.core_v2.source_equation_gate_v1 import (
    BinarySourceColumnSpec,
    evaluate_binary_source_column,
)


def _load_accepted_reference_module():
    path = ROOT / "tools" / "compare_skogestad_dynamic_response.py"
    module_spec = importlib.util.spec_from_file_location(
        "_accepted_skogestad_reference",
        path,
    )
    if module_spec is None or module_spec.loader is None:
        raise RuntimeError(f"could not load accepted reference at {path}")
    module = importlib.util.module_from_spec(module_spec)
    sys.modules[module_spec.name] = module
    module_spec.loader.exec_module(module)
    return module


def _load_source_profile(workbook_path: Path) -> np.ndarray:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook["Initial Conditions"]
        headers = {
            str(cell.value): index
            for index, cell in enumerate(next(sheet.iter_rows()), start=1)
        }
        rows: list[tuple[int, float]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            source_stage = row[headers["Source Stage"] - 1]
            if source_stage is None:
                continue
            rows.append(
                (
                    int(source_stage),
                    float(
                        row[
                            headers[
                                "Liquid Composition Component 1"
                            ]
                            - 1
                        ]
                    ),
                )
            )
    finally:
        workbook.close()
    return np.asarray([x for _, x in sorted(rows)], dtype=float)


def _core_spec(reference_case) -> BinarySourceColumnSpec:
    return BinarySourceColumnSpec(
        n_stages=int(reference_case.n_stages),
        feed_stage_from_bottom=int(reference_case.feed_stage_bottom_based),
        relative_volatility=float(reference_case.alpha),
        liquid_hydraulic_tau_min=float(reference_case.taul_min),
        nominal_feed_kmol_min=float(reference_case.f0_kmol_min),
        nominal_feed_liquid_fraction=float(reference_case.qf0),
        nominal_rectifying_liquid_kmol_min=float(reference_case.l0_kmol_min),
        nominal_boilup_kmol_min=float(reference_case.v0_kmol_min),
        liquid_vapor_coupling=float(reference_case.lambda_k2),
        reflux_kmol_min=float(reference_case.reflux_kmol_min),
        boilup_kmol_min=float(reference_case.boilup_kmol_min),
        distillate_kmol_min=float(reference_case.distillate_kmol_min),
        bottoms_kmol_min=float(reference_case.bottoms_kmol_min),
        feed_kmol_min=float(reference_case.feed_kmol_min),
        feed_light_mole_fraction=float(reference_case.zf),
        feed_liquid_fraction=float(reference_case.qf),
    )


def _parity_metrics(
    reference_module,
    reference_case,
    x: np.ndarray,
    holdup: np.ndarray,
) -> dict[str, float]:
    core = evaluate_binary_source_column(
        _core_spec(reference_case),
        light_mole_fraction=x,
        liquid_holdup_kmol=holdup,
    )
    state = np.concatenate((x, holdup))
    accepted = np.asarray(
        reference_module.colamod_rhs_min(0.0, state, reference_case),
        dtype=float,
    )
    return {
        "max_abs_parity_error": float(
            np.max(np.abs(core.packed_state_rate - accepted))
        ),
        "max_abs_state_rate_per_min": float(
            np.max(np.abs(core.packed_state_rate))
        ),
        "global_total_balance_error_kmol_min": float(
            np.sum(core.total_holdup_rate_kmol_min)
            - (
                float(reference_case.feed_kmol_min)
                - float(reference_case.distillate_kmol_min)
                - float(reference_case.bottoms_kmol_min)
            )
        ),
        "global_light_balance_error_kmol_min": float(
            np.sum(core.light_inventory_rate_kmol_min)
            - (
                float(reference_case.feed_kmol_min)
                * float(reference_case.zf)
                - float(reference_case.distillate_kmol_min) * float(x[-1])
                - float(reference_case.bottoms_kmol_min) * float(x[0])
            )
        ),
    }


def _render_markdown(report: dict) -> str:
    nominal = report["nominal_profile"]
    feed = report["feed_step"]
    perturbed = report["perturbed_state"]
    return "\n".join(
        (
            "# DD-078 Core V2 Source-Equation Gate",
            "",
            f"- Classification: `{report['classification']}`",
            f"- Decision: `{report['decision']}`",
            f"- Source workbook: `{report['source_workbook']}`",
            f"- Stages: `{report['stage_count']}`",
            f"- Nominal parity error: "
            f"`{nominal['max_abs_parity_error']:.12g}`",
            f"- Published-profile residual: "
            f"`{nominal['max_abs_state_rate_per_min']:.12g} /min`",
            f"- +1% feed parity error: "
            f"`{feed['max_abs_parity_error']:.12g}`",
            f"- Perturbed-state parity error: "
            f"`{perturbed['max_abs_parity_error']:.12g}`",
            f"- Material-conservation gate: "
            f"`{report['material_conservation_pass']}`",
            f"- Residual-parity gate: `{report['residual_parity_pass']}`",
            "",
            "## Scope",
            "",
            "- property-free binary source equations only;",
            "- residual evaluation only;",
            "- no nonlinear solve;",
            "- no dynamic integration;",
            "- no DWSIM or other live property evaluation;",
            "- no clipping, projection, profile forcing, or controller action.",
            "",
            "The tabulated source profile is not machine-exact: its largest "
            "source-equation rate is about 3.7e-8 /min. The gate therefore "
            "uses 1e-7 /min for the published-profile residual while requiring "
            "the new assembly to match the accepted independent translation "
            "to 1e-12.",
            "",
            "## Mini8 Reuse Decision",
            "",
            "- reuse its compact workbook, geometry, feed/terminal data, UV "
            "state-building patterns, and conditioning-audit patterns for "
            "Gates B and C;",
            "- do not use its sampled old-model profile or historical run "
            "trajectory as an independent acceptance reference;",
            "- do not import its clipping, explicit-Euler advance, profile "
            "flow ownership, regularization, or legacy governing balances "
            "into core_v2.",
            "",
            "## Authorization",
            "",
            report["authorization"],
            "",
        )
    )


def run(
    *,
    workbook_path: Path,
    out_prefix: Path,
    parity_tolerance: float = 1.0e-12,
    profile_residual_tolerance: float = 1.0e-7,
) -> dict:
    reference = _load_accepted_reference_module()
    x = _load_source_profile(workbook_path)
    holdup = np.full(x.size, 0.5, dtype=float)
    try:
        workbook_label = str(workbook_path.resolve().relative_to(ROOT))
    except ValueError:
        workbook_label = str(workbook_path)

    nominal_case = reference.SourceCase(feed_kmol_min=1.0)
    feed_case = replace(nominal_case, feed_kmol_min=1.01)
    perturbed_case = replace(
        feed_case,
        qf=0.82,
        lambda_k2=0.11,
    )
    x_perturbed = x + 1.0e-3 * np.sin(np.arange(x.size))
    holdup_perturbed = holdup + 2.0e-3 * np.cos(np.arange(x.size))

    nominal = _parity_metrics(
        reference,
        nominal_case,
        x,
        holdup,
    )
    feed = _parity_metrics(reference, feed_case, x, holdup)
    perturbed = _parity_metrics(
        reference,
        perturbed_case,
        x_perturbed,
        holdup_perturbed,
    )

    parity_pass = all(
        block["max_abs_parity_error"] <= float(parity_tolerance)
        for block in (nominal, feed, perturbed)
    )
    conservation_pass = all(
        abs(block[key]) <= 1.0e-12
        for block in (nominal, feed, perturbed)
        for key in (
            "global_total_balance_error_kmol_min",
            "global_light_balance_error_kmol_min",
        )
    )
    profile_pass = (
        nominal["max_abs_state_rate_per_min"]
        <= float(profile_residual_tolerance)
    )
    passed = parity_pass and conservation_pass and profile_pass

    report = {
        "schema_id": "dd078-core-v2-source-equation-gate-v1",
        "classification": (
            "dd078_source_equation_residual_gate_passed"
            if passed
            else "dd078_source_equation_residual_gate_failed"
        ),
        "decision": (
            "authorize_gate_a_dynamic_integration_increment"
            if passed
            else "stop_before_dynamic_integration"
        ),
        "authorization": (
            "The property-free residual assembly is accepted. A separately "
            "bounded Gate A dynamic-integration comparison may proceed next; "
            "live properties and the five-volume solve remain unauthorized."
            if passed
            else
            "Correct the source-equation residual assembly before adding "
            "integration, properties, or reduced-column physics."
        ),
        "source_workbook": workbook_label,
        "stage_count": int(x.size),
        "parity_tolerance": float(parity_tolerance),
        "profile_residual_tolerance_per_min": float(
            profile_residual_tolerance
        ),
        "nominal_profile": nominal,
        "feed_step": feed,
        "perturbed_state": perturbed,
        "residual_parity_pass": parity_pass,
        "material_conservation_pass": conservation_pass,
        "published_profile_residual_pass": profile_pass,
        "nonlinear_solve_attempted": False,
        "dynamic_integration_attempted": False,
        "live_property_evaluation_attempted": False,
        "mini8_reuse_authorized_for_later_gates": True,
        "mini8_historical_trajectory_is_acceptance_reference": False,
    }
    out_prefix.parent.mkdir(parents=True, exist_ok=True)
    out_prefix.with_suffix(".json").write_text(
        json.dumps(report, indent=2),
        encoding="utf-8",
    )
    out_prefix.with_suffix(".md").write_text(
        _render_markdown(report),
        encoding="utf-8",
    )
    return report


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook",
        type=Path,
        default=ROOT
        / "validation_skogestad_column_a_relative_volatility.xlsx",
    )
    parser.add_argument(
        "--out-prefix",
        type=Path,
        default=ROOT
        / "logs"
        / "dd078_core_v2_source_equation_gate_20260718",
    )
    return parser


if __name__ == "__main__":
    args = _parser().parse_args()
    result = run(
        workbook_path=args.workbook,
        out_prefix=args.out_prefix,
    )
    print(json.dumps(result, indent=2))
    raise SystemExit(
        0
        if result["classification"]
        == "dd078_source_equation_residual_gate_passed"
        else 2
    )
