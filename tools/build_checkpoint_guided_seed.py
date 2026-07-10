#!/usr/bin/env python
"""
Build a seed workbook by blending an Excel seed toward a quiet run profile.

This is a diagnostic bridge between raw steady-state imports and a full
least-squares initializer. It keeps the operation generic: stage rows are
updated by stage number, and only the explicit top/bottom boundary sheets are
handled specially.
"""

from __future__ import annotations

import argparse
import json
import math
from pathlib import Path
import shutil
import sys
from typing import Any, Dict, Iterable, List, Optional

import numpy as np
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]


def _resolve(raw: str | Path) -> Path:
    p = Path(str(raw))
    if p.is_absolute():
        return p.resolve()
    return (PROJECT_ROOT / p).resolve()


def _norm_label(value: Any) -> str:
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())


def _normalize(v: Iterable[float], fallback: Optional[np.ndarray] = None) -> np.ndarray:
    arr = np.asarray(list(v), dtype=float).reshape((-1,))
    arr = np.where(np.isfinite(arr), arr, 0.0)
    arr = np.clip(arr, 0.0, None)
    total = float(np.sum(arr))
    if total > 1.0e-300:
        return arr / total
    if fallback is not None:
        return _normalize(fallback)
    return np.full(arr.size, 1.0 / max(float(arr.size), 1.0), dtype=float)


def _blend_scalar(old: Any, new: Any, blend: float) -> float:
    try:
        old_f = float(old)
    except Exception:
        old_f = math.nan
    try:
        new_f = float(new)
    except Exception:
        new_f = math.nan
    if not math.isfinite(new_f):
        return old_f
    if not math.isfinite(old_f):
        return new_f
    a = float(np.clip(float(blend), 0.0, 1.0))
    return (1.0 - a) * old_f + a * new_f


def _blend_comp(old: np.ndarray, new: np.ndarray, blend: float) -> np.ndarray:
    a = float(np.clip(float(blend), 0.0, 1.0))
    return _normalize((1.0 - a) * _normalize(old) + a * _normalize(new), fallback=old)


def _find_header(ws: Any, header: str) -> tuple[int, int]:
    target = _norm_label(header)
    for row in ws.iter_rows():
        for cell in row:
            if _norm_label(cell.value) == target:
                return int(cell.row), int(cell.column)
    raise ValueError(f"Could not find header {header!r} in sheet {ws.title!r}")


def _find_header_optional(ws: Any, header: str) -> Optional[tuple[int, int]]:
    try:
        return _find_header(ws, header)
    except ValueError:
        return None


def _component_columns(ws: Any, prefix: str, n_components: int) -> List[int]:
    cols: List[int] = []
    for k in range(1, int(n_components) + 1):
        _row, col = _find_header(ws, f"{prefix} Component {k}")
        cols.append(col)
    return cols


def _component_names(wb: Any, n_components: int) -> List[str]:
    if "Specifications" in wb.sheetnames:
        ws = wb["Specifications"]
        for r in range(1, ws.max_row + 1):
            if _norm_label(ws.cell(r, 1).value) == "componentname":
                names = [str(ws.cell(r, k + 2).value or f"Component {k + 1}") for k in range(n_components)]
                if all(name.strip() for name in names):
                    return names
    return [f"Component {k + 1}" for k in range(n_components)]


def _component_key_candidates(prefix: str, name: str) -> List[str]:
    variants = {
        str(name),
        str(name).replace("N-", "n-"),
        str(name).replace("n-", "N-"),
        str(name).replace("-", "_"),
        str(name).replace("N-", "n_"),
        str(name).replace("n-", "N_"),
    }
    out: List[str] = []
    for variant in variants:
        for key in (f"{prefix}_{variant}", f"{prefix}_{variant.replace(' ', '_')}"):
            if key not in out:
                out.append(key)
    return out


def _load_profile_rows(path: Path, time_s: Optional[float]) -> tuple[List[Dict[str, Any]], Dict[str, Any]]:
    import pandas as pd

    df = pd.read_csv(path)
    if "time_s" not in df.columns:
        raise ValueError(f"profile CSV lacks time_s column: {path}")
    if time_s is None:
        t_sel = float(np.nanmax(pd.to_numeric(df["time_s"], errors="coerce").to_numpy(dtype=float)))
    else:
        times = pd.to_numeric(df["time_s"], errors="coerce").to_numpy(dtype=float)
        idx = int(np.nanargmin(np.abs(times - float(time_s))))
        t_sel = float(times[idx])
    at_t = df[np.isclose(pd.to_numeric(df["time_s"], errors="coerce"), t_sel, rtol=0.0, atol=1.0e-6)].copy()
    stage_rows = at_t[at_t.get("node_type", "").eq("stage")].to_dict("records")
    if not stage_rows:
        # Some older logs may not have node_type; fall back to numbered stages.
        at_t["stage_num"] = pd.to_numeric(at_t.get("stage"), errors="coerce")
        stage_rows = at_t[at_t["stage_num"].between(1, 10_000)].to_dict("records")
    summary: Dict[str, Any] = {}
    for row in at_t.to_dict("records"):
        for key, value in row.items():
            if key not in summary and value == value:
                summary[key] = value
    return stage_rows, summary


def _row_component(row: Dict[str, Any], prefix: str, component_names: List[str]) -> Optional[np.ndarray]:
    vals: List[float] = []
    for name in component_names:
        found = None
        for key in _component_key_candidates(prefix, name):
            if key in row:
                found = row[key]
                break
        if found is None:
            return None
        vals.append(float(found))
    return _normalize(vals)


def _summary_component(summary: Dict[str, Any], prefix: str, component_names: List[str]) -> Optional[np.ndarray]:
    vals: List[float] = []
    for name in component_names:
        found = None
        for key in _component_key_candidates(prefix, name):
            if key in summary:
                found = summary[key]
                break
        if found is None:
            return None
        vals.append(float(found))
    return _normalize(vals)


def _stage_map(ws: Any) -> Dict[int, int]:
    hdr = _find_header(ws, "Stage")
    stage_col = hdr[1]
    out: Dict[int, int] = {}
    for r in range(hdr[0] + 1, ws.max_row + 1):
        value = ws.cell(r, stage_col).value
        if value is None:
            continue
        try:
            out[int(value)] = int(r)
        except Exception:
            continue
    return out


def _read_comp(ws: Any, row: int, cols: List[int]) -> np.ndarray:
    return _normalize([float(ws.cell(row, c).value or 0.0) for c in cols])


def _write_comp(ws: Any, row: int, cols: List[int], values: np.ndarray) -> None:
    comp = _normalize(values)
    for c, value in zip(cols, comp):
        ws.cell(row, c).value = float(value)


def _update_initial_conditions(
    wb: Any,
    stage_rows: List[Dict[str, Any]],
    component_names: List[str],
    *,
    composition_blend: float,
    holdup_blend: float,
    flow_blend: float,
) -> Dict[str, int]:
    ws = wb["Initial Conditions"]
    n_components = len(component_names)
    rows_by_stage = _stage_map(ws)
    x_cols = _component_columns(ws, "Liquid Composition", n_components)
    y_cols = _component_columns(ws, "Vapor Composition", n_components)
    liquid_holdup_col = _find_header_optional(ws, "Liquid Holdup (lbmol)")
    vapor_holdup_col = _find_header_optional(ws, "Vapor Holdup (lbmol)")
    liquid_flow_col = _find_header_optional(ws, "Liquid Flow (lbmol/h)")
    vapor_flow_col = _find_header_optional(ws, "Vapor Flow (lbmol/h)")
    counts = {"composition_rows": 0, "liquid_holdup_rows": 0, "vapor_holdup_rows": 0, "flow_rows": 0}

    for ref in stage_rows:
        try:
            stage = int(ref.get("stage"))
        except Exception:
            continue
        row = rows_by_stage.get(stage)
        if row is None:
            continue
        x_ref = _row_component(ref, "x", component_names)
        y_ref = _row_component(ref, "y", component_names)
        if x_ref is not None:
            _write_comp(ws, row, x_cols, _blend_comp(_read_comp(ws, row, x_cols), x_ref, composition_blend))
            counts["composition_rows"] += 1
        if y_ref is not None:
            _write_comp(ws, row, y_cols, _blend_comp(_read_comp(ws, row, y_cols), y_ref, composition_blend))
        if liquid_holdup_col is not None and "ML_lbmol" in ref:
            col = liquid_holdup_col[1]
            ws.cell(row, col).value = _blend_scalar(ws.cell(row, col).value, ref.get("ML_lbmol"), holdup_blend)
            counts["liquid_holdup_rows"] += 1
        if vapor_holdup_col is not None and "MV_lbmol" in ref:
            col = vapor_holdup_col[1]
            ws.cell(row, col).value = _blend_scalar(ws.cell(row, col).value, ref.get("MV_lbmol"), holdup_blend)
            counts["vapor_holdup_rows"] += 1
        flow_updated = False
        if liquid_flow_col is not None and "L_out_used_lbmolph" in ref:
            col = liquid_flow_col[1]
            ws.cell(row, col).value = _blend_scalar(ws.cell(row, col).value, ref.get("L_out_used_lbmolph"), flow_blend)
            flow_updated = True
        if vapor_flow_col is not None and "V_out_lbmolph" in ref:
            col = vapor_flow_col[1]
            ws.cell(row, col).value = _blend_scalar(ws.cell(row, col).value, ref.get("V_out_lbmolph"), flow_blend)
            flow_updated = True
        if flow_updated:
            counts["flow_rows"] += 1
    return counts


def _write_boundary_state(
    wb: Any,
    component_names: List[str],
    summary: Dict[str, Any],
    *,
    boundary_blend: float,
) -> Dict[str, Any]:
    n_components = len(component_names)
    existing: Dict[str, np.ndarray] = {}
    if "Boundary State" in wb.sheetnames:
        ws_old = wb["Boundary State"]
        for r in range(2, ws_old.max_row + 1):
            key = str(ws_old.cell(r, 1).value or "").strip()
            if key:
                existing[key] = np.asarray([float(ws_old.cell(r, k + 2).value or 0.0) for k in range(n_components)], dtype=float)
        del wb["Boundary State"]

    top_total = float(summary.get("Distillate_L_lbmol", np.nan))
    bottom_total = float(summary.get("Bottoms_L_lbmol", np.nan))
    top_x = _summary_component(summary, "x_Distillate", component_names)
    if top_x is None:
        top_x = _summary_component(summary, "Distillate_x", component_names)
    bottom_x = _summary_component(summary, "Bottoms_sump_x", component_names)
    if bottom_x is None:
        bottom_x = _summary_component(summary, "Bottoms_x", component_names)
    if bottom_x is None:
        bottom_x = _summary_component(summary, "x_Bottoms", component_names)

    rows: Dict[str, np.ndarray] = {
        "top_L": existing.get("top_L", np.zeros(n_components)),
        "top_V": existing.get("top_V", np.zeros(n_components)),
        "bottom_L": existing.get("bottom_L", np.zeros(n_components)),
        "bottom_V": existing.get("bottom_V", np.zeros(n_components)),
    }
    if top_x is not None and math.isfinite(top_total) and top_total > 0.0:
        old_total = float(np.sum(rows["top_L"]))
        total = _blend_scalar(old_total, top_total, boundary_blend)
        comp = _blend_comp(rows["top_L"], top_x, boundary_blend)
        rows["top_L"] = float(total) * comp
    if bottom_x is not None and math.isfinite(bottom_total) and bottom_total > 0.0:
        old_total = float(np.sum(rows["bottom_L"]))
        total = _blend_scalar(old_total, bottom_total, boundary_blend)
        comp = _blend_comp(rows["bottom_L"], bottom_x, boundary_blend)
        rows["bottom_L"] = float(total) * comp

    ws = wb.create_sheet("Boundary State")
    for c, value in enumerate(["State", *component_names], start=1):
        ws.cell(1, c).value = value
    for r, key in enumerate(["top_L", "top_V", "bottom_L", "bottom_V"], start=2):
        ws.cell(r, 1).value = key
        arr = np.asarray(rows[key], dtype=float).reshape((n_components,))
        for k, value in enumerate(arr, start=2):
            ws.cell(r, k).value = float(value)
    return {
        "top_L_total": float(np.sum(rows["top_L"])),
        "bottom_L_total": float(np.sum(rows["bottom_L"])),
        "top_L_x": _normalize(rows["top_L"]).tolist(),
        "bottom_L_x": _normalize(rows["bottom_L"]).tolist(),
    }


def main() -> int:
    ap = argparse.ArgumentParser(description="Blend an Excel seed toward a quiet profile/checkpoint-derived run.")
    ap.add_argument("--input", required=True, help="Input seed workbook.")
    ap.add_argument("--profile-csv", required=True, help="Reference column_profile_*.csv from a quiet run.")
    ap.add_argument("--output", required=True, help="Output candidate workbook.")
    ap.add_argument("--time-s", type=float, default=None, help="Reference time; default is latest logged time.")
    ap.add_argument("--composition-blend", type=float, default=1.0)
    ap.add_argument("--holdup-blend", type=float, default=1.0)
    ap.add_argument("--flow-blend", type=float, default=1.0)
    ap.add_argument("--boundary-blend", type=float, default=1.0)
    ap.add_argument("--summary-json", default=None)
    args = ap.parse_args()

    input_path = _resolve(args.input)
    output_path = _resolve(args.output)
    profile_path = _resolve(args.profile_csv)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if input_path.resolve() != output_path.resolve():
        shutil.copy2(input_path, output_path)

    wb = load_workbook(output_path)
    if "Initial Conditions" not in wb.sheetnames:
        raise ValueError("Workbook lacks Initial Conditions sheet.")
    ws_ic = wb["Initial Conditions"]
    n_components = 0
    for row in ws_ic.iter_rows():
        for cell in row:
            label = _norm_label(cell.value)
            if label.startswith("liquidcompositioncomponent"):
                try:
                    n_components = max(n_components, int(label.replace("liquidcompositioncomponent", "")))
                except Exception:
                    pass
    if n_components <= 0:
        raise ValueError("Could not infer component count from Initial Conditions headers.")
    component_names = _component_names(wb, n_components)
    stage_rows, summary = _load_profile_rows(profile_path, args.time_s)
    counts = _update_initial_conditions(
        wb,
        stage_rows,
        component_names,
        composition_blend=float(args.composition_blend),
        holdup_blend=float(args.holdup_blend),
        flow_blend=float(args.flow_blend),
    )
    boundary_info = _write_boundary_state(
        wb,
        component_names,
        summary,
        boundary_blend=float(args.boundary_blend),
    )
    wb.save(output_path)

    report = {
        "input": str(input_path),
        "profile_csv": str(profile_path),
        "output": str(output_path),
        "time_s": float(summary.get("time_s", args.time_s if args.time_s is not None else math.nan)),
        "component_names": component_names,
        "composition_blend": float(args.composition_blend),
        "holdup_blend": float(args.holdup_blend),
        "flow_blend": float(args.flow_blend),
        "boundary_blend": float(args.boundary_blend),
        "update_counts": counts,
        "boundary": boundary_info,
    }
    summary_path = _resolve(args.summary_json) if args.summary_json else output_path.with_suffix(".checkpoint_guided_summary.json")
    summary_path.write_text(json.dumps(report, indent=2, sort_keys=True), encoding="utf-8")
    print(f"Wrote: {output_path}")
    print(f"Summary: {summary_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
