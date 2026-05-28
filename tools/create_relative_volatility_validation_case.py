from __future__ import annotations

from pathlib import Path

import numpy as np
from openpyxl import Workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUT_PATH = PROJECT_ROOT / "validation_relative_volatility_energy_30stage.xlsx"


def _y_from_x(x_light: float, alpha: float) -> tuple[float, float]:
    x = np.array([float(x_light), 1.0 - float(x_light)], dtype=float)
    k = np.array([float(alpha), 1.0], dtype=float)
    y = k * x
    y /= float(np.sum(y))
    return float(y[0]), float(y[1])


def build_workbook(path: Path = OUT_PATH) -> Path:
    n_stages = 30
    alpha = 1.6
    feed_stage = 16
    feed_flow = 120.0
    dist_flow = 60.0
    bottom_flow = 60.0
    reflux = 42.0
    boilup = 102.0
    top_x_light = 0.95
    bottom_x_light = 0.05
    tray_holdup = 5.0
    tray_vapor_holdup = 0.5
    top_accumulator_holdup = 10.0
    bottom_holdup = 10.0

    wb = Workbook()
    ws = wb.active
    ws.title = "Specifications"
    specs = [
        ("Parameter", "Value"),
        ("Number of Stages", n_stages),
        ("Number of Components", 2),
        ("Condenser Type", "Total"),
        ("Component Name", "N-butane", "n-Pentane"),
        ("Thermo Mode", "relative-volatility"),
        ("Relative Volatility", alpha),
        ("Runtime Mode", "parity"),
        ("Include Energy", True),
        ("Condenser Duty Mode", "specified"),
        ("Condenser Duty (Btu/h)", -650000.0),
        ("Reboiler Duty (Btu/h)", 650000.0),
        ("Simulation Length (min)", 5.0),
        ("Timestep (sec)", 0.2),
        ("Log Frequency (timesteps)", 300),
        ("Top Accumulator Holdup (lbmol)", top_accumulator_holdup),
        ("Bottom Holdup (lbmol)", bottom_holdup),
        ("Pressure Model", "static"),
        ("Vapor Flow Model", "profile"),
        ("Stage time constant [tau] (sec)", 10.0),
        ("Vapor Holdup Relaxation (sec)", 0.0),
        ("Equilibrium Relaxation Mode", "composition-only"),
        ("Equilibrium Tau (sec)", 0.5),
    ]
    for r, row in enumerate(specs, start=1):
        for c, value in enumerate(row, start=1):
            ws.cell(r, c).value = value

    ws = wb.create_sheet("Initial Conditions")
    headers = [
        "Stage",
        "Temperature (F)",
        "Pressure (psia)",
        "Vapor Flow (lbmol/h)",
        "Liquid Flow (lbmol/h)",
        "Liquid Holdup (lbmol)",
        "Vapor Holdup (lbmol)",
        "Vapor Composition Component 1",
        "Vapor Composition Component 2",
        "Liquid Composition Component 1",
        "Liquid Composition Component 2",
    ]
    for c, header in enumerate(headers, start=1):
        ws.cell(1, c).value = header

    for stage in range(1, n_stages + 1):
        frac = (stage - 1) / max(n_stages - 1, 1)
        x_light = top_x_light + (bottom_x_light - top_x_light) * frac
        y_light, y_heavy = _y_from_x(x_light, alpha)
        temp_f = 180.0 + (230.0 - 180.0) * frac
        vapor_flow = 0.0 if stage == 1 else boilup
        liquid_flow = reflux if stage < feed_stage else reflux + feed_flow
        holdup = tray_holdup
        row = [
            stage,
            temp_f,
            14.7,
            vapor_flow,
            liquid_flow,
            holdup,
            0.0 if stage == 1 else tray_vapor_holdup,
            y_light,
            y_heavy,
            x_light,
            1.0 - x_light,
        ]
        for c, value in enumerate(row, start=1):
            ws.cell(stage + 1, c).value = value

    ws = wb.create_sheet("Streams")
    rows = [
        ("Stream", "Feed", "Distillate", "Bottom"),
        ("Stage", feed_stage, 1, n_stages),
        ("Pressure (psia)", 14.7, 14.7, 14.7),
        ("Vapour fraction", 0.0, 0.0, 0.0),
        ("Temperature (F)", 205.0, 180.0, 230.0),
        ("Total molar flow (lbmol/h)", feed_flow, dist_flow, bottom_flow),
        ("Mole flows (lbmol/h)", None, None, None),
        ("N-butane", 0.5 * feed_flow, top_x_light * dist_flow, bottom_x_light * bottom_flow),
        ("n-Pentane", 0.5 * feed_flow, (1.0 - top_x_light) * dist_flow, (1.0 - bottom_x_light) * bottom_flow),
    ]
    for r, row in enumerate(rows, start=1):
        for c, value in enumerate(row, start=1):
            ws.cell(r, c).value = value

    ws = wb.create_sheet("Components")
    ws.cell(1, 1).value = "Component Name"
    ws.cell(2, 1).value = "N-butane"
    ws.cell(3, 1).value = "n-Pentane"

    ws = wb.create_sheet("Boundary State")
    boundary_rows = [
        ("State", "N-butane", "n-Pentane"),
        (
            "top_L",
            top_accumulator_holdup * top_x_light,
            top_accumulator_holdup * (1.0 - top_x_light),
        ),
        ("top_V", 0.0, 0.0),
        (
            "bottom_L",
            bottom_holdup * bottom_x_light,
            bottom_holdup * (1.0 - bottom_x_light),
        ),
        ("bottom_V", 0.0, 0.0),
    ]
    for r, row in enumerate(boundary_rows, start=1):
        for c, value in enumerate(row, start=1):
            ws.cell(r, c).value = value

    path = Path(path)
    wb.save(path)
    return path


if __name__ == "__main__":
    print(build_workbook())
