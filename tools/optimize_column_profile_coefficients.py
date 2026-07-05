#!/usr/bin/env python
"""
Smooth profile-coefficient initializer for dynamic column seeds.

This experimental tool starts from a mechanically synchronized workbook and
fits low-order whole-column correction curves instead of independent tray
variables. The goal is to avoid the moving-interface artifacts seen with local
terminal-window optimizers.
"""

from __future__ import annotations

import argparse
from dataclasses import replace
import json
import math
from pathlib import Path
import shutil
import subprocess
import sys
from typing import Any, Dict, List, Optional

import numpy as np
from openpyxl import load_workbook
from scipy.optimize import least_squares

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_PATH = PROJECT_ROOT / "src"
if str(SRC_PATH) not in sys.path:
    sys.path.insert(0, str(SRC_PATH))

from dynamic_distillation.column_rhs_v1 import BoundaryFlows, column_rhs  # noqa: E402
from dynamic_distillation.column_spec_builder_v1 import build_column_spec_from_case  # noqa: E402
from dynamic_distillation.dynamic_run_scaffold_v1 import (  # noqa: E402
    RunnerConfig,
    _clear_initial_tray_vapor_holdup,
    _initialize_vapor_holdup_from_spec_pressure,
    build_inputs_for_runner,
)
from dynamic_distillation.excel_case_loader_v1 import load_case_from_excel  # noqa: E402
from dynamic_distillation.state_vector_layout_v1 import StateVectorLayout  # noqa: E402


def _resolve(raw: str | Path) -> Path:
    p = Path(str(raw))
    if p.is_absolute():
        return p.resolve()
    return (PROJECT_ROOT / p).resolve()


def _normalize(v: np.ndarray, fallback: Optional[np.ndarray] = None) -> np.ndarray:
    arr = np.asarray(v, dtype=float).reshape((-1,))
    arr = np.where(np.isfinite(arr), arr, 0.0)
    arr = np.clip(arr, 0.0, None)
    s = float(np.sum(arr))
    if s > 1.0e-300:
        return arr / s
    if fallback is not None:
        return _normalize(fallback)
    return np.full(arr.size, 1.0 / float(arr.size), dtype=float)


def _softmax(logits: np.ndarray) -> np.ndarray:
    z = np.asarray(logits, dtype=float).reshape((-1,))
    z = z - float(np.max(z))
    e = np.exp(np.clip(z, -60.0, 60.0))
    return e / max(float(np.sum(e)), 1.0e-300)


def _logit(comp: np.ndarray) -> np.ndarray:
    c = _normalize(comp)
    return np.log(np.clip(c, 1.0e-12, 1.0))


def _basis(n: int, degree: int) -> np.ndarray:
    z = np.linspace(-1.0, 1.0, int(n), dtype=float)
    cols = [np.ones_like(z)]
    if int(degree) >= 1:
        cols.append(z)
    if int(degree) >= 2:
        cols.append(2.0 * z * z - 1.0)
    for k in range(3, int(degree) + 1):
        cols.append(z ** k)
    return np.vstack(cols).T


def _find_header(ws: Any, header: str) -> tuple[int, int]:
    target = str(header).strip().lower()
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value).strip().lower() == target:
                return int(cell.row), int(cell.column)
    raise ValueError(f"Could not find header {header!r} in sheet {ws.title!r}")


def _component_columns(ws: Any, prefix: str, n_components: int) -> List[int]:
    cols: List[int] = []
    for k in range(1, int(n_components) + 1):
        _row, col = _find_header(ws, f"{prefix} Component {k}")
        cols.append(col)
    return cols


def _update_spec_value(wb: Any, label: str, value: Any) -> None:
    if "Specifications" not in wb.sheetnames:
        return
    ws = wb["Specifications"]
    target = "".join(ch for ch in str(label).strip().lower() if ch.isalnum())
    for r in range(1, ws.max_row + 1):
        key = "".join(ch for ch in str(ws.cell(r, 1).value).strip().lower() if ch.isalnum())
        if key == target:
            ws.cell(r, 2).value = value
            return
    row = ws.max_row + 1
    ws.cell(row, 1).value = label
    ws.cell(row, 2).value = value


def _build_case_runtime(args: argparse.Namespace) -> tuple[Any, StateVectorLayout, Any, np.ndarray]:
    excel_path = _resolve(args.input)
    case = load_case_from_excel(str(excel_path))
    col = build_column_spec_from_case(case)
    layout = StateVectorLayout(
        n_stages=int(col.n_stages),
        n_components=int(col.n_components),
        include_top=True,
        include_bottom=True,
        include_vapor=True,
        include_temperature=bool(args.include_temperature),
        include_energy=bool(args.include_energy),
    )
    cfg = RunnerConfig(
        excel_path=str(excel_path),
        runtime_mode=str(args.runtime_mode),
        include_temperature=bool(args.include_temperature),
        include_energy=bool(args.include_energy),
        include_boundary_states=True,
        include_vapor_states=True,
        thermo_mode=str(args.thermo),
        clapeyron_model=str(args.clapeyron_model),
        condenser_duty_mode=str(args.condenser_duty_mode),
        condenser_duty_btu_per_h=args.condenser_duty_btuph,
        enable_equilibrium_relaxation=not bool(args.no_equilibrium),
        flash_feed_at_stage_conditions=(False if bool(args.no_flash_feed_at_stage_conditions) else None),
        vapor_holdup_relaxation_sec=float(args.vapor_holdup_relaxation_sec),
        write_logs=False,
    )
    inputs, _provider = build_inputs_for_runner(case, col, cfg)
    inputs = replace(
        inputs,
        feed_stage_flash_prev=None,
        feed_stage_flash_reuse_dT_F=0.0,
        feed_stage_flash_reuse_dP_psia=0.0,
        feed_stage_flash_reuse_dx=0.0,
    )
    y0 = layout.pack_y0(col)
    if not bool(args.use_excel_vapor_holdup):
        y0 = _clear_initial_tray_vapor_holdup(y0, layout)
    y0 = _initialize_vapor_holdup_from_spec_pressure(
        col=col,
        layout=layout,
        y=y0,
        inputs=inputs,
        include_temperature=bool(args.include_temperature),
        preserve_tray_vapor_holdup=bool(args.use_excel_vapor_holdup),
    )
    return col, layout, inputs, np.asarray(y0, dtype=float)


def main() -> int:
    ap = argparse.ArgumentParser(description="Fit smooth whole-column initialization profile coefficients.")
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--degree", type=int, default=2)
    ap.add_argument("--max-nfev", type=int, default=12)
    ap.add_argument("--max-logit-delta", type=float, default=0.08)
    ap.add_argument("--max-temp-shift-f", type=float, default=2.0)
    ap.add_argument("--max-energy-rel-delta", type=float, default=0.08)
    ap.add_argument("--max-boilup-log-delta", type=float, default=0.01)
    ap.add_argument("--max-condenser-duty-rel-delta", type=float, default=0.02)
    ap.add_argument("--denom-floor-lbmol", type=float, default=1.0)
    ap.add_argument("--energy-denom-floor-btu", type=float, default=10000.0)
    ap.add_argument("--total-mass-weight", type=float, default=100.0)
    ap.add_argument("--component-weight", type=float, default=1.0)
    ap.add_argument("--energy-weight", type=float, default=1.0)
    ap.add_argument("--coefficient-penalty", type=float, default=0.08)
    ap.add_argument("--vary-temperature", action="store_true")
    ap.add_argument("--vary-liquid", action="store_true")
    ap.add_argument("--vary-vapor", action="store_true")
    ap.add_argument("--vary-liquid-energy", action="store_true")
    ap.add_argument("--vary-boilup", action="store_true")
    ap.add_argument("--vary-condenser-duty", action="store_true")
    ap.add_argument("--runtime-mode", default="hydraulic")
    ap.add_argument("--thermo", default="clapeyron")
    ap.add_argument("--clapeyron-model", default="PR")
    ap.add_argument("--condenser-duty-mode", default="specified")
    ap.add_argument("--condenser-duty-btuph", type=float, default=None)
    ap.add_argument("--no-equilibrium", action="store_true")
    ap.add_argument("--no-flash-feed-at-stage-conditions", action="store_true")
    ap.add_argument("--vapor-holdup-relaxation-sec", type=float, default=0.0)
    ap.add_argument("--use-excel-vapor-holdup", action="store_true")
    ap.add_argument("--no-temperature", dest="include_temperature", action="store_false")
    ap.add_argument("--include-energy", dest="include_energy", action="store_true")
    ap.set_defaults(include_temperature=True, include_energy=False)
    ap.add_argument("--audit-output-dir", default=None)
    args = ap.parse_args()
    if not any((args.vary_temperature, args.vary_liquid, args.vary_vapor, args.vary_liquid_energy)):
        args.vary_liquid = True
        args.vary_vapor = True
        args.vary_liquid_energy = bool(args.include_energy)

    col, layout, inputs, y_base = _build_case_runtime(args)
    n = int(col.n_stages)
    nc = int(col.n_components)
    deg = max(0, int(args.degree))
    b = _basis(n, deg)
    nb = int(b.shape[1])
    sl = layout.slices()
    u_base = layout.unpack(y_base)
    tray_l_base = np.asarray(u_base["tray_L"], dtype=float).reshape((n, nc))
    tray_v_base = np.asarray(u_base["tray_V"], dtype=float).reshape((n, nc))
    ml = np.sum(tray_l_base, axis=1)
    mv = np.sum(tray_v_base, axis=1)
    x_base = np.asarray(u_base["x_tray"], dtype=float).reshape((n, nc))
    x_fallback = np.asarray(getattr(col, "x0", x_base), dtype=float).reshape((n, nc))
    for i in range(n):
        if ml[i] <= float(layout.epsilon_lbmol):
            x_base[i, :] = _normalize(x_fallback[i, :])
    yv_base = np.asarray(u_base["y_tray"], dtype=float).reshape((n, nc))
    yv_fallback = np.asarray(getattr(col, "y0", yv_base), dtype=float).reshape((n, nc))
    for i in range(n):
        if mv[i] <= float(layout.epsilon_lbmol):
            yv_base[i, :] = _normalize(yv_fallback[i, :])
    t_base = np.asarray(u_base.get("tray_T_f", getattr(col, "T_f", np.zeros(n))), dtype=float).reshape((n,))
    el_base = np.asarray(u_base.get("tray_EL_BTU", np.zeros(n)), dtype=float).reshape((n,))

    condenser_duty_base = args.condenser_duty_btuph
    if condenser_duty_base is None:
        condenser_duty_base = getattr(inputs, "condenser_duty_btu_per_h", None)
    if condenser_duty_base is None:
        condenser_duty_base = 0.0
    condenser_duty_base = float(condenser_duty_base)
    boilup_base = float(np.asarray(getattr(col, "V_lbmolph", np.zeros(n)), dtype=float).reshape((n,))[-1])

    blocks: List[tuple[str, int]] = []
    if bool(args.vary_temperature):
        blocks.extend(("T", k) for k in range(nb))
    if bool(args.vary_liquid):
        blocks.extend(("X", j * nb + k) for j in range(nc) for k in range(nb))
    if bool(args.vary_vapor):
        blocks.extend(("Y", j * nb + k) for j in range(nc) for k in range(nb))
    if bool(args.vary_liquid_energy) and bool(args.include_energy):
        blocks.extend(("EL", k) for k in range(nb))
    if bool(args.vary_boilup):
        blocks.append(("B", 0))
    if bool(args.vary_condenser_duty) and abs(condenser_duty_base) > 1.0e-12:
        blocks.append(("QC", 0))

    z0 = np.zeros(len(blocks), dtype=float)
    lb = np.zeros_like(z0)
    ub = np.zeros_like(z0)
    for idx, (kind, _sub) in enumerate(blocks):
        if kind in ("X", "Y"):
            lb[idx] = -abs(float(args.max_logit_delta))
            ub[idx] = abs(float(args.max_logit_delta))
        elif kind == "T":
            lb[idx] = -abs(float(args.max_temp_shift_f))
            ub[idx] = abs(float(args.max_temp_shift_f))
        elif kind == "EL":
            lb[idx] = -abs(float(args.max_energy_rel_delta))
            ub[idx] = abs(float(args.max_energy_rel_delta))
        elif kind == "B":
            lb[idx] = -abs(float(args.max_boilup_log_delta))
            ub[idx] = abs(float(args.max_boilup_log_delta))
        elif kind == "QC":
            lb[idx] = -abs(float(args.max_condenser_duty_rel_delta))
            ub[idx] = abs(float(args.max_condenser_duty_rel_delta))

    best = {"norm": float("inf"), "z": z0.copy(), "max_rel": float("inf")}
    eval_count = 0

    def _curves(z: np.ndarray) -> Dict[str, np.ndarray]:
        curves: Dict[str, np.ndarray] = {}
        idx = 0
        if bool(args.vary_temperature):
            c = np.asarray(z[idx: idx + nb], dtype=float)
            idx += nb
            curves["T"] = b @ c
        if bool(args.vary_liquid):
            cx = np.asarray(z[idx: idx + nc * nb], dtype=float).reshape((nc, nb))
            idx += nc * nb
            curves["X"] = b @ cx.T
        if bool(args.vary_vapor):
            cy = np.asarray(z[idx: idx + nc * nb], dtype=float).reshape((nc, nb))
            idx += nc * nb
            curves["Y"] = b @ cy.T
        if bool(args.vary_liquid_energy) and bool(args.include_energy):
            ce = np.asarray(z[idx: idx + nb], dtype=float)
            curves["EL"] = b @ ce
        return curves

    def _make_state(z: np.ndarray) -> np.ndarray:
        y_state = y_base.copy()
        curves = _curves(z)
        tray_l = tray_l_base.copy()
        tray_v = tray_v_base.copy()
        if "X" in curves:
            for i in range(n):
                x = _softmax(_logit(x_base[i, :]) + curves["X"][i, :])
                tray_l[i, :] = float(ml[i]) * x
        if "Y" in curves:
            for i in range(n):
                if mv[i] <= float(layout.epsilon_lbmol):
                    tray_v[i, :] = 0.0
                    continue
                yy = _softmax(_logit(yv_base[i, :]) + curves["Y"][i, :])
                tray_v[i, :] = float(mv[i]) * yy
        y_state[sl["tray_L"]] = tray_l.reshape(-1)
        y_state[sl["tray_V"]] = tray_v.reshape(-1)
        if "T" in curves and "tray_T_f" in sl:
            y_state[sl["tray_T_f"]] = (t_base + curves["T"]).reshape(-1)
        if "EL" in curves and "tray_EL_BTU" in sl:
            scale = np.maximum(np.abs(el_base), float(args.energy_denom_floor_btu))
            y_state[sl["tray_EL_BTU"]] = (el_base + curves["EL"] * scale).reshape(-1)
        return y_state

    def _inputs_for(z: np.ndarray) -> Any:
        idx = 0
        for kind, _sub in blocks:
            if kind in ("T", "X", "Y", "EL"):
                idx += 1
                continue
            break
        # Simpler and robust: scan block list for scalar handles.
        boilup = boilup_base
        qcond = condenser_duty_base
        for val, (kind, _sub) in zip(z, blocks):
            if kind == "B":
                boilup = float(boilup_base) * math.exp(float(val))
            elif kind == "QC":
                qcond = float(condenser_duty_base) * (1.0 + float(val))
        boundary = replace(
            getattr(inputs, "boundary"),
            boilup_lbmolph=float(boilup),
        )
        return replace(
            inputs,
            boundary=boundary,
            condenser_duty_mode=str(args.condenser_duty_mode),
            condenser_duty_btu_per_h=(float(qcond) if abs(float(qcond)) > 1.0e-12 else getattr(inputs, "condenser_duty_btu_per_h", None)),
        )

    def residual(z: np.ndarray) -> np.ndarray:
        nonlocal eval_count, best
        eval_count += 1
        y_eval = _make_state(z)
        inputs_eval = _inputs_for(z)
        dydt, _diag = column_rhs(0.0, y_eval, col, layout, inputs=inputs_eval)
        u = layout.unpack(y_eval)
        du = layout.unpack(np.asarray(dydt, dtype=float))
        parts: List[np.ndarray] = []
        max_rel = 0.0
        tray_l_val = np.asarray(u["tray_L"], dtype=float).reshape((n, nc))
        tray_v_val = np.asarray(u["tray_V"], dtype=float).reshape((n, nc))
        tray_l_rate = np.asarray(du["tray_L"], dtype=float).reshape((n, nc))
        tray_v_rate = np.asarray(du["tray_V"], dtype=float).reshape((n, nc))
        total_inv = np.sum(np.abs(tray_l_val), axis=1) + np.sum(np.abs(tray_v_val), axis=1)
        total_l_rel = np.sum(tray_l_rate, axis=1) / np.maximum(np.sum(np.abs(tray_l_val), axis=1) + float(args.denom_floor_lbmol), 1.0e-300)
        total_v_rel = np.sum(tray_v_rate, axis=1) / np.maximum(np.sum(np.abs(tray_v_val), axis=1) + float(args.denom_floor_lbmol), 1.0e-300)
        total_comb_rel = (np.sum(tray_l_rate, axis=1) + np.sum(tray_v_rate, axis=1)) / np.maximum(total_inv + float(args.denom_floor_lbmol), 1.0e-300)
        parts.append(float(args.total_mass_weight) * total_l_rel)
        parts.append(float(args.total_mass_weight) * total_v_rel)
        parts.append(float(args.total_mass_weight) * total_comb_rel)
        for val, rate in ((tray_l_val, tray_l_rate), (tray_v_val, tray_v_rate)):
            rel = rate / np.maximum(np.abs(val) + float(args.denom_floor_lbmol), 1.0e-300)
            parts.append(float(args.component_weight) * rel.reshape(-1))
            max_rel = max(max_rel, float(np.max(np.abs(rel))))
        if bool(args.include_energy) and "tray_EL_BTU" in u and "tray_EL_BTU" in du:
            e_val = np.asarray(u["tray_EL_BTU"], dtype=float).reshape((n,))
            e_rate = np.asarray(du["tray_EL_BTU"], dtype=float).reshape((n,))
            e_rel = e_rate / np.maximum(np.abs(e_val) + float(args.energy_denom_floor_btu), 1.0e-300)
            parts.append(float(args.energy_weight) * e_rel)
            max_rel = max(max_rel, float(np.max(np.abs(e_rel))))
        if float(args.coefficient_penalty) > 0.0:
            parts.append(float(args.coefficient_penalty) * np.asarray(z, dtype=float).reshape(-1))
        r = np.concatenate(parts)
        norm = float(np.linalg.norm(r))
        if norm < float(best["norm"]):
            best = {"norm": norm, "z": np.asarray(z, dtype=float).copy(), "max_rel": max_rel}
        if eval_count == 1 or eval_count % 5 == 0:
            print(f"eval={eval_count} norm={norm:.8g} max_rel={max_rel:.8g} best_max_rel={best['max_rel']:.8g}")
        return r

    result = least_squares(
        residual,
        z0,
        bounds=(lb, ub),
        max_nfev=max(int(args.max_nfev), 1),
        x_scale="jac",
        verbose=1,
    )
    z_best = np.asarray(best["z"], dtype=float)
    y_best = _make_state(z_best)
    u_best = layout.unpack(y_best)
    x_best = np.asarray(u_best["x_tray"], dtype=float).reshape((n, nc))
    yv_best = np.asarray(u_best["y_tray"], dtype=float).reshape((n, nc))
    for i in range(n):
        if ml[i] <= float(layout.epsilon_lbmol):
            x_best[i, :] = _normalize(x_fallback[i, :])
        else:
            x_best[i, :] = _normalize(x_best[i, :], fallback=x_base[i, :])
    for i in range(n):
        if mv[i] <= float(layout.epsilon_lbmol):
            yv_best[i, :] = _normalize(yv_fallback[i, :])
        else:
            yv_best[i, :] = _normalize(yv_best[i, :], fallback=yv_base[i, :])
    t_best = np.asarray(u_best.get("tray_T_f", t_base), dtype=float).reshape((n,))
    el_best = np.asarray(u_best.get("tray_EL_BTU", el_base), dtype=float).reshape((n,))

    input_path = _resolve(args.input)
    output_path = _resolve(args.output)
    if input_path.resolve() != output_path.resolve():
        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(input_path, output_path)
    wb = load_workbook(output_path)
    ws = wb["Initial Conditions"]
    stage_hdr_row, stage_col = _find_header(ws, "Stage")
    _row, temp_col = _find_header(ws, "Temperature (F)")
    x_cols = _component_columns(ws, "Liquid Composition", nc)
    y_cols = _component_columns(ws, "Vapor Composition", nc)
    stage_to_row: Dict[int, int] = {}
    for row in range(stage_hdr_row + 1, ws.max_row + 1):
        val = ws.cell(row, stage_col).value
        if val is None:
            continue
        try:
            stage_to_row[int(float(val))] = int(row)
        except Exception:
            pass
    for i in range(n):
        row = stage_to_row.get(i + 1)
        if row is None:
            continue
        ws.cell(row, temp_col).value = float(t_best[i])
        for k, col_idx in enumerate(x_cols):
            ws.cell(row, col_idx).value = float(x_best[i, k])
        for k, col_idx in enumerate(y_cols):
            ws.cell(row, col_idx).value = float(yv_best[i, k])
    if bool(args.include_energy) and "Energy State" in wb.sheetnames:
        ws_e = wb["Energy State"]
        energy_stage_hdr_row, energy_stage_col = _find_header(ws_e, "Stage")
        _row, el_col = _find_header(ws_e, "Tray EL (BTU)")
        e_stage_to_row: Dict[int, int] = {}
        for row in range(energy_stage_hdr_row + 1, ws_e.max_row + 1):
            val = ws_e.cell(row, energy_stage_col).value
            if val is None:
                continue
            try:
                e_stage_to_row[int(float(val))] = int(row)
            except Exception:
                pass
        for i in range(n):
            row = e_stage_to_row.get(i + 1)
            if row is not None:
                ws_e.cell(row, el_col).value = float(el_best[i])
    inputs_best = _inputs_for(z_best)
    boundary_best = getattr(inputs_best, "boundary")
    if bool(args.vary_boilup):
        _row, vapor_flow_col = _find_header(ws, "Vapor Flow (lbmol/h)")
        row = stage_to_row.get(n)
        if row is not None:
            ws.cell(row, vapor_flow_col).value = float(boundary_best.boilup_lbmolph)
    if bool(args.vary_condenser_duty):
        _update_spec_value(wb, "Condenser Duty Mode", str(args.condenser_duty_mode))
        _update_spec_value(wb, "Condenser Duty (Btu/h)", float(inputs_best.condenser_duty_btu_per_h))
    wb.save(output_path)

    audit_dir = args.audit_output_dir
    if audit_dir is None:
        audit_dir = str(output_path.with_suffix("")) + "_profile_coeff_audit"
    audit_dir_path = _resolve(audit_dir)
    audit_cmd = [
        sys.executable,
        str(PROJECT_ROOT / "tools" / "column_initialization_residual_audit.py"),
        "--excel",
        str(output_path),
        "--thermo",
        str(args.thermo),
        "--clapeyron-model",
        str(args.clapeyron_model),
        "--runtime-mode",
        str(args.runtime_mode),
        "--condenser-duty-mode",
        str(args.condenser_duty_mode),
        "--vapor-holdup-relaxation-sec",
        str(float(args.vapor_holdup_relaxation_sec)),
        "--output-dir",
        str(audit_dir_path),
    ]
    if args.condenser_duty_btuph is not None or bool(args.vary_condenser_duty):
        audit_cmd.extend(["--condenser-duty-btuph", str(float(inputs_best.condenser_duty_btu_per_h or 0.0))])
    if bool(args.no_equilibrium):
        audit_cmd.append("--no-equilibrium")
    if bool(args.no_flash_feed_at_stage_conditions):
        audit_cmd.append("--no-flash-feed-at-stage-conditions")
    if bool(args.include_energy):
        audit_cmd.append("--include-energy")
    if bool(args.use_excel_vapor_holdup):
        audit_cmd.append("--use-excel-vapor-holdup")
    if not bool(args.include_temperature):
        audit_cmd.append("--no-temperature")
    subprocess.run(audit_cmd, cwd=str(PROJECT_ROOT), check=True)

    curves_best = _curves(z_best)
    summary = {
        "input": str(input_path),
        "output": str(output_path),
        "audit_dir": str(audit_dir_path),
        "degree": deg,
        "n_variables": int(len(blocks)),
        "blocks": blocks,
        "eval_count": eval_count,
        "least_squares_success": bool(result.success),
        "least_squares_message": str(result.message),
        "best_objective_norm": float(best["norm"]),
        "best_max_relative_rate_per_s_inprocess": float(best["max_rel"]),
        "max_abs_delta_x": float(np.max(np.abs(x_best - x_base))),
        "max_abs_delta_y": float(np.max(np.abs(yv_best - yv_base))),
        "max_abs_delta_T_F": float(np.max(np.abs(t_best - t_base))),
        "max_rel_delta_EL": float(np.max(np.abs(el_best - el_base) / np.maximum(np.abs(el_base), float(args.energy_denom_floor_btu)))),
        "boilup_base_lbmolph": boilup_base,
        "boilup_opt_lbmolph": float(boundary_best.boilup_lbmolph or boilup_base),
        "condenser_duty_base_BTUph": condenser_duty_base,
        "condenser_duty_opt_BTUph": float(inputs_best.condenser_duty_btu_per_h or condenser_duty_base),
        "max_abs_curve_T_F": float(np.max(np.abs(curves_best.get("T", np.zeros(n))))),
        "max_abs_curve_X_logit": float(np.max(np.abs(curves_best.get("X", np.zeros((n, nc)))))),
        "max_abs_curve_Y_logit": float(np.max(np.abs(curves_best.get("Y", np.zeros((n, nc)))))),
        "max_abs_curve_EL_rel": float(np.max(np.abs(curves_best.get("EL", np.zeros(n))))),
    }
    summary_path = output_path.with_suffix(".profile_coeff_summary.json")
    summary_path.write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(f"Wrote: {output_path}")
    print(f"Summary: {summary_path}")
    print(f"Audit: {audit_dir_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
