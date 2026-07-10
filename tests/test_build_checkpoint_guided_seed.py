from importlib.util import module_from_spec, spec_from_file_location
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook


_TOOL_PATH = Path(__file__).resolve().parents[1] / "tools" / "build_checkpoint_guided_seed.py"
_SPEC = spec_from_file_location("build_checkpoint_guided_seed", _TOOL_PATH)
assert _SPEC is not None
assert _SPEC.loader is not None
_MODULE = module_from_spec(_SPEC)
_SPEC.loader.exec_module(_MODULE)


def _make_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Specifications"
    ws.append(["Parameter", "Value", None])
    ws.append(["Component Name", "A", "B"])
    ws = wb.create_sheet("Initial Conditions")
    ws.append(
        [
            "Stage",
            "Liquid Flow (lbmol/h)",
            "Vapor Flow (lbmol/h)",
            "Liquid Holdup (lbmol)",
            "Liquid Composition Component 1",
            "Liquid Composition Component 2",
            "Vapor Composition Component 1",
            "Vapor Composition Component 2",
        ]
    )
    ws.append([1, 100.0, 200.0, 10.0, 0.8, 0.2, 0.7, 0.3])
    ws.append([2, 110.0, 210.0, 11.0, 0.3, 0.7, 0.4, 0.6])
    ws = wb.create_sheet("Boundary State")
    ws.append(["State", "A", "B"])
    ws.append(["top_L", 8.0, 2.0])
    ws.append(["top_V", 1.0, 0.0])
    ws.append(["bottom_L", 3.0, 7.0])
    ws.append(["bottom_V", 0.0, 1.0])
    wb.save(path)


def _make_hyphenated_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Specifications"
    ws.append(["Parameter", "Value", None])
    ws.append(["Component Name", "n-Propane", "n-Butane"])
    ws = wb.create_sheet("Initial Conditions")
    ws.append(
        [
            "Stage",
            "Liquid Composition Component 1",
            "Liquid Composition Component 2",
            "Vapor Composition Component 1",
            "Vapor Composition Component 2",
        ]
    )
    ws.append([1, 0.8, 0.2, 0.7, 0.3])
    ws = wb.create_sheet("Boundary State")
    ws.append(["State", "n-Propane", "n-Butane"])
    ws.append(["top_L", 8.0, 2.0])
    ws.append(["top_V", 1.0, 0.0])
    ws.append(["bottom_L", 3.0, 7.0])
    ws.append(["bottom_V", 0.0, 1.0])
    wb.save(path)


def test_build_checkpoint_guided_seed_blends_stage_and_boundary_state(tmp_path):
    input_path = tmp_path / "seed.xlsx"
    output_path = tmp_path / "out.xlsx"
    profile_path = tmp_path / "profile.csv"
    _make_workbook(input_path)
    pd.DataFrame(
        [
            {
                "time_s": 10.0,
                "stage": 1,
                "node_type": "stage",
                "ML_lbmol": 20.0,
                "L_out_used_lbmolph": 120.0,
                "V_out_lbmolph": 220.0,
                "x_A": 0.6,
                "x_B": 0.4,
                "y_A": 0.5,
                "y_B": 0.5,
                "Distillate_L_lbmol": 30.0,
                "Bottoms_L_lbmol": 40.0,
                "x_Distillate_A": 0.25,
                "x_Distillate_B": 0.75,
                "Bottoms_sump_x_A": 0.1,
                "Bottoms_sump_x_B": 0.9,
            },
            {
                "time_s": 10.0,
                "stage": 2,
                "node_type": "stage",
                "ML_lbmol": 22.0,
                "L_out_used_lbmolph": 130.0,
                "V_out_lbmolph": 230.0,
                "x_A": 0.2,
                "x_B": 0.8,
                "y_A": 0.3,
                "y_B": 0.7,
                "Distillate_L_lbmol": 30.0,
                "Bottoms_L_lbmol": 40.0,
                "x_Distillate_A": 0.25,
                "x_Distillate_B": 0.75,
                "Bottoms_sump_x_A": 0.1,
                "Bottoms_sump_x_B": 0.9,
            },
        ]
    ).to_csv(profile_path, index=False)

    rc = _MODULE.main.__wrapped__ if hasattr(_MODULE.main, "__wrapped__") else None
    assert rc is None
    import sys

    old_argv = sys.argv
    try:
        sys.argv = [
            "build_checkpoint_guided_seed.py",
            "--input",
            str(input_path),
            "--profile-csv",
            str(profile_path),
            "--output",
            str(output_path),
            "--time-s",
            "10",
            "--composition-blend",
            "1",
            "--holdup-blend",
            "0.5",
            "--flow-blend",
            "1",
            "--boundary-blend",
            "1",
        ]
        assert _MODULE.main() == 0
    finally:
        sys.argv = old_argv

    wb = load_workbook(output_path, data_only=True)
    ws = wb["Initial Conditions"]
    assert ws.cell(2, 4).value == 15.0
    assert ws.cell(2, 2).value == 120.0
    assert ws.cell(2, 5).value == 0.6
    assert ws.cell(2, 7).value == 0.5
    assert ws.cell(3, 4).value == 16.5
    ws = wb["Boundary State"]
    assert ws.cell(2, 2).value == 7.5
    assert ws.cell(2, 3).value == 22.5
    assert ws.cell(4, 2).value == 4.0
    assert ws.cell(4, 3).value == 36.0


def test_build_checkpoint_guided_seed_matches_sanitized_component_names(tmp_path):
    input_path = tmp_path / "seed.xlsx"
    output_path = tmp_path / "out.xlsx"
    profile_path = tmp_path / "profile.csv"
    _make_hyphenated_workbook(input_path)
    pd.DataFrame(
        [
            {
                "time_s": 10.0,
                "stage": 1,
                "node_type": "stage",
                "x_n_Propane": 0.55,
                "x_n_Butane": 0.45,
                "y_n_Propane": 0.65,
                "y_n_Butane": 0.35,
                "Distillate_L_lbmol": 20.0,
                "Distillate_x_n_Propane": 0.6,
                "Distillate_x_n_Butane": 0.4,
                "Bottoms_L_lbmol": 30.0,
                "Bottoms_sump_x_n_Propane": 0.2,
                "Bottoms_sump_x_n_Butane": 0.8,
            },
        ]
    ).to_csv(profile_path, index=False)

    import sys

    old_argv = sys.argv
    try:
        sys.argv = [
            "build_checkpoint_guided_seed.py",
            "--input",
            str(input_path),
            "--profile-csv",
            str(profile_path),
            "--output",
            str(output_path),
            "--time-s",
            "10",
        ]
        assert _MODULE.main() == 0
    finally:
        sys.argv = old_argv

    wb = load_workbook(output_path, data_only=True)
    ws = wb["Initial Conditions"]
    assert ws.cell(2, 2).value == 0.55
    assert ws.cell(2, 4).value == 0.65
    ws = wb["Boundary State"]
    assert ws.cell(2, 2).value == 12.0
    assert ws.cell(2, 3).value == 8.0
    assert ws.cell(4, 2).value == 6.0
    assert ws.cell(4, 3).value == 24.0
