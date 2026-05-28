#!/usr/bin/env python
"""Update the Gani Excel workbook with converged ChemSep seed profiles."""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import shutil
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
RESULTS_DIR = PROJECT_ROOT / "logs" / "gani_1986_chemsep_rb_seed_results"
EXCEL_PATH = PROJECT_ROOT / "validation_gani_1986_debutanizer.xlsx"

K_TO_F_SCALE = 9.0 / 5.0
K_TO_F_OFFSET = -459.67
PA_TO_PSIA = 0.00014503773773
KMOLH_TO_LBMOLH = 2.20462262185
W_TO_BTUH = 3.41214163312794


def _tag() -> str:
    return dt.datetime.now().strftime("%Y%m%d_%H%M%S")


def _k_to_f(value_k: float) -> float:
    return float(value_k) * K_TO_F_SCALE + K_TO_F_OFFSET


def _pa_to_psia(value_pa: float) -> float:
    return float(value_pa) * PA_TO_PSIA


def _kmolh_to_lbmolh(value_kmolh: float) -> float:
    return float(value_kmolh) * KMOLH_TO_LBMOLH


def _read_csv_rows(path: Path) -> list[dict[str, float]]:
    with path.open("r", newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    out: list[dict[str, float]] = []
    for row in rows:
        out.append({k: float(v) for k, v in row.items()})
    return out


def _normalize(values: Iterable[float]) -> list[float]:
    vals = [max(float(v), 0.0) for v in values]
    total = sum(vals)
    if total <= 0.0:
        raise ValueError("Cannot normalize non-positive composition.")
    return [v / total for v in vals]


def _find_row(ws, label: str) -> int:
    target = str(label).strip().lower()
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, 1).value
        if value is not None and str(value).strip().lower() == target:
            return row
    raise KeyError(f"Could not find row label {label!r} in {ws.title}")


def _set_or_append_spec(ws, label: str, value) -> None:
    try:
        row = _find_row(ws, label)
    except KeyError:
        row = ws.max_row + 1
        ws.cell(row, 1).value = label
    ws.cell(row, 2).value = value


def _append_note(wb, field: str, value: str) -> None:
    if "Notes" not in wb.sheetnames:
        ws = wb.create_sheet("Notes")
        ws.cell(1, 1).value = "Field"
        ws.cell(1, 2).value = "Value"
    ws = wb["Notes"]
    row = ws.max_row + 1
    ws.cell(row, 1).value = field
    ws.cell(row, 2).value = value


def _component_names(wb) -> list[str]:
    if "Components" in wb.sheetnames:
        ws = wb["Components"]
        names = []
        for row in range(2, ws.max_row + 1):
            value = ws.cell(row, 1).value
            if value:
                names.append(str(value))
        if names:
            return names
    ws = wb["Specifications"]
    row = _find_row(ws, "Component Name")
    names = []
    col = 2
    while ws.cell(row, col).value:
        names.append(str(ws.cell(row, col).value))
        col += 1
    return names


def main() -> int:
    ap = argparse.ArgumentParser(description="Apply ChemSep RB-seed profiles to Gani Excel initial conditions.")
    ap.add_argument("--excel", default=str(EXCEL_PATH))
    ap.add_argument("--results-dir", default=str(RESULTS_DIR))
    ap.add_argument("--backup-dir", default="logs")
    ap.add_argument("--no-backup", action="store_true")
    args = ap.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.is_absolute():
        excel_path = (PROJECT_ROOT / excel_path).resolve()
    results_dir = Path(args.results_dir)
    if not results_dir.is_absolute():
        results_dir = (PROJECT_ROOT / results_dir).resolve()

    profiles = _read_csv_rows(results_dir / "profiles.csv")
    liquid_x = _read_csv_rows(results_dir / "liquid_x.csv")
    vapor_y = _read_csv_rows(results_dir / "vapor_y.csv")
    summary = json.loads((results_dir / "summary.json").read_text(encoding="utf-8"))

    if len(profiles) != len(liquid_x) or len(profiles) != len(vapor_y):
        raise ValueError("ChemSep result row counts do not match.")

    backup_path = None
    if not args.no_backup:
        backup_dir = Path(args.backup_dir)
        if not backup_dir.is_absolute():
            backup_dir = (PROJECT_ROOT / backup_dir).resolve()
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup_path = backup_dir / f"{excel_path.stem}__before_chemsep_seed_{_tag()}{excel_path.suffix}"
        shutil.copy2(excel_path, backup_path)

    wb = load_workbook(excel_path)
    component_names = _component_names(wb)
    n_components = len(component_names)

    ws = wb["Initial Conditions"]
    vapor_start_col = 8
    liquid_start_col = vapor_start_col + n_components
    for idx, profile in enumerate(profiles):
        stage = int(profile["stage"])
        row = stage + 1
        x = _normalize(liquid_x[idx][name] for name in component_names)
        y = _normalize(vapor_y[idx][name] for name in component_names)
        ws.cell(row, 1).value = stage
        ws.cell(row, 2).value = _k_to_f(profile["temperature_K"])
        ws.cell(row, 3).value = _pa_to_psia(profile["pressure_Pa"])
        ws.cell(row, 4).value = _kmolh_to_lbmolh(profile["vapor_flow_kmol_h"])
        ws.cell(row, 5).value = _kmolh_to_lbmolh(profile["liquid_flow_kmol_h"])
        for k, value in enumerate(y):
            ws.cell(row, vapor_start_col + k).value = value
        for k, value in enumerate(x):
            ws.cell(row, liquid_start_col + k).value = value

    if "Streams" in wb.sheetnames:
        ws = wb["Streams"]
        temp_row = _find_row(ws, "Temperature (F)")
        total_row = _find_row(ws, "Total molar flow (lbmol/h)")
        pressure_row = _find_row(ws, "Pressure (psia)")
        mf_row = _find_row(ws, "Mole flows (lbmol/h)")

        top_x = [float(summary["top_composition"][name]) for name in component_names]
        bottom_x = [float(summary["bottom_composition"][name]) for name in component_names]
        top_flow_lbmolh = _kmolh_to_lbmolh(summary["top_flow_kmol_h"])
        bottom_flow_lbmolh = _kmolh_to_lbmolh(summary["bottom_flow_kmol_h"])
        ws.cell(temp_row, 3).value = _k_to_f(summary["top_temperature_K"])
        ws.cell(temp_row, 4).value = _k_to_f(summary["bottom_temperature_K"])
        ws.cell(pressure_row, 3).value = _pa_to_psia(profiles[0]["pressure_Pa"])
        ws.cell(pressure_row, 4).value = _pa_to_psia(profiles[-1]["pressure_Pa"])
        ws.cell(total_row, 3).value = top_flow_lbmolh
        ws.cell(total_row, 4).value = bottom_flow_lbmolh
        for k, name in enumerate(component_names, start=1):
            row = mf_row + k
            ws.cell(row, 1).value = name
            ws.cell(row, 3).value = top_flow_lbmolh * top_x[k - 1]
            ws.cell(row, 4).value = bottom_flow_lbmolh * bottom_x[k - 1]

    if "Boundary State" in wb.sheetnames:
        ws = wb["Boundary State"]
        top_row = _find_row(ws, "top_L")
        bottom_row = _find_row(ws, "bottom_L")
        top_total = sum(float(ws.cell(top_row, c).value or 0.0) for c in range(2, 2 + n_components))
        bottom_total = sum(float(ws.cell(bottom_row, c).value or 0.0) for c in range(2, 2 + n_components))
        top_x = _normalize(summary["top_composition"][name] for name in component_names)
        bottom_x = _normalize(summary["bottom_composition"][name] for name in component_names)
        for k in range(n_components):
            ws.cell(top_row, 2 + k).value = top_total * top_x[k]
            ws.cell(bottom_row, 2 + k).value = bottom_total * bottom_x[k]

    ws = wb["Specifications"]
    _set_or_append_spec(ws, "Initial Conditions Source", "ChemSep RB seed, PR-76, reflux ratio plus bottoms product flow")
    _set_or_append_spec(ws, "ChemSep Seed Reboiler Duty (Btu/h)", summary["reboiler_duty_MW"] * 1.0e6 * W_TO_BTUH)
    _set_or_append_spec(ws, "ChemSep Seed Condenser Duty (Btu/h)", summary["condenser_duty_MW"] * 1.0e6 * W_TO_BTUH)
    _set_or_append_spec(ws, "ChemSep Seed Reboiler Duty Difference vs Source (%)", summary["reboiler_duty_vs_source_pct"])

    _append_note(
        wb,
        "ChemSep seed initial conditions",
        (
            f"{_tag()}: Updated Initial Conditions, product stream conditions, and boundary liquid "
            f"compositions from {results_dir}. Holdups and geometry were preserved. "
            f"ChemSep RB-seed reboiler duty is {summary['reboiler_duty_MW']:.6g} MW, "
            f"{summary['reboiler_duty_vs_source_pct']:.2f}% above the source duty."
        ),
    )

    wb.save(excel_path)
    print(f"updated: {excel_path}")
    if backup_path is not None:
        print(f"backup:  {backup_path}")
    print(f"stages_updated={len(profiles)}")
    print(f"chemsep_reboiler_duty_MW={summary['reboiler_duty_MW']:.9g}")
    print(f"duty_difference_vs_source_pct={summary['reboiler_duty_vs_source_pct']:.6g}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
